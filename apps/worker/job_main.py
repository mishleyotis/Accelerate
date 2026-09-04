"""dmai-worker Cloud Run Job — the package scan (TRD §07, ten steps).

One execution: walk the intake tree, diff against import_scans, and for
every client folder whose scoring workbook is new or changed, assemble
the package (manifest + workbook + report), persist it into the ingested
tier and embed the bundle. Idempotent: an unchanged tree creates nothing.

Env: INTAKE_FOLDER_ID (the General DMAs tree), DB via the Cloud SQL
connector (or LOCAL_DATABASE_URL), EMBED_MODEL_DIR to enable the
embedding pass, MAX_PACKAGES to bound one execution (default 3 — the
Scheduler fires every 30 minutes; steady drain beats a marathon run).
"""
from __future__ import annotations

import json
import os
import re
import sys
import tempfile
import traceback

import openpyxl
from datetime import datetime, timezone

from dma_worker import drive, intake_status, persist
from dma_worker.counts import recount_run, recount_where
from dma_worker.persist import _institution, persist_package
from dma_worker.report_parser import parse_report
from dma_worker.scan_runner import (SCAN_FAILED, SCAN_SUCCEEDED, finish_scan,
                                    open_scan, run_scan)
from dma_worker.workbook_parser import (Observation,
                                        _stated_overall_grain,
                                        mine_evidence_from_rationales,
                                        parse_evidence_master,
                                        parse_grain_summaries,
                                        parse_run_metadata,
                                        parse_peer_benchmarks,
                                        parse_evidence_index,
                                        parse_recommendations,
                                        parse_research_workbook,
                                        parse_scoring_workbook,
                                        parse_technographic_scan,
                                        parse_tech_register,
                                        workbook_tab_coverage,
                                        merge_evidence_sources)


# ── ONE Cloud SQL Connector, imported ────────────────────────────────
# Per-connection `Connector()` is one Cloud SQL ADMIN API call per
# connection; under NullPool that is one per checkout, and it produced a
# 429 on sqladmin.googleapis.com/.../connectSettings on 2026-08-31.
import sys as _sys
from pathlib import Path as _Path


def _shared_roots():
    here = _Path(__file__).resolve()
    roots = [here.parent / "shared", here.parent.parent / "shared"]
    if len(here.parents) > 3:
        roots.append(here.parents[3] / "packages" / "shared")
    return roots


for _cand in _shared_roots():
    if _cand.exists() and str(_cand) not in _sys.path:
        _sys.path.insert(0, str(_cand))

from cloudsql import connect as _cloudsql_connect  # noqa: E402

def _connect():
    return _cloudsql_connect(
        local_user="dmai-worker@digital-maturity-assessor.iam")


# Artefact naming across the shipped corpus is not standardised — every
# synthesis run named its own files. These lists come from the intake tree
# itself (76 of 105 client folders were being skipped on naming alone).
_WB_DECOYS = ("research", "techstack", "tech_stack", "tech stack",
              "technology_stack", "technology stack", "explorium", "toolkit",
              "weight", "appendix", "template", "tracker")
_RPT_DECOYS = ("template",)

#: The folder a superseded package is archived into, INSIDE the client
#: folder. Kept in sync with `assemble.ARCHIVE_DIR` on the engine side; the
#: two are one name and this comment is the only place that says so.
ARCHIVE_SEGMENT = "_superseded"

#: Directories inside a client folder that hold COPIES, not the package.
#:
#: `_superseded` is the engine's own archive and was always excluded. These
#: are the ones agents create while working, and nothing excluded them:
#: measured 2026-09-03 on Bank of Travelers Rest, one client folder held four
#: workbooks at three depths — root, `DMAI - <client>/`, and
#: `DMAI - <client>/memory-backup/` — three of them byte-identical. The scan
#: reads the whole tree at any depth and keeps ONE artefact per kind, so
#: every copy was a candidate to be chosen over the current one, and the copy
#: it chose was a research-stage workbook with zero scored cells while the
#: assessment workbook holding all 688 sat in `memory-backup`.
#:
#: Matched on a whole path SEGMENT, case-insensitively, so a client legitimately
#: named "Backup Bancorp" is untouched.
COPY_SEGMENTS = ("memory-backup", "memory_backup", "backup", "backups",
                 "archive", "archived", "old", "_old", "superseded",
                 "previous", "versions", ".trash")

#: The Client Research Profile, in the spelling `classification.py` already
#: uses for it (priority 3). One artefact, one pattern, two classifiers that
#: now agree.
_PROFILE_RE = re.compile(r"client[_ ]profile.*\.docx$", re.I)

#: Every section kind from the client research profile wears this prefix.
#: The two reports' heading vocabularies overlap — both produce
#: `evidence_sources`, both produce `findings` — so without it a consumer
#: reading `document_sections` gets two documents' answers under one key.
PROFILE_KIND_PREFIX = "client_research:"


def _classify_artefact(f):
    """(kind, rank) for a package artefact, or None. Lower rank wins when a
    folder holds more than one candidate of a kind: the canonical name beats
    a variant, and a scoring workbook beats an assessment workbook."""
    name = f.name.lower()
    # A SUPERSEDED package, kept inside the client folder rather than in a
    # second one. The engine archives a previous run there when a new run
    # opens the same folder (assemble.ARCHIVE_DIR), because
    # `runs.source_folder_id` keys on the folder and forking it would orphan
    # every run before this one. The scan reads the whole tree at any depth
    # and keeps ONE artefact per kind, so without this an archived workbook
    # is a candidate to be chosen over the current one — the retention would
    # have created the very ambiguity it exists to remove.
    if any(seg.strip().lower() == ARCHIVE_SEGMENT for seg in f.path_segments):
        return None
    # The same rule for the directories AGENTS leave copies in. Without it a
    # backup of last week's workbook competes with this week's on equal
    # terms, and filename order decides which the client sees.
    if any(seg.strip().lower() in COPY_SEGMENTS for seg in f.path_segments[:-1]):
        return None
    if name.endswith(".json") and "manifest" in name:
        # run_manifest.json canonical; L1_run_manifest.json / MANIFEST.json seen.
        return "manifest", (0 if name == "run_manifest.json" else 1)
    if name.endswith(".json") and "evidence_index" in name:
        # AUD-0091: the richest evidence store in every package was
        # classified `package_structured` by classification.py, recorded into
        # import_files.classified_kind by the scanner — and then DROPPED,
        # because this function accepted nothing but manifest.json, .xlsx,
        # .xlsm and .docx, and `_package_groups` keys only on this function.
        #
        # Gate M exists because of exactly this file: a client shipped with
        # 85% of its evidence unURLed while `01_evidence/evidence_index.json`
        # carried 752 items with 748 URLs. The link was in the package the
        # whole time and nothing read it.
        return "evidence_index", (0 if name == "evidence_index.json" else 1)
    if "technographic" in name and name.endswith((".json", ".docx")):
        # The fourth final output (engine/assemble.py's package contract).
        # The .json sidecar outranks the .docx because it is the machine
        # copy the ingest parses; classifying the .docx too means a package
        # that shipped only the document is still SEEN, and the miss of its
        # sidecar is reportable rather than invisible.
        return "techscan", (0 if name.endswith(".json") else 1)
    if name.endswith((".xlsx", ".xlsm")):
        # The research workbook is its own artefact, not a decoy. It carries
        # the evidence tier's authority — per-subcap linkage at fact grain,
        # the verbatim passage behind each fact, and the ERS/date ledger the
        # scoring workbook omits — so excluding it left every ingested item
        # undated, unranked and with a scraped excerpt. It never supplies a
        # score; `scoring` remains the only authority for that.
        in_research_folder = any("research" in s.lower() for s in f.path_segments)
        if "research" in name or in_research_folder:
            if "workbook" in name or "research" in name:
                return "research", (0 if "research_workbook" in name else 1)
            return None
        if any(d in name for d in _WB_DECOYS):
            return None
        if "scoring" in name:
            return "workbook", 0
        if "assessment" in name:      # DMA_Assessment_Workbook_<client>.xlsx
            return "workbook", 1
        if "workbook" in name:        # DMA_Workbook_<client>.xlsx
            return "workbook", 2
        return None
    if name.endswith(".docx"):
        # THE CLIENT RESEARCH PROFILE, and it is its own artefact rather than
        # a decoy. `_RPT_DECOYS` used to carry "profile", so every .docx whose
        # name contained it returned None — and the engine's own filename for
        # this report is `Client_Profile_Research_<entity>_<date>.docx`. All
        # eight of its sections therefore reached no table in the app, while
        # four page packs named "Client Profile DOCX" as their source of
        # truth for firmographics, the leadership roster, focus-area quotes
        # and the financial series.
        #
        # Worse, the OTHER classifier in this same service already recognised
        # it — `classification.ARTEFACT_REGISTRY` matches it as
        # `client_profile` priority 3 and the scanner writes that into
        # `import_files.classified_kind`. Classified, recorded, then dropped:
        # the AUD-0091 shape this codebase names by number.
        if _PROFILE_RE.search(name):
            return "profile", 0
        if any(d in name for d in _RPT_DECOYS):
            return None
        if "assessment_report" in name or "assessment report" in name:
            return "report", 0
        if name == "report.docx":
            return "report", 1
        if "report" in name:
            return "report", 2
        return None
    return None


def package_key(tree):
    """A stable display key per client folder, from the tree as a whole.

    The key is the folder's NAME, because that is what `runs.source_folder_id`
    stores and what every downstream lookup (backfill, intake status,
    FORCE_FOLDER) matches on — except where a name is not unique. The
    production intake tree carries two distinct folders both called
    "Corporate America Credit Union - DMA", each with its own scoring
    workbook and its own report, and grouping by name merged them into one
    package: one client's workbook was silently discarded on every firing.
    A colliding name is qualified by its source folder id, so both packages
    ingest and both say which folder they came from.
    """
    ids_by_name: dict = {}
    for f in tree:
        name = f.path_segments[0] if f.path_segments else "?"
        ids_by_name.setdefault(name, set()).add(
            f.parent_ids[0] if f.parent_ids else None)
    collided = {n for n, ids in ids_by_name.items() if len(ids) > 1}

    def key(f):
        name = f.path_segments[0] if f.path_segments else "?"
        if name not in collided:
            return name
        return f"{name} [{(f.parent_ids[0] if f.parent_ids else '?')}]"
    key.collisions = {n: sorted(x for x in ids_by_name[n] if x)
                      for n in sorted(collided)}
    return key



def _mtime(f) -> str:
    """A file's modified time as a sortable string, "" when the source gives
    none. RFC-3339 from Drive sorts lexicographically, so no parsing is
    needed and a missing value loses every comparison — which is right: an
    undated candidate should never displace a dated one."""
    return getattr(f, "modified_time", "") or ""

def _package_groups(to_process, key=None):
    """Group changed files by client folder and keep the best candidate per
    artefact kind. Returns ALL groups — the caller splits ingestable packages
    from partial ones (a folder whose manifest arrived before its workbook
    must retry, not vanish)."""
    key = key or package_key(to_process)
    groups: dict = {}
    ranks: dict = {}
    for f in to_process:
        root = key(f)
        g = groups.setdefault(root, {"folder": root})
        r = ranks.setdefault(root, {})
        c = _classify_artefact(f)
        if not c:
            continue
        kind, rank = c
        # Lowest rank wins; among EQUAL ranks the most recently modified
        # does. The tie used to fall to whichever file the walk met first —
        # stable, arbitrary, and unrelated to which copy is current. That is
        # the "agents using old cached reports" report of 2026-09-03: an
        # agent rewrites the workbook, an older sibling of the same rank is
        # still met first, and the run reads the stale one while the scores
        # look missing.
        prev = g.get(kind)
        if prev is None or rank < r[kind] or (
                rank == r[kind] and _mtime(f) > _mtime(prev)):
            g[kind], r[kind] = f, rank
        # The runners-up are KEPT, under `<kind>__alt`, in rank order.
        #
        # Precedence here is decided by FILENAME, and a filename is not a
        # claim about content. Bank of Travelers Rest ships both
        # `DMA_Scoring_Workbook_*` (rank 0 — and a RESEARCH-stage v5 file:
        # 688 rows seen, column D empty by contract, 0 scored) and
        # `DMA_Assessment_Workbook_*` (rank 1, 688 scored, composite 1.71 at
        # Pillar_Summary!C6). The name won, the scores lost, and eighteen of
        # that entity's nineteen runs landed with `scored_cells = 0`.
        #
        # Discarding the runner-up made that unrecoverable without a human
        # renaming files in Drive. Kept, the ingest can fall through to it
        # when the chosen workbook states no scores — see `_pick_workbook`.
        g.setdefault(f"{kind}__alt", []).append((rank, f))
    for gk in groups.values():
        for k in [k for k in gk if k.endswith("__alt")]:
            gk[k] = [f for _, f in sorted(gk[k], key=lambda t: t[0])][1:]
            if not gk[k]:
                del gk[k]
    return groups


def _requeue(conn, parts, folder, reason):
    """Blank the stored checksums of this package's artefacts so the next
    scan classifies them as CHANGED and retries. Rows are kept (FKs from
    document_sections / parser_observations may point at them)."""
    ids = [parts[k].file_id
           for k in ("manifest", "workbook", "report", "profile")
           if k in parts]
    cur = conn.cursor()
    for fid in ids:
        cur.execute("UPDATE import_files SET checksum = '' WHERE artefact_id = %s",
                    (fid,))
    conn.commit()
    print(f"requeue: {folder} — {reason} ({len(ids)} artefact(s) rescan next firing)")


# A package that cannot be parsed cannot be parsed on the next firing either.
# Blanking its checksums every 30 minutes retried Zions Bancorporation 140
# times against the same ValueError and reported nothing anywhere. Three
# attempts, then the package is quarantined: recorded by name, left out of the
# diff, and surfaced by intake_status. A new upload changes the checksum and
# starts the budget over; FORCE_FOLDER retries it on demand.
MAX_INGEST_ATTEMPTS = int(os.environ.get("MAX_INGEST_ATTEMPTS", "3"))


def _prior_attempts(conn, artefact_id, checksum) -> int:
    """Failed ingests already recorded against this artefact AT THIS CHECKSUM.

    The requeue blanks the checksum, so the live value is '' between the
    failure and the next scan; both the blank and the real checksum count as
    the same package, and only a genuinely different checksum resets."""
    cur = conn.cursor()
    cur.execute(
        """SELECT detail FROM parser_observations
            WHERE artefact_id = %s AND kind = %s ORDER BY occurred_at""",
        (artefact_id, intake_status.PACKAGE_FAILURE_KIND))
    n = 0
    for (detail,) in cur.fetchall():
        d = json.loads(detail) if isinstance(detail, (str, bytes)) else (detail or {})
        if checksum and d.get("checksum") and d["checksum"] not in ("", checksum):
            n = 0                       # a different upload: budget starts over
            continue
        n += 1
    return n


def _record_package_failure(conn, parts, folder, exc) -> bool:
    """Record one failed ingest and decide whether to retry it.

    Returns True when the package was requeued, False when it was
    quarantined. Either way the failure is a row in parser_observations, so
    "why has this folder never produced a run" has an answer that outlives
    the log retention window.
    """
    wb = parts.get("workbook")
    error = f"{type(exc).__name__}: {exc}"[:400]
    attempts = 1
    if wb is not None:
        attempts = _prior_attempts(conn, wb.file_id, wb.checksum) + 1
    quarantined = attempts >= MAX_INGEST_ATTEMPTS
    if wb is not None:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO parser_observations (artefact_id, kind, detail, occurred_at)
               VALUES (%s,%s,%s, now())""",
            (wb.file_id, intake_status.PACKAGE_FAILURE_KIND,
             json.dumps({"folder": folder, "error": error, "attempt": attempts,
                         "checksum": wb.checksum, "quarantined": quarantined})))
        conn.commit()
    if quarantined:
        print(f"quarantine: {folder} — {error} (attempt {attempts} of "
              f"{MAX_INGEST_ATTEMPTS}; not requeued, named by intake status)")
        return False
    _requeue(conn, parts, folder, f"failed: {exc!r} (attempt {attempts} of "
                                  f"{MAX_INGEST_ATTEMPTS})")
    return True



def _pick_workbook(token, td, parts):
    """The workbook that actually CARRIES SCORES, not the one whose filename
    ranked first.

    `_classify_artefact` ranks `scoring` above `assessment` above `workbook`,
    on the filename alone. That is a reasonable default and it is not a claim
    about content. Bank of Travelers Rest ships both:

        DMA_Scoring_Workbook_...xlsx      rank 0   research-stage v5:
                                                   688 rows seen, column D
                                                   empty BY CONTRACT, 0 scored
        DMA_Assessment_Workbook_...xlsx   rank 1   688 scored, composite 1.71
                                                   at Pillar_Summary!C6

    The name won and the scores lost: eighteen of that entity's nineteen runs
    landed with `scored_cells = 0`, each one a promoted-looking package with
    nothing in it. Nothing in the pipeline could recover from it, because the
    runner-up was discarded at grouping time.

    So: parse the chosen file, and if it yields NO scored cell while an
    alternate exists, parse the alternate. Take the first one that states
    scores. An empty workbook is still a legitimate answer (a research-stage
    package genuinely has no scores yet) — this only prefers a sibling that
    HAS them, and never prefers a smaller score set over a larger one.

    Returns `(path, note)`; `note` is None when the first choice stood, and
    otherwise an Observation-shaped dict recording which file was read
    instead and why, so the substitution is never silent.
    """
    def _fetch(stat, name):
        path = os.path.join(td, name)
        with open(path, "wb") as fh:
            fh.write(drive.download(token, stat.file_id))
        return path

    chosen = parts["workbook"]
    path = _fetch(chosen, "wb.xlsx")
    alts = parts.get("workbook__alt") or []
    if not alts:
        return path, None
    try:
        scored = len({s.subcap_id for s in parse_scoring_workbook(path).scores})
    except Exception:                      # noqa: BLE001 — a bad parse is the caller's to report
        return path, None
    if scored:
        return path, None

    for i, alt in enumerate(alts):
        try:
            apath = _fetch(alt, f"wb_alt{i}.xlsx")
            n = len({s.subcap_id for s in parse_scoring_workbook(apath).scores})
        except Exception:                  # noqa: BLE001
            continue
        if n:
            note = {"kind": "workbook_substituted",
                    "detail": {"chosen_by_name": chosen.name, "chosen_scored": 0,
                               "read_instead": alt.name, "scored": n,
                               "why": ("the filename-ranked workbook states no "
                                       "scored cell and a sibling does; the "
                                       "name is not a claim about content")}}
            print(f"workbook: {chosen.name} states 0 scored cells — "
                  f"reading {alt.name} instead ({n} cells)")
            return apath, note
    return path, None

def _ingest_one(conn, token, folder, parts, remint=False):
    """Download, parse and persist one package. Atomic: persist_package
    commits once at the end, so an exception anywhere leaves nothing."""
    with tempfile.TemporaryDirectory() as td:
        # The scoring workbook is the authority; a package that ships no
        # manifest still ingests, with identity from the folder name (the
        # cascade's signal 4 → PENDING_REVIEW) and an observation recorded.
        if "manifest" in parts:
            # utf-8-sig: some shipped manifests carry a BOM
            manifest = json.loads(
                drive.download(token, parts["manifest"].file_id).decode("utf-8-sig"))
        else:
            manifest = {}
        wb_path, wb_note = _pick_workbook(token, td, parts)
        # Declared BEFORE the report parse, which appends to it.
        companion: list = []
        if wb_note:
            # A substitution the run must be able to explain later: which
            # file was read, which was ranked first, and why.
            companion.append(Observation(wb_note["kind"], None, wb_note["detail"]))
        sections = []
        if "report" in parts:
            rp = os.path.join(td, "report.docx")
            with open(rp, "wb") as fh:
                fh.write(drive.download(token, parts["report"].file_id))
            # `companion` collects what the readers could not
            # recognise; the report parser now files an unmapped
            # Heading1 under its own name and says so, instead of
            # under whichever numbered kind the count reached.
            sections = parse_report(rp, companion)
            for sec in sections:
                sec.artefact_id = parts["report"].file_id

        # THE SECOND REPORT. Both are client-facing artefacts of the same
        # run, and until 2026-08-30 only one of them was ever read. Its
        # sections are NAMESPACED by report because the two vocabularies
        # overlap head-on — the profile's "Evidence base" and the assessment
        # report's "Evidence and its limits" both resolve to
        # `evidence_sources`, and its "Negative findings and what they bound"
        # resolves to `findings` — so an un-namespaced merge would put two
        # documents' answers under one key with no way to tell which said
        # what. `document_sections.section_kind` is plain TEXT with no enum
        # and no unique constraint, so the prefix is storable; `embed.py`
        # binds the kind and never reads it, and `bundle.py` passes it
        # through, so neither consumer breaks on the longer name.
        if "profile" in parts:
            pp = os.path.join(td, "client_profile.docx")
            with open(pp, "wb") as fh:
                fh.write(drive.download(token, parts["profile"].file_id))
            profile_sections = parse_report(pp, companion)
            for sec in profile_sections:
                sec.section_kind = f"{PROFILE_KIND_PREFIX}{sec.section_kind}"
                sec.artefact_id = parts["profile"].file_id
            sections = list(sections) + profile_sections
            print(f"ingest: {folder} client profile — "
                  f"{len(profile_sections)} section(s)")

        research = {}
        evidence_index: list = []
        if "evidence_index" in parts:
            ei_path = os.path.join(td, "evidence_index.json")
            with open(ei_path, "wb") as fh:
                fh.write(drive.download(token, parts["evidence_index"].file_id))
            evidence_index = parse_evidence_index(ei_path, companion)
            print(f"ingest: {folder} evidence index — {len(evidence_index)} "
                  f"items, "
                  f"{sum(1 for e in evidence_index if e.get('source_url'))} "
                  f"with a URL")
        if "techscan" in parts:
            ts_name = parts["techscan"].name.lower()
            ts_path = os.path.join(td, "technographic_scan"
                                       + (".json" if ts_name.endswith(".json")
                                          else ".docx"))
            with open(ts_path, "wb") as fh:
                fh.write(drive.download(token, parts["techscan"].file_id))
            n = parse_technographic_scan(ts_path, companion)
            print(f"ingest: {folder} technographic scan — {n} detection(s) "
                  f"recorded from {parts['techscan'].name}")
        if "research" in parts:
            rw_path = os.path.join(td, "research.xlsx")
            with open(rw_path, "wb") as fh:
                fh.write(drive.download(token, parts["research"].file_id))
            research = parse_research_workbook(rw_path, companion)
            print(f"ingest: {folder} research workbook — "
                  f"{len(research.get('ledger') or [])} ledger rows, "
                  f"{len(research.get('links') or [])} linked cells, "
                  f"{len(research.get('absent') or [])} recorded absences")

        wb = parse_scoring_workbook(wb_path)

        # WHICH TABS DOES ANYTHING READ? Recorded on every ingest, before the
        # readers run, because the answer changed under us: the Golden 1
        # workbook ships 43 tabs against readers that claimed 12, and the
        # gap was invisible. A surface written from a tab in this census's
        # worklist renders empty because nothing read it, and a producer who
        # cannot tell that apart from a client with nothing to say will write
        # the absence.
        workbook_tab_coverage(wb_path, companion)

        # The technology register at the grain the techstack contract wants.
        # Run for its OBSERVATIONS as well as its rows: CG-20 (product and
        # vendor stating one string) and CG-12 (a detection_basis over the
        # face-slot budget) are properties of the workbook, and naming them
        # here reports them against the run instead of surfacing them as a
        # producer defect at validation time.
        tech_register = parse_tech_register(wb_path, companion)
        if tech_register:
            print(f"ingest: tech register {len(tech_register)} product row(s)")

        # Every companion tab appends what it could not read to one list; the
        # persist writes them as parser_observations against the run, so a tab
        # the parser did not recognise leaves a record naming the tab, the
        # column and the spelling it expected — not an absent section.
        res = persist_package(
            conn,
            manifest=manifest,
            workbook=wb,
            source_folder_id=folder,
            # The workbook's ledger FIRST, the package index second: the
            # index fills gaps (a URL, a date, a longer excerpt) and never
            # overwrites what the workbook stated. AUD-0091 / gate M.
            evidence=merge_evidence_sources(
                parse_evidence_master(wb_path, companion), evidence_index,
                companion),
            # WHOSE assessment this is, so the peer parser can keep the
            # subject out of its own cohort. `Peer_Benchmarks` carries the
            # entity's own score in a named column (`FUB_Score`) beside the
            # cohort's, and a parser told nothing about the client stored it
            # as a peer institution — the client benchmarked against itself.
            # Both signals are passed because either may be absent: the
            # manifest's institution block is heterogeneous across the corpus
            # and some packages ship no manifest at all, in which case the
            # folder name is the identity the rest of the pipeline uses too.
            peers=parse_peer_benchmarks(
                wb_path, companion,
                subject_names=(_institution(manifest).get("name"), folder)),
            recommendations=parse_recommendations(wb_path, companion),
            companion_observations=companion,
            artefact_id=parts["workbook"].file_id,
            # The bytes this run was read from. A requeue blanks the live
            # checksum in import_files, so the run keeps its own copy and a
            # retry of an unchanged package resolves to the run it already
            # produced instead of minting a second one.
            artefact_checksum=parts["workbook"].checksum,
            remint=remint,
            sections=sections,
            report_artefact_id=(parts["report"].file_id
                                if "report" in parts else None),
            grains=parse_grain_summaries(wb_path, companion),
            # The workbook's Run_Metadata tab, stored beside the
            # manifest so `last_written_at` can resolve a date the
            # package's own manifest never stated.
            wb_metadata=parse_run_metadata(wb_path),
            research=research,
        )
        rationales = {s.subcap_id: s.rationale for s in wb.scores if s.rationale}
        return res, rationales




def backfill_workbook_metadata(conn, token, groups) -> int:
    """Fill `run_manifest.payload.workbook_metadata` for runs ingested before
    anything read the workbook's own `Run_Metadata` tab.

    THE DEFECT. `run_assessment_date` walks six manifest keys and then the
    YYYYMMDD token in the request id. Golden 1 carries none of the six, and
    `DMA-2026-GOLDEN1-001` has no eight-digit token, so the run resolved
    UNKNOWN and served no assessment date — while its own workbook states
    `last_written_at = 2026-08-31T09:33:59Z` on a tab nothing read for a date.

    WHAT THAT COSTS. Not one header line: the freshness dot has nothing to
    draw, and the same candidate list feeds `runs.completed_at`, which is
    every evidence row's `reference_date`. With it null the generated
    `age_months` is null and `recency_band` falls to UNVERIFIED for EVERY
    item — 537 rows on this run, each rendering "unverified" beside evidence
    the package dated.

    Migration 0058 is the read half (the probe row, and the view handing the
    resolver `workbook_metadata || manifest`). This is the other half: a run
    ingested before it cannot benefit without re-reading its own workbook,
    and the package scan is idempotent, so an unchanged tree re-ingests
    nothing.

    Writes ONLY the `workbook_metadata` key — `payload["manifest"]` is never
    touched, because a key written into the package's own artefact after the
    fact is indistinguishable from one the package shipped.

    `completed_at` is filled in the same pass and only where it is NULL: it
    is the column the evidence bands hang off, and leaving it behind would
    fix the dot while every row still read UNVERIFIED.
    """
    cur = conn.cursor()
    cur.execute("""SELECT r.id, r.source_folder_id
                     FROM runs r
                     JOIN run_manifest m ON m.run_id = r.id
                    WHERE r.source_folder_id IS NOT NULL
                      AND m.payload -> 'workbook_metadata' IS NULL
                    ORDER BY r.source_folder_id""")
    todo = cur.fetchall()
    print(f"backfill-wbmeta: {len(todo)} run(s) hold no workbook metadata")
    filled = skipped = empty = dated = failed = 0
    for run_id, folder in todo:
        parts = groups.get(folder) or {}
        if "workbook" not in parts:
            skipped += 1
            continue
        try:
            with tempfile.TemporaryDirectory() as td:
                wp = os.path.join(td, "wb.xlsx")
                with open(wp, "wb") as fh:
                    fh.write(drive.download(token, parts["workbook"].file_id))
                md = parse_run_metadata(wp)
            if not md:
                empty += 1
                print(f"backfill-wbmeta: {folder}: workbook states no Run_Metadata")
                continue
            cur.execute(
                """UPDATE run_manifest
                      SET payload = jsonb_set(payload, '{workbook_metadata}',
                                              %s::jsonb, true)
                    WHERE run_id = %s""",
                (json.dumps(md), run_id))
            # The date the bands hang off, only where nothing stated one.
            stamp = persist._stated_completed_at(md)
            if stamp:
                cur.execute(
                    "UPDATE runs SET completed_at = %s "
                    " WHERE id = %s AND completed_at IS NULL",
                    (stamp, run_id))
                dated += cur.rowcount or 0
            conn.commit()
            print(f"backfill-wbmeta: {folder} -> {len(md)} key(s)"
                  + (f", completed_at {stamp}" if stamp else ""))
            filled += 1
        except Exception as exc:  # noqa: BLE001 — one bad workbook must not sink the pass
            conn.rollback()
            failed += 1
            print(f"backfill-wbmeta FAILED: {folder}: {exc!r}")
    print(f"backfill-wbmeta: {filled} filled ({dated} gained a completed_at), "
          f"{skipped} have no workbook artefact, {empty} state none, "
          f"{failed} failed")
    return 1 if failed else 0

#: How many null-composite runs one scheduled firing will re-read. The pass
#: downloads a workbook per run, and the Job's task timeout is 3600s; a
#: backlog must not be able to spend the whole firing here and starve the
#: scan that is this Job's actual purpose. What is left over is NAMED in the
#: log rather than dropped silently — a cap nobody can see is a cap that
#: reads as "there was nothing else".
COMPOSITE_REPAIR_PER_FIRING = 25


def composite_reader_fingerprint() -> str:
    """A hash of the composite reader itself.

    WHY A HASH AND NOT A VERSION NUMBER. The repair below records
    `composite_absent` for a run whose workbook states no overall, so the
    next firing does not download it again. That record is only safe while
    the reader is unchanged: improve the reader and every one of those runs
    must be looked at again, or the improvement reaches new packages only
    and every existing one stays silently empty — which is precisely the
    failure this whole change exists to end, reintroduced by the fix for it.

    A hand-maintained `READER_VERSION = 3` would work exactly as long as
    everyone remembers to bump it. Hashing the reader's own source and the
    constants it reads cannot be forgotten. A comment-only edit re-opens the
    work list too; re-reading is cheap and correct, being silently stale is
    neither.
    """
    import hashlib
    import inspect
    from dma_worker import workbook_parser as _wp
    src = inspect.getsource(_wp._stated_overall_grain)
    consts = repr((_wp._GRAIN_TABS["pillars"], _wp._GRAIN_ANCHORS["pillars"],
                   _wp._GRAIN_SCORE_KEYS, _wp._OVERALL_LABELS))
    return hashlib.sha256((src + consts).encode()).hexdigest()[:16]


#: Workbook downloads are ~1.5 MB each and this pass re-parses whole
#: workbooks, so the per-firing cap is small. What is left over is NAMED.
EVIDENCE_REPAIR_PER_FIRING = 5


#: An evidence row the CONNECTOR minted carries its own `E-CC-nnn` id, so the
#: numeric-suffix join below cannot reach it — `E-CC-569` is the 569th id the
#: connector allocated, not the 569th row of anyone's workbook. Those rows do
#: say which workbook row they came from, in words, inside `source_name`:
#:     "DFPI Regulated Entity Record … [package evidence id E-5123]"
#:     "Banking Dive — How Golden 1 used AI … — package id E-055;"
#: MEASURED 2026-09-04 on Golden 1 Credit Union: 223 served rows had no URL
#: after the suffix pass, 222 of them named a workbook row this way, and every
#: one of those 222 resolved to a ledger row that states a URL. That marker is
#: the package's own statement of provenance — the honest join key — and it is
#: the difference between a drawer with a source and a drawer without one.
_PACKAGE_ID_MARKER = re.compile(
    r"package\s+(?:evidence\s+)?id[:\s]\s*([A-Za-z0-9][A-Za-z0-9._-]*)",
    re.IGNORECASE)


def stated_package_id(source_name: str | None) -> str | None:
    """The workbook evidence id a served row names in its own source_name."""
    m = _PACKAGE_ID_MARKER.search(source_name or "")
    return m.group(1).rstrip(".-_") if m else None


#: How much of the shorter quote has to lead the longer one for the two to be
#: the same passage. The ingest clips an excerpt into the 50-500 character
#: window (invariant 4), so the served row routinely holds a PREFIX of what
#: the Evidence_Master tab states — same passage, fewer characters.
_EXCERPT_HEAD = 60


def excerpt_agrees(served: str | None, stated: str | None) -> bool | None:
    """Do the served row and the workbook row quote the same passage?

    THE MARKER IS TEXT A PRODUCER WROTE. `[package evidence id E-5123]` is a
    claim about provenance, and a claim can be wrong — a mistyped digit would
    hang a real, checkable URL under somebody else's quote, which is worse
    than the blank drawer it replaced. The excerpt is the independent check:
    both rows carry the VERBATIM passage, so if they agree, the marker points
    where it says it does.

    MEASURED 2026-09-04 on Golden 1 Credit Union: 258 served rows name a
    workbook row, and under this rule all 258 agree with the row they name
    and none disagree. So the guard costs nothing on good data and is the
    only thing standing between a typo and a mis-sourced citation.

    True agrees, False contradicts, None when one side has no quote to
    compare — never guessed either way."""
    a = " ".join((served or "").split()).lower()
    b = " ".join((stated or "").split()).lower()
    if not a or not b:
        return None
    if a == b:
        return True
    return a.startswith(b[:_EXCERPT_HEAD]) or b.startswith(a[:_EXCERPT_HEAD])


def evidence_reader_fingerprint() -> str:
    """A hash of the evidence reader, on `composite_reader_fingerprint`'s
    reasoning: a run this reader has already been through is not work until
    the reader itself changes, and nobody should have to remember to say so.

    It covers the MATCHERS as well as the parser: teaching the pass a second
    way to reach a row is an improvement to the reader, and every run that
    was given up on under the old one has to re-open by itself."""
    import hashlib
    import inspect
    from dma_worker import workbook_parser as _wp
    src = inspect.getsource(_wp.parse_evidence_master)
    src += inspect.getsource(stated_package_id)
    src += inspect.getsource(excerpt_agrees)
    consts = repr((_wp._EV_TABS, _wp._EV_ID_ANCHORS, _wp._FILLABLE,
                   _PACKAGE_ID_MARKER.pattern, _EXCERPT_HEAD))
    return hashlib.sha256((src + consts).encode()).hexdigest()[:16]


def adopt_orphan_runs(conn, token, groups) -> int:
    """Give a run with no `source_folder_id` its package back, by IDENTITY.

    MEASURED 2026-09-04. goeasy Ltd. carries eighteen runs under request id
    `DMA-RES-GSY-20260830-0002`, and NONE of them — including the promoted
    one the client directory reads — records a source folder. Every repair
    pass in this file finds a run's package through
    `runs.source_folder_id`, so all of them are blind to that client
    entirely: the composite repair reported one candidate run in the whole
    database and it was a different institution. The card renders the word
    "maturity" over an empty slot and no amount of folder fallback reaches
    it, because there is no folder to fall back to.

    A NAME IS NOT AN IDENTITY, and this deliberately does not use one. The
    package's own `run_manifest.json` states `run_id`, and that is the same
    string the ingest stored as `runs.request_id`. Matching those two is the
    package saying which run it produced, not this code guessing from a
    folder title — the distinction that `repair_evidence_namespace` was
    written about after a name-derived token was found owned by fourteen
    different entities.

    REFUSES AMBIGUITY. If two packages state the same run id, neither is
    adopted and both are named: one of them is wrong and picking either
    would attach a client's scores to another client's workbook.

    Writes only `source_folder_id`, only where it is NULL, and only from a
    manifest that names the run. Every other repair works afterwards because
    the run is finally traceable to the package it came from.
    """
    cur = conn.cursor()
    cur.execute("""SELECT r.id, r.request_id
                     FROM runs r
                    WHERE r.source_folder_id IS NULL
                      AND r.request_id IS NOT NULL
                    ORDER BY r.id""")
    orphans = cur.fetchall()
    if not orphans:
        return 0
    wanted = {req for _rid, req in orphans}
    print(f"adopt: {len(orphans)} run(s) carry no source folder "
          f"({len(wanted)} distinct request id(s))")

    # One manifest read per folder, and only folders that ship one.
    stated: dict = {}
    clashes: dict = {}
    for folder, parts in sorted(groups.items()):
        if "manifest" not in parts:
            continue
        try:
            raw = drive.download(token, parts["manifest"].file_id)
            run_id = (json.loads(raw.decode("utf-8-sig")) or {}).get("run_id")
        except Exception as exc:      # noqa: BLE001 — one manifest, not the pass
            print(f"adopt: {folder}: manifest unreadable ({exc!r})")
            continue
        if not run_id or run_id not in wanted:
            continue
        if run_id in stated and stated[run_id] != folder:
            clashes.setdefault(run_id, {stated[run_id]}).add(folder)
            continue
        stated[run_id] = folder

    for run_id, folders in clashes.items():
        stated.pop(run_id, None)
        print(f"adopt: REFUSED {run_id} — stated by {len(folders)} packages "
              f"({', '.join(sorted(folders))}); adopting either would attach "
              f"one client's run to another's workbook")

    adopted = 0
    for rid, req in orphans:
        folder = stated.get(req)
        if not folder:
            continue
        cur.execute("UPDATE runs SET source_folder_id = %s "
                    " WHERE id = %s AND source_folder_id IS NULL", (folder, rid))
        adopted += cur.rowcount or 0
    conn.commit()
    unplaced = len(orphans) - adopted
    print(f"adopt: {adopted} run(s) adopted by the package that names them, "
          f"{unplaced} still unplaced (no manifest states their request id)")
    return 0


def backfill_composite(conn, token, groups, *, forced: bool = True) -> int:
    """Fill `runs.composite` for a run whose workbook states it and whose
    ingest had nowhere to read it.

    THE DEFECT. `composite` is written once, at INSERT, from
    `WorkbookParse.composite` — and that field was set in exactly one place,
    `_parse_scorecard`, off the `2_Scorecard` tab. Only the claude_dma
    generation ships that tab. Every general_dma workbook (`P{n}_Subcap_Scoring`)
    and every rollup-only one took a different branch of
    `parse_scoring_workbook`, so the field came back None for the whole
    generation and `runs.composite` was written NULL.

    WHAT THAT COSTS, seen on Golden 1 (run 40971653) 2026-09-02. The
    directory card reads `serving_directory.composite` for its header
    figure. Golden 1's read NULL, so the card rendered the word "maturity"
    over an empty slot — beside its own four pillar bars, which resolve, and
    beside Axos Bank, which ships the other generation and shows 1.9. The
    workbook states the figure FOUR times: `Pillar_Summary!C6`,
    `Pillar_Rollup!C6`, `Executive_Summary` "Overall Maturity", and the
    OVERALL row's weighted contribution. No reader claimed any of them.

    `_stated_overall_grain` is the reader half and is already fixed. This is
    the other half: a run ingested before that fix cannot benefit from it
    without re-reading its own workbook, and the package scan is idempotent
    — an unchanged tree creates nothing to re-ingest.

    READ, never derived. The value written is the one on the row the
    workbook labels OVERALL. A run whose workbook states no overall is left
    NULL and says so; a mean of the pillars would be a derived figure in a
    column whose contract is that it was read, and indistinguishable from
    one afterwards.

    Additive and idempotent, on `backfill_grains`' pattern: a run already
    holding a composite is not in the work list, a folder shipping no
    workbook is skipped, and no run is created, deleted or re-scored.
    """
    cur = conn.cursor()
    reader = composite_reader_fingerprint()
    # A run whose workbook THIS reader already found nothing in is not work.
    # Without that exclusion the scheduled pass re-downloads every genuinely
    # composite-less workbook every thirty minutes, for ever; with it the
    # steady state is zero downloads and a reader change re-opens the lot.
    # `forced` (the BACKFILL_COMPOSITE mode) ignores the record and re-reads
    # everything, which is what a human reaching for the manual pass wants.
    #
    # THE RUN THAT SERVES IS THE ONE THAT MATTERS, and it is not always the
    # one holding the folder. MEASURED 2026-09-04: goeasy Ltd. carries
    # EIGHTEEN runs under one request id (`DMA-RES-GSY-20260830-0002`), every
    # one with a null composite, and this query — which required the run's own
    # `source_folder_id` — matched exactly ONE of them across the whole
    # database. A run re-ingested from the same package does not always carry
    # the folder forward, so the promoted run, the only one the directory
    # reads, was invisible to its own repair.
    #
    # A sibling under the SAME request id and the same entity is the same
    # package by definition — that is what a request id identifies — so its
    # folder is the right place to look, and the newest one wins. Not the
    # entity alone: two assessments of one client are two packages, and
    # reading one's composite out of the other's workbook would be a figure
    # from the wrong run wearing the right name.
    cur.execute("""SELECT r.id,
                          COALESCE(r.source_folder_id, sib.source_folder_id),
                          r.request_id
                     FROM runs r
                     LEFT JOIN LATERAL (
                           SELECT s.source_folder_id
                             FROM runs s
                            WHERE s.request_id = r.request_id
                              AND s.entity_id = r.entity_id
                              AND s.source_folder_id IS NOT NULL
                            ORDER BY s.run_seq DESC
                            LIMIT 1) sib ON TRUE
                    WHERE r.composite IS NULL
                      AND COALESCE(r.source_folder_id,
                                   sib.source_folder_id) IS NOT NULL
                      AND (%s OR NOT EXISTS (
                            SELECT 1 FROM parser_observations o
                             WHERE o.run_id = r.id
                               AND o.kind = 'composite_absent'
                               AND o.detail->>'reader' = %s))
                    ORDER BY 2, r.id""", (forced, reader))
    todo = [(rid, folder) for rid, folder, _req in cur.fetchall()]
    deferred = 0
    if not forced and len(todo) > COMPOSITE_REPAIR_PER_FIRING:
        deferred = len(todo) - COMPOSITE_REPAIR_PER_FIRING
        todo = todo[:COMPOSITE_REPAIR_PER_FIRING]
    print(f"backfill-composite: {len(todo)} run(s) hold no composite "
          f"(reader {reader}"
          + (f", {deferred} deferred to the next firing)" if deferred else ")"))
    filled = skipped = empty = failed = 0
    for run_id, folder in todo:
        parts = groups.get(folder) or {}
        if "workbook" not in parts:
            # NAMED, not tallied. The first production firing reported
            # "1 have no workbook artefact" and stopped there: which run,
            # and under what key, went unsaid — and the key is exactly what
            # goes wrong when a folder is renamed or a package moves.
            skipped += 1
            why = ("no folder by that key in this scan" if folder not in groups
                   else "folder is in the scan but ships no workbook")
            print(f"backfill-composite: run {run_id} skipped — {why} "
                  f"(key {folder!r})")
            continue
        try:
            with tempfile.TemporaryDirectory() as td:
                wp = os.path.join(td, "wb.xlsx")
                with open(wp, "wb") as fh:
                    fh.write(drive.download(token, parts["workbook"].file_id))
                wb = openpyxl.load_workbook(wp, read_only=True, data_only=True)
                try:
                    value, cell = _stated_overall_grain(wb)
                finally:
                    wb.close()
            if value is None:
                # A workbook that states no overall is a fact about the
                # workbook. Absent beats a number nobody wrote down.
                #
                # RECORDED, so the next firing does not pay for the same
                # download to learn the same thing. Stamped with the reader
                # that concluded it: a better reader has a different
                # fingerprint and the run returns to the work list on its own.
                cur.execute(
                    """INSERT INTO parser_observations
                           (run_id, kind, detail, occurred_at)
                       VALUES (%s,'composite_absent',%s, now())""",
                    (run_id, json.dumps({"reader": reader,
                                         "reason": "workbook states no overall"})))
                conn.commit()
                empty += 1
                print(f"backfill-composite: {folder}: workbook states none")
                continue
            cur.execute("UPDATE runs SET composite = %s WHERE id = %s",
                        (value, run_id))
            conn.commit()
            print(f"backfill-composite: {folder} -> {value} (from {cell})")
            filled += 1
        except Exception as exc:  # noqa: BLE001 — one bad workbook must not sink the pass
            conn.rollback()
            failed += 1
            print(f"backfill-composite FAILED: {folder}: {exc!r}")
    # PUBLISH IT. `serving_directory` is materialised: a repaired
    # `runs.composite` is invisible to every client until the view is
    # rebuilt, and until 0059 the worker's role could not ask for that — so
    # the repair committed a correct value that nothing could show. Once per
    # pass, and only when something actually changed; a refresh costs a full
    # rebuild and a pass that filled nothing has nothing to publish.
    if filled:
        try:
            cur.execute("SELECT refresh_serving_directory()")
            conn.commit()
            print("backfill-composite: directory refreshed")
        except Exception as exc:  # noqa: BLE001 — the values are committed
            conn.rollback()
            print(f"backfill-composite: filled {filled} but could NOT refresh "
                  f"the directory ({exc!r}) — the figures are in `runs` and "
                  f"will surface on the next refresh")
    print(f"backfill-composite: {filled} filled, {skipped} have no workbook "
          f"artefact, {empty} state none, {failed} failed")
    return 1 if failed else 0

def backfill_grains(conn, token, groups) -> int:
    """Fill the STATED pillar/category grains a run was ingested without.

    `run_manifest.workbook_grains` is written once, at ingest, and there is
    no update path — so a run ingested while the grain reader could not find
    its Score column keeps `pillars: 0, categories: 0` for ever, and
    re-scanning cannot repair it: the diff is idempotent, so an unchanged
    tree creates nothing to re-ingest.

    WHAT THAT COSTS, measured on Golden 1 (run 40971653) 2026-09-02. The
    workbook states 2.40 / 2.11 / 2.25 / 2.25 in Pillar_Summary and again in
    Pillar_Rollup. The reader looked for a column literally named `score` in
    a tab that names it `Weighted_Score`, so the run stored no grain. With no
    STATED grain, CG-07 resolves a quoted pillar figure against the mean of
    the run's own cells (2.1115 / 2.0345 / 2.0920 / 2.0585) and refuses the
    workbook's weighted figure at 0.05 — so the overview hero rendered four
    EMPTY BARS beside a peer tick, and the client directory card with them,
    on a run whose own report states all four scores.

    The aliases are already fixed in `parse_grain_summaries`. This is the
    other half: a run ingested before that fix cannot benefit from it without
    re-reading its own workbook.

    Additive and idempotent, on the pattern `backfill_sections` set. It
    re-parses the workbook and UPDATES the existing run's manifest in place:
    a run already holding grains is left alone, a folder shipping no workbook
    is skipped, and no run is created, deleted or re-scored. Only the
    `workbook_grains` key is written — `payload["manifest"]` is untouched.
    """
    cur = conn.cursor()
    cur.execute("""SELECT r.id, r.source_folder_id
                     FROM runs r
                     JOIN run_manifest m ON m.run_id = r.id
                    WHERE r.source_folder_id IS NOT NULL
                      AND COALESCE(
                            jsonb_array_length(
                              COALESCE(m.payload->'workbook_grains'->'pillars',
                                       '[]'::jsonb)), 0) = 0
                    ORDER BY r.source_folder_id""")
    todo = cur.fetchall()
    print(f"backfill-grains: {len(todo)} run(s) hold no stated pillar grain")
    filled = skipped = empty = failed = 0
    for run_id, folder in todo:
        parts = groups.get(folder) or {}
        if "workbook" not in parts:
            skipped += 1
            continue
        try:
            obs: list = []
            with tempfile.TemporaryDirectory() as td:
                wp = os.path.join(td, "wb.xlsx")
                with open(wp, "wb") as fh:
                    fh.write(drive.download(token, parts["workbook"].file_id))
                grains = parse_grain_summaries(wp, obs)
            n_p = len(grains.get("pillars") or [])
            n_c = len(grains.get("categories") or [])
            if not n_p and not n_c:
                # The reader still finds nothing. That is a real answer about
                # this workbook rather than a failure to record — the
                # observations say which of the three ways it came back empty.
                empty += 1
                for o in obs:
                    print(f"backfill-grains: {folder}: {o.kind} {o.detail}")
                continue
            cur.execute(
                """UPDATE run_manifest
                      SET payload = jsonb_set(payload, '{workbook_grains}',
                                              %s::jsonb, true)
                    WHERE run_id = %s""",
                (json.dumps(grains), run_id))
            conn.commit()
            print(f"backfill-grains: {folder} -> {n_p} pillar(s), "
                  f"{n_c} categor(ies)")
            filled += 1
        except Exception as exc:  # noqa: BLE001 — one bad workbook must not sink the pass
            conn.rollback()
            failed += 1
            print(f"backfill-grains FAILED: {folder}: {exc!r}")
    print(f"backfill-grains: {filled} filled, {skipped} have no workbook "
          f"artefact, {empty} state none, {failed} failed")
    return 1 if failed else 0


def backfill_sections(conn, token, groups) -> int:
    """One-time recovery for runs ingested before packages were assembled
    from the whole tree: those runs hold no document_sections because their
    report was never in the changed set. Re-parsing the report and
    inserting the sections against the EXISTING run is additive — nothing
    else in the ingested tier is touched, and no duplicate run is created.

    Matched on runs.source_folder_id, which stores the client folder's
    name. A run that already has sections is left alone (idempotent), and
    so is a folder that ships no report."""
    cur = conn.cursor()
    cur.execute("""SELECT r.id, r.source_folder_id
                     FROM runs r
                    WHERE r.source_folder_id IS NOT NULL
                      AND NOT EXISTS (SELECT 1 FROM document_sections d
                                       WHERE d.run_id = r.id)
                    ORDER BY r.source_folder_id""")
    todo = cur.fetchall()
    print(f"backfill: {len(todo)} run(s) with no report sections")
    filled = skipped = failed = 0
    for run_id, folder in todo:
        parts = groups.get(folder) or {}
        if "report" not in parts:
            skipped += 1
            continue
        try:
            with tempfile.TemporaryDirectory() as td:
                rp = os.path.join(td, "report.docx")
                with open(rp, "wb") as fh:
                    fh.write(drive.download(token, parts["report"].file_id))
                # No observation sink on the backfill path: it inserts
                # sections against an EXISTING run and writes no
                # parser_observations, so an unmapped-heading note would have
                # nowhere to land. The kinding itself is the same.
                sections = parse_report(rp)
            for sec in sections:
                cur.execute(
                    """INSERT INTO document_sections
                         (run_id, section_kind, pillar_id, heading, body, page,
                          artefact_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                    (run_id, sec.section_kind, sec.pillar_id, sec.heading,
                     sec.body, sec.page, parts["report"].file_id))
            conn.commit()
            print(f"backfill: {folder} -> {len(sections)} sections")
            filled += 1
        except Exception as exc:  # noqa: BLE001 — one bad report must not sink the pass
            conn.rollback()
            failed += 1
            print(f"backfill FAILED: {folder}: {exc!r}")
    print(f"backfill: {filled} filled, {skipped} have no report artefact, "
          f"{failed} failed")
    return 1 if failed else 0


def backfill_evidence(conn, token, groups, *, forced: bool = True,
                      only: str = "") -> int:
    """Fill in the evidence text that already-ingested rows never got.

    Evidence rows insert with ON CONFLICT DO NOTHING, so re-ingesting a
    package cannot repair one — and every row ingested before the ledger's real
    column names were understood holds a tier and nothing else. This re-parses
    each run's workbook and fills ONLY the columns that are still null:
    source_name, source_url, excerpt, claim_type. A value the package already
    stated is never overwritten, so the pass is additive and idempotent.

    Matched on runs.source_folder_id, which stores the client folder's name.
    """
    cur = conn.cursor()
    reader = evidence_reader_fingerprint()
    # THE SCHEDULED PASS IS INCREMENTAL. Unbounded, this re-downloads and
    # re-parses every run's workbook every thirty minutes for ever. A run
    # with no unlinked citation left is not work, and a run THIS reader has
    # already been through is not work until the reader changes.
    # `forced` (BACKFILL_EVIDENCE) re-reads everything, which is what a
    # human reaching for the manual pass is asking for.
    cur.execute("""SELECT r.id, r.entity_id, r.source_folder_id, r.run_seq
                     FROM runs r
                    WHERE r.source_folder_id IS NOT NULL
                      AND (%s OR (
                            -- BY ENTITY, not by run: `evidence_index` is
                            -- keyed (e_id) and scoped (entity_id) — it has
                            -- NO run_id, and asking for one aborted the
                            -- whole transaction in production on
                            -- 2026-09-04T12:13:20Z, taking the scan that
                            -- runs after it down with a 25P02. The UPDATE
                            -- below is entity-scoped for the same reason.
                            EXISTS (SELECT 1 FROM evidence_index i
                                     WHERE i.entity_id = r.entity_id
                                       AND i.source_url IS NULL)
                        AND NOT EXISTS (SELECT 1 FROM parser_observations o
                                         WHERE o.run_id = r.id
                                           AND o.kind = 'evidence_reader_pass'
                                           AND o.detail->>'reader' = %s)))
                      AND (%s = '' OR r.source_folder_id ILIKE %s)
                    ORDER BY r.source_folder_id""",
                (forced, reader, only or '', f"%{only}%"))
    todo = cur.fetchall()

    # ONE WORKBOOK PER CLIENT, NOT PER RUN. The fill is entity-scoped —
    # every UPDATE below keys on `entity_id`, because `evidence_index` has
    # no run column — so a client's second run has nothing left to give and
    # downloading its 1.5 MB workbook again to discover that is pure waste.
    # MEASURED in the first production firing, 2026-09-04T13:13:20Z:
    #   backfill-evidence: Amalgamated Bank - DMA -> 34 row(s) filled
    #   backfill-evidence: Amalgamated Bank - DMA -> 34 row(s) filled
    #   backfill-evidence: ATB - DMA -> 352 row(s) filled
    #   backfill-evidence: ATB - DMA -> 352 row(s) filled
    # two downloads and two parses each, for one client's worth of work, out
    # of a per-firing budget of five. The cap now counts DOWNLOADS, which is
    # what actually costs, and the pass is recorded against every run the
    # client has so none of them comes back asking again.
    by_client: dict = {}
    for run_id, entity_id, folder, run_seq in todo:
        by_client.setdefault((folder, entity_id), []).append(run_id)
    clients = list(by_client.items())
    deferred = 0
    if not forced and len(clients) > EVIDENCE_REPAIR_PER_FIRING:
        deferred = len(clients) - EVIDENCE_REPAIR_PER_FIRING
        clients = clients[:EVIDENCE_REPAIR_PER_FIRING]
    print(f"backfill-evidence: {len(clients)} client(s), "
          f"{sum(len(v) for _k, v in clients)} run(s) (reader {reader}"
          + (f", {only!r} only" if only else "")
          + (f", {deferred} client(s) deferred to the next firing)"
             if deferred else ")"))
    filled = skipped = failed = 0
    for (folder, entity_id), run_ids in clients:
        run_id = run_ids[0]
        parts = groups.get(folder) or {}
        if "workbook" not in parts:
            skipped += 1
            continue
        try:
            with tempfile.TemporaryDirectory() as td:
                p = os.path.join(td, "wb.xlsx")
                with open(p, "wb") as fh:
                    fh.write(drive.download(token, parts["workbook"].file_id))
                ledger = parse_evidence_master(p)
                wb = parse_scoring_workbook(p)
            mined = mine_evidence_from_rationales(wb.scores)
            n = clashes = 0
            for ev in ledger:
                m = mined.get(ev["e_id"]) or {}
                excerpt = ev.get("excerpt") or m.get("excerpt")
                # Match on ORIGIN and the id's numeric suffix, not on a
                # reconstructed entity token: the ingest takes the token from
                # the manifest, which the backfill does not read, and reading
                # one back off an arbitrary existing id picks up the
                # connector's own E-CC-nnn namespace instead. origin='package'
                # is exactly the set the ingest created from this ledger.
                suffix = ev["e_id"].split("-")[-1]
                # A savepoint per row: filling in an excerpt can collide with
                # the (entity_id, content_hash) dedup index when two ledger
                # rows cite the same url and fact. That is a real duplicate,
                # recorded and skipped — it must not sink the whole run.
                cur.execute("SAVEPOINT ev_row")
                try:
                    cur.execute(
                        """UPDATE evidence_index
                              SET source_name = COALESCE(source_name, %s),
                                  source_url  = COALESCE(source_url, %s),
                                  excerpt     = COALESCE(excerpt, %s),
                                  claim_type  = COALESCE(claim_type, %s)
                            WHERE entity_id = %s
                              AND origin = 'package'
                              AND split_part(e_id, '-', 3) = %s""",
                        (ev.get("source_name"), ev.get("source_url"), excerpt,
                         ev.get("claim_type"), entity_id, suffix))
                    n += cur.rowcount or 0
                    cur.execute("RELEASE SAVEPOINT ev_row")
                except Exception:       # noqa: BLE001 — one row, not the run
                    cur.execute("ROLLBACK TO SAVEPOINT ev_row")
                    clashes += 1
            # SECOND PASS, by the id the row itself names. The suffix join
            # above reaches only the rows the INGEST minted from this ledger;
            # a row the connector registered wears `E-CC-nnn` and states its
            # workbook origin in `source_name` instead. Read what it says.
            by_ledger_id = {ev["e_id"]: ev for ev in ledger}
            cur.execute("""SELECT e_id, source_name, excerpt
                             FROM evidence_index
                            WHERE entity_id = %s AND origin = 'package'
                              AND source_url IS NULL
                              AND source_name IS NOT NULL""", (entity_id,))
            stated = named = conflicts = uncorroborated = 0
            for e_id, source_name, served_excerpt in cur.fetchall():
                pid = stated_package_id(source_name)
                if pid is None:
                    continue
                named += 1
                ev = by_ledger_id.get(pid)
                if ev is None or not (ev.get("source_url") or "").strip():
                    continue
                # THE QUOTE HAS TO AGREE. Both rows carry the verbatim
                # passage, so the Evidence_Master excerpt is what confirms
                # the marker points where it claims. A contradiction is a
                # mis-aimed marker, and hanging a real URL under the wrong
                # quote is worse than the blank drawer it would replace.
                verdict = excerpt_agrees(served_excerpt, ev.get("excerpt"))
                if verdict is False:
                    conflicts += 1
                    continue
                if verdict is None:
                    uncorroborated += 1
                m = mined.get(pid) or {}
                cur.execute("SAVEPOINT ev_named")
                try:
                    cur.execute(
                        """UPDATE evidence_index
                              SET source_url = COALESCE(source_url, %s),
                                  excerpt    = COALESCE(excerpt, %s),
                                  claim_type = COALESCE(claim_type, %s)
                            WHERE entity_id = %s AND e_id = %s""",
                        (ev.get("source_url"),
                         ev.get("excerpt") or m.get("excerpt"),
                         ev.get("claim_type"), entity_id, e_id))
                    stated += cur.rowcount or 0
                    cur.execute("RELEASE SAVEPOINT ev_named")
                except Exception:       # noqa: BLE001 — one row, not the run
                    cur.execute("ROLLBACK TO SAVEPOINT ev_named")
                    clashes += 1
            n += stated
            if conflicts:
                # THE MARKER AND THE QUOTE DISAGREE. Recorded, never
                # resolved: whichever is wrong, a person has to look.
                cur.execute(
                    """INSERT INTO parser_observations (run_id, kind, detail, occurred_at)
                       VALUES (%s,'evidence_stated_id_excerpt_conflict',%s, now())""",
                    (run_id, json.dumps({"rows": str(conflicts),
                                         "reason": "source_name names a package "
                                                   "evidence id whose Evidence_Master "
                                                   "excerpt is a different passage; "
                                                   "no url was attached"})))
            if named and stated < named:
                # NAMED A ROW WE COULD NOT FIND. Say so rather than leaving a
                # blank drawer with no account of why it is blank.
                cur.execute(
                    """INSERT INTO parser_observations (run_id, kind, detail, occurred_at)
                       VALUES (%s,'evidence_stated_id_unresolved',%s, now())""",
                    (run_id, json.dumps({"named": str(named),
                                         "filled": str(stated),
                                         "conflicts": str(conflicts),
                                         "uncorroborated": str(uncorroborated),
                                         "reason": "source_name names a package "
                                                   "evidence id the workbook "
                                                   "ledger does not carry"})))
            if clashes:
                cur.execute(
                    """INSERT INTO parser_observations (run_id, kind, detail, occurred_at)
                       VALUES (%s,'evidence_backfill_dedup_clash',%s, now())""",
                    (run_id, json.dumps({"rows_skipped": clashes,
                                         "reason": "filling the excerpt would "
                                                   "duplicate (entity, content_hash)"})))
            conn.commit()
            # RECORDED, stamped with the reader that did it, ON EVERY RUN
            # THIS CLIENT HAS. Without this the scheduled pass re-downloads a
            # 1.5 MB workbook every thirty minutes to re-learn that it has
            # nothing left to give; with it the steady state is one query,
            # and improving the reader re-opens every run by itself.
            #
            # Every run, not just the one that did the download: the fill is
            # entity-scoped, so a sibling run genuinely has nothing left to
            # give — and stamping only the first leaves the siblings in the
            # work list, asking for the same download again next firing.
            for _rid in run_ids:
                cur.execute(
                    """INSERT INTO parser_observations
                           (run_id, kind, detail, occurred_at)
                       VALUES (%s,'evidence_reader_pass',%s, now())""",
                    (_rid, json.dumps({"reader": reader, "filled": str(n),
                                       "by_run": str(run_id)})))
            conn.commit()
            print(f"backfill-evidence: {folder} -> {n} row(s) filled "
                  f"({stated} by a stated package id"
                  + (f", {conflicts} refused on an excerpt conflict"
                     if conflicts else "") + ", "
                  f"{sum(1 for v in mined.values() if v.get('excerpt'))} mined "
                  f"excerpts{f', {clashes} dedup clash(es)' if clashes else ''})")
            filled += 1
        except Exception as exc:  # noqa: BLE001 — one bad workbook sinks nothing
            conn.rollback()
            failed += 1
            print(f"backfill-evidence FAILED: {folder}: {exc!r}")
    print(f"backfill-evidence: {filled} runs filled, {skipped} without a "
          f"workbook, {failed} failed")
    return 1 if failed else 0



def dedup_remint_links(conn) -> int:
    """Every superseded evidence row that still links where its re-mint does.

    Migration 0043 did this once, as a migration must; this is the same
    move as a re-runnable pass, because the condition is not a one-time
    state. It is produced by any code path that mints a `-R` row without
    carrying — 0043 cleaned 30,269 and the very next corpus repair created
    65,425, because the repair pass was one such path. The pass that
    creates them is fixed; a cleanup that can only run inside a migration
    would mean the next occurrence waits for a schema change.

    A link ONLY the base row holds is kept. This removes duplicates and
    never removes information.
    """
    cur = conn.cursor()
    pairs = """
      SELECT copy.e_id AS copy_id, base.e_id AS base_id
        FROM evidence_index copy
        JOIN evidence_index base
          ON base.e_id = regexp_replace(copy.e_id, '-R[0-9]+$', '')
         AND base.entity_id IS NOT DISTINCT FROM copy.entity_id
       WHERE copy.e_id ~ '-R[0-9]+$'
    """
    cur.execute(f"""
        SELECT count(*) FROM ({pairs}) p
          JOIN evidence_subcap_links k ON k.e_id = p.base_id
         WHERE EXISTS (SELECT 1 FROM evidence_subcap_links m
                        WHERE m.e_id = p.copy_id AND m.subcap_id = k.subcap_id
                          AND m.run_id = k.run_id)""")
    before = cur.fetchone()[0]
    print(f"dedup: {before} duplicated link(s) — a base row and its re-mint "
          f"both linking one cell in one run")
    if not before:
        return 0
    cur.execute(f"""
        DELETE FROM evidence_subcap_links k
         USING ({pairs}) p
         WHERE k.e_id = p.base_id
           AND EXISTS (SELECT 1 FROM evidence_subcap_links m
                        WHERE m.e_id = p.copy_id AND m.subcap_id = k.subcap_id
                          AND m.run_id = k.run_id)""")
    removed = cur.rowcount
    recounted = recount_where(
        cur, f"""sc.run_id IN (SELECT DISTINCT l.run_id
                                FROM ({pairs}) p
                                JOIN evidence_subcap_links l ON l.e_id = p.copy_id)""")
    conn.commit()
    print(f"dedup: removed {removed} duplicate base-row link(s); "
          f"linked_evidence_count recomputed on {recounted} subcap_scores row(s). "
          "Base rows are retained — an older payload's citation still resolves.")
    return 0


def backfill_evidence_urls(conn, groups, *, entity: str | None = None,
                           dry_run: bool = False) -> int:
    """Give stored evidence rows back the URL their package already states.

    For runs ingested before 2026-08-18 the ingest did not write source_url at
    all, so the drawer serves citations nobody can open — 757 of 894 on
    T. Rowe Price. The package still holds them; this joins the stored id back
    to its package-local one and fills ONLY what is NULL.

        EVIDENCE_URL_BACKFILL=1 [EVIDENCE_URL_BACKFILL_ENTITY=<display_id>]
                                [EVIDENCE_URL_BACKFILL_DRY=1]

    Dry run first, always: it prints the same counts and writes nothing.
    """
    from pathlib import Path

    from dma_worker.url_backfill import apply as _apply
    from dma_worker.url_backfill import plan as _plan
    from dma_worker.url_backfill import urls_from_package

    cur = conn.cursor()
    where, args = "", []
    if entity:
        where = " WHERE e.display_id = %s"
        args = [entity]
    cur.execute(
        "SELECT e.display_id, i.e_id, i.source_url"
        "  FROM evidence_index i JOIN entities e ON e.id = i.entity_id"
        f"{where} ORDER BY e.display_id, i.e_id", args)
    by_entity: dict = {}
    for display_id, e_id, url in cur.fetchall():
        by_entity.setdefault(display_id, []).append(
            {"e_id": e_id, "source_url": url})

    total_filled = total_unanswered = 0
    for display_id, rows in sorted(by_entity.items()):
        pkg_dir = Path(f"/root/.dma/packages/{display_id}")
        if not pkg_dir.is_dir():
            print(f"BACKFILL {display_id}: package not pulled locally, skipped "
                  f"({len(rows)} rows left as they are)")
            continue
        urls = urls_from_package(pkg_dir)
        rep = _plan(rows, urls)
        print(f"BACKFILL {display_id}: considered {rep['considered']} | "
              f"already had one {rep['already_had_one']} | "
              f"fillable {len(rep['fills'])} | "
              f"unanswered {len(rep['unanswered'])}")
        if not dry_run and rep["fills"]:
            n = _apply(cur, rep["fills"])
            conn.commit()
            print(f"BACKFILL {display_id}: filled {n}")
            total_filled += n
        total_unanswered += len(rep["unanswered"])
    print(f"BACKFILL total: filled {total_filled} | "
          f"still unanswered {total_unanswered}"
          + ("  (DRY RUN, nothing written)" if dry_run else ""))
    return 0


def repair_evidence_namespace(conn, token, groups, limit: int = 0) -> int:
    """Re-land the package evidence the id collision left unpersistable.

    The ingest qualified workbook-local ids with a token folded out of the
    institution's NAME, and a name is not an identity: 13 tokens were owned
    by more than one entity in production and one of them, `UNK` — the
    fallback for a package that ships no manifest — by 14. The second client
    to land an `E-007` hit the first client's primary key, refused (rightly)
    to alias across entities, exhausted its one `-R{run_seq}` retry and
    recorded the item as unpersistable: 5,019 items across 61 runs, leaving
    50 runs with no citable evidence at all and both Northern Trust and
    Kitsap unable to be produced.

    Migration 0036 gives the local ids an entity-scoped home, and the mint
    ladder now has an escape that cannot collide. This pass re-reads the
    source workbooks and lands what never landed, ADDITIVELY: against the
    EXISTING run, minting no new run and touching no score. A run whose
    evidence is already complete resolves every id to the row it already has
    and writes nothing — the pass is idempotent by construction, because it
    uses the same lander the ingest does.

    What it cannot do: `completed_at` is still null for a package that
    states no date anywhere, so those rows still band UNVERIFIED. That is
    honest (invariant 9) and is fixed by the package shipping a manifest,
    not here.
    """
    from dma_worker.evidence_ids import EvidenceLander

    cur = conn.cursor()
    cur.execute(
        """SELECT r.id, r.entity_id, r.source_folder_id, r.run_seq,
                  r.completed_at, m.payload -> 'manifest'
             FROM runs r LEFT JOIN run_manifest m ON m.run_id = r.id
            WHERE r.source_folder_id IS NOT NULL
              AND (EXISTS (SELECT 1 FROM parser_observations o
                            WHERE o.run_id = r.id
                              AND o.kind IN ('evidence_unpersistable',
                                             'evidence_id_collision',
                                             'evidence_conflict_unresolved'))
                   OR NOT EXISTS (SELECT 1 FROM evidence_index e
                                   WHERE e.entity_id = r.entity_id
                                     AND e.origin = 'package')
                   -- The third arm: the evidence LANDED and cannot be cited.
                   -- One package's 127 rows all carried null excerpts because
                   -- the parser knew one generation's excerpt format and read
                   -- the other as empty; with the parser fixed, a re-read
                   -- re-mints fuller rows. Keyed on the MAPPING TARGETS so a
                   -- repaired run stops being selected: after the repair the
                   -- local ids point at the -R rows that carry the excerpts.
                   OR EXISTS (SELECT 1 FROM evidence_package_ids m
                               JOIN evidence_index e ON e.e_id = m.e_id
                              WHERE m.entity_id = r.entity_id
                                AND e.excerpt IS NULL))
            ORDER BY r.source_folder_id, r.run_seq""")
    todo = cur.fetchall()
    print(f"repair: {len(todo)} run(s) carry a collision or hold no evidence")
    repaired = skipped = failed = 0
    landed_total = links_total = 0
    for run_id, entity_id, folder, run_seq, completed_at, manifest in todo:
        if limit and repaired >= limit:
            print(f"repair: stopping at the {limit}-run bound")
            break
        parts = groups.get(folder) or {}
        if "workbook" not in parts:
            skipped += 1
            continue
        try:
            with tempfile.TemporaryDirectory() as td:
                wb_path = os.path.join(td, "wb.xlsx")
                with open(wb_path, "wb") as fh:
                    fh.write(drive.download(token, parts["workbook"].file_id))
                companion: list = []
                ledger = parse_evidence_master(wb_path, companion)
                wb = parse_scoring_workbook(wb_path)
                research = {}
                if "research" in parts:
                    rw = os.path.join(td, "research.xlsx")
                    with open(rw, "wb") as fh:
                        fh.write(drive.download(token, parts["research"].file_id))
                    research = parse_research_workbook(rw, companion)

            n_obs = [0]

            def _observe(kind, detail, _rid=run_id, _n=n_obs):
                cur.execute(
                    """INSERT INTO parser_observations (run_id, kind, detail, occurred_at)
                       VALUES (%s,%s,%s, now())""",
                    (_rid, kind, json.dumps({"pass": "evidence_namespace_repair",
                                             **detail})))
                _n[0] += 1

            # The same token the ingest used, read back off the run's own
            # retained manifest — never re-derived from somewhere else, or
            # the repair would mint under a different name than the ingest.
            tok = persist._entity_token(manifest or {})
            lander = EvidenceLander(cur, entity_id=entity_id, run_id=run_id,
                                    run_seq=run_seq, token=tok,
                                    reference_date=completed_at,
                                    observe=_observe)
            merged = _merge_ledger(ledger, research, wb)
            alias, before = {}, len(lander.landed)
            for ev in merged:
                alias[ev["e_id"]] = lander.land(ev)
            links = _write_links(cur, run_id, alias, merged, research, wb)
            # The re-mint carry, which `persist_package` does and this pass
            # did not. A repair that re-reads a source with a fuller excerpt
            # mints `-R<seq>`; without this, the base row keeps its links
            # AND the mint gets its own, so one document votes twice in
            # every count that reads them. Measured after the first
            # corpus-wide run of this pass: 65,425 duplicated links, having
            # cleaned 30,269 an hour earlier. `carry_links_across_remint`
            # carries what the write missed and removes the base copies.
            carried = 0
            for minted, superseded in sorted(lander.superseded.items()):
                carried += persist.carry_links_across_remint(
                    cur, superseded, minted)
            if lander.superseded:
                _observe("evidence_remint_links_carried",
                         {"pairs": len(lander.superseded), "carried": carried})
            recount_run(cur, run_id)
            landed = len(lander.landed) - before
            _observe("evidence_namespace_repaired",
                     {"items_read": len(merged), "rows_landed": landed,
                      "links_written": links, "token": tok})
            conn.commit()
            repaired += 1
            landed_total += landed
            links_total += links
            print(f"repair: {folder} run {run_id} -> {landed} row(s), "
                  f"{links} link(s) from {len(merged)} ledger item(s)")
        except Exception as exc:  # noqa: BLE001 — one bad package sinks nothing
            conn.rollback()
            failed += 1
            print(f"repair FAILED: {folder}: {exc!r}")
    print(f"repair: {repaired} run(s) repaired ({landed_total} evidence rows, "
          f"{links_total} links), {skipped} without a workbook, {failed} failed")
    return 1 if failed else 0


def _merge_ledger(ledger, research, wb):
    """The Evidence_Master rows, enriched exactly as the ingest enriches
    them: the research workbook's ERS/date/verbatim passage win where the
    master is blank, and a mined rationale fragment fills what is still
    empty. One merge, so the repair reads the same item the ingest would."""
    rledger = {}
    for item in ((research or {}).get("ledger") or []):
        rledger.setdefault(item["e_id"], item)
    mined = mine_evidence_from_rationales(wb.scores)
    out = []
    for ev in (ledger or []):
        r = rledger.get(ev["e_id"]) or {}
        for field in ("ers", "published_date", "stated_recency", "tier",
                      "fact_count", "claim_type"):
            if r.get(field) is not None and ev.get(field) is None:
                ev = {**ev, field: r[field]}
        # The LONGER copy of a name or URL wins wherever the two ledgers
        # disagree. Both are display-truncated at source — Evidence_Master
        # hard-caps Source_Name at 40 and URL at 50 (measured: 119 of 127
        # names and 89 of 127 URLs at exactly the cap on one package), the
        # research matrix at 60/80 — so "longer" is never a different value,
        # it is more of the same one, and a URL cut mid-path resolves for
        # nobody while its 80-char copy is complete for 89 of 127.
        for field in ("source_name", "source_url"):
            if r.get(field) and len(r[field]) > len(ev.get(field) or ""):
                ev = {**ev, field: r[field]}
        if r.get("excerpt"):
            ev = {**ev, "excerpt": r["excerpt"]}
        m = mined.get(ev["e_id"]) or {}
        if not ev.get("excerpt") and m.get("excerpt"):
            ev = {**ev, "excerpt": m["excerpt"]}
        if not ev.get("subcaps") and m.get("subcaps"):
            ev = {**ev, "subcaps": m["subcaps"]}
        out.append(ev)
    return out


def _write_links(cur, run_id, alias, merged, research, wb) -> int:
    """The three link bases the ingest writes, for the rows this pass landed.
    ON CONFLICT DO NOTHING throughout: a link the run already carries keeps
    its own basis."""
    written = 0

    def link(e_id, subcap_id, basis):
        nonlocal written
        cur.execute(
            """INSERT INTO evidence_subcap_links (e_id, subcap_id, run_id, link_basis)
               VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
            (e_id, subcap_id, run_id, basis))
        written += cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0

    for ev in merged:
        resolved = alias.get(ev["e_id"])
        if resolved:
            for sid in ev.get("subcaps", []):
                link(resolved, sid, "package")
    for l in ((research or {}).get("links") or []):
        for e_id in l.get("e_ids", []):
            resolved = alias.get(e_id)
            if resolved:
                link(resolved, l["subcap_id"], "research_workbook")
    for s in wb.scores:
        for e_id in s.evidence_refs:
            resolved = alias.get(e_id)
            if resolved:
                link(resolved, s.subcap_id, "score_row")
    return written


def dump_headers(token, groups, needle: str) -> int:
    """Print every tab's first non-empty rows for one package's workbook.

    A diagnostic, not a pipeline step: the corpus does not standardise column
    names any more than it standardises file names, and a parser cannot be
    made tolerant of spellings nobody has read. Prints the first three rows of
    each tab so a header row that sits under a title row is still visible.
    """
    import openpyxl
    matches = [k for k in groups if needle.lower() in k.lower()
               and "workbook" in groups[k]]
    if not matches:
        print(f"dump: no package folder matching {needle!r}")
        return 1
    folder = sorted(matches)[0]
    print(f"dump: {folder}")
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, "wb.xlsx")
        with open(p, "wb") as fh:
            fh.write(drive.download(token, groups[folder]["workbook"].file_id))
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        try:
            print(f"dump: tabs = {wb.sheetnames}")
            for name in wb.sheetnames:
                ws = wb[name]
                print(f"--- {name}")
                for i, row in enumerate(
                        ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
                    cells = [str(v).strip() for v in row if v is not None
                             and str(v).strip()]
                    if cells:
                        print(f"    r{i}: {cells[:22]}")
        finally:
            wb.close()
    return 0


def report_intake_status(conn, tree) -> int:
    """Name every intake folder that is not yet serving, and why.

    The routine's first question — "is there anything to synthesise, and is
    anything stuck?" — had no answer anywhere in the system: the folder list
    lives in Drive and the progress lives in Postgres, and nothing joined
    them. INTAKE_STATUS=1 (or =json) prints that join.
    """
    artefacts = {k: set(v) for k, v in _package_groups(tree).items()}
    states = intake_status.intake_status(conn, sorted(artefacts), artefacts)
    if (os.environ.get("INTAKE_STATUS") or "").lower() == "json":
        print(intake_status.as_json(states))
    else:
        print(intake_status.render(states))
    return 0


def main() -> int:
    intake = os.environ["INTAKE_FOLDER_ID"]
    limit = int(os.environ.get("MAX_PACKAGES", "3"))

    # The header dump touches Drive and nothing else, so it runs before the
    # connection and before the scan lock. Taking the lock for a read-only
    # diagnostic just means it loses a race with the Scheduler and reports
    # nothing.
    if os.environ.get("DUMP_HEADERS"):
        print(f"scan: walking intake tree {intake}")
        tree = drive.walk_tree(intake)
        return dump_headers(drive.metadata_token(), _package_groups(tree),
                            os.environ["DUMP_HEADERS"])

    conn = _connect()

    # EVIDENCE_NAMESPACE=report — the package-evidence id namespace, measured.
    # Read-only and lock-free: it is a question about what is already stored,
    # not a firing, and it must be answerable while a scan is running.
    # EVIDENCE_NAMESPACE_RUNS=<uuid,uuid> probes named runs end to end.
    # LINK_DEDUP=1 — remove the duplicate links a re-mint without a carry
    # leaves behind. Lock-free and idempotent: it is a repair of stored
    # state, not a firing, and it must be runnable while a scan holds the
    # scan lock (the pass that creates the duplicates is the scan's own).
    if os.environ.get("LINK_DEDUP"):
        rc = dedup_remint_links(conn)
        conn.close()
        return rc

    if os.environ.get("EVIDENCE_NAMESPACE") == "report":
        from dma_worker.evidence_ids import namespace_report
        probes = [x.strip() for x in
                  (os.environ.get("EVIDENCE_NAMESPACE_RUNS") or "").split(",")
                  if x.strip()]
        print(json.dumps(namespace_report(conn, probes), indent=1, default=str))
        conn.close()
        return 0

    # One scan at a time: the Scheduler fires every 30 minutes and manual
    # executions overlap it. The session-level lock releases when this
    # connection closes (or the container dies) — a second execution
    # exits clean instead of racing the diff into duplicate runs.
    #
    # Which kind of pass this is decides which lock it takes, so it is
    # settled BEFORE the lock rather than beside the scan row below.
    diagnostic = bool(os.environ.get("BACKFILL_SECTIONS")
                      or os.environ.get("BACKFILL_GRAINS")
                      or os.environ.get("BACKFILL_COMPOSITE")
                      or os.environ.get("BACKFILL_WBMETA")
                      or os.environ.get("BACKFILL_EVIDENCE")
                      or os.environ.get("EVIDENCE_NAMESPACE") == "repair")

    # A DIAGNOSTIC pass takes a DIFFERENT lock (815003).
    #
    # It used to take 815002, the scan's own, which made every manual
    # backfill a coin toss against a job that fires every thirty minutes:
    # measured 2026-09-03, `BACKFILL_WBMETA` lost twice in a row and printed
    # "another execution holds the scan lock; exiting" — a clean exit that
    # reads exactly like a completed pass and wrote nothing. On a repair the
    # operator is watching, that is the difference between "done" and
    # "silently did nothing".
    #
    # Sharing the lock was never necessary. A backfill updates columns of
    # runs that already exist and are already NULL there; the scan creates
    # new runs from changed artefacts and writes the checksums. They do not
    # contend for a row. What DOES need serialising is two diagnostics
    # against each other — both re-read workbooks and update the same
    # manifests — and 815003 gives them exactly that.
    lock_id = 815003 if diagnostic else 815002
    what = "backfill" if diagnostic else "scan"
    cur = conn.cursor()
    cur.execute("SELECT pg_try_advisory_lock(%s)", (lock_id,))
    if not cur.fetchone()[0]:
        print(f"{what}: another execution holds the {what} lock; exiting")
        conn.close()
        return 0

    # W2 — evidence→subcap link proposals: an ON-DEMAND pass for the
    # scheduled session, never part of the scan flow. LINK_PROPOSE_RUN_ID=
    # <run uuid> (plus LINK_PROPOSE_DRY_RUN=1 to report without writing)
    # runs the propose-only matcher against that run and exits without
    # touching the intake tree. CLI equivalent:
    # python -m dma_worker.link_propose --run-id <uuid>.
    if os.environ.get("LINK_PROPOSE_RUN_ID"):
        from dma_worker.link_propose import run_from_env
        rc = run_from_env(conn)
        conn.close()
        return rc

    if os.environ.get("RESET_SCAN"):
        # One-time recovery: blank every stored checksum so the whole tree
        # rescans as CHANGED. Rows are kept — FKs may point at them, and
        # already-ingested artefacts re-persist as a new run only if their
        # packages truly reprocess.
        cur = conn.cursor()
        cur.execute("UPDATE import_files SET checksum = ''")
        conn.commit()
        print(f"RESET_SCAN: blanked {cur.rowcount} stored checksums; "
              "full tree rescans this execution")

    # Read-only pass over the intake tree and the ingested tier: what has
    # not been processed, by folder name and with the blocker named. Runs
    # before the scan row is opened — it is a question, not a firing.
    if os.environ.get("INTAKE_STATUS"):
        print(f"scan: walking intake tree {intake}")
        rc = report_intake_status(conn, drive.walk_tree(intake))
        conn.close()
        return rc

    # A backfill is not a scan and must not write a scan row; everything
    # else from here on is one, and gets its row BEFORE the walk so a walk
    # that dies is recorded rather than invisible (the 403 that killed the
    # tree recursion left no import_scans row at all).
    started_at = datetime.now(timezone.utc)
    scan_id = None if diagnostic else open_scan(conn, started_at)
    tally = {"ingested": 0, "failed": 0, "deferred": 0, "quarantined": []}
    try:
        print(f"scan: walking intake tree {intake}")
        tree = drive.walk_tree(intake)
        print(f"scan: {len(tree)} files")

        # Assemble packages from the WHOLE tree, and let the diff decide which
        # folders to process. Assembling from the changed set only meant a
        # folder whose workbook changed but whose report did not lost its
        # report — no run has ever landed its twelve report sections — and a
        # folder with a changed report but an unchanged workbook looked
        # incomplete on every firing.
        key = package_key(tree)
        groups = _package_groups(tree, key)
        for name, ids in key.collisions.items():
            print(f"folder-name collision: {name!r} is {len(ids)} distinct "
                  f"folders ({', '.join(ids)}); each ingests under its own "
                  "id-qualified key rather than one overwriting the other")

        # Backfill runs BEFORE the scan and returns: run_scan stores the new
        # checksums, so a backfill executed after it would swallow that
        # firing's diff and the changed packages would never be ingested.
        if os.environ.get("BACKFILL_SECTIONS"):
            rc = backfill_sections(conn, drive.metadata_token(), groups)
            conn.close()
            return rc

        if os.environ.get("BACKFILL_GRAINS"):
            rc = backfill_grains(conn, drive.metadata_token(), groups)
            conn.close()
            return rc

        if os.environ.get("BACKFILL_WBMETA"):
            rc = backfill_workbook_metadata(conn, drive.metadata_token(), groups)
            conn.close()
            return rc

        if os.environ.get("BACKFILL_COMPOSITE"):
            # The MANUAL pass: re-read every null-composite run, including the
            # ones already recorded as stating none. A human reaching for this
            # is asking for a full re-read, not the incremental one.
            rc = backfill_composite(conn, drive.metadata_token(), groups)
            conn.close()
            return rc

        if os.environ.get("BACKFILL_EVIDENCE"):
            # The manual pass is unbounded by design, so it takes the same
            # client filter: `BACKFILL_EVIDENCE="Golden 1"` repairs that
            # client and reads no other workbook. A bare truthy value still
            # means the whole corpus — 99 runs and as many downloads — which
            # is a thing to ask for on purpose, not by default.
            want = os.environ["BACKFILL_EVIDENCE"].strip()
            only = "" if want.lower() in ("1", "true", "yes", "all") else want
            rc = backfill_evidence(conn, drive.metadata_token(), groups,
                                   only=only)
            conn.close()
            return rc

        if os.environ.get("EVIDENCE_URL_BACKFILL"):
            rc = backfill_evidence_urls(
                conn, groups,
                entity=os.environ.get("EVIDENCE_URL_BACKFILL_ENTITY") or None,
                dry_run=bool(os.environ.get("EVIDENCE_URL_BACKFILL_DRY")))
            conn.close()
            return rc

        if os.environ.get("EVIDENCE_NAMESPACE") == "repair":
            rc = repair_evidence_namespace(
                conn, drive.token_provider(), groups,
                limit=int(os.environ.get("EVIDENCE_NAMESPACE_LIMIT", "0")))
            conn.close()
            return rc

        # ── THE REPAIR RUNS ON THE SCHEDULE, NOT ON SOMEONE REMEMBERING ──
        #
        # `runs.composite` is written exactly once, at INSERT. Every path that
        # could repair a NULL was behind an env var no schedule sets, so a run
        # ingested under an older reader kept its null for ever: the package
        # scan is idempotent, so an unchanged tree re-reads nothing.
        #
        # MEASURED 2026-09-04. goeasy Ltd. (`DMA-RES-GSY-20260830-0002`) served
        # `overall: null` in the client directory beside its own four pillar
        # bars — 2.09 / 2.19 / 2.01 / 2.16 — all of which resolved. Its
        # workbook was last read by this Job at 04:14:58 on 2026-09-03; the
        # reader that can find its composite merged at 06:08 the same day. The
        # run was ingested under the older reader and nothing ever looked
        # again. `backfill_composite` had three tests proving it works and not
        # one proving it runs, and no worker firing has ever logged a line
        # from it.
        #
        # So it runs here, incrementally: null composites only, minus the ones
        # this reader has already found nothing in, capped per firing. The
        # steady state is one query and no downloads. It is deliberately NOT
        # fatal to the scan — a repair that cannot reach Drive must not stop
        # this Job doing the thing it exists to do.
        # "NON-FATAL" HAS TO MEAN IT. Catching the exception is only half:
        # a failed statement leaves PostgreSQL's transaction aborted, and
        # every command after it dies `25P02 current transaction is aborted`
        # — including the scan this Job exists to run. MEASURED
        # 2026-09-04T12:13:20Z: one bad column name in the evidence work
        # list, and `_scan_and_ingest` fell over on its first SELECT. Roll
        # back, then carry on.
        def _repair(label, fn):
            try:
                fn()
            except Exception as exc:        # noqa: BLE001 — see above
                print(f"{label} skipped this firing: {exc!r}")
                try:
                    conn.rollback()
                except Exception:           # noqa: BLE001 — nothing left to do
                    pass

        # FIRST: a run that does not know its own package cannot be repaired
        # by anything below. goeasy's eighteen runs had no source folder at
        # all, which is why two rounds of fixing the folder LOOKUP moved
        # nothing.
        _repair("orphan adoption",
                lambda: adopt_orphan_runs(conn, drive.metadata_token(), groups))
        _repair("composite repair",
                lambda: backfill_composite(conn, drive.metadata_token(),
                                           groups, forced=False))
        # THE EVIDENCE DRAWER'S LINKS — FOR THE CLIENT NAMED, AND NO OTHER.
        #
        # Golden 1 Credit Union served 728 citations with 193 URLs while
        # Baxter served 154 of 154; the missing ones were in the package's
        # own `Evidence_Detail` tab. The first firing that could run this
        # pass reported the true size of that: `99 run(s)` across the whole
        # corpus, and went off repairing 1st Security Bank, Amalgamated and
        # ATB — none of which anybody had asked about. Owner's instruction,
        # 2026-09-04: strictly Golden 1, do not add clients.
        #
        # So the pass does NOTHING unless `EVIDENCE_REPAIR_ONLY` names a
        # client. Unset is off — not "on for everyone" — because the corpus
        # is 99 runs and a repair nobody asked for is still work nobody
        # asked for. `infra/deploy.sh` sets it; emptying it turns the pass
        # off, it does not widen it.
        only = (os.environ.get("EVIDENCE_REPAIR_ONLY") or "").strip()
        if only:
            _repair("evidence repair",
                    lambda: backfill_evidence(conn, drive.metadata_token(),
                                              groups, forced=False, only=only))
        else:
            print("evidence repair: EVIDENCE_REPAIR_ONLY names no client, "
                  "so nothing is repaired this firing")

        _scan_and_ingest(conn, scan_id, tree, groups, started_at, limit, tally,
                         key)
    except BaseException as exc:            # noqa: BLE001 — record, then re-raise
        # Anything that escapes — the Drive walk that 403'd mid-recursion, the
        # empty-tree guard, a diff that could not commit — closes the scan row
        # as failed with the exception naming itself. This is the whole point:
        # a scan row must never say `succeeded` about an execution that raised.
        traceback.print_exc()
        try:
            finish_scan(conn, scan_id, status=SCAN_FAILED,
                        error=f"{type(exc).__name__}: {exc}"[:2000],
                        runs_created=tally["ingested"])
        except Exception as inner:          # noqa: BLE001
            print(f"scan: could not record the failure: {inner!r}")
        conn.close()
        if isinstance(exc, Exception):
            return 1
        raise

    # The traversal completed. It "succeeded" only if nothing inside it
    # failed: a package that raised is a failed firing, named in the row.
    reasons = []
    if tally["failed"]:
        reasons.append(f"{tally['failed']} package(s) failed to ingest")
    if tally["quarantined"]:
        reasons.append("quarantined: " + ", ".join(sorted(tally["quarantined"])))
    finish_scan(conn, scan_id,
                status=SCAN_FAILED if tally["failed"] else SCAN_SUCCEEDED,
                error="; ".join(reasons)[:2000] or None,
                runs_created=tally["ingested"])
    conn.close()
    print(f"done: {tally['ingested']} ingested, {tally['failed']} failed, "
          f"{tally['deferred']} deferred, {len(tally['quarantined'])} quarantined")
    return 1 if tally["failed"] else 0


def _scan_and_ingest(conn, scan_id, tree, groups, started_at, limit, tally,
                     key=None) -> None:
    """The diff and the ingest loop. Counters land in `tally` as they happen,
    so a mid-loop exception still reports the runs it had already created."""
    summary = run_scan(conn, tree, started_at, scan_id=scan_id)
    print(f"scan: new={summary['files_new']} changed={summary['files_changed']}")
    key = key or package_key(tree)
    touched = {key(f) for f in summary["to_process"]}
    # Re-ingest one named folder even though its tree has not changed. The
    # diff is the right default — an unchanged tree must create nothing — but
    # after a parser fix the tree is unchanged and the extraction is not, and
    # RESET_SCAN is too blunt: it rescans every client and mints runs for
    # whichever three the bound happens to reach. Substring match, so the
    # folder's display name is enough.
    force = (os.environ.get("FORCE_FOLDER") or "").strip()
    forced: set = set()
    if force:
        matched = {k for k in groups if force.lower() in k.lower()}
        forced = set(matched)
        if matched:
            touched |= matched
            print(f"FORCE_FOLDER={force!r}: re-ingesting {sorted(matched)}")
        else:
            print(f"FORCE_FOLDER={force!r}: no folder matched; nothing forced")
    # The scoring workbook is what makes a package ingestable; the manifest
    # and report are enriching, not gating.
    packages = {k: v for k, v in groups.items() if "workbook" in v and k in touched}
    partial = {k: v for k, v in groups.items()
               if k in touched and "workbook" not in v
               and any(a in v for a in ("manifest", "report"))}
    if not packages and not partial:
        print("scan: nothing to ingest (unchanged tree creates nothing)")
        return

    token = drive.token_provider()
    encoder = None
    for folder, parts in sorted(packages.items()):
        if tally["ingested"] >= limit:
            _requeue(conn, parts, folder, "over per-execution bound")
            tally["deferred"] += 1
            continue
        print(f"ingest: {folder}")
        try:
            # A named re-ingest is a deliberate re-read after a parser fix
            # and mints a run. Everything else is the scan retrying itself,
            # and an unchanged package must resolve to the run it already
            # produced rather than duplicate it.
            res, rationales = _ingest_one(conn, token, folder, parts,
                                          remint=folder in forced)
        except Exception as exc:  # noqa: BLE001 — one bad package must not sink the batch
            conn.rollback()
            tally["failed"] += 1
            traceback.print_exc()
            if not _record_package_failure(conn, parts, folder, exc):
                tally["quarantined"].append(folder)
            continue
        print(f"ingest: {folder} -> run {res.run_id} "
              f"({res.scored_cells} cells, {res.observations} observations)")
        tally["ingested"] += 1

        # Embedding failures never requeue: the run is persisted, and V4
        # abstains (recorded NOT_RUN) where centroids are thin. Requeueing
        # here would re-persist the package as a duplicate run.
        if os.environ.get("EMBED_MODEL_DIR"):
            try:
                if encoder is None:
                    from dma_worker.embed import minilm_encoder
                    encoder = minilm_encoder(os.environ["EMBED_MODEL_DIR"])
                from dma_worker.embed import embed_run
                stats = embed_run(conn, res.run_id, encoder, rationales)
                print(f"embed: {stats['embeddings']} vectors, "
                      f"{stats['centroids']} centroids")
            except Exception as exc:  # noqa: BLE001
                conn.rollback()
                print(f"embed FAILED (run kept, V4 will abstain): {folder}: {exc!r}")

    for folder, parts in sorted(partial.items()):
        # NOT requeued: the folder holds no scoring workbook at all, and
        # blanking its checksums would re-detect the same folder on every
        # firing forever. A real change in Drive puts it back in the diff.
        print(f"no workbook: {folder} — {sorted(k for k in parts if k != 'folder')} "
              "present, nothing to score")


if __name__ == "__main__":
    sys.exit(main())
