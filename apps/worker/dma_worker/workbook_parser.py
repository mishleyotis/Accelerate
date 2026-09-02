"""Scoring-workbook parser (stage 1.3) — the two fields that cannot be
backfilled originate here: the workbook cell each score was read from,
and (upstream) the artefact bytes.

Shape (Claude-DMA generation, learned from a real fixture):
- `2_Scorecard` carries a "Subcapability scorecard" section — one row per
  assessed subcap with Effective_Score, the SERVED score. Scores are READ
  from the workbook, never re-derived (Backend Schema §04); source_cell
  records exactly which cell ("2_Scorecard!S16").
- `3_Assessment` carries one row per facet (Question_ID = subcap + facet
  suffix) with per-facet score, evidence quality, E-id references and the
  assessor's rationale — the grounding detail behind each subcap row.

Three row states, three different meanings:
- scored             → a subcap_scores row with source_cell + grain ids
- attempted, unscored ("NO_EVIDENCE - ladder run" / missing Effective_Score)
                     → SKIPPED, never defaulted to zero, with a parser
                       observation (a zero renders as the lowest band and
                       is indistinguishable from a real assessment)
- toggled out        → variant cells whose facet rows are entirely empty:
                       excluded by the sub-vertical toggle cascade. Not an
                       observation — they are WHY scored_cells differs
                       from catalogue_cells.
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import openpyxl

# The clip rule is shared with the connector, not restated here. See the
# `_excerpt_clip_scan` call below for why the worker needs the CORPUS half
# of it and `register_evidence` needs the single-string half.
import sys
from pathlib import Path as _Path


def _shared_roots():
    here = _Path(__file__).resolve()
    roots = [here.parent / "shared", here.parent.parent / "shared"]
    if len(here.parents) > 3:
        roots.append(here.parents[3] / "packages" / "shared")
    return roots


for _cand in _shared_roots():
    if _cand.exists() and str(_cand) not in sys.path:
        sys.path.insert(0, str(_cand))

try:
    import excerpt_clip
except ImportError as exc:                                       # pragma: no cover
    raise ImportError(
        "excerpt_clip is not in this image — deploy.sh stages packages/shared "
        "into the worker's build context. A parser that cannot run the clip "
        "check must not read a clipped corpus in silence; that is the exact "
        "shape of MEM-0129, where nothing said the excerpts were cuts and a "
        "producer named nine vendors their own citable spans do not contain."
    ) from exc

SUBCAP_RE = re.compile(r"^P\d+C\d+(?:\.(?:\d+|[A-Z]+\d+))+$")
GRAIN_RE = re.compile(r"^(P\d+)(C\d+)\.(\d+|[A-Z]+\d+)")
# Two evidence-id families, both published by the upstream dma-assessment
# skill's column spec (`references/workbook_specification.md` §Column F):
# `E-\d{3}` for public evidence and `INT-[DOC_ABBREV]-\d{3}` for internal
# documents the client supplied. The internal form was unmatched here, so
# every INT- citation was dropped without an observation — 341 of them in
# one shipped package (ATB), whose cells then read as uncited.
EID_RE = re.compile(r"\b(?:E-[A-Z]{0,3}\d{2,4}(?::F\d+)?|INT-[A-Z0-9]{2,12}-\d{1,4}(?::F\d+)?)\b")

# The rubric the whole product bands on is M1–M5. A workbook cell outside it
# is not a score at a different scale, it is a defect: 0 renders as the lowest
# band and is indistinguishable from a real assessment of "barely started".
SCORE_MIN, SCORE_MAX = Decimal("1"), Decimal("5")


def _norm(name: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9]+", "_", str(name).strip())).strip("_").lower()


def _grain(subcap_id: str):
    m = GRAIN_RE.match(subcap_id)
    if not m:
        return None, None, None
    pillar = m.group(1)
    return pillar, pillar + m.group(2), f"{pillar}{m.group(2)}.{m.group(3)}"


def _decimal(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return "UNPARSEABLE"


@dataclass
class ParsedFacet:
    question_id: str
    facet: str
    score: Decimal | None
    source_cell: str
    evidence_quality: Decimal | None
    contradiction_status: str | None
    rationale: str | None
    evidence_refs: list


@dataclass
class ParsedScore:
    subcap_id: str
    pillar_id: str
    category_id: str
    capability_id: str
    name: str | None
    tier: str | None
    score: Decimal
    source_cell: str
    evidence_quality: Decimal | None
    confidence: str | None = None      # HIGH · MEDIUM · LOW where the workbook carries it
    facets: list = field(default_factory=list)
    evidence_refs: list = field(default_factory=list)
    rationale: str | None = None       # scorer's grounding text; embedded, never stored


@dataclass
class Observation:
    kind: str          # missing_score · unparseable_cell · column_not_found · …
    subcap_id: str
    detail: dict


def _column_not_found(tab: str, field: str, tried, headers) -> Observation:
    """The observation an unrecognised column MUST produce.

    A reader that does not recognise its input carries on as though the input
    were empty, and a null column is indistinguishable from a column of nulls.
    So every lookup that fails names the tab, the field it was looking for,
    the spellings it accepts and the headers actually present — enough for the
    next spelling to be added without reading the workbook first.
    """
    return Observation("column_not_found", None, {
        "tab": tab, "field": field, "expected_any_of": list(tried),
        "headers_present": sorted(headers)[:30]})


def _pick_col(headers: dict, names) -> int | None:
    """First alias present wins; None when the column is not there at all."""
    for n in names:
        if n in headers:
            return headers[n]
    return None


# Column spellings the shipped corpus uses, per field. Measured across the 153
# packages under the production intake tree — `scoring_rationale` alone appears
# in 35 scoring tabs across 9 clients, and matching "rationale" by PREFIX
# missed every one of them, losing the only verbatim excerpt text the
# general_dma generation carries.
_EVIDENCE_ID_KEYS = ("evidence_ids", "evidence_id", "evidence_refs",
                     "evidence_references", "e_ids", "evidence_ids_cited",
                     "citations", "evidence")
_RATIONALE_KEYS = ("rationale", "scoring_rationale", "assessor_rationale",
                   "score_rationale", "rationale_150_chars", "justification",
                   "rationale_evidence", "scoring_notes", "evidence_rationale",
                   "narrative")


@dataclass
class WorkbookParse:
    scores: list
    observations: list
    toggled_out: list           # variant cells excluded by the toggle cascade
    scored_cells: int = 0
    composite: Decimal | None = None
    composite_source_cell: str | None = None
    #: In scope for this engagement and NOT YET SCORED. A different fact from
    #: `toggled_out`, and the parser used to have nowhere to put it.
    #:
    #: AUD-0014: on a real research-stage workbook — where empty scores ARE
    #: the contract — the parser reported 0 of 49 scores and reclassified 44
    #: of them as `toggled_out`, "variant cells excluded by the toggle
    #: cascade". A FOCUSED engagement rendered as an assessment where 90% of
    #: the requested capabilities had been declared inapplicable rather than
    #: pending. Emptiness is not a scope signal; it never was.
    in_scope_unscored: list = field(default_factory=list)


#: A sub-vertical VARIANT cell ends in a lettered suffix (P1C1.3.CU1); a
#: UNIVERSAL cell ends in a number (P1C1.3.1). The catalogue's own shape, and
#: the only scope signal available from an id alone.
_VARIANT_SUFFIX = re.compile(r"\.[A-Z]{2,4}\d+$")


def _is_variant(subcap_id: str) -> bool:
    return bool(_VARIANT_SUFFIX.search(subcap_id or ""))


def _unscored_bucket(result, sid: str) -> None:
    """File an unscored, contentless row under scope, not under emptiness.

    A universal cell applies to every engagement, so an empty one is in
    scope and unscored — which is what a research-stage workbook looks like
    on every row by contract. Only a sub-vertical variant can be excluded by
    the toggle cascade, and that is what `toggled_out` has always meant
    (AUD-0014)."""
    (result.toggled_out if _is_variant(sid)
     else result.in_scope_unscored).append(sid)


def _header_map(ws, anchor: str, marker: str | None = None, max_scan: int = 30):
    """Find the header row (optionally only after a section marker row) and
    return ({normalised name: 0-based col}, first_data_row)."""
    seen_marker = marker is None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        if not seen_marker:
            if any(v and marker.lower() in str(v).lower() for v in row):
                seen_marker = True
            continue
        names = {}
        for i, v in enumerate(row):
            if v is not None and str(v).strip():
                names.setdefault(_norm(v), i)
        if _norm(anchor) in names:
            return names, r + 1
    raise ValueError(f"header row with {anchor!r} not found (marker={marker!r})")


#: The label the OVERALL row of a grain tab carries. `_PILLAR_RE` rejects it
#: — correctly, it is not a pillar — which is why the row is read here and
#: not in the grain loop.
_OVERALL_LABELS = ("overall", "total", "composite", "overall_score",
                   "weighted_total")


def _stated_overall_grain(wb):
    """The composite the workbook STATES on its own rollup tab.

    THE SILENCE THIS CLOSES. `composite` is set in exactly one place —
    `_parse_scorecard`, from the cell under "Overall Effective Score" on
    `2_Scorecard`. That tab exists only in the claude_dma generation. Every
    general_dma workbook (`P{n}_Subcap_Scoring` tabs) and every rollup-only
    one takes a different branch, so `WorkbookParse.composite` came back
    None for all of them, `runs.composite` was written NULL, and
    `serving_directory` served a header with no maturity figure on a run
    whose six pages had promoted.

    Measured on Golden 1 CU (`DMA-2026-GOLDEN1-001`, 43 tabs): the overall
    is stated FOUR times — `Pillar_Summary!C6`, `Pillar_Rollup!C6`,
    `Executive_Summary` "Overall Maturity", and again as the OVERALL row's
    weighted contribution — and no reader claimed any of them. The directory
    card rendered the word "maturity" over an empty slot while the same
    card's four pillar bars resolved.

    READ, never derived: the value is the one on the row labelled OVERALL,
    not a mean of the pillars above it. Where both tabs are present the
    first spelling in `_GRAIN_TABS["pillars"]` wins and the source cell
    records which — they agree on every package measured, and a disagreement
    must surface as two readings of one figure, not be averaged away.

    Returns `(Decimal, "Tab!C6")`, or `(None, None)` when the tab, the
    header row, the score column or the OVERALL row is absent — a workbook
    that states no composite is a fact, and inventing one from the pillars
    would be exactly the derivation the contract forbids.
    """
    present = {_tab_key(n): n for n in wb.sheetnames}
    name = next((present[_tab_key(c)] for c in _GRAIN_TABS["pillars"]
                 if _tab_key(c) in present), None)
    if name is None:
        return None, None
    ws = wb[name]
    for anchor in _GRAIN_ANCHORS["pillars"]:
        try:
            headers, first = _header_map(ws, anchor)
            break
        except ValueError:
            continue
    else:
        return None, None
    score_key = next((k for k in _GRAIN_SCORE_KEYS if k in headers), None)
    if score_key is None:
        return None, None
    label_col = next((headers[k] for k in ("pillar", "pillar_id")
                      if k in headers), None)
    if label_col is None:
        return None, None
    score_col = headers[score_key]
    for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
        label = row[label_col] if label_col < len(row) else None
        if _norm(label or "") not in _OVERALL_LABELS:
            continue
        value = _decimal(row[score_col] if score_col < len(row) else None)
        if value in (None, "UNPARSEABLE"):
            return None, None
        letter = openpyxl.utils.get_column_letter(score_col + 1)
        return value, f"{ws.title}!{letter}{r}"
    return None, None


def parse_scoring_workbook(path: str) -> WorkbookParse:
    """Two shipped generations, detected by tab set:
    - claude_dma:  2_Scorecard (Effective_Score) + 3_Assessment facets
    - general_dma: P{n}_Subcap_Scoring tabs (one row per subcap) +
                   Evidence_Master / Peer_Benchmarks (parsed separately)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "2_Scorecard" in wb.sheetnames:
            facets = _parse_assessment(wb["3_Assessment"]) if "3_Assessment" in wb.sheetnames else {}
            return _parse_scorecard(wb["2_Scorecard"], facets)
        pillar_tabs = [t for t in wb.sheetnames if _is_pillar_tab(t)]
        if pillar_tabs:
            out = _parse_pillar_scoring(wb, pillar_tabs)
            if not out.scores and not out.observations and not out.toggled_out \
                    and not out.in_scope_unscored:
                # Structurally impossible to reach silently: a workbook whose
                # pillar tabs yielded nothing at all still names itself.
                out.observations.append(Observation(
                    "workbook_yielded_nothing", None,
                    {"tabs_read": sorted(pillar_tabs),
                     "reason": "pillar-grain tabs were found and read, and no "
                               "cell, observation or toggled-out variant came "
                               "out of any of them"}))
            # The composite is stated on the rollup tab, not on a tab this
            # generation has. Read there or the header serves no figure.
            out.composite, out.composite_source_cell = _stated_overall_grain(wb)
            return out
        # Rollup-only variant: recognisably a DMA workbook (stated pillar/
        # category grains present) but carrying no subcap-grain tabs at
        # all. The package lands with zero scored cells and its stated
        # grains — recorded, never guessed, never a crash-requeue loop.
        if "Pillar_Summary" in wb.sheetnames or "Category_Detail" in wb.sheetnames:
            out = WorkbookParse(scores=[], observations=[], toggled_out=[])
            out.observations.append(Observation(
                "no_subcap_grain_tabs", None,
                {"tabs": list(wb.sheetnames)[:20],
                 "note": "rollup-only workbook: stated grains land, no cells"}))
            out.scored_cells = 0
            out.composite, out.composite_source_cell = _stated_overall_grain(wb)
            return out
        # A generation nobody has taught this parser. Raising is deliberate:
        # the caller records it and quarantines the package by name after
        # three attempts, which is louder than a run with no cells. The
        # message is the quarantine reason, so it names what was looked for.
        raise ValueError(
            "unrecognised scoring workbook generation: "
            f"tabs={wb.sheetnames}; expected one of "
            "'2_Scorecard' (claude_dma), a P<n>_* subcapability tab "
            "(general_dma), or 'Pillar_Summary'/'Category_Detail' "
            "(rollup-only)")
    finally:
        wb.close()


CONFIDENCE_WORDS = {"HIGH", "MEDIUM", "LOW"}


# Score-column names across the shipped variants, in preference order —
# post-critic beats pre-critic where a critic pass shipped both.
_SCORE_KEYS = ("score", "post_critic_score", "score_1_to_5",
               "effective_score", "final_score")
_NAME_KEYS = ("subcap_name", "subcapability", "sub_cap_name")


# Which pillar tab wins when a workbook carries more than one for the same
# pillar, most authoritative first. Matched as a substring of the lowercased
# tab name; anything unlisted ranks last, in tab order.
#
# MEASURED, not assumed. 23 of the 154 corpus workbooks are MERGED files
# carrying both the assessment's `P{n}_Subcap_Scoring` and the research
# layer's `P{n}_Scoring_Detail`, so the parser read every cell twice: 1,420
# rows for a 710-cell assessment. Which tab is authoritative was settled by
# asking the workbook: aggregating `P*_Subcap_Scoring` reproduces its own
# `Pillar_Summary` (2.13 / 2.45 / 2.08 / 2.26 against a stated 2.13 / 2.44 /
# 2.08 / 2.25), and `_Scoring_Detail` is the calculation chain behind it —
# pre-critic, intermediate, and not what the summary sheets cite.
# Most specific FIRST — every token must be reachable. `scoring_detail`
# contains `scoring`, so listing the generic one earlier would make the
# specific one dead and rank the research tab as though it were the
# assessment's.
_TAB_PRECEDENCE = ("subcap_scoring", "subcapability_scoring",
                   "scoring_detail", "detail", "scoring")


def _tab_rank(tab: str) -> int:
    low = tab.lower()
    for i, token in enumerate(_TAB_PRECEDENCE):
        if token in low:
            return i
    return len(_TAB_PRECEDENCE)


def _dedupe_scores(result: "WorkbookParse") -> None:
    """One score per cell, chosen by tab authority rather than by row order.

    23 of the 154 corpus workbooks are MERGED files carrying both the
    assessment's `P{n}_Subcap_Scoring` and the research layer's
    `P{n}_Scoring_Detail`, so this parser emitted two `ParsedScore` rows per
    cell — 1,420 for a 710-cell assessment, 12,461 redundant rows across the
    corpus.

    WHAT WAS ALREADY SAFE, stated so this is not read as a bigger fix than it
    is. `persist` deduplicates on its own (`seen_subcaps`, first row wins,
    the repeat recorded as `duplicate_subcap_row` carrying its skipped
    score), and it counts `len({s.subcap_id ...})`, so the run's stored cell
    count was never inflated and no duplicate row was ever inserted twice.

    WHAT THIS CHANGES, and it is three narrower things:

      * `WorkbookParse.scored_cells` was the raw row count, so the parse
        result reported 1,420 cells for an assessment that scored 710. Only
        logging and intake status read it, but a field that disagrees with
        the database by a factor of two is a trap for the next reader.
      * persist's first-wins is ALPHABETICAL — `sorted()` puts
        `P2_Scoring_Detail` ahead of `P2_Subcap_Scoring`, so the research
        calculation chain outranked the tab the workbook's own
        `Pillar_Summary` agrees with. Measured, the two tabs never disagree
        today (13 disagreements, all of them two rows on ONE sheet), so
        nothing served was wrong — but the precedence was accidental, and
        the next merged generation would have inherited it.
      * a benign repeat and a genuine contradiction were the same
        observation kind. They are now `superseded_duplicate` and
        `duplicate_score_disagreement`, because a workbook stating one cell
        twice with two different scores is a finding about the workbook.
    """
    best: dict = {}
    for score in result.scores:
        tab = (score.source_cell or "").split("!")[0]
        rank = _tab_rank(tab)
        prior = best.get(score.subcap_id)
        if prior is None:
            best[score.subcap_id] = (rank, score)
            continue
        prior_rank, prior_score = prior
        keep, drop = ((prior_score, score) if prior_rank <= rank
                      else (score, prior_score))
        if str(prior_score.score) != str(score.score):
            result.observations.append(Observation(
                "duplicate_score_disagreement", score.subcap_id, {
                    "kept": {"source_cell": keep.source_cell,
                             "score": str(keep.score)},
                    "dropped": {"source_cell": drop.source_cell,
                                "score": str(drop.score)},
                    "resolution": "the tab the workbook's own Pillar_Summary "
                                  "agrees with is authoritative; the other "
                                  "reading is recorded, not averaged"}))
        else:
            result.observations.append(Observation(
                "superseded_duplicate", score.subcap_id,
                {"kept": keep.source_cell, "dropped": drop.source_cell}))
        best[score.subcap_id] = (min(prior_rank, rank), keep)
    if len(best) != len(result.scores):
        # Order is meaning (rule 10): keep first-seen order, not dict order.
        seen, ordered = set(), []
        for score in result.scores:
            if score.subcap_id in seen:
                continue
            seen.add(score.subcap_id)
            ordered.append(best[score.subcap_id][1])
        result.scores = ordered


def _report_stage(result, stage) -> None:
    """One observation stating what the workbook IS, instead of one per row.

    A research-stage workbook has no scores by contract (rule 4). Reporting
    that as N missing scores buries the one fact a reader needs — which
    stage this is — under N copies of a non-defect, and AUD-0014 measured
    the alternative reading of the same emptiness: 44 of 49 in-scope rows
    declared inapplicable."""
    if stage is None:
        return
    seen = (len(result.scores) + len(result.in_scope_unscored)
            + len(result.toggled_out))
    result.observations.append(Observation(
        "workbook_stage", None, {
            "contract": str(stage.get("workbook_contract") or "unknown"),
            "scope_mode": str(stage.get("scope_mode") or "unknown"),
            "declared_selected": str(stage.get("subcaps_selected") or ""),
            "rows_seen": str(seen),
            "scored": str(len(result.scores)),
            "in_scope_unscored": str(len(result.in_scope_unscored)),
            "toggled_out": str(len(result.toggled_out)),
            "stage": ("research — column D is empty by contract"
                      if not result.scores else "assessment"),
        }))


#: Run_Metadata keys that identify a contract-v3 workbook and its stage.
_STAGE_KEYS = ("workbook_contract", "scope_mode", "subcaps_selected")


def _declared_stage(wb) -> dict | None:
    """What the workbook says about itself, or None if it does not say.

    A contract-v3 workbook carries its own engagement set and its own stage:
    `scope_mode` names how the set was chosen, `subcaps_selected` how many
    rows were seeded, and at the RESEARCH stage column D is empty on every
    row BY CONTRACT (rule 4). Reading that is what lets the parser tell an
    unscored research workbook from a broken assessment one — AUD-0014's
    whole cost was having no way to ask."""
    if "Run_Metadata" not in wb.sheetnames:
        return None
    md = {}
    for row in wb["Run_Metadata"].iter_rows(min_row=1, values_only=True):
        if row and row[0] and len(row) > 1:
            md[_norm(row[0])] = row[1]
    if not any(k in md for k in _STAGE_KEYS):
        return None
    return md


def _parse_pillar_scoring(wb, pillar_tabs) -> WorkbookParse:
    result = WorkbookParse(scores=[], observations=[], toggled_out=[])
    stage = _declared_stage(wb)
    for tab in sorted(pillar_tabs):
        ws = wb[tab]
        headers = first = None
        for anchor in ("SubCap_ID", "Sub_Cap_ID", "SubCapability_ID", "Subcap ID"):
            try:
                headers, first = _header_map(ws, anchor)
                break
            except ValueError:
                continue
        if headers is None:
            # A pillar-shaped tab with no subcapability header is a rollup
            # or a layout this parser does not read: recorded, never fatal.
            result.observations.append(Observation(
                "unrecognised_pillar_tab", None, {"tab": tab}))
            continue
        sid_col = next((headers[k] for k in
                        ("subcap_id", "sub_cap_id", "subcapability_id", "subcap")
                        if k in headers), None)
        if sid_col is None:
            result.observations.append(Observation(
                "unrecognised_pillar_tab", None, {"tab": tab}))
            continue
        score_col = next((headers[k] for k in _SCORE_KEYS if k in headers), None)
        if score_col is None:
            result.observations.append(Observation(
                "missing_score_column", None,
                {"tab": tab, "headers": sorted(headers)[:20]}))
            continue
        score_letter = openpyxl.utils.get_column_letter(score_col + 1)
        conf_col = headers.get("confidence")
        name_col = next((headers[k] for k in _NAME_KEYS if k in headers), None)
        eids_col = _pick_col(headers, _EVIDENCE_ID_KEYS)
        if eids_col is None:
            result.observations.append(
                _column_not_found(tab, "evidence_ids", _EVIDENCE_ID_KEYS, headers))
        rationale_col = _pick_col(headers, _RATIONALE_KEYS)
        if rationale_col is None:  # e.g. "Rationale (≥150 chars)"
            rationale_col = next((v for k, v in sorted(headers.items())
                                  if "rationale" in k), None)
        if rationale_col is None:
            result.observations.append(
                _column_not_found(tab, "rationale", _RATIONALE_KEYS, headers))
        id_col_header = next((k for k, v in headers.items() if v == sid_col), "?")
        seen_ids = unrecognised = 0
        samples: list = []
        for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
            def cell_at(i, _row=row):
                return _row[i] if i is not None and i < len(_row) else None
            sid = cell_at(sid_col)
            sid = str(sid).strip() if sid else None
            if not sid:
                continue
            seen_ids += 1
            if not SUBCAP_RE.match(sid):
                unrecognised += 1
                if len(samples) < 5 and sid not in samples:
                    samples.append(sid)
                continue
            pillar, category, capability = _grain(sid)
            score = _decimal(cell_at(score_col))
            cell = f"{tab}!{score_letter}{r}"
            if score == "UNPARSEABLE":
                result.observations.append(Observation(
                    "unparseable_cell", sid, {"source_cell": cell,
                                              "raw": str(cell_at(score_col))[:80]}))
                continue
            refs_raw = str(cell_at(eids_col) or "")
            rationale = (str(cell_at(rationale_col)).strip()
                         if cell_at(rationale_col) else None)
            if score is None:
                if stage is not None:
                    # The workbook declares its own scope, so an unscored row
                    # is a POSITION in a run, not a defect in one. Filed by
                    # scope; the stage is reported ONCE below, not 851 times.
                    _unscored_bucket(result, sid)
                elif refs_raw.strip() or rationale:
                    result.observations.append(Observation(
                        "missing_score", sid, {"source_cell": cell}))
                else:
                    _unscored_bucket(result, sid)
                continue
            if not (SCORE_MIN <= score <= SCORE_MAX):
                # Out of the M1–M5 rubric: never stored, never rescaled. A 0
                # banded as Activating and a 7 banded as Differentiating are
                # both real maturity on the heatmap, and neither was assessed.
                result.observations.append(Observation(
                    "score_out_of_range", sid,
                    {"source_cell": cell, "stated": str(score),
                     "rubric": f"{SCORE_MIN}–{SCORE_MAX}",
                     "resolution": "cell not scored; the workbook states a "
                                   "value outside the maturity rubric"}))
                continue
            conf = None
            if cell_at(conf_col) is not None:
                c = str(cell_at(conf_col)).strip().upper()
                conf = c if c in CONFIDENCE_WORDS else None
            result.scores.append(ParsedScore(
                subcap_id=sid, pillar_id=pillar, category_id=category,
                capability_id=capability,
                name=(str(cell_at(name_col)).strip() if cell_at(name_col) else None),
                tier=None, score=score, source_cell=cell,
                evidence_quality=None, confidence=conf,
                facets=[], evidence_refs=sorted({m.split(":")[0] for m in EID_RE.findall(refs_raw)}),
                rationale=rationale,
            ))
        if seen_ids and unrecognised == seen_ids:
            # THE silent-zero case. Two shipped packages (American Homes,
            # Wescom Financial) state CATEGORY ids — `P1C1` — in the
            # SubCap_ID column of a subcapability tab, so 1,401 populated
            # rows matched nothing and the whole workbook parsed to zero
            # scores, zero observations and zero toggled-out cells: a record
            # indistinguishable from an assessment nobody had done. A shape
            # the parser cannot read is a named refusal, never an empty tab.
            result.observations.append(Observation(
                "unrecognised_cell_id_format", None, {
                    "tab": tab, "column": id_col_header,
                    "expected": SUBCAP_RE.pattern,
                    "found_examples": samples,
                    "rows_dropped": unrecognised,
                    "reason": "every populated id on this tab failed the "
                              "subcapability id pattern; no cell could be "
                              "attributed, so none was scored"}))
    # One score per cell, BEFORE the count is taken — `scored_cells` is what
    # the directory reports as a run's coverage, so counting duplicates
    # overstates every merged workbook by the size of its second tab set.
    _dedupe_scores(result)
    result.scored_cells = len(result.scores)
    _report_stage(result, stage)
    return result


def _parse_assessment(ws) -> dict:
    headers, first = _header_map(ws, "Sub_Cap_ID")
    col = {k: headers.get(k) for k in
           ("question_id", "sub_cap_id", "facet", "facet_maturity_score",
            "evidence_quality", "contradiction_status", "assessor_rationale",
            "evidence_references")}
    per_subcap: dict[str, list[ParsedFacet]] = {}
    for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
        sid = row[col["sub_cap_id"]] if col["sub_cap_id"] is not None else None
        sid = str(sid).strip() if sid else None
        if not sid or not SUBCAP_RE.match(sid):
            continue
        def v(key):
            i = col.get(key)
            return row[i] if i is not None and i < len(row) else None
        score = _decimal(v("facet_maturity_score"))
        score_col_letter = openpyxl.utils.get_column_letter(col["facet_maturity_score"] + 1)
        refs_raw = str(v("evidence_references") or "")
        per_subcap.setdefault(sid, []).append(ParsedFacet(
            question_id=str(v("question_id") or ""),
            facet=str(v("facet") or ""),
            score=None if score == "UNPARSEABLE" else score,
            source_cell=f"3_Assessment!{score_col_letter}{r}",
            evidence_quality=None if (eq := _decimal(v("evidence_quality"))) == "UNPARSEABLE" else eq,
            contradiction_status=(str(v("contradiction_status")).strip() or None) if v("contradiction_status") else None,
            rationale=(str(v("assessor_rationale")).strip() or None) if v("assessor_rationale") else None,
            evidence_refs=EID_RE.findall(refs_raw),
        ))
    return per_subcap


def _parse_scorecard(ws, facets: dict) -> WorkbookParse:
    result = WorkbookParse(scores=[], observations=[], toggled_out=[])

    # The hero composite: the cell under "Overall Effective Score".
    label_at = None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        for i, v in enumerate(row):
            if v and _norm(v) == "overall_effective_score":
                label_at = (r, i)
        if label_at and r == label_at[0] + 1:
            i = label_at[1]
            comp = _decimal(row[i] if i < len(row) else None)
            if comp not in (None, "UNPARSEABLE"):
                result.composite = comp
                result.composite_source_cell = f"2_Scorecard!{openpyxl.utils.get_column_letter(i + 1)}{r}"

    headers, first = _header_map(ws, "Effective_Score", marker="Subcapability scorecard")
    sid_col = headers["sub_cap_id"]
    score_col = headers["effective_score"]
    eq_col = headers.get("evidence_quality")
    name_col = headers.get("sub_cap_name")
    tier_col = headers.get("tier")
    score_letter = openpyxl.utils.get_column_letter(score_col + 1)

    for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
        sid = row[sid_col] if sid_col < len(row) else None
        sid = str(sid).strip() if sid else None
        if not sid or not SUBCAP_RE.match(sid):
            continue
        pillar, category, capability = _grain(sid)
        sub_facets = facets.get(sid, [])
        score = _decimal(row[score_col] if score_col < len(row) else None)

        if score == "UNPARSEABLE":
            result.observations.append(Observation(
                "unparseable_cell", sid,
                {"source_cell": f"2_Scorecard!{score_letter}{r}",
                 "raw": str(row[score_col])[:80]}))
            continue
        if score is None:
            attempted = any(f.score is not None or f.evidence_refs or
                            "NO_EVIDENCE" in str(f.rationale or "") for f in sub_facets) or any(
                            "NO_EVIDENCE" in str(getattr(f, "rationale", "") or "") for f in sub_facets)
            # Distinguish attempted-but-unscored from toggled-out variants:
            # a toggled-out cell has NO facet content at all.
            has_any_facet_content = any(
                f.score is not None or f.evidence_refs or f.rationale for f in sub_facets)
            if has_any_facet_content or attempted:
                result.observations.append(Observation(
                    "missing_score", sid,
                    {"source_cell": f"2_Scorecard!{score_letter}{r}",
                     "facets_with_content": sum(1 for f in sub_facets
                                                if f.score is not None or f.evidence_refs)}))
            else:
                _unscored_bucket(result, sid)
            continue

        refs = sorted({e for f in sub_facets for e in f.evidence_refs})
        eq = _decimal(row[eq_col]) if eq_col is not None and eq_col < len(row) else None
        result.scores.append(ParsedScore(
            subcap_id=sid, pillar_id=pillar, category_id=category,
            capability_id=capability,
            name=(str(row[name_col]).strip() if name_col is not None and row[name_col] else None),
            tier=(str(row[tier_col]).strip() if tier_col is not None and row[tier_col] else None),
            score=score, source_cell=f"2_Scorecard!{score_letter}{r}",
            evidence_quality=None if eq == "UNPARSEABLE" else eq,
            facets=sub_facets, evidence_refs=refs,
        ))

    result.scored_cells = len(result.scores)
    return result


# ── General-DMA companion tabs ─────────────────────────────────────────

DATE_EXACT = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})(?:\D|$)")
DATE_FUZZY = re.compile(r"^(\d{4})(?:-(?:Q([1-4])|(\d{2})))?")


def parse_fuzzy_date(value):
    """'2025-07' → 2025-07-01; '2025-Q4' → quarter END (H7 rule); '2025' →
    None is NOT returned for a bare year — the year is a date at year grain,
    resolved to Jan 1 conservatively. Unparseable → None (UNVERIFIED)."""
    if value is None:
        return None
    from datetime import date, datetime
    # A DATE THAT ARRIVES AT DAY GRAIN KEEPS ITS DAY. The fuzzy expression
    # below reads only year, year-month and year-quarter, so "2026-03-15" —
    # by far the commonest form, and what openpyxl hands back for a real date
    # cell as "2026-03-15 00:00:00" — resolved to 2026-03-01 and every
    # evidence row silently aged by up to a month. Never enough to sink a
    # band on its own; always enough to make a stated date and a served one
    # disagree, which is the kind of difference nobody can explain later.
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    m = DATE_EXACT.match(str(value).strip())
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass          # "2025-02-31": the DAY is wrong, the month is not —
                          # fall through to month grain rather than to nothing
    m = DATE_FUZZY.match(str(value).strip())
    if not m:
        return None
    try:
        year = int(m.group(1))
        if m.group(2):
            q = int(m.group(2))
            month, day = 3 * q, (31, 30, 30, 31)[q - 1]
            return date(year, month, day)
        if m.group(3):
            return date(year, int(m.group(3)), 1)
        return date(year, 1, 1)
    except ValueError:
        # e.g. "2025-13" — a month that isn't one. Unparseable → None
        # (UNVERIFIED); a mangled date never sinks the package.
        return None


# The shipped corpus spells the evidence ledger's columns several ways. The
# canonical general_dma tab reads
#   Evidence_ID · Source · URL · Tier · Recency · Claim_Type · Fact_Count · SubCaps
# while earlier generations used source_name / publish_date / fact_summary /
# subcaps_supported. First alias present wins.
_EV_ALIASES = {
    "source_name": ("source_name", "source", "source_title", "publisher"),
    # `url_or_citation` is the dma-assessment skill's OWN published column
    # name (skills/dma-assessment/templates/01_evidence_index_template.json):
    # "[URL or full citation if not URL-accessible]". Packages built from that
    # template landed every evidence row with source_url NULL, so the drawer
    # showed a quote with nothing to open — while the workbook carried the
    # link the whole time, one column over under a name nothing read.
    "source_url": ("url", "source_url", "link", "url_or_citation"),
    "tier": ("tier", "evidence_tier"),
    "ers": ("ers_score", "ers"),
    # `date_published` FIRST, and the bare `date` last. Both orderings are
    # deliberate: this file's own linkage-matrix reader (below) and
    # plugins/dma-insights/scripts/evidence_normalize.py already rank a
    # publication-flavoured column above a bare `date`, because at fact grain
    # `date` carries EVENT dates — E-083's 1979 is a timeline fact, not when
    # its source was published. This reader was the one place `date_published`
    # was missing, and the cost is the highest of any alias in this table: a
    # date that does not resolve bands the row UNVERIFIED, which weights 1.0,
    # tied with ARCHIVAL. Every row of such a package looks equally worthless.
    "published": ("date_published", "publish_date", "published_date",
                  "published", "date"),
    # A separate column, and a different KIND of value: "Recency" states a
    # band word (CURRENT / RECENT / …), never a date. Kept apart so a band is
    # never parsed as a date nor a date mistaken for a band.
    "recency": ("recency", "recency_band"),
    "claim_type": ("claim_type", "claim"),
    "fact_count": ("fact_count", "facts"),
    "subcaps": ("subcaps_supported", "subcaps", "subcap_ids"),
    # Only some generations carry the verbatim text here; where they do not,
    # it is mined out of the scoring tabs' Rationale column (see below).
    #
    # ORDER IS THE CONTRACT, and it used to be backwards: `fact_summary`
    # came first, so a register carrying BOTH a real quotation column and a
    # summary served the summary. Measured 2026-08-22 on the intake tree,
    # one package holds 899 facts with both a paraphrase and an
    # `anchor_quote` and not one pair is identical — the paraphrase is the
    # assessor's sentence about the source, the quote is the source. A
    # column that NAMES a quotation now outranks every summary spelling;
    # the summaries stay last because for some generations they are the only
    # text that exists and an evidence drawer cannot ship empty.
    "excerpt": ("excerpt", "anchor_quote", "verbatim", "quote", "passage",
                "fact_summary", "summary"),
}

# The excerpt tag the scoring rationales use: "[E-012:F1] Board committees: …"
# — one fact of one evidence item, verbatim, and the only place in the
# general_dma workbook the excerpt text exists at all.
_FACT_TAG = re.compile(r"\[(E-\d+)(?::(F\d+))?\]\s*")


def _stated_band(recency, published) -> str | None:
    """The band word the workbook asserted, carried for the record only.

    The database generates recency_band from a date, and undated evidence is
    UNVERIFIED, never current (charter invariant 9) — so a package that says
    "CURRENT" without a publication date does not get to say it. The claim is
    kept so the disagreement is visible rather than lost."""
    for cand in (recency, published):
        if cand is None or str(cand).strip() == "":
            continue
        if parse_fuzzy_date(cand) is not None:
            continue                    # a real date, not a band word
        return str(cand).strip().upper()
    return None


def _pick(headers: dict, names) -> int | None:
    for n in names:
        if n in headers:
            return headers[n]
    return None


def _pick_all(headers: dict, names) -> list:
    """EVERY alias present, in alias order — not just the winner.

    `_pick` answers "which column is this field", which is the right question
    for a field with one home. It is the wrong question for the excerpt,
    because a workbook can carry TWO columns that both claim to hold one:
    `Excerpt` and `Anchor_Quote`. Measured 2026-08-22 on the intake tree, one
    package holds 899 facts with both, and not one pair is identical.

    MEM-0162 measured which is which on richwood-bank: the `Excerpt` column
    held the assessor's paraphrase and only `Anchor_Quote` was verbatim. A
    fixed order gets that wrong half the time, so the row picks by CONDITION
    instead — see `_best_excerpt`.
    """
    return [headers[n] for n in names if n in headers]


def _best_excerpt(values) -> str | None:
    """Of the excerpt-class columns this row carries, the one to keep.

    Order is the tie-break, never the rule. A clipped span outranked a whole
    one for as long as the alias order decided it, and the whole point of
    MEM-0129 is that a cut excerpt is WORSE than a short one: a producer
    reads a vendor name out of it that the citable span does not contain.

    So: the first candidate that is not a hard clip wins; if every candidate
    is clipped, the first non-empty one is kept and the corpus census names
    the package. Keeping something clipped rather than nothing is deliberate
    — an evidence drawer that ships empty tells a client less than one that
    ships a cut, and the census is what makes the cut visible.
    """
    texts = [str(v).strip() for v in values
             if v is not None and str(v).strip()
             and not str(v).strip().isdigit()]
    if not texts:
        return None
    for t in texts:
        if excerpt_clip.clause_truncated(t) is None:
            return t
    return texts[0]


# The ledger tab, under every name the corpus gives it. Measured: of the 153
# packages carrying a workbook, 15 have no `Evidence_Master` at all — eleven
# of those name the same tab `Evidence_Index`, `Evidence_Linkage_Matrix`,
# `Evidence_Linkage`, `Evidence_Detail` or `Evidence_Register`, and reading
# only the one spelling left every one of them with no evidence rows and a
# NULL linked-evidence counter instead of a computed zero.
_EV_TABS = ("Evidence_Master", "Evidence_Index", "Evidence_Register",
            "Evidence_Ledger", "Evidence_Detail", "Evidence_Linkage_Matrix",
            "Evidence_Linkage", "Evidence_Inventory")
_EV_ID_ANCHORS = ("Evidence_ID", "E_ID", "Evidence Id", "EvidenceID", "ID")
_EV_E_ID_KEYS = ("evidence_id", "e_id", "id", "evidenceid")

# What a MISSING ledger column actually costs, per field — recorded WITH the
# observation, because the misses are not equivalent and a reader triaging a
# list of them needs to know which one sank the run.
_EV_MISS_COST = {
    "e_id": "no row can be identified, so the whole ledger reads as empty",
    "source_name": "every row lands unattributed",
    "source_url": "every row lands unlinkable — a drawer shows a quote with "
                  "no source to open",
    "tier": "every row lands untiered, so thin evidence cannot be judged",
    "ers": "none: ERS is computed server-side and a sent value is ignored",
    "published": "EVERY row bands UNVERIFIED (weight 1.0, tied with ARCHIVAL) "
                 "and the whole package reads as undated",
    "recency": "the workbook's own band claim is lost; the date still governs",
    "claim_type": "FACT and HYPOTHESIS become indistinguishable",
    "fact_count": "none: grounded_on is the length of the citation list",
    "subcaps": "no row links to a subcap, so no cell drawer can cite one",
    "excerpt": "no verbatim text from the ledger — EXPECTED on general_dma, "
               "which carries it in the scoring tabs' Rationale column and is "
               "mined from there instead",
}


def _read_ev_tab(wb, tab: str, observe) -> dict | None:
    """Read ONE ledger tab into rows keyed by e_id, plus the columns it carried.

    Returns None when the tab has no locatable id column, so a workbook that
    merely *has* the tab is not mistaken for one that could be read.
    """
    ws = wb[tab]
    headers = first = None
    for anchor in _EV_ID_ANCHORS:
        try:
            headers, first = _header_map(ws, anchor)
            break
        except ValueError:
            continue
    if headers is None:
        # No recognisable ledger: the package lands without this tab
        # (links absent, counts computed zero) rather than failing
        # wholesale — but never without saying so.
        observe("evidence_ledger_header_not_found",
                {"tab": tab, "expected_any_of": list(_EV_ID_ANCHORS),
                 "reason": "the ledger tab exists and its id column could "
                           "not be located; no evidence row was read"})
        return None
    cols = {k: _pick(headers, names) for k, names in _EV_ALIASES.items()}
    cols["e_id"] = _pick(headers, _EV_E_ID_KEYS)
    # Every excerpt-class column, because a row can carry two and the
    # first alias is not always the verbatim one (MEM-0162).
    excerpt_cols = _pick_all(headers, _EV_ALIASES["excerpt"])

    rows: dict = {}
    order: list = []
    rows_seen = 0
    for row in ws.iter_rows(min_row=first, values_only=True):
        def v(key):
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None
        e_id = str(v("e_id") or "").strip()
        if e_id:
            rows_seen += 1
        if not (e_id.startswith("E-") or e_id.startswith("INT-")):
            continue
        tier = str(v("tier") or "").strip().upper()
        ers = _decimal(v("ers"))
        claim = str(v("claim_type") or "").strip().upper() or None
        facts = _decimal(v("fact_count"))
        rec = {
            "e_id": e_id,
            "source_name": (str(v("source_name")).strip()
                            if v("source_name") else None),
            "source_url": (str(v("source_url")).strip()
                           if v("source_url") else None),
            "tier": tier if tier in ("T1", "T2", "T3", "T4", "T5") else None,
            "ers": None if ers in (None, "UNPARSEABLE") else ers,
            # "Recency" ships a BAND word ("CURRENT"), not a date. A band
            # asserted without a date cannot be honoured: undated evidence
            # is UNVERIFIED, never current (charter invariant 9), so the
            # stated word is carried for the record and the date stays null.
            "published_date": parse_fuzzy_date(v("published")),
            "stated_recency": _stated_band(v("recency"), v("published")),
            "claim_type": claim if claim in
                          ("FACT", "INFERENCE", "HYPOTHESIS",
                           "CEILING_ESTIMATE") else None,
            "fact_count": None if facts in (None, "UNPARSEABLE") else int(facts),
            "excerpt": _best_excerpt(
                [row[i] if i < len(row) else None for i in excerpt_cols]),
            "subcaps": [s for s in
                        (x.strip() for x in str(v("subcaps") or "").split(","))
                        if SUBCAP_RE.match(s)],
        }
        if e_id not in rows:
            order.append(e_id)
        rows[e_id] = rec
    return {"tab": tab, "cols": cols, "rows": rows,
            "order": order, "rows_seen": rows_seen}


_EV_EMPTY = (None, "", [], {})


def parse_evidence_master(path: str, obs: list | None = None) -> list:
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # EVERY ledger tab present, not the first one that matches.
        #
        # `_EV_TABS` was written as an either/or — 15 of 153 packages have no
        # `Evidence_Master` and name the same tab something else — and reading
        # `next(...)` served that case correctly while silently losing the
        # other one: a workbook carrying `Evidence_Master` AND a richer
        # `Evidence_Detail` beside it. Those are not alternate spellings of one
        # tab, they are a thin index and the ledger it indexes, and precedence
        # order put the thin one first.
        #
        # Measured on the Golden 1 package (DMA-2026-GOLDEN1-001, 43 tabs):
        # `Evidence_Master` carries 731 rows over 8 columns and NO excerpt,
        # date or subcap column at all, while `Evidence_Detail` carries 727 of
        # those same ids over 17 columns with an excerpt on 727 of 727 — every
        # one inside the 50-500 verbatim window — a `Date_Published` on 727 and
        # `SubCap_IDs` on 723. Reading only the first tab landed 589 rows
        # excerpt-less, banded the whole package UNVERIFIED for want of a date
        # column that was one tab over, and left the producer to be refused by
        # ET-04 and CG-50 for citing evidence the workbook had all along.
        #
        # So: the first tab by precedence establishes the row set (it is the
        # widest — the 4 ids `Evidence_Detail` lacks are enrichment rows), and
        # every other present tab FILLS FIELDS IT LEFT EMPTY. A tab is never
        # allowed to overwrite a value another tab already stated; merging only
        # into holes keeps this from re-ordering anyone's evidence.
        present = [t for t in _EV_TABS if t in wb.sheetnames]
        if not present:
            observe("evidence_ledger_tab_not_found",
                    {"expected_any_of": list(_EV_TABS),
                     "tabs_present": list(wb.sheetnames)[:30],
                     "reason": "no evidence ledger tab: this package lands "
                               "with no evidence rows at all"})
            return []
        reads = [r for r in (_read_ev_tab(wb, t, observe) for t in present) if r]
        if not reads:
            return []
        primary, secondaries = reads[0], reads[1:]
        tab = primary["tab"]

        # A column counts as FOUND when any present tab carries it; the census
        # below reports only what no tab had. Reporting the primary's misses
        # alone is what made a missing `excerpt` look like a property of the
        # package rather than of which tab was opened.
        cols = dict(primary["cols"])
        supplied_by = {}
        for r in reads[1:]:
            for field, idx in r["cols"].items():
                if cols.get(field) is None and idx is not None:
                    cols[field] = idx
                    supplied_by[field] = r["tab"]

        # A COLUMN NO TAB RECOGNISED IS NAMED, ALWAYS.
        #
        # Until this loop existed, `_pick` returning None was indistinguishable
        # from a column full of blanks: every row landed with the field null,
        # the run recorded nothing, and the producer was left to explain an
        # evidence set the parser had half-read. That is how `date_published`
        # went unread across a whole generation of packages — nothing anywhere
        # said a date column had been looked for and not found.
        #
        # Each miss carries what it COSTS, because the consequences are not
        # comparable: an absent `ers` changes nothing (the server recomputes
        # it), while an absent date bands every row UNVERIFIED.
        for field, names in list(_EV_ALIASES.items()) + [("e_id", _EV_E_ID_KEYS)]:
            if cols.get(field) is None:
                miss = _column_not_found(tab, field, names, primary["cols"])
                miss.detail["consequence"] = _EV_MISS_COST[field]
                miss.detail["tabs_searched"] = [r["tab"] for r in reads]
                if obs is not None:
                    obs.append(miss)

        out = []
        merged_from: dict = {}
        for e_id in primary["order"]:
            rec = dict(primary["rows"][e_id])
            for sec in secondaries:
                other = sec["rows"].get(e_id)
                if not other:
                    continue
                for field, val in other.items():
                    if field == "e_id":
                        continue
                    if rec.get(field) in _EV_EMPTY and val not in _EV_EMPTY:
                        rec[field] = val
                        merged_from.setdefault(sec["tab"], {})
                        merged_from[sec["tab"]][field] = \
                            merged_from[sec["tab"]].get(field, 0) + 1
            out.append(rec)

        # A row only a secondary tab carries is still evidence. Appended rather
        # than dropped, and counted, because a ledger that indexes 731 of 735
        # facts is exactly the silent loss this reader exists to refuse.
        seen = set(primary["order"])
        for sec in secondaries:
            extra = [i for i in sec["order"] if i not in seen]
            if not extra:
                continue
            for e_id in extra:
                out.append(dict(sec["rows"][e_id]))
                seen.add(e_id)
            observe("evidence_ledger_rows_only_in_secondary", {
                "tab": sec["tab"], "primary_tab": tab, "rows_added": len(extra),
                "example": extra[:5],
                "reason": "these ids appear in a ledger tab the primary does "
                          "not index. They are carried rather than dropped: a "
                          "row missing from the index is still a row."})

        if merged_from:
            observe("evidence_ledger_merged", {
                "primary_tab": tab,
                "tabs_present": present,
                "filled_from": {t: dict(f) for t, f in merged_from.items()},
                "columns_supplied_by": supplied_by,
                "rows": len(out),
                "reason": "fields the primary ledger tab left empty were "
                          "filled from another ledger tab in the same "
                          "workbook. Only holes were filled; no tab "
                          "overwrote a value another had already stated."})

        if primary["rows_seen"] and not out:
            observe("evidence_ledger_ids_unrecognised", {
                "tab": tab, "rows_seen": primary["rows_seen"],
                "expected": "E-… or INT-…-…",
                "reason": "the ledger has rows and not one id was in a form "
                          "this parser recognises; no evidence was read"})

        # PER-COLUMN CENSUS, because a row count cannot see this defect.
        #
        # The loop above names a header this reader could not FIND. It says
        # nothing about a header it found whose column then landed empty on
        # every row — an alias that matched the wrong column, an index off by
        # one, values that failed to parse. Those produce a full row count, a
        # clean parse, no `column_not_found`, and a field that is null all the
        # way to a client's page.
        #
        # MEM-0006, third sighting: "a header spelling the parser does not
        # know drops a column with the row count unchanged … assert per-COLUMN
        # non-null counts after a parse, not row counts. A row count cannot
        # see this defect and never will."
        #
        # Reported, never raised. A column can be legitimately empty — a
        # package that carries no `claim_type` at all is thin, not broken —
        # so this states the denominator and lets the vetter and the run
        # decide, which is the same discipline every absence here keeps.
        # IS THIS WHOLE CORPUS A CLIP RATHER THAN A QUOTATION?
        #
        # MEM-0129 and MEM-0143, both BLOCKER, both this reader's tier. The
        # package arrived with every clause cut at exactly 140 characters and
        # joined with " | ". Three of those total 426 and pass the 50-500
        # verbatim window without a murmur, so nothing fired — and 1,960 of
        # 2,063 served evidence items across 583 of 595 cells showed a client
        # a quotation cut mid-word. `register_evidence` now refuses one at
        # the door, but PACKAGE-ORIGIN EVIDENCE NEVER GOES THROUGH THAT DOOR:
        # it arrives here. A rule enforced in one of the two places it is
        # needed is the half-fix MEM-0143 recorded the first time round,
        # where the repair covered only the ids one surface cited and left
        # the 752-record corpus the heatmap cites untouched.
        #
        # The corpus check works the width out for itself rather than being
        # told it. That matters: `_RATIONALE_KEYS` above already carries a
        # `rationale_150_chars` spelling, so a second clip width is a column
        # name in the shipped corpus, not a hypothetical, and a rule that
        # only knows 140 would walk straight past it.
        #
        # REPORTED, NOT RAISED — the same discipline every absence here
        # keeps. Refusing the package would take down every run whose only
        # evidence is clipped, which is most of the T. Rowe corpus; what the
        # vetter and the producer need is to KNOW, so the producer stops
        # citing past the cut and the package can be re-ingested whole.
        if out:
            scan = excerpt_clip.clip_signature(r.get("excerpt") for r in out)
            if scan["verdict"] == "CLIPPED":
                observe("evidence_excerpts_clause_truncated", {
                    "tab": tab, **scan,
                    "consequence":
                        "every excerpt in this ledger is a CUT, not a "
                        "quotation. A producer reading one names what fell "
                        "past the cut: measured on one register as 9 product "
                        "names present in zero of their own cited excerpts, "
                        "and repairing it against that test took the register "
                        "from 41 rows to 27 and CONFIRMED from 9 to 3. The "
                        "50-500 length window cannot see this and never "
                        "will — three clipped clauses joined by ' | ' total "
                        "426 and look healthy.",
                    "fix": "re-ingest this package's evidence with whole "
                           "spans; until then no excerpt here may be read "
                           "for anything the visible text does not itself "
                           "say."})

        if out:
            for field in list(_EV_ALIASES) + ["e_id"]:
                if cols.get(field) is None:
                    continue            # already reported as not found
                filled = sum(1 for r in out
                             if r.get(field) not in (None, "", [], {}))
                if filled == 0:
                    observe("column_mapped_but_empty", {
                        "tab": tab, "field": field,
                        "header_index": cols[field],
                        "rows_emitted": len(out), "non_null": 0,
                        "consequence": _EV_MISS_COST.get(field),
                        "reason": "a header for this field WAS recognised and "
                                  "every row still landed null. That is not a "
                                  "column this reader failed to find — it is "
                                  "one it found and read nothing from, which "
                                  "a row count cannot see. Either the "
                                  "workbook's column is genuinely empty (thin, "
                                  "not broken) or the alias matched the wrong "
                                  "column; the two look identical downstream, "
                                  "so the denominator is stated here rather "
                                  "than discovered on a page."})
        return out
    finally:
        wb.close()


# A fact-grain citation and the header the research workbook puts in front of
# an excerpt block: "[ERS: 4.20] [FACT] [E-012:F1] Source (T2, CURRENT): text".
#
# The fact suffix is OPTIONAL. One research generation writes "[E-012:F1]"
# per fact; another writes a bare "[E-008]" per item. The mandatory-suffix
# version of this expression read the second generation's 709 populated
# Evidence_Excerpt cells and produced zero fragments and zero observations
# — so 127 evidence rows landed excerpt-less, 517 cells had nothing citable
# behind them, and the run's producer was blamed for a package the parser
# had silently half-read. Measured on the Odlum Brown research workbook,
# 2026-08-09: 0 of 709 cells parsed before, 709 of 709 after.
_RW_FACT_RE = re.compile(r"\[(E-\d+)(?::(F\d+))?\]\s*")
# Ledger markup EMBEDDED in a passage — "[E-051 (T2, CURRENT): corroborates.]",
# "[CEILING: L2 …]", "[ERS: 4.0]" — is the researcher annotating, not the
# document speaking. A fragment is cut at the first such annotation: 75 of 78
# excerpts on one package carried this markup INSIDE what rendered as a
# quotation (MEM-0034), and an annotation a reader mistakes for source text
# is worse than a shorter excerpt.
_RW_ANNOTATION_RE = re.compile(
    r"\s*\[(?:E-\d+[^\]]*|CEILING\b[^\]]*|ERS\b[^\]]*|FACT\b|INFERENCE\b|"
    r"HYPOTHESIS\b|CEILING_ESTIMATE\b|T[1-5]\b[^\]]*)\]")
_RW_HEAD_RE = re.compile(r"^\s*\[ERS:\s*([\d.]+)\]\s*(?:\[([A-Z_]+)\]\s*)?")
_RW_SRC_PREFIX_RE = re.compile(r"^[^:]{0,120}?\((T[1-5]),\s*[A-Z_]+\):\s*")
# The research workbook's per-subcap tabs, under every naming convention the
# corpus uses — `P1_Scoring_Detail` in 50 workbooks, `P1_Subcap_Scoring` in 10,
# a bare `P1` in 13. Pinned to the first spelling, 85 of the 135 research
# workbooks in the intake tree yielded nothing at all: no ledger, no fact-grain
# links, no verbatim passages, no absence register, and no word about it.
def _rw_detail_tabs(sheetnames):
    return [t for t in sheetnames if _is_pillar_tab(t)]


def _rw_split_excerpt(blob) -> dict:
    """{fact_id: verbatim text} out of one Evidence_Excerpt cell.

    Each fact's passage runs to the next fact tag. The first passage carries
    the source and tier as a prefix ("BCU 2024 Annual Report (PDF) (T2,
    CURRENT): …") which is provenance, not the quotation, so it is stripped —
    an excerpt has to be what the document says, byte for byte.
    """
    if not blob:
        return {}
    text = _RW_HEAD_RE.sub("", str(blob).strip())
    hits = list(_RW_FACT_RE.finditer(text))
    out = {}
    for n, m in enumerate(hits):
        end = hits[n + 1].start() if n + 1 < len(hits) else len(text)
        # A bare "[E-008]" cell may repeat its own "[ERS: …] [FACT]" header
        # before each subsequent block; strip it wherever a block starts.
        frag = _RW_HEAD_RE.sub("", text[m.end():end].strip())
        frag = _RW_SRC_PREFIX_RE.sub("", frag).strip(" ;·")
        # …and the passage ends where the researcher starts annotating.
        cut = _RW_ANNOTATION_RE.search(frag)
        if cut:
            frag = frag[:cut.start()].strip(" ;·")
        if frag:
            # A bare item id gets F1: the consumer keys the drawer by the
            # e_id half, and an unnumbered fact is still that item's fact.
            out[f"{m.group(1)}:{m.group(2) or 'F1'}"] = frag
    return out


def parse_research_workbook(path: str, obs: list | None = None) -> dict:
    """The research workbook — the evidence tier's real authority.

    The scoring workbook's Evidence_Master carries a Fact_Count but no ERS,
    no publication date and no fact text, which is why every ingested item
    reached the drawer undated, unranked and with an excerpt scraped out of a
    rationale. This tab set carries all three, plus the per-subcap linkage at
    FACT grain and the register of searches that found nothing.

    It never supplies a score: `03_scoring_workbook` is the only authority for
    a score, and this workbook's Score columns are deliberately ignored.

    → {ledger, links, caps, absent} where `ledger` items match
    parse_evidence_master's shape so the two merge on e_id.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = {"ledger": [], "links": [], "caps": [], "absent": []}

        def tab(name, id_col):
            if name not in wb.sheetnames:
                return None, None, None
            ws = wb[name]
            try:
                headers, first = _header_map(ws, id_col)
            except ValueError:
                return None, None, None
            return ws, headers, first

        # ── the ledger: ERS, dates, fact counts, claim classes ────────────
        ws = headers = first = None
        # Linkage matrix first: it is the research generation's own ledger and
        # the only one carrying ERS and a publication date per item.
        for name in ("Evidence_Linkage_Matrix",) + _EV_TABS:
            for anchor in _EV_ID_ANCHORS:
                ws, headers, first = tab(name, anchor)
                if headers is not None:
                    break
            if headers is not None:
                break
        if headers is not None:
            for row in ws.iter_rows(min_row=first, values_only=True):
                def v(*keys, _row=row):
                    for k in keys:
                        i = headers.get(k)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                e_id = str(v("evidence_id", "e_id") or "").strip()
                if not e_id.startswith("E-"):
                    continue
                tier = str(v("tier") or "").strip().upper()
                ers = _decimal(v("ers_total", "ers", "ers_score"))
                facts = _decimal(v("fact_count", "facts"))
                claim = str(v("claim_types", "claim_type") or "").split(",")[0].strip().upper()
                out["ledger"].append({
                    "e_id": e_id,
                    "source_name": (str(v("source_name")).strip() if v("source_name") else None),
                    "source_url": (str(v("source_url", "url")).strip() if v("source_url", "url") else None),
                    "tier": tier if tier in ("T1", "T2", "T3", "T4", "T5") else None,
                    "ers": None if ers in (None, "UNPARSEABLE") else ers,
                    "published_date": parse_fuzzy_date(v("date_published", "published")),
                    "stated_recency": _stated_band(v("recency"),
                                                   v("date_published", "published")),
                    "claim_type": claim if claim in ("FACT", "INFERENCE", "HYPOTHESIS",
                                                     "CEILING_ESTIMATE") else None,
                    "fact_count": None if facts in (None, "UNPARSEABLE") else int(facts),
                    # READ THE COLUMN IF THIS GENERATION HAS ONE, and fall
                    # back to the fact-tagged detail tabs if it does not.
                    #
                    # This was a flat `None` with the comment "filled from the
                    # detail tabs below", which is true of the generation that
                    # writes its excerpts as a fact-tagged blob under
                    # `Evidence_Excerpt` on a SubCap-anchored tab — and false
                    # of the one measured here. T. Rowe's research workbook
                    # carries an `Evidence_Detail` tab anchored on Evidence_ID
                    # with PLAIN `Excerpt` and `Anchor_Quote` columns: 1,642
                    # values, the longest 480 characters. The ledger read set
                    # them all to None, the fill-in below never matched the
                    # tab shape, and the parse produced 0 excerpts out of
                    # 1,642 while reporting nothing.
                    #
                    # The consequence is what makes it worth the change rather
                    # than a note. The ingest then fell through to the SCORING
                    # workbook's `fact_summary`, which is hard-clipped at 140,
                    # and that is what reached the store and the client: 1,964
                    # of 2,063 served evidence items on the promoted heatmap
                    # cut mid-word. The 480-character source was in the
                    # package the whole time. `_best_excerpt` picks the
                    # unclipped candidate between the two columns present.
                    "excerpt": _best_excerpt(
                        [v(*_EV_ALIASES["excerpt"][:1]),
                         v("anchor_quote"), v("verbatim"), v("quote"),
                         v("passage")]),
                    # THE ALIAS TABLE, not a hand-kept pair.
                    #
                    # AUD-0067: this read `v("subcap_mappings", "subcaps")`,
                    # a two-name lookup, while the contract-v3 column is
                    # `SubCap_IDs`. The module's own `_EV_ALIASES["subcaps"]`
                    # already listed `subcap_ids` — it was used for the
                    # per-column census and not for the read, so the census
                    # reported the column PRESENT while every row lost its
                    # linkage and no observation was emitted. Measured on a
                    # generated workbook: 4 of 4 ledger rows returned
                    # `subcaps: []` while Evidence_Detail carried
                    # "P1C1.1.1, P1C1.1.5" in 4 of 4. Every fact ingested
                    # from a contract-v3 workbook arrived unlinked to any
                    # capability — the workbook's own stated failure for this
                    # join, happening by construction on every run.
                    "subcaps": [s for s in
                                (x.strip() for x in
                                 re.split(r"[,;]",
                                          str(v(*_EV_ALIASES["subcaps"],
                                                "subcap_mappings") or "")))
                                if SUBCAP_RE.match(s)],
                    "signal_direction": (str(v("signal_direction")).strip().upper()
                                         if v("signal_direction") else None),
                })

        # ── per-subcap linkage at fact grain, with its verbatim passages ──
        for name in _rw_detail_tabs(wb.sheetnames):
            ws = headers = first = None
            for anchor in ("SubCap_ID", "Sub_Cap_ID", "SubCapability_ID"):
                ws, headers, first = tab(name, anchor)
                if headers is not None:
                    break
            if headers is None:
                continue
            for row in ws.iter_rows(min_row=first, values_only=True):
                def v(*keys, _row=row):
                    for k in keys:
                        i = headers.get(k)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                sid = str(v("subcap_id") or "").strip()
                if not SUBCAP_RE.match(sid):
                    continue
                facts = [f.strip() for f in
                         re.split(r"[,;]", str(v("evidence_ids") or "")) if f.strip()]
                out["links"].append({
                    "subcap_id": sid,
                    # fact ids kept whole — "E-012:F1" is finer than "E-012"
                    # and the drawer can say which fact carried the claim
                    "fact_ids": [f for f in facts if f.upper().startswith("E-")],
                    "e_ids": sorted({f.split(":")[0] for f in facts
                                     if f.upper().startswith("E-")}),
                    "excerpts": _rw_split_excerpt(v("evidence_excerpt")),
                    "_had_excerpt_cell": bool(str(v("evidence_excerpt") or "").strip()),
                    "source_document": (str(v("source_document")).strip()
                                        if v("source_document") else None),
                    "urls": [u.strip() for u in
                             re.split(r"[,;\s]+", str(v("evidence_urls") or ""))
                             if u.strip().startswith("http")],
                    "evidence_tier": (str(v("evidence_tier")).strip().upper()
                                      if v("evidence_tier") else None),
                    "diagnostic_question": (str(v("diagnostic_question")).strip()
                                            if v("diagnostic_question") else None),
                    "caps_applied": (str(v("caps_applied")).strip()
                                     if v("caps_applied") else None),
                })

        # ── the caps log and the absence register ─────────────────────────
        # The caps log states which cells a safeguard held down and why. Its
        # id column is spelled four ways across the corpus and its score
        # columns five, so a single-spelling read returned zero caps from
        # 1,457 populated rows — an assessment whose safeguards left no trace.
        ws = headers = first = None
        for anchor in ("SubCap_ID", "Sub_Cap_ID", "SubCaps_Affected",
                       "Subcap_IDs"):
            ws, headers, first = tab("Caps_Applied_Log", anchor)
            if headers is not None:
                break
        if headers is not None:
            caps_rows = 0
            for row in ws.iter_rows(min_row=first, values_only=True):
                def v(*keys, _row=row):
                    for k in keys:
                        i = headers.get(k)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                raw = str(v("subcap_id", "sub_cap_id", "subcaps_affected",
                            "subcap_ids") or "").strip()
                if not raw:
                    continue
                caps_rows += 1
                # One row may name several cells ("P1C1.1, P1C1.2"); a cap is
                # an assertion about each of them.
                for sid in (x.strip() for x in re.split(r"[,;]", raw)):
                    if not SUBCAP_RE.match(sid):
                        continue
                    out["caps"].append({
                        "subcap_id": sid,
                        "cap_type": (str(v("cap_type")).strip() if v("cap_type") else None),
                        "cap_reason": (str(v("cap_reason", "cap_rule", "rationale",
                                             "trigger", "reason")).strip()
                                       if v("cap_reason", "cap_rule", "rationale",
                                            "trigger", "reason") else None),
                        "pre_cap_score": _num(v("pre_cap_score", "prior_score",
                                                "original_score", "score_before")),
                        "post_cap_score": _num(v("post_cap_score", "capped_score",
                                                 "cap_value", "max_score",
                                                 "ceiling", "cap_ceiling",
                                                 "score_after")),
                        "e_id": (str(v("evidence_id", "e_id", "evidence_ids",
                                       "evidence_refs")).strip()
                                 if v("evidence_id", "e_id", "evidence_ids",
                                      "evidence_refs") else None),
                    })
            if caps_rows and not out["caps"] and obs is not None:
                obs.append(Observation("caps_log_ids_unrecognised", None, {
                    "tab": "Caps_Applied_Log", "rows_seen": caps_rows,
                    "expected": SUBCAP_RE.pattern,
                    "reason": "the caps log has rows and none names a cell "
                              "this parser recognises; no cap was read"}))

        # This is the absence protocol's own ladder, already run by the
        # assessment: which cells were searched, how hard, and what the
        # highest tier reached was. Emitting a thin alert without it is
        # emitting an absence with no recorded search.
        ws, headers, first = tab("Absent_Evidence_Log", "SubCap_ID")
        if headers is not None:
            for row in ws.iter_rows(min_row=first, values_only=True):
                def v(*keys, _row=row):
                    for k in keys:
                        i = headers.get(k)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                sid = str(v("subcap_id") or "").strip()
                if not sid or not sid.upper().startswith("P"):
                    continue
                out["absent"].append({
                    "subcap_id": sid,
                    "name": (str(v("subcapability")).strip() if v("subcapability") else None),
                    "diagnostic_question": (str(v("diagnostic_question")).strip()
                                            if v("diagnostic_question") else None),
                    "search_count": (str(v("search_count")).strip() if v("search_count") else None),
                    "tiers_searched": (str(v("tiers_searched")).strip() if v("tiers_searched") else None),
                    "highest_tier_found": (str(v("highest_tier_found")).strip()
                                           if v("highest_tier_found") else None),
                    "reason": (str(v("reason")).strip() if v("reason") else None),
                    "discovery_question": (str(v("discovery_question")).strip()
                                           if v("discovery_question") else None),
                    "impact_note": (str(v("impact_note")).strip() if v("impact_note") else None),
                })

        # Hoist the best passage per evidence item onto the ledger: the
        # drawer cites an ITEM, so it needs one verbatim excerpt, and the
        # longest passage is the one that carries the claim rather than a
        # fragment of it. Floor of 50 chars is the contract's.
        best: dict = {}
        cells_with_text = sum(1 for l in out["links"] if l.get("_had_excerpt_cell"))
        for link in out["links"]:
            link.pop("_had_excerpt_cell", None)
            for fact_id, text in link["excerpts"].items():
                e_id = fact_id.split(":")[0]
                if len(text) >= 50 and len(text) > len(best.get(e_id, "")):
                    best[e_id] = text
        for item in out["ledger"]:
            if best.get(item["e_id"]):
                item["excerpt"] = best[item["e_id"]][:500]
        if obs is not None and cells_with_text and not best:
            # REF-0004's rule, applied to the one reader it missed: a reader
            # that does not recognise its input must produce a NAMED
            # observation, never an empty result. This exact silence cost a
            # promoted run its entire citable evidence base — the cells were
            # populated, the expression could not see them, and nothing said
            # so anywhere.
            obs.append(Observation("research_excerpt_format_unrecognised", None, {
                "excerpt_cells_with_text": cells_with_text,
                "fragments_parsed": 0,
                "expected": "[E-nnn:Fn] or [E-nnn] fact tags inside "
                            "Evidence_Excerpt",
                "reason": "the detail tabs carry populated Evidence_Excerpt "
                          "cells and not one parsed into a fragment; every "
                          "ledger row will land excerpt-less and nothing "
                          "downstream will be citable"}))
        if obs is not None and not any(out.values()):
            obs.append(Observation("research_workbook_yielded_nothing", None, {
                "tabs_present": list(wb.sheetnames)[:30],
                "expected_ledger_any_of": ["Evidence_Linkage_Matrix", *_EV_TABS],
                "expected_detail": "a P<n>_* subcapability tab",
                "reason": "the research workbook was read and produced no "
                          "ledger row, no fact-grain link, no cap and no "
                          "recorded absence"}))
        return out
    finally:
        wb.close()


def mine_evidence_from_rationales(scores: list) -> dict:
    """Verbatim excerpts and subcap links, mined out of the scoring rationales.

    The general_dma Evidence_Master carries a Fact_Count but no fact TEXT, so
    without this every ingested evidence row reaches the evidence drawer with
    an empty excerpt — and an evidence drawer with no excerpt is the one thing
    the product cannot ship (invariant 4 requires a verbatim excerpt of
    50–500 chars behind every citation).

    The text does exist: each subcap's Rationale opens with tagged fragments,
    `[E-012:F1] Board committees: Technology Committee (…)`, one per cited
    fact. Returns `{e_id: {"excerpt": longest fragment, "fragments": [...],
    "subcaps": [ids that cite it]}}`.

    Nothing is composed or paraphrased here — a fragment is stored exactly as
    the assessor wrote it, and the longest is chosen only because the excerpt
    column holds one.
    """
    out: dict = {}
    for s in scores or []:
        text = str(getattr(s, "rationale", "") or "")
        for e_id in getattr(s, "evidence_refs", None) or []:
            rec = out.setdefault(str(e_id), {"fragments": [], "subcaps": []})
            if s.subcap_id not in rec["subcaps"]:
                rec["subcaps"].append(s.subcap_id)
        if not text:
            continue
        # Split on the tags, keeping each tag with the text that follows it.
        marks = list(_FACT_TAG.finditer(text))
        for i, m in enumerate(marks):
            e_id = m.group(1)
            end = marks[i + 1].start() if i + 1 < len(marks) else len(text)
            frag = text[m.end():end]
            # A fragment ends at the next evidence tag OR at the next section
            # label ([MATURITY]: / [GAP]: / [CEILING]: / [SO WHAT]: …), whichever
            # comes first. Without the second cut, one fact's excerpt swallows
            # the assessor's maturity reasoning and stops being verbatim.
            cut = re.search(r"\[[A-Z][A-Z ]{2,}\]\s*:?", frag)
            if cut:
                frag = frag[:cut.start()]
            frag = frag.strip().strip(" .;·")
            # 50, not 20. A citation needs a 50-500 character span: that is
            # what the column comment states, what `register_evidence`
            # refuses outside, and what ET-04 blocks a citation for. A mined
            # 20-49 character fragment used to land, link to cells, and then
            # refuse the first producer who tried to cite it — so the miner
            # was manufacturing a defect that surfaced two stages later,
            # wearing the appearance of evidence the whole way.
            if len(frag) < 50:
                continue
            rec = out.setdefault(e_id, {"fragments": [], "subcaps": []})
            if frag not in rec["fragments"]:
                rec["fragments"].append(frag)
            if s.subcap_id not in rec["subcaps"]:
                rec["subcaps"].append(s.subcap_id)
    for e_id, rec in out.items():
        best = max(rec["fragments"], key=len) if rec["fragments"] else None
        # The column is bounded at 500 chars by the registration gate; a
        # longer fragment is truncated on a word boundary rather than dropped.
        if best and len(best) > 500:
            best = best[:500].rsplit(" ", 1)[0]
        rec["excerpt"] = best
    return out


# Pillar-grain scoring tabs, across every shipped naming convention seen in
# the corpus: P1_Subcap_Scoring · P1_Scoring_Detail · P1_Scoring · P1 ·
# P1_RIAs_Broker_Dealers. Anything starting with a pillar token is a
# candidate; these suffixes mark the tabs that are rollups, logs or
# metadata rather than subcapability rows.
_NOT_SCORING = ("rollup", "summary", "benchmark", "peer", "log", "chain",
                "metadata", "caps", "issue", "priority", "index", "check",
                "contradiction", "validation", "absent", "linkage", "revision",
                "recommendation", "weight", "taxonomy", "estimate")


def _is_pillar_tab(tab: str) -> bool:
    if not re.match(r"^P\d+($|[_ ])", tab):
        return False
    low = tab.lower()
    return not any(x in low for x in _NOT_SCORING)


_CATEGORY_RE = re.compile(r"^P\d+C\d+$")
_PILLAR_RE = re.compile(r"^P\d+$")

# The same category id, as the peer tabs actually label their rows:
# `P1C1`, `P1C1: Digital Strategy & Roadmap`, `P1C1_digital_strategy`. The
# lookahead keeps a SUBCAP id (`P1C1.1`) out — that is a different grain.
_CATEGORY_LABEL_RE = re.compile(r"^(P\d+C\d+)(?=$|[^0-9.])")


def _category_id(value) -> str | None:
    m = _CATEGORY_LABEL_RE.match(str(value or "").strip())
    return m.group(1) if m else None

# Peer_Benchmarks stat columns, under every header spelling the corpus uses.
# Anything NOT in here is a named peer institution, so a missing alias does
# not degrade gracefully — it invents a peer. Keyed by _norm() output.
_STAT_ALIASES = {
    "category": "category", "category_id": "category", "cat": "category",
    "category_name": "category_name", "name": "category_name",
    "entity_score": "entity_score", "entity": "entity_score",
    "score": "entity_score", "our_score": "entity_score",
    # `Weighted_Score` is what the v5 workbook contract calls the entity's
    # own figure on Pillar_Summary, and it was the one spelling this table
    # did not know. Measured on the Golden 1 package: the tab carries
    # Weighted_Score 2.25 and Peer_Median 3.05 side by side, `peer_median`
    # resolved and `score` did not, so the run recorded a peer median with
    # no score to compare it against — six `column_not_found` observations
    # and a grain summary that could state the gap but not the position.
    "weighted_score": "entity_score", "overall_score": "entity_score",
    # The gap column under the spelling that ships beside those two.
    "gap_to_peer": "delta", "gap_vs_peer": "delta",
    "peer_median": "median", "median": "median", "cohort_median": "median",
    "peer_p25": "p25", "p25": "p25", "q1": "p25", "percentile_25": "p25",
    "peer_p75": "p75", "p75": "p75", "q3": "p75", "percentile_75": "p75",
    "peer_min": "min", "min": "min", "peer_max": "max", "max": "max",
    "delta_vs_median": "delta", "delta": "delta", "vs_median": "delta",
    "gap": "delta", "priority": "priority", "subcap_count": "subcap_count",
    "pillar": "pillar", "level": "level", "confidence": "confidence",
    # The research engine's own fixed-width tab (contract v4). Its peers ride
    # as a PAIRED LIST in two columns rather than as one column per peer,
    # because `contract.SHEETS` compares the header row as an ordered tuple
    # and a dynamic-width sheet cannot be expressed there. Named here so
    # they are read as statistics rather than mistaken for five institutions
    # called Peer_Names, Peer_Scores, Peer_N, Source_Cell and As_Of.
    "peer_names": "peer_names", "peer_scores": "peer_scores",
    "peer_n": "peer_n", "source_cell": "note", "as_of": "note",
    "peer_basis": "note",
}


# Stat headers the corpus writes freely rather than from a fixed list:
# `Gap_vs_Median`, `Gap_to_Median`, `vs_Peer`, `Position`, `Cat_ID`,
# `Peer_Name`, `Unknown`, `Percentile_Rank`, … Measured across the corpus,
# 28 clients had at least one of these read as a named peer institution.
_STAT_PATTERNS = (
    # NOT a bare `delta*`: `Delta_Community` is a real credit union, and a
    # name-only rule that swallowed it would drop a peer instead of inventing
    # one — the same defect facing the other way.
    (re.compile(r"^(gap|diff|variance)(_|$)"), "delta"),
    (re.compile(r"^delta_(vs|to|from|median|peer)"), "delta"),
    (re.compile(r"^vs_"), "delta"),
    (re.compile(r"_vs_"), "delta"),
    (re.compile(r"^(position|rank|ranking|quartile|percentile\w*)$"), "priority"),
    (re.compile(r"^(cat|cat_id|category\w*)$"), "category"),
    (re.compile(r"^(peer_name|peer|peers|institution|entity_name|unknown|"
                r"n_a|na|blank|notes?|comments?|status|source|basis|cohort|"
                r"method\w*)$"), "note"),
)


# Words that carry no identity and must not reach an acronym: "First United
# Bank, Inc." and "First United Bank" have to produce the same key.
_ENTITY_NOISE = {"inc", "llc", "lp", "ltd", "limited", "corp", "corporation",
                 "company", "co", "na", "plc", "sa", "the", "and", "of",
                 "group", "holdings", "holding", "dma"}


def _subject_keys(names) -> set:
    """Every spelling of the SUBJECT that a peer column might be wearing.

    THE ENTITY IS NOT ITS OWN PEER. `Peer_Benchmarks` puts the subject's own
    score in a named column beside the cohort's — `FUB_Score` for First United
    Bank — and the parser, which was never told whose assessment it was
    reading, stored it as a peer institution. The client then appeared in its
    own cohort, its own score set the benchmark it was being measured against,
    and every cohort statistic computed from that column was measuring the
    subject against itself.

    So: the full normalised name, the acronym of its meaningful words, and any
    parenthesised short form the name carries ("Baxter Credit Union (BCU)").
    Conservative on purpose — refusing a REAL peer is the mirror defect, and
    it is the more expensive one, so nothing here matches on a prefix or a
    substring. Only an exact key, after a trailing score/rating suffix is
    stripped, counts.
    """
    keys = set()
    for raw in names or ():
        if raw is None or not str(raw).strip():
            continue
        s = re.sub(r"\s*-\s*DMA\s*$", "", str(raw).strip(), flags=re.I)
        for short in re.findall(r"\(([^)]{2,12})\)", s):
            if _norm(short):
                keys.add(_norm(short))
        n = _norm(s)
        if not n:
            continue
        keys.add(n)
        words = [w for w in n.split("_") if w and w not in _ENTITY_NOISE]
        acronym = "".join(w[0] for w in words)
        # Three letters minimum. A two-letter acronym collides too easily with
        # a real institution's short name, and dropping a genuine peer costs
        # more than keeping a subject column that a later gate can still see.
        if len(acronym) >= 3:
            keys.add(acronym)
    return keys


_SCORE_SUFFIX = re.compile(r"_?(?:score|scores|maturity|rating)$")


def _is_subject_column(header: str, keys: set) -> bool:
    if not keys:
        return False
    n = _norm(header)
    return n in keys or (_SCORE_SUFFIX.sub("", n) or n) in keys


def _paired_peers(names, scores, num, obs=None):
    """Peers carried as two parallel lists in one row.

    The research engine writes a FIXED-WIDTH Peer_Benchmarks (its contract
    compares the header as an ordered tuple, so it cannot grow a column per
    peer). It puts the cohort in `Peer_Names` and their figures in
    `Peer_Scores`, comma-separated and positional. A name with no figure
    yet is a real peer with a null score — which is what a frozen cohort
    looks like before the assessment stage fills it — and is kept, exactly
    as an unscored peer COLUMN is kept.
    """
    ns = [n.strip() for n in str(names or "").split(",") if n.strip()]
    if not ns:
        return []
    vs = [v.strip() for v in str(scores or "").split(",") if v.strip()]
    if vs and len(vs) != len(ns):
        # Positional lists that disagree in length cannot be zipped without
        # guessing which peer got which score, and a mis-attributed peer
        # score is worse than none.
        if obs is not None:
            obs.append(Observation("peer_paired_list_mismatch", None, {
                "tab": "Peer_Benchmarks", "names": len(ns), "scores": len(vs),
                "reason": "Peer_Names and Peer_Scores are positional lists of "
                          "different lengths; the scores were dropped and the "
                          "peers kept unscored rather than mis-attributed"}))
        vs = []
    return [(n, num(vs[i]) if i < len(vs) else None)
            for i, n in enumerate(ns)]


def _stat_key(header: str):
    """The canonical stat a header names, or None when it names a peer."""
    n = _norm(header)
    hit = _STAT_ALIASES.get(n)
    if hit:
        return hit
    for pattern, canonical in _STAT_PATTERNS:
        if pattern.search(n):
            return canonical
    return None


#: Tab spellings for the two stated grains. One literal name was read until
#: 2026-08-18 and a workbook that spells it any other way lost both grains in
#: silence. Matched case- and separator-insensitively, so `Pillar Summary`,
#: `pillar_summary` and `PillarSummary` are one name.
#: The entity's own figure, under every spelling the corpus gives it. The
#: gate below used to test `"score" not in headers` literally, so a tab
#: heading the column `Weighted_Score` — which is what the v5 workbook
#: contract calls it — lost BOTH grains and every figure on them. Measured on
#: the Golden 1 package: Pillar_Summary states Weighted_Score 2.25 beside
#: Peer_Median 3.05, `peer_median` resolved, `score` did not, and the run
#: landed pillars: 0 with a peer median it had nothing to compare against.
_GRAIN_SCORE_KEYS = ("score", "weighted_score", "entity_score", "overall_score",
               "our_score")

_GRAIN_TABS = {
    "pillars": ("Pillar_Summary", "Pillar Summary", "Pillar_Scores",
                "Pillar Scores", "Pillar_Rollup", "Pillar Rollup",
                "Pillar_Detail", "1_Pillar_Summary", "Summary_Pillar"),
    "categories": ("Category_Detail", "Category Detail", "Category_Scores",
                   "Category Scores", "Category_Scorecard",
                   "Category Scorecard", "Category_Summary",
                   "Category Summary", "Category_Rollup", "2_Category_Detail"),
}
#: The header cell each grain's table is anchored on, in preference order.
_GRAIN_ANCHORS = {
    "pillars": ("Pillar", "Pillar_ID", "Pillar ID"),
    "categories": ("Category_ID", "Category ID", "Category"),
}


def _tab_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").lower())


def parse_grain_summaries(path: str, observations: list | None = None) -> dict:
    """The workbook's own STATED pillar and category grains
    (Pillar_Summary / Category_Detail tabs, cached formula values). H4's
    grain lock forbids recomputing these by averaging subcaps — cap
    logic, weighting and analyst override are applied when they are
    struck. The rubric Level column (M1-M5) is deliberately not read:
    display banding is the app's four-band rule over raw scores.

    THE SILENCE THIS CLOSES, measured 2026-08-18 on the third client.
    Every other companion parser here takes the `observations` list and
    appends what it could not read; this one was the exception, and it had
    three distinct ways to return nothing — tab absent, header row not
    locatable, no Score column — all spelled `(None, None, None)` and none
    of them recorded. The client's run landed with `pillars: 0,
    categories: 0` against the reference client's 4 and 17, its bundle note
    still saying both grains are stated with source cells, and no row
    anywhere naming what had happened. Downstream: the workbook zoom served
    an empty state at both grains, the overview pillar bars had to derive
    means the workbook already stated, and no peer median existed at any
    grain to compare against. One unrecognised tab name, four surfaces.

    The observation is the fix that matters more than the aliases: a
    spelling nobody has met yet still loses the grains, and now it says so
    by name instead of looking like a workbook that states none.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"pillars": [], "categories": []}
    obs = observations if observations is not None else []
    present = {_tab_key(n): n for n in wb.sheetnames}

    def _tab_headers(grain: str):
        """The tab for this grain, or nothing plus the reason why nothing.

        Each branch names a DIFFERENT next move for whoever reads the
        observation: add a spelling, fix the header row, or find the score
        column. "no stated grains" means all three and therefore none.
        """
        name = next((present[_tab_key(c)] for c in _GRAIN_TABS[grain]
                     if _tab_key(c) in present), None)
        if name is None:
            obs.append(Observation(
                "grain_tab_not_found", None,
                {"grain": grain, "tried": list(_GRAIN_TABS[grain]),
                 "tabs": list(wb.sheetnames)[:30],
                 "consequence": (
                     f"no stated {grain} grain lands for this run. The "
                     "workbook zoom serves an empty state at this grain and "
                     "no peer median exists to compare against; add the "
                     "tab's spelling to _GRAIN_TABS rather than letting the "
                     "run look like a workbook that states none")}))
            return None, None, None, None
        ws = wb[name]
        for anchor in _GRAIN_ANCHORS[grain]:
            try:
                headers, first = _header_map(ws, anchor)
                break
            except ValueError:
                continue
        else:
            obs.append(Observation(
                "grain_header_not_found", None,
                {"grain": grain, "tab": name,
                 "anchors_tried": list(_GRAIN_ANCHORS[grain])}))
            return None, None, None, None
        score_key = next((k for k in _GRAIN_SCORE_KEYS if k in headers), None)
        if score_key is None:
            obs.append(_column_not_found(name, "score", _GRAIN_SCORE_KEYS, headers))
            return None, None, None, None
        return ws, headers, first, score_key

    try:
        ws, headers, first, score_key = _tab_headers("pillars")
        if headers is not None:
            for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
                def v(*keys, _row=row):
                    # first alias that resolves: the corpus spells these
                    # headers several ways and a single-alias lookup silently
                    # nulls the column (Pillar_Summary.Weight was read only
                    # as `weight_ib`, so every pillar weight was null)
                    for key in keys:
                        i = headers.get(key)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                pid = str(v("pillar", "pillar_id") or "").strip()
                if not _PILLAR_RE.match(pid):
                    continue
                score_col = openpyxl.utils.get_column_letter(headers[score_key] + 1)
                out["pillars"].append({
                    "pillar_id": pid,
                    "name": (str(v("pillar_name")).strip() if v("pillar_name") else None),
                    "score": _num(v(score_key)),
                    "weight": _num(v("weight_ib", "weight", "weight_pct")),
                    "peer_median": _num(v("peer_median", "median")),
                    "source_cell": f"{ws.title}!{score_col}{r}",
                })
        ws, headers, first, score_key = _tab_headers("categories")
        if headers is not None:
            for r, row in enumerate(ws.iter_rows(min_row=first, values_only=True), first):
                def v(*keys, _row=row):
                    for key in keys:
                        i = headers.get(key)
                        if i is not None and i < len(_row) and _row[i] is not None:
                            return _row[i]
                    return None
                cid = str(v("category_id", "category") or "").strip()
                if not _CATEGORY_RE.match(cid):
                    continue
                score_col = openpyxl.utils.get_column_letter(headers[score_key] + 1)
                out["categories"].append({
                    "category_id": cid,
                    "name": (str(v("category_name")).strip() if v("category_name") else None),
                    "pillar_id": (str(v("pillar")).strip() if v("pillar") else cid.split("C")[0]),
                    "score": _num(v(score_key)),
                    "peer_median": _num(v("peer_median", "median")),
                    "priority_score": _num(v("priority_score", "priority")),
                    "priority_tier": (str(v("priority_tier")).strip() if v("priority_tier") else None),
                    "source_cell": f"{ws.title}!{score_col}{r}",
                })
        return out
    finally:
        wb.close()


def _num(value):
    d = _decimal(value)
    return None if d in (None, "UNPARSEABLE") else float(d)


def _peer_header_row(rows: list):
    """Index of the header row, or None. The header is NOT reliably row 1:
    seven packages put a title, a methodology note or a run id above it
    (`Peer Benchmarking — METHODOLOGY NOTE`, `DMA-RES-APGFCU-… | Peer
    Benchmarks | SV2 Credit Union Medium | Maryland`), and reading row 1 as
    the header made every peer column a fragment of that sentence."""
    best, best_hits = None, 0
    for i, row in enumerate(rows[:10]):
        if not row:
            continue
        hits = sum(1 for h in row if h is not None and _stat_key(str(h)))
        named = sum(1 for h in row if h is not None and str(h).strip())
        if hits and named >= 2 and hits > best_hits:
            best, best_hits = i, hits
    return best


def parse_peer_benchmarks(path: str, obs: list | None = None,
                          subject_names=()) -> list:
    """Peer_Benchmarks is CATEGORY grain with named-peer columns after the
    stat block. Only the per-peer scores are data — Entity_Score and the
    stat columns (median/quartiles/min/max/delta) are derivable, so they
    are read solely to verify, never to store (counts are computed, never
    stored, where a source of truth exists). Stops at the footer notes.

    A column is a peer only if it BOTH fails to name a known stat and holds
    values on the maturity scale. The name test alone invents institutions:
    28 clients in the corpus carried a peer called `Gap_vs_Median`,
    `Position`, `Peer_Name`, `Cat_ID` or `Unknown`, and every one of those
    would have rendered in the cohort as a bank that does not exist.

    `subject_names` is who this assessment is ABOUT — the manifest's
    institution name and the client folder name. Without it the parser cannot
    tell the subject's own score column (`FUB_Score`) from a peer's, and the
    client joins its own cohort. Passing nothing is allowed, because a caller
    may genuinely not know, but it is RECORDED: a cohort read without knowing
    whose it is may contain the subject, and that has to be visible in the run
    rather than inferred from a page months later."""
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Peer_Benchmarks" not in wb.sheetnames:
            # AUD-0042: the category-grain peer store has no feeder, and the
            # missing-tab path RECORDED NOTHING — a package with no peer tab
            # was indistinguishable from one whose peers all parsed, so every
            # peer median downstream silently had no source to reconcile
            # against. An absence is a finding; a silent return is not.
            observe("peer_tab_absent", {
                "expected": "Peer_Benchmarks",
                "tabs_present": list(wb.sheetnames)[:30],
                "consequence": "this run has NO category-grain peer scores. "
                               "Every peer median served for it is a producer "
                               "assertion with nothing to reconcile against, "
                               "and ET-09's peer allow-list is empty, so a "
                               "legitimate peer who is also a corpus client "
                               "reads as foreign-entity contamination."})
            return []
        ws = wb["Peer_Benchmarks"]
        rows = [r for r in ws.iter_rows(values_only=True) if r]
        if not rows:
            return []
        h = _peer_header_row(rows)
        if h is None:
            observe("peer_header_not_found", {
                "tab": "Peer_Benchmarks",
                "row_1": [str(v)[:60] for v in rows[0] if v is not None][:8],
                "reason": "no row in the first ten named a recognised peer "
                          "statistic; no peer row was read"})
            return []
        header, body = rows[h], rows[h + 1:]
        candidates = [(i, str(x).strip()) for i, x in enumerate(header)
                      if x is not None and _stat_key(str(x)) is None
                      and str(x).strip()]

        def col(canonical, row):
            """The row's value for a canonical stat, under any of its aliases."""
            for i, x in enumerate(header):
                if x is not None and _stat_key(str(x)) == canonical:
                    return row[i] if i < len(row) else None
            return None
        # Non-numeric cells ("N/A", footnotes) are None here — a peer grid
        # is data or nothing, and downstream median verification sorts
        # these values.
        num = (lambda v: None if (d := _decimal(v)) in (None, "UNPARSEABLE") else d)
        graded = [r for r in body if _category_id(r[0])]
        if body and not graded:
            observe("peer_rows_unrecognised", {
                "tab": "Peer_Benchmarks", "rows_seen": len(body),
                "first_column_examples": sorted(
                    {str(r[0]).strip()[:40] for r in body if r and r[0]})[:5],
                "expected": _CATEGORY_LABEL_RE.pattern,
                "reason": "the tab has rows and none is at category grain "
                          "(a peer-per-row layout reads this way); no peer "
                          "score was read"})
            return []

        # The value test. A peer is scored on the same M1–M5 scale as the
        # entity; a column of gaps, ranks or words is a statistic or a label
        # whatever it calls itself, and is refused BY NAME rather than
        # stored as an institution.
        subject_keys = _subject_keys(subject_names)
        if candidates and not subject_keys:
            observe("peer_subject_unknown", {
                "tab": "Peer_Benchmarks",
                "candidate_columns": [n for _, n in candidates][:12],
                "reason": "the cohort was read without knowing whose "
                          "assessment this is, so the subject's own score "
                          "column cannot be told from a peer's and may have "
                          "been stored as an institution"})

        peer_cols, refused, unscored, subject_cols = [], [], [], []
        for i, name in candidates:
            # The subject is not its own peer — checked BEFORE the value test,
            # because the subject's column holds perfectly valid scores and
            # would sail through it.
            if _is_subject_column(name, subject_keys):
                subject_cols.append(name)
                continue
            values = [row[i] for row in graded if i < len(row)]
            populated = [v for v in values if v is not None and str(v).strip()]
            nums = [d for d in (num(v) for v in populated) if d is not None]
            out_of_band = [d for d in nums if not (SCORE_MIN <= d <= SCORE_MAX)]
            if not populated:
                # A peer the tab NAMES but never scores. It is a real
                # institution and it keeps its column; every score is null,
                # which is what the cohort should say about it.
                peer_cols.append((i, name))
                unscored.append(name)
            elif nums and not out_of_band:
                peer_cols.append((i, name))
            else:
                refused.append({
                    "column": name,
                    "values_seen": len(populated),
                    "numeric": len(nums),
                    "outside_rubric": [str(d) for d in out_of_band][:4],
                    "examples": [str(v)[:30] for v in populated][:4]})
        if refused:
            observe("peer_column_unrecognised", {
                "tab": "Peer_Benchmarks", "columns": refused,
                "rubric": f"{SCORE_MIN}–{SCORE_MAX}",
                "reason": "column names no known statistic and does not hold "
                          "scores on the maturity scale; not stored as a peer"})
        if subject_cols:
            observe("peer_column_is_the_subject", {
                "tab": "Peer_Benchmarks", "columns": sorted(subject_cols),
                "subject": sorted(subject_keys),
                "reason": "the column names the entity this assessment is "
                          "about; kept out of the cohort so the client is not "
                          "benchmarked against itself"})
        if unscored:
            observe("peer_column_unscored", {
                "tab": "Peer_Benchmarks", "columns": sorted(unscored),
                "reason": "named in the cohort and scored in no category; "
                          "kept with null scores, never a computed zero"})

        out = []
        for row in graded:
            cat = _category_id(row[0])
            # The name comes from the column that declares itself a name.
            # Reading row[1] positionally put a peer's SCORE in the name
            # field whenever the tab has no Category_Name column.
            cat_name = col("category_name", row)
            peers = [(name, num(row[i]) if i < len(row) else None)
                     for i, name in peer_cols]
            peers += _paired_peers(col("peer_names", row),
                                   col("peer_scores", row), num, obs)
            out.append({
                "category_id": cat,
                "category_name": (str(cat_name).strip() or None) if cat_name is not None else None,
                "entity_score": num(col("entity_score", row)),
                "stated_median": num(col("median", row)),
                "peers": peers,
            })
        return out
    finally:
        wb.close()


# Column names a Recommendations header row uses. A row is the header when it
# names at least two of them; anything above it is a title, a deferral note or
# a status line, and reading THAT as the header destroyed the tab.
_REC_HEADER_WORDS = {
    "rec_id", "recommendation", "recommendations", "title", "priority", "rank",
    "category", "category_id", "pillar", "theme", "initiative", "owner",
    "effort", "impact", "horizon", "timeline", "phase", "status", "rationale",
    "description", "subcap_id", "capability", "sequence", "value", "outcome",
    "zennify_solution", "solution", "offering", "dependency", "cost",
}


def _rec_header_row(rows: list):
    """Index of the header row, or None when the tab starts straight into
    data (`CI Segal Bryant & Hammill` opens with `1 | FSC/Sales Cloud`)."""
    for i, row in enumerate(rows[:10]):
        if not row:
            continue
        names = {_norm(str(v)) for v in row if v is not None and str(v).strip()}
        if len(names & _REC_HEADER_WORDS) >= 2:
            return i
    return None


def parse_recommendations(path: str, obs: list | None = None) -> list:
    """The Recommendations tab lands raw: rec_id as it arrived (the raw
    tier preserves package identifiers) plus the full row as payload.

    Neither the header row nor the id column can be assumed. Measured across
    the corpus, twenty packages put a title or deferral note above the header
    and 26 lost rows to a de-duplication keyed on the raw first column —
    Amarillo National Bank's 29 recommendations became 0, Cetera's 24 became
    4 — because that column repeats a priority rank or a phase label."""
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Recommendations" not in wb.sheetnames:
            return []
        ws = wb["Recommendations"]
        all_rows = [r for r in ws.iter_rows(values_only=True) if r]
        if not all_rows:
            return []
        h = _rec_header_row(all_rows)
        if h is None:
            observe("recommendations_header_not_found", {
                "tab": "Recommendations",
                "row_1": [str(v)[:60] for v in all_rows[0] if v is not None][:6],
                "expected_any_of": sorted(_REC_HEADER_WORDS)[:12],
                "reason": "no row in the first ten named two recognised "
                          "recommendation columns; columns land positionally"})
            first = ()
            rows = iter(all_rows)
        else:
            first = all_rows[h]
            rows = iter(all_rows[h + 1:])
        header = [(_norm(str(x)) if x is not None else None) for x in first]
        # The first column is a rec id in some packages and a bare priority
        # rank in others (BCU's header is `Priority` with values 1..8).
        # Requiring a REC- prefix dropped every recommendation in those
        # packages silently, which is why the platform page served none.
        # A row is a recommendation when it carries content, and the raw
        # tier keeps whatever identifier arrived — synthesising REC-n from
        # the rank where the package states no id.
        id_is_rank = header and header[0] in ("priority", "rank", "order", "no", "num")
        out = []
        seen = set()
        collided = 0
        content_rows = 0
        for n, row in enumerate(rows, 1):
            if not row:   # read-only mode yields () for blank rows
                continue
            raw = str(row[0] or "").strip()
            if not raw:
                continue
            payload = {header[i]: (str(row[i]).strip() if i < len(row) and row[i] is not None else None)
                       for i in range(len(header)) if header[i]}
            if not payload:      # no usable header: keep the row positionally
                payload = {f"col_{i + 1}": (str(v).strip() if v is not None else None)
                           for i, v in enumerate(row)}
            # a row whose only populated cell is the id is a footer or spacer
            if not any(v for k, v in payload.items()
                       if k != (header[0] if header else "col_1")):
                continue
            content_rows += 1
            rec_id = f"REC-{raw}" if id_is_rank or not raw.upper().startswith("REC-") else raw
            if rec_id in seen:
                # The id repeats; the ROW does not. Dropping it here is how
                # 26 packages lost recommendations the tab plainly listed —
                # the raw tier keeps every stated row, and an id that cannot
                # be unique is qualified by its position instead of deciding
                # which recommendation the client does not get to read.
                collided += 1
                rec_id = f"{rec_id}#{n}"
            if rec_id in seen:
                continue
            seen.add(rec_id)
            out.append({"rec_id": rec_id, "payload": payload})
        if collided:
            observe("recommendation_id_not_unique", {
                "tab": "Recommendations", "id_column": (header[0] if header else None),
                "rows_qualified_by_position": collided,
                "rows_kept": len(out), "content_rows": content_rows,
                "reason": "the first column repeats, so it is not an id; rows "
                          "are kept under a position-qualified id rather than "
                          "de-duplicated away"})
        if content_rows and not out:
            observe("recommendations_all_dropped", {
                "tab": "Recommendations", "content_rows": content_rows,
                "reason": "the tab has populated rows and none survived "
                          "parsing"})
        return out
    finally:
        wb.close()


# ── the package evidence index: 01_evidence/evidence_index.json ──────────

#: What an evidence-index item may call each field. The workbook's aliases
#: reused, plus the JSON spellings the corpus actually ships.
_EI_ALIASES = {
    "e_id": ("e_id", "evidence_id", "id"),
    "source_name": ("source_name", "source", "title", "source_title",
                    "publisher"),
    "source_url": ("url", "source_url", "link", "url_or_citation"),
    "tier": ("tier", "evidence_tier"),
    "ers": ("ers", "ers_score"),
    "published": ("date_published", "published_date", "publish_date",
                  "published", "date"),
    "recency": ("recency", "recency_band"),
    "claim_type": ("claim_type", "claim"),
    "fact_count": ("fact_count", "facts"),
    "subcaps": ("subcaps_supported", "subcaps", "subcap_ids",
                "subcap_mappings"),
    "excerpt": ("excerpt", "anchor_quote", "verbatim", "quote", "passage",
                "fact_summary", "summary"),
}


def parse_evidence_index(path: str, obs: list | None = None) -> list:
    """Read `01_evidence/evidence_index.json` into evidence rows.

    WHY THIS READER EXISTS. AUD-0091: this file is the richest evidence
    store in every package, `classification.py` has recognised it as
    `package_structured` since stage 1.1, and the scanner records it into
    `import_files.classified_kind` — and the ingest then dropped it, because
    the artefact grouper accepted only manifest.json and the Office formats.

    Gate M was built after the consequence: one client shipped with 85% of
    its evidence unURLed while this file carried 752 items and 748 URLs.

    The shape is heterogeneous across the corpus — a bare list, `{"items":
    [...]}`, `{"evidence": [...]}` — so all three are read, and anything
    else is REPORTED rather than returning an empty list that looks like an
    empty index."""
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    try:
        with open(path, encoding="utf-8-sig") as fh:
            doc = json.load(fh)
    except (OSError, ValueError) as e:
        observe("evidence_index_unreadable",
                {"path": os.path.basename(path), "error": str(e)[:200],
                 "consequence": "the package's own evidence index was not "
                                "read; rows keep whatever the workbook gave "
                                "them"})
        return []

    items = None
    if isinstance(doc, list):
        items = doc
    elif isinstance(doc, dict):
        for key in ("items", "evidence", "evidence_index", "records"):
            if isinstance(doc.get(key), list):
                items = doc[key]
                break
    if items is None:
        observe("evidence_index_shape_unrecognised",
                {"top_level": (list(doc)[:12] if isinstance(doc, dict)
                               else type(doc).__name__),
                 "expected": "a list, or an object with items/evidence/records",
                 "consequence": "no item was read from an index that exists"})
        return []

    def pick(item, field):
        for k in _EI_ALIASES[field]:
            for cand in (k, k.replace("_", ""), k.upper()):
                if isinstance(item, dict) and item.get(cand) not in (None, ""):
                    return item[cand]
        return None

    out, skipped = [], 0
    for item in items:
        if not isinstance(item, dict):
            skipped += 1
            continue
        e_id = str(pick(item, "e_id") or "").strip()
        if not (e_id.startswith("E-") or e_id.startswith("INT-")):
            skipped += 1
            continue
        tier = str(pick(item, "tier") or "").strip().upper()
        ers = _decimal(pick(item, "ers"))
        claim = str(pick(item, "claim_type") or "").strip().upper() or None
        facts = _decimal(pick(item, "fact_count"))
        subs = pick(item, "subcaps")
        if isinstance(subs, str):
            subs = re.split(r"[,;]", subs)
        out.append({
            "e_id": e_id,
            "source_name": (str(pick(item, "source_name")).strip()
                            if pick(item, "source_name") else None),
            "source_url": (str(pick(item, "source_url")).strip()
                           if pick(item, "source_url") else None),
            "tier": tier if tier in ("T1", "T2", "T3", "T4", "T5") else None,
            "ers": None if ers in (None, "UNPARSEABLE") else ers,
            "published_date": parse_fuzzy_date(pick(item, "published")),
            "stated_recency": _stated_band(pick(item, "recency"),
                                           pick(item, "published")),
            "claim_type": claim if claim in
                          ("FACT", "INFERENCE", "HYPOTHESIS",
                           "CEILING_ESTIMATE") else None,
            "fact_count": None if facts in (None, "UNPARSEABLE") else int(facts),
            "excerpt": _best_excerpt([pick(item, "excerpt")]),
            "subcaps": [s for s in (str(x).strip() for x in (subs or []))
                        if SUBCAP_RE.match(s)],
        })
    if skipped:
        observe("evidence_index_items_skipped",
                {"skipped": skipped, "kept": len(out),
                 "expected": "an object carrying an E-… or INT-… id"})
    return out


#: Fields the index may FILL when the workbook left them empty. Never a
#: field the workbook stated: the workbook is the artefact under assessment,
#: and an index that disagrees with it is a disagreement to record, not to
#: resolve in the index's favour.
_FILLABLE = ("source_url", "source_name", "published_date", "stated_recency",
             "tier", "ers", "claim_type", "fact_count", "excerpt")


def merge_evidence_sources(workbook_rows: list, index_rows: list,
                           obs: list | None = None) -> list:
    """Workbook rows first; the index fills what they left blank.

    Two rules, and the asymmetry is the point:

      * A field the workbook STATED is never overwritten. If the index
        disagrees, that is recorded as an observation and the workbook wins —
        the workbook is the artefact being assessed.
      * A field the workbook left blank is filled from the index, and the
        fill is recorded. This is the 748 URLs that were in the package the
        whole time (AUD-0091).

    An id only the index carries is ADDED, because dropping it would repeat
    the original defect one level down."""
    by_id = {r["e_id"]: dict(r) for r in workbook_rows}
    filled, disagreed, added = {}, [], 0
    for item in index_rows:
        eid = item["e_id"]
        row = by_id.get(eid)
        if row is None:
            by_id[eid] = dict(item)
            added += 1
            continue
        for f in _FILLABLE:
            have, come = row.get(f), item.get(f)
            if come in (None, ""):
                continue
            if have in (None, ""):
                row[f] = come
                filled[f] = filled.get(f, 0) + 1
            elif str(have).strip() != str(come).strip():
                disagreed.append({"e_id": eid, "field": f,
                                  "workbook": str(have)[:120],
                                  "index": str(come)[:120]})
        merged = {s for s in (row.get("subcaps") or [])} | \
                 {s for s in (item.get("subcaps") or [])}
        if merged != set(row.get("subcaps") or []):
            row["subcaps"] = sorted(merged)
    if obs is not None:
        if filled or added:
            obs.append(Observation("evidence_index_merged", None, {
                "filled": {k: str(v) for k, v in sorted(filled.items())},
                "items_only_in_index": str(added),
                "reason": "the package's evidence index carried values the "
                          "workbook ledger left blank"}))
        if disagreed:
            obs.append(Observation("evidence_index_disagreement", None, {
                "count": str(len(disagreed)),
                "examples": disagreed[:5],
                "resolution": "the workbook's value is kept; the index's is "
                              "recorded, never averaged and never preferred"}))
    return list(by_id.values())


# ── the technographic scan: the package's fourth final output ────────────

def parse_technographic_scan(path: str, obs: list | None = None) -> int:
    """Read the scan's machine copy and record its shape as observations.

    The scan's DETECTIONS reach the serving tier through the connector's
    techstack producers; the worker's job here is narrower and honest —
    record that the scan arrived, what it holds, and which layers it never
    looked at, so a producer reading get_run_progress can tell 'clean
    estate' from 'never scanned' (the AUD-0115 distinction) without opening
    the file. Returns the detection count.

    A .docx (the human copy arriving without its sidecar) is recorded as
    exactly that — present, unparsed, and the sidecar named as what is
    missing — never silently skipped."""
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    if str(path).lower().endswith(".docx"):
        observe("technographic_scan_docx_only", {
            "file": os.path.basename(path),
            "expected_sidecar": "technographic_scan.json",
            "consequence": "the scan's detections are not machine-readable "
                           "from the document alone; the package shipped the "
                           "human copy without the sidecar the assembler "
                           "writes beside it"})
        return 0
    try:
        with open(path, encoding="utf-8-sig") as fh:
            doc = json.load(fh)
    except (OSError, ValueError) as e:
        observe("technographic_scan_unreadable", {
            "file": os.path.basename(path), "error": str(e)[:200]})
        return 0
    detections = doc.get("detections") or []
    by_status: dict = {}
    by_layer: dict = {}
    for d in detections:
        by_status[str(d.get("status"))] = by_status.get(str(d.get("status")), 0) + 1
        by_layer[str(d.get("layer"))] = by_layer.get(str(d.get("layer")), 0) + 1
    never = ((doc.get("counts") or {}).get("layers_never_looked_at")
             or [l for l in ("OPS", "CUST", "DATA", "INFRA")
                 if l not in by_layer])
    observe("technographic_scan_summary", {
        "run_id": str(doc.get("run_id") or ""),
        "detections": str(len(detections)),
        "by_status": {k: str(v) for k, v in sorted(by_status.items())},
        "by_layer": {k: str(v) for k, v in sorted(by_layer.items())},
        "layers_never_looked_at": never,
        "note": ("layers in layers_never_looked_at were NOT SCANNED — that "
                 "is a gap in the scan, not a clean estate, and nothing may "
                 "be read as ABSENT there"),
    })
    return len(detections)


# ── the technology register, at the grain the techstack contract asks for ──
#
# `parse_technographic_scan` above reads the Technographic_Scan DOCX. The
# scoring workbook carries the same estate as a TAB, one row per product with
# every field the T1/T3 contract names — and nothing read it. Measured on the
# Golden 1 package: `Tech_Register` holds 42 rows over 14 columns
# (TS_ID, Product, Vendor, Layer, Status, Evidence_Level, Detection_Basis,
# Detection_Method, Providers, SubCap_IDs, Evidence_IDs, Source_URLs, As_Of,
# DMA_Impact), 42 of 42 carrying both SubCap_IDs and Evidence_IDs, statuses
# already in the four-value vocabulary (CONFIRMED 17 · CLAIMED 17 ·
# INFERRED 8) and layers already OPS/CUST/DATA/INFRA rather than the
# prototype's L2-L5. A producer writing the techstack page had to reconstruct
# all of it from prose.
_TECH_TABS = ("Tech_Register", "Technographic_Scan", "Technology_Register",
              "Tech_Stack")
_TECH_PEER_TABS = ("Tech_Peer_Deployments", "Platform_Peer_Adoption")

#: layer -> the pillar that absorbs it, per the techstack contract.
_LAYER_PILLAR = {"OPS": "P3", "CUST": "P2", "DATA": "P4", "INFRA": "P4"}
_TECH_STATUS = ("CONFIRMED", "INFERRED", "CLAIMED", "ABSENT")

#: One clause, printed in the register row AND the T3 detail header. The
#: budget is the contract's, not this reader's; it is checked HERE so an
#: over-long clause is named at ingest instead of at the gate, where it
#: reads as a producer defect rather than as the package's own prose.
_DETECTION_BASIS_BUDGET = 160

_TECH_ALIASES = {
    "ts_id": ("ts_id", "id", "tech_id"),
    "product": ("product", "product_name"),
    "vendor": ("vendor", "supplier", "provider"),
    "layer": ("layer", "stack_layer"),
    "status": ("status", "presence"),
    "evidence_level": ("evidence_level", "level"),
    "detection_basis": ("detection_basis", "basis"),
    "detection_method": ("detection_method", "method"),
    "providers": ("providers", "detected_by"),
    "subcaps": ("subcap_ids", "subcaps", "linked_subcap_ids", "cells"),
    "e_ids": ("evidence_ids", "e_ids", "evidence"),
    "source_urls": ("source_urls", "source_url", "urls"),
    "as_of": ("as_of", "as_at", "as at", "asof"),
    "dma_impact": ("dma_impact", "impact"),
}

_TECH_PEER_ALIASES = {
    "ts_id": ("ts_id", "id", "product_layer", "product / layer", "product"),
    "peer": ("peer", "institution"),
    "deployed": ("deployed", "verdict"),
    "basis": ("basis",),
    "source_url": ("source_url", "source"),
    "as_of": ("as_of", "as_at", "as at", "asof"),
}


def _tech_split(value) -> list:
    """A multi-value cell, under the separators the corpus actually uses."""
    if value is None:
        return []
    return [p for p in (x.strip() for x in
                        re.split(r"[,;|\n]+", str(value))) if p]


def _tri_state(value):
    """`deployed` is THREE-valued and the third value is the point: a peer
    nobody could establish is `null`, never False. A coverage figure of 2/5
    with three unknowns is not 2/5, and the card has to be able to say so."""
    s = str(value or "").strip().lower()
    if s in ("yes", "true", "y", "deployed", "confirmed", "1"):
        return True
    if s in ("no", "false", "n", "not deployed", "absent", "0"):
        return False
    return None



def _kept(parts, ok, raw, sink: list, ts_id: str) -> list:
    """Filter to the values that ARE identifiers, recording a cell that held
    something and yielded none."""
    out = [p for p in parts if ok(p)]
    if not out and str(raw or "").strip():
        sink.append({"ts_id": ts_id or None, "stated": str(raw).strip()[:60]})
    return out


def parse_tech_register(path: str, obs: list | None = None) -> list:
    """The workbook's technology register, shaped to the techstack contract.

    Returns one dict per product: {ts_id, product, vendor, layer, pillar_id,
    status, evidence_level, detection_basis, detection_method, providers[],
    linked_subcap_ids[], e_ids[], source_urls[], as_of, dma_impact,
    peer_deployments[]}.

    Two contract rules are checked HERE rather than left to the gates,
    because a defect the package itself carries must not read as one the
    producer introduced. On the Golden 1 register both fired at exactly the
    counts the run was later refused on: 12 rows state the same string as
    Product AND Vendor (CG-20), and 7 detection_basis clauses exceed the
    160-character face-slot budget (CG-12). Naming them at ingest turns 19
    late refusals into 19 rows a producer can see before writing anything.
    """
    def observe(kind, detail):
        if obs is not None:
            obs.append(Observation(kind, None, detail))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        tab = next((t for t in _TECH_TABS if t in wb.sheetnames), None)
        if tab is None:
            observe("tech_register_tab_not_found", {
                "expected_any_of": list(_TECH_TABS),
                "tabs_present": list(wb.sheetnames)[:30],
                "reason": "no technology register tab: the techstack page "
                          "has no register rows from this workbook"})
            return []
        ws = wb[tab]
        try:
            headers, first = _header_map(ws, "TS_ID")
        except ValueError:
            try:
                headers, first = _header_map(ws, "Product")
            except ValueError:
                observe("tech_register_header_not_found", {
                    "tab": tab, "expected_any_of": ["TS_ID", "Product"],
                    "reason": "the register tab exists and its id column "
                              "could not be located; no row was read"})
                return []
        cols = {k: _pick(headers, names) for k, names in _TECH_ALIASES.items()}
        for field, names in _TECH_ALIASES.items():
            if cols.get(field) is None:
                miss = _column_not_found(tab, field, names, headers)
                if obs is not None:
                    obs.append(miss)

        out = []
        id_rows = {}
        # A cell that HELD something and yielded no id. `category-level` and
        # `see Technographic_Scan` are cross-references, not identifiers, and
        # a filter that drops them into an empty list reports a row with no
        # cells exactly like a row whose cells nothing could parse.
        unparsed = {"subcaps": [], "e_ids": []}
        for row in ws.iter_rows(min_row=first, values_only=True):
            def v(key):
                i = cols.get(key)
                return row[i] if i is not None and i < len(row) else None
            ts_id = str(v("ts_id") or "").strip()
            product = str(v("product") or "").strip()
            if not (ts_id or product):
                continue
            # NOT de-duplicated on ts_id. Measured on the Golden 1 register:
            # 42 rows carry 28 distinct ids because the numbering restarts per
            # layer block — TS-021 alone names Modelshop, Salesforce Marketing
            # Cloud, AML RightSource, Azure APIM and Okta. Collapsing on the
            # id drops 14 products the client actually runs, which is the
            # silent loss this reader exists to refuse. The collision is
            # reported instead, and every row is carried.
            if ts_id:
                id_rows.setdefault(ts_id, []).append(product or "(unnamed)")
            raw_sub, raw_eid = v("subcaps"), v("e_ids")
            layer = str(v("layer") or "").strip().upper()
            status = str(v("status") or "").strip().upper()
            out.append({
                "ts_id": ts_id or None,
                "product": product or None,
                "vendor": str(v("vendor") or "").strip() or None,
                "layer": layer if layer in _LAYER_PILLAR else None,
                "pillar_id": _LAYER_PILLAR.get(layer),
                # REQUIRED on every row by the contract — the landscape strip
                # recomputes its four counts from it and is uncomputable
                # without it. Carried as null rather than defaulted: a status
                # this reader invented would be indistinguishable from one the
                # assessment made.
                "status": status if status in _TECH_STATUS else None,
                "evidence_level": str(v("evidence_level") or "").strip().upper() or None,
                "detection_basis": str(v("detection_basis") or "").strip() or None,
                "detection_method": str(v("detection_method") or "").strip() or None,
                "providers": _tech_split(v("providers")),
                "linked_subcap_ids": _kept(_tech_split(raw_sub),
                                           SUBCAP_RE.match, raw_sub,
                                           unparsed["subcaps"], ts_id),
                "e_ids": _kept(_tech_split(raw_eid),
                               lambda e: e.startswith(("E-", "INT-")), raw_eid,
                               unparsed["e_ids"], ts_id),
                "source_urls": _tech_split(v("source_urls")),
                "as_of": str(v("as_of") or "").strip() or None,
                "dma_impact": str(v("dma_impact") or "").strip() or None,
                "peer_deployments": [],
            })

        _attach_peer_deployments(wb, out, observe)

        collisions = {k: v for k, v in id_rows.items() if len(v) > 1}
        if collisions:
            observe("tech_register_ts_id_collision", {
                "tab": tab, "ids": len(collisions),
                "rows_affected": sum(len(v) for v in collisions.values()),
                "example": {k: v for k, v in list(collisions.items())[:3]},
                "reason": "one ts_id names several DIFFERENT products, so the "
                          "register's numbering is not unique across the "
                          "sheet — it restarts per layer block. Every row is "
                          "carried; a reader that keyed on the id would drop "
                          "the products sharing it. ts_id is agent-minted, so "
                          "the repair belongs in the package."})
        for field, rows_ in unparsed.items():
            if rows_:
                observe("tech_register_reference_not_an_id", {
                    "tab": tab, "field": field, "rows": len(rows_),
                    "example": rows_[:4],
                    "reason": "the cell states a cross-reference rather than "
                              "identifiers, so the row lands with an empty "
                              "list. That is not a product with no cells and "
                              "no evidence — it is one whose links were "
                              "written somewhere a reader cannot follow, and "
                              "a techstack row that cites nothing is refused "
                              "by CG-50 whatever the register says."})

        # ── contract defects the PACKAGE carries, named here ──────────────
        same = [r["ts_id"] for r in out
                if r["product"] and r["product"] == r["vendor"]]
        if same:
            observe("tech_register_vendor_equals_product", {
                "tab": tab, "rows": len(same), "example": same[:6],
                "gate": "CG-20",
                "reason": "product and vendor state the same string, so one "
                          "of the two is unstated. A register row names a "
                          "company AND the thing it supplies; repeating the "
                          "company in both renders as a product nobody "
                          "sells. Stated by the WORKBOOK, not introduced by "
                          "a producer — re-ingesting will not change it."})
        longs = [(r["ts_id"], len(r["detection_basis"])) for r in out
                 if r["detection_basis"]
                 and len(r["detection_basis"]) > _DETECTION_BASIS_BUDGET]
        if longs:
            observe("tech_register_detection_basis_over_budget", {
                "tab": tab, "rows": len(longs), "budget": _DETECTION_BASIS_BUDGET,
                "example": longs[:6], "gate": "CG-12",
                "reason": "detection_basis renders in the register row and "
                          "the T3 detail header and holds ONE CLAUSE. The "
                          "repair is to MOVE the prose into dma_impact, not "
                          "to trim it; a paragraph in a face slot overflows "
                          "its container."})
        missing_status = [r["ts_id"] for r in out if r["status"] is None]
        if missing_status:
            observe("tech_register_status_missing", {
                "tab": tab, "rows": len(missing_status),
                "example": missing_status[:6], "expected_any_of": list(_TECH_STATUS),
                "reason": "status is REQUIRED on every register row: the "
                          "landscape strip recomputes its four counts from "
                          "it and cannot be computed without it"})

        observe("tech_register_summary", {
            "tab": tab, "rows": len(out),
            "by_status": {s: sum(1 for r in out if r["status"] == s)
                          for s in _TECH_STATUS
                          if any(r["status"] == s for r in out)},
            "by_layer": {ly: sum(1 for r in out if r["layer"] == ly)
                         for ly in _LAYER_PILLAR
                         if any(r["layer"] == ly for r in out)},
            "with_cells": sum(1 for r in out if r["linked_subcap_ids"]),
            "with_evidence": sum(1 for r in out if r["e_ids"]),
            "with_peer_rows": sum(1 for r in out if r["peer_deployments"]),
        })
        return out
    finally:
        wb.close()


def _attach_peer_deployments(wb, items: list, observe) -> None:
    """Per-peer rows behind a product's coverage share, keyed on TS_ID.

    The contract wants one row per peer INCLUDING the peers nobody could
    establish, so `deployed` stays tri-state and an unmatched key is
    reported rather than dropped.
    """
    tab = next((t for t in _TECH_PEER_TABS if t in wb.sheetnames), None)
    if tab is None:
        return
    ws = wb[tab]
    headers = first = None
    for anchor in ("TS_ID", "Product / Layer", "Product", "Peer"):
        try:
            headers, first = _header_map(ws, anchor)
            break
        except ValueError:
            continue
    if headers is None:
        observe("tech_peer_header_not_found", {
            "tab": tab, "reason": "peer deployment rows exist and their key "
                                  "column could not be located; no peer row "
                                  "was attached"})
        return
    cols = {k: _pick(headers, names) for k, names in _TECH_PEER_ALIASES.items()}
    by_id, by_product = {}, {}
    for it in items:
        if it["ts_id"]:
            by_id[it["ts_id"]] = it
        if it["product"]:
            by_product[it["product"].strip().lower()] = it
    attached = unmatched = 0
    orphans = []
    for row in ws.iter_rows(min_row=first, values_only=True):
        def v(key):
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None
        key = str(v("ts_id") or "").strip()
        peer = str(v("peer") or "").strip()
        if not (key and peer):
            continue
        target = by_id.get(key) or by_product.get(key.lower())
        if target is None:
            unmatched += 1
            if len(orphans) < 6:
                orphans.append(key)
            continue
        target["peer_deployments"].append({
            "peer": peer,
            "deployed": _tri_state(v("deployed")),
            "basis": str(v("basis") or "").strip() or None,
            "source_url": str(v("source_url") or "").strip() or None,
            "as_of": str(v("as_of") or "").strip() or None,
        })
        attached += 1
    observe("tech_peer_deployments_attached", {
        "tab": tab, "attached": attached, "unmatched": unmatched,
        "unmatched_examples": orphans,
        "reason": "peer rows whose key matches no register row are reported "
                  "rather than dropped: a coverage share computed over a "
                  "peer set the register cannot name is not a share."})


# ── which tabs does anything actually read? ────────────────────────────────
#
# The Golden 1 workbook ships 43 tabs and the readers above claim 12 of them.
# The other 31 are not empty — Tech_Register (42 product rows), Focus_Areas,
# Entity_Timeline, Firmographics, Enrichment_Needed and the rest carry the
# material five of the six pages are written from — and nothing anywhere said
# so. A producer met them as blank surfaces and wrote absences over live data.
#
# This census is the standing answer: for any package, which tabs a reader
# claims, and which carry rows that nothing will ever read.
_TAB_READERS = {
    "Run_Metadata": "parse_scoring_workbook",
    "Pillar_Summary": "parse_grain_summaries",
    "Category_Detail": "parse_grain_summaries",
    "Pillar_Rollup": "parse_grain_summaries",
    "Category_Rollup": "parse_grain_summaries",
    "Peer_Benchmarks": "parse_peer_benchmarks",
    "Recommendations": "parse_recommendations",
    "Caps_Applied_Log": "parse_scoring_workbook",
    "Evidence_Master": "parse_evidence_master",
    "Evidence_Detail": "parse_evidence_master",
    "Evidence_Register": "parse_evidence_master",
    "Evidence_Index": "parse_evidence_master",
    "Evidence_Ledger": "parse_evidence_master",
    "Evidence_Linkage": "parse_evidence_master",
    "Evidence_Linkage_Matrix": "parse_evidence_master",
    "Evidence_Inventory": "parse_evidence_master",
    "Tech_Register": "parse_tech_register",
    "Technographic_Scan": "parse_tech_register",
    "Tech_Peer_Deployments": "parse_tech_register",
    "Platform_Peer_Adoption": "parse_tech_register",
}


#: Where a tab's rows BELONG, so the census can say more than "nothing reads
#: this": it can name the surface that is rendering empty because of it.
#: `verified` mappings were checked field-by-field against the page contract
#: returned by the connector's get_page_contract; `proposed` ones are read off
#: the tab's own shape and are the worklist, not a promise. Nothing here
#: parses anything — it is the map a reader consults before writing the next
#: parser, kept beside the readers so the two cannot drift apart.
_TAB_TARGET = {
    # verified against get_page_contract
    "Tech_Register": ("techstack.techstack.items", "verified"),
    "Technographic_Scan": ("techstack.techstack.items", "verified"),
    "Tech_Peer_Deployments":
        ("techstack.techstack.items[].peer_deployments", "verified"),
    "Platform_Peer_Adoption":
        ("techstack.techstack.items[].peer_deployments", "verified"),
    "Recommendations": ("platform.recommendations.recommendations", "verified"),
    # proposed from the tab's own shape — the worklist
    "Focus_Areas": ("insights (H1 focus areas)", "proposed"),
    "Entity_Timeline": ("context", "proposed"),
    "Firmographics": ("overview", "proposed"),
    "Subcap_Scores": ("heatmap cells", "proposed"),
    "Coverage": ("overview coverage posture", "proposed"),
    "Coverage_Map": ("overview coverage posture", "proposed"),
    "Solution_Catalogue": ("platform.platform_story candidate set", "proposed"),
    "Issue_Register":
        ("platform.stairstep.ladder.steps[].blocking_findings", "proposed"),
    "Enrichment_Needed": ("enrichment facets", "proposed"),
    "Report_Narrative": ("page narrative_thread / report sections", "proposed"),
    "Challenge_Log": ("internal_only provenance", "proposed"),
    "Gate_Log": ("internal_only provenance", "proposed"),
    "Provenance": ("internal_only provenance", "proposed"),
    "Search_Log": ("internal_only provenance", "proposed"),
    # run configuration and method, not a client-facing surface
    "Maturity_Rubric": ("run config", "not_client_facing"),
    "Pillar_Weights": ("run config", "not_client_facing"),
    "Catalogue_Meta": ("run config", "not_client_facing"),
    "Handoff_Lock": ("run config", "not_client_facing"),
    "Cap_Triggers": ("run config", "not_client_facing"),
    "Capability_Definitions": ("run config", "not_client_facing"),
    "REF_Method": ("run config", "not_client_facing"),
    "DQ_Bank": ("run config", "not_client_facing"),
    "00_README": ("run config", "not_client_facing"),
    "Executive_Summary": ("run config", "not_client_facing"),
}


def workbook_tab_coverage(path: str, obs: list | None = None) -> dict:
    """Name every tab in the package and say what reads it.

    Emitted at ingest so an unmapped tab is a recorded fact rather than a
    surface that renders empty for reasons nobody can see. `unread_with_rows`
    is the worklist: tabs carrying data no reader claims.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        read, unread = {}, {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = max((ws.max_row or 1) - 1, 0)   # less the header
            reader = _TAB_READERS.get(name)
            if reader:
                read[name] = {"reader": reader, "rows": rows}
            elif _is_pillar_tab(name):
                read[name] = {"reader": "_parse_pillar_scoring", "rows": rows}
            else:
                unread[name] = rows
        with_rows = {k: v for k, v in unread.items() if v > 0}
        ordered = dict(sorted(with_rows.items(), key=lambda kv: -kv[1]))
        # Worst first, and CLIENT-FACING first within that: a run-config tab
        # nothing reads costs nothing, while an unread Focus_Areas is a page
        # rendering empty over live rows.
        targets, unmapped = {}, []
        for name in ordered:
            hit = _TAB_TARGET.get(name)
            if hit is None:
                unmapped.append(name)
            elif hit[1] != "not_client_facing":
                targets[name] = {"feeds": hit[0], "confidence": hit[1]}
        report = {
            "tabs_total": len(wb.sheetnames),
            "tabs_read": len(read),
            "tabs_unread": len(unread),
            "unread_with_rows": ordered,
            # The subset that costs a surface, with the surface named.
            "unread_client_facing": targets,
            # A tab nobody has even classified. Worth a look before the next
            # package arrives carrying more of them.
            "unread_unmapped": unmapped,
        }
        if obs is not None and with_rows:
            obs.append(Observation("workbook_tabs_unread", None, {
                **report,
                "reason": "these tabs carry rows and no reader claims them. "
                          "A surface written from a tab in this list renders "
                          "empty because nothing read it, NOT because the "
                          "client has nothing to say — which is the absence "
                          "a producer must never write."}))
        return report
    finally:
        wb.close()
