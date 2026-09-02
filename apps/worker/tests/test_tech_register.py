"""The technology register tab, and the census that finds tabs nothing reads.

Measured on the Golden 1 package (DMA-2026-GOLDEN1-001), whose scoring
workbook ships 43 tabs. Before this reader existed the parsers claimed 12 of
them, and `Tech_Register` — 42 product rows already carrying every field the
techstack T1/T3 contract names — was not one. A producer writing that page
met a blank surface and reconstructed the estate from prose.

Every shape here is a real one from that package, rebuilt synthetically; no
client data lives in the repo.
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dma_worker.workbook_parser import (  # noqa: E402
    parse_tech_register, workbook_tab_coverage,
)

_HDR = ["TS_ID", "Product", "Vendor", "Layer", "Status", "Evidence_Level",
        "Detection_Basis", "Detection_Method", "Providers", "SubCap_IDs",
        "Evidence_IDs", "Source_URLs", "As_Of", "DMA_Impact"]


def _row(ts_id, product, vendor, layer="CUST", status="CONFIRMED",
         basis="named in the 2024 readout as the digital banking platform",
         subcaps="P1C1.1.1", eids="E-001"):
    return [ts_id, product, vendor, layer, status, "L2", basis,
            "internal_document", "drive", subcaps, eids,
            "https://example.invalid", "2026-08-31", "supports the estate"]


def _book(tmp_path, rows, peers=None, extra_tabs=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tech_Register"
    ws.append(_HDR)
    for r in rows:
        ws.append(r)
    if peers is not None:
        p = wb.create_sheet("Tech_Peer_Deployments")
        p.append(["TS_ID", "Peer", "Deployed", "Basis", "Source_URL", "As_Of"])
        for r in peers:
            p.append(r)
    for name, n in (extra_tabs or {}).items():
        t = wb.create_sheet(name)
        t.append(["Field", "Value"])
        for i in range(n):
            t.append([f"f{i}", "v"])
    path = tmp_path / "tech.xlsx"
    wb.save(path)
    return str(path)


def _obs(obs, kind):
    return [o for o in obs if o.kind == kind]


def test_golden1_a_reused_ts_id_never_collapses_two_products(tmp_path):
    """THE regression. The register's numbering restarts per layer block, so
    42 rows carry 28 distinct ids — TS-021 alone names Modelshop, Salesforce
    Marketing Cloud, AML RightSource, Azure APIM and Okta. A reader keyed on
    the id drops 14 products the client actually runs."""
    rows = [
        _row("TS-021", "Modelshop", "Modelshop", layer="CUST"),
        _row("TS-021", "Salesforce Marketing Cloud", "Salesforce", layer="CUST"),
        _row("TS-021", "Okta", "Okta", layer="INFRA"),
        _row("TS-022", "MuleSoft Anypoint", "Salesforce", layer="INFRA"),
    ]
    obs = []
    out = parse_tech_register(_book(tmp_path, rows), obs)
    assert len(out) == 4, "every row is a product; none may be collapsed"
    assert [o["product"] for o in out] == [
        "Modelshop", "Salesforce Marketing Cloud", "Okta", "MuleSoft Anypoint"]
    hit = _obs(obs, "tech_register_ts_id_collision")
    assert len(hit) == 1
    assert hit[0].detail["ids"] == 1 and hit[0].detail["rows_affected"] == 3


def test_golden1_contract_fields_land_at_the_shape_the_page_wants(tmp_path):
    out = parse_tech_register(_book(tmp_path, [_row("TS-001", "NCR D3", "NCR")]))
    it = out[0]
    assert it["layer"] == "CUST" and it["pillar_id"] == "P2", \
        "the layer carries the pillar that absorbs it"
    assert it["status"] == "CONFIRMED"
    assert it["linked_subcap_ids"] == ["P1C1.1.1"] and it["e_ids"] == ["E-001"]
    assert it["as_of"] == "2026-08-31"


def test_golden1_a_status_outside_the_four_values_is_null_not_defaulted(tmp_path):
    """The landscape strip recomputes its four counts from status. A status
    this reader invented would be indistinguishable from one the assessment
    made, so an unrecognised value lands null and is named."""
    obs = []
    out = parse_tech_register(
        _book(tmp_path, [_row("TS-001", "Thing", "Vendor", status="PROBABLY")]), obs)
    assert out[0]["status"] is None
    assert _obs(obs, "tech_register_status_missing")[0].detail["rows"] == 1


def test_golden1_package_origin_contract_defects_are_named_at_ingest(tmp_path):
    """CG-20 and CG-12 refused the Golden 1 run on 12 and 7 rows. Both counts
    are properties of the WORKBOOK, so naming them here turns 19 late gate
    refusals into 19 rows a producer sees before writing anything."""
    long_basis = "x" * 200
    rows = [
        _row("TS-002", "Zest AI", "Zest AI"),                    # CG-20
        _row("TS-003", "Snowflake", "Snowflake", basis=long_basis),  # both
        _row("TS-004", "NCR D3", "NCR"),                          # clean
    ]
    obs = []
    parse_tech_register(_book(tmp_path, rows), obs)
    same = _obs(obs, "tech_register_vendor_equals_product")[0].detail
    assert same["rows"] == 2 and same["gate"] == "CG-20"
    over = _obs(obs, "tech_register_detection_basis_over_budget")[0].detail
    assert over["rows"] == 1 and over["gate"] == "CG-12" and over["budget"] == 160


def test_golden1_a_cross_reference_is_not_an_empty_link_list(tmp_path):
    """`category-level` and `see Technographic_Scan` are pointers, not ids.
    A row landing with an empty list because its links were written somewhere
    unfollowable is not a row with no cells, and CG-50 refuses it either way."""
    rows = [_row("TS-001", "NCR D3", "NCR",
                 subcaps="category-level", eids="see Technographic_Scan")]
    obs = []
    out = parse_tech_register(_book(tmp_path, rows), obs)
    assert out[0]["linked_subcap_ids"] == [] and out[0]["e_ids"] == []
    named = {o.detail["field"] for o in _obs(obs, "tech_register_reference_not_an_id")}
    assert named == {"subcaps", "e_ids"}


def test_golden1_peer_deployments_keep_the_unknown_peer(tmp_path):
    """A coverage figure of 2/5 with three unknowns is not 2/5. `deployed` is
    tri-state and the third value is the point."""
    peers = [["TS-001", "SchoolsFirst", "yes", "press release", "https://x", "2026-01-01"],
             ["TS-001", "First Tech", "no", "not found", "https://y", "2026-01-01"],
             ["TS-001", "America First", "", "could not establish", "", ""]]
    out = parse_tech_register(
        _book(tmp_path, [_row("TS-001", "NCR D3", "NCR")], peers=peers))
    got = {p["peer"]: p["deployed"] for p in out[0]["peer_deployments"]}
    assert got == {"SchoolsFirst": True, "First Tech": False, "America First": None}


def test_golden1_a_peer_row_matching_no_product_is_reported_not_dropped(tmp_path):
    peers = [["TS-999", "SchoolsFirst", "yes", "b", "https://x", "2026-01-01"]]
    obs = []
    parse_tech_register(
        _book(tmp_path, [_row("TS-001", "NCR D3", "NCR")], peers=peers), obs)
    d = _obs(obs, "tech_peer_deployments_attached")[0].detail
    assert d["attached"] == 0 and d["unmatched"] == 1
    assert d["unmatched_examples"] == ["TS-999"]


# ── class 19: a tab carrying rows that no reader claims ────────────────────
# The Golden 1 workbook ships 43 tabs against 12 claimed readers, and nothing
# anywhere said so. A producer met the other 31 as blank surfaces.

def test_golden1_tabs_no_reader_claims_are_named_with_their_row_counts(tmp_path):
    path = _book(tmp_path, [_row("TS-001", "NCR D3", "NCR")],
                 extra_tabs={"Focus_Areas": 4, "Entity_Timeline": 16,
                             "Empty_Sheet": 0})
    obs = []
    cov = workbook_tab_coverage(path, obs)
    assert cov["tabs_read"] >= 1, "Tech_Register has a reader now"
    unread = cov["unread_with_rows"]
    assert unread["Entity_Timeline"] == 16 and unread["Focus_Areas"] == 4
    assert "Empty_Sheet" not in unread, "a tab with no rows is not a worklist item"
    assert "Tech_Register" not in unread
    # Sorted worst-first: the worklist leads with the biggest unread tab.
    assert list(unread)[0] == "Entity_Timeline"
    hit = _obs(obs, "workbook_tabs_unread")
    assert len(hit) == 1 and hit[0].detail["tabs_total"] == cov["tabs_total"]


def test_golden1_a_fully_mapped_workbook_reports_no_worklist(tmp_path):
    obs = []
    cov = workbook_tab_coverage(
        _book(tmp_path, [_row("TS-001", "NCR D3", "NCR")]), obs)
    assert cov["unread_with_rows"] == {}
    assert _obs(obs, "workbook_tabs_unread") == [], \
        "nothing to say when every tab with rows has a reader"
