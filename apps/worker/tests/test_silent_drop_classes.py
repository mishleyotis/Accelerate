"""The silent-drop classes, each named for the package that exposed it.

Measured on 2026-08-08 against the 171 client folders under the production
intake tree (153 carry a workbook the classifier recognises). Every case here
is a real shape from that corpus, rebuilt synthetically — no client data lives
in the repo. The governing rule they all encode:

    a reader that does not recognise its input must produce a NAMED
    observation, never an empty result.

A parse that yields zero of something the tab plainly contains is a loud
failure. `test_no_workbook_can_parse_to_a_silent_zero` is the general form;
the rest are the specific spellings measured in the corpus.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dma_worker.workbook_parser import EID_RE, parse_scoring_workbook


def _pillar_workbook(tmp_path, header, rows, tab="P1_Subcap_Scoring",
                     name="wb.xlsx"):
    """One general_dma pillar-scoring tab, header on row 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    path = tmp_path / name
    wb.save(path)
    return str(path)


_CANON = ["SubCap_ID", "SubCap_Name", "Category", "Score", "Confidence",
          "Evidence_IDs", "Source_URLs", "Evidence_Ceiling", "Caps_Applied",
          "Rationale", "Proxy_Searched"]


def _obs(parse, kind):
    return [o for o in parse.observations if o.kind == kind]


# ── class 6: cell-id format drift parsed to a totally silent zero ──────────
# American Homes - DMA and Wescom Financial - DMA both state CATEGORY ids in
# the SubCap_ID column of a subcapability tab. 1,401 populated rows across
# eight tabs matched nothing: no scores, no observations, not even a
# toggled_out entry — a record indistinguishable from an empty assessment.

def test_american_homes_category_grain_ids_name_themselves(tmp_path):
    path = _pillar_workbook(tmp_path, _CANON, [
        ["P1C1", "Digital Strategy & Vision", "P1C1", 3, "HIGH",
         "E-001,E-002", "https://x", "tier_ceil:5.0", "IR-cap:3.0",
         "[EVIDENCE]: E-001 — roadmap", "Yes"],
        ["P1C1", "Digital Roadmap", "P1C1", 2.5, "HIGH", "E-003", "https://x",
         "tier_ceil:5.0", "", "[EVIDENCE]: E-003 — board deck", "Yes"],
    ])
    p = parse_scoring_workbook(path)
    assert p.scores == []
    named = _obs(p, "unrecognised_cell_id_format")
    assert len(named) == 1
    d = named[0].detail
    assert d["tab"] == "P1_Subcap_Scoring"
    assert d["column"] == "subcap_id"
    assert d["rows_dropped"] == 2
    assert d["found_examples"] == ["P1C1"]
    assert "P" in d["expected"]          # the pattern it wanted, spelled out


def test_no_workbook_can_parse_to_a_silent_zero(tmp_path):
    """The general form: pillar tabs read, nothing came out, said so."""
    path = _pillar_workbook(tmp_path, _CANON, [])
    p = parse_scoring_workbook(path)
    assert not p.scores and not p.toggled_out
    assert p.observations, "an empty parse must name itself"
    assert _obs(p, "workbook_yielded_nothing")[0].detail["tabs_read"] == \
        ["P1_Subcap_Scoring"]


# ── class 5: the rationale column matched by prefix only ──────────────────
# `Scoring_Rationale` is used by 35 scoring tabs across nine clients (ATB,
# AmeriCU, Bell Bank, Cathay Bank, Compeer, Farm Credit Mid America, GESA,
# Wawanesa, Zions). It is the ONLY source of verbatim excerpt text in the
# general_dma generation, so every one of those clients reached the evidence
# drawer with nothing to quote.

@pytest.mark.parametrize("spelling", ["Rationale", "Scoring_Rationale",
                                      "Assessor_Rationale", "Justification",
                                      "Rationale (>=150 chars)"])
def test_zions_scoring_rationale_spelling_is_read(tmp_path, spelling):
    header = [h if h != "Rationale" else spelling for h in _CANON]
    path = _pillar_workbook(tmp_path, header, [
        ["P1C1.1.1", "Board packs", "P1C1", 3, "HIGH", "E-001", "https://x",
         "", "", "[E-001:F1] The board approved a three-year roadmap.", "Yes"],
    ])
    p = parse_scoring_workbook(path)
    assert len(p.scores) == 1
    assert p.scores[0].rationale.startswith("[E-001:F1]")
    assert not _obs(p, "column_not_found")


def test_a_column_the_parser_cannot_find_names_the_tab_and_the_spellings(tmp_path):
    header = [h for h in _CANON if h not in ("Rationale", "Evidence_IDs")]
    path = _pillar_workbook(tmp_path, header, [
        ["P1C1.1.1", "Board packs", "P1C1", 3, "HIGH", "https://x", "", "",
         "Yes"],
    ])
    p = parse_scoring_workbook(path)
    assert len(p.scores) == 1                      # the score still lands
    fields = {o.detail["field"]: o.detail for o in _obs(p, "column_not_found")}
    assert set(fields) == {"rationale", "evidence_ids"}
    for detail in fields.values():
        assert detail["tab"] == "P1_Subcap_Scoring"
        assert detail["expected_any_of"], "must name the spellings it accepts"
        assert detail["headers_present"], "must name what it actually found"


# ── class 4: internal INT- evidence ids silently discarded ────────────────
# The upstream dma-assessment skill's published column spec
# (references/workbook_specification.md, Column F) mandates
# `E-\d{3}` OR `INT-[DOC_ABBREV]-\d{3}`. One shipped package (ATB - DMA)
# carries 341 cells citing internal documents that way.

def test_atb_internal_int_evidence_ids_are_kept(tmp_path):
    path = _pillar_workbook(tmp_path, _CANON, [
        ["P1C1.1.1", "Board packs", "P1C1", 3, "HIGH",
         "E-001, INT-BOARD-003, INT-REQ-045", "https://x", "", "",
         "cited", "Yes"],
    ])
    p = parse_scoring_workbook(path)
    assert p.scores[0].evidence_refs == ["E-001", "INT-BOARD-003", "INT-REQ-045"]


def test_the_published_id_spec_is_what_the_regex_accepts():
    assert EID_RE.findall("E-001, INT-BOARD-002, E-015") == \
        ["E-001", "INT-BOARD-002", "E-015"]
    assert EID_RE.findall("[E-012:F1] and [INT-TECH-001:F2]") == \
        ["E-012:F1", "INT-TECH-001:F2"]


# ── class 7: no score-range check anywhere ────────────────────────────────
# Payments Canada - DMA states 0 in 38 cells. Zero is not a maturity level —
# the rubric is M1..M5 — but it banded as Activating and rendered as a real
# assessment of a barely-started capability.

@pytest.mark.parametrize("stated", [0, -1, 5.5, 7])
def test_payments_canada_score_outside_the_rubric_is_refused(tmp_path, stated):
    path = _pillar_workbook(tmp_path, _CANON, [
        ["P2C3.9.1", "Settlement", "P2C3", stated, "HIGH", "E-001", "", "",
         "", "cited", "No"],
    ])
    p = parse_scoring_workbook(path)
    assert p.scores == [], "an out-of-rubric value must never become a band"
    out = _obs(p, "score_out_of_range")
    assert len(out) == 1
    assert out[0].subcap_id == "P2C3.9.1"
    assert out[0].detail["stated"] == str(stated)
    assert out[0].detail["source_cell"] == "P1_Subcap_Scoring!D2"


@pytest.mark.parametrize("stated", [1, 2.5, 5])
def test_a_score_inside_the_rubric_still_lands(tmp_path, stated):
    path = _pillar_workbook(tmp_path, _CANON, [
        ["P2C3.9.1", "Settlement", "P2C3", stated, "HIGH", "E-001", "", "",
         "", "cited", "No"],
    ])
    p = parse_scoring_workbook(path)
    assert len(p.scores) == 1 and not _obs(p, "score_out_of_range")


# ── class 1: the peer parser invents institutions ─────────────────────────
# Any header outside the fixed 33 spellings became a named peer. Measured
# across the corpus, 28 clients carried at least one peer that is not an
# institution: `Gap_vs_Median`, `Position`, `Peer_Name`, `Cat_ID`,
# `Unknown`, `Quality_Grade`, `ANB_vs_Median`, …

def _peer_tab(tmp_path, rows, name="peers.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peer_Benchmarks"
    for r in rows:
        ws.append(list(r))
    path = tmp_path / name
    wb.save(path)
    return str(path)


def test_a_gap_column_is_a_statistic_not_a_bank(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "Fake Trust Bank", "Gap_vs_Median", "Position",
         "Unknown", "Quality_Grade", "Peer_Median"],
        ["P1C1", 2.8, -0.42, "Below", "n/a", "B+", 3.2],
        ["P1C2", 3.1, 0.10, "Above", "n/a", "A-", 3.0],
    ])
    obs = []
    out = parse_peer_benchmarks(path, obs)
    assert [n for n, _ in out[0]["peers"]] == ["Fake Trust Bank"]
    # Gap_vs_Median, Position and Unknown name statistics the parser knows.
    # Quality_Grade names nothing it knows AND holds no score, so it is
    # refused BY NAME instead of joining the cohort as an institution.
    refused = {c["column"] for o in obs
               if o.kind == "peer_column_unrecognised" for c in o.detail["columns"]}
    assert refused == {"Quality_Grade"}


def test_a_real_peer_whose_name_starts_with_delta_is_not_a_statistic(tmp_path):
    """Delta Community is a credit union. A name-only rule that swallowed it
    would drop a peer instead of inventing one — the same defect facing the
    other way, which is why the value on the maturity scale decides."""
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "Delta_Community", "Peer_Median"],
        ["P1C1", 3.4, 3.2],
    ])
    assert [n for n, _ in parse_peer_benchmarks(path)[0]["peers"]] == \
        ["Delta_Community"]


def test_first_united_zero_filled_peer_grid_is_refused_by_name(tmp_path):
    """Five named peers, every score literally 0. Zero is not a maturity
    level; storing it puts five banks on the cohort at M1.

    `FUB_Score` is First United Bank's OWN column, and this test used to
    assert it INTO the cohort — see class 17 below, which is the defect that
    assertion was pinning in place.
    """
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "FUB_Score", "Fake Peer One", "Peer_Median"],
        ["P1C1", 1.77, 0, 2.5],
        ["P1C2", 1.86, 0, 2.5],
    ])
    obs = []
    out = parse_peer_benchmarks(path, obs,
                                subject_names=("First United Bank",))
    assert [n for n, _ in out[0]["peers"]] == [], (
        "the only two columns are the subject's own and a zero-filled peer")
    refused = [c for o in obs if o.kind == "peer_column_unrecognised"
               for c in o.detail["columns"]]
    assert refused[0]["column"] == "Fake Peer One"
    assert refused[0]["outside_rubric"] == ["0", "0"]


# ── class 17: the entity is stored as a peer of itself ────────────────────
# `Peer_Benchmarks` carries the subject's own score in a NAMED column beside
# the cohort's — `FUB_Score` for First United Bank. The parser was never told
# whose assessment it was reading (`parse_peer_benchmarks(path, obs)` took no
# identity, and job_main passed none), so that column passed every test a peer
# has to pass: it names no known statistic, and it holds perfectly valid
# scores on the maturity scale. It was stored as an institution.
#
# The consequence is not a stray row. The client joins its own cohort, its own
# score contributes to the benchmark it is being measured against, and the
# cohort statistics computed downstream measure the subject partly against
# itself. Nothing anywhere said so.

def test_the_subject_is_not_stored_as_a_peer_of_itself(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "FUB_Score", "Fake Trust Bank", "Peer_Median"],
        ["P1C1", 1.77, 2.4, 2.5],
    ])
    obs = []
    out = parse_peer_benchmarks(path, obs,
                                subject_names=("First United Bank",))
    assert [n for n, _ in out[0]["peers"]] == ["Fake Trust Bank"]
    named = [o for o in obs if o.kind == "peer_column_is_the_subject"]
    assert named and named[0].detail["columns"] == ["FUB_Score"]


@pytest.mark.parametrize("column,subject", [
    ("FUB_Score", "First United Bank"),
    ("FUB", "First United Bank, Inc."),          # the suffix is not identity
    ("First_United_Bank", "First United Bank"),
    ("BCU_Score", "Baxter Credit Union (BCU)"),  # the parenthesised short form
    ("Fake Bank Of Elsewhere", "Fake Bank of Elsewhere - DMA"),
])
def test_the_subject_is_recognised_under_the_spellings_the_corpus_uses(
        tmp_path, column, subject):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", column, "Fake Trust Bank", "Peer_Median"],
        ["P1C1", 1.77, 2.4, 2.5],
    ])
    out = parse_peer_benchmarks(path, [], subject_names=(subject,))
    assert [n for n, _ in out[0]["peers"]] == ["Fake Trust Bank"]


@pytest.mark.parametrize("peer,subject", [
    # Refusing a REAL peer is the mirror defect and the more expensive one,
    # so nothing matches on a prefix or a substring.
    ("First United Bancorp", "First United Bank"),
    ("Fake United Bank", "First United Bank"),
    ("FUBAR Savings", "First United Bank"),
    ("Delta_Community", "Delta Air Lines"),
])
def test_a_real_peer_that_merely_resembles_the_subject_is_kept(
        tmp_path, peer, subject):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", peer, "Peer_Median"],
        ["P1C1", 2.4, 2.5],
    ])
    out = parse_peer_benchmarks(path, [], subject_names=(subject,))
    assert [n for n, _ in out[0]["peers"]] == [peer]


def test_a_cohort_read_without_knowing_the_subject_says_so(tmp_path):
    """Passing no identity stays legal — a caller may genuinely not know — but
    it is never silent. "This cohort may contain the subject" has to be
    readable in the run, not inferred from a page months later."""
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "FUB_Score", "Fake Trust Bank", "Peer_Median"],
        ["P1C1", 1.77, 2.4, 2.5],
    ])
    obs = []
    parse_peer_benchmarks(path, obs)
    named = [o for o in obs if o.kind == "peer_subject_unknown"]
    assert named, "an unidentified cohort must name itself"
    assert "FUB_Score" in named[0].detail["candidate_columns"]


def test_the_ingest_path_tells_the_peer_parser_who_the_subject_is():
    """Asserted over the CALL, not over the file's text: the comment beside
    that call explains `subject_names`, and a substring check would match its
    own explanation. Three earlier tests in this repo failed exactly that way.
    """
    import ast
    from pathlib import Path
    src = (Path(__file__).resolve().parents[1] / "job_main.py").read_text()
    calls = [n for n in ast.walk(ast.parse(src))
             if isinstance(n, ast.Call)
             and getattr(n.func, "id", None) == "parse_peer_benchmarks"]
    assert calls, "job_main no longer calls the peer parser"
    for call in calls:
        assert "subject_names" in {kw.arg for kw in call.keywords}, (
            "the ingest path knows the client and must say so, or the "
            "subject rejoins its own cohort")


def test_a_named_peer_nobody_scored_keeps_its_column_and_says_so(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category_ID", "Fake Corporate FCU", "Peer_Median"],
        ["P1C1", None, 2.8],
    ])
    obs = []
    out = parse_peer_benchmarks(path, obs)
    assert out[0]["peers"] == [("Fake Corporate FCU", None)]
    assert [o.detail["columns"] for o in obs if o.kind == "peer_column_unscored"] \
        == [["Fake Corporate FCU"]]


# ── class 2: the header is assumed to be physically row 1 ─────────────────
# Seven Peer_Benchmarks tabs and twenty Recommendations tabs put a title, a
# methodology note or a run id above the header. Reading row 1 as the header
# made every column a fragment of that sentence — one client's cohort was an
# institution called "Peer Benchmarks — 5 Locked Corporate-CU Peers".

def test_corporate_america_title_row_above_the_peer_header(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Peer Benchmarks — 5 Locked Corporate-CU Peers"],
        [],
        ["Category_ID", "Category_Name", "Fake Corporate FCU",
         "Fake Central CU", "Median"],
        ["P1C1", "Digital Strategy", 2.8, 2.9, 2.8],
        ["P1C2", "Governance", 3.3, 3.4, 3.3],
    ])
    out = parse_peer_benchmarks(path)
    assert len(out) == 2
    assert [n for n, _ in out[0]["peers"]] == ["Fake Corporate FCU",
                                               "Fake Central CU"]
    assert out[0]["category_name"] == "Digital Strategy"


def test_payments_canada_category_label_rows_are_category_grain(tmp_path):
    """`P1C1: Digital Strategy & Roadmap` and `P1C1_digital_strategy` are the
    same grain as `P1C1`; requiring the bare id lost the whole cohort."""
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Category", "Fake Clearing House", "Peer_Median"],
        ["P1C1: Digital Strategy & Roadmap", 3.0, 3.1],
        ["P1C2_governance", 2.5, 2.6],
        ["P1C1.1", 2.0, 2.0],          # subcap grain — a different grain
    ])
    out = parse_peer_benchmarks(path)
    assert [r["category_id"] for r in out] == ["P1C1", "P1C2"]


def test_a_peer_tab_with_no_findable_header_refuses_with_a_reason(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    path = _peer_tab(tmp_path, [
        ["Peer_Benchmarks — see Phase 6/7 for full population"],
    ])
    obs = []
    assert parse_peer_benchmarks(path, obs) == []
    assert [o.kind for o in obs] == ["peer_header_not_found"]


# ── class 10: recommendations de-duplicated on the raw first column ───────
# Amarillo National Bank's 29 rows became 0 and Cetera's 24 became 4, because
# the first column repeats a rank or a phase label rather than an id.

def _rec_tab(tmp_path, rows, name="recs.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    for r in rows:
        ws.append(list(r))
    path = tmp_path / name
    wb.save(path)
    return str(path)


def test_amarillo_title_row_above_the_recommendations_header(tmp_path):
    from dma_worker.workbook_parser import parse_recommendations
    path = _rec_tab(tmp_path, [
        ["PHASE 5 — RECOMMENDATIONS"],
        ["Priority", "Title", "Category", "Horizon"],
        [1, "Unify the agent desktop", "P2C1", "0-6mo"],
        [2, "Retire the batch feed", "P3C2", "6-12mo"],
    ])
    out = parse_recommendations(path)
    assert [r["rec_id"] for r in out] == ["REC-1", "REC-2"]
    assert out[0]["payload"]["title"] == "Unify the agent desktop"


def test_cetera_repeating_first_column_keeps_every_row(tmp_path):
    from dma_worker.workbook_parser import parse_recommendations
    path = _rec_tab(tmp_path, [
        ["Priority", "Title", "Category"],
        ["High", "Unify the agent desktop", "P2C1"],
        ["High", "Retire the batch feed", "P3C2"],
        ["High", "Publish the data contract", "P4C1"],
    ])
    obs = []
    out = parse_recommendations(path, obs)
    assert len(out) == 3, "a repeating rank is not an id"
    assert len({r["rec_id"] for r in out}) == 3
    assert {r["payload"]["title"] for r in out} == {
        "Unify the agent desktop", "Retire the batch feed",
        "Publish the data contract"}
    named = [o for o in obs if o.kind == "recommendation_id_not_unique"]
    assert named[0].detail["rows_qualified_by_position"] == 2


def test_a_recommendations_tab_with_no_header_row_says_so(tmp_path):
    from dma_worker.workbook_parser import parse_recommendations
    path = _rec_tab(tmp_path, [
        [1, "FSC/Sales Cloud"],
        [2, "Data Cloud"],
    ])
    obs = []
    out = parse_recommendations(path, obs)
    assert [o.kind for o in obs] == ["recommendations_header_not_found"]
    assert len(out) == 2 and out[0]["payload"]["col_2"] == "FSC/Sales Cloud"


# ── class 8: the evidence ledger recognised under exactly one tab name ────
# 15 packages ship no `Evidence_Master`; eleven of them name the same tab
# `Evidence_Index`, `Evidence_Linkage_Matrix`, `Evidence_Linkage`,
# `Evidence_Detail` or `Evidence_Register`. Reading one spelling left every
# one of them with zero evidence rows and a NULL linked-evidence counter.

@pytest.mark.parametrize("tab", ["Evidence_Master", "Evidence_Index",
                                 "Evidence_Register", "Evidence_Detail",
                                 "Evidence_Linkage_Matrix"])
def test_zions_evidence_ledger_under_every_shipped_tab_name(tmp_path, tab):
    from dma_worker.workbook_parser import parse_evidence_master
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab
    ws.append(["Evidence_ID", "Source", "URL", "Tier", "Claim_Type"])
    ws.append(["E-001", "Fake Bank 2026 annual report", "https://x", "T2", "FACT"])
    ws.append(["INT-BOARD-003", "Board minutes", None, "T2", "FACT"])
    path = tmp_path / "ev.xlsx"
    wb.save(path)
    obs = []
    out = parse_evidence_master(str(path), obs)
    assert [e["e_id"] for e in out] == ["E-001", "INT-BOARD-003"]
    # Nothing went wrong with the TAB or the ids under any spelling.
    assert [o.kind for o in obs if o.kind != "column_not_found"] == []
    # The column_not_found observations that do appear are correct and are the
    # point: this fixture carries five columns, so the ledger's other fields
    # genuinely have no column, and each one now says so rather than landing
    # as a null indistinguishable from a column of blanks.
    missed = {o.detail["field"] for o in obs if o.kind == "column_not_found"}
    assert missed == {"ers", "published", "recency", "fact_count", "subcaps",
                      "excerpt"}
    assert not missed & {"e_id", "source_name", "source_url", "tier",
                         "claim_type"}, "a column that IS present must not be "\
                                        "reported missing"


def test_a_package_with_no_evidence_tab_names_what_it_looked_for(tmp_path):
    from dma_worker.workbook_parser import parse_evidence_master
    wb = openpyxl.Workbook()
    wb.active.title = "Executive_Summary"
    path = tmp_path / "ev.xlsx"
    wb.save(path)
    obs = []
    assert parse_evidence_master(str(path), obs) == []
    assert obs[0].kind == "evidence_ledger_tab_not_found"
    assert "Evidence_Master" in obs[0].detail["expected_any_of"]
    assert obs[0].detail["tabs_present"] == ["Executive_Summary"]


def test_a_ledger_whose_ids_are_all_unrecognised_is_not_an_empty_ledger(tmp_path):
    from dma_worker.workbook_parser import parse_evidence_master
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidence_Master"
    ws.append(["Evidence_ID", "Source", "Tier"])
    ws.append(["EV0001", "Fake source", "T2"])
    ws.append(["EV0002", "Fake source two", "T3"])
    path = tmp_path / "ev.xlsx"
    wb.save(path)
    obs = []
    assert parse_evidence_master(str(path), obs) == []
    # By kind, not by position: the reader also names the columns this
    # three-column fixture does not carry, and those observations are emitted
    # at column-resolution time, before any row is read.
    named = [o for o in obs if o.kind == "evidence_ledger_ids_unrecognised"]
    assert len(named) == 1
    assert named[0].detail["rows_seen"] == 2


# ── class 16: ledger columns read under one spelling, misses unrecorded ───
# The tab-name fix above (class 8) got the reader to the right sheet. Inside
# it, two columns were still read under one spelling each:
#
#   url_or_citation   the dma-assessment skill's OWN published column name
#                     (skills/dma-assessment/templates/
#                     01_evidence_index_template.json). Packages built from
#                     that template landed every row with source_url NULL.
#   date_published    the name used by the dma-research CSV schema, by
#                     evidence_index_schema.json, by vet_workbooks.py, and by
#                     this very module's linkage-matrix reader — everywhere
#                     except here.
#
# The date is the worse of the two: an unresolved date bands the row
# UNVERIFIED, which weights 1.0, tied with ARCHIVAL. Every row of such a
# package reads as equally worthless while the workbook holds the dates.
#
# And neither miss was recorded, which is what let both survive: `_pick`
# returning None looked exactly like a column of blanks.

def _ledger(tmp_path, header, rows, tab="Evidence_Master"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = tab
    ws.append(header)
    for r in rows:
        ws.append(r)
    path = tmp_path / "ev.xlsx"
    wb.save(path)
    return str(path)


def test_the_assessment_template_url_column_is_read(tmp_path):
    from dma_worker.workbook_parser import parse_evidence_master
    path = _ledger(tmp_path,
                   ["Evidence_ID", "Source_Name", "URL or Citation", "Tier"],
                   [["E-001", "Fake Bank 2026 annual report",
                     "https://example.test/ar2026", "T2"]])
    obs = []
    out = parse_evidence_master(path, obs)
    assert out[0]["source_url"] == "https://example.test/ar2026"
    assert "source_url" not in {o.detail["field"] for o in obs
                                if o.kind == "column_not_found"}


def test_the_research_schema_date_column_is_read(tmp_path):
    from dma_worker.workbook_parser import parse_evidence_master
    path = _ledger(tmp_path,
                   ["Evidence_ID", "Source_Name", "Date_Published", "Tier"],
                   [["E-001", "Fake Bank 2026 annual report", "2026-03-15",
                     "T2"]])
    out = parse_evidence_master(path, [])
    assert out[0]["published_date"] is not None, (
        "an unread date bands the row UNVERIFIED at weight 1.0, tied with "
        "ARCHIVAL — indistinguishable from having no evidence at all")
    assert out[0]["published_date"].isoformat() == "2026-03-15"


def test_a_publication_date_outranks_a_bare_date_column(tmp_path):
    """At fact grain a bare `date` carries EVENT dates — E-083's 1979 is a
    timeline fact, not when its source was published. evidence_normalize.py
    already ranks the aliases this way; this reader now does too."""
    from dma_worker.workbook_parser import parse_evidence_master
    path = _ledger(tmp_path,
                   ["Evidence_ID", "Source_Name", "Date", "Date_Published"],
                   [["E-001", "Fake source", "1979-01-01", "2026-03-15"]])
    out = parse_evidence_master(path, [])
    assert out[0]["published_date"].isoformat() == "2026-03-15"


def test_a_missing_ledger_column_is_named_with_what_it_costs(tmp_path):
    """The observation is the fix. Without it a null column and a column of
    nulls are the same thing downstream, and the next spelling cannot be added
    without opening the workbook first."""
    from dma_worker.workbook_parser import parse_evidence_master
    path = _ledger(tmp_path, ["Evidence_ID", "Source_Name", "Tier"],
                   [["E-001", "Fake source", "T2"]])
    obs = []
    parse_evidence_master(path, obs)
    misses = {o.detail["field"]: o.detail for o in obs
              if o.kind == "column_not_found"}
    assert "published" in misses and "source_url" in misses
    for field, detail in misses.items():
        assert detail["tab"] == "Evidence_Master"
        assert detail["expected_any_of"], "must name the spellings it accepts"
        assert detail["headers_present"], "must name what it actually found"
        assert detail["consequence"], "a miss without its cost cannot be triaged"
    assert "UNVERIFIED" in misses["published"]["consequence"]
    assert "date_published" in misses["published"]["expected_any_of"]
    assert "url_or_citation" in misses["source_url"]["expected_any_of"]


def test_the_cost_table_covers_every_alias_so_a_new_one_cannot_be_silent(tmp_path):
    """A field added to _EV_ALIASES without a cost entry would raise KeyError
    on the miss path — which only runs when a column is absent, so it would
    ship green and fail on the first package that lacked it."""
    from dma_worker.workbook_parser import _EV_ALIASES, _EV_MISS_COST
    assert set(_EV_ALIASES) | {"e_id"} == set(_EV_MISS_COST)


# ── class 14: packages grouped by folder NAME, not folder id ──────────────
# The production intake tree carries two distinct folders both called
# "Corporate America Credit Union - DMA" — different ids, different scoring
# workbooks, different reports. Grouping by name merged them and one client's
# workbook was silently discarded on every firing.

def test_corporate_america_two_folders_one_name_stay_two_packages():
    import job_main
    from dma_worker.scan_diff import FileStat

    def f(fid, folder_id, name):
        return FileStat(fid, ("Fake Credit Union - DMA", name), name,
                        "abc", 10, "", (folder_id,))

    tree = [f("a1", "folder-A", "DMA_Scoring_Workbook_A.xlsx"),
            f("a2", "folder-A", "run_manifest.json"),
            f("b1", "folder-B", "Scoring_Workbook_B.xlsx")]
    key = job_main.package_key(tree)
    assert key.collisions == {"Fake Credit Union - DMA": ["folder-A", "folder-B"]}
    groups = job_main._package_groups(tree, key)
    assert len(groups) == 2
    workbooks = sorted(g["workbook"].file_id for g in groups.values())
    assert workbooks == ["a1", "b1"], "neither workbook may be discarded"
    assert all("folder-" in k for k in groups), "the key names the folder id"


def test_a_folder_name_that_is_unique_keeps_its_plain_name():
    """`runs.source_folder_id` stores this string and the backfill matches on
    it, so the common case must not change shape."""
    import job_main
    from dma_worker.scan_diff import FileStat

    tree = [FileStat("a1", ("Fake Bank - DMA", "DMA_Scoring_Workbook.xlsx"),
                     "DMA_Scoring_Workbook.xlsx", "abc", 10, "", ("folder-A",))]
    groups = job_main._package_groups(tree, job_main.package_key(tree))
    assert list(groups) == ["Fake Bank - DMA"]


# ── class 15: _ingest_one is not idempotent ───────────────────────────────
# `persist_package` inserted into `runs` unconditionally, and nothing on the
# run recorded which artefact produced it. `_requeue` blanks a failed
# package's stored checksum so it rescans, which makes a byte-identical
# workbook look CHANGED — so every retry minted a second run. Six entities in
# production carry duplicates that way.

class _ScriptedCursor:
    """Answers the statements persist_package issues, records them all."""

    def __init__(self, db):
        self.db = db
        self._rows = []
        self.rowcount = 0

    def execute(self, sql, params=None):
        s = " ".join(str(sql).split())
        self.db.statements.append((s, params))
        self._rows = []
        if s.startswith("SELECT id FROM entities"):
            self._rows = [("entity-1",)]
        elif s.startswith("SELECT id, run_seq, scored_cells FROM runs"):
            self._rows = list(self.db.prior_run)
        elif s.startswith("SELECT COALESCE(max(run_seq)"):
            self._rows = [(4,)]
        elif s.startswith("INSERT INTO runs"):
            self._rows = [("run-new",)]

    def fetchone(self):
        return self._rows[0] if self._rows else None

    def fetchall(self):
        return list(self._rows)


class _ScriptedConn:
    def __init__(self, prior_run=()):
        self.statements = []
        self.prior_run = prior_run
        self.commits = 0

    def cursor(self):
        return _ScriptedCursor(self)

    def commit(self):
        self.commits += 1


def _persist(conn, **kw):
    from dma_worker.persist import persist_package
    from dma_worker.workbook_parser import WorkbookParse
    return persist_package(
        conn, manifest={"institution": {"name": "Fake Bank"}},
        workbook=WorkbookParse(scores=[], observations=[], toggled_out=[]),
        source_folder_id="Fake Bank - DMA", artefact_id="wb-artefact-1",
        artefact_checksum="md5-aaaa", **kw)


def _issued(conn, prefix):
    return [s for s, _ in conn.statements if s.startswith(prefix)]


def test_a_requeued_package_at_the_same_checksum_mints_no_second_run():
    conn = _ScriptedConn(prior_run=[("run-existing", 2, 706)])
    res = _persist(conn)
    assert res.run_id == "run-existing" and res.run_seq == 2
    assert res.scored_cells == 706
    assert _issued(conn, "INSERT INTO runs") == [], \
        "an unchanged package must resolve to the run it already produced"


def test_a_changed_workbook_still_mints_a_run():
    conn = _ScriptedConn(prior_run=[])         # no run at this checksum
    res = _persist(conn)
    assert res.run_id == "run-new"
    inserts = [p for s, p in conn.statements if s.startswith("INSERT INTO runs")]
    assert len(inserts) == 1
    assert "wb-artefact-1" in inserts[0] and "md5-aaaa" in inserts[0], \
        "the run must record which artefact and which bytes produced it"


def test_a_deliberate_re_ingest_after_a_parser_fix_still_mints():
    """FORCE_FOLDER is a re-read, not a retry: the tree is unchanged and the
    extraction is not, so it must be able to mint."""
    conn = _ScriptedConn(prior_run=[("run-existing", 2, 706)])
    res = _persist(conn, remint=True)
    assert res.run_id == "run-new"
    assert len(_issued(conn, "INSERT INTO runs")) == 1


# ── class 9: the research-workbook reader pinned to one set of tab names ──
# 135 packages ship a research workbook — the evidence tier's real authority
# (ERS, publication dates, per-fact linkage, the absence register). Pinned to
# `Evidence_Linkage_Matrix` + `P<n>_Scoring_Detail`, 85 of them yielded
# nothing at all: no ledger, no links, no absences, and no word about it.

def _research_workbook(tmp_path, ledger_tab, detail_tab):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ledger_tab
    ws.append(["Evidence_ID", "Source_Name", "Tier", "ERS_Total",
               "Date_Published"])
    ws.append(["E-001", "Fake Bank 2026 annual report", "T2", 4.2, "2026-03"])
    d = wb.create_sheet(detail_tab)
    d.append(["SubCap_ID", "Evidence_IDs", "Evidence_Excerpt"])
    d.append(["P1C1.1.1", "E-001:F1",
              "[E-001:F1] The board approved a three-year digital roadmap "
              "with quarterly accountability to the technology committee."])
    path = tmp_path / "research.xlsx"
    wb.save(path)
    return str(path)


@pytest.mark.parametrize("ledger_tab", ["Evidence_Linkage_Matrix",
                                        "Evidence_Index", "Evidence_Master"])
@pytest.mark.parametrize("detail_tab", ["P1_Scoring_Detail",
                                        "P1_Subcap_Scoring", "P1"])
def test_research_workbook_under_every_shipped_tab_naming(tmp_path, ledger_tab,
                                                          detail_tab):
    from dma_worker.workbook_parser import parse_research_workbook
    obs = []
    out = parse_research_workbook(
        _research_workbook(tmp_path, ledger_tab, detail_tab), obs)
    assert [e["e_id"] for e in out["ledger"]] == ["E-001"]
    assert out["ledger"][0]["ers"] is not None
    assert out["ledger"][0]["published_date"] is not None
    assert out["links"][0]["subcap_id"] == "P1C1.1.1"
    assert out["ledger"][0]["excerpt"].startswith("The board approved")
    assert obs == []


def test_a_research_workbook_that_yields_nothing_names_its_tabs(tmp_path):
    from dma_worker.workbook_parser import parse_research_workbook
    wb = openpyxl.Workbook()
    wb.active.title = "Coverage_Summary"
    path = tmp_path / "research.xlsx"
    wb.save(path)
    obs = []
    out = parse_research_workbook(str(path), obs)
    assert not any(out.values())
    assert obs[0].kind == "research_workbook_yielded_nothing"
    assert obs[0].detail["tabs_present"] == ["Coverage_Summary"]
    assert "Evidence_Linkage_Matrix" in obs[0].detail["expected_ledger_any_of"]


def test_a_caps_log_of_placeholder_rows_is_not_an_absence_of_caps(tmp_path):
    """Most shipped caps logs hold one row that says the tab is empty. Zero
    caps read from a tab with rows in it is a named observation, not a fact
    about the assessment's safeguards."""
    from dma_worker.workbook_parser import parse_research_workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caps_Applied_Log"
    ws.append(["SubCap_ID", "Cap_Type", "Cap_Reason", "Pre_Cap_Score",
               "Post_Cap_Score"])
    ws.append(["(empty — Layer 1 dma-assessment populates this)",
               None, None, None, None])
    path = tmp_path / "research.xlsx"
    wb.save(path)
    obs = []
    out = parse_research_workbook(str(path), obs)
    assert out["caps"] == []
    assert [o.kind for o in obs] == ["caps_log_ids_unrecognised",
                                     "research_workbook_yielded_nothing"]


def test_a_caps_row_naming_several_cells_caps_each_of_them(tmp_path):
    from dma_worker.workbook_parser import parse_research_workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caps_Applied_Log"
    ws.append(["SubCaps_Affected", "Cap_Type", "Cap_Rule", "Original_Score",
               "Cap_Ceiling"])
    ws.append(["P1C1.1.1, P1C1.1.2", "TIER_CEILING", "T4 evidence only",
               4.0, 3.0])
    path = tmp_path / "research.xlsx"
    wb.save(path)
    out = parse_research_workbook(str(path))
    assert [c["subcap_id"] for c in out["caps"]] == ["P1C1.1.1", "P1C1.1.2"]
    assert out["caps"][0]["cap_reason"] == "T4 evidence only"
    assert out["caps"][0]["pre_cap_score"] == 4.0
    assert out["caps"][0]["post_cap_score"] == 3.0
