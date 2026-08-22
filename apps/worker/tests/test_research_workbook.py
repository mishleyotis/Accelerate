"""The research workbook is the evidence tier's authority.

Excluding it as a decoy is what left every ingested evidence item undated
(published_date 0/82), unranked (ers 0/82), banded UNVERIFIED, and carrying an
excerpt scraped out of a rationale — median 80 characters against a contract
floor of 50–500. It also holds the per-cell linkage at FACT grain and the
register of searches that found nothing.

These build the shipped tab shapes from the BCU package and assert the parser
takes what is there, and that a score is never taken from it: the scoring
workbook is the only authority for a score.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dma_worker.workbook_parser import (_rw_split_excerpt, _stat_key,
                                        parse_peer_benchmarks,
                                        parse_recommendations,
                                        parse_research_workbook)

_DETAIL_HEADERS = ["Category_ID", "Category_Name", "Cap_ID", "Capability",
                   "SubCap_ID", "SubCapability", "Tier", "Diagnostic_Question",
                   "Weight_Pct", "Score_1_to_5", "Evidence_IDs", "Evidence_URLs",
                   "Evidence_Tier", "Confidence", "Caps_Applied", "Final_Score",
                   "Prior_Score", "Scoring_Rationale", "Proof_Claims",
                   "Proof_Links", "Evidence_Excerpt", "Source_Document"]

_EXCERPT = ("[ERS: 4.20] [FACT] [E-012:F1] BCU 2024 Annual Report (PDF) "
            "(T2, CURRENT): Board committees: Technology Committee (Paul Martin "
            "chair, 7 members), Supervisory Committee, Nominating Governance "
            "Committee [E-012:F2] NPS 79.81 in 2024 (up from 78.62 in 2023), "
            "nearly double large national banks")


def _research_wb(tmp_path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P1_Scoring_Detail"
    ws.append(_DETAIL_HEADERS)
    row = [None] * len(_DETAIL_HEADERS)
    row[0], row[4], row[5] = "P1C1", "P1C1.1.1", "Board oversight"
    row[9] = 4.0                      # Score_1_to_5 — must be IGNORED
    row[10] = "E-012:F1, E-012:F3, E-012:F2"
    row[11] = "https://www.bcu.org/annual-report-2024.pdf"
    row[12] = "T2"
    row[20], row[21] = _EXCERPT, "BCU 2024 Annual Report (PDF), CURRENT."
    ws.append(row)

    m = wb.create_sheet("Evidence_Linkage_Matrix")
    m.append(["Evidence_ID", "Source_Name", "Source_URL", "Tier", "Recency",
              "ERS_Total", "Date_Published", "Fact_Count", "SubCap_Mappings",
              "Claim_Types", "Signal_Direction", "Batch"])
    m.append(["E-012", "BCU 2024 Annual Report (PDF)",
              "https://www.bcu.org/annual-report-2024.pdf", "T2", "CURRENT",
              "4.2", "2025-03", "3", "P1C1.1.1, P1C1.1.2", "FACT", "POSITIVE", "1"])
    m.append(["E-001", "NCUSO.org NCUA Data", "https://ncuso.org/x", "T1",
              "CURRENT", "4.5", "2025-09", "2", "ENTITY_PROFILE", "FACT",
              "POSITIVE", "1"])

    a = wb.create_sheet("Absent_Evidence_Log")
    a.append(["SubCap_ID", "SubCapability", "Diagnostic_Question", "Search_Count",
              "Tiers_Searched", "Highest_Tier_Found", "Proxy_Attempts", "Reason",
              "Discovery_Question", "Impact_Note"])
    a.append(["P1C3.x.7", "IP/patents", None, "6+", "T1-T10", "NO_EVIDENCE",
              "YES", "NO_EVIDENCE", "INT-020: patents?", "Internal validation"])

    p = tmp_path / "research.xlsx"
    wb.save(p)
    return str(p)


def test_ledger_supplies_the_ers_and_date_the_scoring_workbook_omits(tmp_path):
    out = parse_research_workbook(_research_wb(tmp_path))
    by_id = {x["e_id"]: x for x in out["ledger"]}
    assert set(by_id) == {"E-012", "E-001"}
    e = by_id["E-012"]
    assert e["ers"] is not None and float(e["ers"]) == 4.2
    assert e["published_date"].isoformat().startswith("2025-03")
    assert e["tier"] == "T2" and e["stated_recency"] == "CURRENT"
    assert e["fact_count"] == 3 and e["claim_type"] == "FACT"
    # ENTITY_PROFILE is a corpus token, not a cell id, and must not become one
    assert by_id["E-001"]["subcaps"] == []
    assert by_id["E-012"]["subcaps"] == ["P1C1.1.1", "P1C1.1.2"]


def test_links_are_fact_grain_and_carry_a_verbatim_passage(tmp_path):
    out = parse_research_workbook(_research_wb(tmp_path))
    assert len(out["links"]) == 1
    link = out["links"][0]
    assert link["subcap_id"] == "P1C1.1.1"
    assert link["fact_ids"] == ["E-012:F1", "E-012:F3", "E-012:F2"]
    assert link["e_ids"] == ["E-012"]          # deduped to item grain too
    assert link["urls"] == ["https://www.bcu.org/annual-report-2024.pdf"]
    assert set(link["excerpts"]) == {"E-012:F1", "E-012:F2"}


def test_a_score_is_never_taken_from_the_research_workbook(tmp_path):
    """Score_1_to_5 is populated in the fixture and must appear nowhere."""
    out = parse_research_workbook(_research_wb(tmp_path))
    assert "score" not in out["links"][0]
    assert not any("score" in k for k in out["links"][0])


def test_the_absence_register_lands_with_its_ladder(tmp_path):
    out = parse_research_workbook(_research_wb(tmp_path))
    assert len(out["absent"]) == 1
    a = out["absent"][0]
    # the id carries an 'x' placeholder rather than a real cell number, so it
    # is kept verbatim for the alert queue rather than silently dropped
    assert a["subcap_id"] == "P1C3.x.7"
    assert a["reason"] == "NO_EVIDENCE"
    assert a["search_count"] == "6+" and a["tiers_searched"] == "T1-T10"


def test_excerpt_split_strips_provenance_and_keeps_the_quotation():
    got = _rw_split_excerpt(_EXCERPT)
    assert set(got) == {"E-012:F1", "E-012:F2"}
    # the "[ERS: 4.20] [FACT]" header and the "Source (T2, CURRENT):" prefix
    # are provenance, not the quotation
    assert got["E-012:F1"].startswith("Board committees:")
    assert "ERS" not in got["E-012:F1"] and "T2, CURRENT" not in got["E-012:F1"]
    assert got["E-012:F2"].startswith("NPS 79.81")
    assert _rw_split_excerpt(None) == {} and _rw_split_excerpt("no tags") == {}


# ── the three parser bugs, pinned ────────────────────────────────────────────

def test_a_bare_stat_header_is_not_a_peer(tmp_path):
    """`Median`/`P25`/`P75` were read as institutions, so 54 of BCU's 144
    peer rows arrived as peers named after quartiles."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peer_Benchmarks"
    ws.append(["Category", "CEFCU", "Alliant CU", "Median", "P25", "P75"])
    ws.append(["P1C1", 3.0, 3.5, 3.0, 2.5, 3.5])
    p = tmp_path / "s.xlsx"
    wb.save(p)
    rows = parse_peer_benchmarks(str(p))
    assert len(rows) == 1
    assert [n for n, _ in rows[0]["peers"]] == ["CEFCU", "Alliant CU"]
    assert rows[0]["stated_median"] is not None and float(rows[0]["stated_median"]) == 3.0
    # with no Category_Name column the name is null, never a peer's score
    assert rows[0]["category_name"] is None
    assert _stat_key("Median") == "median" and _stat_key("CEFCU") is None


def test_a_priority_ranked_recommendation_is_not_dropped(tmp_path):
    """Requiring a REC- prefix dropped all 8 of BCU's recommendations, whose
    first column is `Priority` = 1..8 — the platform page served none."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    ws.append(["Priority", "Capability", "Gap", "Root_Cause", "Zennify_Solution"])
    ws.append([1, "P4C1 Data Governance", -0.55, "Patchwork warehouse", "Data Cloud"])
    ws.append([2, "P4C3 Architecture", -0.81, "MuleSoft absent", "MuleSoft"])
    ws.append([3, None, None, None, None])        # spacer: id only, no content
    p = tmp_path / "s.xlsx"
    wb.save(p)
    got = parse_recommendations(str(p))
    assert [r["rec_id"] for r in got] == ["REC-1", "REC-2"]
    assert got[0]["payload"]["capability"] == "P4C1 Data Governance"
