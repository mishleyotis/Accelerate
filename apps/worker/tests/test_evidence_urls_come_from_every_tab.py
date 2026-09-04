"""Every evidence tab is read, not the first one that answers.

MEASURED 2026-09-04 from the promoted client surfaces. Golden 1 Credit
Union served 728 evidence items and 193 of them carried a URL — 497
package-origin citations (Banking Dive, The Financial Brand, a Fiserv case
study, an AML RightSource press release) rendered in the drawer as a quote
nobody can open. Baxter Credit Union, the same surface, served 154 of 154
WITH a URL.

The URLs were in the workbook the whole time. `parse_research_workbook`
walked `("Evidence_Linkage_Matrix",) + _EV_TABS` and BROKE on the first tab
that yielded headers, with the linkage matrix first because it is "the only
one carrying ERS and a publication date per item". That is true, and it is
not the only one carrying a URL: `contract.py` names `Evidence_Detail` as
the research generation's evidence table, `ledger.py` writes `Source_URL`
into it on every banked item, and a workbook shipping BOTH tabs had
Evidence_Detail read by nothing at all.

NOT FROM THE SCORING TABS. The pillar tabs do carry `Source_URLs` (column G)
beside `Evidence_IDs` (column F), and pairing them by position would be
wrong: `ledger.py` appends an id for every item but a URL only when the item
has one, and dedupes URLs. The two lists are a set and a list, not a
mapping, so position means nothing. A URL attached to the wrong evidence id
is worse than a missing one.

Run with `pytest apps/worker/tests/test_evidence_urls_come_from_every_tab.py`.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dma_worker.workbook_parser import parse_research_workbook


def _linkage(wb):
    """The tab the parser prefers: ERS and dates, and URLs on SOME rows."""
    ws = wb.create_sheet("Evidence_Linkage_Matrix")
    ws.append(["Evidence_ID", "Source_Name", "Source_URL", "Tier", "ERS",
               "Date_Published", "Claim_Type", "SubCap_IDs", "Excerpt"])
    ws.append(["E-CC-001", "Banking Dive", None, "T3", 0.72,
               "2024-05-01", "FACT", "P4C1.1.3", "A" * 60])
    ws.append(["E-CC-002", "The Financial Brand", None, "T3", 0.68,
               "2024-07-02", "FACT", "P4C1.1.6", "B" * 60])
    ws.append(["E-CC-003", "Fiserv case study", "https://fiserv.example/g1",
               "T3", 0.80, "2025-01-09", "FACT", "P2C1.1.1", "C" * 60])
    return ws


def _detail(wb):
    """contract.py's `Evidence_Detail` — a URL per banked item."""
    ws = wb.create_sheet("Evidence_Detail")
    ws.append(["E_ID", "Fact_ID", "Source_Name", "Source_URL", "Tier",
               "ERS", "Date_Published", "Claim_Type", "SubCap_IDs", "Excerpt"])
    ws.append(["E-CC-001", "F1", "Banking Dive",
               "https://bankingdive.example/golden1-ai", "T3", 0.72,
               "2024-05-01", "FACT", "P4C1.1.3", "A" * 60])
    ws.append(["E-CC-002", "F1", "The Financial Brand",
               "https://thefinancialbrand.example/g1-martech", "T3", 0.68,
               "2024-07-02", "FACT", "P4C1.1.6", "B" * 60])
    ws.append(["E-CC-003", "F1", "Fiserv case study",
               "https://fiserv.example/SOMETHING-ELSE", "T3", 0.80,
               "2025-01-09", "FACT", "P2C1.1.1", "C" * 60])
    return ws


def _book(tmp_path, *builders):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for b in builders:
        b(wb)
    # a scoring tab, so the research generation is recognised
    ws = wb.create_sheet("P4_Subcap_Scoring")
    ws.append(["SubCap_ID", "SubCap_Name", "Category", "Score", "Confidence",
               "Evidence_IDs", "Source_URLs"])
    ws.append(["P4C1.1.3", "A capability", None, None, "MED",
               "E-CC-001:F1", "https://bankingdive.example/golden1-ai"])
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return str(path)


def _by_id(ledger):
    return {r["e_id"]: r for r in ledger}


def test_a_url_only_on_evidence_detail_still_reaches_the_drawer(tmp_path):
    """GOLDEN 1'S EXACT SHAPE: both tabs present, the preferred one silent on
    URLs. Against the code that shipped, these come back None."""
    out = parse_research_workbook(_book(tmp_path, _linkage, _detail), [])
    rows = _by_id(out["ledger"])

    assert rows["E-CC-001"]["source_url"] == \
        "https://bankingdive.example/golden1-ai"
    assert rows["E-CC-002"]["source_url"] == \
        "https://thefinancialbrand.example/g1-martech"


def test_the_primary_tab_still_wins_where_it_states_one(tmp_path):
    """Fill the blanks, never overwrite. The linkage matrix states a URL for
    E-CC-003 and Evidence_Detail states a different one; the primary is the
    artefact under assessment and keeps its value."""
    out = parse_research_workbook(_book(tmp_path, _linkage, _detail), [])
    assert _by_id(out["ledger"])["E-CC-003"]["source_url"] == \
        "https://fiserv.example/g1"


def test_the_disagreement_is_recorded_rather_than_resolved_silently(tmp_path):
    obs = []
    parse_research_workbook(_book(tmp_path, _linkage, _detail), obs)
    kinds = {getattr(o, "kind", None) for o in obs}
    assert "evidence_index_merged" in kinds or \
           "evidence_index_disagreement" in kinds, \
        f"the merge left no trace: {sorted(k for k in kinds if k)}"


def test_a_workbook_with_only_the_linkage_matrix_is_unchanged(tmp_path):
    """No regression for the generation that ships one evidence tab."""
    out = parse_research_workbook(_book(tmp_path, _linkage), [])
    rows = _by_id(out["ledger"])
    assert rows["E-CC-003"]["source_url"] == "https://fiserv.example/g1"
    assert rows["E-CC-001"]["source_url"] is None, \
        "a URL nobody stated must stay absent"


def test_a_url_is_never_taken_from_the_scoring_tab_pairing(tmp_path):
    """`Evidence_IDs` and `Source_URLs` are a list and a set, not a mapping:
    ledger.py appends an id for every item and a URL only when there is one,
    and dedupes. E-CC-002 has no URL in either evidence tab here, and the
    scoring tab's single URL belongs to E-CC-001 — it must not migrate."""
    out = parse_research_workbook(_book(tmp_path, _linkage), [])
    assert _by_id(out["ledger"])["E-CC-002"]["source_url"] is None, \
        "a URL was taken from the scoring tab and attached by position"


# ── and the repair has to RUN, like the composite one ────────────────────

def test_the_scheduled_scan_runs_the_evidence_repair(monkeypatch):
    """Golden 1's 497 unlinked citations were repairable the whole time:
    `backfill_evidence` re-parses the workbook and fills ONLY nulls, and
    `parse_evidence_master` already reads every evidence tab. It sat behind
    `if os.environ.get("BACKFILL_EVIDENCE")`, which no schedule sets."""
    import job_main
    src = Path(job_main.__file__).read_text()
    scan = src[src.index("_scan_and_ingest(conn, scan_id, tree") - 4000:
               src.index("_scan_and_ingest(conn, scan_id, tree")]
    assert "backfill_evidence(conn" in scan, \
        "the scheduled scan never runs the evidence repair"
    assert "forced=False" in scan, \
        "the scheduled pass must be incremental, not a full re-read"


def test_the_evidence_work_list_is_bounded_and_stamped(monkeypatch):
    """Unbounded it re-downloads every 1.5 MB workbook every thirty minutes.
    A run with no unlinked citation is not work; a run this reader has
    already been through is not work until the reader changes."""
    import job_main

    class _Cur:
        def __init__(self): self.sql = None
        def execute(self, sql, params=None): self.sql, self.params = sql, params
        def fetchall(self): return []

    class _Conn:
        def __init__(self): self.cur = _Cur()
        def cursor(self): return self.cur
        def commit(self): pass
        def rollback(self): pass

    conn = _Conn()
    job_main.backfill_evidence(conn, "t", {}, forced=False)
    assert "i.source_url IS NULL" in conn.cur.sql, \
        "runs with every citation linked are still being re-read"
    assert "evidence_reader_pass" in conn.cur.sql, \
        "the work list is not stamped with the reader that did it"
    assert job_main.evidence_reader_fingerprint() in conn.cur.params


def test_a_better_evidence_reader_reopens_every_run(monkeypatch):
    import job_main
    from dma_worker import workbook_parser as wp
    before = job_main.evidence_reader_fingerprint()
    monkeypatch.setattr(wp, "_EV_TABS", wp._EV_TABS + ("Evidence_Extra",))
    assert job_main.evidence_reader_fingerprint() != before, \
        "the fingerprint ignored a change to the reader's own tab list"
