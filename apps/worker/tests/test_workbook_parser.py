"""Stage 1.3 QA bullets as tests, on a SANITIZED synthetic workbook built
in-test (modeled on the real Claude-DMA shape; no client data in the repo):

- Every score row carries a source cell and all four grain ids.
- A missing score is skipped, never defaulted to zero.
- scored_cells is stamped and differs from catalogue cells where toggles
  applied (toggled-out variants are not observations).
"""
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dma_worker.workbook_parser import parse_scoring_workbook


@pytest.fixture()
def synthetic_workbook(tmp_path):
    wb = openpyxl.Workbook()
    sc = wb.active
    sc.title = "2_Scorecard"
    sc["A1"] = "Maturity Scorecard"
    sc["A4"] = "Overall Effective Score"
    sc["A5"] = 2.4517
    sc["A14"] = "Subcapability scorecard"
    sc.append([])  # r15 placeholder; real header written below
    hdr = ["Pillar", "Pillar_Category", "L1_Capability", "Sub_Cap_ID", "Sub_Cap_Name",
           "Tier", "Works", "Fails", "Value", "Contradicts", "Corroborates",
           "Diagnostic_Score", "Evidence_Quality", "Descriptor_Fit_Level",
           "Descriptor_Fit_Rationale", "Contradiction_Check", "Final_Maturity_Override",
           "Override_Rationale", "Effective_Score"]
    for i, h in enumerate(hdr, 1):
        sc.cell(row=15, column=i, value=h)
    rows = [
        # scored
        ("Pillar 1", "Strategy", "Decisions", "P1C1.1.1", "Fake Board Packs", "T1",
         3, 2, 2, 3, 4, 2.65, 2.8, 3.0, "M2-ish", "Checked", None, None, 2.97),
        # attempted, unscored (ladder ran, nothing found) -> observation
        ("Pillar 2", "CX", "Onboarding", "P2C1.1.2", "Fake Onboarding", "T1",
         None, None, None, None, None, None, None, None, None, None, None, None, None),
        # toggled-out sub-vertical variant -> excluded, not an observation
        ("Pillar 2", "CX", "Channels", "P2C2.9.CU1", "Fake CU Variant", "T2",
         None, None, None, None, None, None, None, None, None, None, None, None, None),
        # unparseable score cell -> observation, row skipped
        ("Pillar 4", "Data", "Quality", "P4C1.2.1", "Fake Data Quality", "T1",
         2, 2, 1, 2, 2, 1.9, 2.0, 2.0, "M2", "Checked", None, None, "n/a"),
    ]
    for r in rows:
        sc.append(list(r))

    a = wb.create_sheet("3_Assessment")
    a["A1"] = "Diagnostic Assessment"
    ahdr = ["Question_ID", "Pillar", "Pillar_Category", "L1_Capability", "Sub_Cap_ID",
            "Sub_Cap_Name", "Tier", "Facet", "Weight", "Diagnostic_Question",
            "Curve_Placement_Guide", "Why_It_Matters", "Internal_Evidence_to_Request",
            "Public_Evidence_to_Inspect", "Public_Search_Probes", "Public_Visibility",
            "Discovery_Notes", "Evidence_References", "Facet_Maturity_Score",
            "Evidence_Quality", "Contradiction_Status", "Assessor_Rationale"]
    for i, h in enumerate(ahdr, 1):
        a.cell(row=4, column=i, value=h)
    a.append(["P1C1.1.1-W", "Pillar 1", "Strategy", "Decisions", "P1C1.1.1",
              "Fake Board Packs", "T1", "Works", 0.3, "q", "g", "w", "i", "p", "s",
              "v", "n", "E-C001, E-C002", 3.0, 4.0, "None", "Solid packs"])
    a.append(["P1C1.1.1-F", "Pillar 1", "Strategy", "Decisions", "P1C1.1.1",
              "Fake Board Packs", "T1", "Fails", 0.2, "q", "g", "w", "i", "p", "s",
              "v", "n", "E-C001", 2.0, 3.0, "None", "Some gaps"])
    a.append(["P2C1.1.2-W", "Pillar 2", "CX", "Onboarding", "P2C1.1.2",
              "Fake Onboarding", "T1", "Works", 0.3, "q", "g", "w", "i", "p", "s",
              "v", "n", None, None, None, None, "NO_EVIDENCE - ladder run"])
    a.append(["P2C2.9.CU1-W", "Pillar 2", "CX", "Channels", "P2C2.9.CU1",
              "Fake CU Variant", "T2", "Works", 0.3, "q", "g", "w", "i", "p", "s",
              "v", None, None, None, None, None, None])
    a.append(["P4C1.2.1-W", "Pillar 4", "Data", "Quality", "P4C1.2.1",
              "Fake Data Quality", "T1", "Works", 0.3, "q", "g", "w", "i", "p", "s",
              "v", "n", "E-C009", 2.0, 2.0, "None", "ok"])
    path = tmp_path / "synthetic_scoring_workbook.xlsx"
    wb.save(path)
    return str(path)


def test_every_score_row_carries_source_cell_and_all_four_grain_ids(synthetic_workbook):
    p = parse_scoring_workbook(synthetic_workbook)
    assert p.scored_cells == 1
    s = p.scores[0]
    assert s.subcap_id == "P1C1.1.1" and s.score == Decimal("2.97")
    assert s.source_cell == "2_Scorecard!S16"
    assert (s.pillar_id, s.category_id, s.capability_id) == ("P1", "P1C1", "P1C1.1")
    assert s.evidence_refs == ["E-C001", "E-C002"]
    assert len(s.facets) == 2 and s.facets[0].source_cell.startswith("3_Assessment!S")


def test_missing_score_is_skipped_never_zeroed_and_observed(synthetic_workbook):
    p = parse_scoring_workbook(synthetic_workbook)
    assert all(s.score is not None for s in p.scores)          # nothing defaulted
    missing = [o for o in p.observations if o.kind == "missing_score"]
    assert [o.subcap_id for o in missing] == ["P2C1.1.2"]
    unparseable = [o for o in p.observations if o.kind == "unparseable_cell"]
    assert [o.subcap_id for o in unparseable] == ["P4C1.2.1"]


def test_toggled_out_variant_is_not_an_observation(synthetic_workbook):
    p = parse_scoring_workbook(synthetic_workbook)
    assert p.toggled_out == ["P2C2.9.CU1"]
    assert all(o.subcap_id != "P2C2.9.CU1" for o in p.observations)
    # scored_cells (1) < the 4 catalogue rows present — the toggle explains it
    assert p.scored_cells < 4


def test_composite_read_from_its_cell_raw(synthetic_workbook):
    p = parse_scoring_workbook(synthetic_workbook)
    assert p.composite == Decimal("2.4517")                    # raw; rounded ONCE at persist
    assert p.composite_source_cell == "2_Scorecard!A5"


def test_grain_summaries_tolerate_unfamiliar_tab_shapes(tmp_path):
    """Prod regression (APG Federal Credit Union): a workbook whose
    Pillar_Summary / Category_Detail tabs carry no recognisable header row
    must yield NO stated grains — never a ValueError that sinks the whole
    package. H4's grain lock then simply has nothing to serve at that
    grain (derived values are computed or null, never defaulted)."""
    from dma_worker.workbook_parser import parse_grain_summaries

    wb = openpyxl.Workbook()
    ps = wb.active
    ps.title = "Pillar_Summary"
    ps["A1"] = "A narrative block, not a header"
    ps["B2"] = 3.1
    cd = wb.create_sheet("Category_Detail")
    cd["A1"] = "Category rollup (chart source, no Category_ID column)"
    cd["B3"] = 2.5
    path = tmp_path / "odd_shape.xlsx"
    wb.save(path)

    out = parse_grain_summaries(str(path))
    assert out == {"pillars": [], "categories": []}


def test_grain_summaries_read_stated_values_when_tabs_are_regular(tmp_path):
    from dma_worker.workbook_parser import parse_grain_summaries

    wb = openpyxl.Workbook()
    ps = wb.active
    ps.title = "Pillar_Summary"
    for i, h in enumerate(["Pillar", "Pillar_Name", "Score", "Weight_IB", "Peer_Median"], 1):
        ps.cell(row=1, column=i, value=h)
    ps.append(["P1", "Strategy, Governance & Culture", 2.1, 0.25, 2.4])
    cd = wb.create_sheet("Category_Detail")
    for i, h in enumerate(["Category_ID", "Category_Name", "Pillar", "Score", "Peer_Median"], 1):
        cd.cell(row=1, column=i, value=h)
    cd.append(["P1C1", "Fake Category", "P1", 1.9, 2.2])
    path = tmp_path / "regular.xlsx"
    wb.save(path)

    out = parse_grain_summaries(str(path))
    assert [p["pillar_id"] for p in out["pillars"]] == ["P1"]
    assert out["pillars"][0]["score"] == 2.1
    assert out["pillars"][0]["source_cell"] == "Pillar_Summary!C2"
    assert [c["category_id"] for c in out["categories"]] == ["P1C1"]
    assert out["categories"][0]["score"] == 1.9


def test_scoring_detail_generation_and_score_key_variants(tmp_path):
    """Third shipped generation: P{n}_Scoring_Detail tabs, plus the
    corpus's score-column variants (Post-Critic Score / Score_1_to_5).
    Same subcap-grain parse; M-level labels never read."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P1_Scoring_Detail"
    ws.append(["Pillar", "Category_Name", "Cap_ID", "SubCap_ID", "SubCap_Name",
               "M_Level_Label", "Score", "Evidence_IDs", "Confidence"])
    ws.append(["P1", "Strategy", "P1C1.1", "P1C1.1.1", "Fake Strategy Doc",
               "M3 (Standardized)", 3.1, "E-029:F1, E-039:F1", "HIGH"])
    ws.append([])  # blank row must not crash
    p = tmp_path / "detail_gen.xlsx"
    wb.save(p)
    out = parse_scoring_workbook(str(p))
    assert [s.subcap_id for s in out.scores] == ["P1C1.1.1"]
    assert out.scores[0].score == Decimal("3.1")
    assert out.scores[0].evidence_refs == ["E-029", "E-039"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P1_Subcap_Scoring"
    ws.append(["SubCap ID", "SubCap Name", "Pre-Critic Score", "Post-Critic Score",
               "Confidence", "Evidence IDs", "Rationale (≥150 chars)"])
    ws.append(["P1C1.1.1", "Fake", 3.5, 3.0, "HIGH", "E-074", "long enough rationale"])
    p2 = tmp_path / "critic_gen.xlsx"
    wb.save(p2)
    out2 = parse_scoring_workbook(str(p2))
    assert out2.scores[0].score == Decimal("3.0")   # post-critic wins
    assert out2.scores[0].rationale == "long enough rationale"


def test_fuzzy_date_impossible_month_is_unverified():
    from dma_worker.workbook_parser import parse_fuzzy_date
    assert parse_fuzzy_date("2025-13") is None
    assert parse_fuzzy_date("2025-07") is not None


def test_peer_grid_nonnumeric_cells_become_none(tmp_path):
    from dma_worker.workbook_parser import parse_peer_benchmarks
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peer_Benchmarks"
    ws.append(["Category", "Category_Name", "Entity_Score", "Peer_Median",
               "Fake Peer A", "Fake Peer B"])
    ws.append([])  # blank row must not crash
    ws.append(["P1C1", "Strategy", 2.5, "N/A", 2.1, "not scored"])
    p = tmp_path / "peers.xlsx"
    wb.save(p)
    out = parse_peer_benchmarks(str(p))
    assert out[0]["stated_median"] is None
    assert out[0]["peers"] == [("Fake Peer A", Decimal("2.1")), ("Fake Peer B", None)]


def test_pillar_tab_naming_variants_across_the_corpus():
    """Five naming conventions for the same tab kind appear in the intake
    tree (P1_Subcap_Scoring · P1_Scoring_Detail · P1_Scoring · P1 ·
    P1_RIAs_Broker_Dealers), and rollups/logs share the pillar prefix
    without carrying subcapability rows."""
    from dma_worker.workbook_parser import _is_pillar_tab
    for tab in ("P1_Subcap_Scoring", "P1_Scoring_Detail", "P1_Scoring", "P1",
                "P1 Scoring", "P1_RIAs_Broker_Dealers", "P4_Scoring_Detail"):
        assert _is_pillar_tab(tab), tab
    for tab in ("Category_Rollup", "Capability_Rollup", "Peer_Benchmark",
                "Calculation_Chain", "Run_Metadata", "Caps_Applied_Log",
                "Pillar_Summary", "Executive_Summary", "Gap_Priority",
                "Issues_Caps", "Critic_Log", "Evidence_Index",
                "Contradiction_Log", "Absent_Evidence_Log", "P1C1_Detail"):
        assert not _is_pillar_tab(tab), tab


def test_a_pillar_shaped_tab_without_subcap_rows_is_observed_not_fatal(tmp_path):
    """One package's only pillar tab is a rollup with no SubCap_ID column.
    It must be recorded and skipped, not raise — a workbook with one odd
    tab still has three good ones."""
    from dma_worker.workbook_parser import parse_scoring_workbook
    wb = openpyxl.Workbook()
    bad = wb.active
    bad.title = "P1"                       # pillar-shaped, rollup content
    bad.append(["Category", "Mean", "Notes"])
    bad.append(["P1C1", 3.1, "rollup only"])
    good = wb.create_sheet("P2_Scoring")
    good.append(["SubCap_ID", "SubCap_Name", "Score", "Confidence"])
    good.append(["P2C1.1.1", "Fake Onboarding", 2.5, "HIGH"])
    p = tmp_path / "mixed_tabs.xlsx"
    wb.save(p)
    out = parse_scoring_workbook(str(p))
    assert [s.subcap_id for s in out.scores] == ["P2C1.1.1"]
    assert any(o.kind == "unrecognised_pillar_tab" for o in out.observations)


def test_evidence_master_reads_the_corpus_real_column_names():
    """The shipped general_dma ledger reads
    Evidence_ID · Source · URL · Tier · Recency · Claim_Type · Fact_Count · SubCaps.
    The parser was looking for source_name / publish_date / fact_summary /
    subcaps_supported, so 75 of Baxter's 84 evidence rows landed with a tier and
    nothing else — and an evidence drawer with no source and no excerpt is not a
    drawer."""
    import openpyxl, tempfile, os
    from dma_worker.workbook_parser import parse_evidence_master
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidence_Master"
    ws.append(["Evidence_ID", "Source", "URL", "Tier", "Recency", "Claim_Type",
               "Fact_Count", "SubCaps"])
    ws.append(["E-001", "NCUSO.org NCUA Data", "https://ncuso.org/x", "T1",
               "CURRENT", "FACT", 2, "ENTITY_PROFILE"])
    with tempfile.TemporaryDirectory() as td:
        p = os.path.join(td, "wb.xlsx")
        wb.save(p)
        rows = parse_evidence_master(p)
    assert len(rows) == 1
    r = rows[0]
    assert r["source_name"] == "NCUSO.org NCUA Data"
    assert r["source_url"] == "https://ncuso.org/x"
    assert r["tier"] == "T1" and r["claim_type"] == "FACT"
    assert r["fact_count"] == 2
    # "Recency" ships a BAND word, not a date. Undated evidence is UNVERIFIED,
    # never current — so the date stays null and the stated word is recorded.
    assert r["published_date"] is None
    assert r["stated_recency"] == "CURRENT"
    # Fact_Count is a number: it must never be mistaken for the excerpt text.
    assert r["excerpt"] is None
    # ENTITY_PROFILE is not a cell id.
    assert r["subcaps"] == []


def test_excerpts_are_mined_verbatim_and_stop_at_the_next_label():
    """The excerpt text exists only inside the scoring Rationale, tagged per
    evidence id. A fragment must end at the next evidence tag OR the next
    section label — otherwise one fact's excerpt swallows the assessor's
    maturity reasoning and stops being verbatim."""
    from dma_worker.workbook_parser import mine_evidence_from_rationales

    class S:
        def __init__(self, sid, rationale, refs):
            self.subcap_id, self.rationale, self.evidence_refs = sid, rationale, refs

    text = ("[EVIDENCE]: [E-012:F1] Board committees: Technology Committee "
            "(Paul Martin chair, 7 members), Supervisory Committee. "
            "[SECOND]: [E-014:F1] Jim Block: responsible for product lines since 1995. "
            "[MATURITY]: Maps to M3 because a documented strategy exists. "
            "[GAP]: Needs industry recognition.")
    out = mine_evidence_from_rationales([S("P1C1.1.1", text, ["E-012", "E-014", "E-016"])])

    assert out["E-012"]["excerpt"].startswith("Board committees: Technology Committee")
    assert "MATURITY" not in out["E-012"]["excerpt"]
    assert out["E-014"]["excerpt"] == ("Jim Block: responsible for product lines "
                                       "since 1995")
    assert "Maps to M3" not in out["E-014"]["excerpt"], "the reasoning is not an excerpt"
    # cited but untagged: linked to the cell, with no excerpt invented for it
    assert out["E-016"]["excerpt"] is None
    assert out["E-016"]["subcaps"] == ["P1C1.1.1"]
