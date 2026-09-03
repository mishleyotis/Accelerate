"""The composite a workbook states, on the generation that has no 2_Scorecard.

Reported 2026-09-02 from the promoted client directory: Golden 1 Credit Union's
card rendered the word "maturity" over an EMPTY slot, beside its own four
pillar bars (which resolve) and beside Axos Bank, which shows 1.9.

`serving_directory.composite` is `runs.composite`, and `runs.composite` is
written once, at INSERT, from `WorkbookParse.composite`. That field was set in
exactly one place — `_parse_scorecard`, off the cell under "Overall Effective
Score" on the `2_Scorecard` tab.

`parse_scoring_workbook` has three branches and only the first reaches it:

    2_Scorecard present          -> _parse_scorecard        composite SET
    P{n}_Subcap_Scoring present  -> _parse_pillar_scoring    composite None
    Pillar_Summary only          -> rollup-only stub         composite None

So the figure was null for the whole general_dma generation, not for one
package. Golden 1's own workbook states it four times — `Pillar_Summary!C6`,
`Pillar_Rollup!C6`, `Executive_Summary` "Overall Maturity", and the OVERALL
row's weighted contribution — and no reader claimed any of them.

READ, never derived, is the property these tests pin hardest: the value must
be the one on the row labelled OVERALL, and a workbook that states none must
stay None rather than acquire the mean of the pillars above it. A derived
figure in a column whose contract says it was read is indistinguishable from
a stated one afterwards, which is the failure that outlives the empty card.

Run with `pytest apps/worker/tests/test_stated_composite.py`.
"""
import openpyxl
import pytest


def _pillar_tab(wb, title="Pillar_Summary", score_header="Weighted_Score",
                overall_label="OVERALL", overall_score=2.25):
    """Golden 1's own tab shape: four pillars then an OVERALL row, with the
    score column named `Weighted_Score` rather than `Score`."""
    ps = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ps.title = title
    for i, h in enumerate(["Pillar", "Name", score_header, "Maturity",
                           "Peer_Median", "Gap_to_Peer"], 1):
        ps.cell(row=1, column=i, value=h)
    ps.append(["P1", "Strategy, Governance & Culture", 2.40, "M2", 3.10, -0.70])
    ps.append(["P2", "Member Experience", 2.11, "M2", 3.00, -0.89])
    ps.append(["P3", "Operations, Risk & Compliance", 2.25, "M2", 3.00, -0.75])
    ps.append(["P4", "Data, Analytics & Technology", 2.25, "M2", 3.10, -0.85])
    if overall_label is not None:
        ps.append([overall_label, "Golden 1 CU (CU weights)", overall_score,
                   "M2", 3.05, -0.80])
    return ps


def _subcap_tab(wb, pillar="P1"):
    """One scored row, enough to take the general_dma branch."""
    ws = wb.create_sheet(f"{pillar}_Subcap_Scoring")
    for i, h in enumerate(["SubCap_ID", "SubCap_Name", "Effective_Score"], 1):
        ws.cell(row=1, column=i, value=h)
    ws.append([f"{pillar}C1.1.1", "A capability", 2.0])
    return ws


def test_the_stated_overall_lands_on_a_workbook_with_no_scorecard_tab(tmp_path):
    """The reported defect, at its own grain: a general_dma workbook now
    carries the composite its rollup tab states."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb)
    _subcap_tab(wb)
    path = tmp_path / "general_dma.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert "2_Scorecard" not in openpyxl.load_workbook(path).sheetnames
    assert float(out.composite) == pytest.approx(2.25), (
        "the OVERALL row states 2.25 and the header serves nothing without it")
    assert out.composite_source_cell == "Pillar_Summary!C6", (
        "the source cell must name where the figure was read, not that it was")


def test_the_composite_is_read_not_derived_from_the_pillars(tmp_path):
    """The guard that matters most. The four pillars mean 2.2525; the OVERALL
    row states 3.90, which no averaging of them produces. The stated figure
    must win, or the column silently holds a derivation."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb, overall_score=3.90)
    _subcap_tab(wb)
    path = tmp_path / "stated_wins.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert float(out.composite) == pytest.approx(3.90)
    assert abs(float(out.composite) - 2.2525) > 0.01, (
        "the mean of the four pillars must not be what lands here")


def test_a_workbook_stating_no_overall_keeps_a_null_composite(tmp_path):
    """Absent beats invented. Four pillars and no OVERALL row is a workbook
    that states no composite; the mean of the four is NOT the answer."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb, overall_label=None)
    _subcap_tab(wb)
    path = tmp_path / "no_overall.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert out.composite is None
    assert out.composite_source_cell is None


def test_an_unparseable_overall_is_null_rather_than_a_crash(tmp_path):
    """A mangled figure must not sink the package — the run lands with no
    composite and the rest of the parse intact."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb, overall_score="n/a")
    _subcap_tab(wb)
    path = tmp_path / "mangled.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert out.composite is None
    assert out.scores, "the subcap rows must still parse"


def test_the_rollup_only_generation_also_lands_its_overall(tmp_path):
    """The third branch: recognisably a DMA workbook with no subcap tabs at
    all. It still states a composite and must still serve one."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb)
    path = tmp_path / "rollup_only.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert out.scored_cells == 0
    assert float(out.composite) == pytest.approx(2.25)


def test_pillar_rollup_is_read_when_pillar_summary_is_absent(tmp_path):
    """The alias set is the grain reader's own `_GRAIN_TABS["pillars"]`, so a
    package shipping only `Pillar_Rollup` resolves the same figure."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pillar_Rollup"
    for i, h in enumerate(["pillar_id", "pillar_name", "score", "weight",
                           "peer_median"], 1):
        ws.cell(row=1, column=i, value=h)
    ws.append(["P1", "Strategy", 2.40, 0.25, 3.10])
    ws.append(["OVERALL", "Golden 1 Credit Union", 2.25, 1.0, 3.05])
    _subcap_tab(wb)
    path = tmp_path / "rollup_alias.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert float(out.composite) == pytest.approx(2.25)
    assert out.composite_source_cell == "Pillar_Rollup!C3"


def test_the_scorecard_generation_is_unchanged(tmp_path):
    """The claude_dma branch must keep reading its own cell. This change adds
    a reader to two branches; it must not move the one that worked."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    sc = wb.active
    sc.title = "2_Scorecard"
    sc["A1"] = "Overall Effective Score"
    sc["A2"] = 3.40
    sc["A4"] = "Subcapability scorecard"
    for i, h in enumerate(["Sub_Cap_ID", "Sub_Cap_Name", "Effective_Score"], 1):
        sc.cell(row=5, column=i, value=h)
    sc.append(["P1C1.1.1", "A capability", 2.0])
    # A rollup tab stating something DIFFERENT, to prove precedence.
    _pillar_tab(wb, overall_score=2.25)
    path = tmp_path / "scorecard.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert float(out.composite) == pytest.approx(3.40), (
        "2_Scorecard's own cell still wins where that tab exists")
    assert out.composite_source_cell.startswith("2_Scorecard!")


def test_every_candidate_tab_is_tried_not_just_the_first_that_exists(tmp_path):
    """goEasy Ltd., measured 2026-09-03.

    That workbook ships BOTH `Pillar_Summary` and `Pillar_Rollup`, and only
    the second carries an OVERALL row — the first stops at P4. Taking the
    first tab that merely EXISTS and giving up when it stated no overall
    returned None for a workbook that states 2.11 plainly, on a tab this
    reader already knew about.

    Golden 1's `Pillar_Summary` happens to carry the row, which is why the
    shape held until a client shipped the other arrangement. Every
    `return None, None` inside a tab means "not on THIS tab", never a
    statement about the workbook.
    """
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    ps = wb.active
    ps.title = "Pillar_Summary"
    for i, h in enumerate(["Pillar", "Pillar_Name", "Score", "Peer_Median"], 1):
        ps.cell(row=1, column=i, value=h)
    for pid, sc in (("P1", 2.09), ("P2", 2.19), ("P3", 2.00), ("P4", 2.16)):
        ps.append([pid, f"Pillar {pid}", sc, 2.5])
    # no OVERALL row here — the whole point

    pr = wb.create_sheet("Pillar_Rollup")
    for i, h in enumerate(["pillar_id", "pillar_name", "score", "weight",
                           "weighted_contribution", "peer_median"], 1):
        pr.cell(row=1, column=i, value=h)
    for pid, sc in (("P1", 2.09), ("P2", 2.19), ("P3", 2.00), ("P4", 2.16)):
        pr.append([pid, f"Pillar {pid}", sc, 0.25, sc * 0.25, 2.5])
    pr.append(["OVERALL", "Overall (equal-weighted)", 2.11, 1.0, 2.11, 2.5])

    _subcap_tab(wb)
    path = tmp_path / "two_tabs.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert float(out.composite) == pytest.approx(2.11), (
        "the reader stopped at Pillar_Summary and never looked at "
        "Pillar_Rollup, which is where this generation states its overall")
    assert out.composite_source_cell.startswith("Pillar_Rollup!"), \
        out.composite_source_cell


def test_the_first_tab_still_wins_when_it_does_carry_an_overall(tmp_path):
    """Precedence is unchanged: falling through is for a tab that states
    NOTHING, not a licence to prefer a later tab's figure."""
    from dma_worker.workbook_parser import parse_scoring_workbook

    wb = openpyxl.Workbook()
    _pillar_tab(wb, overall_score=2.25)              # Pillar_Summary, has OVERALL
    pr = wb.create_sheet("Pillar_Rollup")
    for i, h in enumerate(["pillar_id", "pillar_name", "score"], 1):
        pr.cell(row=1, column=i, value=h)
    pr.append(["P1", "Strategy", 2.40])
    pr.append(["OVERALL", "Overall", 9.99])          # must NOT be preferred
    _subcap_tab(wb)
    path = tmp_path / "both.xlsx"
    wb.save(path)

    out = parse_scoring_workbook(str(path))
    assert float(out.composite) == pytest.approx(2.25)
    assert out.composite_source_cell.startswith("Pillar_Summary!")
