"""ccg_loader tab-alias mapping + zero-row gate (2026-07 D3 remediation).

The audit root-caused "value chains 0/94": the shipped v7.0 workbooks
drifted on 15/25 canonical tab names (`21_Value_Chain_Mapping` ships as
`21_VC_Mapping_PerSubcap`) and the loader silently recorded `missing_tab`
forever. These tests pin:

  1. the TAB_ALIASES table against the REAL sheet-name inventory
     enumerated from all 4 committed v7.0 workbooks (frozen fixture);
  2. `resolve_sheet_name` exact-first / alias-order semantics;
  3. the v7.0-shaped parsers (title-block VC sheet with full-name
     subvertical columns; WIDE maturity descriptors; header-embedded
     `_R1` prior version);
  4. the loader's zero-row gate — a registered tab present in a workbook
     (canonically or via alias) that parses 0 rows must fail the run.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

# Make the workers/ package importable for tests.
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from workers.ccg_loader.canonical_tabs import (  # noqa: E402
    CANONICAL_TABS,
    TAB_ALIASES,
    expected_tab_keys,
    resolve_sheet_name,
)
from workers.ccg_loader.main import load_workbooks  # noqa: E402
from workers.ccg_loader.parsers import (  # noqa: E402
    PARSER_FOR_TAB,
    parse_alias_bridge,
    parse_maturity_bands,
    parse_value_chain,
    parse_workbook_tabs,
)

# The REAL sheet names enumerated from all 4 committed v7.0 workbooks
# (docs/reference/catalogue/v7.0/Pillar_*_v7.0.xlsx) — identical across
# pillars; P4 additionally carries `_R3_P4_Specific_Backup`.
V7_SHEET_NAMES: frozenset[str] = frozenset({
    "1_Overview", "2_Capability_Map", "3_User_Stories_Catalogue",
    "4_L3_Detailed", "5_L4_Detailed_Features", "6_Maturity_Descriptors",
    "7_Product_Catalogue", "8_Agentforce_Agents_List",
    "9_Platform_Constructs_Library", "Z1_QA_Gates", "Z2_Plan_Revisions",
    "10_Productized_Offerings", "11_Data_Products",
    "12_Offering_SubCap_Matrix", "13_DataProduct_SubCap_Matrix",
    "14_CrossPillar_Stories", "15_Theme_SubCap_Mapping",
    "16_SubCap_CrossPillar_Coverage", "17_Toggle_Control_Panel",
    "18_SubCap_Completeness_Profile", "19_Toggle_Cascade_Simulation",
    "20_Final_QA_Report", "21_VC_Mapping_PerSubcap",
    "_R1_Source_Reference", "_R2_Dropped_Stories",
    "_R3_P4_Specific_Backup",
})


class TestAliasTable:
    def test_every_alias_key_is_canonical(self) -> None:
        canonical = expected_tab_keys()
        for key in TAB_ALIASES:
            assert key in canonical, f"alias key {key!r} not a canonical tab"

    def test_every_alias_target_exists_in_v7_workbooks(self) -> None:
        for key, aliases in TAB_ALIASES.items():
            for alias in aliases:
                assert alias in V7_SHEET_NAMES, (
                    f"{key} aliases {alias!r} which is not a real v7.0 sheet"
                )

    def test_all_canonical_tabs_resolve_against_v7_except_subverticals(self) -> None:
        """Every canonical tab except the informational 19_Subverticals
        (not shipped; target None) must resolve to a real v7.0 sheet."""
        for key, _, _target in CANONICAL_TABS:
            resolved = resolve_sheet_name(key, set(V7_SHEET_NAMES))
            if key == "19_Subverticals":
                assert resolved is None
            else:
                assert resolved is not None, f"{key} does not resolve on v7.0"

    def test_value_chain_alias_is_the_persubcap_sheet(self) -> None:
        assert resolve_sheet_name(
            "21_Value_Chain_Mapping", set(V7_SHEET_NAMES),
        ) == "21_VC_Mapping_PerSubcap"

    def test_every_registered_parser_tab_resolves_on_v7(self) -> None:
        """The tabs that actually feed ccg_* tables must all be readable
        from the shipped workbooks — the drift class that silently
        zeroed value chains."""
        for tab in PARSER_FOR_TAB:
            assert resolve_sheet_name(tab, set(V7_SHEET_NAMES)) is not None

    def test_exact_name_wins_over_alias(self) -> None:
        names = {"3_Maturity_Scoring_Bands", "6_Maturity_Descriptors"}
        assert resolve_sheet_name("3_Maturity_Scoring_Bands", names) \
            == "3_Maturity_Scoring_Bands"

    def test_unresolvable_returns_none(self) -> None:
        assert resolve_sheet_name("21_Value_Chain_Mapping", {"random"}) is None


def _v7_vc_workbook() -> Workbook:
    """v7.0-shaped VC sheet: 4-row title block, header row 5, FULL-name
    subvertical columns, ▌-prefixed multi-line stage cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "21_VC_Mapping_PerSubcap"
    ws.append(["Value Chain Mapping — Per-Subcap"])
    ws.append(["Each cell shows the value chain stages"])
    ws.append(["Source: Value Chain Reference"])
    ws.append([])
    ws.append(["Category", "L1_Capability", "Sub_Cap_ID", "Sub_Cap_Name",
               "Tier", "Retail Banking", "Credit Unions"])
    ws.append(["P1C1", "Strategy", "P1C1.1.1", "Digital Strategy", "T1",
               "▌ MARKET\n▌ BACK OFFICE OPS", "▌ MEMBER SERVICING"])
    ws.append(["P1C1", "Strategy", "P1C1.1.2", "Business Alignment", "T1",
               None, "Not applicable — credit unions follow NCUA framework"])
    return wb


class TestV7ValueChainParser:
    def test_parses_full_name_columns_to_codes(self) -> None:
        wb = _v7_vc_workbook()
        res = parse_value_chain(wb["21_VC_Mapping_PerSubcap"], "v7.0", "P1")
        assert len(res.rows) == 3
        by_key = {(r["subcap_id"], r["subvertical_code"]): r for r in res.rows}
        rb = by_key[("P1C1.1.1", "RB")]
        assert rb["value_chain_stages"] == ["MARKET", "BACK OFFICE OPS"]
        cu = by_key[("P1C1.1.1", "CU")]
        assert cu["value_chain_stages"] == ["MEMBER SERVICING"]
        # The N/A footnote row still parses (filtering is a read-path
        # concern); empty cells contribute nothing.
        assert ("P1C1.1.2", "RB") not in by_key
        assert ("P1C1.1.2", "CU") in by_key

    def test_legacy_code_columns_still_parse(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "21_Value_Chain_Mapping"
        ws.append(["SubCap ID", "RB", "CU"])
        ws.append(["P1C1.1.1", "ONBOARD│SERVICE", None])
        res = parse_value_chain(ws, "v7.0", "P1")
        assert len(res.rows) == 1
        assert res.rows[0]["value_chain_stages"] == ["ONBOARD", "SERVICE"]


class TestV7MaturityDescriptors:
    def test_wide_form_expands_to_band_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "6_Maturity_Descriptors"
        ws.append(["Sub_Cap_ID", "Sub_Cap_Name", "Category", "L1_Capability",
                   "M1_Foundational", "M1_Foundational_Features",
                   "M2_Developing", "M2_Developing_Features",
                   "M3_Established_AI_Assisted", "M3_Agentic_Features",
                   "M4_Advanced_Hybrid_Agentic", "M4_Cross_System",
                   "M5_Transformational_Headless", "M5_Headless_360",
                   "Zennify_Effective_Status"])
        ws.append(["P1C1.1.1", "Digital Strategy", "P1C1", "Strategy",
                   "doc exists", "• artifacts", "captured", "• custom",
                   "data-grounded", "• research", "operational", "• align",
                   "autonomous", "• pubsub", "Active"])
        res = parse_maturity_bands(ws, "v7.0", "P1")
        assert len(res.rows) == 5
        bands = {r["band"]: r for r in res.rows}
        assert set(bands) == {"M1", "M2", "M3", "M4", "M5"}
        assert bands["M1"]["narrative"] == "doc exists"
        assert bands["M1"]["features"] == "• artifacts"
        assert bands["M5"]["narrative"] == "autonomous"


class TestV7AliasBridge:
    def test_header_embedded_prior_version(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "_R1_Source_Reference"
        ws.append(["Sub_Cap_ID (post-rename)", "Sub_Cap_ID (v5.0 original)",
                   "v5.0_Pillar_Weight"])
        ws.append(["P1C1.1.1", "P1C1.1.1", "20%"])
        ws.append(["P1C1.3.AM1", "P1C1.3.WM1", "20%"])
        res = parse_alias_bridge(ws, "v7.0", "P1")
        assert len(res.rows) == 2
        identity, renamed = res.rows
        assert identity["migration_action"] == "MIGRATED"
        assert identity["prior_version"] == "v5.0"
        assert renamed["migration_action"] == "RENAMED"
        assert renamed["prior_subcap_id"] == "P1C1.3.WM1"
        assert renamed["current_subcap_id"] == "P1C1.3.AM1"


class TestAliasDispatch:
    def test_parse_workbook_tabs_resolves_via_alias(self) -> None:
        wb = _v7_vc_workbook()
        out = parse_workbook_tabs(
            wb, version="v7.0", pillar_id="P1",
            tab_names=["21_Value_Chain_Mapping"],
        )
        res = out["21_Value_Chain_Mapping"]
        assert len(res.rows) == 3
        assert any(w.get("kind") == "aliased_tab" for w in res.warnings)

    def test_missing_tab_still_warns(self) -> None:
        wb = Workbook()
        wb.active.title = "unrelated"
        out = parse_workbook_tabs(
            wb, version="v7.0", pillar_id="P1",
            tab_names=["21_Value_Chain_Mapping"],
        )
        assert any(w.get("kind") == "missing_tab"
                   for w in out["21_Value_Chain_Mapping"].warnings)


class TestZeroRowGate:
    def _write_pillar_workbook(self, tmp_path: Path, *, empty_vc: bool) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2_Capability_Map"
        ws.append(["Category", "L1_Capability", "Sub_Cap_ID", "Sub_Cap_Name",
                   "Description", "Solution_Type", "Tier"])
        ws.append(["P1C1", "Strategy", "P1C1.1.1", "Digital Strategy",
                   "desc", "Hybrid", "T1"])
        vc = wb.create_sheet("21_VC_Mapping_PerSubcap")
        vc.append(["Value Chain Mapping — Per-Subcap"])
        vc.append([])
        vc.append([])
        vc.append([])
        vc.append(["Category", "L1_Capability", "Sub_Cap_ID", "Sub_Cap_Name",
                   "Tier", "Retail Banking"])
        if not empty_vc:
            vc.append(["P1C1", "Strategy", "P1C1.1.1", "Digital Strategy",
                       "T1", "▌ MARKET"])
        wb.save(tmp_path / "Pillar_1_Comprehensive_Capability_Mapping_v7.0.xlsx")

    def test_gate_fails_on_present_but_zero_row_tab(self, tmp_path: Path) -> None:
        self._write_pillar_workbook(tmp_path, empty_vc=True)
        result = load_workbooks(version="v7.0", workbooks_dir=tmp_path)
        assert "P1/21_Value_Chain_Mapping" in result.zero_row_gate_failures

    def test_gate_clean_when_rows_parse(self, tmp_path: Path) -> None:
        self._write_pillar_workbook(tmp_path, empty_vc=False)
        result = load_workbooks(version="v7.0", workbooks_dir=tmp_path)
        assert "P1/21_Value_Chain_Mapping" not in result.zero_row_gate_failures
        assert result.tab_row_counts["21_Value_Chain_Mapping"]["P1"] == 1
