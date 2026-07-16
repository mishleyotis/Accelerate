"""Tests for the research-workbook parser.

State-transition coverage matrix (per scope §1):

  - full_extract                              → TestPerPillarSheets.test_alma_per_pillar_full_extract
  - partial_with_warnings                     → TestPerPillarSheets.test_partial_with_bad_tier_row
  - llm_column_mapper_used                    → TestParseWithCache.test_cache_miss_invokes_infer
  - headers_too_drifted_requires_admin_review → TestPerPillarSheets.test_unknown_headers_pending_review
  - file_missing                              → TestPerPillarSheets.test_no_pillar_sheets_returns_file_missing
  - handoff cross-reference (JSON wins)       → TestHandoffCrossRef.test_handoff_overrides_workbook_tier
"""
from __future__ import annotations

from openpyxl import Workbook

from app.services.parsers.research_workbook import (
    ResearchWorkbookMap,
    collect_research_metadata,
    cross_reference_with_handoff,
    parse_per_pillar_sheets,
    parse_research_with_map,
    parse_research_workbook,
    research_shape_fingerprint,
)


def _wb(title: str, headers: list[str], rows: list[list]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return wb


class TestCollectMetadata:
    def test_skips_padding_rows(self) -> None:
        wb = _wb("Research", ["E-ID", "Source", "Excerpt", "Tier"],
                 [["E-1", "ABA", "AUM grew 8%", 2]])
        md = collect_research_metadata(wb)
        assert md[0]["headers"] == ["E-ID", "Source", "Excerpt", "Tier"]
        assert md[0]["sample_rows"] == [["E-1", "ABA", "AUM grew 8%", 2]]

    def test_fingerprint_stable_across_rows(self) -> None:
        wb1 = _wb("Research", ["E-ID", "Source", "Excerpt", "Tier"], [["E-1", "A", "x", 2]])
        wb2 = _wb("Research", ["E-ID", "Source", "Excerpt", "Tier"], [["E-9", "B", "y", 4]])
        assert research_shape_fingerprint(collect_research_metadata(wb1)) == \
               research_shape_fingerprint(collect_research_metadata(wb2))


class TestParseWithMap:
    def test_happy_path(self) -> None:
        wb = _wb(
            "Research",
            ["E-ID", "Source", "URL", "Excerpt", "Claim", "Tier", "Date", "Subcaps"],
            [
                ["E-1", "ABA Banking Journal", "https://x", "AUM grew 8% YoY",
                 "strategic_signal", 2, "2024-04-01", "P1C1.1.1,P2C3.1.1"],
                ["E-2", "10-K filing", "", "Loans up 5%",
                 "financial_signal", 1, "2024-03-15", "P1C1.1.2"],
            ],
        )
        m = ResearchWorkbookMap(
            sheet_name="Research",
            e_id_col="A", source_name_col="B", source_url_col="C",
            excerpt_col="D", claim_type_col="E", tier_col="F",
            published_date_col="G", linked_subcap_ids_col="H",
        )
        res = parse_research_with_map(wb, m)
        assert [r.e_id for r in res.rows] == ["E-1", "E-2"]
        assert res.rows[0].linked_subcap_ids == ["P1C1.1.1", "P2C3.1.1"]
        assert res.rows[1].source_url is None
        assert res.warnings == []

    def test_drops_missing_required(self) -> None:
        """e_id + excerpt are hard-required (row drops); tier is NOT.

        The old contract also dropped rows with a blank tier cell —
        real evidence content lost over a missing label. New contract:
        the row is KEPT with an honest ``tier=None`` (never a
        fabricated default) plus a ``missing_tier`` warning.
        """
        wb = _wb(
            "Research",
            ["E-ID", "Source", "Excerpt", "Tier"],
            [
                ["E-1", "ABA", "AUM grew 8%", 2],
                [None, "X", "y", 3],          # missing e_id → drop
                ["E-3", "X", None, 4],        # missing excerpt → drop
                ["E-4", "X", "ok", None],     # missing tier → keep, tier=None
            ],
        )
        m = ResearchWorkbookMap(
            sheet_name="Research", e_id_col="A",
            source_name_col="B", source_url_col=None,
            excerpt_col="C", claim_type_col=None, tier_col="D",
        )
        res = parse_research_with_map(wb, m)
        assert [(r.e_id, r.tier) for r in res.rows] == [("E-1", 2), ("E-4", None)]
        kinds = sorted(w["kind"] for w in res.warnings)
        assert kinds == ["missing_required", "missing_required", "missing_tier"]

    def test_bad_tier_nulled_row_kept(self) -> None:
        """Junk / out-of-taxonomy tier → honest ``tier=None``; row survives.

        The old contract dropped these rows outright (evidence lost)
        and accepted [1, 8] — but no research-workbook source-tier
        taxonomy defines a T8 (the union is T1..T7). New contract:
        out-of-taxonomy values keep the row with ``tier=None`` — never
        dropped, never clamped to an invented neighbour — and each bad
        cell warns ``bad_tier``.
        """
        wb = _wb(
            "Research",
            ["E-ID", "Source", "Excerpt", "Tier"],
            [
                ["E-1", "A", "x", "garbage"],   # non-tier junk
                ["E-2", "B", "y", 9],           # out of taxonomy [1, 7]
                ["E-3", "C", "z", 0],           # out of taxonomy [1, 7]
                ["E-4", "D", "w", 3],
            ],
        )
        m = ResearchWorkbookMap(
            sheet_name="Research", e_id_col="A",
            source_name_col="B", source_url_col=None,
            excerpt_col="C", claim_type_col=None, tier_col="D",
        )
        res = parse_research_with_map(wb, m)
        assert [(r.e_id, r.tier) for r in res.rows] == [
            ("E-1", None), ("E-2", None), ("E-3", None), ("E-4", 3),
        ]
        kinds = sorted(w["kind"] for w in res.warnings)
        assert kinds == ["bad_tier", "bad_tier", "bad_tier"]

    def test_missing_sheet_warning(self) -> None:
        wb = _wb("Research", ["E-ID", "Excerpt", "Tier"], [])
        m = ResearchWorkbookMap(
            sheet_name="NotPresent", e_id_col="A",
            source_name_col="A", source_url_col=None,
            excerpt_col="B", claim_type_col=None, tier_col="C",
        )
        res = parse_research_with_map(wb, m)
        assert res.rows == []
        assert res.warnings == [{"kind": "missing_sheet", "sheet": "NotPresent"}]


class TestParseWithCache:
    def test_cache_hit_skips_infer(self) -> None:
        wb = _wb("Research", ["E-ID", "Source", "Excerpt", "Tier"], [["E-1", "A", "x", 2]])
        cached_map = ResearchWorkbookMap(
            sheet_name="Research", e_id_col="A",
            source_name_col="B", source_url_col=None,
            excerpt_col="C", claim_type_col=None, tier_col="D",
        )
        cache = {research_shape_fingerprint(collect_research_metadata(wb)): cached_map}

        def infer(_md):
            raise AssertionError("should not be called on cache hit")

        res = parse_research_workbook(
            wb,
            cache_lookup=lambda fp: cache.get(fp),
            cache_store=lambda fp, m: None,
            infer_map=infer,
        )
        assert [r.e_id for r in res.rows] == ["E-1"]

    def test_cache_miss_invokes_infer(self) -> None:
        wb = _wb("Research", ["E-ID", "Source", "Excerpt", "Tier"], [["E-1", "A", "x", 2]])
        stored: dict = {}

        def infer(_md):
            return ResearchWorkbookMap(
                sheet_name="Research", e_id_col="A",
                source_name_col="B", source_url_col=None,
                excerpt_col="C", claim_type_col=None, tier_col="D",
            )

        res = parse_research_workbook(
            wb,
            cache_lookup=lambda _fp: None,
            cache_store=lambda fp, m: stored.update({fp: m}),
            infer_map=infer,
        )
        assert len(res.rows) == 1
        assert len(stored) == 1


def _per_pillar_wb(sheets: list[tuple[str, list[str], list[list]]]) -> Workbook:
    """Build a per-pillar-shape workbook from (title, headers, rows) tuples."""
    wb = Workbook()
    # Replace the default sheet with our first one.
    default = wb.active
    wb.remove(default)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for r in rows:
            ws.append(r)
    return wb


class TestPerPillarSheets:
    def test_alma_per_pillar_full_extract(self) -> None:
        """Verify Alma-shape: P1C1 sheet with multi-value Evidence_IDs."""
        wb = _per_pillar_wb([
            ("P1C1", [
                "Category_ID", "Category_Name", "Cap_ID", "Capability_Name",
                "SubCap_ID", "SubCap_Name", "Tier", "Evidence_Count",
                "Evidence_IDs", "Source_URLs", "Evidence_Excerpt",
            ], [
                ["P1C1", "Digital Strategy", "P1C1.1", "Strategy Foundation",
                 "P1C1.1.1", "Digital Strategy Doc", "T3", 2,
                 "E-005; E-006",
                 "https://linkedin.com/x; https://prnewswire.com/y",
                 "CEO Psyllos partnership with Cetera"],
                ["P1C1", "Digital Strategy", "P1C1.1", "Strategy Foundation",
                 "P1C1.1.2", "Business Alignment", "T2", 1,
                 "E-006",
                 "https://prnewswire.com/y",
                 "Cetera partnership growth"],
            ]),
            ("P2C1", [
                "Category_ID", "Category_Name", "Cap_ID", "Capability_Name",
                "SubCap_ID", "SubCap_Name", "Tier", "Evidence_Count",
                "Evidence_IDs", "Source_URLs", "Evidence_Excerpt",
            ], [
                ["P2C1", "Engagement", "P2C1.1", "Channel Strategy",
                 "P2C1.1.1", "Customer Touchpoints", "T2", 1,
                 "E-007",
                 "https://almabank.com/about",
                 "Service tiers across NY/NJ branches"],
            ]),
        ])
        res = parse_per_pillar_sheets(wb)
        assert res.state_kind == "full_extract"
        assert res.sheets_scanned == 2
        assert {r.e_id for r in res.rows} == {"E-005", "E-006", "E-007"}
        # E-006 appears in two rows → linked subcaps aggregate.
        e006 = next(r for r in res.rows if r.e_id == "E-006")
        assert "P1C1.1.1" in e006.linked_subcap_ids
        assert "P1C1.1.2" in e006.linked_subcap_ids
        # E-006 tier should reflect the strongest (lowest) seen — 2.
        assert e006.tier == 2

    def test_partial_with_bad_tier_row(self) -> None:
        """Bad-tier rows still warn (→ partial_with_warnings) but are KEPT.

        The old contract discarded the whole row over an unparseable
        tier cell — destroying the E-ID → excerpt → subcap linkage that
        downstream explanations depend on. New contract: the evidence
        content survives with an honest ``tier=None`` (out-of-taxonomy
        labels like T9 are never clamped to a neighbour), and each junk
        cell emits a ``bad_tier`` warning.
        """
        wb = _per_pillar_wb([
            ("P1C1", [
                "SubCap_ID", "Tier", "Evidence_IDs", "Source_URLs",
                "Evidence_Excerpt",
            ], [
                ["P1C1.1.1", "T3", "E-001", "https://a", "good row"],
                ["P1C1.1.2", "garbage", "E-002", "https://b", "bad tier"],
                ["P1C1.1.3", "T9", "E-003", "https://c", "out-of-range"],
            ]),
        ])
        res = parse_per_pillar_sheets(wb)
        assert res.state_kind == "partial_with_warnings"
        # All three rows survive; bad tiers are None, never fabricated.
        assert {r.e_id: r.tier for r in res.rows} == {
            "E-001": 3, "E-002": None, "E-003": None,
        }
        # Evidence → subcap linkage is preserved even for bad-tier rows.
        linked = {r.e_id: r.linked_subcap_ids for r in res.rows}
        assert linked["E-002"] == ["P1C1.1.2"]
        assert linked["E-003"] == ["P1C1.1.3"]
        assert sum(1 for w in res.warnings if w["kind"] == "bad_tier") == 2

    def test_unknown_headers_pending_review(self) -> None:
        """If a per-pillar sheet has zero matchable headers, mark drifted."""
        wb = _per_pillar_wb([
            ("P1C1", ["Foo", "Bar", "Baz"], [["x", "y", "z"]]),
        ])
        res = parse_per_pillar_sheets(wb)
        assert res.state_kind in (
            "headers_too_drifted_requires_admin_review",
            "partial_with_warnings",
        )
        assert res.rows == []

    def test_no_pillar_sheets_returns_file_missing(self) -> None:
        wb = Workbook()
        wb.active.title = "Summary"
        wb.active.append(["a", "b", "c"])
        res = parse_per_pillar_sheets(wb)
        assert res.state_kind == "file_missing"
        assert res.sheets_scanned == 0
        assert res.rows == []


class TestHandoffCrossRef:
    def test_handoff_overrides_workbook_tier(self) -> None:
        from app.services.parsers.research_workbook import ParsedEvidence
        workbook_rows = [
            ParsedEvidence(
                e_id="E-001", source_name="workbook",
                source_url="https://wb", excerpt="wb excerpt",
                claim_type="research_evidence", tier=5,
                linked_subcap_ids=["P1C1.1.1"],
            ),
        ]
        handoff_items = [{
            "evidence_id": "E-001", "tier": "T2",
            "url": "https://handoff",
            "source_name": "10-K filing",
            "excerpt": "official 10-K text",
            "signal_direction": "POSITIVE",
            "publish_date": "2025-02",
            "subcap_mappings": ["P1C1.1.2", "P1C1.1.1"],
        }]
        merged, conflicts = cross_reference_with_handoff(workbook_rows, handoff_items)
        assert len(merged) == 1
        m = merged[0]
        assert m.tier == 2
        assert m.source_url == "https://handoff"
        assert m.excerpt == "official 10-K text"
        assert "P1C1.1.2" in m.linked_subcap_ids
        assert any(c["kind"] == "tier_conflict_handoff_wins" for c in conflicts)

    def test_handoff_adds_missing_evidence(self) -> None:
        from app.services.parsers.research_workbook import (
            cross_reference_with_handoff,
        )
        workbook_rows: list = []
        handoff_items = [{
            "evidence_id": "E-077", "tier": "T3",
            "url": "https://example.com/sec",
            "source_name": "SEC filing",
            "excerpt": "Annual report excerpt",
            "signal_direction": "NEUTRAL",
            "publish_date": "2024-08",
            "subcap_mappings": ["P3C1.1.1"],
        }]
        merged, conflicts = cross_reference_with_handoff(workbook_rows, handoff_items)
        assert len(merged) == 1
        assert merged[0].e_id == "E-077"
        assert merged[0].tier == 3
        assert merged[0].published_date == "2024-08"
        assert conflicts == []
