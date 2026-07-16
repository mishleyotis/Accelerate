"""Derive the D1 "Evidence & benchmarks" card surfaces (plan 4.6).

Writes three JSONB columns on the ACTIVE run (migration 045) that were
0/94 populated at audit time:

  runs.evidence_summary   → EvidenceTierCard: tier histogram + claim +
                            polarity-signal mix + connector counts, computed
                            from evidence_index (universal — every scored run
                            has an index; research_handoff's own summary block
                            was only shipped by 65/113 packages).
  runs.coverage_stats     → CoverageByPillarCard: per-pillar scored/total/thin
                            from subcap_scores (the in-scope denominators the
                            run actually assessed; the 80% gate is a UI const).
  runs.uncertainty_bands  → CeilingEstimateCard: per-category
                            {ceiling, band, modifiers[], evidence[], rationale}
                            normalized from (1) the ingested uncertainty
                            register (client_knowledge_sections artifact_kind=
                            'uncertainty' / the raw ingested array), else
                            (2) synthesized from REAL modifiers — score
                            headroom, thin-evidence density, applied caps with
                            their stated reasons — never generic filler.

The ingested `runs.uncertainty_bands` raw arrays (24 runs) are REPLACED by
the normalized dict; the raw rows remain durable in
client_knowledge_sections (verified 1:1 row parity), so nothing is lost and
the pass is idempotent. Every payload stamps `derived_from`.

NLP (plan 3.5 matrix): quantities (ceiling "L2.5 ±0.5" parsing), polarity
(signal mix), patterns (register row shapes).

Usage: DATABASE_URL=... python -m app.scripts.derive_evidence_surfaces
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import sys
from collections import Counter

from sqlalchemy import text

from app.database import get_sessionmaker
from app.services.nlp.polarity import signal as polarity_signal

# ── evidence_summary ─────────────────────────────────────────────────────────
_CONNECTOR_RE = (
    ("Explorium", re.compile(r"explorium", re.I)),
    ("Clay", re.compile(r"\bclay\b", re.I)),
    ("Indeed", re.compile(r"indeed", re.I)),
    ("LinkedIn", re.compile(r"linkedin", re.I)),
    ("Glassdoor", re.compile(r"glassdoor", re.I)),
)
# Enrichment-connector rows (post-research automation) are NOT the analyst's
# evidence trail — they inflated Bank of Utah's tier card to 95 items vs the
# workbook's 81 (2026-07-06 deploy review). They are excluded from the
# LAST-RESORT index histogram and reported under `connectors` only.
_ENRICHMENT_CONNECTORS = frozenset({"Explorium", "Clay", "Indeed"})
_TIER_KEY_RE = re.compile(r"^T[1-8]$")


def _clean_tier_dist(raw: object) -> dict[str, int] | None:
    """A {'T1': n, …} mapping with int counts, or None when malformed."""
    if not isinstance(raw, dict):
        return None
    out: dict[str, int] = {}
    for k, v in raw.items():
        key = str(k).strip().upper()
        if not _TIER_KEY_RE.match(key):
            continue
        try:
            n = int(v)
        except (TypeError, ValueError):
            continue
        if n > 0:
            out[key] = n
    return dict(sorted(out.items())) or None


def build_evidence_summary(rows: list[dict], *,
                           workbook_tiers: dict | None = None,
                           handoff_summary: dict | None = None) -> dict | None:
    """`[{tier, claim_type, excerpt, source_name, linked_n}]` → the
    EvidenceTierCard shape. None when nothing is available (honest-null).

    Tier-histogram provenance ladder (user mandate: the card shows the
    WORKBOOKS' evidence tiers, 2026-07-06 deploy review):
      1. ``workbook_tiers``   — the research workbook's own flat evidence
         sheet (research_workbook.evidence_tier_histogram);
      2. ``handoff_summary``  — research_handoff.json evidence_summary
         (matches the workbook wherever both exist);
      3. the evidence_index rows — LAST RESORT, with enrichment-connector
         rows (Explorium/Clay/Indeed) excluded from the histogram and
         listed under ``connectors`` only.
    ``derived_from`` stamps the rung that produced the histogram; claim/
    signal/connector mixes always come from the index rows (the drawer's
    own content)."""
    if not rows and not workbook_tiers and not handoff_summary:
        return None
    tiers: Counter = Counter()
    claims: Counter = Counter()
    signals: Counter = Counter()
    connectors: Counter = Counter()
    total_facts = 0
    index_items = 0
    for r in rows:
        c = str(r.get("claim_type") or "").strip().upper()
        if c:
            claims[c] += 1
        sig = polarity_signal(str(r.get("excerpt") or ""))
        signals[sig.upper()] += 1
        total_facts += int(r.get("linked_n") or 0)
        src = str(r.get("source_name") or "")
        connector = next(
            (label for label, pat in _CONNECTOR_RE if pat.search(src)), None)
        if connector:
            connectors[connector] += 1
        if connector in _ENRICHMENT_CONNECTORS:
            continue    # connector rows never enter the tier histogram
        t = r.get("tier")
        if isinstance(t, int) and 1 <= t <= 8:
            tiers[f"T{t}"] += 1
            index_items += 1
    wb_tiers = _clean_tier_dist((workbook_tiers or {}).get("tiers"))
    ho_tiers = _clean_tier_dist((handoff_summary or {}).get("tier_distribution"))
    if wb_tiers:
        tier_hist = wb_tiers
        total_items = int((workbook_tiers or {}).get("total_items")
                          or sum(wb_tiers.values()))
        derived_from = "research_workbook"
    elif ho_tiers:
        tier_hist = ho_tiers
        try:
            total_items = int((handoff_summary or {}).get("total_items")
                              or sum(ho_tiers.values()))
        except (TypeError, ValueError):
            total_items = sum(ho_tiers.values())
        with_facts = (handoff_summary or {}).get("total_facts")
        if isinstance(with_facts, int) and with_facts > 0:
            total_facts = max(total_facts, with_facts)
        derived_from = "research_handoff"
    else:
        if not tiers and not rows:
            return None
        tier_hist = dict(sorted(tiers.items()))
        total_items = index_items or len(rows)
        derived_from = "evidence_index"
    return {
        "total_items": total_items,
        "total_facts": total_facts or total_items,
        "tiers": tier_hist,
        "claims": dict(claims.most_common()),
        "signals": dict(signals.most_common()),
        "connectors": dict(connectors.most_common()),
        "derived_from": derived_from,
    }


# ── coverage_stats ───────────────────────────────────────────────────────────
def build_coverage_stats(pillar_rows: list[dict]) -> dict | None:
    """`[{pillar, subcaps, scored, thin}]` → CoverageByPillarCard shape.
    pct = scored share of the run's in-scope subcaps per pillar."""
    by_pillar = []
    tot = scr = 0
    for r in sorted(pillar_rows, key=lambda x: str(x.get("pillar"))):
        n, k = int(r.get("subcaps") or 0), int(r.get("scored") or 0)
        if n == 0:
            continue
        by_pillar.append({
            "pillar": r["pillar"], "pct": round(100.0 * k / n),
            "subcaps": n, "scored": k, "thin": int(r.get("thin") or 0),
        })
        tot += n
        scr += k
    if not by_pillar:
        return None
    return {"overall_pct": round(100.0 * scr / tot), "by_pillar": by_pillar,
            "gate_pct": 80, "derived_from": "subcap_scores"}


# ── uncertainty_bands ────────────────────────────────────────────────────────
_CEILING_RE = re.compile(r"[LM]?\s*(\d(?:\.\d{1,2})?)\s*(?:±\s*(\d(?:\.\d{1,2})?))?")
_CAT_RE = re.compile(r"^(P[1-4]C\d+)", re.I)


def _parse_ceiling(raw: object) -> tuple[float | None, float | None]:
    m = _CEILING_RE.search(str(raw or ""))
    if not m:
        return None, None
    ceil = float(m.group(1))
    band = float(m.group(2)) if m.group(2) else None
    if not (1.0 <= ceil <= 5.0):
        return None, None
    return ceil, band


def normalize_register_rows(raw_rows: list[dict]) -> dict[str, dict]:
    """Ingested uncertainty-register rows (heterogeneous shapes: cap_id /
    ceiling_estimate 'L2.5 ±0.5' / band / base / total / evidence_count /
    coverage_pct / no_evidence / tier_dist / note) → per-CATEGORY
    {ceiling, band, modifiers[], evidence[], rationale} aggregates."""
    per_cat: dict[str, dict] = {}
    for r in raw_rows:
        if not isinstance(r, dict):
            continue
        cap = str(r.get("cap_id") or r.get("category") or "")
        m = _CAT_RE.match(cap)
        if not m:
            continue
        cat = m.group(1).upper()
        agg = per_cat.setdefault(cat, {"ceils": [], "bands": [], "mods": [],
                                       "notes": [], "n": 0})
        agg["n"] += 1
        ceil, band = _parse_ceiling(r.get("ceiling_estimate") or r.get("total")
                                    or r.get("base"))
        if ceil is not None:
            agg["ceils"].append(ceil)
        for b_raw in (band, r.get("band")):
            try:
                b = float(b_raw)  # type: ignore[arg-type]
            except (TypeError, ValueError):
                continue
            if 0 < b <= 2:
                agg["bands"].append(b)
                break
        cov = str(r.get("coverage_pct") or "").rstrip("%")
        try:
            if cov and float(cov) < 50:
                agg["mods"].append(
                    f"coverage {cov}% on {r.get('capability') or cap}")
        except ValueError:
            pass
        ne = r.get("no_evidence")
        try:
            if ne is not None and int(ne) >= 3:
                agg["mods"].append(f"{ne} subcaps without evidence in "
                                   f"{r.get('capability') or cap}")
        except (TypeError, ValueError):
            pass
        note = str(r.get("note") or "").strip()
        if note and len(note) > 12:
            agg["notes"].append(note)
    out: dict[str, dict] = {}
    for cat, agg in per_cat.items():
        if not agg["ceils"]:
            continue
        ceiling = round(sum(agg["ceils"]) / len(agg["ceils"]), 1)
        band = round(sum(agg["bands"]) / len(agg["bands"]), 1) if agg["bands"] else 0.4
        mods = list(dict.fromkeys(agg["mods"]))[:3]
        rationale = (f"Register-sourced ceiling across {agg['n']} capability rows"
                     + (f"; analyst note: {agg['notes'][0][:140]}" if agg["notes"] else "")
                     + ".")
        out[cat] = {"ceiling": ceiling, "band": band, "modifiers": mods,
                    "evidence": [], "rationale": rationale,
                    "derived_from": "uncertainty_register"}
    return out


def synthesize_band(*, cat: str, cat_name: str, avg_score: float,
                    n_subcaps: int, thin_n: int, caps: list[dict],
                    eids: list[str]) -> dict:
    """Ceiling/band from REAL facts only: score headroom, thin-evidence
    density, applied caps with their stated reasons. The rationale names the
    numbers and the cap reasons — never generic filler."""
    cap_levels = []
    mods: list[str] = []
    for c in caps[:2]:
        reason = str(c.get("reason") or "").strip()
        applied = c.get("cap") is True     # cap_applied is a BOOLEAN flag
        if not applied and not reason:
            continue
        # the LEVEL, when stated, lives in the reason text ("cap M3", "3.0")
        lm = re.search(r"\bM([1-5])\b|\bcap(?:ped)?(?: at| to)?\s*([1-5](?:\.\d)?)",
                       reason, re.I)
        lvl = float(lm.group(1) or lm.group(2)) if lm else None
        if lvl is not None:
            cap_levels.append(lvl)
            mods.append(f"cap M{lvl:g} applied: {reason[:90]}" if reason
                        else f"score cap M{lvl:g} applied")
        elif reason:
            mods.append(f"cap applied: {reason[:90]}")
        elif applied:
            mods.append("score cap applied (level in issue register)")
    if thin_n:
        mods.append(f"{thin_n} of {n_subcaps} subcaps on thin evidence")
    # headroom shrinks as evidence thins; a hard cap binds the ceiling.
    density = 1.0 - (thin_n / n_subcaps if n_subcaps else 0)
    headroom = 0.5 + 0.5 * density
    ceiling = min(5.0, round(avg_score + headroom, 1))
    if cap_levels:
        ceiling = min(ceiling, min(cap_levels))
    band = round(0.3 + 0.4 * (1.0 - density), 1)
    rationale = (f"{cat_name} averages {avg_score:.1f}/5 over {n_subcaps} scored "
                 f"subcaps; ceiling {ceiling:.1f} reflects "
                 + (f"the binding cap ({mods[0]})" if cap_levels else
                    f"{'thin' if thin_n else 'solid'} evidence density "
                    f"({n_subcaps - thin_n}/{n_subcaps} full-evidence cells)")
                 + f"; band ±{band} tracks evidence coverage.")
    return {"ceiling": ceiling, "band": band, "modifiers": mods[:3],
            "evidence": eids[:2], "rationale": rationale,
            "derived_from": "synthesized:scores+caps"}


async def _load_workbook_tiers(session, eid: str) -> dict | None:
    """Rung 1: the research workbook's own flat-sheet tier histogram, read
    from the raw_artifacts store (the compressed originals ingest kept).
    Best-effort — any unreadable/tierless workbook falls through."""
    from io import BytesIO

    from app.services.parsers.research_workbook import evidence_tier_histogram
    from app.services.raw_artifact_store import decompress_payload
    rows = (await session.execute(text(
        r"""
        SELECT rel_path, codec, content FROM raw_artifacts
        WHERE entity_id=CAST(:e AS uuid)
          AND rel_path ~* '02_research_workbook/[^/]+\.xlsx$'
        ORDER BY rel_path LIMIT 6
        """), {"e": eid})).all()
    best: dict | None = None
    for r in rows:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(
                BytesIO(decompress_payload(bytes(r.content), r.codec)),
                read_only=True, data_only=True)
            hist = evidence_tier_histogram(wb)
            wb.close()
        except Exception:
            continue
        if hist and (best is None
                     or hist["total_items"] > best["total_items"]):
            best = hist
    return best


async def _load_handoff_summary(session, eid: str) -> dict | None:
    """Rung 2: research_handoff.json's evidence_summary block (shipped by
    65/113 packages; matches the workbook wherever both exist)."""
    from app.services.raw_artifact_store import decompress_payload
    rows = (await session.execute(text(
        r"""
        SELECT rel_path, codec, content FROM raw_artifacts
        WHERE entity_id=CAST(:e AS uuid)
          AND rel_path ~* '[^/]*research_handoff[^/]*\.json$'
        ORDER BY length(rel_path) LIMIT 4
        """), {"e": eid})).all()
    for r in rows:
        try:
            data = json.loads(
                decompress_payload(bytes(r.content), r.codec).decode(
                    "utf-8", errors="replace"))
        except Exception:
            continue
        es = data.get("evidence_summary") if isinstance(data, dict) else None
        if isinstance(es, dict) and _clean_tier_dist(es.get("tier_distribution")):
            return es
    return None


async def _amain() -> int:
    if not os.environ.get("DATABASE_URL"):
        print("ERROR: DATABASE_URL not set", file=sys.stderr)
        return 2
    sm = get_sessionmaker()
    ev_n = cov_n = unc_reg = unc_syn = 0
    src_counts: Counter = Counter()
    async with sm() as session:
        runs = (await session.execute(text(
            """
            SELECT r.id::text rid, e.id::text eid, e.display_id
            FROM runs r JOIN entities e ON e.id=r.entity_id
            WHERE r.status='ACTIVE' AND e.status='ACTIVE'
            ORDER BY e.display_id
            """))).all()
        for run in runs:
            # 1. evidence_summary
            ev_rows = (await session.execute(text(
                """
                SELECT tier, claim_type, COALESCE(excerpt,'') excerpt,
                       COALESCE(source_name,'') source_name,
                       cardinality(linked_subcap_ids) linked_n
                FROM evidence_index WHERE run_id=CAST(:rid AS uuid) LIMIT 2000
                """), {"rid": run.rid})).mappings().all()
            # provenance ladder: workbook flat sheet → handoff summary →
            # index histogram (each rung best-effort; see build_evidence_
            # summary docstring for the mandate).
            workbook_tiers = await _load_workbook_tiers(session, run.eid)
            handoff_summary = (None if workbook_tiers
                               else await _load_handoff_summary(session, run.eid))
            summary = build_evidence_summary(
                [dict(r) for r in ev_rows],
                workbook_tiers=workbook_tiers,
                handoff_summary=handoff_summary)
            if summary:
                src_counts[summary["derived_from"]] += 1

            # 2. coverage_stats
            cov_rows = (await session.execute(text(
                """
                SELECT substring(subcap_id,1,2) pillar,
                       count(*) subcaps,
                       count(*) FILTER (WHERE score IS NOT NULL AND score > 0) scored,
                       count(*) FILTER (WHERE is_thin_evidence) thin
                FROM subcap_scores WHERE run_id=CAST(:rid AS uuid)
                GROUP BY 1 ORDER BY 1
                """), {"rid": run.rid})).mappings().all()
            coverage = build_coverage_stats([dict(r) for r in cov_rows])

            # 3. uncertainty_bands — register-normalized else synthesized
            reg_rows = (await session.execute(text(
                """
                SELECT provenance->'raw_row' raw FROM client_knowledge_sections
                WHERE entity_id=CAST(:e AS uuid) AND artifact_kind='uncertainty'
                LIMIT 1500
                """), {"e": run.eid})).scalars().all()
            raw_rows = [r for r in reg_rows if isinstance(r, dict)]
            if not raw_rows:
                cur = (await session.execute(text(
                    "SELECT uncertainty_bands FROM runs WHERE id=CAST(:rid AS uuid)"
                ), {"rid": run.rid})).scalar()
                if isinstance(cur, list):    # the ingested raw array shape
                    raw_rows = [r for r in cur if isinstance(r, dict)]
            bands = normalize_register_rows(raw_rows) if raw_rows else {}
            source = "uncertainty_register" if bands else None

            cat_rows = (await session.execute(text(
                """
                SELECT COALESCE(s.parent_category_id, LEFT(s.subcap_id,4)) cat,
                       ROUND(AVG(s.score)::numeric,2) sc,
                       count(*) n,
                       count(*) FILTER (WHERE s.is_thin_evidence) thin,
                       (ARRAY_AGG(COALESCE(cs.name, s.subcap_id) ORDER BY s.score))[1] cname,
                       jsonb_agg(DISTINCT jsonb_build_object(
                           'cap', s.cap_applied, 'reason', s.cap_reason))
                           FILTER (WHERE s.cap_applied IS TRUE
                                   OR s.cap_reason IS NOT NULL) caps
                FROM subcap_scores s
                LEFT JOIN ccg_subcaps cs ON cs.subcap_id = s.subcap_id
                WHERE s.run_id=CAST(:rid AS uuid) AND s.score IS NOT NULL
                GROUP BY 1 ORDER BY 1
                """), {"rid": run.rid})).mappings().all()
            ev_by_cat_rows = (await session.execute(text(
                """
                SELECT DISTINCT ON (substring(sid,1,4)) substring(sid,1,4) cat, e_id
                FROM (SELECT unnest(linked_subcap_ids) sid, e_id, tier
                      FROM evidence_index WHERE run_id=CAST(:rid AS uuid)) x
                ORDER BY substring(sid,1,4), tier
                """), {"rid": run.rid})).all()
            eids_by_cat: dict[str, list[str]] = {}
            for er in ev_by_cat_rows:
                eids_by_cat.setdefault(er.cat, []).append(er.e_id)
            for cr in cat_rows:
                cat = str(cr["cat"] or "")[:4].upper()
                if not cat or cat in bands:
                    continue
                caps = [c for c in (cr["caps"] or []) if isinstance(c, dict)]
                bands[cat] = synthesize_band(
                    cat=cat, cat_name=str(cr["cname"] or cat),
                    avg_score=float(cr["sc"]), n_subcaps=int(cr["n"]),
                    thin_n=int(cr["thin"]), caps=caps,
                    eids=eids_by_cat.get(cat, []))
                source = source or "synthesized"
            # attach category evidence to register-sourced bands too
            for cat, b in bands.items():
                if not b.get("evidence"):
                    b["evidence"] = eids_by_cat.get(cat, [])[:2]

            await session.execute(text(
                """
                UPDATE runs SET
                    evidence_summary = CAST(:es AS jsonb),
                    coverage_stats = CAST(:cs AS jsonb),
                    uncertainty_bands = CAST(:ub AS jsonb),
                    updated_at = NOW()
                WHERE id = CAST(:rid AS uuid)
                """), {
                    "es": json.dumps(summary) if summary else None,
                    "cs": json.dumps(coverage) if coverage else None,
                    "ub": json.dumps(bands) if bands else None,
                    "rid": run.rid,
                })
            ev_n += 1 if summary else 0
            cov_n += 1 if coverage else 0
            if bands:
                if source == "uncertainty_register":
                    unc_reg += 1
                else:
                    unc_syn += 1
        await session.commit()

    tier_src = " ".join(f"{k}={v}" for k, v in sorted(src_counts.items()))
    print(f"# derive_evidence_surfaces: evidence_summary={ev_n} "
          f"[tier source: {tier_src or 'n/a'}] "
          f"coverage_stats={cov_n} uncertainty_bands="
          f"{unc_reg + unc_syn} (register={unc_reg} synthesized={unc_syn}) "
          f"(tier ladder: research workbook → handoff → evidence_index)",
          flush=True)
    return 0


def main() -> None:
    raise SystemExit(asyncio.run(_amain()))


if __name__ == "__main__":
    main()


# re-exported for tests
__all__ = [
    "build_coverage_stats",
    "build_evidence_summary",
    "normalize_register_rows",
    "synthesize_band",
]
