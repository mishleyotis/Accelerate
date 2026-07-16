"""Scoring workbook parser.

Workbooks vary in shape per analyst per DMA. The parser is two-pass:

  1. **Read sheet metadata** (sheet names, first-row headers, sample row
     types) via openpyxl.
  2. **LLM column-map inference** — a Gemini Pro structured-output call
     emits a `ScoringWorkbookMap` (which sheet is the subcap-scores sheet,
     which column is the ID / score / confidence / evidence). The map is
     CACHED by `sha256(normalized_header_tuple)` so subsequent workbooks
     with the same shape skip the LLM entirely.
  3. **Deterministic ingest** — iterate using the column map, emit rows
     ready for `subcap_scores` insertion.

Every subcap_id is checked against `ccg_subcaps` (current version) or
`ccg_subcap_aliases` (older version). Unknown IDs > 5% of the workbook
flip the import_files row to PENDING_REVIEW.

This module exposes pure functions; IO is at the loader's edge.
"""
from __future__ import annotations

import hashlib
from collections.abc import Callable
from dataclasses import dataclass, field
from typing import Any

try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover
    Worksheet = Any  # type: ignore[assignment,misc]


@dataclass
class ScoringWorkbookMap:
    """Where each scalar lives within a particular workbook shape."""
    subcap_score_sheet: str
    subcap_id_col: str          # column letter (A, B, …)
    score_col: str
    confidence_col: str | None = None
    evidence_col: str | None = None
    peer_median_col: str | None = None
    rationale_col: str | None = None


@dataclass
class SheetMetadata:
    sheet_name: str
    headers: list[str]
    sample_rows: list[list[Any]]


@dataclass
class ParsedScore:
    subcap_id: str
    score: float
    confidence: float | None = None
    evidence_excerpt: str | None = None
    peer_median: float | None = None
    rationale: str | None = None


@dataclass
class WorkbookParseResult:
    rows: list[ParsedScore] = field(default_factory=list)
    warnings: list[dict[str, Any]] = field(default_factory=list)
    used_map: ScoringWorkbookMap | None = None


def collect_sheet_metadata(workbook: Any, *, sample_size: int = 20) -> list[SheetMetadata]:
    """Pure: gather (sheet_name, headers, samples) for every sheet."""
    out: list[SheetMetadata] = []
    for ws in workbook.worksheets:
        headers: list[str] = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
            headers.append(str(cell).strip() if cell is not None else "")
        samples: list[list[Any]] = []
        for row in ws.iter_rows(min_row=2, max_row=1 + sample_size, values_only=True):
            if all(c is None for c in row):
                continue  # skip the padding empties openpyxl emits up to max_row
            samples.append(list(row))
        out.append(SheetMetadata(sheet_name=ws.title, headers=headers, sample_rows=samples))
    return out


def shape_fingerprint(metadata: list[SheetMetadata]) -> str:
    """Stable hash of (sheet_name, headers) tuples — used to cache the LLM map."""
    serialized = []
    for m in metadata:
        serialized.append(m.sheet_name)
        serialized.extend(m.headers)
        serialized.append("|")
    blob = "␟".join(serialized).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()


def col_letter(idx: int) -> str:
    """0-indexed column → Excel-style letter (0→A, 25→Z, 26→AA, …)."""
    if idx < 0:
        raise ValueError("col_letter: idx must be >= 0")
    out = ""
    n = idx
    while True:
        out = chr(ord("A") + n % 26) + out
        n = n // 26 - 1
        if n < 0:
            break
    return out


def header_to_col(headers: list[str], header: str) -> str | None:
    for i, h in enumerate(headers):
        if h.strip().lower() == header.strip().lower():
            return col_letter(i)
    return None


def col_to_index(letter: str) -> int:
    """Excel-style letter → 0-indexed column."""
    idx = 0
    for ch in letter.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"col_to_index: bad letter {letter!r}")
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def parse_with_map(
    workbook: Any,
    workbook_map: ScoringWorkbookMap,
) -> WorkbookParseResult:
    """Deterministic ingest given a column map."""
    result = WorkbookParseResult(used_map=workbook_map)
    ws = workbook[workbook_map.subcap_score_sheet]
    cols = {
        "id": col_to_index(workbook_map.subcap_id_col),
        "score": col_to_index(workbook_map.score_col),
        "confidence": col_to_index(workbook_map.confidence_col) if workbook_map.confidence_col else None,
        "evidence": col_to_index(workbook_map.evidence_col) if workbook_map.evidence_col else None,
        "peer_median": col_to_index(workbook_map.peer_median_col) if workbook_map.peer_median_col else None,
        "rationale": col_to_index(workbook_map.rationale_col) if workbook_map.rationale_col else None,
    }

    def _safe(row: tuple[Any, ...], idx: int | None) -> Any:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in ws.iter_rows(min_row=2, values_only=True):
        subcap_raw = _safe(row, cols["id"])
        score_raw = _safe(row, cols["score"])
        if subcap_raw is None or score_raw is None:
            continue
        subcap_id = str(subcap_raw).strip()
        try:
            score = float(score_raw)
        except (TypeError, ValueError):
            result.warnings.append({"kind": "bad_score_cell", "subcap_id": subcap_id,
                                    "value": str(score_raw)})
            continue
        if not (1.0 <= score <= 5.0):
            result.warnings.append({"kind": "score_out_of_range", "subcap_id": subcap_id,
                                    "value": score})
            continue
        confidence: float | None = None
        conf_raw = _safe(row, cols["confidence"])
        if conf_raw is not None:
            try:
                c = float(conf_raw)
                if 0.0 <= c <= 1.0:
                    confidence = c
            except (TypeError, ValueError):
                pass
        peer_median: float | None = None
        pm_raw = _safe(row, cols["peer_median"])
        if pm_raw is not None:
            try:
                pm = float(pm_raw)
                if 1.0 <= pm <= 5.0:
                    peer_median = pm
            except (TypeError, ValueError):
                pass

        result.rows.append(
            ParsedScore(
                subcap_id=subcap_id,
                score=score,
                confidence=confidence,
                evidence_excerpt=_str_or_none(_safe(row, cols["evidence"])),
                peer_median=peer_median,
                rationale=_str_or_none(_safe(row, cols["rationale"])),
            )
        )
    return result


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse(
    workbook: Any,
    *,
    map_cache_lookup: Callable[[str], ScoringWorkbookMap | None],
    map_cache_store: Callable[[str, ScoringWorkbookMap], None],
    infer_map: Callable[[list[SheetMetadata]], ScoringWorkbookMap],
) -> WorkbookParseResult:
    """End-to-end parse: classify shape → reuse cached map or call LLM → ingest."""
    metadata = collect_sheet_metadata(workbook)
    fp = shape_fingerprint(metadata)
    workbook_map = map_cache_lookup(fp)
    if workbook_map is None:
        workbook_map = infer_map(metadata)
        map_cache_store(fp, workbook_map)
    return parse_with_map(workbook, workbook_map)


def score_to_band(score: float) -> str:
    """Map a 1-5 score to its M1-M5 maturity band.

    Bands are inclusive at the lower bound, exclusive at the upper, with M5
    capturing scores 4.5..5.0 inclusive. This matches the wireframe rule on
    the heatmap cells.
    """
    if score < 1.0 or score > 5.0:
        raise ValueError(f"score_to_band: out of range: {score}")
    if score >= 4.5:
        return "M5"
    if score >= 3.5:
        return "M4"
    if score >= 2.5:
        return "M3"
    if score >= 1.5:
        return "M2"
    return "M1"


# ── Category-level workbook fallback (no LLM, no DB) ───────────────────────
import re as _re  # noqa: E402
from pathlib import Path as _Path  # noqa: E402

_CAT_RE = _re.compile(r"P\d+C\d+")


def _norm_hdr(h: object) -> str:
    return _re.sub(r"[^a-z0-9]+", "_", str(h or "").strip().lower()).strip("_")


def category_scores_from_workbook(root: _Path) -> list[dict[str, Any]]:
    """Extract CATEGORY-level scores from a scoring workbook's per-pillar
    sheets (e.g. MidFirst `P#_Scoring_Detail` with Category_ID / Category_Name
    / Score columns). Some packages score at category granularity in the
    workbook rather than shipping subcap rows or an export_category_summary.

    Recursive discovery (workbook may sit at the package root). Returns dicts
    matching CategoryScoreRow(**row). [] when no category-bearing sheet found.
    Pure / no DB / no LLM.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return []
    out: dict[str, dict[str, Any]] = {}
    seen_files: set[_Path] = set()
    for pat in ("**/*[Ss]coring*[Ww]orkbook*.xlsx", "**/*[Ss]coring*.xlsx"):
        for xlsx in sorted(root.glob(pat)):
            if xlsx in seen_files:
                continue
            seen_files.add(xlsx)
            try:
                wb = load_workbook(xlsx, read_only=True, data_only=True)
            except Exception:
                continue
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                idx = {_norm_hdr(h): i for i, h in enumerate(header)}
                cat_i = idx.get("category_id")
                score_i = next(
                    (idx[k] for k in ("score", "final_score", "category_score",
                                      "post_critic_score") if k in idx),
                    None,
                )
                if cat_i is None or score_i is None:
                    continue  # not a category-score sheet (e.g. Summary)
                name_i = idx.get("category_name")
                for r in rows:
                    if cat_i >= len(r) or score_i >= len(r):
                        continue
                    m = _CAT_RE.match(str(r[cat_i] or "").strip())
                    if not m:
                        continue
                    try:
                        score = float(str(r[score_i]).strip())
                    except (TypeError, ValueError):
                        continue
                    cat = m.group(0)
                    if cat in out:
                        continue
                    pm = _re.match(r"P\d+", cat)
                    out[cat] = {
                        "category_id": cat,
                        "category_name": (str(r[name_i]).strip()
                                          if name_i is not None and name_i < len(r)
                                          and r[name_i] else None),
                        "pillar_id": pm.group(0) if pm else cat[:2],
                        "score": round(score, 2),
                    }
            if out:
                return list(out.values())
    return list(out.values())
