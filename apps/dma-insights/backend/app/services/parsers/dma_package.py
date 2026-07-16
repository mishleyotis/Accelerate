"""DMA package orchestrator.

Walks a `{Entity}_DMA_Complete_Package/` folder (or zip-extracted
equivalent) and produces an `IngestedPackage` envelope by dispatching
each file to the appropriate leaf parser.

Canonical folder layout (per the project's package contract, verified
against the AlmaBank + WSFS reference packages):

    {root}/
      MANIFEST.json
      01_evidence/
        evidence_index.csv | evidence_index.json
      02_research_workbook/
        research_handoff.json
        *.xlsx                                  (raw Layer-0 workbook)
      03_scoring_workbook/
        export_scoring_detail.csv               (Alma canonical)
        export_pillar_summary.csv
        export_category_summary.csv
        final_scores.json
        *.xlsx                                  (raw Layer-1 workbook; F3
                                                 XLSX-fallback path for
                                                 packages w/o export CSVs —
                                                 Calprivate / Amalgamated /
                                                 AmeriCU)
      04_reports/
        *_Assessment_Report.docx
        *_Client_Profile_Research_Report.docx   (F5b leadership + F5c
                                                 narrative_md sources)
      05_narrative_deck/                        (often empty)
      06_peers/
        peer_scores_*.json
        peer_comparison_table.csv
        peer_synthesis.md
      07_governance/
        run_manifest.json | qa_verdict.json | audit_summary.json
        layer1_issue_register.json (ALMA) | assessment_issue_register.json (WSFS)
        L1_qa_verdict.json | L2_qa_verdict.json (Odlum 2-stage variant)
        Layer1_qa_verdict.json | GOV_qa_verdict.json (Calprivate variant)
        NicolaWealth_L2_QAVerdict.json          (Nicola — only L2)
        recommendations_register.json           (Odlum variant; canonical
                                                 lives in 08_appendices/)
        scoring_exports/export_*.csv            (Odlum F2 variant — final
                                                 cap-applied scoring lives
                                                 here, not 03_)
        issue_register.csv
      08_appendices/
        recommendations_detail.json             (Alma canonical)
        assessment_analysis.json
        *_Explorium_Tech_Stack.xlsx
        report_synthesis.md
        run_manifest.json                       (WSFS variant)
        A1_Evidence_Index.csv … A9_*.csv        (Calprivate appendix CSVs;
                                                 NOT scoring exports — F2
                                                 globs cannot false-positive
                                                 because none match
                                                 `export_scoring_detail*.csv`)

State-branch contract:
  - MANIFEST.json missing      → falls back to scanning 07/08 for
                                  run_manifest.json; warning added.
  - run_manifest.json missing  → `_synthesize_run_manifest_from_handoff`
                                  reads `02_research_workbook/research_handoff.json`
                                  (F1 — Nicola path); falls back further
                                  to `_synthesize_run_manifest_from_exports`
                                  (RegionsBank path).
  - 03_scoring_workbook missing→ subcap_scores empty; warning; downstream
                                  pages render "no scoring data" state.
  - 03 has only XLSX (no CSV)  → `_scoring_from_xlsx_fallback` mines
                                  per-pillar sheets (F3 — Calprivate);
                                  emits `scoring loaded from xlsx fallback`
                                  warning.
  - 03 CSVs absent + 07_governance/scoring_exports/ present
                               → F2: loop iterates 03 → 07_/scoring_exports/
                                  → 08_appendices/ in priority order; first
                                  non-empty wins. Variant-location warning
                                  emitted when non-canonical dir used.
  - 06_peers missing           → peers=[]; D3 peer overlay hidden.
  - 07_governance missing      → qa_verdict=None; issue_register=[];
                                  D6 Health page shows "no governance data".
  - 08_appendices missing      → recommendations=[]; tech_stack=[];
                                  D4 Platform page shows "no recommendations".
  - 08_appendices present but no recommendations source
                               → recommendations=[]; F4 emits explicit
                                  `no_recommendations_source` warning so
                                  admin import-audit can flag the source-
                                  side gap (WSFS / Nicola / Calprivate).
  - 04_reports/*_Client_Profile_Research_Report.docx present
                               → F5b: leadership extracted into
                                  `firmographics.leadership` (canonical
                                  `Name`+`Title` cols OR combined
                                  `Executive` cells split on `\\n` / `—`
                                  / `-` / `|`; iterates ALL matching
                                  tables and dedupes by name).
                               → F5c: `firmographics_narrative_md` (the
                                  analyst-prose Entity-Profile paragraph,
                                  200-1600 chars) threaded onto
                                  `firmographics.narrative_md`. Both flow
                                  through additively even when handoff
                                  JSON pre-populated `firm`.
"""
from __future__ import annotations

import contextlib
import json
import re
from pathlib import Path

import structlog

from app.schemas.package import (
    CategoryScoreRow,
    FocusAreaRow,
    IngestedPackage,
    ReportSectionRow,
    RunManifest,
    TechStackRow,
)
from app.services.parsers.assessment_report import (
    find_assessment_reports,
    parse_assessment_report,
)
from app.services.parsers.facts_extractor import extract_timeline_events
from app.services.parsers.package_csvs import (
    parse_category_summary_csv,
    parse_evidence_csv,
    parse_issue_register_csv,
    parse_pillar_summary_csv,
    parse_scoring_detail_csv,
)
from app.services.parsers.package_financials import (
    load_financial_baseline,
    load_financial_trends,
    load_sentiment,
    sentiment_from_entity_profile,
)
from app.services.parsers.package_json import (
    parse_firmographics,
    parse_peer_score,
    parse_qa_verdict,
    parse_run_manifest,
    parse_top_manifest,
)
from app.services.parsers.package_peers import load_peer_benchmarks
from app.services.parsers.package_recommendations import (
    _rec_id,
    parse_recommendations_any,
)
from app.services.parsers.package_techstack import (
    extract_tech_from_text,
    load_tech_stack,
    sanitize_tech_rows,
)
from app.services.parsers.recommendation_validation import parse_rec_prerequisites
from app.services.parsers.report_synthesis import (
    build_derived_scqa,
    find_report_synthesis,
    parse_report_synthesis_md,
)
from app.services.parsers.run_id import parse_run_id
from app.services.parsers.section_analysis import (
    combine_insight_rungs,
    insights_from_category_gaps,
    insights_from_profile_findings,
    insights_from_recommendations,
    insights_from_zennify_opportunities,
    parse_section_analyses,
)

log = structlog.get_logger()


# ── Warning-severity taxonomy (Part 12.1) ───────────────────────────────
# Every parser warning is one of three severities:
#   INFO      — expected variance (variant file locations, derived
#               fallbacks that carry full fidelity, reconciliations).
#   DEGRADED  — the package parsed, but a source artifact was corrupt /
#               unreadable / schema-drifted and a fallback rung supplied
#               a poorer substitute (or none).
#   DATA_LOSS — user-visible data is missing (zero evidence, zero recs,
#               dropped scores) — the fail-loud gate routes these runs
#               to PENDING_REVIEW unless DMA_ALLOW_HOLLOW=1.
#
# Backward compat: `parser_warnings` stays list[str] (the audit scripts,
# admin import-audit previews, and the overview surface all consume the
# strings via substring matching). New warnings are PREFIXED with
# "{SEVERITY}/{code}: "; `structure_warnings()` parses both the prefixed
# form and the legacy unprefixed strings into structured
# {code, severity, detail} dicts for the parallel `structured` list the
# backfill writes into runs.parser_warnings JSONB.
SEVERITY_INFO = "INFO"
SEVERITY_DEGRADED = "DEGRADED"
SEVERITY_DATA_LOSS = "DATA_LOSS"

_SEVERITY_PREFIX_RE = re.compile(
    r"^(INFO|DEGRADED|DATA_LOSS)/([A-Za-z0-9_.-]+):\s*(.*)$", re.DOTALL
)

# Legacy (unprefixed) warning families → severity. Matched by substring
# against the warning string, FIRST match wins — order most-specific
# first. Anything unmatched defaults to INFO (the corpus baseline showed
# the remaining families are informational provenance notes).
_LEGACY_SEVERITY_RULES: tuple[tuple[str, str], ...] = (
    # DATA_LOSS — user-visible surface is empty / rows were dropped.
    ("no_recommendations_source", SEVERITY_DATA_LOSS),
    ("01_evidence missing", SEVERITY_DATA_LOSS),
    ("no evidence_index", SEVERITY_DATA_LOSS),
    ("evidence_index missing", SEVERITY_DATA_LOSS),
    ("catalogue_empty_for_version", SEVERITY_DATA_LOSS),
    ("pillar_level_scores_dropped", SEVERITY_DATA_LOSS),
    ("zero_evidence", SEVERITY_DATA_LOSS),
    ("zero_recommendations", SEVERITY_DATA_LOSS),
    # DEGRADED — corrupt/unreadable/drifted source; fallback used.
    ("json_corrupt", SEVERITY_DEGRADED),
    ("io_error", SEVERITY_DEGRADED),
    ("schema_mismatch", SEVERITY_DEGRADED),
    ("download_failed", SEVERITY_DEGRADED),
    ("export_too_large", SEVERITY_DEGRADED),
    ("drive_zip_failed", SEVERITY_DEGRADED),
    ("drive_zip_slip", SEVERITY_DEGRADED),
    ("drive_zip_symlink_rejected", SEVERITY_DEGRADED),
    ("drive_zip_oversize_entry", SEVERITY_DEGRADED),
    ("drive_zip_bomb_guard", SEVERITY_DEGRADED),
    ("_failed", SEVERITY_DEGRADED),          # *_extract_failed / *_heal_failed …
    ("failed:", SEVERITY_DEGRADED),
    ("catalogue_unresolved", SEVERITY_DEGRADED),
    ("subcap coverage", SEVERITY_DEGRADED),
    ("institution_name_junk", SEVERITY_DEGRADED),
    ("scoring loaded from xlsx fallback", SEVERITY_DEGRADED),
    ("run_id format", SEVERITY_DEGRADED),
    ("synthesized run manifest", SEVERITY_DEGRADED),
    ("synthesized run_manifest", SEVERITY_DEGRADED),
    ("run_manifest.json missing", SEVERITY_DEGRADED),
    ("synth_run_id", SEVERITY_DEGRADED),
    ("docx_only_package_no_manifest", SEVERITY_DEGRADED),
    ("lenient_mode_deep_extract", SEVERITY_DEGRADED),
)


def warn(
    warnings: list[str], code: str, severity: str, detail: str,
) -> str:
    """Append a severity-tagged warning string and return it.

    Format: ``{SEVERITY}/{code}: {detail}`` — substring-compatible with
    every existing consumer (they match on fragments, not equality).
    """
    entry = f"{severity}/{code}: {detail}"
    warnings.append(entry)
    return entry


def classify_warning(w: object) -> dict:
    """One warning (string or dict) → {code, severity, detail}."""
    if isinstance(w, dict):
        # Already structured (e.g. a PATTERN_GAP entry).
        return {
            "code": str(w.get("code") or "unknown"),
            "severity": str(w.get("severity") or SEVERITY_INFO),
            "detail": str(w.get("detail") or w.get("reason") or ""),
        }
    s = str(w)
    m = _SEVERITY_PREFIX_RE.match(s)
    if m:
        return {"code": m.group(2), "severity": m.group(1), "detail": m.group(3)}
    low = s.lower()
    for needle, severity in _LEGACY_SEVERITY_RULES:
        if needle.lower() in low:
            # code = the first token of the warning up to ':' (best effort)
            code = s.split(":", 1)[0].strip()[:64] or "unknown"
            return {"code": code, "severity": severity, "detail": s}
    code = s.split(":", 1)[0].strip()[:64] or "unknown"
    return {"code": code, "severity": SEVERITY_INFO, "detail": s}


def structure_warnings(warnings: list) -> list[dict]:
    """Parallel structured list for runs.parser_warnings JSONB."""
    return [classify_warning(w) for w in (warnings or [])]


def severity_counts(structured: list[dict]) -> dict[str, int]:
    counts = {SEVERITY_INFO: 0, SEVERITY_DEGRADED: 0, SEVERITY_DATA_LOSS: 0}
    for entry in structured:
        counts[entry.get("severity", SEVERITY_INFO)] = (
            counts.get(entry.get("severity", SEVERITY_INFO), 0) + 1
        )
    return counts


def _read_text(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="replace")


# Matches canonical numbered subfolders `01_evidence` … `11_issues`.
# Used by _find_root to avoid mis-rooting onto a numbered subfolder that
# happens to contain a MANIFEST.json (First Citizens: 06_peers/MANIFEST.json).
_CANON_SUBFOLDER_RE = re.compile(r"^\d{2}_")

# Directories that are never a package root themselves — they hold
# leaf artifacts (charts/exports) or sub-batched copies. Skipped by the
# bounded-depth descent so it can't mis-root onto e.g. `03_scoring_workbook/exports`.
_DESCENT_SKIP_DIRS = frozenset({
    "charts", "exports", "sub_batches", "__pycache__", ".git", "scoring_exports",
})

_CANONICAL_SUBFOLDERS = frozenset({
    "01_evidence", "02_research_workbook", "03_scoring_workbook",
    "04_reports", "05_narrative_deck", "06_peers", "07_governance",
    # OZK-shaped packages ship a dedicated `08_qa/` for the QA verdict
    # + results (qa_verdict.json / qa_results_*.json) alongside an
    # `08_appendices/`; list both so neither is flagged as unexpected.
    "08_appendices", "08_qa",
})
_MANIFEST_BEARING = frozenset({
    "07_governance", "08_appendices", "03_scoring_workbook",
})


def _score_package_dir(p: Path) -> int:
    """Heuristic score for how 'package-root-like' a directory is.

    Higher = more likely to be the effective package root. Combines the
    count of canonical numbered subfolders, presence of a manifest-bearing
    subfolder, and presence of key files (scores JSON, scoring workbook,
    assessment DOCX) anywhere directly inside. A score of 0 means the
    directory shows no DMA-package signal at all.

    Pure structural inspection (names + globs only) — no file parsing — so
    it is safe to call across an entire tree during discovery.
    """
    if not p.is_dir():
        return 0
    try:
        children = list(p.iterdir())
    except (PermissionError, OSError):
        return 0
    dir_names = {c.name for c in children if c.is_dir()}
    score = 0
    canon_present = dir_names & _CANONICAL_SUBFOLDERS
    score += len(canon_present) * 10
    if dir_names & _MANIFEST_BEARING:
        score += 5
    # Key files directly under p (cheap, no recursion).
    if (p / "MANIFEST.json").exists() or list(p.glob("package_manifest.json")):
        score += 8
    if list(p.glob("*[Ss]cores*.json")) or list(p.glob("run_manifest*.json")):
        score += 4
    if list(p.glob("*[Ss]coring*[Ww]orkbook*.xlsx")):
        score += 4
    for f in children:
        if f.is_file() and f.suffix.lower() in (".docx", ".pdf"):
            nl = f.name.lower()
            if any(tok in nl for tok in _DMA_DOCX_NAME_TOKENS):
                score += 3
                break
    return score


def _is_acceptable_descended_root(p: Path) -> bool:
    """Gate for the bounded-depth descent: a directory only qualifies as a
    package root if it shows REAL package structure — the same threshold as
    `_find_root`'s `_has_numbered_layout` (≥2 canonical folders AND a
    manifest-bearing one) OR a key authoritative file (MANIFEST / scores /
    run_manifest / scoring workbook). This deliberately EXCLUDES weak
    signals — a couple of non-manifest folders (e.g. evidence + narrative
    only) or a stray assessment DOCX — so the descent never pre-empts the
    legacy reject (FileNotFoundError) or the docx-only branch, which own
    those cases.
    """
    if not p.is_dir():
        return False
    try:
        dir_names = {c.name for c in p.iterdir() if c.is_dir()}
    except (PermissionError, OSError):
        return False
    canon = dir_names & _CANONICAL_SUBFOLDERS
    manifest_bearing = bool(dir_names & _MANIFEST_BEARING)
    has_key_file = (
        (p / "MANIFEST.json").exists()
        or bool(list(p.glob("*[Ss]cores*.json")))
        or bool(list(p.glob("run_manifest*.json")))
        or bool(list(p.glob("*[Ss]coring*[Ww]orkbook*.xlsx")))
    )
    return (len(canon) >= 2 and manifest_bearing) or has_key_file


def _descend_to_best_root(root: Path, max_depth: int = 4) -> Path | None:
    """Bounded-depth BFS for the most package-root-like directory.

    The legacy `_find_root` checks only `root` and its direct children, so
    deeply-nested corpus layouts -- `IMA Financial - DMA/IMA Financial/IMA
    Financial DMA/`, `Navy Federal ... - DMA/DMA/DMA 2026-02-20/`, the
    `ATB - DMA/ATB DMA v2/` hybrid -- at depth 2-4 are never found and the
    whole package is dropped. This descent walks to `max_depth`, scores every
    directory via `_score_package_dir`, and returns the highest-scoring one
    (ties broken toward the shallowest). Returns None when nothing in the
    tree shows any package signal (caller then falls through to the existing
    docx-only / FileNotFoundError paths).
    """
    best: Path | None = None
    best_score = 0
    best_depth = max_depth + 1
    # BFS queue of (dir, depth).
    queue: list[tuple[Path, int]] = [(root, 0)]
    while queue:
        cur, depth = queue.pop(0)
        # Only REAL package roots are eligible (gate); score ranks among them.
        if _is_acceptable_descended_root(cur):
            s = _score_package_dir(cur)
            if s > best_score or (s == best_score and depth < best_depth):
                best, best_score, best_depth = cur, s, depth
        if depth >= max_depth:
            continue
        try:
            for child in sorted(cur.iterdir()):
                if (
                    child.is_dir()
                    and child.name not in _DESCENT_SKIP_DIRS
                    and not _CANON_SUBFOLDER_RE.match(child.name)
                ):
                    queue.append((child, depth + 1))
        except (PermissionError, OSError):
            continue
    return best


def _find_root(root: Path) -> Path:
    """Locate the actual package root inside `root`.

    Accepts THREE layouts (in priority order):
      1. `{root}/MANIFEST.json` exists — canonical AlmaBank-style.
      2. `{root}/<one-child>/MANIFEST.json` exists — same, wrapped in a
         per-entity folder (the zip layout).
      3. The numbered-subfolder pattern at `{root}` or one level down,
         without any MANIFEST.json — RegionsBank-style.

    The numbered-subfolder check requires ≥3 of the 8 canonical
    `NN_<name>/` directories to be present. We don't require all 8
    because real packages routinely omit `05_narrative_deck/`,
    `06_peers/`, or `08_appendices/`.
    """
    canonical_subfolders = {
        "01_evidence", "02_research_workbook", "03_scoring_workbook",
        "04_reports", "05_narrative_deck", "06_peers", "07_governance",
        "08_appendices",
    }

    def _has_numbered_layout(p: Path) -> bool:
        if not p.is_dir():
            return False
        present = {c.name for c in p.iterdir() if c.is_dir()}
        # Threshold relaxed from ≥3 → ≥2 so packages with sparse layout
        # (AmeriCU at the time of this commit only ships 01_evidence +
        # 03_scoring_workbook + 06_peers + 07_governance regularly; older
        # vintages may ship fewer). Any package with ≥2 numbered
        # subfolders AND at least one of the manifest-bearing kinds
        # (07_governance or 08_appendices or 03_scoring_workbook) is
        # accepted.
        manifest_bearing = {
            "07_governance", "08_appendices", "03_scoring_workbook",
        }
        return len(present & canonical_subfolders) >= 2 and bool(
            present & manifest_bearing
        )

    def _has_manifest_anywhere(p: Path, depth: int = 0) -> bool:
        """True if ANY of MANIFEST.json / run_manifest*.json / *qa_verdict*.json
        exists at `p` or one level down. Final fallback for malformed packages
        the operator hand-uploaded."""
        if not p.is_dir() or depth > 2:
            return False
        for name in ("MANIFEST.json",):
            if (p / name).exists():
                return True
        for pattern in ("run_manifest*.json", "*qa_verdict*.json"):
            if list(p.glob(pattern)):
                return True
        if depth >= 2:
            return False
        for child in p.iterdir():
            if child.is_dir() and _has_manifest_anywhere(child, depth + 1):
                return True
        return False

    if (root / "MANIFEST.json").exists():
        return root

    # 2026-06-07 corpus fix (First Citizens): if `root` ITSELF has the
    # numbered-subfolder layout, `root` is the package root — even if a
    # MANIFEST.json lives inside one of those numbered subfolders (First
    # Citizens ships `06_peers/MANIFEST.json`). The old code's
    # "child has MANIFEST -> child is root" scan ran first and wrongly
    # returned `06_peers/` as the root, so `03_scoring_workbook` +
    # `01_evidence` (siblings of 06_peers) were reported missing and the
    # package ingested 0 evidence / 0 subcaps. Checking the root's own
    # numbered layout first preempts that mis-root.
    if _has_numbered_layout(root):
        return root

    # Wrapper-folder case (AlmaBank-style): root has no numbered layout
    # but a single child holds the MANIFEST.json. Skip canonical
    # numbered subfolders so a MANIFEST buried in `0N_*` never makes
    # that subfolder the root.
    for child in root.iterdir():
        if (
            child.is_dir()
            and not _CANON_SUBFOLDER_RE.match(child.name)
            and (child / "MANIFEST.json").exists()
        ):
            return child

    # No MANIFEST anywhere — fall back to numbered-subfolder detection
    # in a child (RegionsBank variant: no top-level manifest, package
    # nested one level down without a wrapper MANIFEST).
    for child in root.iterdir():
        if child.is_dir() and _has_numbered_layout(child):
            return child

    # Original 03_scoring_workbook-only check (legacy fallback).
    if (root / "03_scoring_workbook").exists():
        return root
    for child in root.iterdir():
        if child.is_dir() and (child / "03_scoring_workbook").exists():
            return child

    # Last resort: any directory at depth ≤ 2 that contains a manifest-
    # bearing file. Lets the parser ingest a wider variety of folder
    # shapes operators upload by hand.
    if _has_manifest_anywhere(root):
        return root
    for child in root.iterdir():
        if child.is_dir() and _has_manifest_anywhere(child):
            return child

    # 2026-06-09 corpus-coverage fix: deeply-nested package layouts.
    # ~49% of the historical corpus buries the canonical 01_..08_ folders
    # 2-4 levels below the entity folder (`<Entity> - DMA/<Entity>/<Entity>
    # DMA/`, `... - DMA/DMA/DMA <date>/`, the `ATB - DMA/ATB DMA v2/` hybrid).
    # The depth-0/1 checks above miss them, so the whole package was dropped.
    # The bounded-depth descent finds the most package-root-like directory
    # anywhere within depth 4 BEFORE falling back to a bare docx match, so a
    # nested full package is preferred over a stray report DOCX.
    descended = _descend_to_best_root(root)
    if descended is not None:
        return descended

    # 2026-05-28 H6 hotfix: DOCX-only Drive folders.
    # Many production Drive folders contain valid DMA report DOCX files
    # but no MANIFEST.json and no canonical 01_..08_ subfolder layout.
    # Without this branch, `_find_root` raised FileNotFoundError and the
    # historical backfill counted those folders as parse failures — 21
    # of 115 folders in the 2026-05-28 backfill (16 "no DMA package
    # detected" + 5 "no run manifest"). All of them contained at least
    # one assessment/research/profile DOCX.
    #
    # Accept the folder when at least one DMA-shaped DOCX exists at
    # depth ≤ 3. The parser will still emit
    # `docx_only_package_no_manifest` in `parser_warnings` so the
    # import audit can distinguish docx-only ingest from canonical
    # full-package ingest, and downstream synthesis/score steps see
    # an empty `subcap_scores` list (no false scores from a missing
    # workbook).
    if _has_dma_docx_reports(root):
        return root
    for child in root.iterdir():
        if child.is_dir() and _has_dma_docx_reports(child):
            return child

    raise FileNotFoundError(
        f"no DMA package detected under {root} "
        "(expected MANIFEST.json, ≥2 of 01_..08_ canonical subfolders, "
        "or at least one DMA-shaped *.docx within depth 3)"
    )


_DMA_DOCX_NAME_TOKENS = (
    "assessment_report",
    "assessmentreport",
    "client_profile",
    "clientprofile",
    "research_report",
    "researchreport",
    "research_handoff",
    "_dma_",
    "dma_assessment",
    "dma_report",
    "dma_complete",
    # 2026-06-10 corpus-coverage: the 8 remaining discovery failures.
    # IMA/TBOM/YNCU ship `*Digital_Maturity_Assessment*.docx`; Amegy a
    # `*Background Research.docx` — all real DMA deliverables the token
    # list missed.
    "digital_maturity",
    "maturity_assessment",
    "background_research",
)

# Research-artifact CSVs (appendix exports). A folder with >=2 of these
# is a DMA research package even with no DOCX/manifest at all (AAA Club
# Alliance + Midfirst ship exactly this shape: A1_Evidence_Inventory /
# A3_Financial_Trends / A4_Technology_Stack_Map + VIZ PNGs).
_DMA_RESEARCH_CSV_TOKENS = (
    "evidence_inventory",
    "tech_stack_map",
    "technology_stack_map",
    "financial_trends",
    "sentiment_data",
    "assumptions_register",
)


def _has_dma_docx_reports(p: Path, depth: int = 0, max_depth: int = 3) -> bool:
    """True if `p` (or a descendant at depth ≤ max_depth) contains at
    least one *.docx whose filename looks like a DMA report artifact.

    Filename-only check — fast, no DOCX parsing. We deliberately match a
    broad set of name tokens (Assessment_Report, Client_Profile,
    Research_Report, *_DMA_*, etc.) so a folder containing ANY of the
    canonical DMA report shapes is accepted. Unrelated DOCX files
    (e.g. meeting notes, contracts) are skipped — those folders will
    still raise the "no DMA package detected" error, which is the
    correct classification (`SKIPPED_NO_DMA_REPORT`).
    """
    if not p.is_dir() or depth > max_depth:
        return False
    try:
        entries = list(p.iterdir())
    except (PermissionError, OSError):
        return False
    csv_hits = 0
    for entry in entries:
        if not entry.is_file():
            continue
        # Space/underscore drift is rampant in hand-named Drive files
        # ("Amegy_Bank_Background Research.docx") — normalise both.
        name_norm = entry.name.lower().replace(" ", "_")
        suffix = entry.suffix.lower()
        # DOCX report OR the bot's xlsx deliverable ("..._DMA_<date>_x.xlsx"
        # — Navy Federal / Navacord ship ONLY these; the xlsx scoring
        # fallback mines them downstream).
        if suffix in (".docx", ".xlsx") and any(
            tok in name_norm for tok in _DMA_DOCX_NAME_TOKENS
        ):
            return True
        if suffix == ".csv" and any(
            tok in name_norm for tok in _DMA_RESEARCH_CSV_TOKENS
        ):
            csv_hits += 1
    if csv_hits >= 2:
        return True
    if depth < max_depth:
        for entry in entries:
            if entry.is_dir() and _has_dma_docx_reports(
                entry, depth + 1, max_depth
            ):
                return True
    return False


def _maybe(parser, path: Path, warnings: list[str], label: str):
    """Run a parser; on failure, append a TYPED parser warning instead
    of dropping the source error.

    Before 2026-05-26 the helper appended `f"{label}: {e!s}"` only —
    so a JSONDecodeError surfaced as "Expecting value: line 1 column
    1 (char 0)" with no signal that the file was *corrupt* vs
    *missing*. Operators tracking down a failed import couldn't tell
    a missing-manifest from a corrupt-manifest from a schema-mismatch.

    State branches:
      json_corrupt    → "{label}: json_corrupt: <details>"
      file_unreadable → "{label}: io_error: <details>"
      schema_mismatch → "{label}: schema_mismatch: <details>"
      generic         → "{label}: <details>"
    """
    import json as _json
    try:
        text = _read_text(path)
    except OSError as e:
        warnings.append(f"{label}: io_error: {e!s}")
        return None
    try:
        return parser(text)
    except _json.JSONDecodeError as e:
        warnings.append(f"{label}: json_corrupt: line {e.lineno} col {e.colno}: {e.msg}")
        return None
    except (KeyError, ValueError, TypeError) as e:
        # Most parsers raise ValueError/TypeError for schema mismatch;
        # KeyError when a required field is missing post-load.
        warnings.append(
            f"{label}: schema_mismatch: {type(e).__name__}: "
            f"{str(e).split(chr(10), 1)[0][:120]}"
        )
        return None
    except Exception as e:
        # First line only, capped — keeps pydantic's `errors.pydantic.dev`
        # docs URL (emitted on a later line) out of parser_warnings.
        warnings.append(
            f"{label}: {type(e).__name__}: {str(e).split(chr(10), 1)[0][:120]}"
        )
        return None


_ISSUE_CONTENT_WORD_RE = re.compile(r"[a-z][a-z0-9'-]{3,}")


def _issue_text_tokens(text: str) -> set[str]:
    return set(_ISSUE_CONTENT_WORD_RE.findall((text or "").lower()))


def _issue_texts_similar(a: str, b: str) -> bool:
    """Cheap token-overlap similarity for issue dedup/matching — the
    register texts are short compressed clauses ("Barracuda ESG breach:
    34,515 individuals…" vs "Barracuda breach (34.5K affected…)") so a
    content-word overlap ratio over the smaller set is the robust
    signal. Pure + import-light (the TF-IDF path in nlp.similarity is
    overkill for ≤ 25-row registers)."""
    ta, tb = _issue_text_tokens(a), _issue_text_tokens(b)
    if not ta or not tb:
        return False
    overlap = len(ta & tb) / min(len(ta), len(tb))
    return overlap >= 0.5


def _merge_profile_issue_rows(
    issues: list,
    cp_issue_rows: list[dict],
    cp_issue_triggers: list[dict],
    warnings: list[str],
) -> list:
    """Merge Client-Profile-DOCX issue material into the register.

    - DOCX rows matching an existing CLIENT row (same issue_id, or
      similar text) ENRICH it: missing caps / affected ids / status /
      regulator / dates fill in; nothing is overwritten.
    - Unmatched DOCX rows APPEND as lower-priority kind='client' rows
      (namespaced RPT- ids so they can't collide with CSV ids).
    - Trigger-table rows ("Barracuda breach" → P4C4.7 @ 3.0) push cap
      attribution to SUBCAP grain on the row whose text matches.
    """
    from app.schemas.package import IssueRow
    from app.services.parsers.package_csvs import (
        canonical_issue_status,
        compose_dma_impact,
        enrich_issue_row,
        mine_cap_levels,
        mine_p_codes,
        normalize_issue_severity,
    )

    client_rows = [r for r in issues if r.kind == "client"]
    other_rows = [r for r in issues if r.kind != "client"]
    appended = enriched = 0

    def _find_match(d: dict):
        did = (d.get("issue_id") or "").strip()
        for r in client_rows:
            if did and r.issue_id.strip() == did:
                return r
        for r in client_rows:
            if _issue_texts_similar(r.description, d.get("description") or ""):
                return r
        return None

    for d in cp_issue_rows:
        desc = (d.get("description") or "").strip()
        if not desc:
            continue
        caps = dict(d.get("caps") or {})
        cap_cell = " ; ".join(
            s for s in (d.get("capability_impact"), d.get("cap_value")) if s
        )
        if not caps:
            caps = mine_cap_levels(cap_cell)
        affected = mine_p_codes(cap_cell) or list(caps)
        # A flat numeric cap value applies to every affected id.
        if not caps and affected:
            m = re.search(r"\b(\d(?:\.\d+)?)\b", str(d.get("cap_value") or ""))
            if m and 1.0 <= float(m.group(1)) <= 5.0:
                caps = {c: float(m.group(1)) for c in affected}
        match = _find_match(d)
        if match is not None:
            before = (len(match.caps), len(match.affected_categories),
                      match.status, match.regulator)
            for code, level in caps.items():
                match.caps.setdefault(code, level)
            for code in affected:
                if code not in match.affected_categories:
                    match.affected_categories.append(code)
            if not match.status and d.get("status"):
                match.status = d["status"]
            if not match.regulator and d.get("regulator"):
                match.regulator = d["regulator"]
            if match.opened_on is None and d.get("date"):
                from app.services.parsers.package_csvs import _fuzzy_date_iso
                match.opened_on = _fuzzy_date_iso(d["date"])
            match.dma_impact = compose_dma_impact(match) or match.dma_impact
            if before != (len(match.caps), len(match.affected_categories),
                          match.status, match.regulator):
                enriched += 1
            continue
        rid = (d.get("issue_id") or f"RPT-{appended + 1:03d}").strip()
        if not rid.startswith("RPT-"):
            rid = f"RPT-{rid}"
        existing_ids = {r.issue_id for r in client_rows}
        while rid in existing_ids:
            rid = f"{rid}b"
        from app.services.parsers.package_csvs import _fuzzy_date_iso
        row = IssueRow(
            issue_id=rid,
            type=d.get("type"),
            severity=normalize_issue_severity(d.get("severity")),
            status=d.get("status"),
            description=desc,
            evidence_ids=[
                s.strip() for s in (d.get("evidence") or "").split(",")
                if s.strip()
            ],
            cap_ceiling=None,
            affected_categories=affected,
            kind="client",
            regulator=d.get("regulator"),
            opened_on=_fuzzy_date_iso(d.get("date")),
            resolved_on=(
                _fuzzy_date_iso(d.get("date"))
                if canonical_issue_status(d.get("status")) == "RESOLVED"
                else None
            ),
            caps=caps,
        )
        client_rows.append(enrich_issue_row(row))
        appended += 1

    # Trigger table → subcap-grain caps on the matching row.
    trigger_hits = 0
    for trig in cp_issue_triggers:
        text_t = trig.get("trigger") or ""
        level = trig.get("max_score")
        subcaps = trig.get("subcap_ids") or []
        if not text_t or not subcaps:
            continue
        target = None
        for r in client_rows:
            if _issue_texts_similar(r.description, text_t):
                target = r
                break
        if target is None:
            continue
        for sid in subcaps:
            if level is not None:
                target.caps.setdefault(sid, level)
            if sid not in target.affected_categories:
                target.affected_categories.append(sid)
        target.dma_impact = compose_dma_impact(target) or target.dma_impact
        trigger_hits += 1
    if appended or enriched or trigger_hits:
        warnings.append(
            "issues_from_client_profile_docx: "
            f"appended={appended} enriched={enriched} "
            f"trigger_caps={trigger_hits}"
        )
    return client_rows + other_rows


_TECH_STACK_VENDOR_KEYS = (
    "vendor", "company", "technology", "name",
    # 2026-05-28 audit: Calprivate uses a combined `Vendor / Product`
    # column; Nicola starts data 2+ rows into the sheet behind a
    # human-readable preamble. We match by *containment* so combined
    # headers register as vendor.
    "vendor / product", "vendor/product",
    "vendor / technology", "vendor/technology",
)


def _is_tech_stack_header_row(row: tuple) -> bool:
    """A row is the header row if it contains AT LEAST 2 known column
    labels (case-insensitive, whitespace-normalized). Avoids treating
    preamble rows like 'Source: Explorium Live …' as the header.
    """
    if not row:
        return False
    cells = [
        str(c).strip().lower() if c is not None else ""
        for c in row
    ]
    hits = 0
    known = {
        "category", "vendor", "vendor / product", "vendor/product",
        "technology", "product", "company", "name", "layer",
        "confidence", "tier", "evidence id", "evidence_id",
        "deploy status", "deployment_confirmed", "deployment confirmed",
        "presence", "validation method", "validation_method",
    }
    for c in cells:
        if c in known or any(c.startswith(k) for k in known):
            hits += 1
    return hits >= 2


def _parse_explorium_xlsx(path: Path) -> list[TechStackRow]:
    """Parses `08_appendices/*_Explorium_Tech_Stack.xlsx` and its
    variants:
      Alma:    AlmaBank_Explorium_Tech_Stack.xlsx
               sheet 'Alma_Bank_Tech_Stack' (cols Vendor, Product, ...)
      WSFS:    WSFS_Explorium_Tech_Stack.xlsx
               sheet 'WSFS_Tech_Stack'
      Calprivate: CalPrivate_Technographic_Stack_Explorium.xlsx
                  sheet 'Confirmed_Tech_Stack' (col 'Vendor / Product')
      Nicola:  NicolaWealth_Explorium_TechStack_Evidence.xlsx
               sheet 'Confirmed_Tech_Stack' (preamble + header row 3)
      Odlum:   OdlumBrown_Explorium_TechStack.xlsx
               sheet 'Confirmed_Tech_Stack' (col 'Technology')
               first sheet is 'Explorium_Match' — not tech data; skip

    Sheet selection priority:
      1. 'Confirmed_Tech_Stack' (Calprivate / Nicola / Odlum canonical)
      2. any *_Tech_Stack sheet (Alma / WSFS variant)
      3. wb.active fallback (legacy behavior)

    Header detection: scan the first ~6 rows for a row with >= 2
    known column names. Skips human-readable preambles.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if wb is None:
        return []

    ws_candidate = None
    if "Confirmed_Tech_Stack" in wb.sheetnames:
        ws_candidate = wb["Confirmed_Tech_Stack"]
    else:
        for sn in wb.sheetnames:
            if sn.lower().endswith("_tech_stack") and not sn.lower().endswith(
                "_recommendations_map"
            ):
                ws_candidate = wb[sn]
                break
    ws = ws_candidate or wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Locate the real header row within the first 6 rows.
    header_idx = 0
    for i in range(min(6, len(rows))):
        if _is_tech_stack_header_row(rows[i]):
            header_idx = i
            break

    headers = [
        str(c).strip().lower() if c is not None else ""
        for c in rows[header_idx]
    ]

    def _find_value_in(cells: dict, keys: tuple[str, ...]) -> str | None:
        for k in keys:
            v = cells.get(k)
            if v:
                return str(v).strip()
        # Substring fallback: 'vendor / product' header → key contains 'vendor'.
        for hk, hv in cells.items():
            for needle in keys:
                if hv and needle in hk:
                    return str(hv).strip()
        return None

    out: list[TechStackRow] = []
    for raw in rows[header_idx + 1:]:
        if not raw or all(c is None for c in raw):
            continue
        cells = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        vendor = _find_value_in(cells, _TECH_STACK_VENDOR_KEYS)
        if not vendor:
            continue
        # Skip rows where 'vendor' looks like a preamble note.
        if vendor.lower().startswith(("source:", "note:", "—")):
            continue
        out.append(TechStackRow(
            vendor=vendor,
            product=str(
                cells.get("product") or cells.get("subcategory") or ""
            ).strip() or None,
            category=str(cells.get("category") or "").strip() or None,
            layer=str(cells.get("layer") or "").strip() or None,
            confidence=_safe_float(cells.get("confidence")),
            source=str(cells.get("source") or "Explorium").strip(),
        ))
    return out


def _scoring_from_xlsx_fallback(
    scoring_dir: Path,
) -> tuple[list, list[str]]:
    """Extract subcap scores from the per-pillar sheets of the
    assessment workbook (`DMA_*_Workbook_*.xlsx`).

    Covers two real package shapes the export_*.csv path misses:
      - Amalgamated: sheets `P{1..4}_Subcap_Scoring` with cols
        `SubCap ID`, `Category`, `Post-Critic Score`, `Confidence`.
      - AmeriCU: sheets `P{1..4}_Scoring_Detail` with cols
        `SubCap_ID`, `Category_ID`, `Score_1_to_5`.

    Both shapes are matched via flexible header normalisation
    (lowercase, underscores, strip punctuation). Returns `[]` when no
    candidate XLSX is found — caller emits an empty-subcaps warning.
    """
    warnings: list[str] = []
    candidates = list(scoring_dir.glob("*.xlsx"))
    if not candidates:
        return [], warnings
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["openpyxl missing for xlsx scoring fallback"]
    from app.schemas.package import SubcapScoreRow

    def _norm(h: str | None) -> str:
        if h is None:
            return ""
        s = str(h).strip().lower()
        for ch in (" ", "-", "δ", "(", ")", "/", ":"):
            s = s.replace(ch, "_")
        while "__" in s:
            s = s.replace("__", "_")
        return s.strip("_")

    # Column aliases — keys are the canonical fields the parser wants,
    # values are the header-norm forms we accept.
    ALIASES = {
        "subcap_id": {"subcap_id", "sub_cap_id", "subcapid"},
        "subcap_name": {
            "subcap_name", "sub_cap_name", "subcapname",
            "capability_name", "sub_capability_name", "name",
        },
        "category_id": {
            "category", "category_id", "cat_id",
            # Some packages put category in a different column AFTER
            # subcap_id; we extract it from the subcap_id prefix as a
            # final fallback.
        },
        "score": {
            "post_critic_score", "score_1_to_5", "post_critic", "score",
            "final_score", "weighted_score",
        },
        "confidence": {"confidence", "confidence_band", "conf"},
        "evidence_ceiling": {
            "evidence_ceiling", "proxy_ceiling", "ceiling",
        },
        "rationale": {
            "rationale", "score_rationale", "score_rationale_md",
            "notes", "justification",
        },
    }

    rows: list[SubcapScoreRow] = []
    for xlsx in candidates:
        try:
            wb = load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:
            warnings.append(f"xlsx fallback open failed {xlsx.name}: {e}")
            continue
        per_workbook = 0
        for sheet_name in wb.sheetnames:
            sn = sheet_name.lower()
            # Scoring sheets: per-pillar (`P{n}_Subcap_Scoring` /
            # `P{n}_Scoring_Detail`) AND the consolidated single-sheet
            # shape (`Scoring_Workbook` — Zions). The real gate is the
            # header check below (requires subcap_id + score columns),
            # so accepting any "scoring"-named sheet is safe: a non-
            # subcap "scoring" summary sheet yields no header_row and
            # contributes nothing.
            if not (
                "scoring" in sn
                or sn.startswith(("p1_", "p2_", "p3_", "p4_"))
            ):
                continue
            ws = wb[sheet_name]
            header_row = None
            col_idx: dict[str, int] = {}
            for r_idx, raw in enumerate(ws.iter_rows(values_only=True)):
                if header_row is None:
                    headers = [_norm(c) for c in raw]
                    # Need at least subcap_id + score columns to count
                    # as a scoring sheet header.
                    for canon, aliases in ALIASES.items():
                        for i, h in enumerate(headers):
                            if h in aliases:
                                col_idx[canon] = i
                                break
                    if "subcap_id" in col_idx and "score" in col_idx:
                        header_row = r_idx
                    continue
                if not raw or all(c is None for c in raw):
                    continue
                sid = raw[col_idx["subcap_id"]]
                sc = raw[col_idx["score"]]
                if sid is None or sc is None:
                    continue
                sid_s = str(sid).strip()
                try:
                    score = float(sc)
                except (TypeError, ValueError):
                    continue
                # Clamp to the 1-5 maturity domain — drop obvious
                # garbage rows (totals, headers further down).
                if score < 1.0 or score > 5.0:
                    continue
                cat = None
                if "category_id" in col_idx:
                    raw_cat = raw[col_idx["category_id"]]
                    cat = str(raw_cat).strip() if raw_cat is not None else None
                if not cat:
                    # Derive category from subcap_id prefix:
                    # `P1C1.1.1` → `P1C1`; `P3C2.4.2` → `P3C2`.
                    m = _SUBCAP_CAT_RE.match(sid_s)
                    if m:
                        cat = m.group(1)
                if not cat:
                    continue
                conf = None
                if "confidence" in col_idx:
                    cv = raw[col_idx["confidence"]]
                    conf = str(cv).strip() if cv is not None else None
                name = None
                if "subcap_name" in col_idx:
                    nv = raw[col_idx["subcap_name"]]
                    name = str(nv).strip() if nv is not None else None
                    if name == "":
                        name = None
                rationale = None
                if "rationale" in col_idx:
                    rv = raw[col_idx["rationale"]]
                    rationale = str(rv).strip() if rv is not None else None
                    if rationale == "":
                        rationale = None
                rows.append(SubcapScoreRow(
                    subcap_id=sid_s,
                    category_id=cat,
                    score=score,
                    confidence=conf,
                    name=name,
                    rationale=rationale,
                ))
                per_workbook += 1
        if per_workbook:
            warnings.append(
                f"xlsx fallback extracted {per_workbook} subcaps from {xlsx.name}"
            )
    return rows, warnings


# `P1C1.1.1` / `P1C1-1-1` / `P1C1 1.1` → cat `P1C1`. Tolerant of
# whitespace; rejects anything that doesn't start with `P{1..4}C\d+`.
_SUBCAP_CAT_RE = re.compile(r"^(P\d+C\d+)\b")


_NON_PERSON_LEADER = re.compile(
    r"^\s*(no\s|none\b|n/?a\b|tbd\b|vacant\b|unknown\b|leadership\b|"
    r"gap\b|missing\b|without\b|the\s|key\s+leader|executive\s+team|n/a)",
    re.IGNORECASE,
)
_LEADER_PROSE_MARKERS = (
    "gap", "without", "impacts", "ceiling", "absence", "no chief", " no ",
    "lacks", "not detected", "unconfirmed",
)


def _is_person_name(name: object) -> bool:
    """True only for a plausible PERSON name — filters the "Leadership Gaps"
    absence statements ("No CDO", "No Chief Digital Officer", "Leadership
    Gaps") that the DOCX leadership extractor otherwise emits as fake people.
    Leadership is Clay-enriched downstream, so a clean (possibly empty) seed
    beats a list polluted with non-people."""
    n = str(name or "").strip()
    if not (3 <= len(n) <= 60):
        return False
    low = n.lower()
    if _NON_PERSON_LEADER.match(n) or any(m in low for m in _LEADER_PROSE_MARKERS):
        return False
    tokens = [t for t in re.split(r"\s+", n) if t]
    # Reject capability-id-shaped strings ("P4C2.1, P4C2.2, P4C2.6") and
    # digit-heavy tokens that the table extractor confuses for a person row
    # (2026-06-24: ci-financial leaked "P4C2.1, P4C2.2, P4C2.6" as a name).
    if any(re.match(r"^[Pp]\d+C\d", t) for t in tokens):
        return False
    if sum(1 for t in tokens if any(c.isdigit() for c in t)) / len(tokens) > 0.3:
        return False
    capitalised = sum(1 for t in tokens if t[:1].isupper())
    # ≥2 tokens, ≥2 capitalised (first + last name), not a trailing-colon header.
    if not (len(tokens) >= 2 and capitalised >= 2 and not n.rstrip().endswith(":")):
        return False
    # Final gate via the SHARED is_person_name (2026-06-25 leadership remediation,
    # mirrored helper): rejects 'TITLE: Name' colon labels ('CEO: Brandon'),
    # emoji/symbol names ('⚠️ CISO Akerberg'), and status words ('CDO ABSENT')
    # that the checks above let through. Conjunction only tightens — a real
    # person ('Kurt MacAlpine') passes both. Lazy import: stdlib-only, no cycle.
    from app.services import startup_enrich as _se
    return _se.is_person_name(n)


def _safe_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


_PILLAR_NAMES = {
    "P1": "Strategy, Governance & Culture",
    "P2": "Customer Experience & Engagement",
    "P3": "Process Automation & Operations",
    "P4": "Data & AI Enablement",
}


def _evidence_rows_from_json(path: Path, warnings: list[str]) -> list:
    """Build EvidenceRow list from a JSON evidence file.

    Tolerates two shapes surfaced by the 2026-06-07 corpus:
      - items-wrapped: `{"items": [...]}`  (canonical evidence_index.json)
      - top-level list: `[...]`            (Rockland evidence_index_master.json)

    And the common key aliases:
      e_id        <- evidence_id | e_id | id
      source_name <- source_name | source_file | source
      source_url  <- url | source_url | link
      tier        <- tier (int or "T3")
      subcap_mappings <- subcap_mappings | mapped_subcaps | linked_subcap_ids

    Returns [] for missing/malformed input.
    """
    from app.schemas.package import EvidenceRow

    try:
        d = json.loads(_read_text(path))
    except (OSError, json.JSONDecodeError, ValueError) as e:
        warnings.append(f"{path.name}: {e!s}")
        return []
    if isinstance(d, dict):
        # `evidence_items` (Zions shape) + the items-wrapped aliases.
        rows_raw = (
            d.get("items") or d.get("evidence_items") or d.get("evidence")
            or d.get("rows") or d.get("entries") or []
        )
    elif isinstance(d, list):
        rows_raw = d
    else:
        return []
    # Dict-keyed-by-e_id shape (Compeer `batch2_3_consolidated_evidence.json`
    # ships `evidence_items` as {"E-027": {source, finding, …}} rather than a
    # list). Flatten to a list, carrying the key in as the evidence_id so the
    # row keeps its identifier.
    if isinstance(rows_raw, dict):
        rows_raw = [
            {**v, "evidence_id": v.get("evidence_id") or v.get("e_id") or k}
            for k, v in rows_raw.items()
            if isinstance(v, dict)
        ]
    out: list = []
    for raw in rows_raw:
        if not isinstance(raw, dict):
            continue
        e_id = raw.get("evidence_id") or raw.get("e_id") or raw.get("id")
        if not e_id:
            continue
        # Tier may be "T3", a bare int, or a suffixed synthetic tier
        # ("T10-CONTRADICTORY"). Normalize to the canonical source-tier
        # taxonomy ([1, 7]); anything else is honestly None — the prior
        # clamp/default path fabricated tiers (T10 → 8, missing → 5).
        from app.schemas.package import normalize_tier as _norm_tier
        tier = _norm_tier(raw.get("tier", raw.get("ers_tier")))
        subcaps = (
            raw.get("subcap_mappings")
            or raw.get("mapped_subcaps")
            or raw.get("linked_subcap_ids")
            or raw.get("subcap_mapping")   # Compeer (singular)
            or raw.get("maps_to")          # Payments Canada evidence_register
            or []
        )
        if isinstance(subcaps, str):
            subcaps = [s.strip() for s in re.split(r"[;,|]", subcaps) if s.strip()]
        # 2026-06-07 corpus: some JSON-evidence variants (SL Green 525-char,
        # Kitsap 55-char) dump free text into `mapped_subcaps`. The
        # `evidence_index.linked_subcap_ids` column is VARCHAR(32)[], so an
        # over-length / non-subcap-shaped element triggers
        # StringDataRightTruncation and aborts the whole ingest. Keep only
        # well-formed subcap/category IDs (P{1-4}C…) that fit the column.
        if isinstance(subcaps, list):
            subcaps = [
                s for s in subcaps
                if isinstance(s, str) and len(s) <= 32
                and re.match(r"^P\d+C", s.strip())
            ]
        else:
            subcaps = []
        # `evidence_index.e_id` is VARCHAR(16); a malformed long id
        # (Sunflower 58-char) would also truncate-error. Bound it —
        # and WARN, since a truncated id breaks E-ID lookups + dedup
        # matching downstream (2026-06-10 audit, HIGH #7).
        e_id_s = str(e_id)[:16]
        if len(str(e_id)) > 16:
            warnings.append(
                f"evidence_e_id_truncated: {str(e_id)[:40]} -> {e_id_s}")
        with contextlib.suppress(Exception):
            out.append(EvidenceRow(
                e_id=e_id_s,
                source_name=(
                    raw.get("source_name") or raw.get("source_file")
                    or raw.get("source") or raw.get("source_title")  # LPL
                    or "(unnamed)"
                ),
                source_url=(
                    raw.get("url") or raw.get("source_url") or raw.get("link")
                ),
                tier=tier,
                ers=(
                    raw.get("ers") or raw.get("ers_tier_score")
                    or raw.get("ers_score")
                ),
                # Date aliases: `date_published` (Cathay enhanced_evidence),
                # `pub_date` (LPL checkpoint), `date` (Compeer / Payments
                # Canada register — may be a free-text range that the
                # downstream `_publish_date_or_none` coerces to NULL).
                publish_date=str(
                    raw.get("publish_date") or raw.get("date_published")
                    or raw.get("pub_date") or raw.get("date") or ""
                ) or None,
                subcap_mappings=list(subcaps) if isinstance(subcaps, list) else [],
                excerpt=(
                    raw.get("excerpt") or raw.get("summary")     # LPL
                    or raw.get("finding") or raw.get("key_fact")  # Compeer / PMTCAN
                    or raw.get("claim")   # First United per-E-ID claim prose
                    or ""
                ),
                signal_direction=(
                    raw.get("signal_direction") or raw.get("claim_label")
                    or raw.get("claim_type")
                ),
                internal_source=bool(raw.get("internal_source")),
                corroboration_count=raw.get("corroboration_count"),
                # facts[] retained for D5 timeline derivation; tolerant
                # coercion (FactItem.text defaults "") so a malformed fact
                # never aborts the EvidenceRow.
                facts=raw.get("facts") if isinstance(raw.get("facts"), list) else [],
            ))
    return out


def _merge_facts_from_json_twin(
    root_p: Path, evidence: list, warnings: list[str],
) -> None:
    """Fill-if-empty merge of the JSON evidence twin's facts[] +
    per-row excerpt prose into already-assembled evidence rows.

    2026-07 stress-test root cause: when a package ships BOTH
    ``evidence_index.csv`` and ``evidence_index.json``, the CSV wins the
    assembly ladder — but the CSV carries only ``facts_count`` while the
    JSON twin carries the actual fact TEXT (91% of corpus JSON rows).
    The persisted rows then fell to the '(no excerpt)' placeholder,
    capping AE-facing depth everywhere (heatmap synthesis substance,
    EvidenceDrawer excerpts, RAG grounding, insight evidence tabs).

    Strictly additive: only rows with EMPTY facts/excerpt gain the
    twin's values; CSV-carried excerpts are never overwritten.
    """
    rows_needing = [
        ev for ev in evidence
        if not (getattr(ev, "facts", None) or [])
        or not (getattr(ev, "excerpt", "") or "").strip()
    ]
    if not rows_needing:
        return
    # Candidate fact carriers, canonical first. First hit with items wins.
    candidates: list[Path] = [
        root_p / "01_evidence" / "evidence_index.json",
        root_p / "03_scoring_workbook" / "evidence_index.json",
    ]
    for d_sub in ("01_evidence", "08_appendices", "02_research_workbook"):
        sub = root_p / d_sub
        if sub.is_dir():
            candidates.extend(sorted(sub.glob("*[Ee]vidence_[Ii]ndex*.json")))
    twin_rows: list = []
    seen: set[Path] = set()
    for cand in candidates:
        if cand in seen or not cand.is_file():
            continue
        seen.add(cand)
        twin_rows = _evidence_rows_from_json(cand, [])
        if twin_rows:
            break
    if not twin_rows:
        return
    facts_by_eid = {
        r.e_id: r.facts for r in twin_rows if r.e_id and r.facts
    }
    excerpt_by_eid = {
        r.e_id: r.excerpt for r in twin_rows
        if r.e_id and (r.excerpt or "").strip()
    }
    merged_facts = merged_excerpts = 0
    for ev in rows_needing:
        eid = getattr(ev, "e_id", None)
        if not eid:
            continue
        if not (getattr(ev, "facts", None) or []) and eid in facts_by_eid:
            ev.facts = facts_by_eid[eid]
            merged_facts += 1
        if not (getattr(ev, "excerpt", "") or "").strip() \
                and eid in excerpt_by_eid:
            ev.excerpt = excerpt_by_eid[eid]
            merged_excerpts += 1
    if merged_facts or merged_excerpts:
        warn(
            warnings, "evidence_facts_merged_from_json_twin", SEVERITY_INFO,
            f"+{merged_facts} rows gained facts[], +{merged_excerpts} "
            f"gained excerpts from the JSON evidence twin",
        )


def _institution_from_artifacts(
    root_p: Path, warnings: list[str] | None = None,
) -> str | None:
    """Report-derived institution name (2026-06-10 name ladder).

    The institution's real name ships INSIDE the package —
    `entity_profile.json` (`entity_name` in 21 corpus packages,
    `legal_name` in 14) and `research_handoff.json` (`entity` /
    `institution`) — so a missing/garbage manifest name must never
    fall straight through to the Drive FOLDER name (that path put raw
    folder IDs and "… DMA Engagement FINAL" on the live directory).

    Cheap direct JSON reads over the known candidate folders; returns
    the first candidate that passes check_institution_name, else None.
    """
    import json as _json

    from app.services.entity_name_sanity import check_institution_name

    candidate_files: list[Path] = []
    for sub in (
        "00_entity_profile", "01_evidence", "02_research_workbook",
        "08_appendices", "",
    ):
        base = (root_p / sub) if sub else root_p
        candidate_files.extend([
            base / "entity_profile.json",
            base / "research_handoff.json",
        ])
        if base.is_dir():
            candidate_files.extend(sorted(base.glob("*research_handoff*.json")))
    name_keys = (
        "entity_name", "legal_name", "entity", "institution",
        "institution_name",
    )
    for path in candidate_files:
        if not path.is_file():
            continue
        try:
            data = _json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            # Previously a truly-silent drop (audit 12.1): a corrupt
            # entity_profile/research_handoff JSON vanished without a
            # trace and the name ladder skipped a rung invisibly.
            if warnings is not None:
                _msg = (
                    f"{SEVERITY_DEGRADED}/artifact_json_unreadable: "
                    f"{path.name} skipped during institution-name "
                    f"derivation ({type(e).__name__}: {str(e)[:120]})"
                )
                if _msg not in warnings:
                    warnings.append(_msg)
            continue
        if not isinstance(data, dict):
            continue
        for key in name_keys:
            val = data.get(key)
            # research_handoff variants nest the entity block:
            # {"entity": {"name": "...", "legal_name": "..."}}
            if isinstance(val, dict):
                val = val.get("name") or val.get("entity_name") \
                    or val.get("legal_name")
            if isinstance(val, str) and val.strip():
                cand = val.strip()
                junk, _ = check_institution_name(cand)
                if not junk:
                    return cand
    return None


# ── Canonical entity-name resolver (2026-07-06 operator report) ─────────
# Operator: "Some client names are off — is there a client known as
# 'Greenstone Farm Credit Services ACA'? Similarly, which client is TII,
# CCU? Please use full forms as stated within the reports. Always confirm
# these names against the reports."
#
# Root cause: check_institution_name only rejects names of <= 2 chars as
# fragments, so a BARE ACRONYM the pipeline resolved as the institution
# name ("TII", "CCU", "VNO", "AAFCU") sails through the name ladder and
# lands on the live directory as the client's display name. The full
# legal name IS in the package — entity_profile.legal_name/entity_name,
# the research handoff entity block, or the 04_reports DOCX title — so a
# bare acronym must be expanded from those sources before it persists.
#
# The four confirmed cases below were each read out of their own package
# reports (see _KNOWN_ACRONYM_EXPANSIONS provenance comments). The map is
# the report-confirmed FALLBACK for packages that ship no clean self-name.

# Bare acronym: 2-5 upper-case letters (optionally trailing digits), no
# spaces — "TII", "CCU", "VNO", "AAFCU". Real short clients that must NOT
# be treated as acronyms carry spaces / punctuation / mixed case ("IMA
# Financial Group", "Bank OZK", "iQ Credit Union").
_BARE_ACRONYM_RE = re.compile(r"^[A-Z]{2,5}[0-9]{0,2}$")

# Report-confirmed full legal names for the four bare-acronym packages in
# the live corpus. Keyed by the bare acronym (== the entities.display_id
# prefix: "tii"→tii-0001, "ccu"→ccu-0001, "vno"→vno-0001,
# "aafcu"→aafcu-0001). Provenance (verbatim, confirmed against reports):
#   TII   — TII_DMA_Assessment_Report_FINAL.docx: "Travel Insured
#           International, Inc. (TII) is a leading US travel insurance
#           MGA …"; TII_Client_Profile_FINAL.docx title.
#   CCU   — DMA_Assessment_Report_CCUIL_20260504.docx title: "Consumers
#           Credit Union (CCU)" | Lake Forest, Illinois.
#   VNO   — VNO_Assessment_Report_FINAL.docx title "Vornado Realty Trust"
#           (NYSE: VNO); VNO_Research_Handoff.json entity.name.
#   AAFCU — entity_profile.json entity_name / corporate_identity.legal_name
#           "American Airlines Federal Credit Union".
_KNOWN_ACRONYM_EXPANSIONS: dict[str, str] = {
    "TII": "Travel Insured International, Inc.",
    "CCU": "Consumers Credit Union",
    "VNO": "Vornado Realty Trust",
    "AAFCU": "American Airlines Federal Credit Union",
}

# Boilerplate report-cover lines to skip when mining a DOCX title for the
# entity name (the real name is the heading that is NOT one of these).
_REPORT_TITLE_BOILERPLATE_RE = re.compile(
    r"^(digital\s+maturity\s+assessment(\s+report)?"
    r"|dma\s+assessment\s+report"
    r"|assessment\s+report"
    r"|client\s+profile(\s+research)?(\s+report)?"
    r"|background\s+research\s+report"
    r"|table\s+of\s+contents"
    r"|executive\s+summary"
    r"|\d+\.\s+.*)$",
    re.IGNORECASE,
)


def _looks_like_bare_acronym(name: str | None) -> bool:
    """True when ``name`` is a bare acronym (e.g. ``"AAFCU"``)."""
    return bool(_BARE_ACRONYM_RE.match((name or "").strip()))


def _clean_mined_title(raw: str) -> str:
    """Strip report-title decoration to leave the bare legal name.

    "Consumers Credit Union (CCU)" → "Consumers Credit Union"
    "Travel Insured International, Inc. — Digital Maturity Assessment |
     Client Profile Research Report" → "Travel Insured International, Inc."
    """
    v = raw.strip()
    # Report titles glue the entity name to a deliverable label with an
    # em dash, en dash or a pipe; keep only the leading segment. The two
    # dashes are spelled as \u2014 / \u2013 escapes (not literal glyphs)
    # to avoid ambiguous-character lint while still matching the report
    # punctuation exactly.
    v = re.split(r"\s+[\u2014\u2013|]\s+", v, maxsplit=1)[0].strip()
    # Drop a trailing "(ACRONYM)" that merely echoes the short code.
    v = re.sub(r"\s*\(([A-Z0-9&.\- ]{2,10})\)\s*$", "", v).strip()
    return v


def _entity_name_from_report_title(
    root_p: Path, warnings: list[str] | None = None,
) -> str | None:
    """Mine the client's full name from a 04_reports DOCX cover title.

    Pure-stdlib DOCX read (zipfile + regex over ``word/document.xml``) so
    the name path adds no python-docx import cost and runs on the workers
    image. Scans the first handful of paragraphs of each Assessment /
    Client-Profile DOCX, skips the boilerplate cover lines, and returns
    the first heading that passes check_institution_name and is not itself
    a bare acronym. Conservative by design — returns None rather than a
    guess, leaving _KNOWN_ACRONYM_EXPANSIONS as the backstop.
    """
    import zipfile

    from app.services.entity_name_sanity import check_institution_name

    reports_dir = root_p / "04_reports"
    if not reports_dir.is_dir():
        # Some variant packages nest reports one level down.
        candidates = sorted(root_p.glob("*/04_reports"))
        reports_dir = candidates[0] if candidates else reports_dir
    if not reports_dir.is_dir():
        return None
    # Prefer the Client-Profile / Assessment-Report DOCX cover pages.
    docx_paths: list[Path] = []
    for pat in ("*Client_Profile*.docx", "*Assessment_Report*.docx", "*.docx"):
        for p in sorted(reports_dir.glob(pat)):
            if p not in docx_paths:
                docx_paths.append(p)
    for docx in docx_paths:
        try:
            with zipfile.ZipFile(docx) as zf:
                xml = zf.read("word/document.xml").decode("utf-8", "ignore")
        except Exception as e:
            if warnings is not None:
                _msg = (
                    f"{SEVERITY_INFO}/report_title_unreadable: "
                    f"{docx.name} skipped during acronym expansion "
                    f"({type(e).__name__})"
                )
                if _msg not in warnings:
                    warnings.append(_msg)
            continue
        seen = 0
        for para_xml in xml.split("</w:p>"):
            texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", para_xml)
            line = "".join(texts).strip()
            if not line:
                continue
            seen += 1
            if seen > 12:
                break
            if _REPORT_TITLE_BOILERPLATE_RE.match(line):
                continue
            cand = _clean_mined_title(line)
            if not cand or _looks_like_bare_acronym(cand):
                continue
            junk, _ = check_institution_name(cand)
            if not junk and 4 <= len(cand) <= 120:
                return cand
    return None


def _resolve_canonical_entity_name(
    name: str | None,
    root_p: Path,
    warnings: list[str] | None = None,
) -> str:
    """Expand a bare-acronym institution name to its full legal name.

    Returns ``name`` unchanged when it is not a bare acronym. Otherwise
    resolves the report-stated full name via, in order:
      1. entity_profile.json / research_handoff.json name fields
         (``_institution_from_artifacts`` — the package self-describes;
         most authoritative and collision-free, e.g. AAFCU).
      2. the 04_reports DOCX cover title (``_entity_name_from_report_title``
         — generalises to any acronym package that ships a titled report).
      3. ``_KNOWN_ACRONYM_EXPANSIONS`` — the report-confirmed backstop for
         the four live corpus packages that ship no clean self-name
         (TII / CCU / VNO / AAFCU).
    The returned value is only ever a NON-acronym string, so a genuinely
    unknown acronym (no artifact, no title, not in the map) is returned
    unchanged rather than mangled.
    """
    v = (name or "").strip()
    if not _looks_like_bare_acronym(v):
        return v
    # 1. Package self-name (entity_profile / handoff).
    mined = _institution_from_artifacts(root_p, warnings)
    if mined and not _looks_like_bare_acronym(mined):
        return mined
    # 2. Report DOCX cover title.
    title = _entity_name_from_report_title(root_p, warnings)
    if title and not _looks_like_bare_acronym(title):
        return title
    # 3. Report-confirmed fallback map.
    return _KNOWN_ACRONYM_EXPANSIONS.get(v.upper(), v)


def _subvertical_from_artifacts(
    root_p: Path, warnings: list[str] | None = None,
) -> str | None:
    """Report-derived subvertical label (2026-06-10 cohort ladder).

    45 of 95 corpus packages carry NO subvertical in the manifest the
    parser picked — but ship it in research_handoff.json
    (subvertical / subvertical_id / subvertical_name), nested
    run_manifest variants (sub_vertical / subvertical_initial) or
    section_analysis_1.json ("SV7 (Insurance Brokers)"). A NULL
    entities.subvertical silently blanks EVERY peer surface (D1 ticks,
    D3 overlay, RAG cohorts), so scan the known carriers and return the
    first non-empty raw label; package_persist._canonical_subvertical
    maps it to the canonical code.
    """
    import json as _json

    # Depth-3 filename globs rather than fixed folder names: real
    # packages stash the carriers in non-canonical dirs
    # ("04_Governance/run_manifest.json", "01_Research/handoff/
    # research_handoff.json", "08_appendices/assessment_parameters.json").
    candidate_files: list[Path] = []
    for pat in (
        "*research_handoff*.json", "*run_manifest*.json",
        "subvertical_classification.json", "assessment_parameters.json",
        "*parameters*.json", "entity_profile.json",
        "section_analysis_1.json",
    ):
        for depth in ("", "*/", "*/*/"):
            candidate_files.extend(sorted(root_p.glob(f"{depth}{pat}")))
    candidate_files = candidate_files[:80]
    keys = (
        "subvertical_code", "subvertical_id", "subvertical_name",
        "subvertical", "sub_vertical", "sub_vertical_name",
        "subvertical_initial",
    )
    for path in candidate_files:
        try:
            data = _json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            # Previously a truly-silent drop (audit 12.1): a corrupt
            # carrier JSON silently removed a subvertical source and the
            # entity's peer cohort NULLed with no trace.
            if warnings is not None:
                _msg = (
                    f"{SEVERITY_DEGRADED}/artifact_json_unreadable: "
                    f"{path.name} skipped during subvertical derivation "
                    f"({type(e).__name__}: {str(e)[:120]})"
                )
                if _msg not in warnings:
                    warnings.append(_msg)
            continue
        if not isinstance(data, dict):
            continue
        # Top level first, then ONE nesting level — research_handoff /
        # run_manifest variants wrap the block ({"parameter_lock":
        # {"subvertical": ...}}, {"assessment_parameters":
        # {"sub_vertical": ...}}, {"entity": {...}}).
        scopes = [data] + [v for v in data.values() if isinstance(v, dict)]
        for scope in scopes:
            for key in keys:
                val = scope.get(key)
                # Bridgecrest handoff shape (2026-06-10): the value is
                # itself a dict — {"id": "SV3-SF", "name": "Specialty
                # Finance — Auto Lending/Servicing", "confidence": 72}.
                # Prefer the name (richer for the tolerant mapper),
                # fall back to the id.
                if isinstance(val, dict):
                    val = val.get("name") or val.get("id")
                if (
                    isinstance(val, str)
                    and val.strip()
                    and not val.strip().upper().startswith("TBD")
                ):
                    return val.strip()
    return None


def _clean_institution_from_folder(folder_name: str) -> str:
    """Derive a clean institution name from a package folder name.

    The 2026-06-07 34-package validation corpus surfaced folder names
    like `Ameris Bank - DMA`, `SPG - DMA`, `Valley Bank - DMA`,
    `ZipHQ - DMA` where the old fallback (`replace('_',' ').replace
    ('-',' ')`) produced mangled `Ameris Bank   DMA` (triple space +
    retained DMA suffix) shown verbatim in every page header.

    Strips the common DMA-package suffixes, normalizes separators,
    and collapses runs of whitespace. Suffix-stripping is applied
    AFTER underscore->space so `_DMA_Complete_Package` variants match
    too.
    """
    name = folder_name.strip()
    # Strip recognized DMA-package suffixes first (case-insensitive),
    # including the ` - DMA` and ` DMA` trailing forms.
    for suffix in (
        "_DMA_Complete_Package", "_DMA_Complete", "_DMA_Full_Package",
        "_DMA_Deliverable", "_DMA_Deliverables", " - DMA", "_DMA", " DMA",
        "-DMA",
    ):
        if name.lower().endswith(suffix.lower()):
            name = name[: -len(suffix)]
            break
    # Handle DMA markers EMBEDDED mid-name (nested deliverable folders
    # like `Amalgamated_Bank_DMA_2026`, `TII_DMA_Engagement_Package`,
    # `CCU_DMA_CCUIL_20260504`): everything from the first DELIMITED DMA
    # token onward is package / engagement / date metadata, not the
    # institution name. Requires a leading separator so a name that
    # merely starts with "dma" is untouched.
    m = re.search(r"[ _-]DMA([ _-]|$)", name, flags=re.IGNORECASE)
    if m and m.start() > 0:
        name = name[: m.start()]
    # Normalize separators and collapse whitespace.
    name = name.replace("_", " ").replace("-", " ")
    name = re.sub(r"\s+", " ", name).strip()
    # Re-strip a trailing standalone "DMA" token left after separator
    # normalization (e.g. `Penderfund DMA` folder -> `Penderfund`).
    name = re.sub(r"\s+DMA$", "", name, flags=re.IGNORECASE).strip()
    return name or "(unknown)"


def _synthesize_run_manifest_from_parameters(
    root_p: Path, warnings: list[str]
) -> RunManifest | None:
    """Synthesize a RunManifest from `07_governance/00_parameters.json`.

    The 2026-06-07 34-package validation corpus surfaced a manifest-
    source variant (Alliant_Insurance and likely future handoff-mode
    packages) that ships NO run_manifest.json / qa_verdict.json but
    DOES carry `07_governance/00_parameters.json` with:
        assessment_id, research_id, evidence_mode, entity,
        subvertical, subvertical_name, pillar_weights, ...
    This is a richer manifest source than the export-CSV header
    synthesis, so it's tried first.

    Returns None when the file is absent / malformed / missing the
    required `assessment_id` + `entity` anchors.
    """
    from app.schemas.package import RunManifest

    params_path = root_p / "07_governance" / "00_parameters.json"
    if not params_path.exists():
        return None
    try:
        d = json.loads(params_path.read_text())
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(d, dict):
        return None
    run_id = d.get("assessment_id") or d.get("run_id")
    entity = d.get("entity") or d.get("institution_name")
    if not run_id or not entity:
        return None
    # pillar_weights in 00_parameters are integer percentages (20/35/...);
    # normalize to 0-1 floats to match the RunManifest contract.
    raw_weights = d.get("pillar_weights")
    pillar_weights = None
    if isinstance(raw_weights, dict) and raw_weights:
        try:
            total = sum(float(v) for v in raw_weights.values())
            if total > 0:
                pillar_weights = {
                    str(k): float(v) / total if total > 1.5 else float(v)
                    for k, v in raw_weights.items()
                }
        except (TypeError, ValueError):
            pillar_weights = None
    warnings.append(
        f"synthesized run_manifest from 00_parameters.json: "
        f"run_id={run_id}, institution_name={entity}"
    )
    return RunManifest(
        run_id=str(run_id).strip(),
        institution_name=str(entity).strip(),
        research_run_id=(
            str(d.get("research_id")).strip()
            if d.get("research_id") else None
        ),
        evidence_mode=d.get("evidence_mode"),
        subvertical_code=d.get("subvertical"),
        subvertical_name=d.get("subvertical_name"),
        pillar_weights=pillar_weights,
        overall_score=None,
    )


def _synthesize_run_manifest_from_handoff(
    root_p: Path, warnings: list[str]
) -> RunManifest | None:
    """Synthesize a RunManifest from the research handoff JSON.

    Real-sample target: Nicola Wealth ships with no run_manifest.json
    anywhere. The package carries:
      * `02_research_workbook/NicolaWealth_research_handoff.json`
        with `assessment_id`, `entity` (legal name + subvertical),
        `evidence_mode`, `parameter_lock` (rubric_version + skill_version).

    The synthesized request_id is the `assessment_id` from the handoff,
    so it's stable across temp directories, retries, and Drive backfills.

    Returns None if no recognisable handoff file exists. Other samples
    (Alma/Calprivate/Odlum/WSFS) hit the canonical run_manifest path
    first, so this never fires for them.
    """
    from app.schemas.package import RunManifest

    handoff_candidates: list[Path] = []
    # 2026-06-07: added 01_evidence (Alliant_Insurance ships
    # research_handoff.json there, not in 02_research_workbook).
    for sub in (
        "02_research_workbook", "01_evidence", "07_governance", "08_appendices",
    ):
        subdir = root_p / sub
        if subdir.is_dir():
            for p in sorted(subdir.glob("*research_handoff*.json")):
                handoff_candidates.append(p)
            for p in sorted(subdir.glob("research_handoff*.json")):
                if p not in handoff_candidates:
                    handoff_candidates.append(p)
    if not handoff_candidates:
        return None

    for p in handoff_candidates:
        try:
            d = json.loads(p.read_text())
        except (OSError, json.JSONDecodeError):
            continue
        # Real Nicola handoff shape: assessment_id + entity{legal_name,
        # subvertical, ...} + parameter_lock{rubric_version, skill_version}.
        # WSFS handoff also has assessment_id but in a different nesting.
        rid = (
            d.get("assessment_id")
            or d.get("run_id")
            or d.get("l1_run_id")
            or ""
        )
        entity_obj = d.get("entity") or {}
        inst_name = ""
        if isinstance(entity_obj, dict):
            inst_name = (
                entity_obj.get("legal_name")
                or entity_obj.get("name")
                or entity_obj.get("institution")
                or ""
            )
        elif isinstance(entity_obj, str):
            inst_name = entity_obj
        # Fall back to top-level keys some handoffs use.
        if not inst_name:
            inst_name = (
                d.get("entity_legal_name")
                or d.get("entity_name")
                or d.get("institution_name")
                or d.get("institution")
                or ""
            )
        # Final fall-back: parse from folder name.
        if not inst_name:
            inst_name = root_p.name
            for suffix in (
                "_DMA_Complete_Package", "_DMA_Complete", " - DMA", "_DMA",
            ):
                if inst_name.endswith(suffix):
                    inst_name = inst_name[: -len(suffix)]
                    break
            inst_name = inst_name.replace("_", " ").strip() or "(unknown)"

        if not rid:
            # If we have an institution but no assessment_id, fabricate
            # a deterministic run_id from the folder name so re-ingest
            # is idempotent. The hash includes only the institution name
            # so it survives temp-dir variation.
            import hashlib
            digest = hashlib.sha1(inst_name.encode("utf-8")).hexdigest()[:8].upper()
            rid = f"DMA-SYNTH-{digest}-0001"

        sv = None
        if isinstance(entity_obj, dict):
            sv = entity_obj.get("subvertical") or entity_obj.get("sub_vertical")
        sv = sv or d.get("subvertical") or d.get("sub_vertical")

        # parameter_lock carries rubric/skill versions for Nicola.
        pl = d.get("parameter_lock") or {}
        rubric = None
        skill = None
        if isinstance(pl, dict):
            rubric = pl.get("rubric_version") or pl.get("rubric")
            skill = pl.get("skill_version") or pl.get("skill")

        warnings.append(
            f"synthesized run_manifest from handoff: source={p.name}, "
            f"run_id={rid}, institution_name={inst_name}"
        )
        return RunManifest(
            run_id=rid,
            institution_name=inst_name,
            evidence_mode=d.get("evidence_mode"),
            rubric_version=str(rubric) if rubric else None,
            skill_version=str(skill) if skill else None,
            subvertical_name=sv,
            overall_score=None,
        )
    return None


def _synthesize_run_manifest_from_exports(
    root_p: Path, warnings: list[str]
) -> RunManifest | None:
    """Last-resort run-manifest synthesis for packages with no MANIFEST.json
    and no governance run_manifest.json.

    Strategy: scan the export CSVs (located either at
    `03_scoring_workbook/export_*.csv` or
    `03_scoring_workbook/exports/export_*.csv`) for the `# run_id:`
    header comment. The bot pipeline writes this comment as the first
    line of every export CSV — it's the most reliable cross-variant
    source of the run_id. The institution name falls back to the
    package's directory name (with `_DMA_Complete_Package` /
    ` - DMA` suffix stripped).

    Returns None if no run_id can be recovered (means the package is
    structurally too broken to ingest).
    """
    from app.schemas.package import RunManifest

    scoring_dir = root_p / "03_scoring_workbook"
    if not scoring_dir.is_dir():
        return None

    # Search both flat (AlmaBank) and nested (RegionsBank) layouts.
    candidates = list(scoring_dir.glob("export_*.csv"))
    candidates += list(scoring_dir.glob("exports/export_*.csv"))
    candidates += list(scoring_dir.glob("**/export_*.csv"))

    run_id = None
    for c in candidates:
        try:
            with c.open() as f:
                first = f.readline().strip()
            # Format: `# run_id: DMA-ASM-REGIONS-20260518-0001`
            if first.startswith("#") and "run_id" in first.lower():
                parts = first.split(":", 1)
                if len(parts) == 2:
                    run_id = parts[1].strip()
                    break
        except OSError as e:
            # Previously a truly-silent drop (audit 12.1): an unreadable
            # export CSV silently removed a run_id source, so the whole
            # package could hard-fail "no run manifest found" with no
            # trace of WHY the synthesis rung came up empty.
            warn(
                warnings, "export_csv_unreadable", SEVERITY_DEGRADED,
                f"{c} unreadable during run_id synthesis "
                f"({type(e).__name__}: {str(e)[:120]})",
            )
            continue

    if not run_id:
        return None

    # Best-effort institution name from the folder layout (shared
    # cleaner handles ` - DMA` / `_DMA_Complete_Package` / etc. and
    # collapses whitespace — see _clean_institution_from_folder).
    inst_name = _clean_institution_from_folder(root_p.name)

    warnings.append(
        f"synthesized run_manifest from export header: run_id={run_id}, "
        f"institution_name={inst_name}"
    )
    return RunManifest(
        run_id=run_id,
        institution_name=inst_name,
        overall_score=None,
    )


def _synthesize_pillar_rows(
    pillar_scores: dict[str, float] | None,
    pillar_weights: dict[str, float] | None,
) -> list:
    """If `export_pillar_summary.csv` is missing, derive from run_manifest's
    pillar_scores + pillar_weights so D1 Overview still renders."""
    from app.schemas.package import PillarScoreRow

    if not pillar_scores:
        return []
    # Iterate the manifest's OWN pillar keys (sorted) rather than a
    # hardcoded P1-P4 tuple — a future catalogue pillar (P5) would have
    # silently vanished here (2026-06-10 resilience audit, CRITICAL #5).
    out = []
    for pid in sorted(pillar_scores):
        out.append(PillarScoreRow(
            pillar_id=pid,
            pillar_name=_PILLAR_NAMES.get(pid),
            score=float(pillar_scores[pid]),
            weight=(pillar_weights or {}).get(pid),
        ))
    return out


def _score_like_to_float(v: object) -> float | None:
    """'M1.66' / 'M1' / 2.43 / '2.43' → float; anything else → None.

    Several corpus manifests emit maturity-band-prefixed scores
    (Access CU: pillar_scores {"P1": "M1.66"}; overall_maturity
    "M1.58"). The band prefix carries no extra information — strip it.
    """
    if v is None:
        return None
    if isinstance(v, int | float):
        return float(v)
    s = str(v).strip()
    m = re.match(r"^[Mm]?\s*(\d+(?:\.\d+)?)$", s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _reconcile_run_manifest_dict(d: dict) -> dict:
    """Normalize a schema-drifted run_manifest dict onto the RunManifest
    field set (Part 12.1 — 45 corpus packages carried drifted manifests
    that ValidationError'd into the synthesized-manifest path).

    Drift classes measured on the committed corpus:
      - Amalgamated : `entity` is a NESTED DICT {name, ticker,
        sub_vertical, sub_vertical_code}; scores under
        scores.pillars.P#.{post,pre}_critic; pillar_weights +
        assessment_date + evidence_mode under assessment_parameters.
      - Access CU   : pillar_scores values are band strings ("M1.66");
        overall_maturity instead of overall_score.
      - Haventree   : assessment_run_id (not run_id/assessment_id),
        institution_legal_name, framework_version.
      - Rockland / Spokane / Tristate / Virtuity: bookkeeping-only
        manifests (run_id + phases) with no institution name — valid;
        the name ladder fills institution_name downstream.
    """
    ap = d.get("assessment_parameters") if isinstance(
        d.get("assessment_parameters"), dict) else {}
    entity = d.get("entity") if isinstance(d.get("entity"), dict) else {}
    scores = d.get("scores") if isinstance(d.get("scores"), dict) else {}

    run_id = (
        d.get("run_id") or d.get("l1_run_id") or d.get("assessment_id")
        or d.get("assessment_run_id") or ""
    )
    name = (
        d.get("institution_name")
        or (d.get("entity") if isinstance(d.get("entity"), str) else None)
        or entity.get("name")
        or d.get("institution")
        or d.get("entity_name")
        or d.get("entity_legal_name")
        or d.get("institution_legal_name")
        or ""
    )
    # research_run_id variants; `research_handoff` sometimes carries the
    # run id, sometimes a filesystem path — only accept id-shaped values.
    research = d.get("research_run_id") or d.get("l0_run_id")
    if not research:
        rh = d.get("research_handoff")
        if isinstance(rh, str) and re.match(r"^(DMA-|REQ-)", rh.strip()):
            research = rh.strip()

    # Pillar scores: band-string coercion + Amalgamated nesting.
    raw_ps = d.get("pillar_scores")
    pillar_scores: dict[str, float] | None = None
    if isinstance(raw_ps, dict):
        coerced = {
            k: _score_like_to_float(v)
            for k, v in raw_ps.items()
        }
        pillar_scores = {
            k: v for k, v in coerced.items() if v is not None
        } or None
    elif isinstance(scores.get("pillars"), dict):
        nested = {}
        for pid, pv in scores["pillars"].items():
            if isinstance(pv, dict):
                val = _score_like_to_float(
                    pv.get("post_critic", pv.get("pre_critic"))
                )
            else:
                val = _score_like_to_float(pv)
            if val is not None:
                nested[str(pid)] = val
        pillar_scores = nested or None

    raw_pw = d.get("pillar_weights") or ap.get("pillar_weights")
    pillar_weights: dict[str, float] | None = None
    if isinstance(raw_pw, dict):
        coerced_w = {k: _score_like_to_float(v) for k, v in raw_pw.items()}
        pillar_weights = {
            k: v for k, v in coerced_w.items() if v is not None
        } or None

    overall = _score_like_to_float(
        d.get("overall_score")
        if d.get("overall_score") is not None
        else d.get("pillar_weighted_average")
        if d.get("pillar_weighted_average") is not None
        else d.get("overall_maturity")
        if d.get("overall_maturity") is not None
        else scores.get("overall_post_critic", scores.get("overall_pre_critic"))
    )

    rubric = d.get("rubric_version") or d.get("framework_version") \
        or d.get("framework") or d.get("version")

    return {
        "run_id": str(run_id or ""),
        "research_run_id": research,
        "institution_name": str(name or ""),
        "evidence_mode": d.get("evidence_mode") or ap.get("evidence_mode"),
        "rubric_version": str(rubric) if rubric else None,
        "skill_version": d.get("skill_version")
        or d.get("governance_skill_version"),
        "subvertical_code": d.get("subvertical_code")
        or entity.get("sub_vertical_code"),
        "subvertical_name": d.get("subvertical_name") or d.get("subvertical")
        or d.get("sub_vertical") or entity.get("sub_vertical"),
        "pillar_weights": pillar_weights,
        "pillar_scores": pillar_scores,
        "overall_score": overall,
        "assessment_date": d.get("assessment_date")
        or ap.get("assessment_date"),
    }


def _parse_run_manifest_tolerant(
    blob: str,
    warnings: list[str] | None = None,
    label: str = "run_manifest.json",
) -> RunManifest:
    """parse_run_manifest with a schema-drift reconcile rung.

    1. Strict alias-aware parse (package_json.parse_run_manifest) —
       covers the 5 canonical shapes.
    2. When that raises ValidationError (or yields an empty run_id),
       normalize the drifted dict via `_reconcile_run_manifest_dict`
       and validate the result. Reconciled manifests get an
       INFO-severity note; truly-malformed ones raise (→ the caller's
       `_maybe` records a DEGRADED schema_mismatch and the synthesized-
       manifest rungs take over).
    """
    strict_error: Exception | None = None
    try:
        rm = parse_run_manifest(blob)
        if (rm.run_id or "").strip():
            return rm
    except (ValueError, TypeError, KeyError, AttributeError) as e:
        # pydantic ValidationError subclasses ValueError; AttributeError
        # covers non-object JSON roots (a bare list has no .get).
        strict_error = e
    d = json.loads(blob)
    if not isinstance(d, dict):
        raise ValueError(f"{label}: run_manifest is not a JSON object")
    norm = _reconcile_run_manifest_dict(d)
    if not norm["run_id"]:
        # Nothing id-shaped anywhere — genuinely unusable as a manifest.
        raise ValueError(
            f"{label}: no run_id-shaped field after reconcile "
            f"(strict error: {type(strict_error).__name__ if strict_error else 'empty run_id'})"
        )
    # Dates: RunManifest expects date | None; tolerate bad strings.
    ad = norm.get("assessment_date")
    if isinstance(ad, str):
        try:
            from datetime import date as _d
            norm["assessment_date"] = _d.fromisoformat(ad[:10])
        except ValueError:
            norm["assessment_date"] = None
    rm = RunManifest(**norm)
    if warnings is not None and strict_error is not None:
        warn(
            warnings, "run_manifest_reconciled", SEVERITY_INFO,
            f"{label}: schema drift normalized "
            f"({type(strict_error).__name__}: {str(strict_error)[:160]})",
        )
    return rm


def parse_package(root: str | Path) -> IngestedPackage:
    """Walk `root`, return a fully-typed `IngestedPackage`."""
    root_p = _find_root(Path(root))
    log.info("dma_package.parse", root=str(root_p))
    warnings: list[str] = []
    # Self-improvement observation log harvested from sub-parsers.
    # Cap at 200 / package to prevent a runaway workbook from filling
    # the parser_observations table on a single ingest. Each entry
    # passes through to `IngestedPackage.parser_observations` so
    # `package_persist.persist_package` can flush them post-commit.
    _observations_to_emit: list[dict[str, object]] = []

    # ── Manifests ──────────────────────────────────────────────────────
    manifest = None
    top = root_p / "MANIFEST.json"
    if top.exists():
        manifest = _maybe(parse_top_manifest, top, warnings, "MANIFEST.json")

    # 2026-05-28 H6 hotfix: detect docx-only packages (accepted by
    # `_find_root` via the DOCX-only branch) and surface them in
    # parser_warnings so import_audit can classify them distinctly
    # from canonical-package ingests. The warning string is checked
    # by tests + by the admin UI's import-audit summary.
    if (
        manifest is None
        and not (root_p / "07_governance").exists()
        and not (root_p / "03_scoring_workbook").exists()
        and not list(root_p.glob("*qa_verdict*.json"))
        and not list(root_p.glob("run_manifest*.json"))
    ):
        warnings.append(
            "docx_only_package_no_manifest: no MANIFEST.json or "
            "canonical 01_..08_ layout; ingesting DOCX report(s) only"
        )

    # 2026-05-28 deep-extract: when retry-failed-only is active the
    # worker sets DMA_INGEST_LENIENT=1. If we couldn't find a manifest
    # AND the canonical layout is missing, fall through to the deep
    # extractor — it scrapes DOCX text + OCRs images + PDFs as a last
    # resort before giving up. The result is recorded in parser_warnings
    # so the admin UI shows WHICH strategy succeeded.
    import os as _os
    lenient = bool(_os.environ.get("DMA_INGEST_LENIENT", ""))
    if (
        lenient
        and manifest is None
        and not list(root_p.glob("run_manifest*.json"))
    ):
        try:
            from app.services.parsers.deep_extract import deep_extract_folder
            deep = deep_extract_folder(root_p)
            if deep.strategy != "none":
                warnings.append(
                    f"lenient_mode_deep_extract: strategy={deep.strategy} "
                    f"text_chars={len(deep.scraped_text)} "
                    f"ocr_pages={deep.ocr_pages} "
                    f"docx_count={len(deep.docx_paths_scraped or [])}"
                )
                warnings.extend(deep.warnings or [])
        except Exception as e:
            warnings.append(
                f"lenient_mode_deep_extract_failed: "
                f"{type(e).__name__}: {str(e)[:200]}"
            )

    run_mf: RunManifest | None = None
    # Search for the run manifest across canonical AND variant file
    # naming patterns. AlmaBank uses `07_governance/run_manifest.json`
    # (canonical). RegionsBank uses
    # `07_governance/governance_qa_verdict_<entity>_<date>.json` which
    # has `"$schema": "run_manifest_v2"` and `run_id` but lacks the
    # rich scoring fields. Other variants we've seen:
    #   - `07_governance/qa_verdict.json`           (AlmaBank canonical)
    #   - `07_governance/audit_summary.json`        (older variant)
    #   - `08_appendices/run_manifest.json`         (some bot revisions)
    # We try the canonical names first, then fall back to glob.
    manifest_candidates: list[Path] = []
    # Priority 1: canonical `run_manifest.json` at any of the known
    # locations. AmeriCU drops it in `03_scoring_workbook/`; WSFS uses
    # `08_appendices/`; AlmaBank uses `07_governance/`; some bot revisions
    # write to the root.
    for fixed in (
        root_p / "run_manifest.json",
        root_p / "07_governance" / "run_manifest.json",
        root_p / "08_appendices" / "run_manifest.json",
        root_p / "03_scoring_workbook" / "run_manifest.json",
        root_p / "02_research_workbook" / "run_manifest.json",
        # 2026-06-07 corpus: Pentegra Retirement ships its manifest in
        # `01_evidence/run_manifest.json` (misplaced — canonical is
        # 07_governance / 08_appendices). Without this fallback the
        # parser reaches the synthetic-manifest branch with
        # institution_name="" → entity upserts collide on the
        # _display_id_for("") = "entity-0001" default.
        root_p / "01_evidence" / "run_manifest.json",
    ):
        if fixed.exists() and fixed not in manifest_candidates:
            manifest_candidates.append(fixed)
    # Priority 2: CASE-INSENSITIVE scan for any `*run_manifest*.json`.
    # 2026-06-07 corpus: Chemung_Canal_Trust ships
    # `08_appendices/DMA_CCTRUST_Run_Manifest.json` (CamelCase
    # `Run_Manifest`) which the old case-sensitive `glob("*run_manifest
    # *.json")` missed on Linux → hard FAIL "no run manifest found".
    # iterdir + lowercase-compare catches `run_manifest`, `Run_Manifest`,
    # `RUN_MANIFEST` alike. (Same bug class as the original F1 fix.)
    for sub in (".", "07_governance", "08_appendices",
                "03_scoring_workbook", "02_research_workbook"):
        subdir = root_p if sub == "." else (root_p / sub)
        if not subdir.is_dir():
            continue
        for p in sorted(subdir.iterdir()):
            if not p.is_file() or p.suffix.lower() != ".json":
                continue
            if "run_manifest" in p.name.lower() and p not in manifest_candidates:
                manifest_candidates.append(p)
    # Priority 3: variant names — qa_verdict and audit_summary.
    for fixed in (
        root_p / "07_governance" / "qa_verdict.json",
        root_p / "07_governance" / "audit_summary.json",
    ):
        if fixed.exists() and fixed not in manifest_candidates:
            manifest_candidates.append(fixed)
    # Priority 4: glob fallback for variant naming.
    # NOTE: do NOT extend this to match `*QAVerdict.json` (capital QA,
    # Nicola_Wealth shape) — Nicola is handled by the better
    # `_synthesize_run_manifest_from_handoff` fallback below, which
    # produces a richer RunManifest (institution_name, run_id,
    # research_run_id) than the QAVerdict file can. Picking up the
    # variant-cased file here would short-circuit synthesis and leave
    # institution_name blank.
    for sub in ("07_governance", "08_appendices"):
        subdir = root_p / sub
        if subdir.is_dir():
            for p in sorted(subdir.glob("*qa_verdict*.json")):
                if p not in manifest_candidates:
                    manifest_candidates.append(p)
    for candidate in manifest_candidates:
        def _tolerant(blob: str, _label: str = candidate.name) -> RunManifest:
            return _parse_run_manifest_tolerant(
                blob, warnings=warnings, label=_label,
            )
        run_mf = _maybe(_tolerant, candidate, warnings, candidate.name)
        if run_mf is not None:
            if "run_manifest.json" not in candidate.name:
                warnings.append(
                    f"used variant manifest file: {candidate.relative_to(root_p)}"
                )
            break
    if run_mf is None and manifest is not None:
        # Synthesize a minimal RunManifest from the top-level manifest.
        run_mf = RunManifest(
            run_id=manifest.run_id,
            institution_name=manifest.engagement.split("—")[0].strip(),
            overall_score=manifest.overall_score,
        )
        warnings.append("run_manifest.json missing; synthesized from MANIFEST.json")
    if run_mf is None:
        # 2026-05-28 audit fix: Nicola Wealth ships with no
        # run_manifest.json AND no `# run_id:` header in the export CSVs
        # (its exports are operator-named like
        # `NicolaWealth_Scoring_Detail.csv`). The synthesis path
        # _synthesize_run_manifest_from_handoff reads the research
        # handoff JSON which DOES carry `assessment_id` + `entity` +
        # subvertical + evidence_mode for these packages.
        synth = _synthesize_run_manifest_from_handoff(root_p, warnings)
        if synth is not None:
            run_mf = synth
            # NOTE: do NOT append a second "synthesized run manifest" warning
            # here — `_synthesize_run_manifest_from_handoff` already pushes
            # a richer warning string with source filename + run_id +
            # institution_name (line ~709). Adding a generic redundant
            # warning here pushed Nicola's parser_warnings count to 10
            # (right at the `<=10` test cap in test_parse_audit_local.py).
    if run_mf is None:
        # 2026-06-07: handoff-mode packages (Alliant_Insurance) ship
        # `07_governance/00_parameters.json` (assessment_id + entity +
        # subvertical + pillar_weights) but no run_manifest/qa_verdict.
        # Richer than the export-CSV header synthesis below, so tried
        # first.
        synth = _synthesize_run_manifest_from_parameters(root_p, warnings)
        if synth is not None:
            run_mf = synth
    if run_mf is None:
        # 2026-06-07 stress corpus: CASE-INSENSITIVE qa_verdict fallback.
        # Farm Credit Mid America, Interactive Brokers, and Vornado ship
        # their only manifest as a CamelCase `*_QA_Verdict.json` /
        # `L{1,2}_QA_Verdict.json` (no run_manifest, no handoff, no
        # 00_parameters). The case-sensitive Priority-3/4 globs missed
        # them -> hard FAIL "no run manifest found". This runs AFTER the
        # handoff + 00_parameters synthesis so Nicola (which has a richer
        # research_handoff) is unaffected — its handoff path already won.
        # Match any `*verdict*.json` (case-insensitive): covers
        # `*_QA_Verdict.json` (Farm Credit / IBKR / Vornado),
        # `L{1,2}_QA_Verdict.json` (Vornado), and
        # `DMA_GovernanceVerdict_*.json` (American National Bank of Texas).
        for sub in ("07_governance", "08_appendices", "."):
            subdir = root_p if sub == "." else (root_p / sub)
            if not subdir.is_dir():
                continue
            for p in sorted(subdir.iterdir()):
                if not p.is_file() or p.suffix.lower() != ".json":
                    continue
                if "verdict" not in p.name.lower():
                    continue
                cand = _maybe(parse_run_manifest, p, warnings, p.name)
                if cand is not None and cand.run_id:
                    # The verdict file may carry no institution name
                    # (ANBTX GovernanceVerdict has run_id only) — fall
                    # back to the cleaned folder name so the page header
                    # isn't blank.
                    if not (cand.institution_name or "").strip():
                        cand.institution_name = _clean_institution_from_folder(
                            root_p.name
                        )
                    run_mf = cand
                    warnings.append(
                        f"run_manifest from case-insensitive verdict "
                        f"fallback: {p.relative_to(root_p)}"
                    )
                    break
            if run_mf is not None:
                break
    if run_mf is None:
        # LAST-RESORT synthesis: derive run_id from any export CSV's
        # `# run_id:` header comment, and institution_name from the
        # folder name. This keeps RegionsBank-style packages (no
        # MANIFEST.json AND no run_manifest.json) ingestible.
        synth = _synthesize_run_manifest_from_exports(root_p, warnings)
        if synth is not None:
            run_mf = synth
            warnings.append(
                "synthesized run manifest from export CSV headers + folder name"
            )
    if run_mf is None:
        # 2026-05-28 H6 hotfix: if the package was accepted via the
        # DOCX-only branch (no manifest, no canonical subfolders), we
        # can still produce a valid IngestedPackage from the DOCX
        # content alone. Synthesize a minimal RunManifest from the
        # folder name + first DOCX so downstream persist + UI have
        # something to anchor to. The synthesized run_id is a stable
        # hash of the absolute folder path so re-ingest is idempotent.
        is_docx_only = any(
            "docx_only_package_no_manifest" in w for w in warnings
        )
        if is_docx_only:
            import hashlib

            from app.schemas.package import RunManifest as _RunManifest
            digest = hashlib.sha1(str(root_p.resolve()).encode()).hexdigest()[:8].upper()
            inst_name = _clean_institution_from_folder(root_p.name)
            run_mf = _RunManifest(
                run_id=f"REQ-{digest}",
                institution_name=inst_name,
                evidence_mode="docx_only",
                rubric_version=None,
                skill_version=None,
                subvertical_code=None,
                subvertical_name=None,
                pillar_weights=None,
                pillar_scores=None,
                overall_score=None,
            )
            warnings.append(
                f"synthesized run_manifest from docx-only folder: "
                f"run_id={run_mf.run_id}, institution={inst_name}"
            )
        else:
            # Operator-actionable error: list which candidate paths
            # WERE tried and which warnings were emitted at each
            # attempt. The earlier "no run manifest found" message
            # was uninformative — an operator couldn't tell if (a)
            # no file existed at any candidate path, (b) a file
            # existed but was corrupt, or (c) a file existed but
            # failed schema validation.
            tried = [str(c.relative_to(root_p)) for c in manifest_candidates]
            manifest_warnings = [
                w for w in warnings
                if any(k in w for k in ("run_manifest", "qa_verdict", "json_corrupt"))
            ]
            detail = (
                f"no run manifest found under {root_p}. "
                f"tried_paths={tried or '(none — directory has no canonical layout)'}; "
                f"per-file warnings={manifest_warnings or '(none captured)'}"
            )
            raise ValueError(detail)

    # Universal institution-name resolution LADDER (2026-06-10). The
    # old fallback jumped straight from "manifest empty" to the Drive
    # FOLDER name — which is how raw folder IDs
    # ("1NYe2zU3wmBEvd8ZRFWEHpAGIUuK1O1L2") and deliverable noise
    # ("VNO DMA Engagement FINAL") became live entity names. The names
    # ARE in the reports (operator), so prefer report-derived sources:
    #
    #   1. manifest institution_name        (when present AND clean)
    #   2. entity_profile.json              entity_name / legal_name
    #   3. research_handoff.json            entity / institution
    #   4. cleaned folder name              (last resort)
    #
    # Each rung is gated by check_institution_name; a junk winner at
    # the bottom is persisted but PARKED in the admin PENDING_REVIEW
    # queue by package_persist (never AE-visible).
    from app.services.entity_name_sanity import check_institution_name
    _cur = (run_mf.institution_name or "").strip()
    _cur_junk, _cur_reason = check_institution_name(_cur)
    if _cur_junk:
        derived = _institution_from_artifacts(root_p, warnings)
        if derived:
            run_mf.institution_name = derived
            warnings.append(
                f"institution_name derived from report artifacts: "
                f"{derived!r}"
                + (f" (manifest value rejected: {_cur_reason})" if _cur else "")
            )
        else:
            fallback_name = _clean_institution_from_folder(root_p.name)
            fb_junk, fb_reason = check_institution_name(fallback_name)
            if fallback_name and fallback_name != "(unknown)" and not _cur:
                run_mf.institution_name = fallback_name
                warnings.append(
                    f"institution_name backfilled from folder name: "
                    f"{fallback_name!r}"
                )
            if fb_junk and not _cur:
                warnings.append(
                    f"institution_name_junk:{fb_reason} — entity will be "
                    f"parked in PENDING_REVIEW (admin queue)"
                )
            elif _cur:
                warnings.append(
                    f"institution_name_junk:{_cur_reason} — entity will "
                    f"be parked in PENDING_REVIEW (admin queue)"
                )

    # Acronym-expansion pass (2026-07-06 operator report). A resolved
    # institution_name that is a BARE ACRONYM ("TII", "CCU", "VNO",
    # "AAFCU") passes check_institution_name (3-5 letters, not a
    # <=2-char fragment) yet is NOT the full legal name the operator
    # needs on the directory card. The full name IS in the package —
    # entity_profile.legal_name / research_handoff / the 04_reports DOCX
    # title — so expand it here, before it flows to persist_package and
    # becomes entities.name. Runs on the FINAL name regardless of which
    # ladder rung produced it. See _resolve_canonical_entity_name.
    _expanded = _resolve_canonical_entity_name(
        run_mf.institution_name, root_p, warnings,
    )
    if _expanded and _expanded != (run_mf.institution_name or "").strip():
        warnings.append(
            f"institution_name acronym expanded to report legal name: "
            f"{(run_mf.institution_name or '').strip()!r} -> {_expanded!r}"
        )
        run_mf.institution_name = _expanded

    # Subvertical ladder (2026-06-10): when the chosen manifest carries
    # no subvertical, pull the label from the package's own artifacts so
    # entities.subvertical (the peer-cohort key) never silently NULLs.
    if not (run_mf.subvertical_code or "").strip() \
            and not (run_mf.subvertical_name or "").strip():
        sv_label = _subvertical_from_artifacts(root_p, warnings)
        if sv_label:
            run_mf.subvertical_name = sv_label
            warnings.append(
                f"subvertical derived from report artifacts: {sv_label!r}"
            )

    # Validate the run_id parses (REQ-… or DMA-ASM-…); record on warnings if not.
    try:
        parse_run_id(run_mf.run_id)
    except ValueError as e:
        warnings.append(f"run_id format: {e}")

    if manifest is None:
        from app.schemas.package import PackageManifest
        manifest = PackageManifest(
            engagement=run_mf.institution_name,
            run_id=run_mf.run_id,
            overall_score=run_mf.overall_score,
        )

    # ── Scoring tables ─────────────────────────────────────────────────
    # Look for export_*.csv across multiple candidate directories.
    # Some packages nest the final scoring exports outside
    # 03_scoring_workbook/:
    #   • AlmaBank canonical    : 03_scoring_workbook/export_*.csv
    #   • RegionsBank variant   : 03_scoring_workbook/exports/export_*.csv
    #   • Odlum_Brown variant   : 07_governance/scoring_exports/export_*.csv
    #     (the FINAL cap-applied scoring exports; 03_ has only the raw XLSX)
    # We prefer 03_ first (canonical) then 07_governance/scoring_exports/
    # then 08_appendices/ as a last resort. First non-empty match wins.
    #
    # `canonical_scoring_dir` is captured separately because the XLSX
    # fallback further down (Amalgamated / AmeriCU / Calprivate paths)
    # always looks in 03_scoring_workbook/, never in the variant dirs.
    canonical_scoring_dir = root_p / "03_scoring_workbook"
    scoring_candidate_dirs: list[Path] = [
        canonical_scoring_dir,
        root_p / "07_governance" / "scoring_exports",
        root_p / "08_appendices",
    ]
    subcaps = []
    pillars = []
    categories = []
    any_scoring_dir_present = False
    for cand_scoring_dir in scoring_candidate_dirs:
        if not cand_scoring_dir.exists():
            continue
        any_scoring_dir_present = True

        def _first_glob(*patterns: str, _dir: Path = cand_scoring_dir) -> Path | None:
            for pat in patterns:
                hits = sorted(_dir.glob(pat))
                if hits:
                    return hits[0]
            return None

        if not subcaps:
            detail = _first_glob(
                "export_scoring_detail*.csv",
                "exports/export_scoring_detail*.csv",
                "**/export_scoring_detail*.csv",
            )
            if detail and detail.exists():
                parsed = _maybe(parse_scoring_detail_csv, detail, warnings, detail.name)
                if parsed:
                    subcaps = parsed
                    # Emit warning when scoring came from a non-canonical
                    # location so operator sees the variant in flight.
                    if cand_scoring_dir.name != "03_scoring_workbook":
                        try:
                            rel = detail.relative_to(root_p)
                        except ValueError:
                            rel = detail
                        warnings.append(
                            f"used variant scoring CSV location: {rel}"
                        )
                    # Self-improvement: observe unknown headers in the scoring
                    # detail CSV. The observer is pure; the result lives in
                    # `_observations_to_emit` so `package_persist` can flush.
                    try:
                        from app.services.parsers.package_csvs import (
                            SCORING_DETAIL_ALIASES,
                            observe_csv_unknown_columns,
                        )
                        _observations_to_emit.extend(
                            observe_csv_unknown_columns(
                                _read_text(detail),
                                alias_lookup=SCORING_DETAIL_ALIASES,
                                parser_name="package_csvs.parse_scoring_detail_csv",
                                sample_label=detail.name,
                            )[:30]
                        )
                    except Exception as e:
                        warnings.append(
                            f"scoring_detail_observation_failed:{e!s}"
                        )
        if not pillars:
            pillar_csv = _first_glob(
                "export_pillar_summary*.csv",
                "exports/export_pillar_summary*.csv",
                "**/export_pillar_summary*.csv",
            )
            if pillar_csv and pillar_csv.exists():
                parsed = _maybe(parse_pillar_summary_csv, pillar_csv, warnings, pillar_csv.name)
                if parsed:
                    pillars = parsed
        if not categories:
            cat_csv = _first_glob(
                "export_category_summary*.csv",
                "exports/export_category_summary*.csv",
                "**/export_category_summary*.csv",
            )
            if cat_csv and cat_csv.exists():
                parsed = _maybe(parse_category_summary_csv, cat_csv, warnings, cat_csv.name)
                if parsed:
                    categories = parsed
    # Last-resort root-recursive search: some packages ship the canonical
    # export CSVs under a folder NOT in the candidate list above
    # (ProPartners → `04_scoring/exports/`, others under `03_Assessment/`).
    # Rather than enumerate every variant folder name, search the whole
    # package tree so a non-standard layout never silently drops a
    # client's entire score set (2026-06-08 corpus audit: packages with
    # a 708-row export_scoring_detail.csv were ingesting 0 subcaps).
    if not subcaps:
        root_hits = sorted(root_p.glob("**/export_scoring_detail*.csv"))
        if root_hits:
            detail = root_hits[0]
            parsed = _maybe(parse_scoring_detail_csv, detail, warnings, detail.name)
            if parsed:
                subcaps = parsed
                any_scoring_dir_present = True
                try:
                    rel = detail.relative_to(root_p)
                except ValueError:
                    rel = detail
                warnings.append(
                    f"used variant scoring CSV location (root fallback): {rel}"
                )
                try:
                    from app.services.parsers.package_csvs import (
                        SCORING_DETAIL_ALIASES,
                        observe_csv_unknown_columns,
                    )
                    _observations_to_emit.extend(
                        observe_csv_unknown_columns(
                            _read_text(detail),
                            alias_lookup=SCORING_DETAIL_ALIASES,
                            parser_name="package_csvs.parse_scoring_detail_csv",
                            sample_label=detail.name,
                        )[:30]
                    )
                except Exception as e:
                    warnings.append(f"scoring_detail_observation_failed:{e!s}")
    if not pillars:
        root_pillar = sorted(root_p.glob("**/export_pillar_summary*.csv"))
        if root_pillar:
            parsed = _maybe(parse_pillar_summary_csv, root_pillar[0], warnings, root_pillar[0].name)
            if parsed:
                pillars = parsed
    if not categories:
        root_cat = sorted(root_p.glob("**/export_category_summary*.csv"))
        if root_cat:
            parsed = _maybe(parse_category_summary_csv, root_cat[0], warnings, root_cat[0].name)
            if parsed:
                categories = parsed
    if not any_scoring_dir_present:
        warnings.append("03_scoring_workbook missing — no subcap scores ingested")

    # The XLSX fallback + name enrichment below operate on the canonical
    # 03_scoring_workbook (always — even when the CSV win came from
    # 07_governance/scoring_exports/, the assessment workbook XLSX
    # itself lives in 03_).
    scoring_dir = canonical_scoring_dir

    # XLSX fallback — Amalgamated / AmeriCU only ship the assessment
    # workbook (no export_*.csv). Read scores directly from per-pillar
    # sheets (`P1_Subcap_Scoring` / `P1_Scoring_Detail` / variants).
    # This closes the gap where the parser found the run manifest but
    # returned 0 subcap_scores for these two real packages.
    # Try the canonical 03_scoring_workbook first, then ANY folder that
    # ships a scoring workbook xlsx under a non-standard path (LPL →
    # 03_Assessment/workbook/, MidFirst/ATB → other folders). Name-gated
    # to `*scoring*workbook*.xlsx` and the fallback itself is sheet-gated
    # (only P{1..4}_Scoring_Detail / _Subcap_Scoring sheets yield rows),
    # so research/peer workbooks are never mis-read as scores
    # (2026-06-08 corpus audit: full P{n}_Scoring_Detail workbooks sat in
    # non-canonical folders and ingested 0 subcaps).
    if not subcaps:
        xlsx_dirs: list[Path] = []
        if scoring_dir.exists():
            xlsx_dirs.append(scoring_dir)
        for wbk in sorted(root_p.glob("**/*[Ss]coring*[Ww]orkbook*.xlsx")):
            if wbk.parent not in xlsx_dirs:
                xlsx_dirs.append(wbk.parent)
        for xd in xlsx_dirs:
            xlsx_subcaps, xlsx_warnings = _scoring_from_xlsx_fallback(xd)
            warnings.extend(xlsx_warnings)
            if xlsx_subcaps:
                subcaps = xlsx_subcaps
                any_scoring_dir_present = True
                try:
                    rel = xd.relative_to(root_p)
                except ValueError:
                    rel = xd
                warnings.append(
                    f"scoring loaded from xlsx fallback "
                    f"({len(xlsx_subcaps)} subcaps, dir={rel})"
                )
                break

    # XLSX name enrichment — even when the CSV ALREADY provided scores
    # (ALMA-shape `export_scoring_detail.csv` ships `SubCap_ID,Category,
    # Score,Evidence_Ceiling,Caps_Applied,Confidence` with NO
    # `SubCap_Name`), the assessment workbook's per-pillar sheets carry
    # the human-readable `SubCap_Name` ("Digital Strategy Document",
    # "Audience Segmentation"). Pull those names + merge onto the
    # existing rows so the catalogue auto-bootstrap inserts real names
    # into `ccg_subcaps.name` instead of placeholder `Subcap P1C1.1.1`.
    # This is the FE-facing copy on heatmap drawers + insight cards.
    if subcaps and scoring_dir.exists():
        try:
            xlsx_with_names, _ = _scoring_from_xlsx_fallback(scoring_dir)
            name_lookup = {
                r.subcap_id: r.name
                for r in xlsx_with_names if r.name
            }
            rationale_lookup = {
                r.subcap_id: r.rationale
                for r in xlsx_with_names if r.rationale
            }
            enriched_names = 0
            enriched_rationales = 0
            for sc in subcaps:
                if sc.name is None and sc.subcap_id in name_lookup:
                    sc.name = name_lookup[sc.subcap_id]
                    enriched_names += 1
                if sc.rationale is None and sc.subcap_id in rationale_lookup:
                    sc.rationale = rationale_lookup[sc.subcap_id]
                    enriched_rationales += 1
            if enriched_names or enriched_rationales:
                warnings.append(
                    f"xlsx_name_enrichment: +{enriched_names} subcap names "
                    f"+{enriched_rationales} rationales merged from "
                    f"workbook into CSV-parsed rows"
                )
        except Exception as e:
            warnings.append(f"xlsx_name_enrichment_failed: {e!s}")

    if not pillars:
        pillars = _synthesize_pillar_rows(run_mf.pillar_scores, run_mf.pillar_weights)

    # ── Evidence ───────────────────────────────────────────────────────
    evidence = []
    ev_csv = root_p / "01_evidence" / "evidence_index.csv"
    ev_json = root_p / "01_evidence" / "evidence_index.json"
    if ev_csv.exists():
        parsed = _maybe(parse_evidence_csv, ev_csv, warnings, ev_csv.name)
        evidence = parsed or []
    elif ev_json.exists():
        # JSON fallback — the json file wraps rows under `items`.
        try:
            d = json.loads(_read_text(ev_json))
            from app.schemas.package import EvidenceRow, normalize_tier
            for raw in d.get("items", []):
                evidence.append(EvidenceRow(
                    e_id=raw.get("evidence_id") or raw.get("e_id"),
                    source_name=raw.get("source_name", "(unnamed)"),
                    source_url=raw.get("url") or raw.get("source_url"),
                    tier=normalize_tier(raw.get("tier")),
                    ers=raw.get("ers"),
                    publish_date=str(raw.get("publish_date") or "") or None,
                    subcap_mappings=raw.get("subcap_mappings") or [],
                    excerpt=raw.get("excerpt", ""),
                    signal_direction=raw.get("signal_direction"),
                    internal_source=bool(raw.get("internal_source")),
                    corroboration_count=raw.get("corroboration_count"),
                    # facts[] retained for D5 timeline derivation.
                    facts=raw.get("facts") if isinstance(raw.get("facts"), list) else [],
                ))
        except Exception as e:
            warnings.append(f"evidence_index.json: {e!s}")
    else:
        warnings.append("01_evidence missing — no evidence rows ingested")

    # Evidence variant fallback — many packages ship the per-E-ID inventory
    # under a non-canonical NAME *and* a non-canonical FOLDER:
    #   01_evidence/A1_Evidence_Inventory.csv          (Amalgamated)
    #   08_appendices/A1_Evidence_Inventory.csv        (Acuity, Bank of Utah, …)
    #   02_research_workbook/A1_Evidence_Inventory.csv
    # The prior 01_evidence-only fallback missed ~40% of the corpus, dropping
    # the source's E-ID<->subcap linkage and leaving why-now / findings with no
    # traceable evidence (the 2026-06-26 depth audit's `wn_ev=0.0` cohort).
    # Search all three canonical folders; `parse_evidence_csv` already maps the
    # `subcaps_supported` column the A1 inventory uses.
    if not evidence:
        for ev_dir in (root_p / "01_evidence", root_p / "08_appendices",
                       root_p / "02_research_workbook"):
            if not ev_dir.is_dir():
                continue
            ev_variant_candidates = [
                ev_dir / "A1_Evidence_Inventory.csv",
                ev_dir / "A1_evidence_index_full.csv",
            ]
            ev_variant_candidates.extend(sorted(ev_dir.glob("*[Ee]vidence*[Ii]nventory*.csv")))
            ev_variant_candidates.extend(sorted(ev_dir.glob("*evidence_index*.csv")))
            for cand in ev_variant_candidates:
                if cand.exists():
                    parsed = _maybe(parse_evidence_csv, cand, warnings, cand.name)
                    if parsed:
                        evidence = parsed
                        warnings.append(
                            f"evidence loaded from variant {cand.name} ({ev_dir.name})"
                        )
                        break
            if evidence:
                break

    # JSON evidence variant fallback (2026-06-07 corpus) — Rockland_Trust
    # ships `01_evidence/evidence_index_master.json` as a TOP-LEVEL LIST
    # (not wrapped in `items`) with keys `evidence_id` / `mapped_subcaps`
    # / `ers_tier_score`. Corporate_America ships `evidence_index_L0.json`.
    # `_evidence_rows_from_json` tolerates both the items-wrapped and
    # top-level-list shapes plus the common key aliases.
    if not evidence and (root_p / "01_evidence").is_dir():
        ev_dir = root_p / "01_evidence"
        json_variant_candidates: list[Path] = [
            ev_dir / "evidence_index_master.json",
            ev_dir / "evidence_index_L0.json",
            ev_dir / "evidence_unified_by_capability.json",
        ]
        json_variant_candidates.extend(
            sorted(ev_dir.glob("*evidence_index*.json"))
        )
        seen_json: set[Path] = set()
        for cand in json_variant_candidates:
            if cand in seen_json or not cand.exists():
                continue
            seen_json.add(cand)
            rows = _evidence_rows_from_json(cand, warnings)
            if rows:
                evidence = rows
                warnings.append(
                    f"evidence loaded from json variant {cand.name} "
                    f"({len(rows)} rows)"
                )
                break

    # Misplaced-evidence fallback (2026-06-07 corpus) — First Citizens
    # and Security Finance ship `evidence_index.csv` inside
    # `03_scoring_workbook/` (not `01_evidence/`), with `01_evidence/`
    # holding only research/assessment subdirs or per-pillar JSON. Scan
    # the scoring dir as a last resort so the EvidenceDrawer populates.
    if not evidence:
        for misplaced in (
            root_p / "03_scoring_workbook" / "evidence_index.csv",
            root_p / "03_scoring_workbook" / "evidence_index.json",
        ):
            if not misplaced.exists():
                continue
            if misplaced.suffix == ".csv":
                parsed = _maybe(
                    parse_evidence_csv, misplaced, warnings, misplaced.name
                )
                rows = parsed or []
            else:
                rows = _evidence_rows_from_json(misplaced, warnings)
            if rows:
                evidence = rows
                warnings.append(
                    f"evidence loaded from misplaced "
                    f"03_scoring_workbook/{misplaced.name} ({len(rows)} rows)"
                )
                break

    # ── Non-canonical evidence layouts (2026-06 corpus sweep) ──────────
    # 8 packages shipped a populated evidence index the chain above
    # missed, leaving the EvidenceDrawer empty (QA: zero evidence_index
    # rows for ANB Texas, Cathay, Compeer, Farm Credit Mid-America,
    # Guaranteed Rate, Interactive Brokers, LPL, Payments Canada). Causes:
    #   • case-sensitive glob miss — `ANBTX_Evidence_Index_*.json`,
    #     `FCMA_Evidence_Index.csv`, `Evidence_Index_GRATE_*.csv` (the
    #     chain above globs lowercase `*evidence_index*`, Linux is
    #     case-sensitive so the capitalised names never matched).
    #   • alt 01_evidence names — `*consolidated_evidence*.json`,
    #     `evidence_register.json`.
    #   • alt folders — `08_appendices/enhanced_evidence.json` (Cathay),
    #     `02_research_workbook/*Evidence*.json` (IBKR),
    #     `01_Research/**/*evidence_checkpoint*.json` (LPL).
    # Fires ONLY when evidence is still empty, so the packages that already
    # loaded are untouched. `_evidence_rows_from_json` is self-protecting
    # (no e_id → 0 rows); among candidates the one yielding the MOST rows
    # wins so a short summary/mapping never beats the full index.
    if not evidence:
        alt_candidates: list[Path] = []
        ev_dir = root_p / "01_evidence"
        if ev_dir.is_dir():
            alt_candidates.extend(sorted(ev_dir.glob("*[Ee]vidence*.json")))
            alt_candidates.extend(sorted(ev_dir.glob("*[Ee]vidence*.csv")))
        for sub in ("08_appendices", "02_research_workbook"):
            d_sub = root_p / sub
            if d_sub.is_dir():
                alt_candidates.extend(sorted(d_sub.glob("*[Ee]vidence*.json")))
        # `01_Research` (capital R) + any research-named top dir, recursive
        # — LPL nests its checkpoint under 01_Research/checkpoints/.
        for research_dir in root_p.glob("0?_[Rr]esearch*"):
            if research_dir.is_dir():
                alt_candidates.extend(
                    sorted(research_dir.glob("**/*[Ee]vidence*.json"))
                )
        best_rows: list = []
        best_name = ""
        seen_alt: set[Path] = set()
        for cand in alt_candidates:
            if cand in seen_alt or not cand.exists():
                continue
            seen_alt.add(cand)
            if cand.suffix == ".csv":
                rows = _maybe(parse_evidence_csv, cand, warnings, cand.name) or []
            else:
                rows = _evidence_rows_from_json(cand, warnings)
            if len(rows) > len(best_rows):
                best_rows = rows
                best_name = cand.name
        if best_rows:
            evidence = best_rows
            warnings.append(
                f"evidence loaded from non-canonical layout {best_name} "
                f"({len(best_rows)} rows)"
            )

    # ── Research workbook per-pillar evidence (extends CSV) ────────────
    # Real packages ship a 17-sheet per-pillar research workbook with
    # detailed evidence rows the CSV evidence_index doesn't capture. The
    # CSV is the "headline" 100-row evidence list; the workbook's
    # per-pillar sheets carry the full ~700-row evidence trail with
    # richer subcap_mappings (multi-pillar coverage on a single E-ID).
    #
    # State-branches (merge):
    #   csv_only        — research workbook missing → CSV rows are the
    #                      only source.
    #   union_new_eids  — workbook has E-IDs not in CSV → append as new
    #                      EvidenceRow entries.
    #   merge_subcaps   — workbook E-ID matches a CSV row → union the
    #                      subcap_mappings (workbook usually has more
    #                      complete coverage; CSV may carry sentinels).
    #   workbook_only   — CSV is empty / corrupt → workbook rows act as
    #                      the canonical evidence list.
    rw_dir = root_p / "02_research_workbook"
    if rw_dir.is_dir():
        try:
            from openpyxl import load_workbook as _load_wb

            from app.schemas.package import EvidenceRow as _EvRow
            from app.services.parsers.research_workbook import (
                parse_per_pillar_sheets as _parse_per_pillar,
            )
            rw_xlsx = next(iter(rw_dir.glob("*.xlsx")), None)
            if rw_xlsx is not None:
                wb = _load_wb(rw_xlsx, data_only=True)
                per_pillar = _parse_per_pillar(wb)
                csv_by_eid = {ev.e_id: ev for ev in evidence if ev.e_id}
                added = 0
                merged = 0
                for r in per_pillar.rows:
                    if not r.e_id:
                        continue
                    existing = csv_by_eid.get(r.e_id)
                    if existing is not None:
                        # Union the subcap mappings — workbook usually
                        # carries the full pillar coverage (5-10 subcaps
                        # per E-ID) while CSV may have 1 or sentinel.
                        union = list(dict.fromkeys(
                            (existing.subcap_mappings or [])
                            + (r.linked_subcap_ids or [])
                        ))
                        if len(union) > len(existing.subcap_mappings or []):
                            existing.subcap_mappings = union
                            merged += 1
                    else:
                        evidence.append(_EvRow(
                            e_id=r.e_id,
                            source_name=r.source_name or "(unknown)",
                            source_url=r.source_url,
                            tier=r.tier,
                            ers=None,
                            publish_date=r.published_date,
                            subcap_mappings=list(r.linked_subcap_ids or []),
                            excerpt=r.excerpt or "(no excerpt)",
                            signal_direction=r.claim_type or "EVIDENCE",
                            internal_source=False,
                            corroboration_count=None,
                        ))
                        added += 1
                if added or merged:
                    warnings.append(
                        f"research_workbook_evidence: +{added} new E-IDs, "
                        f"{merged} CSV rows enriched with workbook subcaps "
                        f"(from {rw_xlsx.name})"
                    )
                for w in per_pillar.warnings[:10]:
                    warnings.append(
                        f"research_workbook:{w.get('kind','?')}:"
                        f"{w.get('sheet','-')}"
                    )
                # Self-improvement: harvest unknown-column observations
                # so package_persist can flush them into the
                # parser_observations table after commit. Stored in
                # IngestedPackage so the persist layer (sole owner of
                # the DB session) gets the data without a separate
                # plumbing channel. Cap at 50 / package to defend
                # against degenerate workbooks with thousands of
                # spurious headers.
                _observations_to_emit.extend(per_pillar.observations[:50])
        except Exception as e:
            warnings.append(f"research_workbook_evidence_failed: {e!s}")

    # ── Facts/excerpt merge from the JSON evidence twin (2026-07) ──────
    # Fill-if-empty: rows whose winning source carried no fact text /
    # excerpt gain them from the sibling evidence_index.json so the
    # persist layer can compose real excerpts instead of the
    # '(no excerpt)' placeholder. Never overrides CSV-carried prose.
    try:
        _merge_facts_from_json_twin(root_p, evidence, warnings)
    except Exception as e:
        warn(
            warnings, "evidence_facts_merge_failed", SEVERITY_DEGRADED,
            f"{type(e).__name__}: {str(e)[:160]}",
        )

    # ── Issue register (client-business + assessment-QA) ───────────────
    # 2026-07-06 rework (Context-page defect family): the corpus ships
    # TWO register classes and the old first-non-empty pick let the
    # 07_governance QA checklist (rows ABOUT the deliverable's own
    # files: "Missing governance artifact: caps_applied_log.csv")
    # shadow the client's REAL issues in 08_appendices
    # (A5_issue_register.csv / A6_issues_register.csv — Barracuda ESG
    # breach, FDIC Consent Order…) in 22/113 corpus packages. Now:
    #   1. discovery is recursive + case-insensitive + plural-tolerant
    #      (A5_Issue_Register.csv, A6_issues_register.csv, registers in
    #      01_evidence / 02_research_workbook/exports all found);
    #   2. every candidate is header-classified client vs assessment_qa
    #      (package_csvs.classify_issue_register_headers);
    #   3. the BEST client register wins the client rows; the QA
    #      register's rows are KEPT with kind='assessment_qa' (Health
    #      page material) and namespaced ids — downstream AE surfaces
    #      filter on kind.
    issues = []
    _issue_name_re = re.compile(r"issues?_?register|issueregister", re.I)
    issue_csv_candidates: list[Path] = []
    canonical_csv = root_p / "07_governance" / "issue_register.csv"
    if canonical_csv.exists():
        issue_csv_candidates.append(canonical_csv)
    for p in sorted(root_p.rglob("*.csv")):
        if _issue_name_re.search(p.name) and p not in issue_csv_candidates:
            issue_csv_candidates.append(p)
    client_best: list = []
    client_src: Path | None = None
    qa_best: list = []
    qa_src: Path | None = None

    def _client_rank(path: Path, rows: list) -> tuple:
        rel = str(path.relative_to(root_p)).lower()
        titled = sum(1 for r in rows if (r.description or "").strip())
        attributed = sum(1 for r in rows if r.affected_categories or r.caps)
        # Appendix/evidence/research registers are the analyst's client
        # deliverable; 07_governance is the bot's QA area — prefer the
        # former on ties.
        dir_pref = 0 if "07_governance" not in rel else -1
        return (titled, attributed, dir_pref)

    for issue_csv in issue_csv_candidates:
        parsed = _maybe(parse_issue_register_csv, issue_csv, warnings, issue_csv.name)
        # Self-improvement: observe unknown headers regardless of
        # whether the parse succeeded — even an empty-row CSV with
        # variant headers carries learning signal.
        try:
            from app.services.parsers.package_csvs import (
                ISSUE_REGISTER_ALIASES,
                observe_csv_unknown_columns,
            )
            _observations_to_emit.extend(
                observe_csv_unknown_columns(
                    _read_text(issue_csv),
                    alias_lookup=ISSUE_REGISTER_ALIASES,
                    parser_name="package_csvs.parse_issue_register_csv",
                    sample_label=issue_csv.name,
                )[:30]
            )
        except Exception as e:
            warnings.append(
                f"issue_register_observation_failed:{e!s}"
            )
        if not parsed:
            continue
        if parsed[0].kind == "assessment_qa":
            if qa_src is None or len(parsed) > len(qa_best):
                qa_best, qa_src = parsed, issue_csv
        else:
            if client_src is None or \
                    _client_rank(issue_csv, parsed) > _client_rank(client_src, client_best):
                client_best, client_src = parsed, issue_csv
    if not client_best:
        # JSON client registers (Alma layer1_issue_register.json carries
        # the REAL regulatory issues; WSFS assessment_issue_register.json).
        for j in (
            root_p / "07_governance" / "layer1_issue_register.json",
            root_p / "07_governance" / "assessment_issue_register.json",
        ):
            if j.exists():
                try:
                    d = json.loads(_read_text(j))
                    from app.schemas.package import IssueRow
                    from app.services.parsers.package_csvs import (
                        enrich_issue_row,
                    )
                    for raw in d.get("issues", []):
                        # Severity normalization (WSFS uses MATERIAL).
                        sev = str(raw.get("severity") or "MEDIUM").upper()
                        if sev == "MATERIAL":
                            sev = "HIGH"
                        # Evidence may be a CSV string (ALMA) or a list (WSFS).
                        ev_field = raw.get("evidence") or raw.get("evidence_ids", "")
                        ev_ids = (
                            [s.strip() for s in ev_field.split(",") if s.strip()]
                            if isinstance(ev_field, str) else list(ev_field or [])
                        )
                        # Cap value may be `cap_ceiling` (ALMA) or `cap_value` (WSFS).
                        cap_v = raw.get("cap_ceiling")
                        if cap_v is None:
                            cap_v = raw.get("cap_value")
                        # Affected categories: WSFS calls this `capabilities_affected`.
                        affected = (
                            raw.get("affected_categories")
                            or raw.get("capabilities_affected")
                            or []
                        )
                        desc = str(raw.get("description") or "").strip()
                        if not desc:
                            continue  # untitleable — never a blank row
                        client_best.append(enrich_issue_row(IssueRow(
                            issue_id=raw.get("issue_id") or raw.get("id") or "",
                            type=raw.get("type"),
                            severity=sev,
                            status=raw.get("status"),
                            description=desc,
                            evidence_ids=ev_ids,
                            cap_formula=raw.get("cap_formula") or raw.get("cap_logic"),
                            cap_ceiling=cap_v if isinstance(cap_v, int | float) else None,
                            affected_categories=affected,
                            kind="client",
                        )))
                    if client_best:
                        client_src = j
                    break
                except Exception as e:
                    warnings.append(f"{j.name}: {e!s}")
    # Merge: client rows first; QA rows kept but namespaced so a QA
    # "ISS-001" can never collide with (and silently clobber) the
    # client's "ISS-001" under the (run_id, issue_id) unique key.
    issues = list(client_best)
    client_ids = {r.issue_id for r in issues}
    for r in qa_best:
        if r.issue_id in client_ids or not r.issue_id.startswith("QA-"):
            r.issue_id = f"QA-{r.issue_id}"
        issues.append(r)
    if client_src is not None and str(client_src.name) != "issue_register.csv":
        warnings.append(
            f"used variant issue register: {client_src.relative_to(root_p)}"
        )

    # ── C5 (2026-06-07): L1/L2 governance distinction ──────────────────
    # `qa` (L2/final) keeps the legacy semantics: read the canonical
    # `qa_verdict.json` or any case-insensitive variant. Then load the
    # NEW `qa_l1` from the L1-specific filename variants. 2 of 5 real
    # fixtures (Odlum + Calprivate) ship both; the other 3 only ship
    # L2 (qa_l1 stays None — Gates tab renders "L1 not reported").
    qa = None
    qa_l1 = None

    # L2/final candidates in priority order: canonical 07_governance
    # names; the Calprivate `GOV_` variant; OZK ships its verdict only in
    # `08_qa/qa_verdict.json` and/or `07_governance/governance_verdict.json`
    # (no "qa_verdict" token), so those explicit names come next. Finally a
    # broad `*verdict*.json` sweep over BOTH dirs (07_governance before
    # 08_qa) catches any future layout. L1 files are excluded throughout —
    # they go through the L1 loader below.
    gov_dir = root_p / "07_governance"
    qa_dir = root_p / "08_qa"
    l2_candidates = [
        gov_dir / "qa_verdict.json",
        gov_dir / "GOV_qa_verdict.json",
        gov_dir / "L2_qa_verdict.json",
        # `governance_verdict.json` is the OZK / TrustCo / Penderfund (L2)
        # name — same schema as qa_verdict.json, just a different token.
        gov_dir / "governance_verdict.json",
        gov_dir / "L2_governance_verdict.json",
        gov_dir / "l2_governance_verdict.json",
        # 08_qa explicit names (OZK).
        qa_dir / "qa_verdict.json",
        qa_dir / "GOV_qa_verdict.json",
        qa_dir / "L2_qa_verdict.json",
        qa_dir / "governance_verdict.json",
    ]

    def _is_l1_name(name_lower: str) -> bool:
        return (
            "l1" in name_lower
            or "layer1" in name_lower
            or "layer_1" in name_lower
        )

    # Broad sweep over 07_governance THEN 08_qa (order fixed so a package
    # carrying a verdict in both dirs deterministically prefers
    # 07_governance). Match any `*verdict*.json` that isn't an L1 file and
    # isn't an issue-register / partial / patch artifact. The `.json`
    # suffix guard already rejects `*.md` summaries.
    for sweep_dir in (gov_dir, qa_dir):
        if not sweep_dir.is_dir():
            continue
        for p in sorted(sweep_dir.iterdir()):
            if not p.is_file() or p.suffix.lower() != ".json":
                continue
            name_lower = p.name.lower()
            if (
                ("verdict" in name_lower or "qaverdict" in name_lower)
                and p not in l2_candidates
                and not _is_l1_name(name_lower)
                # Reject non-verdict artifacts that happen to carry a
                # related token (e.g. `issue_register_qa_gov.json`).
                and "issue" not in name_lower
                and "partial" not in name_lower
                and "patch" not in name_lower
            ):
                l2_candidates.append(p)
    for qa_json in l2_candidates:
        if qa_json.exists():
            qa = _maybe(parse_qa_verdict, qa_json, warnings, qa_json.name)
            if qa is not None:
                break

    # L1 candidates — distinct filenames per real fixture:
    #   Odlum     : 07_governance/L1_qa_verdict.json
    #   Calprivate: 07_governance/Layer1_qa_verdict.json
    # Tolerant of case-insensitive variants for future packages.
    l1_named = [
        root_p / "07_governance" / "L1_qa_verdict.json",
        root_p / "07_governance" / "Layer1_qa_verdict.json",
        root_p / "07_governance" / "layer1_qa_verdict.json",
    ]
    for l1_path in l1_named:
        if l1_path.exists():
            qa_l1 = _maybe(parse_qa_verdict, l1_path, warnings, l1_path.name)
            if qa_l1 is not None:
                break
    # Case-insensitive sweep for variants (Nicola pattern would put L1
    # in an entity-prefixed file).
    if qa_l1 is None and gov_dir.is_dir():
        for p in sorted(gov_dir.iterdir()):
            if not p.is_file() or p.suffix.lower() != ".json":
                continue
            name_lower = p.name.lower()
            # Match L1 marker AND verdict marker; reject issue-register
            # files (they have a different schema; not a verdict).
            if (
                ("l1" in name_lower or "layer1" in name_lower or "layer_1" in name_lower)
                and ("qa_verdict" in name_lower or "qaverdict" in name_lower)
                and "issue_register" not in name_lower
            ):
                qa_l1 = _maybe(parse_qa_verdict, p, warnings, p.name)
                if qa_l1 is not None:
                    break

    # Recursive fallback: qa_verdict.json ships in non-canonical locations
    # across the corpus — at the package root (GESA), `04_Governance/`
    # (LPL), `08_qa/` (OZK), lowercase `l1_` (Penderfund). Sweep the whole
    # resolved root, matching case/underscore-insensitively, and classify
    # L1 vs L2/final by name marker. Future-proof: any new layout is caught.
    if qa is None or qa_l1 is None:
        verdict_files = sorted(
            p for p in root_p.rglob("*.json")
            if "qaverdict" in p.name.lower().replace("_", "")
            and "issue" not in p.name.lower()
        )

        def _is_l1(p: Path) -> bool:
            n = p.name.lower().replace("_", "").replace("-", "")
            return "l1" in n or "layer1" in n

        if qa is None:
            for p in verdict_files:
                if _is_l1(p):
                    continue
                qa = _maybe(parse_qa_verdict, p, warnings, p.name)
                if qa is not None:
                    warnings.append(f"qa_verdict_from: {p.relative_to(root_p).as_posix()}")
                    break
        if qa_l1 is None:
            for p in verdict_files:
                if _is_l1(p):
                    qa_l1 = _maybe(parse_qa_verdict, p, warnings, p.name)
                    if qa_l1 is not None:
                        break

    if qa_l1 is not None and qa is not None:
        warnings.append(
            f"qa_verdict_l1_l2_pair: L1={qa_l1.verdict!r} "
            f"L2={qa.verdict!r} (2-stage QA escalation captured)"
        )

    # ── C7 (2026-06-07): bot governance audit logs (D6 Audit tab) ─────
    # 2 of 5 real fixtures ship at least one component:
    #   Nicola : 07_governance/reasoning_chain_log.json (12 subcap chains)
    #            + contradiction_log.csv
    #   Odlum  : 07_governance/contradiction_log.csv
    # Surfaces "the bot reached M2.0 via these 5 reasoning steps;
    # contradiction CONTRA-001 was resolved E-113 wins" defensibility
    # to D6 Audit tab for analyst review.
    from app.services.parsers.governance_audit import (
        parse_governance_audit_logs,
    )
    audit_logs_data = None
    try:
        audit_logs_data = parse_governance_audit_logs(root_p)
        if audit_logs_data is not None:
            n_chain = len(audit_logs_data.reasoning_chain)
            n_contra = len(audit_logs_data.contradictions)
            warnings.append(
                f"governance_audit_logs: reasoning_chain={n_chain} "
                f"contradictions={n_contra}"
            )
    except Exception as e:
        warnings.append(f"governance_audit_logs_parse_failed: {e!s}")

    # ── C11 (2026-06-07): assumptions register (D1 ClientOverview footer) ─
    # 2 of 5 real fixtures ship the analyst's assumptions list:
    #   Calprivate: 08_appendices/assumptions_register.json (5 entries)
    #   Nicola:     07_governance/A9_Assumptions_Register.csv
    # Surfaces "we assumed X because Y" defensible rationale on D1.
    from app.schemas.package import AssumptionRow

    assumptions_data: list[AssumptionRow] = []
    assumption_candidates: list[Path] = [
        root_p / "08_appendices" / "assumptions_register.json",
        root_p / "07_governance" / "assumptions_register.json",
        root_p / "07_governance" / "A9_Assumptions_Register.csv",
        root_p / "07_governance" / "assumptions_register.csv",
        root_p / "08_appendices" / "assumptions_register.csv",
    ]
    # Case-insensitive sweep for variants (entity-prefixed file names).
    for sub in ("07_governance", "08_appendices"):
        d = root_p / sub
        if d.is_dir():
            for p in sorted(d.iterdir()):
                if not p.is_file():
                    continue
                name_lower = p.name.lower()
                if (
                    "assumption" in name_lower
                    and (name_lower.endswith(".json") or name_lower.endswith(".csv"))
                    and p not in assumption_candidates
                ):
                    assumption_candidates.append(p)
    for asm_path in assumption_candidates:
        if asm_path.exists():
            try:
                from app.services.parsers.assumptions_register import (
                    parse_assumptions_register,
                )
                parsed = parse_assumptions_register(asm_path)
                if parsed:
                    assumptions_data = parsed
                    warnings.append(
                        f"assumptions_register: {len(parsed)} entry(ies) "
                        f"parsed from {asm_path.relative_to(root_p)}"
                    )
                    break
            except Exception as e:
                warnings.append(
                    f"assumptions_register_parse_failed: "
                    f"{asm_path.name}: {e!s}"
                )

    # ── C10 (2026-06-07): caps-applied event log (D6 Health Gates tab) ──
    # 4 of 5 real fixtures ship `07_governance/caps_applied_log.csv`
    # (Alma 8 caps, Calprivate 115, Nicola 8, Odlum 10; WSFS embeds
    # equivalent semantics in the per-subcap `caps_applied` string on
    # SubcapScoreRow). Surfaces "this subcap scored M2.5 because IR-003
    # severity capped it" defensibility on D6.
    from app.schemas.package import CapsAppliedRow

    caps_applied_log_data: list[CapsAppliedRow] = []
    # caps_applied_log.csv ships in 07_governance (canonical) OR
    # 03_scoring_workbook / nested dirs — search the resolved root.
    caps_log_path = next(
        (p for p in (
            root_p / "07_governance" / "caps_applied_log.csv",
            *sorted(root_p.glob("**/caps_applied_log.csv")),
        ) if p.exists()),
        root_p / "07_governance" / "caps_applied_log.csv",
    )
    if caps_log_path.exists():
        try:
            from app.services.parsers.caps_applied_log import (
                parse_caps_applied_log,
            )
            caps_applied_log_data = parse_caps_applied_log(caps_log_path)
            if caps_applied_log_data:
                warnings.append(
                    f"caps_applied_log: {len(caps_applied_log_data)} cap "
                    f"event(s) parsed from {caps_log_path.relative_to(root_p)}"
                )
        except Exception as e:
            warnings.append(f"caps_applied_log_parse_failed: {e!s}")

    # ── Recommendations (D4) ───────────────────────────────────────────
    # Real-sample variants observed in the 5 uploaded zips:
    #   Alma:       08_appendices/recommendations_detail.json (canonical, 7 recs)
    #   Odlum:      07_governance/recommendations_register.json (variant, 6 recs)
    #   Calprivate: NO rec source — bot pipeline shipped no recs JSON
    #   WSFS:       NO rec source — same
    #   Nicola:     NO rec source — same
    # The 3 source-side gaps are F4 (2026-06-06 v2 QA): see
    # `docs/qa/qa_5folder_live_findings.md` §F4. We surface a single
    # parser_warning so the absence is observable (admin import audit
    # surfaces it; AE-facing recs panel falls back to its empty state).
    recs = []
    # Recommendations ship under several filenames AND schemas across the
    # corpus; parse_recommendations_any normalises all of them. Ordered:
    # canonical detail/register first, then the broadly-shipped
    # recommendations.json (37 pkgs) / 06_recommendations.json (13 pkgs).
    _rec_dirs = (
        "08_appendices", "07_governance", "02_research_workbook",
        "03_scoring_workbook",
    )
    _rec_names = (
        "recommendations_detail.json", "recommendations_register.json",
        "recommendations.json", "06_recommendations.json",
        # observed one-off variants across the corpus (2026-07-15 residual
        # audit): capitalised-key summary + alternate names. parse_recommendations_any
        # + _normalize_rec normalise all of them.
        "recommendation_summary.json", "recommendations_final.json",
        "recommendations_master.json", "recommendations_complete.json",
        "recommendations_ALL.json", "recommendations_validated.json",
        "phase_6_recommendations.json", "assessment_06_recommendations.json",
        "layer1_recommendations_ranked.json",
    )
    rec_candidates: list[Path] = [
        root_p / d / n for n in _rec_names for d in _rec_dirs
    ]
    # Recursive fallback: rec files also land in non-canonical dirs
    # (ProPartners ships 06_recommendations.json under 06_reports/). Append
    # any matches not already covered by the canonical-dir candidates.
    for n in _rec_names:
        for p in sorted(root_p.glob(f"**/{n}")):
            if p not in rec_candidates:
                rec_candidates.append(p)
    rec_source_found = False
    for rec_json in rec_candidates:
        if rec_json.exists():
            rec_source_found = True
            parsed = _maybe(
                parse_recommendations_any, rec_json, warnings, rec_json.name
            )
            if parsed:
                recs = parsed
                if rec_json.name != "recommendations_detail.json":
                    warnings.append(
                        f"used variant recommendations source: "
                        f"{rec_json.relative_to(root_p)}"
                    )
                break

    # Phase-split fallback: some packages split recs across
    # recommendations_6a/6b/6c.json — combine them when no single file won.
    if not recs:
        split_parts: list[Path] = sorted(
            p for d in _rec_dirs
            for p in (root_p / d).glob("recommendations_6[a-z].json")
            if p.exists()
        )
        for part in split_parts:
            rec_source_found = True
            parsed = _maybe(
                parse_recommendations_any, part, warnings, part.name
            )
            if parsed:
                recs.extend(parsed)
        if recs:
            warnings.append(
                f"recommendations from {len(split_parts)} phase-split files"
            )


    # NOTE: when no JSON rec source loaded, we ALSO try a DOCX-§9
    # fallback that mines the Assessment_Report.docx prose. That
    # fallback runs AFTER `report_sections` is populated below — see
    # the "F4 DOCX fallback" block lower in this function.

    # ── Peer set (D3 overlay) ──────────────────────────────────────────
    peers = []
    peers_dir = root_p / "06_peers"
    if peers_dir.exists():
        for peer_file in sorted(peers_dir.glob("peer_scores_*.json")):
            p = _maybe(parse_peer_score, peer_file, warnings, peer_file.name)
            if p is not None:
                peers.append(p)
        # Variant fallback — multiple bot revisions ship a single
        # benchmarks/peer-set JSON instead of one
        # `peer_scores_<peer>.json` per peer.
        #   Amalgamated / AmeriCU → `peer_set.json`
        #   Calprivate            → `peer_benchmarks.json` (top-level peers list)
        #   Nicola                → `02_peer_benchmarks.json` (peer_set + benchmarks dict)
        #   Odlum                 → `benchmarks.json` (top-level benchmarks dict)
        # Synthesize PeerScore rows so the D3 peer overlay can render.
        if not peers:
            from app.schemas.package import PeerScore
            variants_to_try = (
                peers_dir / "peer_set.json",
                peers_dir / "peer_benchmarks.json",
                peers_dir / "02_l1_peer_benchmarks.json",
                peers_dir / "02_peer_benchmarks.json",
                peers_dir / "benchmarks.json",
            )
            for variant in variants_to_try:
                if not variant.exists():
                    continue
                try:
                    d = json.loads(_read_text(variant))
                    # Nicola/Odlum shape: peer_set is a list of strings
                    # + peer_overall_scores is a dict {peer_name: score}
                    # + benchmarks is a dict of category_id → peer_scores.
                    # Detect by presence of `peer_overall_scores` (Nicola)
                    # OR a peers list of strings (Odlum).
                    peer_set = d.get("peer_set")
                    overall_scores = d.get("peer_overall_scores")
                    benchmarks_dict = d.get("benchmarks")
                    if (
                        isinstance(peer_set, list)
                        and peer_set
                        and isinstance(peer_set[0], str)
                    ):
                        # String-list shape — synthesize one row per name.
                        # Real-sample case: Nicola has peer_set entries like
                        # "Connor, Clark & Lunn Private Capital" but
                        # peer_overall_scores keys are "CC&L_Private_Capital".
                        # Build a normalized index so the lookup still works.
                        def _normalize_peer_key(name: str) -> str:
                            n = (
                                name.lower()
                                .replace(",", "")
                                .replace("+", "")
                                .replace(".", "")
                                .replace("&", " and ")
                            )
                            for tok in ("management", "associates", "limited",
                                        "private capital", "wealth", "ltd",
                                        "investment counsel"):
                                n = n.replace(tok, "")
                            n = " ".join(n.split())
                            return "".join(c for c in n if c.isalnum())

                        scored_index: dict[str, float] = {}
                        if isinstance(overall_scores, dict):
                            for k, v in overall_scores.items():
                                fv = _safe_float(v)
                                if fv is not None:
                                    scored_index[_normalize_peer_key(str(k))] = fv
                        for peer_name in peer_set:
                            scores = {}
                            # Exact match first, then normalized fallback.
                            if isinstance(overall_scores, dict):
                                ov = _safe_float(overall_scores.get(peer_name))
                                if ov is None:
                                    nk = _normalize_peer_key(peer_name)
                                    # Try prefix-substring match in normalized space.
                                    for key, val in scored_index.items():
                                        if key and (key in nk or nk[: max(4, len(key))] == key[: max(4, len(key))]):
                                            ov = val
                                            break
                                if ov is not None:
                                    scores["overall"] = ov
                            # Aggregate by-pillar averages from benchmarks
                            # so D3 overlay has something to render.
                            if isinstance(benchmarks_dict, dict):
                                pillar_acc: dict[str, list[float]] = {}
                                for cat_id, cat in benchmarks_dict.items():
                                    if not isinstance(cat, dict):
                                        continue
                                    pillar = cat.get("pillar") or cat_id[:2]
                                    ps = cat.get("peer_scores") or {}
                                    if isinstance(ps, dict):
                                        v = _safe_float(ps.get(peer_name))
                                        if v is not None:
                                            pillar_acc.setdefault(
                                                pillar, []
                                            ).append(v)
                                for pid, vals in pillar_acc.items():
                                    if vals:
                                        scores[pid] = sum(vals) / len(vals)
                            peers.append(PeerScore(
                                peer_id=peer_name,
                                peer_name=peer_name,
                                ticker=None,
                                assets=None,
                                rationale=None,
                                scores=scores,
                            ))
                    else:
                        # Existing dict-list shape (Amalgamated / Calprivate / etc.)
                        items = (
                            d.get("peers")
                            or peer_set
                            or benchmarks_dict
                            or (d if isinstance(d, list) else [])
                        )
                        for raw in items if not isinstance(items, dict) else items.values():
                            if not isinstance(raw, dict):
                                continue
                            pid = (
                                raw.get("peer_id") or raw.get("ticker")
                                or raw.get("name") or raw.get("entity")
                            )
                            if not pid:
                                continue
                            scores = {}
                            if isinstance(raw.get("pillar_scores"), dict):
                                for k, v in raw["pillar_scores"].items():
                                    f = _safe_float(v)
                                    if f is not None:
                                        scores[str(k)] = f
                            overall = _safe_float(
                                raw.get("overall_score")
                                or raw.get("composite_score")
                                or raw.get("score")
                            )
                            if overall is not None:
                                scores.setdefault("overall", overall)
                            peers.append(PeerScore(
                                peer_id=str(pid),
                                peer_name=str(raw.get("name") or pid),
                                ticker=raw.get("ticker"),
                                assets=str(raw.get("assets") or "") or None,
                                rationale=raw.get("rationale"),
                                scores=scores,
                            ))
                    if peers:
                        warnings.append(
                            f"peers loaded from variant {variant.name} "
                            f"({len(peers)} peers)"
                        )
                        break
                except Exception as e:
                    warnings.append(f"{variant.name}: {e!s}")

    # ── Tech stack (Explorium) ─────────────────────────────────────────
    # Real-sample variants observed:
    #   Alma:        08_appendices/AlmaBank_Explorium_Tech_Stack.xlsx
    #   WSFS:        08_appendices/WSFS_Explorium_Tech_Stack.xlsx
    #   Calprivate:  08_appendices/CalPrivate_Technographic_Stack_Explorium.xlsx
    #                                       ^^^^^^^^^^^^^ different word order
    #   Odlum:       08_appendices/OdlumBrown_Explorium_TechStack.xlsx
    #                                                  ^^^^^^^^^ no underscore
    #   Nicola:      08_appendices/NicolaWealth_Explorium_TechStack_Evidence.xlsx
    tech: list[TechStackRow] = []
    appx = root_p / "08_appendices"
    if appx.exists():
        tech_patterns = (
            "*Explorium*Tech_Stack*.xlsx",
            "*Explorium*TechStack*.xlsx",
            "*Technographic*Stack*Explorium*.xlsx",
            "*Tech_Stack*Explorium*.xlsx",
            "*TechStack*Explorium*.xlsx",
        )
        seen: set[Path] = set()
        for pat in tech_patterns:
            for xlsx in sorted(appx.glob(pat)):
                if xlsx in seen:
                    continue
                seen.add(xlsx)
                try:
                    # Part 9.1: the Explorium xlsx path runs through the same
                    # taxonomy gate as the CSV/JSON/prose paths — this feed
                    # was the audit's heaviest noise source (verbatim cells).
                    parsed = sanitize_tech_rows(
                        _parse_explorium_xlsx(xlsx), warnings=warnings,
                    )
                    if parsed:
                        tech = parsed
                        if "Explorium_Tech_Stack" not in xlsx.name:
                            warnings.append(
                                f"used variant tech stack xlsx: {xlsx.name}"
                            )
                        break
                except Exception as e:
                    warnings.append(f"{xlsx.name}: {e!s}")
            if tech:
                break

    # Structured-variant fallback: most packages ship the tech stack as
    # A4_Tech_Stack_Map.csv / tech_inventory.json / tech_stack.json rather
    # than the Explorium xlsx. load_tech_stack normalises all of them
    # (taxonomy-sanitized inside the leaf parsers).
    if not tech:
        tech = load_tech_stack(root_p, warnings=warnings)
        if tech:
            warnings.append(f"tech_stack_from_variant_source: {len(tech)} entries")

    # ── Firmographics (from research_handoff.json) ─────────────────────
    # Real-sample variant locations:
    #   WSFS:        02_research_workbook/research_handoff.json (canonical)
    #   Nicola:      02_research_workbook/NicolaWealth_research_handoff.json
    #   Calprivate:  07_governance/research_handoff.json (NEW location)
    #   Alma/Odlum:  no handoff JSON; firmographics come from DOCX later.
    firm = None
    # D5: client-research-report financial highlights (captured in the
    # client-profile block below; the A#_financial_trends.csv CSV is
    # preferred when present).
    cp_financials: dict = {}

    # ── C9 (2026-06-07): structured entity_profile.json takes priority ─
    # Some packages (Calprivate confirmed; likely future variants) ship
    # `08_appendices/entity_profile.json` with richly-structured
    # firmographics. When present, it's preferred over the DOCX regex
    # path because:
    #   - JSON is structured (no regex fragility);
    #   - It carries fields the DOCX prose typically omits (ticker,
    #     founded date, branch_count);
    #   - Financial baseline gives `total_assets` directly vs. the
    #     fragile $-amount regex on prose.
    # Subsequent merges (handoff JSON, then DOCX regex) only fill in
    # missing fields via the additive F5b-1 / F5c paths below.
    # entity_profile.json ships in different folders across the corpus:
    # 08_appendices (Calprivate nested variant) AND the flat standalone
    # variant in 01_evidence / 02_research_workbook / 08_appendices.
    # Search all three; first existing wins. The parser auto-detects the
    # nested-vs-flat schema.
    entity_profile_path = next(
        (
            p
            for sub in (
                "08_appendices", "00_entity_profile",
                "01_evidence", "02_research_workbook",
            )
            for p in [root_p / sub / "entity_profile.json"]
            if p.exists()
        ),
        root_p / "08_appendices" / "entity_profile.json",
    )
    if entity_profile_path.exists():
        try:
            from app.schemas.package import Firmographics, LeadershipPerson
            from app.services.parsers.entity_profile import (
                parse_entity_profile_json,
                parse_entity_profile_leadership,
            )
            ep_fields = parse_entity_profile_json(entity_profile_path)
            ep_leadership_raw = parse_entity_profile_leadership(
                entity_profile_path
            )
            if ep_fields:
                # Build a Firmographics from the structured JSON.
                ep_leadership = [
                    LeadershipPerson(
                        name=p["name"],
                        title=p.get("title", ""),
                        last_verified_date=None,
                    )
                    for p in ep_leadership_raw
                    if _is_person_name(p.get("name"))
                ]
                # Pop the `branches` extra so it lands in model_extras.
                branches = ep_fields.pop("branches", None)
                firm = Firmographics(
                    **ep_fields,
                    leadership=ep_leadership,
                )
                if branches:
                    firm.branches = branches  # extra='allow'
                warnings.append(
                    f"firmographics_from_entity_profile_json: "
                    f"fields={sorted(ep_fields.keys())} "
                    f"leadership={len(ep_leadership)}"
                )
        except Exception as e:
            warnings.append(
                f"entity_profile_json_parse_failed: {e!s}"
            )

    handoff_paths: list[Path] = [
        root_p / "02_research_workbook" / "research_handoff.json",
        root_p / "07_governance" / "research_handoff.json",
        root_p / "08_appendices" / "research_handoff.json",
        # 2026-06-07: Alliant_Insurance ships it in 01_evidence.
        root_p / "01_evidence" / "research_handoff.json",
    ]
    # Plus any named-variant handoff (Nicola pattern).
    for sub in ("02_research_workbook", "07_governance", "01_evidence"):
        d = root_p / sub
        if d.is_dir():
            for p in sorted(d.glob("*research_handoff*.json")):
                if p not in handoff_paths:
                    handoff_paths.append(p)
    # Only consult handoff JSON if the entity_profile.json path above
    # didn't already populate `firm`. The entity_profile.json is richer
    # (ticker, founded date, branch_count, total_assets); handoff JSON
    # is the canonical fallback for packages that don't ship the richer
    # structured profile.
    if firm is None:
        for rh in handoff_paths:
            if rh.exists():
                firm = _maybe(parse_firmographics, rh, warnings, rh.name)
                if firm is not None:
                    if rh.name != "research_handoff.json":
                        warnings.append(
                            f"used variant research_handoff: {rh.relative_to(root_p)}"
                        )
                    break

    # ── Client profile DOCX (2026-05-28 audit + 2026-05-29 focus_areas) ─
    # When research_handoff.json is absent, fall back to the client
    # profile DOCX (`04_reports/*ClientProfile*.docx` /
    # `*Client_Profile*.docx`) for firmographics. Independently of
    # firmographics state, we ALWAYS scan the same DOCX for focus_areas
    # — focus areas are a separate artifact (Top Findings / Critical
    # Gaps callouts with verbatim quote + source path + page number),
    # not derived from firmographics. The 2026-05-29 finalization
    # surfaced that focus_areas were extracted by the parser but
    # only the COUNT was logged — never propagated onto the
    # IngestedPackage envelope, so the focus_areas table stayed
    # empty even though every package contained them.
    #
    # Real-sample shape (Alma/Odlum have the DOCX but no handoff):
    #   Alma:  04_reports/AlmaBank_ClientProfile_Research_Report.docx
    #   Odlum: 04_reports/OdlumBrown_ClientProfile_FINAL.docx
    # Calprivate ships BOTH — DOCX still parsed to enrich the handoff
    # if it has missing fields.
    focus_areas_collected: list[FocusAreaRow] = []
    # D5 surfaces mined from the Client Profile report (fallbacks assembled
    # below): the "Digital Evolution Timeline" table, the "Sentiment Overview"
    # table, and the "Acquisition History" table. Used only as fallbacks /
    # additive merges — zero regression risk to packages that already populate.
    cp_timeline_candidates: list = []
    cp_sentiment: dict = {}
    cp_acquisitions: list = []
    # D2 Part 5.1 PRIMARY-rung material: normalized ProfileFinding rows
    # mined from the Client Profile Research Report's Key Findings /
    # Strategic Priorities / Digital Evolution / Technology Landscape
    # sections (82/113 packages ship the report; it fed ZERO cards
    # before this). Collected across every parsed profile DOCX.
    cp_profile_findings: list = []
    cp_issue_rows: list = []
    cp_issue_triggers: list = []
    reports_dir = root_p / "04_reports"
    cp_candidates: list[Path] = []
    if reports_dir.is_dir():
        for pat in (
            "*ClientProfile*.docx",
            "*Client_Profile*.docx",
            "*client_profile*.docx",
        ):
            for p in sorted(reports_dir.glob(pat)):
                if p not in cp_candidates:
                    cp_candidates.append(p)
    # Broaden to non-canonical layouts + variant names (e.g. "<Client>
    # Background Research.docx", profile docs outside 04_reports) when the
    # narrow 04_reports search finds nothing — the 2026-06-23 corpus audit
    # found ~30 packages whose only roster lives in such a DOCX. Reuse the
    # healer's discovery (broad profile regex + peer/evidence exclusions) so a
    # competitor report is never pulled in.
    if not cp_candidates:
        from app.services.entity_healing import _profile_docx_paths
        cp_candidates = _profile_docx_paths(root_p)[:3]
    if cp_candidates:
        for cp_docx in cp_candidates:
            try:
                from app.schemas.package import Firmographics, LeadershipPerson
                from app.services.parsers.client_profile import (
                    parse_client_profile_path,
                )
                cp_result = parse_client_profile_path(cp_docx)
            except Exception as e:
                warnings.append(
                    f"client_profile_docx_parse_failed:{cp_docx.name}:{e!s}"
                )
                continue
            if cp_result.state_kind == "no_docx_found":
                continue
            # Capture the report's Digital Evolution Timeline / Sentiment /
            # Acquisition History (first DOCX that carries each wins) for the
            # D5 fallbacks assembled below.
            if not cp_timeline_candidates and cp_result.timeline_events:
                cp_timeline_candidates = list(cp_result.timeline_events)
            if not cp_sentiment and getattr(cp_result, "sentiment", None):
                cp_sentiment = dict(cp_result.sentiment)
            if not cp_acquisitions and getattr(cp_result, "acquisition_events", None):
                cp_acquisitions = list(cp_result.acquisition_events)
            # 2026-07-06: Risk & Issues mining (first DOCX carrying each
            # wins) — merged into `issues` after this block, deduped
            # against the CSV register rows.
            if not cp_issue_rows and getattr(cp_result, "issue_rows", None):
                cp_issue_rows = list(cp_result.issue_rows)
            if not cp_issue_triggers and getattr(
                    cp_result, "issue_cap_triggers", None):
                cp_issue_triggers = list(cp_result.issue_cap_triggers)
            # D2 Part 5.1: profile findings feed the insight ladder's
            # PRIMARY rung (collected across every parsed DOCX; the
            # builder dedups near-identical observations).
            if getattr(cp_result, "profile_findings", None):
                cp_profile_findings.extend(cp_result.profile_findings)
            # Operator note: the client research report carries financial
            # highlights — capture them as the D5 fallback when no structured
            # A#_financial_trends.csv ships.
            if cp_result.financial_highlights and not cp_financials:
                cp_financials = dict(cp_result.financial_highlights)
            # Collect focus_areas from EVERY successfully parsed DOCX
            # (idempotent re-ingest is handled by the persist layer's
            # DELETE-then-INSERT pattern).
            for fa in (cp_result.focus_areas or []):
                focus_areas_collected.append(
                    FocusAreaRow(
                        title=fa.title,
                        verbatim_quote=fa.verbatim_quote,
                        source_path=fa.source_path,
                        page_number=fa.page_number,
                        involved_subcap_ids=list(fa.involved_subcap_ids or []),
                    )
                )
            # 2026-06-06 F5b-1: even when the handoff JSON pre-populated
            # `firm` (Nicola/WSFS/Calprivate paths), the leadership table
            # from the DOCX is still strictly additive — handoff JSONs
            # rarely carry executive lists. Merge here, before the
            # `firm is None` fallback, so Nicola's 14 leadership entries
            # are no longer dropped on the floor.
            if firm is not None and not firm.leadership and cp_result.leadership:
                from app.schemas.package import LeadershipPerson
                firm.leadership = [
                    LeadershipPerson(
                        name=p.name,
                        title=p.role or "",
                        last_verified_date=None,
                    )
                    for p in cp_result.leadership
                    if _is_person_name(p.name)
                ]
                warnings.append(
                    f"firmographics_leadership_merged_from_docx: "
                    f"source={cp_docx.name} count={len(firm.leadership)}"
                )
            # 2026-06-07 F5c: narrative_md is strictly additive in the
            # same way — handoff JSONs don't carry the analyst's prose
            # paragraph from the Client Profile DOCX's "Entity Profile"
            # section. Merge here so the D5 Context "About" panel reads
            # it regardless of whether the package shipped a handoff.
            if (
                firm is not None
                and not getattr(firm, "narrative_md", None)
                and cp_result.firmographics_narrative_md
            ):
                firm.narrative_md = cp_result.firmographics_narrative_md

            # Firmographics fallback ONLY when handoff JSON was absent;
            # otherwise keep the handoff value.
            if firm is None:
                leadership_blob = [
                    LeadershipPerson(
                        name=p.name,
                        title=p.role or "",
                        last_verified_date=None,
                    )
                    for p in (cp_result.leadership or [])
                    if _is_person_name(p.name)
                ]
                # 2026-06-06 Batch 4.2: mine the firmographics narrative
                # for the structured fields the Overview FirmographicsRows
                # React component reads. Until this, the Firmographics
                # object was created with every field None and the React
                # row layout rendered "—" universally even though the
                # source DOCX had the data.
                from app.services.parsers.client_profile import (
                    _extract_firmographics_facts,
                )
                facts = _extract_firmographics_facts(
                    cp_result.firmographics_narrative_md or "",
                    # strict: a profile narrative can cite an acquired
                    # bank's assets alongside the entity's own (FNBO) —
                    # disagreeing amounts stay empty for Gemini to fill.
                    strict=True,
                )
                # `branches` is not in the Firmographics schema today (it
                # was prototype-only), but Firmographics has `extra="allow"`
                # so we pass it through as kwargs. Pydantic v2 stores it
                # in __pydantic_extra__ and model_dump() emits it, which
                # is what the JSONB persistence + React
                # `(firm as { branches?: ... })?.branches` read both rely
                # on. (Note model_copy(update=...) DOESN'T update extras
                # in Pydantic v2 -- must be passed at construction.)
                firm_kwargs: dict[str, object] = {
                    "legal_name": (run_mf.institution_name if run_mf else None) or None,
                    "ticker": None,
                    "hq": facts.get("hq"),
                    "founded": None,
                    "total_assets": facts.get("total_assets"),
                    "employees_approx": facts.get("employees_approx"),
                    "primary_regulator": facts.get("primary_regulator"),
                    "cra_rating": None,
                    "leadership": leadership_blob,
                    # F5c: thread the analyst-prose narrative through.
                    "narrative_md": cp_result.firmographics_narrative_md or None,
                }
                if facts.get("branches"):
                    firm_kwargs["branches"] = facts["branches"]
                firm = Firmographics(**firm_kwargs)
                warnings.append(
                    f"firmographics_from_client_profile_docx: source={cp_docx.name} "
                    f"leadership={len(leadership_blob)} focus_areas={len(cp_result.focus_areas)} "
                    f"facts_extracted={len(facts)}"
                )

    # ── Standalone financial_baseline.json → firmographics merge ───────
    # 31 packages ship a flat `financial_baseline.json` (01_evidence /
    # 02_research_workbook / 08_appendices) — the only structured source
    # for D1 firmographics + D5 financials on those packages, previously
    # read by nothing. Merge ADDITIVELY: fill only fields the richer
    # sources above left empty, and create `firm` from it when no other
    # firmographics source existed at all.
    fb_path = next(
        (
            p
            for sub in (
                "00_entity_profile", "01_evidence",
                "02_research_workbook", "08_appendices",
            )
            for p in [root_p / sub / "financial_baseline.json"]
            if p.exists()
        ),
        None,
    )
    if fb_path is not None:
        from app.schemas.package import Firmographics
        from app.services.parsers.entity_profile import (
            parse_financial_baseline_json,
        )
        fb_fields = parse_financial_baseline_json(fb_path)
        if fb_fields:
            if firm is None:
                firm = Firmographics(**fb_fields)
                warnings.append(
                    f"firmographics_from_financial_baseline_json: "
                    f"fields={sorted(fb_fields.keys())}"
                )
            else:
                filled = []
                for key, value in fb_fields.items():
                    if not getattr(firm, key, None):
                        setattr(firm, key, value)  # extra='allow'
                        filled.append(key)
                if filled:
                    warnings.append(
                        f"financial_baseline_json_filled: {sorted(filled)}"
                    )

    # ── Client Profile "Risk & Issues" → issue-register merge ───────────
    # 2026-07-06: the research report's Issue Register table / Risk
    # prose is the SECOND grounded issue source (69/80 corpus reports
    # carry the section). CSV appendix rows keep priority; DOCX rows
    # merge by issue_id (enriching subcap/cap attribution the CSV
    # lacked) or append when no near-duplicate CSV row exists. The
    # "Trigger → Capabilities Affected → Maximum Score" table then
    # pushes cap attribution down to SUBCAP grain on matching rows.
    if cp_issue_rows or cp_issue_triggers:
        try:
            issues = _merge_profile_issue_rows(
                issues, cp_issue_rows, cp_issue_triggers, warnings,
            )
        except Exception as e:
            warnings.append(f"profile_issue_merge_failed:{e!s}")

    # ── Assessment-report DOCX → document_sections rows ────────────────
    # The DOCX is the analyst's PROSE. Without it, the surfaces show
    # skeleton-only data; with it, D1 SCQA / D2 IC explanations / D4
    # REC modals / D5 trend overlays / D6 data-gap copy all come from
    # the report. We tolerate its absence (`no_docx_found` state) — the
    # `narrative` subfield on each endpoint returns `null` and the
    # frontend keeps showing the skeleton.
    # 2026-05-28 H7 hotfix: parse ALL assessment-report DOCX
    # candidates, not just the first. Many Drive packages contain
    # both an initial assessment + a refresh, or an assessment +
    # an addendum, and dropping the others lost real content.
    # Per-section `source_path` preserves provenance so downstream
    # synthesis/citation can attribute correctly.
    report_sections: list[ReportSectionRow] = []
    docx_candidates = find_assessment_reports(root_p)
    if docx_candidates:
        # Dedup is scoped ACROSS DOCXs only, not within a single DOCX.
        # The H7 hotfix dedup `{kind}::{heading}` was originally added
        # so that two reports (assessment + addendum) wouldn't double-
        # count their shared "Executive Summary" heading. But within a
        # single DOCX, multiple sections sharing the same (kind, heading)
        # are INTENTIONAL — e.g., WSFS's §9 has 5 cycles of
        # `[ROOT CAUSE]` / `[SOLUTION]` / `[EXPECTED OUTCOMES]` heading-3
        # blocks, one per rec; collapsing them into one lost 4/5 of the
        # rec body content that the F4 DOCX extractor needs.
        # Now we track only keys from PRIOR DOCXs; current-DOCX duplicates
        # pass through.
        seen_kinds_across_docxs: set[str] = set()
        for primary in docx_candidates:
            try:
                parsed_report = parse_assessment_report(primary)
            except Exception as e:
                warnings.append(
                    f"assessment_report DOCX parse error "
                    f"({primary.name}): {e!s}"
                )
                continue
            if parsed_report is None:
                continue
            current_docx_keys: set[str] = set()
            for s in parsed_report.sections:
                key = f"{s.kind}::{(s.heading or '').strip().lower()}"
                # Skip if a PRIOR DOCX already contributed this
                # (kind, heading). Within the current DOCX, we accept
                # duplicates — preserves WSFS-shape `[ROOT CAUSE]` x5.
                if key in seen_kinds_across_docxs:
                    continue
                current_docx_keys.add(key)
                report_sections.append(ReportSectionRow(
                    kind=s.kind,
                    heading=s.heading,
                    body=s.body,
                    ordinal=s.ordinal,
                    page_number=s.page_number,
                    subcap_ids_mentioned=s.subcap_ids_mentioned,
                    e_ids_mentioned=s.e_ids_mentioned,
                    source_path=parsed_report.source_path,
                ))
            # After processing this DOCX, fold its keys into the
            # across-DOCX dedup set so the NEXT DOCX skips them.
            seen_kinds_across_docxs |= current_docx_keys
        # Re-bind `parsed_report` for the downstream
        # `parser_warnings += report state` block below so it sees
        # the LAST successfully-parsed report's state. Pre-fix code
        # only had one report; this preserves behavior when there is
        # one, and reflects the most recent state when there are many.
        parsed_report = None
        for primary in reversed(docx_candidates):
            try:
                pr = parse_assessment_report(primary)
            except Exception:
                continue
            if pr is not None:
                parsed_report = pr
                break
        # Only emit DEGRADATION warnings (not "look, I parsed it
        # successfully" notes) so the parser_warnings list stays
        # accurate. State + coverage are observable from the
        # report_sections themselves.
        if parsed_report is not None and parsed_report.state_kind in (
            "partial_coverage", "llm_fallback_used"
        ):
            log.info(
                    "assessment_report.partial",
                    state=parsed_report.state_kind,
                    sections=len(parsed_report.sections),
                    coverage=parsed_report.coverage_ratio,
                )
    # When no DOCX is found we do NOT emit a parser_warnings row — the
    # narrative subfields will return None and the UI surfaces skeleton
    # fallback. The state-transition `no_docx_found` is observable from
    # the empty `report_sections` list. Tests rely on parser_warnings
    # being clean when only the DOCX is absent.

    # ── report_synthesis.md → D1 SCQA card ─────────────────────────────
    # 41 packages ship a complete 4-part SCQA in report_synthesis.md
    # ("Generated from report_analysis.json — EVERY claim cites E-IDs") that
    # was previously unread. The bot's report_synthesis.md is the canonical
    # SCQA and is consistently FAR richer than the DOCX exec-summary the
    # classifier produces (corpus survey: in 24/24 packages shipping both,
    # the md body is larger — often the DOCX section is a <100-char stub).
    # Feed it through the SAME channel as the DOCX exec-summary — a
    # ReportSectionRow(kind="executive_summary_scqa") → document_sections →
    # section_routing.build_narrative_overview → D1 SCQA card — PREFERRING
    # it whenever its body is longer than any DOCX exec-summary (replacing
    # the thinner one in place). No new endpoint/FE wiring.
    synth_path = find_report_synthesis(root_p)
    synth = parse_report_synthesis_md(synth_path) if synth_path else None
    if synth is not None:
        existing_scqa = [
            rs for rs in report_sections if rs.kind == "executive_summary_scqa"
        ]
        existing_len = max((len(rs.body or "") for rs in existing_scqa), default=0)
        if len(synth.body) > existing_len:
            keep_ord = min((rs.ordinal for rs in existing_scqa), default=None)
            report_sections = [
                rs for rs in report_sections
                if rs.kind != "executive_summary_scqa"
            ]
            ordinal_v = (
                keep_ord if keep_ord is not None
                else max((rs.ordinal for rs in report_sections), default=-1) + 1
            )
            report_sections.append(ReportSectionRow(
                kind="executive_summary_scqa",
                heading=synth.heading,
                body=synth.body,
                ordinal=ordinal_v,
                subcap_ids_mentioned=synth.subcap_ids,
                e_ids_mentioned=synth.e_ids,
                source_path=str(synth_path.relative_to(root_p)),
            ))
            warnings.append(
                f"scqa_from_report_synthesis_md: {synth_path.name} "
                f"e_ids={len(synth.e_ids)} subcaps={len(synth.subcap_ids)} "
                f"replaced_docx_scqa={bool(existing_scqa)}"
            )

    # ── F4 DOCX fallback for recommendations (2026-06-07) ──────────────
    # When neither the canonical `08_appendices/recommendations_detail.json`
    # nor the variant `07_governance/recommendations_register.json`
    # populated `recs`, mine the Assessment_Report.docx §9
    # "Recommendations" section instead. Shapes handled:
    #   - Alma : `REC-NN: <title>` heading-2 per rec (canonical)
    #   - Nicola: `REC-NNN: <title> [ZENNIFY]` heading-2 + sub-blocks
    #   - WSFS / Calprivate / Odlum: `REC-NNN` / `R-NNN` IDs inline
    #     in the parent §9 body text + `[ROOT CAUSE]` / `[SOLUTION]`
    #     / `[EXPECTED OUTCOMES]` sub-block markers.
    if not recs and report_sections:
        try:
            from app.services.parsers.report_recommendations import (
                extract_recommendations_from_report_sections,
            )
            docx_recs = extract_recommendations_from_report_sections(
                report_sections
            )
            if docx_recs:
                recs = docx_recs
                warnings.append(
                    f"used docx-extracted recommendations: "
                    f"{len(docx_recs)} rec(s) from Assessment_Report.docx §9"
                )
        except Exception as e:
            warnings.append(
                f"docx_recommendations_extract_failed: {e!s}"
            )

    if not recs and not rec_source_found:
        # True source-side gap — package shipped no rec JSON AND the
        # DOCX §9 section either is missing or yielded no parseable
        # rec IDs. End-user impact: AE-facing recs panel renders its
        # empty state; admin import-audit flags the gap.
        warnings.append(
            "no_recommendations_source: package ships no "
            "recommendations_detail.json or recommendations_register.json "
            "AND the Assessment_Report.docx §9 had no extractable rec "
            "IDs (see docs/qa/qa_5folder_live_findings.md §F4)"
        )

    # ── Validation: subcap coverage ─────────────────────────────────────
    if run_mf and subcaps:
        expected = _expected_subcap_count(run_mf.subvertical_code)
        coverage = len(subcaps) / expected if expected > 0 else 0.0
        if coverage < 0.80:
            warnings.append(
                f"subcap coverage {coverage:.0%} below 80% threshold "
                f"({len(subcaps)} of {expected} expected for {run_mf.subvertical_code or 'unknown'})"
            )

    # ── Category rollup from subcaps (when no category scoring shipped) ─
    # ≈half the corpus (Greenstone / Haventree / Chemung / …) ships only
    # subcap-level scores — no export_category_summary.csv — leaving
    # `categories` empty, so D1 category bars, top_findings, and the
    # per-category peer overlay all render nothing. Derive the category
    # rollup from the entity's OWN subcap scores (DERIVED tier, mean per
    # P#C# group). Only when `categories` is empty — never override real
    # category-level scoring.
    if not categories and subcaps:
        _cat_groups: dict[str, list[float]] = {}
        for s in subcaps:
            m = re.match(r"(P\d+C\d+)", s.subcap_id or "")
            if m and s.score is not None:
                _cat_groups.setdefault(m.group(1), []).append(s.score)
        for cat_id in sorted(_cat_groups):
            vals = _cat_groups[cat_id]
            pm = re.match(r"(P\d+)", cat_id)
            categories.append(CategoryScoreRow(
                category_id=cat_id,
                pillar_id=pm.group(1) if pm else cat_id[:2],
                score=round(sum(vals) / len(vals), 2),
            ))
        if categories:
            warnings.append(
                f"category_scores_derived_from_subcaps: {len(categories)} "
                "categories (mean rollup)"
            )

    # Category-level workbook fallback: some packages (MidFirst) ship NO
    # subcap rows and NO export_category_summary — only a scoring workbook
    # with per-pillar P#_Scoring_Detail sheets at CATEGORY granularity.
    # Mine those directly so D1 bars + the per-category peer overlay light up.
    if not categories:
        from app.services.parsers.scoring_workbook import (
            category_scores_from_workbook,
        )
        for row in category_scores_from_workbook(root_p):
            categories.append(CategoryScoreRow(**row))
        if categories:
            warnings.append(
                f"category_scores_from_workbook: {len(categories)} categories"
            )

    # ── Peer benchmarks → fill category peer median/p25/p75 ────────────
    # Many packages (e.g. Greenstone) ship no `peer_median` column in
    # export_category_summary.csv but DO carry the peer cohort in
    # 06_peers/{peer_benchmarks.json,peer_comparison_table.csv}. Fill the
    # category rows from there where the workbook/CSV path left them empty
    # (additive — never override an existing value). Feeds D1 ticks + D3
    # overlay + the peer_benchmarks persist.
    peer_benchmarks = load_peer_benchmarks(root_p)
    if peer_benchmarks:
        filled = 0
        for cs in categories:
            # category-keyed benchmark, else a pillar-keyed one broadcast to
            # this category (transposed pillar-level peer table — DERIVED).
            b = (peer_benchmarks.get((cs.category_id or "").upper())
                 or peer_benchmarks.get((cs.pillar_id or "").upper()))
            if b is None:
                continue
            if cs.peer_median is None and b.median is not None:
                cs.peer_median = b.median
                filled += 1
            if cs.peer_p25 is None and b.p25 is not None:
                cs.peer_p25 = b.p25
            if cs.peer_p75 is None and b.p75 is not None:
                cs.peer_p75 = b.p75
        if filled:
            warnings.append(f"peer_benchmarks_filled: {filled} categories")

    # Tech-stack last resort (operator note: tech is also named in the
    # Client-Profile / Assessment report prose). When no structured tech
    # artifact shipped, scan the report section bodies for a curated
    # KNOWN-vendor dictionary (precision-first). Tagged source so the UI /
    # provenance can distinguish a prose mention from a confirmed inventory.
    if not tech and report_sections:
        prose = "\n".join(s.body or "" for s in report_sections)
        tech = extract_tech_from_text(prose)
        if tech:
            warnings.append(
                f"tech_stack_from_report_prose: {len(tech)} vendors mentioned"
            )

    # ── D5 financials + sentiment ──────────────────────────────────────
    # Both surfaces were dead corpus-wide (persist never wrote the columns
    # AND the sources were unread). Financial trajectory: the multi-year
    # A#_financial_trends.csv is preferred; the client research report's
    # flat financial highlights are the fallback. Sentiment grid:
    # A#_sentiment_data.csv. Persist writes them (COALESCE-guarded vs Clay).
    # Financial trajectory source priority: multi-year trends CSV → the
    # client research report's highlights → financial_baseline.json
    # (FDIC/UBPR flat metrics). Sentiment: A#_sentiment_data.csv → the
    # entity_profile sentiment block (Glassdoor / Indeed / app / press).
    csv_fh = load_financial_trends(root_p)
    fh_final = csv_fh or (cp_financials if cp_financials else {}) \
        or load_financial_baseline(root_p)
    sent = (
        load_sentiment(root_p)
        or sentiment_from_entity_profile(root_p)
        or cp_sentiment
    )
    if fh_final or sent:
        if firm is None:
            # ~half the corpus ships no firmographics source but DOES ship
            # the financial/sentiment artifacts — create a minimal row so
            # they aren't dropped (mirrors the financial_baseline path).
            from app.schemas.package import Firmographics
            firm = Firmographics()
        if fh_final and not firm.financial_highlights:
            firm.financial_highlights = fh_final
            warnings.append(
                "financial_highlights_from_trends_csv" if csv_fh
                else "financial_highlights_from_client_report"
            )
        if sent and not firm.sentiment:
            firm.sentiment = sent
            warnings.append(
                f"sentiment_from_csv: {len(sent.get('sources', []))} sources"
            )

    # Firmographics floor: every assessed entity HAS a legal name — the
    # run_manifest institution. When no structured firmographics source
    # parsed a name (handoff schema drift etc.), set it from the canonical
    # institution_name so D1 always shows the entity, then enrich from the
    # client research report below. Faithful (the name is known), universal.
    inst_name = (run_mf.institution_name or "").strip() if run_mf else ""
    if inst_name:
        if firm is None:
            from app.schemas.package import Firmographics
            firm = Firmographics(legal_name=inst_name)
        elif not firm.legal_name:
            firm.legal_name = inst_name

    # ── Report-mining backstop (2026-06-10, Clay NOT in prod) ──────────
    # The operator mandate: D1 firmographics must be drawn from the
    # CLIENT RESEARCH / CLIENT PROFILE REPORTS — Clay enrichment is not
    # available in this version. The structured extraction above only
    # mined the Client Profile's firmographics SECTION, and only when no
    # handoff JSON existed — so assets/employees/branches that the
    # analyst wrote in the exec summary, financial highlights, or the
    # Assessment Report body never reached the panel (13/95 assets
    # coverage). Mine the FULL report prose additively: fill ONLY fields
    # every richer source above left empty. Deterministic regex over the
    # analyst's own words — nothing fabricated.
    if firm is not None:
        from app.services.parsers.client_profile import (
            _extract_firmographics_facts,
        )
        _corpus_parts: list[str] = []
        if getattr(firm, "narrative_md", None):
            _corpus_parts.append(str(firm.narrative_md))
        _fh = getattr(firm, "financial_highlights", None) or {}
        if isinstance(_fh, dict):
            _corpus_parts.extend(
                str(v) for v in _fh.get("lines", []) if v
            )
            _corpus_parts.extend(
                f"{k}: {v}" for k, v in _fh.items()
                if k != "lines" and isinstance(v, str)
            )
        # PRECISION CONTRACT: mine ONLY texts authored ABOUT the
        # assessed entity (the Client Profile firmographics narrative +
        # the financial-highlights lines). Free report prose is OFF the
        # table for numerics — FNBO's exec summary mentions an ACQUIRED
        # bank's "$2.2B assets" before its own $35B, and peer figures
        # litter the pillar deep-dives. Fields these trusted texts
        # don't carry stay empty here and are filled by the Gemini
        # firmographics_extraction enrichment surface (grounded +
        # quota-gated) on the live deployment — Clay is NOT in prod
        # for this version.
        _mined = _extract_firmographics_facts(
            "\n".join(_corpus_parts), strict=True,
        )
        if _mined:
            _dump = firm.model_dump()
            _fill = {
                k: v for k, v in _mined.items()
                if not (_dump.get(k) or "")
            }
            if _fill:
                from app.schemas.package import Firmographics
                firm = Firmographics(**{**_dump, **_fill})
                warnings.append(
                    "firmographics_mined_from_report_prose:"
                    + ",".join(sorted(_fill))
                )

    # DERIVED SCQA fallback (plan ladder): when no report_synthesis.md / DOCX
    # exec-summary produced an SCQA, synthesise a 4-part SCQA from the
    # entity's OWN category scores + recommendations (deterministic, no
    # fabrication). Feeds the D1 SCQA card via the same channel.
    if categories and not any(
        s.kind == "executive_summary_scqa" for s in report_sections
    ):
        derived_scqa = build_derived_scqa(
            run_mf.institution_name if run_mf else None, categories, recs)
        if derived_scqa:
            ordv = max((rs.ordinal for rs in report_sections), default=-1) + 1
            report_sections.append(ReportSectionRow(
                kind="executive_summary_scqa",
                heading="Executive Summary",
                body=derived_scqa,
                ordinal=ordv,
                source_path="derived://scqa-from-scores",
            ))
            warnings.append("scqa_derived_from_scores")

    # D2 Insights: structured top_findings from section_analysis_#.json →
    # insight_cards (the table/endpoint/InsightCard grid exist but were
    # ── Recommendation prerequisites (D4 DependencyMap) ────────────────────
    # The bot's recommendation_validation.json carries a free-text
    # `prerequisite` clause on the recs that must follow others (e.g.
    # Greenstone R8 <- R2 + R5). Parse it, intersect with the package's known
    # rec_ids, and attach to each RecommendationRow so the persist layer can
    # write recommendations.prerequisite_rec_ids. Discovery mirrors the recs
    # file: canonical dirs + recursive fallback (the dir varies by package).
    if recs:
        _val_dirs = ("02_research_workbook", "01_evidence",
                     "07_governance", "08_appendices")
        val_candidates: list[Path] = [
            root_p / d / "recommendation_validation.json" for d in _val_dirs
        ]
        for vp in sorted(root_p.glob("**/recommendation_validation.json")):
            if vp not in val_candidates:
                val_candidates.append(vp)
        known_rec_ids = {_rec_id(getattr(r, "id", "")) for r in recs}
        known_rec_ids.discard("")
        for val_json in val_candidates:
            if not val_json.exists():
                continue
            prereq_map = _maybe(
                lambda txt: parse_rec_prerequisites(txt, known_rec_ids),
                val_json, warnings, "recommendation_validation.json",
            ) or {}
            if prereq_map:
                for rec in recs:
                    rec.prerequisite_rec_ids = prereq_map.get(
                        _rec_id(getattr(rec, "id", "")), [])
                edges = sum(len(v) for v in prereq_map.values())
                warnings.append(
                    f"rec_prerequisites_parsed: {edges} edges across "
                    f"{len(prereq_map)} recs")
            break

    # never populated). Part 5.1 ladder (2026-07-02) — rung priority:
    #   1. client_profile_findings (Client Profile Research Report; the
    #      82/113-package source the audit measured feeding ZERO cards)
    #   2. section_analysis top_findings (6/113)
    #   3. recommendations-derived (46/113 — they carry real facts)
    #   4. category-gaps — LAST RESORT, capped 4 (audit: 74.6% of all
    #      cards were category-gap restatements)
    # Rungs 1-3 COEXIST (near-duplicates deduped, higher rung wins);
    # the shared builders live in parsers/section_analysis.py so this
    # ladder and scripts/derive_insights.py (DB re-derive) cannot drift.
    _sub_scores_by_id = {
        s.subcap_id: s.score for s in subcaps if s.score is not None
    }
    profile_cards = (
        insights_from_profile_findings(
            cp_profile_findings, sub_scores=_sub_scores_by_id,
        )
        if cp_profile_findings else []
    )
    if profile_cards:
        warnings.append(
            f"insights_from_client_profile_findings: {len(profile_cards)}"
        )
    section_cards = parse_section_analyses(root_p)
    rec_cards = insights_from_recommendations(recs) if recs else []
    if rec_cards and not (profile_cards or section_cards):
        warnings.append(
            f"insights_derived_from_recommendations: {len(rec_cards)}"
        )
    insight_cards_data = combine_insight_rungs(
        profile_cards, section_cards, rec_cards,
    )
    # Universal DERIVED last resort: the entity's own below-peer-median /
    # low-maturity categories become insight cards (real scores, no
    # fabrication) — so every scored package surfaces D2 insights.
    if not insight_cards_data and categories:
        insight_cards_data = insights_from_category_gaps(categories)
        if insight_cards_data:
            warnings.append(
                f"insights_derived_from_category_gaps: {len(insight_cards_data)}"
            )

    # Ground every derived card on the package's own evidence — the
    # SHARED evidence ladder (Part 5.1: inline citations ride the cards
    # already; subcap-linked → category roll-up → lexical-similarity
    # attach; the persisted `basis` chip is added by the derive chain
    # for whatever remains). Same rungs as scripts/derive_insights so
    # ingest-created and re-derived cards ground identically.
    if insight_cards_data and evidence:
        from app.services.parsers.section_analysis import (
            attach_evidence_ladder,
            similarity_attach_evidence,
        )
        # Report-cited E-IDs the package's evidence_index doesn't carry
        # would render dead chips — drop them; the ladder re-grounds.
        _valid_eids = {e.e_id for e in evidence}
        for _card in insight_cards_data:
            _card.linked_e_ids = [
                e for e in (_card.linked_e_ids or []) if e in _valid_eids
            ]
        ev_by_subcap: dict[str, list[str]] = {}
        for ev_row in evidence:
            for _sid in (getattr(ev_row, "subcap_mappings", None) or []):
                ev_by_subcap.setdefault(_sid, []).append(ev_row.e_id)
        attach_evidence_ladder(insight_cards_data, ev_by_subcap)
        try:
            similarity_attach_evidence(
                insight_cards_data,
                [{"e_id": e.e_id, "excerpt": e.excerpt,
                  "source_name": e.source_name} for e in evidence],
            )
        except Exception as e:  # sklearn edge cases must never fail a parse
            warn(
                warnings, "insight_similarity_attach_failed",
                SEVERITY_DEGRADED, f"{type(e).__name__}: {str(e)[:160]}",
            )

    # D5 Context timeline: derive dated, classifiable events from the
    # evidence facts[]. Pure/deterministic (DERIVED tier) — empty when no
    # dated evidence carries event-shaped facts.
    timeline_events = extract_timeline_events(evidence)
    # Fallback (plan D5 "mine the reports"): when no dated evidence facts
    # produced timeline events, surface the Client Profile's own "Digital
    # Evolution Timeline" table so D5 lights up for report-only packages.
    if not timeline_events and cp_timeline_candidates:
        timeline_events = cp_timeline_candidates
        warnings.append(
            f"timeline_from_client_profile_docx: {len(timeline_events)} events"
        )
    # Acquisition-history events are ALWAYS merged (deduped) — the D5
    # acquisitions list reads timeline_events kind='acquisition', and the
    # report's dedicated Acquisition History table is the authoritative source.
    if cp_acquisitions:
        existing = {
            (e.kind, e.event_date, (e.title or "")[:60].lower())
            for e in timeline_events
        }
        added = 0
        for ev in cp_acquisitions:
            key = (ev.kind, ev.event_date, (ev.title or "")[:60].lower())
            if key not in existing:
                timeline_events.append(ev)
                existing.add(key)
                added += 1
        if added:
            timeline_events.sort(key=lambda e: e.event_date)
            warnings.append(
                f"acquisitions_from_client_profile_docx: {added} events"
            )

    # Universal leadership fallback (2026-06-23): when none of the structured
    # JSON / Client-Profile-DOCX paths populated a roster, mine ALL collected
    # source docs via the robust roster finder. This recovers the dict-of-roles
    # shapes (`leadership_snapshot` in entity_profile.json, `key_leadership` in
    # research_handoff.json) that the single-source parsers above miss when
    # `firm` was already built from a different artifact — the corpus census
    # showed ~60/113 packages reaching the panel empty before this.
    if firm is not None and not firm.leadership:
        try:
            from app.schemas.package import LeadershipPerson
            from app.services.entity_healing import extract_leadership
            healed = extract_leadership(root_p)
            healed_people = [
                LeadershipPerson(
                    name=p["name"],
                    title=(p.get("title") or ""),
                    last_verified_date=None,
                )
                for p in healed
                if _is_person_name(p.get("name"))
            ]
            if healed_people:
                firm.leadership = healed_people
                warnings.append(
                    f"firmographics_leadership_healed: count={len(healed_people)}"
                )
        except Exception as e:
            warnings.append(f"leadership_heal_failed: {type(e).__name__}: {e!s}"[:200])

    # Backfill the prototype's `l3_id` platform link on tech rows that came
    # from the Explorium XLSX / report-prose paths (the CSV/JSON readers set it
    # inline). Pure keyword resolution — no DB.
    from app.services.parsers.tech_linker import l3_for_tech as _l3_for_tech
    for _ts in tech:
        if _ts.l3_id is None:
            _ts.l3_id = _l3_for_tech(_ts.vendor, _ts.product, _ts.category)

    # ── Unconsumed-artifact knowledge mining + pattern-gap hook ─────────
    # (Part 12.6). AFTER every existing parser so the consumed-artifact
    # set is final: the miner walks MATERIAL artifacts nothing above
    # consumed, runs the registered fingerprint parsers (zennify
    # opportunities / uncertainty register / org capability), falls to
    # the generic section-miner for the rest, and records a
    # nlp.patterns pattern_gap per unmatched shape. Failures are
    # DEGRADED warnings — mining never fails a parse.
    mined_knowledge = None
    try:
        from app.services.parsers.knowledge_artifacts import (
            mine_package_knowledge,
        )
        mined_knowledge = mine_package_knowledge(root_p, warnings)
    except Exception as e:
        warn(
            warnings, "knowledge_mining_failed", SEVERITY_DEGRADED,
            f"{type(e).__name__}: {str(e)[:200]}",
        )

    # D2 Part 5.1: generated OPPORTUNITY cards from the analyst's
    # A#_zennify_opportunities.csv rows the miner just extracted — each
    # fully evidenced via its trigger_evidence E-IDs or not emitted.
    # Appended AFTER the ladder (never displaces report-derived cards);
    # combine_insight_rungs dedups near-identical opportunities.
    if mined_knowledge is not None and mined_knowledge.sections:
        try:
            _ev_excerpts = {
                e.e_id: (getattr(e, "excerpt", "") or "",
                         getattr(e, "source_name", "") or "")
                for e in evidence
            }
            opp_rows = [
                s.get("provenance") or {}
                for s in mined_knowledge.sections
                if s.get("artifact_kind") == "zennify_opportunity"
            ]
            opp_cards = insights_from_zennify_opportunities(
                opp_rows,
                sub_scores=_sub_scores_by_id,
                evidence_excerpts=_ev_excerpts,
            ) if opp_rows else []
            if opp_cards:
                insight_cards_data = combine_insight_rungs(
                    insight_cards_data, opp_cards,
                )
                warnings.append(
                    f"insights_from_zennify_opportunities: {len(opp_cards)}"
                )
        except Exception as e:
            warn(
                warnings, "zennify_opportunity_cards_failed",
                SEVERITY_DEGRADED, f"{type(e).__name__}: {str(e)[:200]}",
            )

    pkg = IngestedPackage(
        manifest=manifest,
        run_manifest=run_mf,
        pillar_weights=run_mf.pillar_weights,
        pillar_scores=pillars,
        category_scores=categories,
        subcap_scores=subcaps,
        evidence=evidence,
        timeline_events=timeline_events,
        insight_cards=insight_cards_data,
        issue_register=issues,
        recommendations=recs,
        peers=peers,
        tech_stack=tech,
        firmographics=firm,
        qa_verdict=qa,
        qa_verdict_l1=qa_l1,
        report_sections=report_sections,
        focus_areas=focus_areas_collected,
        caps_applied_log=caps_applied_log_data,
        assumptions_register=assumptions_data,
        audit_logs=audit_logs_data,
        parser_warnings=warnings,
        parser_observations=_observations_to_emit[:200],
    )
    # Side-channel: the knowledge sections ride the envelope WITHOUT a
    # schema change (IngestedPackage is extra='forbid' and app/schemas is
    # contract-frozen this commit series). `object.__setattr__` writes
    # into the instance __dict__ — invisible to model_dump()/validation;
    # the backfill reads it via getattr(pkg, "_mined_knowledge", None)
    # and persists client_knowledge_sections + runs.uncertainty_bands.
    if mined_knowledge is not None:
        object.__setattr__(pkg, "_mined_knowledge", mined_knowledge)
    log.info(
        "dma_package.parsed",
        run_id=run_mf.run_id,
        subcaps=len(subcaps),
        evidence=len(evidence),
        peers=len(peers),
        recs=len(recs),
        warnings=len(warnings),
    )
    return pkg


# Per-subvertical expected subcap totals (v7.0 baseline).
# Older catalogue versions vary slightly — the alias bridge handles the
# delta; here we use a conservative floor.
_SV_EXPECTED = {
    "RB": 700, "CU": 700, "CL": 708, "CIB": 720, "FC": 700,
    "AM": 720, "RIA": 700, "IC": 720, "IB": 700,
}


def _expected_subcap_count(subvertical_code: str | None) -> int:
    if not subvertical_code:
        return 700  # generic floor
    # SV1/SV2/SV3 legacy notation falls back to floor.
    if re.match(r"^SV\d", subvertical_code):
        return 700
    return _SV_EXPECTED.get(subvertical_code.upper(), 700)
