"""Per-tab parsers for the v7.0 pillar workbooks.

Each parser:
  - takes an openpyxl Worksheet and the pillar_id
  - emits a list of dicts ready for SQLAlchemy bulk insert into the target
    ccg_* table
  - normalizes the P-prefix (P1/P2/P3/P4) so per-pillar workbooks share the
    same loader pipeline
  - records non-fatal anomalies in `warnings` (returned alongside rows)

All parsers are pure (no DB, no IO beyond the sheet). The loader's `main.py`
orchestrates IO + transactions.
"""
from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from typing import Any

try:
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover - openpyxl is a hard dep at runtime
    Worksheet = Any  # type: ignore[assignment,misc]

SLUG_RE = re.compile(r"[^a-z0-9]+")
RE_SUBCAP_ID = re.compile(r"^P\d+C\d+")


def slug(value: str) -> str:
    """Slugify an L1 capability name to a stable id fragment."""
    return SLUG_RE.sub("-", value.lower()).strip("-")


def derive_l1_id(category_id: str, l1_name: str) -> str:
    return f"{category_id}::{slug(l1_name)}"


# The v7.0 workbooks carry NO category display-name column — `Category` IS
# the id (P1C1), so ccg_categories.name loaded blank and every grid label,
# drilled header and synthesis title fell back to the bare mono id (the
# all-94 prototype-parity capture flagged it on the standard heatmap).
# These names are derived from each category's REAL v7.0 L1 composition
# (e.g. P3C2's L1s are Fraud Detection/Investigation/Intelligence — NOT the
# prototype mock's "Loan Origination") so they are honest to the catalogue,
# not to the wireframe's demo data. Applied only when the workbook supplies
# no explicit name; an explicit column always wins.
CATEGORY_DISPLAY_NAMES: dict[str, str] = {
    "P1C1": "Digital Strategy",
    "P1C2": "Governance & Risk",
    "P1C3": "Innovation Operating Model",
    "P1C4": "Talent, Culture & Change",
    "P1C5": "ESG & Community",
    "P2C1": "Marketing & Demand",
    "P2C2": "Onboarding & Origination",
    "P2C3": "Service & Support",
    "P2C4": "Personalisation & Deepening",
    "P3C1": "Process Automation",
    "P3C2": "Fraud & Operational Risk",
    "P3C3": "Compliance Operations",
    "P3C4": "Resilience & Third-Party Risk",
    "P4C1": "Data Foundation",
    "P4C2": "Analytics & AI",
    "P4C3": "Architecture & Cloud",
    "P4C4": "Security & Trust",
}


def derive_category_name(category_id: str, explicit_name: str = "") -> str:
    """Explicit workbook name > curated composition-derived name > ''.

    Never returns the bare id — a blank stays blank so consumers keep
    their honest id-only fallback rather than a fake 'P9C9' name."""
    name = (explicit_name or "").strip()
    if name and name != category_id:
        return name
    return CATEGORY_DISPLAY_NAMES.get(category_id, "")


@dataclass
class ParseResult:
    rows: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[dict[str, Any]] = field(default_factory=list)


def _header_row(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])):
        if cell is None:
            continue
        headers[str(cell).strip()] = idx
    return headers


# The v7.0 workbooks prepend a 3-4 row title/description block to some tabs
# (21_VC_Mapping_PerSubcap). Scan the first rows for the row that carries the
# anchor column and treat THAT as the header row.
_HEADER_SCAN_ROWS = 10


def _find_header_row(
    ws: Worksheet, *, anchors: tuple[str, ...],
) -> tuple[dict[str, int], int]:
    """Locate the header row by anchor column names within the first
    ``_HEADER_SCAN_ROWS`` rows. Returns ``(headers, first_data_row)``;
    falls back to row 1 when no anchor is found (legacy layout)."""
    for row_no, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS, values_only=True),
        start=1,
    ):
        cells = {str(c).strip() for c in row if c is not None}
        if any(a in cells for a in anchors):
            headers = {
                str(c).strip(): idx
                for idx, c in enumerate(row)
                if c is not None and str(c).strip()
            }
            return headers, row_no + 1
    return _header_row(ws), 2


def _val(row: tuple[Any, ...], headers: dict[str, int], *names: str, default: Any = None) -> Any:
    """First-present-header lookup. Allows aliasing column names across pillars."""
    for n in names:
        if n in headers:
            v = row[headers[n]]
            if v is not None and (not isinstance(v, str) or v.strip() != ""):
                return v
    return default


# ---------- 1_Overview ----------

def parse_overview(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """Returns a single ccg_pillars row per pillar workbook."""
    res = ParseResult()
    headers = _header_row(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        res.rows.append(
            {
                "version": version,
                "pillar_id": pillar_id,
                "name": str(_val(row, headers, "Pillar Name", "Name", default=pillar_id)).strip(),
                "description": str(_val(row, headers, "Description", default="")).strip(),
                "category_count": int(_val(row, headers, "Category Count", "Categories", default=0) or 0),
                "l1_capability_count": int(_val(row, headers, "L1 Capability Count", "L1", default=0) or 0),
                "subcap_count": int(_val(row, headers, "SubCap Count", "Sub-Caps", default=0) or 0),
            }
        )
        break  # Overview tab is single-row per workbook
    return res


# ---------- 2_Capability_Map ----------

def parse_capability_map(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """Returns rows for ccg_categories, ccg_l1_capabilities, ccg_subcaps.

    Categories + L1 capabilities are de-duplicated by `(category_id, l1_id)`;
    subcaps are unique by `subcap_id`. The caller bulk-inserts each list into
    its own table.
    """
    res = ParseResult()
    headers = _header_row(ws)
    categories: dict[str, dict[str, Any]] = {}
    l1s: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        # The v7.0 workbooks use a single `Category` column whose value
        # IS the category_id (e.g. "P1C1"). Older schemas had separate
        # `Category ID` + `Category Name` columns; we accept both.
        category_id = str(_val(row, headers, "Category ID", "CategoryID",
                               "Category", default="")).strip()
        # Category name is only populated when an explicit name column
        # exists; otherwise leave blank (1_Overview / 2_Capability_Map
        # use the same column for ID).
        category_name = str(_val(row, headers, "Category Name", "Category_Name",
                                  default="")).strip()
        l1_name = str(_val(row, headers, "L1 Capability", "L1_Capability",
                            "L1", default="")).strip()
        # Prefer canonical L1_ID column if the workbook supplies it (resolved
        # decision 7 in the plan). Otherwise derive a stable slug.
        l1_id_canonical = _val(row, headers, "L1_ID", "L1 ID")
        l1_id = (
            str(l1_id_canonical).strip()
            if l1_id_canonical
            else derive_l1_id(category_id, l1_name)
        )
        subcap_id = str(_val(row, headers, "SubCap ID", "Sub_Cap_ID",
                              "Subcap ID", default="")).strip()
        if not (category_id and l1_id and subcap_id):
            res.warnings.append(
                {"kind": "missing_id_fields", "row": list(row)[:6]}
            )
            continue

        categories.setdefault(
            category_id,
            {
                "version": version,
                "category_id": category_id,
                "pillar_id": pillar_id,
                "name": derive_category_name(category_id, category_name),
            },
        )
        l1s.setdefault(
            l1_id,
            {
                "version": version,
                "l1_id": l1_id,
                "category_id": category_id,
                "name": l1_name,
            },
        )
        tier = str(_val(row, headers, "Tier", default="T1")).strip() or "T1"
        solution_type = str(_val(row, headers, "Solution Type", default="Traditional")).strip()
        if solution_type not in ("Traditional", "Hybrid", "Headless"):
            res.warnings.append(
                {"kind": "unknown_solution_type", "subcap_id": subcap_id, "value": solution_type}
            )
            solution_type = "Traditional"
        personas_cell = _val(row, headers, "Personas", "Persona")
        personas = (
            [p.strip() for p in str(personas_cell).split(";") if p.strip()]
            if personas_cell else None
        )
        l3_cell = _val(row, headers, "L3 Platforms", "L3_Platforms",
                       "L3_Platforms_Addressing_SubCap", "L3")
        l3_platforms = (
            [p.strip() for p in str(l3_cell).split(";") if p.strip()]
            if l3_cell else None
        )
        res.rows.append(
            {
                "version": version,
                "subcap_id": subcap_id,
                "l1_id": l1_id,
                "name": str(_val(row, headers, "SubCap Name", "Sub_Cap_Name", default="")).strip(),
                "description": str(_val(row, headers, "Description", default="")).strip(),
                "solution_type": solution_type,
                "tier": tier,
                "personas": personas,
                "l3_platforms": l3_platforms,
                "use_cases": _val(row, headers, "Use Cases", "UseCases"),
                "story_refs": _val(row, headers, "Story Refs", "StoryRefs"),
                "zennify_status": str(_val(row, headers, "Zennify Status", default="Active")).strip(),
            }
        )

    res.rows = [
        {"__target__": "ccg_categories", **r} for r in categories.values()
    ] + [
        {"__target__": "ccg_l1_capabilities", **r} for r in l1s.values()
    ] + [
        {"__target__": "ccg_subcaps", **r} for r in res.rows
    ]
    return res


# ---------- 3_Maturity_Scoring_Bands (v7.0 sheet: 6_Maturity_Descriptors) ----------

_BAND_COL_RE = re.compile(r"^M([1-5])[_ ]")


def _wide_band_columns(headers: dict[str, int]) -> dict[str, tuple[int, int | None]]:
    """v7.0 ships the descriptors WIDE — one column pair per band
    (``M1_Foundational`` + ``M1_Foundational_Features``, …). Returns
    ``{"M1": (narrative_col, features_col|None), ...}``: per band, the
    first M{n}_-prefixed column (by sheet order) is the narrative, the
    second the features list."""
    per_band: dict[str, list[int]] = {}
    for name, idx in headers.items():
        m = _BAND_COL_RE.match(name)
        if m:
            per_band.setdefault(f"M{m.group(1)}", []).append(idx)
    out: dict[str, tuple[int, int | None]] = {}
    for band, cols in per_band.items():
        cols.sort()
        out[band] = (cols[0], cols[1] if len(cols) > 1 else None)
    return out


def parse_maturity_bands(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """Handles BOTH shapes: the legacy long form (one row per subcap x band
    with a `Band` column) and the v7.0 wide form (`6_Maturity_Descriptors`:
    one row per subcap, M1..M5 narrative/features column pairs)."""
    res = ParseResult()
    headers, first_data_row = _find_header_row(
        ws, anchors=("Sub_Cap_ID", "SubCap ID"))
    wide = _wide_band_columns(headers)
    long_form = any(h in headers for h in ("Band", "Maturity Band"))
    for row in ws.iter_rows(min_row=first_data_row, values_only=True):
        if all(c is None for c in row):
            continue
        subcap_id = str(_val(row, headers, "SubCap ID", "Sub_Cap_ID", default="")).strip()
        if not subcap_id:
            continue
        if long_form:
            band = str(_val(row, headers, "Band", "Maturity Band", default="")).strip()
            if band not in ("M1", "M2", "M3", "M4", "M5"):
                res.warnings.append({"kind": "bad_band_row", "subcap": subcap_id, "band": band})
                continue
            res.rows.append(
                {
                    "version": version,
                    "subcap_id": subcap_id,
                    "band": band,
                    "narrative": str(_val(row, headers, "Narrative", "Description", default="")).strip(),
                    "features": str(_val(row, headers, "Features", "Capabilities", default="")).strip(),
                }
            )
            continue
        if not wide:
            res.warnings.append({"kind": "bad_band_row", "subcap": subcap_id, "band": None})
            continue
        for band in ("M1", "M2", "M3", "M4", "M5"):
            cols = wide.get(band)
            if cols is None:
                continue
            narrative_idx, features_idx = cols
            narrative = row[narrative_idx] if narrative_idx < len(row) else None
            features = (
                row[features_idx]
                if features_idx is not None and features_idx < len(row)
                else None
            )
            if narrative is None and features is None:
                continue
            res.rows.append(
                {
                    "version": version,
                    "subcap_id": subcap_id,
                    "band": band,
                    "narrative": str(narrative or "").strip(),
                    "features": str(features or "").strip(),
                }
            )
    return res


# ---------- 21_Value_Chain_Mapping (v7.0 sheet: 21_VC_Mapping_PerSubcap) ----------

VC_STAGE_SEP = re.compile(r"[│|▌\n]")  # ▌-prefixed / pipe- / newline-separated stages

# The v7.0 per-subcap VC sheet labels its subvertical columns with the FULL
# canonical names (matching ccg_subverticals.name), not the 2-3 letter codes.
SUBVERTICAL_NAME_TO_CODE: dict[str, str] = {
    "Retail Banking": "RB",
    "Credit Unions": "CU",
    "Commercial Lending": "CL",
    "Corp & Investment Banking": "CIB",
    "Farm Credit / Ag Lending": "FC",
    "Asset & Wealth Management": "AM",
    "RIA / Broker-Dealer": "RIA",
    "Insurance Carriers": "IC",
    "Insurance Brokerages": "IB",
}
_SUBVERTICAL_CODES = tuple(SUBVERTICAL_NAME_TO_CODE.values())


def parse_value_chain(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """Long-form mapping: 1 row per (subcap, subvertical) with `value_chain_stages`.

    Handles the shipped v7.0 layout — a 4-row title block, header row 5,
    subvertical columns labelled with FULL names ("Retail Banking" …) and
    ▌-prefixed multi-line stage cells — as well as the legacy code-labelled
    single-header layout."""
    res = ParseResult()
    headers, first_data_row = _find_header_row(
        ws, anchors=("Sub_Cap_ID", "SubCap ID"))
    subvertical_columns: list[tuple[str, int]] = []
    for h, idx in headers.items():
        if h in _SUBVERTICAL_CODES:
            subvertical_columns.append((h, idx))
        elif h in SUBVERTICAL_NAME_TO_CODE:
            subvertical_columns.append((SUBVERTICAL_NAME_TO_CODE[h], idx))
    if not subvertical_columns:
        res.warnings.append({"kind": "no_subvertical_columns",
                             "headers": sorted(headers)[:20]})
        return res
    for row in ws.iter_rows(min_row=first_data_row, values_only=True):
        if all(c is None for c in row):
            continue
        subcap_id = str(_val(row, headers, "SubCap ID", "Sub_Cap_ID", default="")).strip()
        if not subcap_id:
            continue
        for code, idx in subvertical_columns:
            cell = row[idx] if idx < len(row) else None
            if cell is None or (isinstance(cell, str) and not cell.strip()):
                continue
            stages = [s.strip() for s in VC_STAGE_SEP.split(str(cell)) if s.strip()]
            if not stages:
                continue
            res.rows.append(
                {
                    "version": version,
                    "subcap_id": subcap_id,
                    "subvertical_code": code,
                    "value_chain_stages": stages,
                }
            )
    return res


# ---------- _R1_Source_Reference (alias bridge) ----------

_R1_PRIOR_COL_RE = re.compile(r"^Sub_?Cap[ _]?ID \((v?[\d.]+) original\)$", re.IGNORECASE)
_R1_CURRENT_COL = "Sub_Cap_ID (post-rename)"


def parse_alias_bridge(ws: Worksheet, current_version: str, pillar_id: str) -> ParseResult:
    """Each row maps prior_subcap_id → current_subcap_id with an action.

    Two shapes: the legacy explicit-columns form (`Prior Version` /
    `Prior SubCap ID` / `Current SubCap ID` / `Action`) and the shipped
    v7.0 `_R1_Source_Reference` form, whose prior version is embedded in
    the column header (`Sub_Cap_ID (v5.0 original)` → prior_version
    "v5.0"); rows whose IDs differ are RENAMED, identical IDs MIGRATED."""
    res = ParseResult()
    headers = _header_row(ws)
    # ── v7.0 header-embedded shape ─────────────────────────────────────
    prior_col: int | None = None
    prior_version_v7: str | None = None
    for name, idx in headers.items():
        m = _R1_PRIOR_COL_RE.match(name)
        if m:
            prior_col = idx
            v = m.group(1)
            prior_version_v7 = v if v.startswith("v") else f"v{v}"
            break
    if prior_col is not None and _R1_CURRENT_COL in headers:
        current_col = headers[_R1_CURRENT_COL]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            prior_subcap = str(row[prior_col] or "").strip() if prior_col < len(row) else ""
            current_subcap = str(row[current_col] or "").strip() if current_col < len(row) else ""
            if not (prior_subcap and current_subcap):
                res.warnings.append({"kind": "incomplete_alias_row", "row": list(row)[:3]})
                continue
            res.rows.append(
                {
                    "prior_version": prior_version_v7,
                    "prior_subcap_id": prior_subcap,
                    "current_version": current_version,
                    "current_subcap_id": current_subcap,
                    "migration_action": (
                        "MIGRATED" if prior_subcap == current_subcap else "RENAMED"
                    ),
                    "migration_notes": None,
                }
            )
        return res
    # ── legacy explicit-columns shape ──────────────────────────────────
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        prior_version = str(_val(row, headers, "Prior Version", "From Version", default="")).strip()
        prior_subcap = str(_val(row, headers, "Prior SubCap ID", "From ID", default="")).strip()
        current_subcap = str(_val(row, headers, "Current SubCap ID", "To ID", default="")).strip()
        action = str(_val(row, headers, "Action", "Migration Action", default="MIGRATED")).strip().upper()
        if not (prior_version and prior_subcap and current_subcap):
            res.warnings.append({"kind": "incomplete_alias_row", "row": list(row)[:5]})
            continue
        if action not in ("MIGRATED", "RENAMED", "SPLIT", "MERGED", "DROPPED", "L1_ID_PROMOTED"):
            res.warnings.append({"kind": "unknown_alias_action", "action": action})
            action = "MIGRATED"
        res.rows.append(
            {
                "prior_version": prior_version,
                "prior_subcap_id": prior_subcap,
                "current_version": current_version,
                "current_subcap_id": current_subcap,
                "migration_action": (
                    "l1_id_promoted" if action == "L1_ID_PROMOTED" else action
                ),
                "migration_notes": str(_val(row, headers, "Notes", default="")).strip() or None,
            }
        )
    return res


# ---------- Tab dispatcher ----------

def _s(row, headers, *names, limit=2000):
    v = _val(row, headers, *names, default="")
    return (str(v).strip()[:limit] or None) if v is not None else None


def parse_l3_detailed(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """`4_L3_Detailed` — the L3 platform reference (vendor platforms the
    catalogue maps capabilities onto). One row per L3_ID; the sheet is
    identical across the four pillar workbooks, so the promote step
    dedupes on (version, l3_id)."""
    res = ParseResult()
    headers, first = _find_header_row(ws, anchors=("L3_ID",))
    for row in ws.iter_rows(min_row=first, values_only=True):
        if all(c is None for c in row):
            continue
        l3 = _s(row, headers, "L3_ID", limit=64)
        if not l3 or not l3.upper().startswith("L3"):
            continue
        res.rows.append({
            "version": version, "l3_id": l3,
            "vendor": _s(row, headers, "Vendor", limit=120),
            "platform_name": _s(row, headers, "Platform_Name", limit=200),
            "category": _s(row, headers, "Category", limit=120),
            "description": _s(row, headers, "Description"),
            "setup_path": _s(row, headers, "Setup_Path", limit=500),
            "prerequisites": _s(row, headers, "Prerequisites", limit=1000),
            "detailed_capabilities": _s(row, headers,
                                        "Detailed_Capabilities", limit=4000),
        })
    return res


def parse_l4_features(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """`5_L4_Detailed_Features` — subcap → L3 platform feature links, the
    trainable capability→platform affinity layer."""
    res = ParseResult()
    headers, first = _find_header_row(ws, anchors=("Sub_Cap_ID", "SubCap ID"))
    for row in ws.iter_rows(min_row=first, values_only=True):
        if all(c is None for c in row):
            continue
        sid = _s(row, headers, "Sub_Cap_ID", "SubCap ID", limit=32)
        feat = _s(row, headers, "Feature_Name", limit=300)
        if not sid or not feat or not RE_SUBCAP_ID.match(sid):
            continue
        res.rows.append({
            "version": version, "subcap_id": sid,
            "l3_id": _s(row, headers, "L3_Platform_ID", "L3_ID", limit=64),
            "feature_name": feat,
            "vendor": _s(row, headers, "Vendor", limit=120),
            "feature_type": _s(row, headers, "Feature_Type", limit=80),
            "customization_level": _s(row, headers,
                                      "Customization_Level", limit=80),
            "reference_url": _s(row, headers, "Reference_URL",
                                "Configuration_Path", limit=500),
        })
    return res


_CONFIDENCE_WORDS = {"high": 0.9, "medium": 0.6, "med": 0.6, "low": 0.3}


def parse_user_stories(ws: Worksheet, version: str, pillar_id: str) -> ParseResult:
    """`3_User_Stories_Catalogue` — use-case stories per subcap with their
    L4 features and use-case ids; the training corpus for
    capability→platform reasoning."""
    res = ParseResult()
    headers, first = _find_header_row(ws, anchors=("Story_Key",))
    for row in ws.iter_rows(min_row=first, values_only=True):
        if all(c is None for c in row):
            continue
        key = _s(row, headers, "Story_Key", limit=64)
        sid = _s(row, headers, "Sub_Cap_ID", "SubCap ID", limit=32)
        if not key or not sid:
            continue
        raw_conf = str(_val(row, headers, "Match_Confidence", default="") or "").strip()
        try:
            conf = float(raw_conf)
        except ValueError:
            conf = _CONFIDENCE_WORDS.get(raw_conf.lower())
        if conf is not None:
            if 1.0 < conf <= 100.0:  # percentage-styled sheet value
                conf = conf / 100.0
            conf = min(max(conf, 0.0), 1.0)
        res.rows.append({
            "version": version, "story_key": key, "subcap_id": sid,
            "source_type": _s(row, headers, "Source_Type", limit=32),
            "source_ref": _s(row, headers, "Source_Ref", limit=256),
            "use_case_ids": _s(row, headers, "Use_Case_IDs", limit=1000),
            "l4_features_used": _s(row, headers, "L4_Features_Used", limit=2000),
            "match_confidence": conf,
        })
    return res


PARSER_FOR_TAB: dict[str, Any] = {
    "1_Overview": parse_overview,
    "2_Capability_Map": parse_capability_map,
    "3_Maturity_Scoring_Bands": parse_maturity_bands,
    "4_L3_Platforms_Reference": parse_l3_detailed,
    "5_L4_Features": parse_l4_features,
    "6_User_Stories": parse_user_stories,
    "21_Value_Chain_Mapping": parse_value_chain,
    "_R1_Source_Reference": parse_alias_bridge,
}


def parse_workbook_tabs(
    workbook: Any,
    *,
    version: str,
    pillar_id: str,
    tab_names: Iterable[str],
) -> dict[str, ParseResult]:
    """Iterate the named tabs in a workbook; return per-tab ParseResults.

    Tab resolution goes through `canonical_tabs.resolve_sheet_name` so the
    canonical key matches the shipped v7.0 sheet names (15/25 drifted —
    see TAB_ALIASES). An alias hit is recorded as an informational
    ``aliased_tab`` warning so the loader run log shows the mapping."""
    from .canonical_tabs import resolve_sheet_name

    out: dict[str, ParseResult] = {}
    sheet_keys = {ws.title.strip(): ws for ws in workbook.worksheets}
    sheet_names = set(sheet_keys)
    for tab in tab_names:
        parser = PARSER_FOR_TAB.get(tab)
        if parser is None:
            continue  # tab is loaded by a future patch; skip cleanly
        actual = resolve_sheet_name(tab, sheet_names)
        if actual is None:
            out[tab] = ParseResult(warnings=[{"kind": "missing_tab", "tab": tab}])
            continue
        ws = sheet_keys[actual]
        try:
            out[tab] = parser(ws, version, pillar_id)
        except Exception as exc:
            out[tab] = ParseResult(warnings=[{"kind": "parser_exception",
                                              "tab": tab, "error": str(exc)}])
            continue
        if actual != tab:
            out[tab].warnings.append(
                {"kind": "aliased_tab", "tab": tab, "sheet": actual}
            )
    return out
