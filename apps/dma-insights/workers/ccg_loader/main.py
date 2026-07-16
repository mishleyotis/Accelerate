"""Catalogue loader entrypoint — Cloud Run Job.

Usage (local):
  python -m workers.ccg_loader.main \
    --version v7.0 \
    --workbooks-dir /path/to/v7-pillars/ \
    --dry-run

GCS-backed (production): `--workbooks-dir` accepts `gs://bucket/prefix/`
and the loader downloads every `*.xlsx` under that prefix to a tempdir
before parsing.

  python -m workers.ccg_loader.main \
    --version v7.0 \
    --workbooks-dir gs://${PROJECT_ID}-catalogue-staging/v7.0/

Production:
  Triggered by Cloud Scheduler hourly job polling
  gs://dma-insights-catalogue-staging/ for any new workbook set matching the
  filename prefix `Pillar_{1..4}_Comprehensive_Capability_Mapping_v{X}.xlsx`.

Lifecycle:
  1. Read 4 pillar workbooks (P1..P4) + Visualized Schema HTML.
  2. For each pillar workbook:
       - run the 25 canonical-tab parsers
       - assemble row sets keyed by target ccg_* table
  3. Merge per-pillar rows into a single rowset.
  4. Run validators (row counts, FK closure, subvertical canonicality).
  5. If validation passes: write to staging.ccg_* tables, emit diff vs prior
     frozen version, set ccg_loader_runs.status='AWAITING_APPROVAL'.
  6. Admin clicks Approve in /admin/catalogue → atomic copy from staging into
     canonical schema + UPDATE ccg_catalog_versions.frozen_at = NOW().

This module intentionally keeps IO at the edges (reading workbooks, writing
DB rows) and routes all parsing through pure functions in ./parsers.py so the
loader can be unit-tested with synthetic fixtures.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import pathlib
import sys
from dataclasses import dataclass, field
from typing import Any

from .canonical_tabs import expected_tab_keys, resolve_sheet_name
from .parsers import PARSER_FOR_TAB, ParseResult, parse_workbook_tabs
from .validators import (
    merge_reports,
    validate_fk_closure,
    validate_pillar_totals,
    validate_value_chain_subverticals,
)


@dataclass
class LoaderResult:
    version: str
    rows_by_table: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    warnings: list[dict[str, Any]] = field(default_factory=list)
    source_sha256s: dict[str, str] = field(default_factory=dict)
    validation_passed: bool = False
    validation_detail: dict[str, Any] = field(default_factory=dict)
    # Per-(pillar, canonical-tab) parsed row counts for every
    # parser-registered tab — feeds the zero-row gate + the run report.
    tab_row_counts: dict[str, dict[str, int]] = field(default_factory=dict)
    # "P{n}/{tab}" entries where a workbook CONTAINS a registered tab
    # (directly or via TAB_ALIASES) but parsing yielded 0 rows. A drifted
    # tab must never silently load nothing again (the value-chains-0/94
    # incident class) — main() exits 3 when this is non-empty.
    zero_row_gate_failures: list[str] = field(default_factory=list)


def _sha256(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def load_workbooks(
    *,
    version: str,
    workbooks_dir: pathlib.Path,
) -> LoaderResult:
    """Read the 4 pillar workbooks from a directory + run all parsers."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(
            "openpyxl required; install backend deps in workers env"
        ) from e

    result = LoaderResult(version=version)
    pillar_files: dict[str, pathlib.Path] = {}
    for pillar_no in (1, 2, 3, 4):
        candidates = list(
            workbooks_dir.glob(
                f"Pillar_{pillar_no}_Comprehensive_Capability_Mapping_v*.xlsx"
            )
        )
        if not candidates:
            result.warnings.append(
                {"kind": "missing_pillar_workbook", "pillar": f"P{pillar_no}"}
            )
            continue
        pillar_files[f"P{pillar_no}"] = sorted(candidates)[-1]

    tab_names = expected_tab_keys()

    for pillar_id, path in pillar_files.items():
        result.source_sha256s[path.name] = _sha256(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet_names = {ws.title.strip() for ws in wb.worksheets}
        per_tab: dict[str, ParseResult] = parse_workbook_tabs(
            wb, version=version, pillar_id=pillar_id, tab_names=tab_names
        )
        for tab_name, parse_res in per_tab.items():
            result.warnings.extend(
                [{**w, "tab": tab_name, "pillar": pillar_id}
                 for w in parse_res.warnings]
            )
            result.tab_row_counts.setdefault(tab_name, {})[pillar_id] = len(parse_res.rows)
            # Zero-row gate: a registered tab that IS present in this
            # workbook (canonically or via alias) must parse >0 rows.
            if (
                len(parse_res.rows) == 0
                and tab_name in PARSER_FOR_TAB
                and resolve_sheet_name(tab_name, sheet_names) is not None
            ):
                result.zero_row_gate_failures.append(f"{pillar_id}/{tab_name}")
            for row in parse_res.rows:
                target = row.pop("__target__", _default_target_for_tab(tab_name))
                if target is None:
                    continue
                result.rows_by_table.setdefault(target, []).append(row)
        wb.close()

    return result


def _default_target_for_tab(tab_name: str) -> str | None:
    from .canonical_tabs import target_for

    return target_for(tab_name)


def run_validators(result: LoaderResult) -> dict[str, Any]:
    subcaps = result.rows_by_table.get("ccg_subcaps", [])
    vc_rows = result.rows_by_table.get("ccg_vc_mapping", [])
    matrix_refs: list[str] = []
    for table_name in ("ccg_offering_subcap_matrix", "ccg_dataproduct_subcap_matrix",
                       "ccg_maturity_descriptors", "ccg_vc_mapping",
                       "ccg_l4_features", "ccg_theme_subcap_mapping",
                       "ccg_subcap_xpillar_coverage", "ccg_subcap_completeness",
                       "ccg_toggle_cascade"):
        for row in result.rows_by_table.get(table_name, []):
            sid = row.get("subcap_id")
            if sid:
                matrix_refs.append(str(sid))

    report = merge_reports(
        validate_pillar_totals(subcaps),
        validate_fk_closure(subcaps, matrix_refs),
        validate_value_chain_subverticals(vc_rows),
    )
    result.validation_passed = report.ok
    result.validation_detail = {
        "passed": report.passed,
        "failed": report.failed,
        "warnings": report.warnings,
    }
    return result.validation_detail


def _normalize_version(raw: str) -> str:
    """Tolerate `7.0` / `v7.0` / `V7.0` / `7` → canonical `v7.0`."""
    s = raw.strip().lower().lstrip("v")
    if not s:
        raise ValueError(f"invalid --version: {raw!r}")
    if "." not in s:
        s = f"{s}.0"
    return f"v{s}"


def _resolve_workbooks_dir(raw: str) -> pathlib.Path:
    """Accept a local path OR a `gs://bucket/prefix/` URL.

    For GCS, downloads every `*.xlsx` under the prefix into a tempdir
    (cleaned up by the OS on container exit) and returns that local
    path. Imports the google-cloud-storage client lazily so local
    unit tests don't need it.
    """
    if not raw.startswith("gs://"):
        return pathlib.Path(raw)
    import tempfile

    try:
        from google.cloud import storage  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "google-cloud-storage required for gs:// paths; "
            "install it in the workers image"
        ) from e

    bucket_name, _, prefix = raw[len("gs://"):].partition("/")
    if not bucket_name:
        raise ValueError(f"invalid GCS URL (no bucket): {raw!r}")
    prefix = prefix.rstrip("/")  # `gs://b/p/` → `p`; `gs://b` → ``

    tmp = pathlib.Path(tempfile.mkdtemp(prefix="ccg-loader-"))
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    found = 0
    for blob in client.list_blobs(bucket, prefix=prefix or None):
        # Skip directory-marker objects (zero-byte keys ending in /).
        if blob.name.endswith("/"):
            continue
        # Only download workbooks; the loader doesn't read other files.
        if not blob.name.lower().endswith(".xlsx"):
            continue
        local = tmp / pathlib.Path(blob.name).name
        blob.download_to_filename(str(local))
        found += 1
    if found == 0:
        raise FileNotFoundError(
            f"no *.xlsx found under {raw}; upload the 4 pillar workbooks "
            "via `gsutil cp Pillar_*.xlsx gs://…/v7.0/` first"
        )
    print(f"  downloaded {found} workbooks from {raw} → {tmp}", file=sys.stderr)
    return tmp


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="CCG catalogue loader")
    parser.add_argument("--version", required=True,
                        help="Catalogue version (`7.0` / `v7.0` both accepted)")
    parser.add_argument("--workbooks-dir", required=True,
                        help="Local directory OR `gs://bucket/prefix/` URL")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse + validate only; do not write to DB")
    parser.add_argument("--output-json", type=pathlib.Path,
                        help="Where to write the parsed rowset JSON")
    args = parser.parse_args(argv)

    try:
        version = _normalize_version(args.version)
    except ValueError as e:
        print(f"::error::{e}", file=sys.stderr)
        return 2

    try:
        workbooks_dir = _resolve_workbooks_dir(args.workbooks_dir)
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        print(f"::error::{e}", file=sys.stderr)
        return 2

    if not workbooks_dir.exists():
        print(f"workbooks dir not found: {workbooks_dir}", file=sys.stderr)
        return 2

    # Early progress update — operator was reporting "in progress" forever
    # on the admin pill because ccg_loader only flushed counters AFTER the
    # full parse + validation chain (which can run for minutes on a fresh
    # v7 catalogue). Surface "parsing started" immediately so the UI moves
    # off the bare default `result_summary="in progress"`.
    import contextlib as _ctx
    with _ctx.suppress(Exception):
        from workers._runner import get_current_tracker
        _early_ex = get_current_tracker()
        if _early_ex is not None:
            _early_ex.update(
                result_summary=f"parsing workbooks at {workbooks_dir.name}",
                files_parsed=0,
            )

    result = load_workbooks(
        version=version,
        workbooks_dir=workbooks_dir,
    )
    run_validators(result)

    summary = {
        "version": result.version,
        "tables": {t: len(rows) for t, rows in result.rows_by_table.items()},
        "per_tab_rows": result.tab_row_counts,
        "zero_row_gate_failures": result.zero_row_gate_failures,
        "warnings": len(result.warnings),
        "validation_passed": result.validation_passed,
        "validation": result.validation_detail,
        "source_sha256s": result.source_sha256s,
    }
    print(json.dumps(summary, indent=2))

    # ── Zero-row gate (2026-07 D3 remediation) ─────────────────────────
    # A registered+aliased tab that exists in a workbook but parsed 0
    # rows means the sheet layout drifted again. Fail LOUD (exit 3) so
    # the drift can never re-create the silent "value chains 0/94" class.
    if result.zero_row_gate_failures:
        print(
            "::error::zero-row gate: registered tab(s) present in the "
            "workbook parsed 0 rows: "
            + ", ".join(result.zero_row_gate_failures)
            + " — sheet layout drifted; update workers/ccg_loader "
            "TAB_ALIASES/parsers.",
            file=sys.stderr,
        )
        return 3

    # Flush counters to the admin pill — rows_added sums all per-table
    # row counts; parser_warnings carries the warning list for display.
    import contextlib
    with contextlib.suppress(Exception):
        from workers._runner import get_current_tracker
        ex = get_current_tracker()
        if ex is not None:
            total_rows = sum(len(rows) for rows in result.rows_by_table.values())
            ex.update(
                rows_added=total_rows,
                files_parsed=len(result.source_sha256s or {}),
                parser_warnings=result.warnings[:30],   # cap for JSONB
            )

    if args.output_json:
        args.output_json.write_text(
            json.dumps(
                {**summary, "rows": result.rows_by_table, "warnings_list": result.warnings},
                indent=2,
            )
        )

    if args.dry_run:
        return 0 if result.validation_passed else 1

    # ── Persist ccg_loader_runs row so admin UI sees it ──────────────
    #
    # State branches:
    #   validation_passed=True  → status='AWAITING_APPROVAL';
    #                              admin sees it in catalogue queue
    #                              + can promote to APPLIED.
    #   validation_passed=False → status='REJECTED' with the
    #                              validation report attached;
    #                              admin sees the failure reasons.
    #   DATABASE_URL_SYNC unset → log warning + skip; the JSON
    #                              summary is still printed.
    #   db_unreachable          → log + skip; tracker still captures
    #                              the run via job_executions wrapper.
    persist_rc = _persist_loader_run(result)
    if persist_rc != 0:
        return persist_rc
    # ── Canonical promote (2026-06-11 prod incident) ──────────────────
    # The admin Approve endpoint flips ccg_loader_runs to APPLIED and
    # promises "staging→canonical promote on the next tick" — but no
    # code path EVER wrote ccg_subcaps: production ran for weeks with a
    # bare ccg_catalog_versions FK row hand-inserted by the operator
    # ("manual band-aid") while every persisted package landed scores=0
    # (admin issue card "Catalogue placeholder without children").
    # Promote the validated parse directly: idempotent UPSERTs keyed on
    # (version, id) — re-runs and ingest-time auto-bootstrap stubs are
    # overwritten with real workbook names/descriptions. Approval
    # remains the human gate for making the version the default for
    # new ingests; rows existing early is harmless (reads are pinned
    # per-run via ccg_catalog_version).
    if result.validation_passed:
        promote_rc = _persist_canonical_rows(result)
        if promote_rc != 0:
            return promote_rc
    return 0


def _persist_loader_run(result) -> int:
    """Write one `ccg_loader_runs` row capturing this parse pass.

    Idempotent: re-running the same workbook produces a NEW row with
    the same source_sha256s — the admin reviews each one explicitly.
    The migrate-promote (staging.ccg_* → ccg_*) is a separate admin
    action (clicking "Apply" in the admin UI), so we never mutate the
    canonical catalogue tables here.

    Returns:
      0 on success, 0 on DB-unreachable (best-effort), 1 on
      unrecoverable error (caller exits with that code).
    """
    import contextlib
    import json as _json
    import os
    import uuid
    try:
        from sqlalchemy import create_engine, text
    except ImportError:
        print("::warning::sqlalchemy missing; skipping ccg_loader_runs row",
              file=sys.stderr)
        return 0

    # Use the shared resolver: workers historically only had
    # DATABASE_URL (asyncpg) injected by terraform; the sync DSN is
    # derived by replacing the driver suffix. Explicit
    # DATABASE_URL_SYNC always wins. Without this fallback the
    # 2026-05-28 production loader run exited 0 but wrote nothing —
    # the ccg_loader_runs INSERT was skipped silently, ccg_subcaps
    # stayed empty, and every backfill package emitted scores=0.
    try:
        from app.services.sync_dsn import resolve_sync_dsn
    except ImportError:
        # workers ship with app.* on PYTHONPATH; ImportError here
        # would be an env corruption. Fall back to the old behavior
        # so we don't crash a working production deploy.
        resolve_sync_dsn = None  # type: ignore[assignment]
    if resolve_sync_dsn is not None:
        url = resolve_sync_dsn()
    else:
        url = os.environ.get("DATABASE_URL_SYNC")
    if not url:
        print("::warning::No sync DSN available (neither "
              "DATABASE_URL_SYNC nor DATABASE_URL set) — skipping "
              "ccg_loader_runs persist. Run via the admin UI / "
              "Cloud Run Job for full persistence.",
              file=sys.stderr)
        return 0

    status_val = "AWAITING_APPROVAL" if result.validation_passed else "REJECTED"
    row_id = str(uuid.uuid4())
    source_files = _json.dumps([
        {"name": k, "sha256": v}
        for k, v in (result.source_sha256s or {}).items()
    ])
    parse_warnings = _json.dumps(list(result.warnings or []))
    validation_report = _json.dumps(result.validation_detail or {})

    eng = None
    try:
        eng = create_engine(url, pool_pre_ping=True, pool_size=1)
        with eng.begin() as conn:
            conn.execute(text("""
                INSERT INTO ccg_loader_runs (
                    id, version, status,
                    loader_started_at, loader_finished_at,
                    source_files, parse_warnings, validation_report,
                    diff_vs_prior_version
                ) VALUES (
                    CAST(:id AS uuid), :ver, :st,
                    NOW(), NOW(),
                    CAST(:sf AS jsonb), CAST(:pw AS jsonb),
                    CAST(:vr AS jsonb), NULL
                )
            """), {
                "id": row_id, "ver": result.version, "st": status_val,
                "sf": source_files, "pw": parse_warnings,
                "vr": validation_report,
            })
        print(f"ccg_loader_runs: persisted row {row_id} "
              f"(version={result.version}, status={status_val})",
              flush=True)
    except Exception as e:
        msg = str(e).lower()
        if "ccg_loader_runs" in msg and ("does not exist" in msg or "undefinedtable" in msg):
            print("::warning::ccg_loader_runs table missing — migration "
                  "012_ccg_catalogue not applied. Run ./migrate.sh.",
                  file=sys.stderr)
            return 0
        print(f"::warning::ccg_loader_runs persist failed: {e!s}",
              file=sys.stderr)
        return 0
    finally:
        if eng is not None:
            with contextlib.suppress(Exception):
                eng.dispose()
    return 0


def _persist_canonical_rows(result) -> int:
    """UPSERT the validated parse into the canonical ccg_* tables.

    Tables: ccg_catalog_versions (FK parent), ccg_pillars (synthesized
    from category pillar_ids), ccg_categories, ccg_l1_capabilities,
    ccg_subcaps. Conflict keys are (version, <id>); names/descriptions
    always take the workbook value so ingest-time stub bootstraps
    ("Subcap P1C1.1.1") heal on the next load. Best-effort on a missing
    DATABASE_URL_SYNC (same contract as _persist_loader_run).
    """
    import json as _json
    import os
    try:
        from sqlalchemy import create_engine, text
    except ImportError:
        print("::warning::sqlalchemy missing; skipping canonical promote",
              file=sys.stderr)
        return 0
    dsn = os.environ.get("DATABASE_URL_SYNC")
    if not dsn:
        print("::warning::DATABASE_URL_SYNC unset; skipping canonical promote",
              file=sys.stderr)
        return 0
    pillar_names = {"P1": "Strategy", "P2": "Customer Experience",
                    "P3": "Process Automation", "P4": "Data & AI"}
    cats = result.rows_by_table.get("ccg_categories", [])
    l1s = result.rows_by_table.get("ccg_l1_capabilities", [])
    subs = result.rows_by_table.get("ccg_subcaps", [])
    bands = result.rows_by_table.get("ccg_maturity_descriptors", [])
    vc_rows = result.rows_by_table.get("ccg_vc_mapping", [])
    alias_rows = result.rows_by_table.get("ccg_subcap_aliases", [])
    l3_raw = result.rows_by_table.get("ccg_l3_platforms", [])
    l4_raw = result.rows_by_table.get("ccg_l4_features", [])
    story_raw = result.rows_by_table.get("ccg_user_stories", [])
    # The L3 reference sheet ships identically in all four pillar workbooks
    # and the stories sheet repeats keys across pillars — dedupe on each
    # table's PK python-side so one executemany can't hit the same row twice
    # ("ON CONFLICT DO UPDATE cannot affect row a second time").
    l3_rows = list({r["l3_id"]: r for r in l3_raw
                    if r.get("l3_id") and r.get("platform_name")}.values())
    l4_rows = list({(r["subcap_id"], r.get("l3_id") or "", r["feature_name"]): r
                    for r in l4_raw
                    if r.get("subcap_id") and r.get("feature_name")}.values())
    story_rows = list({r["story_key"]: r for r in story_raw
                       if r.get("story_key") and r.get("subcap_id")}.values())
    try:
        eng = create_engine(dsn, pool_pre_ping=True)
        with eng.begin() as cx:
            cx.execute(text(
                """
                INSERT INTO ccg_catalog_versions
                    (version, released_at, source_sha256s, loader_run_id,
                     frozen_at, notes)
                VALUES (:v, NOW(), CAST(:sha AS JSONB), gen_random_uuid(),
                        NOW(), 'canonical promote by ccg_loader')
                ON CONFLICT (version) DO NOTHING
                """),
                {"v": result.version,
                 "sha": _json.dumps(getattr(result, "source_sha256s", {}) or {})})
            for pid in sorted({c.get("pillar_id") for c in cats if c.get("pillar_id")}):
                cx.execute(text(
                    """
                    INSERT INTO ccg_pillars (version, pillar_id, name,
                        description, category_count, l1_capability_count,
                        subcap_count)
                    VALUES (:v, :p, :n, :n, 0, 0, 0)
                    ON CONFLICT (version, pillar_id) DO UPDATE SET name=:n
                    """), {"v": result.version, "p": pid,
                           "n": pillar_names.get(pid, pid)})
            for c in cats:
                cx.execute(text(
                    """
                    INSERT INTO ccg_categories (version, category_id,
                        pillar_id, name)
                    VALUES (:v, :c, :p, :n)
                    ON CONFLICT (version, category_id)
                    DO UPDATE SET name = EXCLUDED.name,
                                  pillar_id = EXCLUDED.pillar_id
                    """), {"v": result.version, "c": c["category_id"],
                           "p": c.get("pillar_id"), "n": c.get("name")})
            for r in l1s:
                cx.execute(text(
                    """
                    INSERT INTO ccg_l1_capabilities (version, l1_id,
                        category_id, name)
                    VALUES (:v, :l, :c, :n)
                    ON CONFLICT (version, l1_id)
                    DO UPDATE SET name = EXCLUDED.name,
                                  category_id = EXCLUDED.category_id
                    """), {"v": result.version, "l": r["l1_id"],
                           "c": r.get("category_id"), "n": r.get("name")})
            for r in subs:
                cx.execute(text(
                    """
                    INSERT INTO ccg_subcaps (version, subcap_id, l1_id,
                        name, description, solution_type, tier,
                        zennify_status)
                    VALUES (:v, :s, :l, :n, :d, :st, :t, 'Active')
                    ON CONFLICT (version, subcap_id)
                    DO UPDATE SET name = EXCLUDED.name,
                                  description = EXCLUDED.description,
                                  l1_id = EXCLUDED.l1_id,
                                  solution_type = EXCLUDED.solution_type,
                                  tier = EXCLUDED.tier
                    """),
                    {"v": result.version, "s": r["subcap_id"],
                     "l": r.get("l1_id"), "n": r.get("name"),
                     "d": (r.get("description") or "")[:4000],
                     "st": r.get("solution_type") or "Hybrid",
                     "t": r.get("tier") or "T1"})
            # ── ccg_maturity_descriptors (FK → ccg_subcaps just upserted) ──
            # Batched executemany; conflict key (version, subcap_id, band).
            if bands:
                cx.execute(text(
                    """
                    INSERT INTO ccg_maturity_descriptors
                        (version, subcap_id, band, narrative, features)
                    VALUES (:version, :subcap_id, :band, :narrative, :features)
                    ON CONFLICT (version, subcap_id, band)
                    DO UPDATE SET narrative = EXCLUDED.narrative,
                                  features = EXCLUDED.features
                    """),
                    [{"version": b["version"], "subcap_id": b["subcap_id"],
                      "band": b["band"],
                      "narrative": (b.get("narrative") or "")[:8000],
                      "features": (b.get("features") or "")[:8000]}
                     for b in bands])
            # ── ccg_vc_mapping (the D3 value-chain view's ONLY source; was
            # never persisted — root cause of "value chains 0/94") ─────────
            if vc_rows:
                cx.execute(text(
                    """
                    INSERT INTO ccg_vc_mapping
                        (version, subcap_id, subvertical_code, value_chain_stages)
                    VALUES (:version, :subcap_id, :subvertical_code, :stages)
                    ON CONFLICT (version, subcap_id, subvertical_code)
                    DO UPDATE SET value_chain_stages = EXCLUDED.value_chain_stages
                    """),
                    [{"version": r["version"], "subcap_id": r["subcap_id"],
                      "subvertical_code": r["subvertical_code"],
                      "stages": r.get("value_chain_stages") or []}
                     for r in vc_rows])
            # ── ccg_subcap_aliases — only rows that actually bridge a
            # rename (prior != current). The v7.0 _R1 sheet also carries
            # placeholder priors ("NEW (v7.0 addition)", "[v7.0-new — split
            # from …]") that are NOT subcap ids — those are dropped, and the
            # batch is deduped on the (prior_version, prior_subcap_id,
            # current_version) conflict key so executemany can't hit the
            # same row twice in one statement. ─────────────────────────────
            import re as _re
            _subcap_id_re = _re.compile(r"^P\d+C\d+")
            dedup: dict[tuple[str, str, str], dict] = {}
            for r in alias_rows:
                ps = (r.get("prior_subcap_id") or "").strip()
                cs = (r.get("current_subcap_id") or "").strip()
                if not (ps and cs) or ps == cs:
                    continue
                if not (_subcap_id_re.match(ps) and _subcap_id_re.match(cs)):
                    continue
                dedup[(r["prior_version"], ps, r["current_version"])] = r
            bridging = list(dedup.values())
            if bridging:
                cx.execute(text(
                    """
                    INSERT INTO ccg_subcap_aliases
                        (prior_version, prior_subcap_id, current_version,
                         current_subcap_id, migration_action, migration_notes)
                    VALUES (:pv, :ps, :cv, :cs, :act, :notes)
                    ON CONFLICT (prior_version, prior_subcap_id, current_version)
                    DO UPDATE SET current_subcap_id = EXCLUDED.current_subcap_id,
                                  migration_action = EXCLUDED.migration_action
                    """),
                    [{"pv": r["prior_version"], "ps": r["prior_subcap_id"],
                      "cv": r["current_version"], "cs": r["current_subcap_id"],
                      "act": r.get("migration_action") or "MIGRATED",
                      "notes": r.get("migration_notes")}
                     for r in bridging])
            # ── ccg_l3_platforms / ccg_l4_features / ccg_user_stories —
            # the platform-affinity training layers (2026-07-12 directive:
            # "train the models using the L3, L4 and use cases"). These
            # sheets parsed but were never promoted, leaving all three
            # tables at 0 rows. ───────────────────────────────────────────
            if l3_rows:
                cx.execute(text(
                    """
                    INSERT INTO ccg_l3_platforms
                        (version, l3_id, vendor, platform_name, category,
                         description, setup_path, prerequisites,
                         detailed_capabilities)
                    VALUES (:version, :l3_id, :vendor, :platform_name,
                            :category, :description, :setup_path,
                            :prerequisites, :detailed_capabilities)
                    ON CONFLICT (version, l3_id)
                    DO UPDATE SET vendor = EXCLUDED.vendor,
                                  platform_name = EXCLUDED.platform_name,
                                  category = EXCLUDED.category,
                                  description = EXCLUDED.description,
                                  setup_path = EXCLUDED.setup_path,
                                  prerequisites = EXCLUDED.prerequisites,
                                  detailed_capabilities =
                                      EXCLUDED.detailed_capabilities
                    """),
                    [{"version": r["version"], "l3_id": r["l3_id"],
                      "vendor": r.get("vendor") or "Unknown",
                      "platform_name": r["platform_name"],
                      "category": r.get("category"),
                      "description": r.get("description"),
                      "setup_path": r.get("setup_path"),
                      "prerequisites": r.get("prerequisites"),
                      "detailed_capabilities": r.get("detailed_capabilities")}
                     for r in l3_rows])
            if l4_rows:
                cx.execute(text(
                    """
                    INSERT INTO ccg_l4_features
                        (version, subcap_id, l3_id, feature_name, vendor,
                         feature_type, customization_level, reference_url)
                    VALUES (:version, :subcap_id, :l3_id, :feature_name,
                            :vendor, :feature_type, :customization_level,
                            :reference_url)
                    ON CONFLICT (version, subcap_id, l3_id, feature_name)
                    DO UPDATE SET vendor = EXCLUDED.vendor,
                                  feature_type = EXCLUDED.feature_type,
                                  customization_level =
                                      EXCLUDED.customization_level,
                                  reference_url = EXCLUDED.reference_url
                    """),
                    [{"version": r["version"], "subcap_id": r["subcap_id"],
                      "l3_id": r.get("l3_id") or "",
                      "feature_name": r["feature_name"],
                      "vendor": r.get("vendor"),
                      "feature_type": r.get("feature_type"),
                      "customization_level": r.get("customization_level"),
                      "reference_url": r.get("reference_url")}
                     for r in l4_rows])
            if story_rows:
                cx.execute(text(
                    """
                    INSERT INTO ccg_user_stories
                        (version, story_key, subcap_id, source_type,
                         source_ref, use_case_ids, l4_features_used,
                         match_confidence)
                    VALUES (:version, :story_key, :subcap_id, :source_type,
                            :source_ref, :use_case_ids, :l4_features_used,
                            :match_confidence)
                    ON CONFLICT (version, story_key)
                    DO UPDATE SET subcap_id = EXCLUDED.subcap_id,
                                  source_type = EXCLUDED.source_type,
                                  source_ref = EXCLUDED.source_ref,
                                  use_case_ids = EXCLUDED.use_case_ids,
                                  l4_features_used = EXCLUDED.l4_features_used,
                                  match_confidence = EXCLUDED.match_confidence
                    """),
                    [{"version": r["version"], "story_key": r["story_key"],
                      "subcap_id": r["subcap_id"],
                      "source_type": r.get("source_type"),
                      "source_ref": r.get("source_ref"),
                      "use_case_ids": r.get("use_case_ids"),
                      "l4_features_used": r.get("l4_features_used"),
                      "match_confidence": r.get("match_confidence")}
                     for r in story_rows])
        print(f"canonical promote: pillars+cats={len(cats)} l1={len(l1s)} "
              f"subcaps={len(subs)} maturity_bands={len(bands)} "
              f"vc_mapping={len(vc_rows)} aliases={len(bridging)} "
              f"l3_platforms={len(l3_rows)} l4_features={len(l4_rows)} "
              f"user_stories={len(story_rows)} upserted for {result.version}")
        return 0
    except Exception as e:  # pragma: no cover — db-unreachable best-effort
        print(f"::warning::canonical promote failed: {type(e).__name__}: {e}",
              file=sys.stderr)
        return 0


# NB: the __main__ guard MUST stay at the BOTTOM of this module. It used to
# sit above `_persist_canonical_rows`, so script execution (python -m
# workers.ccg_loader.main) hit `raise SystemExit(main())` before the promote
# function was even defined — every direct loader run crashed with NameError
# AFTER printing its summary, and the canonical ccg_* tables were never
# written (one more contributor to the "value chains 0/94" incident).
if __name__ == "__main__":
    from workers._runner import track_job_execution
    with track_job_execution("ccg_loader"):
        raise SystemExit(main())
