"""The gold-standard gate, calibrated to the Golden 1 Credit Union reference.

Every check here reproduces a goeasy-Ltd finding (GSY-xx) as a mutation of a
synthetic gold-standard workbook/report and asserts the gate now catches it —
and that a clean gold-standard artefact passes. Self-contained: builds its own
fixtures with openpyxl / python-docx so CI needs no external Drive files.

See docs/goeasy-findings-register.md and docs/GOLD-STANDARD.md.
"""
import zipfile

import openpyxl
import pytest

from engine import gold_standard as GS


# ── a synthetic GOLD-STANDARD workbook (the Golden 1 shape, minimal) ──────

def _gold_workbook(path, n_per_pillar=3):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ex = wb.create_sheet("Executive_Summary")
    ex.append(["Field", "Value"])
    for f, v in [("Institution", "Test CU"), ("Sub-Vertical", "Credit Union"),
                 ("Evidence Mode", "HYBRID"), ("Overall Maturity", "2.25 (M2)"),
                 ("Peer Median (est.)", "3.05"), ("Gap to Peer", "-0.8"),
                 ("Subcaps Scored", "9 evidenced of 12 (75% coverage)"),
                 ("Evidence Gaps (Unknown)", "3"), ("Headline", "A clear one-liner.")]:
        ex.append([f, v])

    for p in ("P1", "P2", "P3", "P4"):
        ws = wb.create_sheet(f"{p}_Subcap_Scoring")
        ws.append(["SubCap_ID", "SubCap_Name", "Category", "Score", "Confidence",
                   "Evidence_IDs", "Source_URLs", "Evidence_Ceiling", "Caps_Applied",
                   "Rationale", "Proxy_Searched"])
        for i in range(1, n_per_pillar + 1):
            ws.append([f"{p}C1.1.{i}", f"Subcap {i}", f"{p}C1", 2.5, "MEDIUM",
                       "E-1", "https://example.com", 3.0, "none applied", "Because.", "No"])

    ps = wb.create_sheet("Pillar_Summary")
    ps.append(["Pillar", "Name", "Weighted_Score", "Maturity", "Gap_to_Peer"])
    ps.append(["OVERALL", "Overall", 2.25, "M2", -0.8])
    for p in ("P1", "P2", "P3", "P4"):
        ps.append([p, f"{p} name", 2.25, "M2", -0.8])

    cd = wb.create_sheet("Category_Detail")
    cd.append(["Category", "Maturity", "Coverage", "Peer_Median_Est", "Priority"])
    cd.append(["P1C1", "2.5 (M2)", "75%", 3.0, "High"])

    cov = wb.create_sheet("Coverage")
    cov.append(["Category", "Subcaps", "Scored", "Unknown_EvidenceGap", "Coverage_Pct"])
    cov.append(["P1C1", 12, 9, 3, 75.0])

    pb = wb.create_sheet("Peer_Benchmarks")
    pb.append(["peer_id", "peer_name", "overall_score_est", "basis"])
    pb.append(["p1", "SchoolsFirst", 3.1, "public digital-maturity signals"])

    fm = wb.create_sheet("Firmographics")
    fm.append(["Field", "Value", "Unit", "As at", "Evidence"])
    fm.append(["founded", "1933", "year", "2026", "E-1"])

    fa = wb.create_sheet("Focus_Areas")
    fa.append(["ID", "Priority", "Verbatim quote", "Document", "Page", "Cells"])
    fa.append(["FA-1", "Member relationship layer", "we will...", "AR2025", "12", "P2C1"])

    ir = wb.create_sheet("Issue_Register")
    ir.append(["ID", "Type", "Severity", "Status", "Description", "Capability impact"])
    ir.append(["IS-1", "regulatory", "S2", "open", "matter", "P1C2"])

    rec = wb.create_sheet("Recommendations")
    rec.append(["Rec_ID", "Title", "Category_ID", "Priority", "Horizon", "Owner"])
    sol = wb.create_sheet("Solution_Catalogue")
    sol.append(["solution_id", "solution_name"])
    sol.append(["SOL-1", "CRM"])

    tr = wb.create_sheet("Tech_Register")
    tr.append(["TS_ID", "Product", "Vendor", "Layer", "Status"])
    tr.append(["TS-1", "Core", "Vendor", "OPS", "CONFIRMED"])

    ft = wb.create_sheet("Financial_Trends")
    ft.append(["Metric", "Unit", "FY2020", "FY2021", "FY2022", "FY2023",
               "FY2024", "CAGR (FY20-24)", "Evidence"])
    for name, vals in [("Revenue", [100, 120, 140, 165, 190]),
                       ("Net income", [10, 14, 12, 18, 22]),
                       ("Total assets", [500, 620, 730, 840, 960]),
                       ("Loans receivable", [400, 520, 640, 760, 900]),
                       ("Return on equity", [12.1, 13.4, 11.9, 14.2, 15.0])]:
        ft.append([name, "$000s"] + vals + ["17.4%", "E-1"])

    wb.save(str(path))
    return path


# ── the clean artefact passes ─────────────────────────────────────────────

def test_gold_standard_workbook_passes(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    assert GS.workbook_findings(wb) == [], "a gold-standard workbook must pass"


# ── GSY-15/17 · research workbook shipped where the assessment belongs ────

def test_missing_assessment_sheets_is_caught(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    for s in ("Executive_Summary", "Firmographics", "Focus_Areas", "Issue_Register"):
        del w[s]
    w.save(wb)
    codes = {f["code"] for f in GS.workbook_findings(wb)}
    assert "GS-WB-STAGE" in codes


# ── GSY-16 · coverage hidden / gaps not disclosed ─────────────────────────

def test_coverage_must_disclose_unknown(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    ws = w["Coverage"]
    # rewrite header without the disclosure columns
    for j, v in enumerate(["Category", "Selected", "Verdict"], start=1):
        ws.cell(1, j).value = v
    ws.cell(1, 4).value = None
    ws.cell(1, 5).value = None
    w.save(wb)
    codes = {f["code"] for f in GS.workbook_findings(wb)}
    assert "GS-WB-COVERAGE" in codes


# ── GSY-01 / GSY-02 · a blank or zero score ───────────────────────────────

@pytest.mark.parametrize("bad,code", [(None, "GS-WB-SCORES"),
                                      ("Unknown", "GS-WB-SCORES"),
                                      (0, "GS-WB-NOZERO")])
def test_every_subcap_must_carry_a_real_score(tmp_path, bad, code):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    w["P1_Subcap_Scoring"].cell(2, 4).value = bad   # column D of the first subcap
    w.save(wb)
    codes = {f["code"] for f in GS.workbook_findings(wb)}
    assert code in codes


# ── GSY-03 · a blank subcap name ──────────────────────────────────────────

def test_blank_subcap_name_is_caught(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    w["P1_Subcap_Scoring"].cell(2, 2, "N/A")
    w.save(wb)
    assert "GS-WB-NAMES" in {f["code"] for f in GS.workbook_findings(wb)}


# ── GSY-04 · a hedge string in a value cell ───────────────────────────────

def test_hedge_string_in_a_value_cell_is_caught(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    w["Peer_Benchmarks"].cell(2, 3, "Not established this run")
    w.save(wb)
    codes = {f["code"] for f in GS.workbook_findings(wb)}
    assert "GS-WB-PEERS" in codes or "GS-WB-NOHEDGE" in codes


# ── REPORT gate, exercised at the string level via a synthetic docx ───────

def _docx(path, paragraphs, add_header=True):
    import docx
    d = docx.Document()
    for style, text in paragraphs:
        d.add_paragraph(text, style=style)
    d.save(str(path))
    if add_header:
        # inject a header part so GS-RPT-BRANDING sees branding chrome
        with zipfile.ZipFile(str(path), "a") as z:
            z.writestr("word/header1.xml", "<hdr/>")
    return path


def _assessment_body(overall="2.25"):
    """A report in the PINNED template's shape: every numbered section as a
    Heading 1 (GS-RPT-SECTIONS checks number AND heading), the four pillar
    deep dives as cards under §5, the REC cards under §8."""
    from engine import report_spec as RS
    body = []
    for h in RS.numbered_headings("assessment"):
        body.append(("Heading 1", h))
        n = h.split(".")[0]
        if n == "1":
            body.append(("Normal", f"Overall {overall} (M2). Coverage 75%, Unknown gaps disclosed."))
        elif n == "5":
            for p in "P1 P2 P3 P4".split():
                body.append(("Heading 2", f"5.{p[1]} Pillar deep dive ({p}): a pillar"))
                body.append(("Normal", "AI and data overlay: models."))
        elif n == "8":
            for i in range(1, 6):
                body.append(("Heading 2", f"REC-0{i}: do a thing"))
                body.append(("Heading 3", "Rebuttal"))
                body.append(("Normal", "Strongest counter. It survives because."))
        else:
            body.append(("Normal", "Section body."))
    # 5-year financial trajectory (GS-RPT-FINANCIALS)
    body.append(("Normal", "Revenue grew across FY2020, FY2021, FY2022, FY2023 and "
                 "FY2024, a 17% CAGR; net income and total assets rose over the "
                 "five-year trajectory."))
    # depth: citations + words, at the FULL-SIZE Golden 1 floors (a bare
    # `report_findings` holds a report to the reference's 690 subcaps: 115
    # distinct citations and the pinned Doc's 8,400-word contract).
    body.append(("Normal", " ".join(f"E-{i}" for i in range(1, 121)) + " " + "word " * 8600))
    return body


def test_gold_standard_assessment_report_passes(tmp_path):
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", _assessment_body())
    assert GS.report_findings(rp, kind="assessment") == []


@pytest.mark.parametrize("inject,code", [
    ("Not established this run", "GS-RPT-NOHEDGE"),
    ("This is Transformational band", "GS-RPT-BANDS"),
    ("{{ENTITY_NAME}}", "GS-RPT-NOTOKENS"),
])
def test_report_hedges_tokens_and_fifth_band_are_caught(tmp_path, inject, code):
    body = _assessment_body() + [("Normal", inject)]
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", body)
    assert code in {f["code"] for f in GS.report_findings(rp, kind="assessment")}


def test_m_level_scale_is_allowed(tmp_path):
    # M1..M5 is the maturity SCALE the reference uses; only a 5th BAND is banned.
    body = _assessment_body() + [("Normal", "The rubric runs M1 through M5.")]
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", body)
    assert "GS-RPT-BANDS" not in {f["code"] for f in GS.report_findings(rp, kind="assessment")}


def test_blank_document_loses_branding(tmp_path):
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", _assessment_body(), add_header=False)
    assert "GS-RPT-BRANDING" in {f["code"] for f in GS.report_findings(rp, kind="assessment")}


def test_missing_ai_overlay_is_caught(tmp_path):
    body = [("Heading 1", "1. Executive Summary"),
            ("Normal", "Overall 2.25. coverage disclosed. " + "word " * 3600
             + " ".join(f"E-{i}" for i in range(1, 70)))]
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", body)
    assert "GS-RPT-AIOVERLAY" in {f["code"] for f in GS.report_findings(rp, kind="assessment")}


# ── GSY-18 · depth: a real 5-year financial trajectory must be present ────

def test_workbook_without_a_5_year_financial_trajectory_is_caught(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    del w["Financial_Trends"]          # remove the only 5-year series
    w.save(wb)
    assert "GS-WB-FINANCIALS" in {f["code"] for f in GS.workbook_findings(wb)}


def test_financial_trends_sheet_with_too_few_years_is_caught(tmp_path):
    wb = _gold_workbook(tmp_path / "wb.xlsx")
    w = openpyxl.load_workbook(wb)
    ws = w["Financial_Trends"]
    for col in (5, 6, 7):              # blank FY2022..FY2024 headers -> < 5 year cols
        ws.cell(1, col).value = None
    w.save(wb)
    assert "GS-WB-FINANCIALS" in {f["code"] for f in GS.workbook_findings(wb)}


def test_report_without_financial_trajectory_is_caught(tmp_path):
    body = [("Heading 1", "1. Executive Summary"),
            ("Normal", "Overall 2.25. coverage disclosed. AI and data overlay. "
             + "word " * 3600 + " ".join(f"E-{i}" for i in range(1, 70)))]
    rp = _docx(tmp_path / "DMA_Assessment_Report_x.docx", body)
    assert "GS-RPT-FINANCIALS" in {f["code"] for f in GS.report_findings(rp, kind="assessment")}


def test_fy_prefixed_years_count_as_fiscal_years(tmp_path):
    # regression: "FY2020" has no \b before the digits — the year regex must
    # still see it, or a real trends sheet reads as zero year columns.
    assert len(GS.re.findall(r"(?<!\d)20[0-3]\d(?!\d)", "FY2020 FY2021 FY2022")) == 3
