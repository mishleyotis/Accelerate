"""Each contract rule fired and not fired, and the handoff that stops lying.

The three rules the audit found inert — added columns, blank Evidence_IDs,
and row identity — each get a test that FIRES them, because a rule with only
a happy-path test is how they stayed inert."""
import json
import pathlib
import shutil

import openpyxl
import pytest

from engine import contract as C
from engine import handoff, ledger as L, patch_validator, strip_working_area
from engine import validator
from engine.workbook import WorkbookError

from fixtures import CAT, bank_evidence, good_synthesis, new_run, synthesise


def _good_run(tmp_path, n=8, prelim=True):
    run = new_run(tmp_path, n=n, prelim=prelim)
    wb = run.open()
    for cell in wb.selected_subcaps():
        synthesise(wb, cell, good_synthesis(cell, bank_evidence(wb, cell)))
    return run, wb


# ── the positive control ─────────────────────────────────────────────────

def test_a_workbook_the_engine_built_passes_its_own_contract(tmp_path):
    run, wb = _good_run(tmp_path)
    assert validator.validate(wb.path, run_id="R-TEST-1") == []


# ── rule 1 ───────────────────────────────────────────────────────────────

def test_rule1_fires_on_a_missing_sheet(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path); del book["Handoff_Lock"]
    book.save(wb.path)
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 1 and "Handoff_Lock" in f["detail"] for f in fails)


# ── rule 2 · AUD-0064, the clause that never fired ───────────────────────

def test_rule2_fires_on_an_added_column(tmp_path):
    """The old validator sliced columns 1..11, so a 12th was invisible — and
    an unstripped 22-column working area passed the only gate on the file."""
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path)
    ws = book["P1_Subcap_Scoring"]
    ws.cell(row=1, column=len(C.PILLAR_COLUMNS) + 1, value="Smuggled_Column")
    book.save(wb.path)
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 2 and "Smuggled_Column" in f["detail"]
               for f in fails), fails


def test_rule2_fires_on_a_renamed_column(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path)
    book["P1_Subcap_Scoring"].cell(row=1, column=4, value="Points")
    book.save(wb.path)
    assert any(f["rule"] == 2 for f in validator.validate(wb.path))


# ── rule 3 · AUD-0014, identity not count ────────────────────────────────

def test_rule3_fires_when_the_rows_are_the_wrong_subcaps(tmp_path):
    """Swapping an in-scope id for an out-of-scope one while HOLDING THE
    COUNT CONSTANT returned FAILS=0 under the old rule."""
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path)
    book["P1_Subcap_Scoring"].cell(row=2, column=1, value="P4C9.9.9")
    book.save(wb.path)
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 3 for f in fails), fails


def test_rule3_fires_on_a_duplicate_id(tmp_path):
    run, wb = _good_run(tmp_path, n=3)
    book = openpyxl.load_workbook(wb.path)
    ws = book["P1_Subcap_Scoring"]
    ws.cell(row=3, column=1, value=ws.cell(row=2, column=1).value)
    book.save(wb.path)
    assert any(f["rule"] == 3 and "duplicate" in f["detail"]
               for f in validator.validate(wb.path))


# ── rule 4 ───────────────────────────────────────────────────────────────

def test_rule4_fires_on_a_score_at_the_research_stage(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    wb.set_scoring(wb.selected_subcaps()[0], {"Score": 3})
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 4 for f in fails)


def test_rule4_fires_when_the_assessment_stage_has_no_scores(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    fails = validator.validate(wb.path, expect_scores=True)
    assert any(f["rule"] == 4 for f in fails)


def test_rule4_reads_the_stage_off_the_file_when_no_caller_says(tmp_path):
    """`expect_scores` used to default to False, and `assemble.package`
    passed the default — so every package this engine built was validated
    with research semantics, an assessment one included. `None` means read
    the workbook's own recorded stage."""
    run, wb = _good_run(tmp_path, n=2)
    wb.set_metadata("stage", "assessment")
    fails = validator.validate(wb.path)          # no expect_scores at all
    assert any(f["rule"] == 4 for f in fails), \
        "an assessment-stage workbook with an empty column D must fail rule 4"


def test_validate_takes_a_PATH_and_never_a_run_workbook(tmp_path):
    """The regression that broke the only route to a strip.

    `validate` opens the file with openpyxl on purpose — rules 1 and 2 judge
    a file that may not BE a valid run workbook. Reading the stage through
    `RunWorkbook.metadata()` therefore raised AttributeError inside
    `handoff.build`, and `strip` refuses without a handoff, so a run could
    finish research and never hand off.
    """
    run, wb = _good_run(tmp_path, n=2)
    assert validator.validate(str(wb.path)) == []
    assert validator.validate(wb.path) == []
    with pytest.raises((TypeError, AttributeError, OSError, ValueError)):
        validator.validate(wb)                   # the RunWorkbook itself


def test_a_pillar_the_run_never_selected_is_not_asked_for_a_score(tmp_path):
    """The regression the stage key introduced and the stress walk caught.

    A run's scope is a SELECTION; most runs do not cover all four pillars, so
    three of the four pillar sheets carry no rows at all. Rule 4 demanded a
    score in every sheet at the assessment stage, which made the assessment
    stage unreachable for those runs — the package refused with 'no scores
    present' for pillars nobody had ever been asked to research. No unit test
    saw it because `expect_scores` defaulted to False until the stage key
    existed, so the branch had never run against a real selection.
    """
    run, wb = _good_run(tmp_path, n=2)
    for cell in wb.selected_subcaps():
        wb.set_scoring(cell, {"Score": 3})
    wb.set_metadata("stage", "assessment")

    scored = {C.PILLAR_SHEET_OF[c.split("C")[0]] if hasattr(
        C, "PILLAR_SHEET_OF") else "P1_Subcap_Scoring"
        for c in wb.selected_subcaps()}
    assert len(scored) < len(C.PILLAR_SHEETS), \
        "this test needs a selection that does NOT cover all four pillars"

    fails = validator.validate(wb.path)
    rule4 = [f for f in fails if f["rule"] == 4]
    assert not rule4, (
        f"an unselected pillar was asked for a score: "
        f"{[f['detail'] for f in rule4]}")


def test_a_selected_pillar_with_no_score_still_fails_at_the_assessment_stage(
        tmp_path):
    """The other direction, which is what rule 4 is actually for."""
    run, wb = _good_run(tmp_path, n=2)
    wb.set_metadata("stage", "assessment")
    fails = [f for f in validator.validate(wb.path) if f["rule"] == 4]
    assert fails and "in scope and none scored" in fails[0]["detail"], fails


def test_the_command_line_has_no_opinion_unless_it_is_given_one(tmp_path):
    """`--expect-scores` was `store_true`, so ABSENT meant `False`, and False
    is not "no opinion" — it is the caller asserting the research stage.
    Passed straight through, the command line hard-coded research semantics
    for every workbook it was ever pointed at."""
    import subprocess
    import sys

    from engine import validator as V

    run, wb = _good_run(tmp_path, n=2)
    wb.set_metadata("stage", "assessment")
    script = str(pathlib.Path(V.__file__))
    r = subprocess.run([sys.executable, script, "--workbook", str(wb.path)],
                       capture_output=True, text=True, timeout=120)
    assert r.returncode == 1, r.stdout + r.stderr
    assert "rule 4" in r.stdout, r.stdout

    r2 = subprocess.run([sys.executable, script, "--workbook", str(wb.path),
                         "--no-expect-scores"],
                        capture_output=True, text=True, timeout=120)
    assert r2.returncode == 0, r2.stdout + r2.stderr


# ── rule 5 · AUD-0064, the vacuous branch ────────────────────────────────

def test_rule5_fires_on_a_blank_evidence_ids_cell(tmp_path):
    """`if fv and ...` turned every blank into a pass. On the real golden
    workbook 44 of 49 rows were blank and it certified clean."""
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path)
    # `.cell(..., value=None)` is "no value given" in openpyxl, not "clear",
    # so the assignment form is the one that actually blanks the cell.
    book["P1_Subcap_Scoring"].cell(
        row=2, column=C.PILLAR_COLUMNS.index("Evidence_IDs") + 1).value = None
    book.save(wb.path)
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 5 and "blank" in f["detail"] for f in fails), fails


def test_rule5_fires_on_cited_evidence_with_no_url(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    book = openpyxl.load_workbook(wb.path)
    book["P1_Subcap_Scoring"].cell(
        row=2, column=C.PILLAR_COLUMNS.index("Source_URLs") + 1, value="")
    book.save(wb.path)
    assert any(f["rule"] == 5 and "no URL" in f["detail"]
               for f in validator.validate(wb.path))


def test_rule5_accepts_the_literal_no_evidence(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    cell = wb.selected_subcaps()[0]
    wb.set_scoring(cell, {"Evidence_IDs": C.NO_EVIDENCE, "Source_URLs": None})
    assert not any(f["rule"] == 5 for f in validator.validate(wb.path))


# ── rule 6 ───────────────────────────────────────────────────────────────
#
# MEM-0467. Rule 6 had a test that FIRED it and none that did not, which is
# the one asymmetry this file's docstring says it does not tolerate — and it
# is exactly how an OVER-firing predicate stayed invisible. The rule tested
# `token in cell`, unanchored, so every banned token that is also an ordinary
# fragment of a URL path refused a real link: "n/a" lives inside
# "linkedin.com/in/an·n/a·mills" and inside "kpmg.com/us/e·n/a·rticles".
# On bank-of-travelers-rest that refused 12 of 12 cited URLs and returned
# FAILS=2 on a workbook with no data defect in it.
#
# So both directions below, and the not-firing half is the half that matters:
# every case marked False here PASSES the fixed predicate and five of them
# FAIL the substring one.


def _rule6_fires(wb, value) -> bool:
    wb.set_scoring(wb.selected_subcaps()[0], {"Source_URLs": value})
    return any(f["rule"] == 6 for f in validator.validate(wb.path))


@pytest.mark.parametrize("bad", [
    "Multiple Searches", "see report", "N/A",          # the original three
    "n/a", "tbd", "various", "see above",              # the rest of the set
    "Various", "See Above", "TBD",                     # mixed casing
    "N/A.", "(TBD)",                                   # trivial punctuation
    "n/a - nothing found",                             # token + prose
    "https://real.example/x, n/a",                     # one bad entry of two
])
def test_rule6_fires_on_a_cell_that_is_a_placeholder(tmp_path, bad):
    """The narrowing must not have weakened it: each of these still FAILS."""
    run, wb = _good_run(tmp_path, n=2)
    assert _rule6_fires(wb, bad), bad


@pytest.mark.parametrize("good", [
    # The two live BTR values. Both FAIL the pre-fix substring predicate.
    "https://www.linkedin.com/in/anna-mills-344a81101/",
    "https://kpmg.com/us/en/articles/2023/third-party-risk-management"
    "-final-interagency-guidance-reg-alert.html",
    # One per banned token, embedded in a path the way a real URL embeds it.
    "https://x.example/multiple-searches",
    "https://x.example/see-report-2024",
    "https://x.example/various-rates",
    "https://x.example/tbd-holdings/",
    "https://x.example/see-above-all",
    # A location need not carry a scheme to be a location.
    "kpmg.com/us/en/articles/x",
])
def test_rule6_does_not_fire_on_a_cell_that_holds_a_link(tmp_path, good):
    """A location is never a placeholder, however its characters spell."""
    run, wb = _good_run(tmp_path, n=2)
    assert not _rule6_fires(wb, good), good


def test_rule6_reports_the_offending_value_not_the_token_that_matched(tmp_path):
    """The old message printed "n/a" for a cell holding a LinkedIn URL, which
    reads as a data defect and sends a conductor hunting one that is not
    there. The reader needs the value that is wrong."""
    run, wb = _good_run(tmp_path, n=2)
    wb.set_scoring(wb.selected_subcaps()[0],
                   {"Source_URLs": "https://real.example/x; TBD (pending)"})
    detail = [f for f in validator.validate(wb.path) if f["rule"] == 6][0]["detail"]
    assert "TBD (pending)" in detail, detail


def test_rule6_predicate_is_directly_testable(tmp_path):
    """The predicate is public so the matrix above can be argued about at the
    unit it actually judges: one entry, not one workbook."""
    assert validator.placeholder_entries(
        "https://www.linkedin.com/in/anna-mills-344a81101/") == []
    assert validator.placeholder_entries("https://real.example/x, n/a") == ["n/a"]
    assert validator.placeholder_entries("") == []


# ── rule 7 ───────────────────────────────────────────────────────────────

def test_rule7_fires_on_a_run_id_mismatch(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    fails = validator.validate(wb.path, run_id="SOME-OTHER-RUN")
    assert any(f["rule"] == 7 for f in fails)


def test_rule7_fires_when_the_catalogue_has_moved(tmp_path, monkeypatch):
    run, wb = _good_run(tmp_path, n=2)
    monkeypatch.setattr(C, "catalogue_hash", lambda: "0" * 64)
    fails = validator.validate(wb.path)
    assert any(f["rule"] == 7 and "catalogue has moved" in f["detail"]
               for f in fails)


# ── AUD-0011 · the strip, and what it refuses to destroy ─────────────────

def test_the_strip_refuses_while_the_analysis_has_no_surviving_copy(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    with pytest.raises(SystemExit, match="Triangulation"):
        strip_working_area.strip(wb.path)


def test_the_strip_proceeds_once_the_handoff_carries_the_three(tmp_path):
    run, wb = _good_run(tmp_path, n=8)
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    hp = run.deliverables / handoff.HANDOFF_NAME
    hp.parent.mkdir(parents=True, exist_ok=True)
    hp.write_text(json.dumps(doc, default=str))
    out = tmp_path / "stripped.xlsx"
    r = strip_working_area.strip(wb.path, handoff=hp, out=out)
    assert r["core_columns_kept"] == 11
    book = openpyxl.load_workbook(out)
    assert book["P1_Subcap_Scoring"].max_column == 11


def test_the_stripped_workbook_still_carries_the_analysis_in_the_handoff(tmp_path):
    """AUD-0065's cost, measured: strip as instructed and three
    gate-required fields were destroyed with no surviving copy."""
    run, wb = _good_run(tmp_path, n=8)
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    rec = next(r for r in doc["subcap_records"] if r["research_synthesis"])
    for f in ("triangulation", "why_it_matters", "dma_impact"):
        assert rec["research_synthesis"][f], f"{f} did not survive"


# ── AUD-0078 · no fabricated band, no float ceiling ──────────────────────

def test_an_unresearched_subcap_carries_a_null_band_not_a_default(tmp_path):
    run = new_run(tmp_path, n=4); wb = run.open()
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    bands = {r["ceiling_band"] for r in doc["subcap_records"]}
    assert bands == {None}, bands


def test_category_ceilings_stay_band_words_and_never_become_scores(tmp_path):
    run, wb = _good_run(tmp_path, n=8)
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    for cat, c in doc["capability_ceilings"].items():
        assert c["ceiling_band"] in (None,) + C.BANDS
        assert not isinstance(c["ceiling_band"], (int, float))


# ── AUD-0138 · a facet that never ran says so ────────────────────────────

def test_a_facet_that_never_ran_is_NOT_RUN_not_an_empty_list(tmp_path):
    run = new_run(tmp_path, n=2); wb = run.open()
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    assert doc["safeguard_gates"]["outcome"] == "NOT_RUN"
    assert doc["safeguard_gates"]["reason"]
    assert doc["org_capability_proxies"]["outcome"] == "NOT_RUN"


def test_a_facet_that_did_run_carries_its_rows(tmp_path):
    from engine import floors_gate
    run, wb = _good_run(tmp_path, n=8)
    floors_gate.run(wb, CAT, qa_dir=run.qa_dir)
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    assert isinstance(doc["safeguard_gates"], list) and doc["safeguard_gates"]


# ── AUD-0002 · the handoff says what it is ───────────────────────────────

def test_the_handoff_names_the_workbook_as_the_authority(tmp_path):
    run, wb = _good_run(tmp_path, n=8)
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    assert doc["_contract"]["authority"] == "the scoring workbook"
    assert doc["_contract"]["handoff_lock"]["catalogue_hash"] == C.catalogue_hash()


def test_a_handoff_cannot_be_built_from_an_invalid_workbook(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    wb.set_scoring(wb.selected_subcaps()[0], {"Score": 4})
    with pytest.raises(SystemExit, match="does not satisfy its own contract"):
        handoff.build(wb, qa_dir=run.qa_dir, strict=True)


# ── AUD-0011 / AUD-0061 · the migration that did not exist ───────────────

def test_patch_validator_reports_a_v3_workbook_as_already_migrated(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    assert patch_validator.plan(wb.path)["already_v3"] is True


def test_patch_validator_folds_the_two_retired_sheets(tmp_path):
    run, wb = _good_run(tmp_path, n=3)
    cells = wb.selected_subcaps()
    legacy = tmp_path / "legacy.xlsx"
    shutil.copy2(wb.path, legacy)
    book = openpyxl.load_workbook(legacy)
    ws = book.create_sheet("Subcap_Synthesis")
    ws.append(["SubCap_ID", "dominant_claim", "triangulation"])
    ws.append([cells[0], "legacy claim carried over", "legacy triangulation"])
    neg = book.create_sheet("Negative_Findings")
    neg.append(["SubCap_ID", "proxy_log"])
    neg.append([cells[1], "peer probe, regulator probe"])
    book.save(legacy)
    p = patch_validator.plan(legacy)
    assert p["already_v3"] is False and len(p["fold"]) == 2
    r = patch_validator.apply(legacy)
    assert r["applied"] and r["rows_folded"] == 2
    book = openpyxl.load_workbook(legacy)
    assert "Subcap_Synthesis" not in book.sheetnames
    ws = book["P1_Subcap_Scoring"]
    idx = C.PILLAR_COLUMNS.index("Dominant_Claim") + 1
    vals = [ws.cell(row=r_, column=idx).value for r_ in range(2, ws.max_row + 1)]
    assert "legacy claim carried over" in vals


def test_a_migration_that_would_drop_a_column_refuses(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    legacy = tmp_path / "legacy2.xlsx"
    shutil.copy2(wb.path, legacy)
    book = openpyxl.load_workbook(legacy)
    ws = book.create_sheet("Subcap_Synthesis")
    ws.append(["SubCap_ID", "dominant_claim", "some_field_with_no_home"])
    ws.append([wb.selected_subcaps()[0], "x", "y"])
    book.save(legacy)
    with pytest.raises(SystemExit, match="no home in the contract"):
        patch_validator.apply(legacy)


# ── AUD-0042 / AUD-0043 · the peer store, and the lock on it ────────────

def test_the_contract_carries_a_peer_store_at_all():
    """AUD-0042: the category-grain peer store had NO FEEDER — the pinned
    workbook removed Peer_Benchmarks — so peer_scores is empty for every new
    run, which also empties ET-09's allow-list."""
    assert "Peer_Benchmarks" in C.SHEETS
    for col in ("Category_ID", "Peer_Median", "Peer_N", "Peer_Basis",
                "Peer_Names", "Source_Cell"):
        assert col in C.PEER_BENCHMARK_COLUMNS


def test_a_peer_figure_must_declare_the_rung_it_came_from():
    assert C.PEER_BASIS == ("table", "recomputed", "inferred",
                            "cannot_estimate")


def test_the_peer_set_locks_once_and_refuses_a_different_cohort(tmp_path):
    """AUD-0043: Handoff_Lock is the immutability mechanism both templates
    depend on and it existed in neither tree."""
    # prelim=False: PRELIM freezes a peer set of its own, and this test is
    # about the lock refusing a SECOND, different cohort.
    run, wb = _good_run(tmp_path, n=2, prelim=False)
    first = wb.lock_peer_set(["Alpha CU", "Beta CU"], basis="table")
    assert first["peer_n"] == 2 and first["already_locked"] is False
    again = wb.lock_peer_set(["Beta CU", "Alpha CU"], basis="table")
    assert again["already_locked"] is True
    with pytest.raises(WorkbookError, match="already locked"):
        wb.lock_peer_set(["Alpha CU", "Gamma CU"], basis="table")


def test_a_peer_basis_outside_the_ladder_is_refused(tmp_path):
    run, wb = _good_run(tmp_path, n=2)
    with pytest.raises(WorkbookError, match="peer basis"):
        wb.lock_peer_set(["Alpha CU"], basis="vibes")


def test_the_locked_set_reaches_the_handoff(tmp_path):
    run, wb = _good_run(tmp_path, n=8, prelim=False)
    wb.lock_peer_set(["Alpha CU", "Beta CU"], basis="recomputed")
    doc = handoff.build(wb, qa_dir=run.qa_dir, strict=False)
    lock = doc["_contract"]["handoff_lock"]
    assert lock["locked_peer_set"] == "Alpha CU|Beta CU"
    assert lock["peer_basis"] == "recomputed"
