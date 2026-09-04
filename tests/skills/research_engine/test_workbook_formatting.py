"""The workbook a client opens — formatted, named, and carrying its trajectory.

Owner, 2026-09-03 (issue 3): "The workbook always defaults to the wrong
structure each time; missing fields etc. Missing formatting; missing subcaps
names." Measured before the fix: a fresh workbook's D2 read `General`, no
sheet carried a data validation, no sheet carried an autofilter, and the
docstring on `_format_sheet` promised a two-decimal score format that nothing
applied. These pin the formatting the contract now guarantees, the
five-year financial trajectory (GSY-18) as a first-class sheet, and the
validator's refusal of a blank SubCap_Name at every stage.
"""
from __future__ import annotations

import openpyxl
import pytest

from engine import contract as C
from engine import profile
from engine import validator

from fixtures import bank_evidence, new_run, scored_run


def _reload(wb):
    return openpyxl.load_workbook(wb.path)


# ── number formats follow the column, not the writer ─────────────────────

def test_number_formats_follow_the_column_not_the_writer(tmp_path):
    run, wb, cells, ev = scored_run(tmp_path)
    x = _reload(wb)
    ws = x["P1_Subcap_Scoring"]
    cols = list(C.PILLAR_COLUMNS)
    score_col = cols.index("Score") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=score_col)
        if isinstance(cell.value, (int, float)):
            assert cell.number_format == "0.00", (r, cell.number_format)
    # the assessment tabs the same
    roll = x["Pillar_Rollup"]
    rcols = list(C.PILLAR_ROLLUP_COLUMNS)
    for r in range(2, roll.max_row + 1):
        c = roll.cell(row=r, column=rcols.index("score") + 1)
        if isinstance(c.value, (int, float)):
            assert c.number_format == "0.00"
    # prose `Value` columns are NOT numbers — Firmographics keeps General
    firm = x["Firmographics"]
    v = firm.cell(row=2, column=list(C.FIRMOGRAPHICS_COLUMNS).index("Value") + 1)
    assert v.number_format == "General"


def test_enum_columns_carry_data_validation_from_the_contract(tmp_path):
    wb = new_run(tmp_path, n=2).open()
    x = _reload(wb)
    ws = x["P1_Subcap_Scoring"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert formulas, "the scoring sheet carries no dropdowns at all"
    joined = " ".join(formulas)
    for label in C.CLAIM_LABELS:
        assert label in joined
    for band in C.BANDS:
        assert band in joined
    assert "YES,NO" in joined                       # Absence_Claimed
    # the fifth band is not offered anywhere
    assert "Transformational" not in joined
    # Status follows the tab: Tech_Register offers the four stack statuses
    tr = " ".join(dv.formula1 for dv in x["Tech_Register"].data_validations.dataValidation)
    for s in C.TECH_STATUS:
        assert s in tr
    # and the Search_Log offers only the contract's tools
    sl = " ".join(dv.formula1 for dv in x["Search_Log"].data_validations.dataValidation)
    for t in C.SEARCH_TOOLS:
        assert t in sl


def test_autofilter_covers_the_rows_after_every_save(tmp_path):
    run, wb, cells, ev = scored_run(tmp_path)
    x = _reload(wb)
    for name in ("P1_Subcap_Scoring", "Evidence_Detail", "Search_Log",
                 "Financial_Trends", "Subcap_Scores"):
        ws = x[name]
        assert ws.auto_filter.ref, f"{name}: no autofilter"
        assert ws.auto_filter.ref.endswith(str(ws.max_row)), (name, ws.auto_filter.ref)
        assert ws.freeze_panes == "A2"
        assert ws.cell(row=1, column=1).font.bold


def test_the_dashboard_is_the_first_tab_once_scoring_opens(tmp_path):
    run, wb, cells, ev = scored_run(tmp_path)
    assert _reload(wb).sheetnames[0] == "Executive_Summary"


# ── the five-year trajectory ─────────────────────────────────────────────

def test_a_financial_series_is_written_through_its_own_refusals(tmp_path):
    run = new_run(tmp_path, n=2, prelim=False)
    wb = run.open()
    cell = wb.selected_subcaps()[0]
    eid = bank_evidence(wb, cell, n=1)[0]
    with pytest.raises(profile.ProfileRefusal, match="fiscal year"):
        profile.financial(wb, metric="total assets", fiscal_year="latest",
                          value=1, unit="USD m", evidence=eid)
    with pytest.raises(profile.ProfileRefusal, match="not a number"):
        profile.financial(wb, metric="total assets", fiscal_year="FY2024",
                          value="about a billion", unit="USD m", evidence=eid)
    with pytest.raises(profile.ProfileRefusal, match="does not resolve"):
        profile.financial(wb, metric="total assets", fiscal_year="FY2024",
                          value=1, unit="USD m", evidence="E-9999")
    with pytest.raises(profile.ProfileRefusal, match="unit"):
        profile.financial(wb, metric="total assets", fiscal_year="FY2024",
                          value=1, unit="", evidence=eid)
    out = profile.financial(wb, metric="total assets", fiscal_year="2024",
                            value="1,234.5", unit="USD m", evidence=eid)
    assert out["action"] == "added" and out["Fiscal_Year"] == "FY2024"
    again = profile.financial(wb, metric="total assets", fiscal_year="FY2024",
                              value=1300, unit="USD m", evidence=eid)
    assert again["action"] == "updated"
    rows = wb.rows("Financial_Trends")
    assert len(rows) == 1 and float(rows[0]["Value"]) == 1300.0


def test_the_depth_floor_is_golden_1s_five_years_or_a_declared_reason(tmp_path):
    from engine import completeness
    run = new_run(tmp_path, n=2, prelim=False)
    wb = run.open()
    d = profile.financial_depth(wb)
    assert d["met"] is False and "five-year trajectory" in d["fix"]
    assert "engine.profile financial" in d["fix"]
    completeness.declare(wb, "Financial_Trends",
                         reason=("a private mutual publishes two fiscal years of "
                                 "audited statements and nothing earlier; both banked"))
    assert profile.financial_depth(wb)["met"] is True


def test_a_prelim_closed_run_carries_the_series(tmp_path):
    wb = new_run(tmp_path, n=2).open()
    d = profile.financial_depth(wb)
    assert d["met"] and d["years"] >= C.FINANCIAL_YEARS_FLOOR
    assert d["metrics"] >= C.FINANCIAL_METRICS_FLOOR
    assert profile.state(wb)["financial_trends"]["met"]


# ── names, at every stage ────────────────────────────────────────────────

def test_rule3_fires_on_a_blank_subcap_name(tmp_path):
    wb = new_run(tmp_path, n=3).open()
    assert validator.validate(wb.path, run_id="R-TEST-1") == []
    cell = wb.selected_subcaps()[1]
    wb.set_scoring(cell, {"SubCap_Name": ""})
    fails = validator.validate(wb.path, run_id="R-TEST-1")
    assert any(f["rule"] == 3 and "SubCap_Name" in f["detail"] for f in fails), fails


# ── the contract bump migrates, never strands ────────────────────────────

def test_a_v6_workbook_opens_under_v7_with_the_new_sheet_added(tmp_path):
    """A run in flight when Financial_Trends landed gets the sheet on open,
    and its Handoff_Lock moves with it — never a HALT."""
    wb = new_run(tmp_path, n=2, prelim=False).open()
    x = openpyxl.load_workbook(wb.path)
    del x["Financial_Trends"]
    for r in x["Handoff_Lock"].iter_rows(min_row=2):
        if r[0].value == "workbook_contract":
            r[1].value = "v6"
    x.save(wb.path)
    from engine.workbook import RunWorkbook
    again = RunWorkbook(wb.path)
    assert "Financial_Trends" in again._wb.sheetnames          # noqa: SLF001
    assert again.handoff_lock()["workbook_contract"] == C.WORKBOOK_CONTRACT
    assert again.verify_handoff_lock() == []
    notes = [r["Value"] for r in again.rows("00_README")
             if str(r["Key"]).startswith("shape_upgraded")]
    assert any("Financial_Trends" in str(n) for n in notes)
