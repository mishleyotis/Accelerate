#!/usr/bin/env python3
"""
Populate the DMA Research Workbook from the evidence index.

Usage:
    python populate_workbook.py <evidence_index.json> <diagnostic_questions.json> \
        --entity "Entity Name" --subvertical "Credit Union" \
        --output /mnt/user-data/outputs/DMA_Research_Workbook_Entity_2025-03-15.xlsx

Creates a 10-sheet workbook with evidence columns (A-I, K-M, U-V) populated
and scoring columns (J, N-T) left empty for dma-assessment.
"""

import argparse
import json
import os
import sys
from datetime import datetime

# The refusal must not depend on the environment (2026-09-04): a retired
# writer that dies on a missing import reads as a crash rather than as the
# refusal it is. The legacy body below is unreachable, so None is enough.
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:                                  # pragma: no cover
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = get_column_letter = None


# Zennify brand colors
TEAL = 'FF27BBAF'
TEAL_DARK = 'FF1A8A80'
TEAL_LIGHT = 'FFE0F5F3'
WHITE = 'FFFFFFFF'
DARK_TEXT = 'FF1E1E1E'
MEDIUM_TEXT = 'FF4A4A4A'
BORDER_GRAY = 'FFCCCCCC'

# Column definitions for P#_Scoring_Detail sheets
COLUMNS = [
    ('A', 'Category_ID', 12),
    ('B', 'Category_Name', 30),
    ('C', 'Cap_ID', 8),
    ('D', 'Capability', 35),
    ('E', 'SubCap_ID', 14),
    ('F', 'SubCapability', 40),
    ('G', 'Tier', 12),
    ('H', 'Diagnostic_Question', 60),
    ('I', 'Weight_Pct', 12),
    ('J', 'Score_1_to_5', 14),       # EMPTY — for dma-assessment
    ('K', 'Evidence_IDs', 30),
    ('L', 'Evidence_URLs', 40),
    ('M', 'Evidence_Tier', 15),
    ('N', 'Confidence', 12),          # EMPTY
    ('O', 'Caps_Applied', 12),        # EMPTY
    ('P', 'Final_Score', 12),         # EMPTY
    ('Q', 'Prior_Score', 12),         # EMPTY
    ('R', 'Scoring_Rationale', 60),   # EMPTY
    ('S', 'Proof_Claims', 40),        # EMPTY
    ('T', 'Proof_Links', 30),         # EMPTY
    ('U', 'Evidence_Excerpt', 80),
    ('V', 'Source_Document', 40),
]

EMPTY_COLUMNS = {'J', 'N', 'O', 'P', 'Q', 'R', 'S', 'T'}

# Pillar names
PILLAR_NAMES = {
    'P1': 'Strategy, Governance & Culture',
    'P2': 'Member/Customer Experience',
    'P3': 'Operations, Risk & Compliance',
    'P4': 'Data, Analytics & Technology',
}

PILLAR_SHEETS = {
    'P1': 'P1_Scoring_Detail',
    'P2': 'P2_Scoring_Detail',
    'P3': 'P3_Scoring_Detail',
    'P4': 'P4_Scoring_Detail',
}


def create_styles():
    """Create reusable styles."""
    header_font = Font(name='DM Sans', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type='solid')
    body_font = Font(name='DM Sans', size=10, color=DARK_TEXT)
    alt_fill = PatternFill(start_color=TEAL_LIGHT, end_color=TEAL_LIGHT, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color=BORDER_GRAY),
        right=Side(style='thin', color=BORDER_GRAY),
        top=Side(style='thin', color=BORDER_GRAY),
        bottom=Side(style='thin', color=BORDER_GRAY),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'body_font': body_font,
        'alt_fill': alt_fill,
        'thin_border': thin_border,
        'wrap_alignment': wrap_alignment,
    }


def create_pillar_sheet(wb, pillar_id, subcaps, evidence_index, styles):
    """Create a P#_Scoring_Detail sheet with evidence populated."""
    sheet_name = PILLAR_SHEETS[pillar_id]
    ws = wb.create_sheet(sheet_name)
    
    # Write headers
    for col_idx, (letter, name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row and subcap ID columns (A-F)
    ws.freeze_panes = 'G2'
    
    # Write subcap rows
    evidence_map = evidence_index.get('subcap_coverage', {})
    evidence_items = {e['evidence_id']: e for e in evidence_index.get('evidence_items', [])}
    
    row_idx = 2
    for sc in subcaps:
        subcap_id = sc['subcap_id']
        coverage = evidence_map.get(subcap_id, {})
        
        # Collect evidence for this subcap
        fact_ids = coverage.get('fact_ids', [])
        e_ids = coverage.get('evidence_ids', [])
        
        # Build evidence columns
        evidence_id_str = ', '.join(fact_ids) if fact_ids else (', '.join(e_ids) if e_ids else 'NO_EVIDENCE')
        
        # Get URLs and tiers from evidence items
        urls = []
        tiers = set()
        excerpts = []
        sources = []
        highest_tier = 'NO_EVIDENCE'
        
        for e_id in e_ids:
            item = evidence_items.get(e_id, {})
            if item:
                if item.get('source_url'):
                    urls.append(item['source_url'])
                tiers.add(item.get('tier', ''))
                sources.append(item.get('source_name', ''))
                # Get the highest-ERS fact for the excerpt
                for fact in item.get('facts', []):
                    if subcap_id in fact.get('subcap_mappings', []):
                        excerpts.append(fact.get('workbook_column_u_text', fact.get('fact_text', '')))
        
        # Determine highest tier
        tier_order = ['T1', 'T2', 'T3', 'T4', 'T5']
        for t in tier_order:
            if t in tiers:
                highest_tier = t
                break
        
        if not e_ids:
            highest_tier = 'NO_EVIDENCE'
            excerpt_text = f"No evidence identified through {coverage.get('searches_executed', 0)} searches targeting: {sc.get('diagnostic_question', 'N/A')}"
            source_text = ''
        else:
            excerpt_text = '; '.join(excerpts[:2]) if excerpts else ''
            source_text = '; '.join(sources[:3]) if sources else ''
        
        # Write row data
        row_data = {
            'A': sc.get('category_id', pillar_id),
            'B': PILLAR_NAMES.get(pillar_id, sc.get('category', '')),
            'C': sc.get('cap_id', ''),
            'D': sc.get('capability', ''),
            'E': subcap_id,
            'F': sc.get('subcap_name', ''),
            'G': highest_tier,
            'H': sc.get('diagnostic_question', ''),
            'I': sc.get('weight', ''),
            'J': '',  # EMPTY — Score
            'K': evidence_id_str,
            'L': '; '.join(urls[:3]),
            'M': ', '.join(sorted(tiers)) if tiers else 'NO_EVIDENCE',
            'N': '',  # EMPTY — Confidence
            'O': '',  # EMPTY — Caps_Applied
            'P': '',  # EMPTY — Final_Score
            'Q': '',  # EMPTY — Prior_Score
            'R': '',  # EMPTY — Scoring_Rationale
            'S': '',  # EMPTY — Proof_Claims
            'T': '',  # EMPTY — Proof_Links
            'U': excerpt_text,
            'V': source_text,
        }
        
        for col_idx, (letter, name, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(letter, ''))
            cell.font = styles['body_font']
            cell.border = styles['thin_border']
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = styles['alt_fill']
            
            # Wrap text for wide columns
            if width >= 40:
                cell.alignment = styles['wrap_alignment']
            
            # Mark empty scoring columns with gray
            if letter in EMPTY_COLUMNS:
                cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
        
        row_idx += 1
    
    return row_idx - 2  # return row count


def create_skeleton_sheets(wb, entity_name, subvertical, styles):
    """Create skeleton sheets that dma-assessment will populate."""
    
    # Summary sheet (skeleton)
    ws = wb.create_sheet('Summary', 0)
    headers = ['Pillar', 'Score', 'Level', 'Peer_Median', 'vs_Median', 'vs_P25', 'vs_P75',
               'Trend', 'Evidence_Coverage_Pct', 'Confidence', 'Key_Findings']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
    for row_idx, pillar in enumerate(['P1', 'P2', 'P3', 'P4', 'Overall'], 2):
        ws.cell(row=row_idx, column=1, value=pillar)
    
    # Calculation_Chain (skeleton)
    ws = wb.create_sheet('Calculation_Chain')
    ws.cell(row=1, column=1, value='— Populated by dma-assessment skill —')
    
    # Evidence_Linkage_Matrix (skeleton — will be filled from evidence index)
    ws = wb.create_sheet('Evidence_Linkage_Matrix')
    elm_headers = ['Evidence_ID', 'Source', 'Tier', 'ERS', 'Date', 'Subcap_Mappings', 'URL']
    for col_idx, header in enumerate(elm_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
    
    # Caps_Applied_Log (skeleton)
    ws = wb.create_sheet('Caps_Applied_Log')
    ws.cell(row=1, column=1, value='— Populated by dma-assessment skill —')
    
    # Absent_Evidence_Log (skeleton — research skill could pre-populate)
    ws = wb.create_sheet('Absent_Evidence_Log')
    absent_headers = ['SubCap_ID', 'SubCapability', 'Diagnostic_Question', 'Searches_Executed',
                      'Max_Tier_Reached', 'Reason', 'Discovery_Questions', 'Impact']
    for col_idx, header in enumerate(absent_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
    
    # QA_Validation_Log (skeleton)
    ws = wb.create_sheet('QA_Validation_Log')
    ws.cell(row=1, column=1, value='— Populated by dma-assessment skill —')


RETIRED = """REFUSED: populate_workbook.py is retired (2026-09-03). The scoring workbook has
ONE writer — the research engine — and this script built a SECOND, 10-sheet
workbook beside it from a JSON plane the engine never wrote, which is the
"workbook defaults to the wrong structure every run" defect (owner issue 3;
goeasy GSY-15/21). Use the engine:

    python3 -m engine.cli start    --run R --root ROOT --entity … --entity-id … \\
                                   --reference-date … --preflight preflight.json
    python3 -m engine.cli evidence --run R --root ROOT --subcap P1C1.1.1 …
    python3 -m engine.cli synthesise / absence / gate …

The workbook it creates IS the pinned template's shape (engine/contract.py,
40 sheets, SubCap_Name seeded from the catalogue), formatted, validated and
bound to the report templates. Nothing else may write it.
"""


def main():
    import sys as _sys
    _sys.stderr.write(RETIRED)
    return 1


def _legacy_main():           # kept for reference; unreachable
    parser = argparse.ArgumentParser(description='Populate DMA Research Workbook')
    parser.add_argument('evidence_index', help='Path to evidence_index.json')
    parser.add_argument('diagnostic_questions', help='Path to diagnostic_questions.json')
    parser.add_argument('--entity', required=True, help='Entity name')
    parser.add_argument('--subvertical', required=True, help='Subvertical classification')
    parser.add_argument('--output', '-o', default=None, help='Output XLSX path')
    args = parser.parse_args()
    
    # Load data
    with open(args.evidence_index) as f:
        evidence_index = json.load(f)
    
    with open(args.diagnostic_questions) as f:
        dq_data = json.load(f)
    
    subcaps_flat = dq_data.get('subcaps_flat', [])
    
    # Group subcaps by pillar
    pillar_subcaps = {'P1': [], 'P2': [], 'P3': [], 'P4': []}
    for sc in subcaps_flat:
        pillar = sc['subcap_id'][:2] if sc['subcap_id'].startswith('P') else ''
        if pillar in pillar_subcaps:
            pillar_subcaps[pillar].append(sc)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    styles = create_styles()
    
    # Create skeleton sheets first
    create_skeleton_sheets(wb, args.entity, args.subvertical, styles)
    
    # Create pillar sheets with evidence
    total_rows = 0
    for pillar_id in ['P1', 'P2', 'P3', 'P4']:
        subcaps = pillar_subcaps[pillar_id]
        if subcaps:
            rows = create_pillar_sheet(wb, pillar_id, subcaps, evidence_index, styles)
            total_rows += rows
            print(f"  {PILLAR_SHEETS[pillar_id]}: {rows} rows")
        else:
            print(f"  WARNING: No subcaps for {pillar_id}")
    
    # Output path
    date_str = datetime.now().strftime('%Y-%m-%d')
    entity_clean = args.entity.replace(' ', '_').replace('/', '_')
    output_path = args.output or f'/mnt/user-data/outputs/DMA_Research_Workbook_{entity_clean}_{date_str}.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"WORKBOOK GENERATED: {output_path}")
    print(f"TOTAL SUBCAP ROWS: {total_rows}")
    print(f"EVIDENCE COLUMNS: Populated (A-I, K-M, U-V)")
    print(f"SCORING COLUMNS: Empty (J, N-T) — for dma-assessment")
    print(f"{'='*60}")


if __name__ == '__main__':
    import sys as _sys
    _sys.exit(main())
