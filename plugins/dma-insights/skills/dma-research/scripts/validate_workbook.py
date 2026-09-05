#!/usr/bin/env python3
"""
Validate DMA Research Workbook against quality schema.

Usage:
    python validate_workbook.py <workbook.xlsx> [--schema <schema.json>] [--report <output.json>]

Checks:
  - Structural: sheets exist, columns correct, row counts within tolerance
  - Evidence: columns K/L/M/U/V populated correctly for evidence rows
  - Empty: scoring columns J/N-T must be empty
  - Quality: ERS annotations, ceiling estimates, claim labels in Column U
  - Cross-reference: ELM completeness, Absent_Evidence_Log entries
"""

import argparse
import json
import os
import re
import sys

# The refusal must not depend on the environment (2026-09-04): a retired
# writer that dies on a missing import reads as a crash rather than as the
# refusal it is. The legacy body below is unreachable, so None is enough.
try:
    import openpyxl
except ImportError:                                  # pragma: no cover
    openpyxl = None


# Default quality thresholds
DEFAULT_THRESHOLDS = {
    'row_count_tolerance_pct': 5,
    'min_excerpt_length_evidence': 50,
    'min_excerpt_length_no_evidence': 100,
    'min_coverage_pct': 80,
    'max_no_evidence_pct': 20,
    'min_t1_t2_pct': 20,
    'max_t5_pct': 30,
    'min_mean_ers': 2.5,
    'min_pct_with_ceiling': 70,
    'min_pct_with_claim_label': 90,
    'min_no_evidence_search_count': 6
}

ROW_TARGETS = {
    'P1_Scoring_Detail': 203,
    'P2_Scoring_Detail': 291,
    'P3_Scoring_Detail': 164,
    'P4_Scoring_Detail': 189
}

REQUIRED_SHEETS = [
    'Summary', 'Calculation_Chain',
    'P1_Scoring_Detail', 'P2_Scoring_Detail', 'P3_Scoring_Detail', 'P4_Scoring_Detail',
    'Evidence_Linkage_Matrix', 'Caps_Applied_Log', 'Absent_Evidence_Log', 'QA_Validation_Log'
]

REQUIRED_COLUMNS = [
    'Category_ID', 'Category_Name', 'Cap_ID', 'Capability',
    'SubCap_ID', 'SubCapability', 'Tier', 'Diagnostic_Question',
    'Weight_Pct', 'Score_1_to_5', 'Evidence_IDs', 'Evidence_URLs',
    'Evidence_Tier', 'Confidence', 'Caps_Applied', 'Final_Score',
    'Prior_Score', 'Scoring_Rationale', 'Proof_Claims',
    'Proof_Links', 'Evidence_Excerpt', 'Source_Document'
]

SCORING_COLUMNS = ['Score_1_to_5', 'Confidence', 'Caps_Applied', 'Final_Score',
                   'Prior_Score', 'Scoring_Rationale', 'Proof_Claims', 'Proof_Links']

VALID_TIERS = {'T1', 'T2', 'T3', 'T4', 'T5', 'NO_EVIDENCE'}

EVIDENCE_ID_PATTERN = re.compile(r'^E-\d{3}(:F\d+)?$')


class ValidationResult:
    def __init__(self):
        self.checks = []
        self.blocks = 0
        self.warnings = 0
        self.passes = 0

    def add(self, check_id, name, status, details, severity='WARNING'):
        self.checks.append({
            'check_id': check_id,
            'name': name,
            'status': status,
            'details': details,
            'severity': severity
        })
        if status == 'FAIL':
            if severity == 'BLOCK':
                self.blocks += 1
            else:
                self.warnings += 1
        else:
            self.passes += 1

    def summary(self):
        return {
            'total_checks': len(self.checks),
            'passed': self.passes,
            'warnings': self.warnings,
            'blocks': self.blocks,
            'verdict': 'BLOCK' if self.blocks > 0 else ('WARNING' if self.warnings > 0 else 'PASS')
        }


def validate_sheets(wb, results):
    """Check all required sheets exist."""
    existing = set(wb.sheetnames)
    for sheet in REQUIRED_SHEETS:
        if sheet in existing:
            results.add(f'SHEET-{sheet}', f'Sheet exists: {sheet}', 'PASS', '', 'BLOCK')
        else:
            results.add(f'SHEET-{sheet}', f'Sheet exists: {sheet}', 'FAIL',
                        f'Missing sheet: {sheet}', 'BLOCK')


def validate_columns(ws, sheet_name, results):
    """Check all 22 columns exist with correct headers."""
    headers = [cell.value for cell in ws[1]]
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        results.add(f'COL-{sheet_name}', f'Columns in {sheet_name}', 'FAIL',
                     f'Missing columns: {", ".join(missing)}', 'BLOCK')
    else:
        results.add(f'COL-{sheet_name}', f'Columns in {sheet_name}', 'PASS',
                     f'All 22 columns present', 'BLOCK')
    return {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}


def validate_row_counts(ws, sheet_name, results, thresholds):
    """Check row count within tolerance."""
    target = ROW_TARGETS.get(sheet_name)
    if not target:
        return 0
    actual = ws.max_row - 1  # subtract header
    tolerance = target * thresholds['row_count_tolerance_pct'] / 100
    if abs(actual - target) <= tolerance:
        results.add(f'ROWS-{sheet_name}', f'Row count {sheet_name}', 'PASS',
                     f'{actual} rows (target {target} ±{thresholds["row_count_tolerance_pct"]}%)', 'BLOCK')
    else:
        results.add(f'ROWS-{sheet_name}', f'Row count {sheet_name}', 'FAIL',
                     f'{actual} rows (target {target} ±{thresholds["row_count_tolerance_pct"]}%)', 'BLOCK')
    return actual


def validate_scoring_columns_empty(ws, col_map, sheet_name, results):
    """Verify scoring columns are empty."""
    violations = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_name in SCORING_COLUMNS:
            if col_name in col_map:
                cell_value = row[col_map[col_name]].value
                if cell_value is not None and str(cell_value).strip():
                    violations += 1
    if violations == 0:
        results.add(f'EMPTY-{sheet_name}', f'Scoring columns empty: {sheet_name}', 'PASS',
                     'All scoring columns (J, N-T) are empty', 'BLOCK')
    else:
        results.add(f'EMPTY-{sheet_name}', f'Scoring columns empty: {sheet_name}', 'FAIL',
                     f'{violations} cells in scoring columns have values', 'BLOCK')


def validate_evidence_quality(ws, col_map, sheet_name, results, thresholds):
    """Validate evidence column quality."""
    stats = {
        'total_rows': 0, 'with_evidence': 0, 'no_evidence': 0,
        'tiers': {'T1': 0, 'T2': 0, 'T3': 0, 'T4': 0, 'T5': 0, 'NO_EVIDENCE': 0},
        'excerpt_lengths': [], 'no_ev_excerpt_lengths': [],
        'has_ceiling': 0, 'has_claim_label': 0, 'has_ers': 0,
        'missing_excerpt': 0, 'missing_source': 0, 'invalid_tier': 0
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if not row[0].value:  # skip empty rows
            continue
        stats['total_rows'] += 1

        tier = str(row[col_map.get('Tier', 6)].value or '').strip()
        excerpt = str(row[col_map.get('Evidence_Excerpt', 20)].value or '').strip()
        source = str(row[col_map.get('Source_Document', 21)].value or '').strip()
        ev_ids = str(row[col_map.get('Evidence_IDs', 10)].value or '').strip()

        # Tier validation
        if tier in VALID_TIERS:
            stats['tiers'][tier] = stats['tiers'].get(tier, 0) + 1
        else:
            stats['invalid_tier'] += 1

        if tier == 'NO_EVIDENCE':
            stats['no_evidence'] += 1
            if excerpt:
                stats['no_ev_excerpt_lengths'].append(len(excerpt))
            else:
                stats['missing_excerpt'] += 1
        else:
            stats['with_evidence'] += 1
            if excerpt:
                stats['excerpt_lengths'].append(len(excerpt))
                if 'CEILING' in excerpt:
                    stats['has_ceiling'] += 1
                if any(label in excerpt for label in ['FACT', 'INFERENCE', 'HYPOTHESIS', 'CEILING_ESTIMATE']):
                    stats['has_claim_label'] += 1
                if 'ERS' in excerpt:
                    stats['has_ers'] += 1
            else:
                stats['missing_excerpt'] += 1
            if not source:
                stats['missing_source'] += 1

    # Report coverage
    if stats['total_rows'] > 0:
        coverage = stats['with_evidence'] / stats['total_rows'] * 100
        results.add(f'COV-{sheet_name}', f'Evidence coverage: {sheet_name}', 
                     'PASS' if coverage >= thresholds['min_coverage_pct'] else 'FAIL',
                     f'{coverage:.1f}% ({stats["with_evidence"]}/{stats["total_rows"]})',
                     'BLOCK' if coverage < 50 else 'WARNING')

    # Report excerpt completeness
    results.add(f'EXCPT-{sheet_name}', f'Excerpt completeness: {sheet_name}',
                'PASS' if stats['missing_excerpt'] == 0 else 'FAIL',
                f'{stats["missing_excerpt"]} rows missing excerpts', 'BLOCK')

    # Report quality indicators
    if stats['with_evidence'] > 0:
        ceiling_pct = stats['has_ceiling'] / stats['with_evidence'] * 100
        claim_pct = stats['has_claim_label'] / stats['with_evidence'] * 100
        ers_pct = stats['has_ers'] / stats['with_evidence'] * 100
        results.add(f'QUAL-{sheet_name}', f'Write-up quality: {sheet_name}', 
                     'PASS' if claim_pct >= thresholds['min_pct_with_claim_label'] else 'FAIL',
                     f'Ceiling: {ceiling_pct:.0f}%, Claim labels: {claim_pct:.0f}%, ERS: {ers_pct:.0f}%',
                     'WARNING')

    return stats


def validate_workbook(workbook_path, schema_path=None, thresholds=None):
    """Main validation function."""
    if thresholds is None:
        thresholds = DEFAULT_THRESHOLDS

    results = ValidationResult()
    wb = openpyxl.load_workbook(workbook_path, read_only=True)

    # 1. Sheet validation
    validate_sheets(wb, results)

    # 2. Per-sheet validation
    all_stats = {}
    total_rows = 0
    for sheet_name in ['P1_Scoring_Detail', 'P2_Scoring_Detail',
                        'P3_Scoring_Detail', 'P4_Scoring_Detail']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_map = validate_columns(ws, sheet_name, results)
        row_count = validate_row_counts(ws, sheet_name, results, thresholds)
        total_rows += row_count
        validate_scoring_columns_empty(ws, col_map, sheet_name, results)
        stats = validate_evidence_quality(ws, col_map, sheet_name, results, thresholds)
        all_stats[sheet_name] = stats

    # 3. Total row count
    target_total = sum(ROW_TARGETS.values())
    tolerance = target_total * thresholds['row_count_tolerance_pct'] / 100
    results.add('ROWS-TOTAL', 'Total row count',
                'PASS' if abs(total_rows - target_total) <= tolerance else 'FAIL',
                f'{total_rows} total (target {target_total})', 'BLOCK')

    # 4. Summary
    summary = results.summary()
    print(f"\n{'='*60}")
    print(f"WORKBOOK VALIDATION: {summary['verdict']}")
    print(f"{'='*60}")
    print(f"Checks: {summary['total_checks']} total")
    print(f"  PASS: {summary['passed']}")
    print(f"  WARNING: {summary['warnings']}")
    print(f"  BLOCK: {summary['blocks']}")
    print()
    for check in results.checks:
        icon = '✅' if check['status'] == 'PASS' else ('⚠️' if check['severity'] == 'WARNING' else '❌')
        print(f"  {icon} {check['check_id']}: {check['name']} — {check['details']}")

    return results


RETIRED = """REFUSED: validate_workbook.py is retired (2026-09-03). It validated the
22-column, P#_Scoring_Detail workbook the research engine replaced, so on the run's real
workbook (engine/contract.py: 41 sheets, P#_Subcap_Scoring with 33 columns) it fails every
structural check and on the retired populate_workbook.py's output it passes — the wrong
answer both ways. The run's ONE workbook is validated by:

    python3 -m engine.cli validate --run <RUN_ID> --root <ROOT>     # shape, vocabularies, rule 8
    python3 -m engine.cli complete check --run <RUN_ID> --root <ROOT>
    python3 -m engine.gold_standard workbook <workbook.xlsx>       # against the Golden 1 reference

This file stays so an old reference fails LOUD, naming the engine, rather than silently.
"""


def main():
    import sys as _sys
    _sys.stderr.write(RETIRED)
    return 1


def _legacy_main():           # kept for reference; unreachable
    parser = argparse.ArgumentParser(description='Validate DMA Research Workbook')
    parser.add_argument('workbook', help='Path to workbook XLSX file')
    parser.add_argument('--schema', help='Path to quality schema JSON', default=None)
    parser.add_argument('--report', help='Output validation report JSON', default=None)
    args = parser.parse_args()

    results = validate_workbook(args.workbook, args.schema)
    if args.report:
        with open(args.report, 'w') as f:
            json.dump({'summary': results.summary(), 'checks': results.checks}, f, indent=2)
        print(f"\nReport saved to: {args.report}")


if __name__ == '__main__':
    sys.exit(main())
