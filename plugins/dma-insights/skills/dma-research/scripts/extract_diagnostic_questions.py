#!/usr/bin/env python3
"""
Extract diagnostic questions from Pillar Scoring Toolkit XLSX files.

Usage:
    python extract_diagnostic_questions.py <pillar_xlsx_path> [--output <json_path>]
    python extract_diagnostic_questions.py --all <directory_with_pillar_files> [--output <json_path>]

Reads the Capability Map sheet from each Pillar XLSX and extracts:
- SubCap_ID, SubCapability name, Diagnostic Question, Weight, Capability, Category

Output: JSON file with all subcapabilities grouped by pillar → capability → subcap.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


def find_header_row(ws, target_keywords=None):
    """Find the header row by looking for known column names."""
    if target_keywords is None:
        target_keywords = ['subcap', 'diagnostic', 'question', 'capability', 'weight']
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        row_text = ' '.join([str(cell.value or '').lower() for cell in row])
        matches = sum(1 for kw in target_keywords if kw in row_text)
        if matches >= 2:
            return row_idx
    return 1  # default to row 1


def find_column_index(header_row, keywords):
    """Find column index by matching header text against keywords."""
    for idx, cell in enumerate(header_row):
        cell_text = str(cell.value or '').lower().strip()
        for kw in keywords:
            if kw in cell_text:
                return idx
    return None


def extract_from_pillar(xlsx_path):
    """Extract all subcapabilities and diagnostic questions from a single Pillar XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    # Find the capability map sheet (try common names)
    sheet_names = wb.sheetnames
    target_sheet = None
    for name in sheet_names:
        name_lower = name.lower().replace(' ', '').replace('_', '')
        if any(kw in name_lower for kw in ['capabilitymap', 'capmap', 'subcap', 'scoring']):
            target_sheet = name
            break
    
    if not target_sheet:
        # Try the first sheet or the one with the most rows
        target_sheet = sheet_names[0]
        print(f"  WARNING: No 'Capability Map' sheet found. Using '{target_sheet}'")
    
    ws = wb[target_sheet]
    
    # Find header row
    header_row_idx = find_header_row(ws)
    header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]
    
    # Map columns
    col_map = {
        'subcap_id': find_column_index(header_row, ['subcap_id', 'subcapability_id', 'sub_cap', 'id']),
        'subcap_name': find_column_index(header_row, ['subcapability', 'subcap_name', 'sub_capability']),
        'capability': find_column_index(header_row, ['capability', 'cap_name']),
        'cap_id': find_column_index(header_row, ['cap_id', 'capability_id']),
        'category': find_column_index(header_row, ['category', 'pillar', 'category_name']),
        'category_id': find_column_index(header_row, ['category_id', 'pillar_id']),
        'diagnostic_q': find_column_index(header_row, ['diagnostic', 'question', 'assessment_question']),
        'weight': find_column_index(header_row, ['weight', 'weight_pct', 'weighting']),
    }
    
    # Report what we found
    found = {k: v for k, v in col_map.items() if v is not None}
    missing = {k: v for k, v in col_map.items() if v is None}
    print(f"  Sheet: {target_sheet}, Header row: {header_row_idx}")
    print(f"  Columns found: {list(found.keys())}")
    if missing:
        print(f"  Columns NOT found (will use defaults): {list(missing.keys())}")
    
    # Extract rows
    subcaps = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        values = [cell.value for cell in row]
        
        # Get subcap_id — skip empty rows
        subcap_id_idx = col_map.get('subcap_id', 0)
        subcap_id = values[subcap_id_idx] if subcap_id_idx is not None and subcap_id_idx < len(values) else None
        
        if not subcap_id or str(subcap_id).strip() == '':
            continue
        
        subcap_id = str(subcap_id).strip()
        
        # Extract all fields with safe indexing
        def safe_get(key, default=''):
            idx = col_map.get(key)
            if idx is not None and idx < len(values) and values[idx] is not None:
                return str(values[idx]).strip()
            return default
        
        subcap = {
            'subcap_id': subcap_id,
            'subcap_name': safe_get('subcap_name'),
            'capability': safe_get('capability'),
            'cap_id': safe_get('cap_id'),
            'category': safe_get('category'),
            'category_id': safe_get('category_id'),
            'diagnostic_question': safe_get('diagnostic_q'),
            'weight': safe_get('weight', '0'),
        }
        
        # Try to parse weight as number
        try:
            subcap['weight_numeric'] = float(subcap['weight'].replace('%', ''))
        except (ValueError, AttributeError):
            subcap['weight_numeric'] = 0.0
        
        subcaps.append(subcap)
    
    wb.close()
    return subcaps


def extract_all_pillars(directory):
    """Extract from all Pillar XLSX files in a directory."""
    all_subcaps = []
    xlsx_files = sorted(Path(directory).glob('*.xlsx'))
    
    if not xlsx_files:
        xlsx_files = sorted(Path(directory).glob('**/*.xlsx'))
    
    pillar_files = [f for f in xlsx_files if any(kw in f.name.lower() for kw in 
                    ['pillar', 'p1', 'p2', 'p3', 'p4', 'scoring', 'toolkit'])]
    
    if not pillar_files:
        pillar_files = xlsx_files  # fall back to all XLSX files
    
    for xlsx_path in pillar_files:
        print(f"\nProcessing: {xlsx_path.name}")
        try:
            subcaps = extract_from_pillar(str(xlsx_path))
            print(f"  Extracted: {len(subcaps)} subcapabilities")
            all_subcaps.extend(subcaps)
        except Exception as e:
            print(f"  ERROR: {e}")
    
    return all_subcaps


def organize_by_hierarchy(subcaps):
    """Organize flat list into pillar → capability → subcap hierarchy."""
    hierarchy = {}
    for sc in subcaps:
        # Infer pillar from subcap_id (e.g., P1C1.1.1 → P1)
        pillar = ''
        if sc['subcap_id'].startswith('P'):
            pillar = sc['subcap_id'][:2]
        elif sc['category_id']:
            pillar = sc['category_id'][:2] if sc['category_id'].startswith('P') else ''
        
        # Infer capability from subcap_id (e.g., P1C1.1.1 → P1C1)
        cap_id = ''
        parts = sc['subcap_id'].split('.')
        if len(parts) >= 1:
            cap_id = parts[0]  # P1C1
        
        if pillar not in hierarchy:
            hierarchy[pillar] = {'capabilities': {}, 'subcap_count': 0}
        
        if cap_id not in hierarchy[pillar]['capabilities']:
            hierarchy[pillar]['capabilities'][cap_id] = {
                'name': sc['capability'],
                'subcaps': []
            }
        
        hierarchy[pillar]['capabilities'][cap_id]['subcaps'].append(sc)
        hierarchy[pillar]['subcap_count'] += 1
    
    return hierarchy


def main():
    parser = argparse.ArgumentParser(description='Extract diagnostic questions from Pillar XLSX files')
    parser.add_argument('path', help='Path to single XLSX file or directory')
    parser.add_argument('--all', action='store_true', help='Process all XLSX files in directory')
    parser.add_argument('--output', '-o', default=None, help='Output JSON path')
    parser.add_argument('--format', choices=['flat', 'hierarchy', 'both'], default='both',
                       help='Output format')
    args = parser.parse_args()
    
    path = Path(args.path)
    
    if path.is_dir() or args.all:
        subcaps = extract_all_pillars(str(path))
    elif path.is_file():
        print(f"Processing: {path.name}")
        subcaps = extract_from_pillar(str(path))
    else:
        print(f"ERROR: {path} not found")
        sys.exit(1)
    
    # Summary
    print(f"\n{'='*60}")
    print(f"TOTAL SUBCAPABILITIES EXTRACTED: {len(subcaps)}")
    with_questions = sum(1 for sc in subcaps if sc['diagnostic_question'])
    print(f"WITH DIAGNOSTIC QUESTIONS: {with_questions} ({100*with_questions/max(len(subcaps),1):.0f}%)")
    without = len(subcaps) - with_questions
    if without > 0:
        print(f"WITHOUT QUESTIONS: {without} — these will need manual query generation")
    
    # Organize
    hierarchy = organize_by_hierarchy(subcaps)
    for pillar, data in sorted(hierarchy.items()):
        cap_count = len(data['capabilities'])
        print(f"  {pillar}: {data['subcap_count']} subcaps across {cap_count} capabilities")
    
    # Output
    output_path = args.output or '/home/claude/dma_checkpoints/diagnostic_questions.json'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    output = {
        'metadata': {
            'total_subcaps': len(subcaps),
            'with_diagnostic_questions': with_questions,
            'source_files': [str(path)]
        }
    }
    
    if args.format in ('flat', 'both'):
        output['subcaps_flat'] = subcaps
    if args.format in ('hierarchy', 'both'):
        output['subcaps_hierarchy'] = hierarchy
    
    with open(output_path, 'w') as f:
        json.dump(output, f, indent=2, default=str)
    
    print(f"\nOutput saved to: {output_path}")
    return subcaps


if __name__ == '__main__':
    main()
