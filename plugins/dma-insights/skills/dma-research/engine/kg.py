#!/usr/bin/env python3
"""The knowledge graph: every DQ mapped, mode-filtered, routed by category.

    python3 -m engine.kg build --run R --toolkits DIR [--mode PUBLIC]
    python3 -m engine.kg route --run R [--category P1C1]
    python3 -m engine.kg show  --run R --subcap P1C1.1.1
    python3 -m engine.kg verify --run R

WHAT THE GRAPH IS. Nodes are the catalogue's own grains — pillar → category
→ capability → subcap — plus one node per DIAGNOSTIC QUESTION. Every subcap
carries nine DQs:

    order 0        `primary`  the Pillar Scoring Toolkit's own question for
                              this subcap (column H), regraded to an open
                              form where the toolkit wrote it closed
    orders 1..5    the five facet probes: works · fails · value ·
                   contradicts · corroborates
    orders 6..8    the AI overlay: ai_deployment · ai_data · ai_constraint —
                   measured from the pinned workbook's own DQ_Bank, whose
                   question set is 851 x 5 + 851 x 3 exactly, so every
                   subcap carries all three

Each DQ node carries `mode_fit` (PUBLIC | INTERNAL | BOTH) and the toolkit's
own source lists — `Internal Evidence Sources` and `Public / External
Evidence Sources`, the columns that name, per subcap, exactly which
artefacts answer the question. That is the highest-value routing signal the
workbook ecosystem holds: a researcher told WHAT to look for stops fishing.

MODE FILTERING IS DISCLOSURE, NOT DELETION. A run declares its evidence
mode (PUBLIC / INTERNAL / HYBRID) at start. The router serves only the DQs
answerable in that mode; the rest are returned as DEFERRED questions with
the reason and the discovery form (`INT-Q:` for an internal-only question in
a public run, `PUB-Q:` for the reverse) — they land in the synthesis's
Discovery_Questions and the report's validation needs, never in a silent
gap. A question nobody could answer is a finding; a question nobody asked
is a defect.

THE WORKBOOK STAYS THE SUBSTRATE. `build` seeds the workbook's own DQ_Bank
sheet — orient's work cards and every downstream reader consume the SHEET.
The kg.json this module also writes is a projection for fast routing, its
checksum recorded in Run_Metadata.kg_checksum so drift between the two is
detectable (`verify`), and a run whose KG was never built says so at resume.
"""
from __future__ import annotations

# Runnable both ways. `python3 -m engine.<mod>` is the documented invocation,
# but every audit and every operator reaches for `python3 <path> --help`
# first, and a relative import dies there. Binding __package__ makes the two
# equivalent instead of making one of them a trap.
if __package__ in (None, ""):  # noqa: E402  (must precede the relative imports)
    import os as _os
    import sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.dirname(
        _os.path.abspath(__file__))))
    __package__ = "engine"

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

from . import contract as C
from . import runstate
from .workbook import RunWorkbook, WorkbookError

#: Toolkit workbook per pillar, and the sheet each sub-vertical reads.
#: Sheet names measured from Pillar1_Scoring_Toolkit.xlsx as shipped.
#: FC -> "Lending" is an ADJUDICATION, not a lookup: the toolkits carry
#: eight sub-vertical sheets for the catalogue's nine codes, and Farm
#: Credit institutions are lenders — if a later toolkit ships a "Farm
#: Credit" sheet, the exact name below wins because it is checked first.
TOOLKIT_FILES = {p: f"Pillar{p[1]}_Scoring_Toolkit.xlsx"
                 for p in ("P1", "P2", "P3", "P4")}
TOOLKIT_SHEETS = {
    "CU": ("Farm Credit", "Credit Unions"),   # exact-name-first probe order
    "RB": ("Regional Banks",),
    "CL": ("Lending",),
    "FC": ("Farm Credit", "Lending"),
    "CIB": ("CIB",),
    "RIA": ("RIAs & Broker-Dealers",),
    "AM": ("Asset Management",),
    "IC": ("Insurance Carriers",),
    "IB": ("Insurance Brokers",),
}
# CU listed with a Farm Credit probe by mistake would be wrong — fix: CU
# reads Credit Unions only.
TOOLKIT_SHEETS["CU"] = ("Credit Unions",)

_MODE_NORM = {
    "both": "BOTH", "public only": "PUBLIC", "public": "PUBLIC",
    "internal only": "INTERNAL", "internal": "INTERNAL",
    "external": "PUBLIC", "public / external": "PUBLIC",
}

#: Closed-question openers and their graded rewrites — the conversion table
#: from references/diagnostic_questions.md § Grading a closed pattern. A
#: closed question cannot produce a graded answer, and a five-level scale
#: needs one that can.
_GRADED = (
    (re.compile(r"^\s*does\s+(.*?)\s+have\s+(.+?)\??$", re.I),
     r"To what extent is \2 established at \1, and since when?"),
    (re.compile(r"^\s*is\s+there\s+(?:a\s+|an\s+)?(.+?)\??$", re.I),
     r"How far does \1 reach, who owns it, and how is it reviewed?"),
    (re.compile(r"^\s*is\s+(.+?)\s+(documented|defined|established|in place)\??$", re.I),
     r"How far does \1's documentation reach, and who owns it?"),
    (re.compile(r"^\s*are\s+(.+?)\s+(defined|documented|measured|established)\??$", re.I),
     r"Which \1 are \2, by whom, and how are they reviewed?"),
    (re.compile(r"^\s*can\s+(.+?)\s+(.+?)\??$", re.I),
     r"Where has \1 demonstrated \2, and where has it not held?"),
    (re.compile(r"^\s*(?:has|have)\s+(.+?)\s+(.+?)\??$", re.I),
     r"What is the arc of \2 at \1 — earliest signal, refreshes, stalls?"),
)


def grade_question(q: str) -> str:
    """The open form of a toolkit question, or the question unchanged.

    Only the leading token decides closed-ness (the G7 lesson: a closed
    question wearing an open opener passes a first-word test — so this
    REWRITES rather than tests, and an already-open question is left alone)."""
    text = " ".join(str(q or "").split())
    if not text:
        return text
    if re.match(r"^(Where|What|How|Which|Who|To what extent|Describe)\b",
                text, re.I):
        return text
    for rx, sub in _GRADED:
        m = rx.match(text)
        if m:
            return rx.sub(sub, text).rstrip("?") + "?"
    return text


# ── reading the toolkits ─────────────────────────────────────────────────

def load_toolkits(toolkit_dir, sub_vertical: str | None) -> tuple[dict, list]:
    """{subcap_id: toolkit row} for one sub-vertical, plus problems.

    Problems are returned, not raised: a missing pillar file costs that
    pillar's primary questions and source lists (the facet probes still
    build), and the build REPORTS the degradation instead of failing the
    whole run or, worse, proceeding silently."""
    import openpyxl
    toolkit_dir = Path(toolkit_dir)
    sv = sub_vertical or "RB"           # universal default: the widest sheet
    sheet_names = TOOLKIT_SHEETS.get(sv)
    problems: list[str] = []
    if sheet_names is None:
        return {}, [f"no toolkit sheet mapping for sub-vertical {sv!r}"]
    out: dict = {}
    for pillar, fname in TOOLKIT_FILES.items():
        path = toolkit_dir / fname
        if not path.is_file():
            problems.append(f"{fname} not found in {toolkit_dir} — {pillar}'s "
                            f"primary questions and source lists are absent")
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = next((s for s in sheet_names if s in wb.sheetnames), None)
            if sheet is None:
                problems.append(
                    f"{fname}: none of {sheet_names} among its sheets "
                    f"{wb.sheetnames[:10]}")
                continue
            ws = wb[sheet]
            hdr_row = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8,
                                                 values_only=True), 1):
                cells = [str(v).strip().lower() for v in row if v]
                if any("sub-cap id" in c or "subcap id" in c for c in cells):
                    hdr_row = i
                    hdr = {str(v).strip().lower(): j for j, v in enumerate(row)
                           if v}
                    break
            if hdr_row is None:
                problems.append(f"{fname}/{sheet}: no header row naming "
                                f"'Sub-Cap ID' in the first 8 rows")
                continue

            def col(*names):
                for n in names:
                    for k, j in hdr.items():
                        if n in k:
                            return j
                return None

            c_id = col("sub-cap id", "subcap id")
            c_name = col("sub-capability", "subcapability")
            c_q = col("diagnostic question")
            c_int = col("internal evidence")
            c_pub = col("public / external", "public/external",
                        "external evidence", "public evidence")
            c_type = col("source type")
            c_w = col("weight")
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                sid = str(row[c_id]).strip() if c_id is not None and \
                    len(row) > c_id and row[c_id] else ""
                if not re.match(r"^P\d+C\d+\.", sid):
                    continue

                def v(j):
                    return (str(row[j]).strip()
                            if j is not None and len(row) > j and row[j]
                            else None)
                raw_type = (v(c_type) or "both").strip().lower()
                out[sid] = {
                    "name": v(c_name),
                    "question": v(c_q),
                    "internal_sources": v(c_int),
                    "public_sources": v(c_pub),
                    "mode_fit": _MODE_NORM.get(raw_type, "BOTH"),
                    "weight_pct": v(c_w),
                    "sheet": f"{fname}/{sheet}",
                }
        finally:
            wb.close()
    return out, problems


# ── the facet probes ─────────────────────────────────────────────────────
#
# Templated per subcap; {entity} stays UNBOUND here on purpose — orient
# binds it at card time, and a query fired with the token still in it is
# refused at the ledger (the AUD-0015 discipline).

def _facet_dq(facet: str, subcap_name: str, primary: str | None) -> str:
    n = subcap_name or "this capability"
    table = {
        "works": (f"Where has {{entity}} demonstrated {n}, from the earliest "
                  f"signal through refreshes, expansions or stalls to today?"),
        "fails": (f"What has {{entity}} attempted on {n} that did not hold — "
                  f"delayed, descoped, abandoned or repeated?"),
        "value": (f"What does {n} change for {{entity}} — which decisions "
                  f"does it make faster, cheaper or more accountable, and "
                  f"where is that measured?"),
        "contradicts": (f"What public record cuts against {{entity}}'s "
                        f"account of {n} — enforcement, complaint, delay, "
                        f"reversal?"),
        "corroborates": (f"Which independent party — regulator, analyst, "
                         f"customer body — has said anything about "
                         f"{{entity}}'s {n}?"),
        "ai_deployment": (f"Where does {{entity}} deploy AI or automation "
                          f"inside {n}, and is it live, piloted or claimed?"),
        "ai_data": (f"Is the data feeding {n} at {{entity}} ready for AI use "
                    f"— governed, joined, fresh — and who says so?"),
        "ai_constraint": (f"What governance, regulatory or capacity "
                          f"constraint bounds AI use in {n} at {{entity}}, "
                          f"and is it stated or inferred?"),
    }
    q = table[facet]
    if facet == "works" and primary:
        # the toolkit's own question is the sharper works probe where it
        # exists; the arc requirement is appended so the graded form still
        # asks for the timeline the works DQ demands.
        q = grade_question(primary).rstrip("?") + \
            " Trace the arc: earliest signal, refreshes, stalls, today."
    return q


#: Which facets are answerable purely from public evidence regardless of the
#: toolkit's primary-question fit. `contradicts` and `corroborates` ARE
#: public by construction — they ask what the outside world says.
_FACET_MODE = {
    "works": None,            # inherits the toolkit's mode_fit
    "fails": None,
    "value": None,
    "contradicts": "PUBLIC",
    "corroborates": "PUBLIC",
    "ai_deployment": "BOTH",
    "ai_data": None,
    "ai_constraint": "BOTH",
}


# ── build ────────────────────────────────────────────────────────────────

def build(wb: RunWorkbook, *, toolkit_dir=None, kg_path: Path | None = None
          ) -> dict:
    """Seed DQ_Bank from the toolkits + facet probes; write kg.json + checksum."""
    md = wb.metadata()
    sv = str(md.get("sub_vertical") or "") or None
    toolkit, problems = ({}, ["no --toolkits directory given; primary "
                              "questions and source lists absent"]) \
        if toolkit_dir is None else load_toolkits(toolkit_dir, sv)

    selected = wb.selected_subcaps()
    tax = C.taxonomy()
    names = _subcap_names(toolkit)

    ws = wb._sheet("DQ_Bank")
    ws.delete_rows(2, max(0, ws.max_row - 1))
    rows = []
    for cell in selected:
        t = toolkit.get(cell) or {}
        base_fit = t.get("mode_fit") or "BOTH"
        primary = t.get("question")
        order = 0
        if primary:
            rows.append({
                "SubCap_ID": cell, "Order": 0, "Facet": C.PRIMARY_FACET,
                "Probe_Tier": "TOOLKIT",
                "Question": grade_question(primary),
                "Mode_Fit": base_fit,
                "Internal_Sources": t.get("internal_sources"),
                "Public_Sources": t.get("public_sources"),
                "Weight_Pct": t.get("weight_pct"),
            })
        for i, facet in enumerate(C.FACETS, start=1):
            fit = _FACET_MODE[facet] or base_fit
            rows.append({
                "SubCap_ID": cell, "Order": i, "Facet": facet,
                "Probe_Tier": "CATALOGUE",
                "Question": _facet_dq(facet, names.get(cell), primary),
                "Mode_Fit": fit,
                "Internal_Sources": t.get("internal_sources"),
                "Public_Sources": t.get("public_sources"),
                "Weight_Pct": None,
            })
        for i, facet in enumerate(C.AI_FACETS, start=6):
            fit = _FACET_MODE[facet] or base_fit
            rows.append({
                "SubCap_ID": cell, "Order": i, "Facet": facet,
                "Probe_Tier": "AI_OVERLAY",
                "Question": _facet_dq(facet, names.get(cell), primary),
                "Mode_Fit": fit,
                "Internal_Sources": None, "Public_Sources": None,
                "Weight_Pct": None,
            })
    for r in rows:
        ws.append([r.get(c) for c in C.DQ_BANK_COLUMNS])
    wb.save()

    checksum = checksum_of(wb)
    wb.set_metadata("kg_checksum", checksum)

    kg = {
        "_contract": {
            "authority": "the workbook's DQ_Bank sheet",
            "this_file": ("a routing projection over that sheet. If the two "
                          "disagree, the sheet is right; `engine.kg verify` "
                          "detects the drift via Run_Metadata.kg_checksum."),
        },
        "run_id": md.get("run_id"),
        "sub_vertical": sv,
        "evidence_mode": md.get("evidence_mode"),
        "catalogue_version": md.get("catalogue_version"),
        "kg_checksum": checksum,
        "counts": {
            "subcaps": len(selected),
            "dqs": len(rows),
            "with_toolkit_primary": sum(1 for r in rows if r["Order"] == 0),
            "categories": len({s.split(".")[0] for s in selected}),
        },
        # Named, not left as arithmetic: these subcaps have no toolkit row,
        # so their works probe is the generic form and they carry no source
        # lists. Usually sub-vertical variant cells the toolkits do not
        # enumerate — a researcher should know which of their cells run thin.
        "subcaps_without_primary": sorted(
            set(selected) - {r["SubCap_ID"] for r in rows if r["Order"] == 0}),
        "toolkit_problems": problems,
        "categories": _route_map(selected, rows,
                                 str(md.get("evidence_mode") or "PUBLIC")),
    }
    out = Path(kg_path) if kg_path else None
    if out is not None:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(kg, indent=2))
        kg["written_to"] = str(out)
    return kg


def _subcap_names(toolkit: dict) -> dict:
    """Subcap display names, from the toolkit's own Sub-Capability column.
    Where the toolkit has no row, the probe says "this capability" — the id
    would read as jargon in a question and an invented name would be worse."""
    return {sid: t.get("name") for sid, t in toolkit.items() if t.get("name")}


def checksum_of(wb: RunWorkbook) -> str:
    """A stable digest of the DQ bank as the sheet actually holds it."""
    payload = "\n".join(
        "\t".join("" if r.get(c) is None else str(r.get(c))
                  for c in C.DQ_BANK_COLUMNS)
        for r in wb.rows("DQ_Bank"))
    return hashlib.sha256(payload.encode()).hexdigest()


def verify(wb: RunWorkbook) -> list[str]:
    """Divergences between the recorded checksum and the sheet, as sentences."""
    recorded = str(wb.metadata().get("kg_checksum") or "").strip()
    if not recorded:
        return ["the KG was never built for this run (kg_checksum is empty); "
                "orient's cards fall back to the default probes with no "
                "toolkit primaries and no source lists"]
    actual = checksum_of(wb)
    if actual != recorded:
        return [f"DQ_Bank has changed since the KG was built: recorded "
                f"{recorded[:12]}… vs sheet {actual[:12]}… — rebuild with "
                f"`engine.kg build` or the routing projection is stale"]
    return []


# ── routing ──────────────────────────────────────────────────────────────

def _route_map(selected, dq_rows, mode: str) -> dict:
    answerable = set(C.MODE_ANSWERABLE.get(mode, ()))
    by_cat: dict = {}
    for cell in selected:
        cat = cell.split(".")[0]
        by_cat.setdefault(cat, {"subcaps": [], "dqs": 0, "deferred": []})
        by_cat[cat]["subcaps"].append(cell)
    for r in dq_rows:
        cat = str(r["SubCap_ID"]).split(".")[0]
        d = by_cat.get(cat)
        if d is None:
            continue
        if (r.get("Mode_Fit") or "BOTH") in answerable:
            d["dqs"] += 1
        else:
            prefix = "INT-Q" if r.get("Mode_Fit") == "INTERNAL" else "PUB-Q"
            d["deferred"].append({
                "subcap": r["SubCap_ID"], "facet": r["Facet"],
                "discovery_question": f"{prefix}: {r['Question']}",
                "why": (f"mode_fit {r.get('Mode_Fit')} is not answerable in a "
                        f"{mode} assessment — carry it as a discovery "
                        f"question, never as a silent gap"),
            })
    for cat, d in by_cat.items():
        d["agent"] = agent_name(cat)
        d["subcaps"].sort()
    return dict(sorted(by_cat.items()))


def agent_name(category: str) -> str:
    """The per-category researcher that owns a category's worklist."""
    return f"research-{category.lower()}-producer"


def route(wb: RunWorkbook, category: str | None = None,
          mode: str | None = None) -> dict:
    md = wb.metadata()
    mode = mode or str(md.get("evidence_mode") or "PUBLIC")
    if mode not in C.ASSESSMENT_MODES:
        raise WorkbookError(f"mode {mode!r} not in {C.ASSESSMENT_MODES}")
    rows = wb.rows("DQ_Bank")
    selected = wb.selected_subcaps()
    full = _route_map(selected, rows, mode)
    drift = verify(wb)
    out = {"run_id": md.get("run_id"), "evidence_mode": mode,
           "kg_drift": drift,
           "categories": ({category: full[category]} if category else full)}
    if category and category not in full:
        raise WorkbookError(
            f"{category} is not in this run's engagement set; its categories "
            f"are {sorted(full)}")
    return out


def dqs_for(wb: RunWorkbook, subcap: str, mode: str | None = None) -> dict:
    """One subcap's questions, split answerable / deferred for the mode."""
    md = wb.metadata()
    mode = mode or str(md.get("evidence_mode") or "PUBLIC")
    answerable = set(C.MODE_ANSWERABLE.get(mode, ()))
    ask, defer = [], []
    for r in wb.rows("DQ_Bank"):
        if str(r.get("SubCap_ID") or "") != subcap:
            continue
        row = {"facet": r.get("Facet"), "order": r.get("Order"),
               "question": r.get("Question"),
               "probe_tier": r.get("Probe_Tier"),
               "mode_fit": r.get("Mode_Fit"),
               "internal_sources": r.get("Internal_Sources"),
               "public_sources": r.get("Public_Sources")}
        ((ask if (r.get("Mode_Fit") or "BOTH") in answerable else defer)
         .append(row))
    return {"subcap": subcap, "mode": mode, "ask": ask, "deferred": defer}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    sub = ap.add_subparsers(dest="cmd", required=True)
    for name in ("build", "route", "show", "verify"):
        s = sub.add_parser(name)
        s.add_argument("--run", required=True)
        s.add_argument("--root")
        if name == "build":
            s.add_argument("--toolkits", help="dir holding the four "
                                              "Pillar*_Scoring_Toolkit.xlsx")
        if name == "route":
            s.add_argument("--category")
            s.add_argument("--mode", choices=C.ASSESSMENT_MODES)
        if name == "show":
            s.add_argument("--subcap", required=True)
            s.add_argument("--mode", choices=C.ASSESSMENT_MODES)
    a = ap.parse_args(argv)
    run = runstate.locate(a.run, Path(a.root) if a.root else None)
    wb = run.open()
    if a.cmd == "build":
        out = build(wb, toolkit_dir=a.toolkits,
                    kg_path=run.root / "00_entity_profile" / "kg.json")
        print(json.dumps({k: out[k] for k in
                          ("run_id", "kg_checksum", "counts",
                           "toolkit_problems", "written_to")
                          if k in out}, indent=2))
        return 0
    if a.cmd == "route":
        print(json.dumps(route(wb, a.category, a.mode), indent=2))
        return 0
    if a.cmd == "show":
        print(json.dumps(dqs_for(wb, a.subcap, a.mode), indent=2))
        return 0
    if a.cmd == "verify":
        drift = verify(wb)
        for d in drift:
            print(f"DRIFT: {d}")
        print("kg verify: " + ("clean" if not drift else f"{len(drift)} issue(s)"))
        return 1 if drift else 0
    return 2


if __name__ == "__main__":
    sys.exit(main())
