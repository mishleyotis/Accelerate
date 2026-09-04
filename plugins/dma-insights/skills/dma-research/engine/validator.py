#!/usr/bin/env python3
"""The seven contract rules, and the three of them that never fired.

WHY THIS EXISTS.

  AUD-0064  rule 2 ("the 11 headers differ in name or order, OR A COLUMN IS
      ADDED") never enforced its third clause, because the header slice
      stopped at column 11: a workbook with 22 extra columns returned
      FAILS=0, exit 0. Rule 5 was guarded by `if fv and ...`, so a row with
      Evidence_IDs blank skipped every content check — and on the real golden
      workbook, 44 of 49 rows were blank, so a 10%-coverage workbook was
      certified clean.
  AUD-0014  the row-count rule compared `max_row-1` to a length and never
      the ID SET, so a workbook carrying entirely the wrong subcaps passed.
  AUD-0012 / AUD-0061  the required sheet set was one the pinned template had
      already retired, so the authority artefact could not pass the gate meant
      to admit it. Here the required set IS the contract's set (one object,
      `contract.REQUIRED_SHEETS`), so the two cannot disagree.
  AUD-0013  and none of it ran in production. `validate()` is called by the
      gate, by the handoff builder and by the CLI; a workbook cannot reach the
      handoff without passing it.

Every rule below has a test that FIRES it and a test that does not.
"""
from __future__ import annotations

# Runnable both ways. `python3 -m engine.<mod>` is the documented invocation,
# but every audit and every operator reaches for `python3 <path> --help`
# first, and a relative import dies there. Binding __package__ makes the two
# equivalent instead of making one of them a trap.
if __package__ in (None, ""):  # noqa: E402  (must precede the relative imports)
    import os as _os
    import sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.dirname(
        _os.path.abspath(__file__))))
    __package__ = "engine"

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from . import contract as C

URL_RE = re.compile(r"https?://", re.I)

#: Source_URLs holds a LIST. These are the separators the workbook writers
#: use between entries, and rule 6 judges each entry on its own.
ENTRY_SEP_RE = re.compile(r"[,;]")

#: A bare host with no scheme — "kpmg.com", "www.example.co.uk/x". Together
#: with URL_RE this is what "the entry is a location" means. No banned token
#: contains a dot, so no placeholder can satisfy this.
HOSTISH_RE = re.compile(
    r"^[a-z0-9](?:[a-z0-9-]*[a-z0-9])?(?:\.[a-z0-9-]+)+(?:[/:?#].*)?$", re.I)

#: Punctuation an entry may be wrapped in and still BE the placeholder:
#: "(TBD)", "N/A.", "- see above". Deliberately excludes "/" so that the
#: "n/a" token keeps its slash and cannot be reached by stripping.
TRIVIAL_EDGE = " \t\r\n.,;:!?-–—_'\"“”‘’()[]{}<>*"


def _is_location(entry: str) -> bool:
    """Does this entry name a place to go? A scheme, or a dotted host.

    This is the whole anti-false-positive guard: a location is never a
    placeholder, however its characters happen to spell.
    """
    return bool(URL_RE.match(entry) or HOSTISH_RE.match(entry))


def placeholder_entries(cell) -> list[str]:
    """The entries in one Source_URLs cell that ARE placeholders.

    THE SEMANTICS, and why they are not `token in cell` (MEM-0467).

      A cell fails rule 6 when the cell — or a comma/semicolon-separated
      entry within it — IS a placeholder. It does NOT fail because a banned
      token appears somewhere inside a longer string. The tokens are short
      and ordinary: "n/a" occurs inside every LinkedIn profile of anyone
      called An-n/a-, and inside every path segment "e-n/a-rticles", so the
      containment test refused 12 of 12 real, resolving URLs on
      bank-of-travelers-rest and blocked a run with no data defect in it.
      "tbd" and "various" are the same hazard one client away
      ("/tbd-holdings/", "/various-rates").

    Order matters: a location is excluded FIRST, so no amount of stripping
    or prefixing can ever reach a URL. Only then is a non-location entry
    judged, and it is judged generously — equal to a banned token after
    trivial punctuation, or opening with one at a word boundary, so
    "N/A.", "(TBD)" and "n/a - nothing found" all still fail.

    Returns the offending entries themselves, because the reader needs the
    value that is wrong, not the token that matched it.
    """
    out = []
    for raw in ENTRY_SEP_RE.split(str(cell or "")):
        entry = raw.strip()
        if not entry or _is_location(entry):
            continue
        core = entry.strip(TRIVIAL_EDGE).casefold()
        for bad in C.BANNED_URL_PLACEHOLDERS:
            if core == bad or re.match(rf"{re.escape(bad)}\b", core):
                out.append(entry)
                break
    return out


class Failure(dict):
    def __init__(self, rule: int, name: str, detail: str, **kw):
        super().__init__(rule=rule, name=name, detail=detail, **kw)

    def __str__(self):
        return f"rule {self['rule']} ({self['name']}): {self['detail']}"


def validate(path, *, run_id: str | None = None,
             expect_scores: bool | None = None) -> list[Failure]:
    """Every rule, against one workbook. An empty list is a pass."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    fails: list[Failure] = []
    try:
        fails += _rule1_sheets(wb)
        if any(f["rule"] == 1 for f in fails):
            return fails            # nothing else is meaningful without sheets
        fails += _rule2_headers(wb)
        fails += _rule3_rows(wb)
        # THE STAGE DECIDES, and it is recorded now. `expect_scores` was a
        # caller's opinion defaulted to False, and `assemble.package` passed
        # the default — hard-coding research semantics into every package it
        # ever built, including an assessment one. `None` means "read the
        # workbook's own stage"; an explicit True/False still wins, because
        # `--expect-scores` is how a caller checks a workbook whose stage
        # they are testing rather than trusting.
        want = (C.stage_of(_metadata(wb)) == "assessment"
                if expect_scores is None else expect_scores)
        fails += _rule4_scores(wb, want)
        fails += _rule5_evidence_and_urls(wb)
        fails += _rule6_placeholders(wb)
        fails += _rule7_run_id(wb, run_id)
        fails += _rule8_absence_declared(wb, want)
    finally:
        wb.close()
    return fails


def _metadata(wb) -> dict:
    """Run_Metadata as a dict, read off the RAW openpyxl workbook.

    `validate` opens the file with openpyxl rather than through
    `RunWorkbook`, deliberately: rules 1 and 2 exist to judge a file that may
    not BE a valid run workbook, and RunWorkbook's accessors assume the shape
    under test. So the two columns are read directly. Every caller is after
    rule 1's early return, so the sheet is known to exist.

    One reader, not two: `_rule7_run_id` carried its own copy of this loop,
    and the stage read added a third by calling `wb.metadata()` on an
    openpyxl workbook, which has no such method — an AttributeError inside
    the handoff builder, which is the only path to a strip.
    """
    md = {}
    for r in wb["Run_Metadata"].iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            md[str(r[0])] = r[1]
    return md


# ── rule 1 · the sheet set ───────────────────────────────────────────────

def _rule1_sheets(wb) -> list[Failure]:
    have = set(wb.sheetnames)
    missing = [s for s in C.REQUIRED_SHEETS if s not in have]
    if missing:
        return [Failure(1, "sheets", f"missing: {sorted(missing)}")]
    return []


# ── rule 2 · headers, INCLUDING an added column ──────────────────────────

def _rule2_headers(wb) -> list[Failure]:
    """The whole header row is compared, not a slice of it.

    The old validator read `range(1, 12)`, so appending a 12th column was
    invisible — which is precisely how an unstripped 33-column working area
    passed the only gate that guards the file."""
    out = []
    for sheet, cols in C.SHEETS.items():
        ws = wb[sheet]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        got = tuple(str(c).strip() for c in row if c is not None
                    and str(c).strip())
        if got == tuple(cols):
            continue
        extra = [c for c in got if c not in cols]
        missing = [c for c in cols if c not in got]
        bits = []
        if extra:
            bits.append(f"added column(s) {extra}")
        if missing:
            bits.append(f"absent column(s) {missing}")
        if not bits:
            bits.append(f"order differs: {list(got)}")
        out.append(Failure(2, "headers", f"{sheet}: " + "; ".join(bits),
                           sheet=sheet))
    return out


# ── rule 3 · the ROWS ARE THE ENGAGEMENT SET, by id ──────────────────────

def _rule3_rows(wb) -> list[Failure]:
    """Compared as a SET OF IDS, not as a count.

    AUD-0014 forced this: swapping an in-scope subcap for an out-of-scope one
    while holding the count constant returned FAILS=0. A count is not an
    identity."""
    tax = C.taxonomy()
    out = []
    seen: set[str] = set()
    for sheet in C.PILLAR_SHEETS:
        pillar = sheet[:2]
        ws = wb[sheet]
        ids = []
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            v = r[0] if r else None
            if v is None or not str(v).strip():
                continue
            ids.append(str(v).strip())
        dupes = sorted({i for i in ids if ids.count(i) > 1})
        if dupes:
            out.append(Failure(3, "rows", f"{sheet}: duplicate ids {dupes}",
                               sheet=sheet))
        unknown = sorted({i for i in ids if i not in tax.tier})
        if unknown:
            out.append(Failure(
                3, "rows",
                f"{sheet}: {len(unknown)} id(s) are not in catalogue "
                f"{tax.version}: {unknown[:8]}", sheet=sheet))
        wrong = sorted({i for i in ids if i in tax.tier
                        and not i.startswith(pillar)})
        if wrong:
            out.append(Failure(
                3, "rows", f"{sheet}: id(s) belonging to another pillar: "
                           f"{wrong[:8]}", sheet=sheet))
        # A seeded row carries its catalogue name (goeasy GSY-03: 656 blank
        # names shipped). The seed refuses an unnamed cell; this makes a
        # name blanked AFTER seeding fail the workbook's own validator too,
        # at every stage, rather than only at `assessment open`.
        unnamed = [str(r[0]).strip() for r in
                   ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
                   if r and r[0] is not None and str(r[0]).strip()
                   and (len(r) < 2 or r[1] is None or not str(r[1]).strip())]
        if unnamed:
            out.append(Failure(
                3, "rows",
                f"{sheet}: {len(unnamed)} row(s) carry no SubCap_Name "
                f"({unnamed[:6]}); names come from the catalogue at seed "
                f"time and are never blank", sheet=sheet))
        seen |= set(ids)
    if not seen:
        out.append(Failure(3, "rows",
                           "no scoring rows at all — the engagement set is "
                           "the workbook's scope declaration and it is empty"))
    return out


# ── rule 4 · scores at the research stage ────────────────────────────────

def _rule4_scores(wb, expect_scores: bool) -> list[Failure]:
    out = []
    for sheet in C.PILLAR_SHEETS:
        ws = wb[sheet]
        col = C.PILLAR_COLUMNS.index("Score") + 1
        scored, rows = [], 0
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip():
                rows += 1
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                scored.append((ws.cell(row=r, column=1).value, v))
        # A pillar with NO ROWS is a pillar the run never selected, and an
        # unselected pillar has nothing to score. Demanding a score there
        # made the assessment stage unreachable for every run whose scope is
        # not all four pillars — which is most of them, and which no unit
        # test caught because `expect_scores` defaulted to False until the
        # stage key existed. A sheet that HAS rows and no scores is still the
        # real failure this rule is for.
        if expect_scores and rows and not scored:
            out.append(Failure(4, "scores",
                               f"{sheet}: assessment stage, {rows} row(s) in "
                               f"scope and none scored",
                               sheet=sheet))
        if not expect_scores and scored:
            out.append(Failure(
                4, "scores",
                f"{sheet}: {len(scored)} scored row(s) at the research stage, "
                f"first {scored[0]}. Column D is the assessment's to write.",
                sheet=sheet))
    return out


# ── rule 8 · an absence is DECLARED, never merely flagged ────────────────

def _rule8_absence_declared(wb, expect_scores: bool) -> list[Failure]:
    """The absence flag has one writer, and the validator can prove it.

    `engine.cli absence` is the only path that writes `Absence_Claimed` AND
    the Provenance row Step == 'absence'; a flag without its row was written
    around the command (a notebook consolidation, a synthesis record, a hand
    edit — all measured 2026-09-03). At the assessment stage an undeclared
    NO_EVIDENCE row is a cell nobody searched, and it must not be scored."""
    out = []
    prov = wb["Provenance"] if "Provenance" in wb.sheetnames else None
    declared: set[str] = set()
    if prov is not None:
        for r in prov.iter_rows(min_row=2, values_only=True):
            if r and len(r) > 1 and str(r[1] or "").strip() == "absence":
                declared.add(str(r[0] or "").strip())
    fi = C.PILLAR_COLUMNS.index("Evidence_IDs") + 1
    ai = C.PILLAR_COLUMNS.index("Absence_Claimed") + 1
    for sheet in C.PILLAR_SHEETS:
        ws = wb[sheet]
        forged, undeclared = [], []
        for r in range(2, ws.max_row + 1):
            sub = str(ws.cell(row=r, column=1).value or "").strip()
            if not sub:
                continue
            flag = str(ws.cell(row=r, column=ai).value or "").strip().upper() \
                if ws.max_column >= ai else ""
            eids = str(ws.cell(row=r, column=fi).value or "").strip()
            if flag in ("YES", "TRUE", "1") and sub not in declared:
                forged.append(sub)
            if expect_scores and eids == C.NO_EVIDENCE and sub not in declared:
                undeclared.append(sub)
        if forged:
            out.append(Failure(
                8, "absence",
                f"{sheet}: {len(forged)} row(s) carry Absence_Claimed with no "
                f"Provenance 'absence' row — the flag was written around "
                f"`engine.cli absence`: {forged[:8]}", sheet=sheet))
        if undeclared:
            out.append(Failure(
                8, "absence",
                f"{sheet}: {len(undeclared)} NO_EVIDENCE row(s) reach the "
                f"assessment stage undeclared — score nothing that was not "
                f"searched: {undeclared[:8]}", sheet=sheet))
    return out


# ── rule 5 · evidence and URLs, with no vacuous branch ───────────────────

def _rule5_evidence_and_urls(wb) -> list[Failure]:
    """Every row is checked. There is no blank that skips the check.

    The contract says Evidence_IDs must be an E-id list or the literal
    NO_EVIDENCE — blank is neither, and the old guard `if fv and ...` turned
    every blank into a silent pass."""
    out = []
    for sheet in C.PILLAR_SHEETS:
        ws = wb[sheet]
        fi = C.PILLAR_COLUMNS.index("Evidence_IDs") + 1
        gi = C.PILLAR_COLUMNS.index("Source_URLs") + 1
        blank, unurled = [], []
        for r in range(2, ws.max_row + 1):
            sub = ws.cell(row=r, column=1).value
            if sub is None or not str(sub).strip():
                continue
            f = str(ws.cell(row=r, column=fi).value or "").strip()
            g = str(ws.cell(row=r, column=gi).value or "").strip()
            if not f:
                blank.append(str(sub))
                continue
            if f == C.NO_EVIDENCE:
                continue
            if not URL_RE.search(g):
                unurled.append(str(sub))
        if blank:
            out.append(Failure(
                5, "evidence", f"{sheet}: {len(blank)} row(s) with a blank "
                f"Evidence_IDs — the contract admits an E-id list or the "
                f"literal {C.NO_EVIDENCE}, and blank is neither: {blank[:8]}",
                sheet=sheet))
        if unurled:
            out.append(Failure(
                5, "evidence", f"{sheet}: {len(unurled)} row(s) cite evidence "
                f"with no URL in Source_URLs: {unurled[:8]}", sheet=sheet))
    return out


# ── rule 6 · banned placeholders ─────────────────────────────────────────

def _rule6_placeholders(wb) -> list[Failure]:
    out = []
    for sheet in C.PILLAR_SHEETS:
        ws = wb[sheet]
        gi = C.PILLAR_COLUMNS.index("Source_URLs") + 1
        hits = []
        for r in range(2, ws.max_row + 1):
            g = str(ws.cell(row=r, column=gi).value or "").strip()
            if not g:
                continue
            bad_entries = placeholder_entries(g)
            if bad_entries:
                # The offending VALUE, not the token that matched it: the old
                # message printed "n/a" for a cell holding a LinkedIn URL,
                # which reads as a data defect and is not one (MEM-0467).
                hits.append((str(ws.cell(row=r, column=1).value),
                             bad_entries[0]))
        if hits:
            out.append(Failure(
                6, "placeholders",
                f"{sheet}: {len(hits)} Source_URLs cell(s) hold a placeholder "
                f"instead of a link: {hits[:5]}", sheet=sheet))
    return out


# ── rule 7 · run_id equality, and the catalogue lock ─────────────────────

def _rule7_run_id(wb, run_id: str | None) -> list[Failure]:
    md = _metadata(wb)
    out = []
    have = str(md.get("run_id") or "").strip()
    if not have:
        out.append(Failure(7, "run_id", "Run_Metadata carries no run_id"))
    elif "{{" in have:
        out.append(Failure(7, "run_id",
                           f"run_id is an unresolved token: {have!r}"))
    elif run_id and have != run_id:
        out.append(Failure(7, "run_id",
                           f"workbook says {have!r}, caller says {run_id!r}"))
    lock = {}
    for r in wb["Handoff_Lock"].iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            lock[str(r[0])] = r[1]
    if not lock:
        out.append(Failure(7, "handoff_lock",
                           "Handoff_Lock is empty — the catalogue lock the "
                           "assessment compares against does not exist"))
    else:
        if str(lock.get("catalogue_hash") or "") != C.catalogue_hash():
            out.append(Failure(
                7, "handoff_lock",
                f"catalogue has moved since this run was locked: workbook "
                f"{lock.get('catalogue_hash')} vs current {C.catalogue_hash()}"))
        if str(lock.get("run_id") or "") != have:
            out.append(Failure(7, "handoff_lock",
                               "Handoff_Lock.run_id disagrees with "
                               "Run_Metadata.run_id"))
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--run-id")
    # BooleanOptionalAction with default None, NOT store_true. store_true
    # defaults to False, and False is not "no opinion" — it is the caller
    # asserting the RESEARCH stage. Passed straight through, that made the
    # command line hard-code research semantics for every workbook it was
    # ever pointed at, which is the same defect the stage key exists to
    # remove, one layer up. Absent now means "read the workbook's own stage".
    ap.add_argument("--expect-scores", action=argparse.BooleanOptionalAction,
                    default=None,
                    help="assessment stage: column D must be POPULATED. "
                         "Omit to read the workbook's recorded stage; pass "
                         "--no-expect-scores to assert the research stage.")
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args(argv)
    fails = validate(a.workbook, run_id=a.run_id, expect_scores=a.expect_scores)
    if a.json:
        print(json.dumps({"fails": len(fails), "failures": fails}, indent=2))
    else:
        for f in fails:
            print(f"FAIL {f}")
        print(f"validate_workbook: FAILS={len(fails)}")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
