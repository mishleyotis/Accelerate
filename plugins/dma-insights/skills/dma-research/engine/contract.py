#!/usr/bin/env python3
"""The one place the DMA research engine states its own shape.

WHY THIS EXISTS. Two audit findings share a single root. AUD-0062 measured
35 occurrences of the literal `836` and 21 of `17 categories` across skills
that feed an assessment, against a settled taxonomy of 851 cells in 16
categories — and the rule that would have prevented it ("counts are derived
by counting catalogue rows at render time; never write one as a literal in
prose") reached no renderer. AUD-0066 measured two incompatible workbook
shapes both called "the workbook", with nothing able to say which was right.

Both are the same defect: a shape asserted in prose in many places instead of
computed in one. So this module computes the counts from the catalogue and
declares the workbook shape as data. Everything downstream — the substrate,
the validator, the renderers, the drift check — reads from here, and a
disagreement becomes a test failure instead of a document that is quietly
wrong.

Nothing in here is a literal that a catalogue change would falsify. The one
irreducible literals are the four band boundaries and the eleven core column
names, and both are stated once, here, with the reason they cannot move.
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

# ── locating the catalogue ────────────────────────────────────────────────
#
# The engine ships inside the plugin, and the plugin may be installed far
# from any repo checkout. The catalogue is a repo artefact, so it is looked
# for in the places it can actually be, and its absence is reported as an
# absence — never defaulted to a number that looks like data (invariant 9).

_HERE = Path(__file__).resolve()
_CATALOGUE_NAME = "catalogue_v70_tier.json"


def _candidate_catalogue_paths() -> list[Path]:
    """Every place the catalogue can legitimately be, nearest first.

    A repo checkout wins over the packaged copy, so a catalogue update lands
    without a plugin release; the packaged copy exists so a trigger-fired
    container with no checkout can still count (AUD-0069's lesson: an input a
    headless producer cannot resolve is an input it does not have)."""
    out: list[Path] = []
    env = os.environ.get("DMA_CATALOGUE")
    if env:
        out.append(Path(env))
    for anc in _HERE.parents:                       # a checkout above us
        out.append(anc / "packages" / "shared" / _CATALOGUE_NAME)
    out.append(_HERE.parent / "data" / _CATALOGUE_NAME)   # packaged fallback
    return out


class CatalogueUnavailable(RuntimeError):
    """The catalogue could not be read, so no count can be stated.

    Raised rather than returning a plausible number. A run that cannot count
    its own universe must say so; every finding in the AUD-0062 family began
    with a number that was available when it should not have been."""


@lru_cache(maxsize=1)
def catalogue_path() -> Path:
    for p in _candidate_catalogue_paths():
        if p.is_file():
            return p
    raise CatalogueUnavailable(
        "catalogue_v70_tier.json not found. Looked in: "
        + ", ".join(str(p) for p in _candidate_catalogue_paths())
    )


#: The subcapability DISPLAY NAMES, one per catalogue cell, extracted from
#: the owner's workbook template (DQ_Bank.SubCap_Name, 851 rows). Owner,
#: 2026-09-03: "the workbook always defaults to the wrong structure each
#: time; missing fields … missing subcaps names". The engine seeded column B
#: empty because the tier catalogue carries no names and the toolkits (which
#: do) are pulled from Drive after the run starts — so every run began with
#: 690 blank names and nothing downstream filled them (goeasy GSY-03: 656
#: blank SubCap_Name rows shipped). Names travel WITH the catalogue now.
_NAMES_NAME = "catalogue_v70_names.json"


def names_path() -> Path:
    for p in _candidate_catalogue_paths():
        q = p.with_name(_NAMES_NAME)
        if q.is_file():
            return q
    raise CatalogueUnavailable(
        f"{_NAMES_NAME} not found beside the tier catalogue. Looked in: "
        + ", ".join(str(p.with_name(_NAMES_NAME))
                    for p in _candidate_catalogue_paths()))


def subcap_names() -> dict[str, str]:
    """cell id -> display name, for every cell in the catalogue."""
    raw = json.loads(names_path().read_text(encoding="utf-8"))
    return {str(k): str(v).strip() for k, v in (raw.get("names") or {}).items()}


def proxy_classes() -> dict[str, str]:
    """cell id -> the proxy class the template names for an absent cell
    (leadership_title, regulator_filing, org_talent, …). The ladder a
    declared absence must climb starts on this rung."""
    raw = json.loads(names_path().read_text(encoding="utf-8"))
    return {str(k): str(v).strip()
            for k, v in (raw.get("proxy_class_if_absent") or {}).items()}


_CELL_RE = re.compile(r"^(P\d)C(\d+)\.(\d+)(?:\.(.+))?$")


@dataclass(frozen=True)
class Taxonomy:
    """The catalogue, counted. Every field is measured, none is asserted."""
    version: str
    cells: tuple[str, ...]
    pillars: tuple[str, ...]
    categories: tuple[str, ...]
    capabilities: tuple[str, ...]
    universal: tuple[str, ...]
    variants: tuple[str, ...]
    tier: dict

    # The counts the prose used to hardcode. Properties, not stored values,
    # so there is no second place for them to drift to.
    @property
    def n_pillars(self) -> int: return len(self.pillars)

    @property
    def n_categories(self) -> int: return len(self.categories)

    @property
    def n_capabilities(self) -> int: return len(self.capabilities)

    @property
    def n_cells(self) -> int: return len(self.cells)

    @property
    def n_universal(self) -> int: return len(self.universal)

    @property
    def n_variants(self) -> int: return len(self.variants)

    def name_of(self, cell: str) -> str:
        """The catalogue's display name for a cell, or '' when it has none.

        Never invents one: an id rendered where a name belongs is jargon, an
        invented name is worse, and '' is a state the seeder and the gates
        can see (the scoring gate refuses an unnamed row)."""
        return subcap_names().get(cell, "")

    def cells_in(self, grain: str) -> tuple[str, ...]:
        """Every cell under a pillar (P1), category (P1C1) or capability
        (P1C1.1). Prefix matching is done on the parsed id, not the string,
        so P1C1 never swallows P1C10."""
        want = grain.strip()
        out = []
        for c in self.cells:
            m = _CELL_RE.match(c)
            if not m:
                continue
            pillar, cat, cap = m.group(1), f"{m.group(1)}C{m.group(2)}", \
                f"{m.group(1)}C{m.group(2)}.{m.group(3)}"
            if want in (pillar, cat, cap):
                out.append(c)
        return tuple(out)

    def sub_vertical_codes(self) -> tuple[str, ...]:
        """The sub-vertical codes the catalogue actually carries, from its
        own tier labels (T2-CU, T2-RB, …). Never a hand-kept list."""
        codes = {v.split("-", 1)[1] for v in self.tier.values() if "-" in v}
        return tuple(sorted(codes))

    def selected(self, sv: str | None, scope: str) -> tuple[str, ...]:
        """The engagement set for a sub-vertical and scope mode.

        AUD-0077: the archive's binder validated neither argument, so an
        unknown scope selected everything and an unknown sub-vertical
        silently produced a 686-cell generic engagement with exit 0. Both
        are refused here — an upstream classification failure must halt,
        not produce a plausible-looking run."""
        if scope not in SCOPE_MODES:
            raise ValueError(
                f"unknown scope {scope!r}; valid: {', '.join(sorted(SCOPE_MODES))}")
        known = self.sub_vertical_codes()
        if sv is not None and sv not in known:
            raise ValueError(
                f"unknown sub-vertical {sv!r}; the catalogue carries: "
                f"{', '.join(known)}")
        univ = list(self.universal)
        if sv is None:
            overlay: list[str] = []
        else:
            overlay = [c for c in self.variants
                       if self.tier.get(c, "").endswith("-" + sv)]
        if scope == "T1_CORE":
            return tuple(univ)
        if scope == "OVERLAY_ONLY":
            return tuple(overlay)
        # FULL and OFFERING both take the overlay. AUD-0077 also measured the
        # overlay landing ALONGSIDE its base sibling, so an entity was
        # researched twice on near-identical capabilities. An overlay
        # supersedes the base cell it varies, so the base is withdrawn.
        superseded = {self.base_of(c) for c in overlay}
        return tuple([c for c in univ if c not in superseded] + overlay)

    def base_of(self, variant: str) -> str | None:
        """The universal cell a sub-vertical variant supersedes, or None.

        `P1C1.3.CU1` varies the capability `P1C1.3`. The archive derived
        `overlay_of` as the parent CAPABILITY id, which is not a cell, so
        0 of 31 targets resolved and no engine code could use it. Resolved
        here to a cell: the lowest-numbered universal sibling under that
        capability, which is the one the variant replaces.

        None is a real answer, not a failure: 68 of the 165 variants sit
        under a capability with no universal sibling at all. Those are
        ADDITIVE — they supersede nothing, and selecting them alongside the
        universal set is correct. Distinguishing the two is the whole point;
        collapsing them is how AUD-0077's double-research happened."""
        m = _CELL_RE.match(variant)
        if not m or (m.group(4) or "").isdigit():
            return None
        cap = f"{m.group(1)}C{m.group(2)}.{m.group(3)}"
        sibs = [c for c in self.universal if c.rsplit(".", 1)[0] == cap]
        return min(sibs) if sibs else None


SCOPE_MODES = ("T1_CORE", "FULL", "OFFERING", "OVERLAY_ONLY")


@lru_cache(maxsize=1)
def taxonomy() -> Taxonomy:
    raw = json.loads(catalogue_path().read_text())
    tier = raw["tier"]
    cells = tuple(sorted(tier))
    pillars, cats, caps = set(), set(), set()
    universal, variants = [], []
    for c in cells:
        m = _CELL_RE.match(c)
        if not m:
            raise CatalogueUnavailable(f"catalogue holds an unparseable id: {c!r}")
        pillars.add(m.group(1))
        cats.add(f"{m.group(1)}C{m.group(2)}")
        caps.add(f"{m.group(1)}C{m.group(2)}.{m.group(3)}")
        (universal if (m.group(4) or "").isdigit() else variants).append(c)
    declared = raw.get("cells")
    if declared is not None and declared != len(cells):
        raise CatalogueUnavailable(
            f"catalogue declares {declared} cells and holds {len(cells)}")
    return Taxonomy(
        version=raw.get("catalogue_version", "unknown"),
        cells=cells,
        pillars=tuple(sorted(pillars)),
        categories=tuple(sorted(cats)),
        capabilities=tuple(sorted(caps)),
        universal=tuple(universal),
        variants=tuple(variants),
        tier=tier,
    )


# ── maturity bands ────────────────────────────────────────────────────────
#
# Four bands, strict less-than, on the RAW score before display rounding
# (charter invariant 6). A fifth level must not exist in code, enum or prose:
# AUD-0071 found an 85-block rubric built on five, and AUD-0144 found 36
# offering mappings naming one. The resolver has four branches and the tuple
# below is the only place a band word is written, so there is nowhere for a
# fifth to go and a test asserts the file never names one.

BANDS = ("Activating", "Building", "Competing", "Differentiating")
_BAND_MAX = (2.0, 3.0, 4.0)


def band_of(score):
    """The band word for a raw score, or None when there is no score.

    None is not a band and not a colour: a null score renders as no swatch
    (AUD-0049 found a grey swatch painted for null, which reads as a
    measurement that was never taken)."""
    if score is None:
        return None
    s = float(score)
    for word, hi in zip(BANDS, _BAND_MAX):
        if s < hi:
            return word
    return BANDS[3]


# ── the contract-v3 scoring workbook ──────────────────────────────────────
#
# PROVENANCE, stated because it is load-bearing. The eleven core columns are
# the set the deployed parser already reads (apps/worker/tests/
# test_silent_drop_classes.py::_CANON) and the set the archive's validator
# slices (columns 1..11). The working area's anchors — L Dominant_Claim,
# T Triangulation, V Why_It_Matters, W DMA_Impact, AG Retrieved_At — are the
# positions AUD-0011 and AUD-0065 measured in the pinned template. The
# columns BETWEEN those anchors are this repository's declaration, because
# the pinned Drive template is not resolvable from any code path (AUD-0069)
# and a shape nobody can fetch cannot be the authority for one that runs.
#
# `compare_to_workbook()` below exists so that when the pinned template IS
# fetched, the divergence is reported rather than silently absorbed.

CORE_COLUMNS = (
    "SubCap_ID",         # A
    "SubCap_Name",       # B
    "Category",          # C
    "Score",             # D  research leaves this EMPTY; a value here is rule 4
    "Confidence",        # E
    "Evidence_IDs",      # F  E-id list, or the literal NO_EVIDENCE. Never blank.
    "Source_URLs",       # G
    "Evidence_Ceiling",  # H
    "Caps_Applied",      # I
    "Rationale",         # J
    "Proxy_Searched",    # K
)

WORKING_AREA = (
    "Dominant_Claim",              # L  ← anchor
    "Claim_Label",                 # M
    "What_We_Found",               # N
    "Facet_Coverage",              # O
    "DQ_Works",                    # P
    "DQ_Fails",                    # Q
    "DQ_Value",                    # R
    "DQ_Corroborates",             # S
    "Triangulation",               # T  ← anchor
    "Ceiling_Reasoning",           # U
    "Why_It_Matters",              # V  ← anchor
    "DMA_Impact",                  # W  ← anchor
    "DQ_Contradicts",              # X
    "Contradiction_Disposition",   # Y
    "Absence_Claimed",             # Z
    "Proxy_Log",                   # AA
    "Negative_Ladder",             # AB
    "Discovery_Questions",         # AC
    "Challenge_Verdict",           # AD
    "Ceiling_Band",                # AE
    "Uncertainty",                 # AF
    "Retrieved_At",                # AG ← anchor
)

PILLAR_COLUMNS = CORE_COLUMNS + WORKING_AREA

#: Anchors the pinned template fixes by position. Asserted in a test so a
#: reordering of WORKING_AREA cannot slip through unnoticed.
WORKING_AREA_ANCHORS = {
    "L": "Dominant_Claim", "T": "Triangulation", "V": "Why_It_Matters",
    "W": "DMA_Impact", "AG": "Retrieved_At",
}

EVIDENCE_COLUMNS = (
    "E_ID", "Fact_ID", "Source_Name", "Source_URL", "Tier", "ERS",
    "Date_Published", "Recency", "Claim_Type", "Fact_Count", "SubCap_IDs",
    "Excerpt", "Anchor_Quote", "Retrieved_At", "Origin", "Access_Status",
    "Conflict",
)

#: The digital-evolution timeline (C1 on the served CONTEXT page).
#:
#: WIDENED 2026-08-30, and the widening is the app's vocabulary rather than a
#: preference. The tab carried one `Signal` column drawn from a nine-token
#: EVENT-CLASS list, while `context.timeline` needs BOTH a three-token
#: DIRECTION (`signal`) and an eight-token CLASS (`kind`) — plus the body,
#: the maturity effect and the claim label it renders. A run's own dated
#: events are stronger ground for C1 than a re-search, and they could not
#: reach it: the vocabularies did not map and the tab had no reader at all.
TIMELINE_COLUMNS = ("Event_Date", "Title", "Body", "Kind", "Signal",
                    "Maturity_Effect", "Claim_Label", "SubCap_IDs",
                    "Evidence_IDs")

#: The event's DIRECTION for maturity — what D5 clusters on. The consequence
#: sentence belongs in Maturity_Effect, not here.
TIMELINE_SIGNALS = ("POSITIVE", "NEUTRAL", "NEGATIVE")

#: The event's CLASS — what D5 filters on. These are the app's eight, exactly:
#: a near-miss ('TECHNOLOGY' for PLATFORM, 'CAPABILITY' for DATA) is not a
#: synonym, it is an event no filter can reach. Measured on a served run, 4
#: of 11 events carried a kind outside the eight and were invisible on a page
#: that rendered them.
TIMELINE_KINDS = ("PLATFORM", "LEADERSHIP", "M&A", "REGULATORY", "CHANNEL",
                  "DATA", "SECURITY", "STRATEGY")

#: The nine event classes the tab used to carry, and where each lands in the
#: app's eight. Kept as a MAP rather than deleted: a run pinned to an earlier
#: engine wrote these words, and a bridge is how a reader of that workbook
#: still gets a filterable event.
TIMELINE_KIND_BRIDGE = {
    "INVESTMENT": "STRATEGY", "DIVESTMENT": "STRATEGY",
    "LEADERSHIP": "LEADERSHIP", "PLATFORM": "PLATFORM",
    "REGULATORY": "REGULATORY", "MERGER": "M&A", "INCIDENT": "SECURITY",
    "LAUNCH": "CHANNEL", "PARTNERSHIP": "STRATEGY",
}

#: The technographic register — the research-stage record behind the fourth
#: deliverable (the Technographic Scan). Vocabulary is the CHARTER's, not the
#: prototype's: four layers OPS/CUST/DATA/INFRA (never L2-L5), four statuses
#: with CLAIMED present and required per row.
TECH_REGISTER_COLUMNS = (
    "TS_ID", "Product", "Vendor", "Layer", "Status", "Evidence_Level",
    "Detection_Basis", "Detection_Method", "Providers", "SubCap_IDs",
    "Evidence_IDs", "Source_URLs", "As_Of", "DMA_Impact",
)
TECH_LAYERS = ("OPS", "CUST", "DATA", "INFRA")
TECH_STATUS = ("CONFIRMED", "INFERRED", "CLAIMED", "ABSENT")
#: How a detection was made — the scan must say, per row.
TECH_METHODS = ("technographic_scan", "public_document", "job_posting",
                "vendor_announcement", "internal_document", "client_stated")

#: WHO produced the row. The deployed app's techstack facet declares its
#: sources as exactly {explorium, clay} (apps/api/dma_api/computed.py, and
#: the mcp server's `record_enrichment` source vocabulary), so a register
#: assembled from whatever a web search happened to surface is an estate the
#: app cannot reconcile against its own contract. `Providers` is REQUIRED per
#: row and is a list, because a row is often seen by more than one — and
#: which ones is exactly the question `Status` turns on.
TECH_PROVIDERS = ("clay", "explorium", "indeed", "exa", "tavily", "web",
                  "drive", "internal", "client")

#: The two data brokers. A broker asserts a deployment with no primary
#: source behind it, which is the definition of CLAIMED — so a row whose
#: ONLY providers are brokers may not wear CONFIRMED however many brokers
#: agree, because two brokers reselling one crawl is one observation.
TECH_BROKERS = ("clay", "explorium")

#: The tools a Search_Log row may name. Owner, 2026-09-03: cells were closed
#: "no evidence" with no enrichment effort — and the log could not show it,
#: because `--tool` was free text (measured: "my-made-up-tool" accepted). A
#: closed vocabulary is what lets the floors gate COUNT which connectors were
#: asked before a cell is declared absent. `web_search`/`web_fetch` are the
#: built-in tools; everything after them is an enrichment connector.
SEARCH_TOOLS = ("web_search", "web_fetch", "exa", "tavily", "clay",
                "explorium", "vibe", "indeed", "quartr", "drive", "internal")
#: The connectors whose absence from a cell's searches means "no enrichment
#: was attempted" — a declared absence must show at least one of these.
ENRICHMENT_TOOLS = tuple(t for t in SEARCH_TOOLS
                         if t not in ("web_search", "web_fetch"))

#: T3, the drilldown. A register row answers "what do they run"; the detail
#: page a click opens answers "so what" — and it has three content cards, two
#: of which render from nothing the research run used to capture. The
#: 2026-08-30 audit measured the last real run at 0 of 32 rows carrying a
#: peer deployment and 0 of 32 impacts naming a pathway, which is why the
#: drawer opened onto its own empty states.
#:
#: `DMA_Impact` (on Tech_Register) is the 40–90 word answer to "what does
#: this product do to the assessment", and it belongs to the research run
#: because that is where the subcap evidence is. Peer deployments get their
#: own sheet because they are one-to-many per product, and a many packed
#: into one cell is the shape nothing can query.
TECH_PEER_COLUMNS = (
    "TS_ID", "Peer", "Deployed", "Basis", "Source_URL", "As_Of",
)

#: THE ASSESSMENT STAGE'S THREE, and the column names are the APP's own —
#: the same technique PEER_BENCHMARK_COLUMNS uses, so a workbook this engine
#: writes is readable by the parser that already exists. All three are read
#: today (`parse_grain_summaries`, `parse_recommendations`), both land
#: server-side, and both already back live gates: the 0.05 grain tolerance
#: reads the stated grains, and CG-39 reads `recommendations_raw`. None of
#: them was ever WRITTEN, so every engine package landed with zero
#: recommendations and both stated grains absent.
#:
#: They are ASSESSMENT-stage sheets. Column D is empty at the research stage
#: by contract rule 4, so a research workbook cannot honestly fill a score
#: grain — and declaring them required at every stage would give every
#: research run three tabs it can never fill, which is the same defect
#: facing the other way.
#:
#: `Pillar` is the anchor the parser looks for, and `Score` has NO aliases at
#: grain level: a grain whose score column is missing is dropped wholesale.
PILLAR_SUMMARY_COLUMNS = (
    "Pillar", "Pillar_Name", "Score", "Weight_Pct", "Peer_Median",
    # Added 2026-09-03 (contract v6): the gold-standard workbook states the
    # gap and the M-band beside the score (Golden 1: Weighted_Score /
    # Maturity / Peer_Median / Gap_to_Peer) and the app's `_STAT_ALIASES`
    # already reads `gap_to_peer` as the delta. Additive; `_upgrade_shape`
    # migrates a v5 sheet on open.
    "Gap_to_Peer", "Maturity",
)

#: `Category_ID` is the anchor; `Pillar` lets the parser skip deriving it
#: from the id prefix. `Priority_Tier` is a string and `Priority_Score` a
#: number, which is why they are two columns rather than one.
CATEGORY_DETAIL_COLUMNS = (
    "Category_ID", "Category_Name", "Pillar", "Score", "Peer_Median",
    "Priority_Score", "Priority_Tier",
    # v6: the gap, the M-band and the coverage disclosure per category
    # (Golden 1's Category_Detail carries Maturity / Coverage / Peer_Median_Est).
    "Gap_to_Peer", "Maturity", "Coverage",
)

# ── THE ASSESSMENT-STAGE SHEETS (contract v6, 2026-09-03) ────────────────
#
# Owner: "the workbook always defaults to the wrong structure each time;
# missing fields etc." The gold-standard workbook (Golden 1, 43 sheets) is the
# research substrate PLUS the tabs the scoring stage and the client profile
# fill — and the app READS them (`workbook_parser._TAB_TARGET`, 29 tabs each
# bound to the surface it feeds). None of them had a writer in this engine, so
# the assessment skill built a SEPARATE 11-sheet workbook from a scratchpad
# and a package shipped two half-workbooks (Bank of Travelers Rest: 20 + 23
# tabs, 11 and 13 read-tabs, eighteen of nineteen runs with zero scored cells).
#
# Every sheet below is written by `engine.profile` (research stage: the
# client's own facts) or `engine.assessment` (scoring stage), through
# refusals, into the ONE workbook. Column names are the app parser's own
# where it has one (Recommendations, the grains, Caps_Applied_Log) and the
# template's where the report reads it (Subcap_Scores' six AI-overlay
# columns are the Doc's §5 contract, verbatim).

#: The dashboard an executive reads first (gold gate GS-WB-DASHBOARD).
EXECUTIVE_SUMMARY_COLUMNS = ("Field", "Value")
EXECUTIVE_SUMMARY_FIELDS = (
    "Institution", "Sub-Vertical", "Evidence Mode", "Overall Maturity",
    "Peer Median (est.)", "Gap to Peer", "Subcaps Scored",
    "Evidence Gaps (Unknown)", "P1", "P2", "P3", "P4", "Headline",
)
#: One row per scored subcap, with the AI-and-data overlay the assessment
#: report's §5 renders from ("These six columns are the contract between the
#: workbook and this section. A pillar cannot render its overlay if they are
#: absent."). `source_cell` is the pillar-sheet row the score was struck on.
SUBCAP_SCORES_COLUMNS = (
    "subcap_id", "subcap_name", "category", "source_cell", "score",
    "confidence", "evidence_ids", "source_urls", "evidence_ceiling",
    "caps_applied", "rationale",
    "ai_applicability", "data_dependency", "data_readiness",
    "ai_evidence_ids", "ai_blocker", "peer_ai_signal",
)
AI_APPLICABILITY = ("NONE", "ASSISTIVE", "AUGMENTED", "AUTONOMOUS")
DATA_READINESS = ("RED", "AMBER", "GREEN", "UNKNOWN")
PILLAR_ROLLUP_COLUMNS = (
    "pillar_id", "pillar_name", "score", "weight", "weighted_contribution",
    "peer_median", "gap", "level",
)
CATEGORY_ROLLUP_COLUMNS = (
    "category_id", "category_name", "pillar_id", "score", "peer_median",
    "gap", "level", "coverage",
)
PILLAR_WEIGHTS_COLUMNS = ("weight_set_id", "pillar_id", "pillar_name", "weight")
#: The maturity SCORE scale's rubric tab (level / name / range / meaning).
#: The levels themselves live in engine/rubric.py, built without a literal
#: top-level token so this file keeps the four-band invariant test honest:
#: the four display BANDS are the only band words prose may carry.
MATURITY_RUBRIC_COLUMNS = ("level", "name", "range", "meaning")
CATALOGUE_META_COLUMNS = ("key", "value")
CAP_TRIGGERS_COLUMNS = ("rule_id", "severity", "max_score", "effect", "horizon")
CAP_TRIGGERS = (
    ("CAP-S3", "CRITICAL (active enforcement <12mo)", 2.0,
     "Caps affected capabilities at M2", "while active; lifts on remediation"),
    ("CAP-S2", "MATERIAL (issue active/terminated <24mo)", 3.0,
     "Caps affected capabilities at M3",
     "lifts when the issue closes or ages past 24 months"),
    ("CAP-S1", "MINOR (resolved)", 4.0, "Caps affected capabilities at M4",
     "lifted on resolution"),
    ("CAP-T5", "T5-only evidence", 2.0,
     "Marketing or claims alone cannot exceed M2", "lifts on corroborating evidence"),
    ("CAP-T45", "T4/T5-only evidence", 2.5,
     "Unvalidated internal narrative plus claims cannot exceed M2.5",
     "lifts on a T1-T3 source"),
    ("CAP-SS", "Single-source", 3.0, "A single source cannot exceed M3",
     "lifts on triangulation"),
)
CAPS_APPLIED_LOG_COLUMNS = (
    "subcap_id", "category", "final_score", "evidence_ceiling", "caps_applied",
)
ISSUE_REGISTER_COLUMNS = (
    "ID", "Type", "Severity", "Status", "Description", "Capability impact",
    "Cap", "Evidence_IDs", "As_Of",
)
ISSUE_SEVERITIES = ("CRITICAL", "MATERIAL", "MINOR")
ISSUE_STATUSES = ("Active", "Resolved", "Terminated", "Monitoring")
FIRMOGRAPHICS_COLUMNS = ("Field", "Value", "Unit", "As at", "Evidence",
                         "Conf.", "State", "Reason", "Route")
FIRMOGRAPHIC_STATES = ("STATED", "ABSENT", "QUARANTINED")
#: The Client Profile §1.1 must-present set. `website` is load-bearing in the
#: app and required on every sub-vertical; the rest are stated or carry a
#: reason, never blank.
FIRMOGRAPHIC_MUST_PRESENT = (
    "website", "employees", "assets_or_aum_or_revenue", "cagr", "branches",
    "headquarters", "founded", "primary_regulator", "charter", "ownership",
)
FOCUS_AREAS_COLUMNS = ("ID", "Priority in the client's words", "Verbatim quote",
                       "Document", "Page", "Cells", "Evidence_IDs",
                       "Currency_Status", "Currency_Note")
CURRENCY_STATUSES = ("CONFIRMED_CURRENT", "AGING", "SUPERSEDED", "UNCONFIRMED")
SOLUTION_CATALOGUE_COLUMNS = ("solution_id", "solution_name", "platform",
                              "categories", "rec_id")
PLATFORM_PEER_ADOPTION_COLUMNS = ("Product / Layer", "Peer", "Verdict", "Basis",
                                  "Source", "As at")
CAPABILITY_DEFINITIONS_COLUMNS = ("category_id", "category_name", "pillar",
                                  "assessed_through")
#: The gold coverage DISCLOSURE (GS-WB-COVERAGE): scored, unknown, percent.
COVERAGE_MAP_COLUMNS = ("category_id", "category_name", "subcaps", "evidenced",
                        "evidence_gap", "coverage_pct", "confidence_posture")
ENRICHMENT_NEEDED_COLUMNS = ("Area", "Field / cell", "Status", "What would close it")

#: The multi-year financial trajectory (research stage; Client Profile §3/§4,
#: assessment §1/§6; gold gate GS-WB-FINANCIALS / GSY-18). LONG format — one
#: row per (metric, fiscal year) — because a fixed-width contract cannot carry
#: a variable number of year columns; the report renderer pivots it wide and
#: computes the CAGR at render time. Golden 1 carries a five-year series; the
#: floor is FINANCIAL_YEARS_FLOOR fiscal years across FINANCIAL_METRICS_FLOOR
#: metrics, declarable away only through `engine.completeness declare` with a
#: reason (an institution that publishes fewer years).
FINANCIAL_TRENDS_COLUMNS = ("Metric", "Fiscal_Year", "Value", "Unit",
                            "Evidence_IDs", "Source_URL", "Basis")
FINANCIAL_YEARS_FLOOR = 5
FINANCIAL_METRICS_FLOOR = 3

#: The Recommendations tab. Its header row is recognised when at least TWO
#: of its cells are in the parser's 29-token vocabulary — these are seven of
#: them — and the FIRST column becomes the rec_id, which is why `Rec_ID`
#: leads. A row whose first cell is blank is skipped as a spacer.
RECOMMENDATIONS_COLUMNS = (
    "Rec_ID", "Title", "Category_ID", "Priority", "Horizon", "Owner",
    "Rationale",
)

COVERAGE_COLUMNS = (
    "Category_ID", "Selected", "Researched", "Items", "Floor_Pass",
    "Floor_Pass_Pct", "Synthesised", "Verdict",
)

SEARCH_LOG_COLUMNS = (
    "Seq", "Timestamp", "SubCap_ID", "Facet", "Query", "Tool", "Hits",
    "Kept", "Outcome",
)

GATE_LOG_COLUMNS = (
    "Timestamp", "Gate", "Scope", "Verdict", "Detail", "Blocking",
)

#: WHO did each step. AUD-0018 / AUD-0024: the repository solved reviewer
#: independence BY CONSTRUCTION in one agent — `learning-grader` carries no
#: Write/Edit and no connector write tool, so it cannot touch the change it
#: is scoring — and then explicitly inverted it in the research challenge,
#: where the same actor writes a synthesis and its own verdict on that
#: synthesis. A verdict you wrote on your own work is a feeling.
#:
#: Independence is only checkable if authorship is RECORDED, so every write
#: names its actor here and the gate compares them.
PROVENANCE_COLUMNS = ("SubCap_ID", "Step", "Actor", "At", "Detail", "Session")

#: The steps whose authorship is load-bearing.
PROVENANCE_STEPS = ("synthesis", "challenge", "repair", "enrichment",
                    # v6: who struck each score and who declared each absence
                    "score", "absence")

#: The challenge, in full — the verdict on the scoring row is a
#: denormalised copy of the latest row here, and the gate reconciles them.
CHALLENGE_LOG_COLUMNS = (
    "SubCap_ID", "Verdict", "Actor", "Dimensions", "Rationale",
    "Ceiling_Band_Delta", "At", "Session",
)

#: A challenge verdict says one of these. AUD-0102: the schema required a
#: `dimensions` object with no required keys, so a ZERO-dimension verdict
#: validated and the whole challenge could be a token.
CHALLENGE_VERDICTS = ("PASS", "FAIL", "NOT_RUN")

#: The dimensions a challenge must actually address. `synthesis_quality` is
#: the one the shipped card silently omitted, and it carries ten
#: sub-conditions — so it is required by name, not by count.
CHALLENGE_DIMENSIONS = (
    "evidence_sufficiency", "claim_label_fit", "facet_coverage",
    "contradiction_handling", "ceiling_reasoning", "recency",
    "synthesis_quality",
)

HANDOFF_LOCK_COLUMNS = ("Key", "Value")

#: The category-grain peer store, which AUD-0042 found had no feeder at all:
#: the pinned workbook removed `Peer_Benchmarks` and the app's missing-tab
#: path recorded nothing, so `peer_scores` is empty for every new run — which
#: also empties ET-09's allow-list, making a legitimate peer who is also a
#: corpus client read as foreign-entity contamination.
#:
#: The column names are the app's own (`workbook_parser._STAT_ALIASES` and
#: the named-peer columns after them), so a workbook this engine writes is
#: readable by the parser that already exists rather than needing a new one.
PEER_BENCHMARK_COLUMNS = (
    "Category_ID", "Category_Name", "Entity_Score", "Peer_Median",
    "Peer_P25", "Peer_P75", "Peer_N", "Peer_Basis", "Source_Cell",
    "Peer_Names", "Peer_Scores", "As_Of",
)

#: What `Peer_Basis` may say. The app serves a five-rung ladder and discloses
#: which rung a figure came from; a figure with no basis cannot be disclosed.
PEER_BASIS = ("table", "recomputed", "inferred", "cannot_estimate")

#: One row per diagnostic question. `Mode_Fit`, `Internal_Sources` and
#: `Public_Sources` come from the Pillar Scoring Toolkit's own columns I/J/K
#: — the toolkit names, per subcap, exactly which internal artefacts and
#: which public artefacts answer its question, which is the highest-value
#: routing information the workbook carries: a researcher told WHAT to look
#: for stops fishing.
DQ_BANK_COLUMNS = ("SubCap_ID", "Order", "Facet", "Probe_Tier", "Question",
                   "Mode_Fit", "Internal_Sources", "Public_Sources",
                   "Weight_Pct")

#: Where the two client-facing reports are WRITTEN, so they can be CURATED.
#:
#: AUD-0052 measured the reports rendered from the JSON plane and never from
#: the workbook — "the report and the register are produced from DIFFERENT
#: SUBSTRATES with no join, so nothing could have caught it", which is the
#: mechanism behind 6 of 21 cited ids not resolving in a delivered report.
#: An agent writes its narrative HERE, beside the evidence it cites, and the
#: renderer curates from the same sheets the gates read.
#: One row per report section. The first eight columns are the PROSE; the
#: rest are the ARGUMENT behind it, and they exist because the 2026-08-30
#: audit asked four questions a body of text cannot answer on its own —
#: how the argument was weighed, how an absence was confirmed rather than
#: assumed, where an inference is doing work a fact is not, and what bias
#: the author is aware of carrying. A section that cannot fill them is a
#: section that was written rather than reasoned.
REPORT_NARRATIVE_COLUMNS = (
    "Report", "Section_ID", "Heading", "Body", "Evidence_IDs", "Kind",
    "Author", "Written_At",
    # How the claims in Body were weighed: the evidence mass behind them
    # (ERS), what was weighed AGAINST them, and why the balance fell where
    # it did. Not a restatement of Body — the reasoning Body concludes.
    "Weighing",
    # Every absence the section asserts, with the proxy ladder that
    # establishes it. "We found no evidence of X" without a ladder is an
    # admission about the search, not a finding about the client.
    "Absence_Basis",
    # What the author assumed, and which way each assumption cuts. Named,
    # because an unnamed assumption reads as a fact.
    "Assumptions",
    # The section's own account of what would bias it — sub-vertical
    # priors, source availability skew, the client's own publishing habits.
    "Bias_Notes",
    # Which statements are INFERENCE rather than fact, and what would
    # confirm each. Tagged in-line as [INF] in Body and enumerated here.
    "Inference_Tags",
    # The measurable accuracy claim: citation density, ERS mass, share of
    # claims that survived challenge. Computed, never asserted.
    "Accuracy_Basis",
    # The independent verdict. Written by an actor that did not author the
    # section, exactly as a synthesis challenge is.
    "Review_Verdict", "Review_Actor", "Review_At",
    # WHICH CARD, for the three sections that are a LIST rather than a
    # passage — insight cards, findings, recommendations. Until 2026-08-30
    # `narrative.write` overwrote a section's single row on every write, so
    # the eight-card blocking minimum the same module enforces was
    # arithmetically unreachable: the writer could not produce what the
    # checker demanded. Blank on a prose section; the card's own id on a
    # list one, and (Report, Section_ID, Card_ID) is the row's identity.
    "Card_ID",
)

#: `Kind` vocabulary for a Report_Narrative row.
NARRATIVE_KINDS = ("section", "insight_card", "recommendation", "finding")

RUN_METADATA_COLUMNS = ("Key", "Value")

#: The Run_Metadata keys the chain depends on. `run_id` and `catalogue_hash`
#: are the two anti-drift anchors: AUD-0010 found both shipping as unresolved
#: template tokens ({{RUN_ID}}, {{CHECKSUM}}), and AUD-0060 found the
#: catalogue-hash lock built nowhere at all.
RUN_METADATA_KEYS = (
    "run_id", "entity_name", "entity_id", "sub_vertical", "scope_mode",
    "catalogue_version", "catalogue_hash", "taxonomy_pillars",
    "taxonomy_categories", "taxonomy_capabilities", "taxonomy_cells",
    "subcaps_selected", "reference_date", "engine_version", "workbook_contract",
    # WHICH STAGE this workbook is at — `research` or `assessment`. Recorded
    # rather than inferred from the emptiness of column D, because a sheet
    # that is required at one stage and meaningless at the other cannot be
    # gated against a stage nobody wrote down. See STAGES / SHEET_STAGE.
    "stage",
    "evidence_mode", "sv_basis", "mode_basis", "lob_census", "kg_checksum",
    "created_at", "last_written_at", "checkpoint",
    # The binding preflight's digest. `sv_basis` and `mode_basis` are
    # RENDERED from the preflight document this hashes, so a basis that
    # drifts from the financial review and the recorded owner answer behind
    # it is detectable rather than merely unlikely.
    "preflight_sha",
    # The client folder, opened at run start rather than assembled at the
    # end: a run that stops early must still be findable in the intake tree.
    "client_folder", "client_folder_opened_at",
    # WHERE THE REQUEST CAME FROM, so the answer can go back to it.
    #
    # Assessment requests arrive in a Slack thread (#deal-desk) and are
    # finished when somebody replies IN THAT THREAD with the folder link. The
    # run that answers a request is started by one firing and finished by
    # another — often days later, certainly in another container — so the
    # thread has to travel with the run or the completion reply has nowhere
    # to go and the requester is left watching a thread that never closes.
    #
    # Additive: a run started any other way simply carries neither, and
    # `slack_thread_ts` empty means "there is no thread to answer", which is
    # the manual path and not a defect. No contract bump — the Run_Metadata
    # sheet is key/value rows, so an older workbook is missing a row rather
    # than the wrong shape.
    "slack_channel", "slack_thread_ts", "requested_by",
    # PRELIM: the preliminary research pass that grounds the Client Research
    # Profile. Category dispatch refuses while it is open.
    "prelim_status", "prelim_completed_at",
    # Sheets that are legitimately empty for THIS run, each with the reason.
    # The completeness gate reads this; an unlisted empty sheet blocks.
    "empty_sheet_reasons",
    # The templates this run is BOUND to: the sha256 of
    # references/templates/report_templates.json + workbook_template.json +
    # gold_reference.json at `start`. `engine.template bind` writes it;
    # orient refuses to serve a card while it is blank, so no run can begin
    # without the shape of its own deliverables pinned in the workbook.
    "template_binding",
    # THE RUN'S OWN CLOCK AND BILL (2026-09-03, owner issue 9: "the assessment
    # takes more than 6 hours" — and nothing recorded how long any stage took,
    # so the claim could be neither measured nor regressed). `engine.pipeline`
    # writes stage_timings at every stage boundary; `engine.cost record`
    # folds each dispatched lane into cost_summary. Both are JSON objects.
    "stage_timings", "cost_summary",
    # THE CONNECTOR SIDE OF SHIP-AS-YOU-GO (owner issue 7). The pipeline
    # ingests exactly twice — the SCORING_PASS checkpoint and the package —
    # and each ingest is a connector run version; these anchor which version
    # the staged pages belong to, and `promoted_at` is the run's end.
    "connector_run_id", "connector_run_id_prev", "connector_ingest_after_seq",
    "promoted_at",
    # Which engine.pipeline drove the run (blank for a hand-narrated one), so
    # a resumed run knows whether the STAGE_* gates in Gate_Log are its own.
    "pipeline_version",
)

# ── THE STAGE, which nothing recorded ────────────────────────────────────
#
# Measured 2026-08-30: there was no stage key ANYWHERE. `RUN_METADATA_KEYS`
# had none, `_write_metadata` wrote none, and the app INFERRED it from the
# emptiness of column D —
#
#     "stage": "research — column D is empty by contract" if not scores
#              else "assessment"
#
# — while the engine side made it a CLI opinion: `validator.validate` takes
# `expect_scores=False` unless `--expect-scores` is passed, and
# `assemble.package` calls it with the default, hard-coding research
# semantics for every package it ever builds.
#
# Inference is fine for a reader and useless for a GATE. A sheet that is
# required at one stage and meaningless at the other cannot be expressed
# against a stage nobody wrote down, and `REQUIRED_SHEETS = tuple(SHEETS)`
# made every declared sheet required at every stage. So the stage is a
# recorded fact now, and the inference stays as the app's fallback for the
# workbooks written before it.
STAGES = ("research", "assessment")

#: Sheets that belong to ONE stage. Everything not named here belongs to
#: both. A sheet out of its stage is NOT_APPLICABLE — neither populated nor
#: an omission, because there is nothing at this stage that could fill it.
SHEET_STAGE = {
    "Pillar_Summary": "assessment",
    "Category_Detail": "assessment",
    "Recommendations": "assessment",
    "Executive_Summary": "assessment",
    "Subcap_Scores": "assessment",
    "Pillar_Rollup": "assessment",
    "Category_Rollup": "assessment",
    "Pillar_Weights": "assessment",
    "Maturity_Rubric": "assessment",
    "Catalogue_Meta": "assessment",
    "Cap_Triggers": "assessment",
    "Caps_Applied_Log": "assessment",
    "Coverage_Map": "assessment",
    "Capability_Definitions": "assessment",
    "Solution_Catalogue": "assessment",
    "Platform_Peer_Adoption": "assessment",
}


def stage_of(metadata: dict) -> str:
    """The workbook's stage: recorded if it says, inferred if it is older.

    The inference is the app's own — scores present means assessment — and
    it is here so a v4 workbook upgraded in place reads correctly rather
    than reading as an assessment workbook with three empty tabs.
    """
    got = str((metadata or {}).get("stage") or "").strip().lower()
    return got if got in STAGES else "research"

#: Sub-vertical pillar weights (dma-assessment SKILL.md § Sub-Vertical Pillar
#: Weights), the weight set the rollup states. Percent, summing to 100.
PILLAR_WEIGHTS = {
    "CU":  {"P1": 25, "P2": 30, "P3": 20, "P4": 25},
    "RB":  {"P1": 25, "P2": 30, "P3": 20, "P4": 25},
    "CL":  {"P1": 20, "P2": 20, "P3": 35, "P4": 25},
    "CIB": {"P1": 20, "P2": 20, "P3": 35, "P4": 25},
    "IC":  {"P1": 20, "P2": 20, "P3": 30, "P4": 30},
    "IB":  {"P1": 20, "P2": 35, "P3": 20, "P4": 25},
    "WM":  {"P1": 25, "P2": 30, "P3": 20, "P4": 25},
    "RIA": {"P1": 25, "P2": 30, "P3": 20, "P4": 25},
    "AM":  {"P1": 20, "P2": 30, "P3": 25, "P4": 25},
    "FC":  {"P1": 25, "P2": 30, "P3": 20, "P4": 25},
}
DEFAULT_PILLAR_WEIGHTS = {"P1": 25, "P2": 25, "P3": 25, "P4": 25}

#: Pillar display names, as the catalogue and the reports name them.
PILLAR_NAMES = {
    "P1": "Strategy, Governance & Culture",
    "P2": "Member/Customer Experience & Engagement",
    "P3": "Operations, Risk & Compliance",
    "P4": "Data, Analytics & Technology",
}


#: v6 (2026-09-03) is v5 plus the sheets the gold-standard workbook carries
#: and the app reads — the client's own facts (Firmographics, Focus_Areas,
#: Issue_Register, Enrichment_Needed) at the research stage and the scoring
#: stage's dashboard, rollups, weights, rubric, caps, overlay and catalogue
#: (Executive_Summary … Platform_Peer_Adoption) — plus Gap_to_Peer / Maturity
#: on the two stated grains. Additive; a v5 workbook upgrades in place.
WORKBOOK_CONTRACT = "v7"
ENGINE_VERSION = "7.0.0"

SHEETS = {
    "00_README": ("Key", "Value"),
    "DQ_Bank": DQ_BANK_COLUMNS,
    "P1_Subcap_Scoring": PILLAR_COLUMNS,
    "P2_Subcap_Scoring": PILLAR_COLUMNS,
    "P3_Subcap_Scoring": PILLAR_COLUMNS,
    "P4_Subcap_Scoring": PILLAR_COLUMNS,
    "Evidence_Detail": EVIDENCE_COLUMNS,
    "Entity_Timeline": TIMELINE_COLUMNS,
    "Tech_Register": TECH_REGISTER_COLUMNS,
    "Tech_Peer_Deployments": TECH_PEER_COLUMNS,
    "Pillar_Summary": PILLAR_SUMMARY_COLUMNS,
    "Category_Detail": CATEGORY_DETAIL_COLUMNS,
    "Recommendations": RECOMMENDATIONS_COLUMNS,
    "Coverage": COVERAGE_COLUMNS,
    "Search_Log": SEARCH_LOG_COLUMNS,
    "Gate_Log": GATE_LOG_COLUMNS,
    "Report_Narrative": REPORT_NARRATIVE_COLUMNS,
    "Provenance": PROVENANCE_COLUMNS,
    "Challenge_Log": CHALLENGE_LOG_COLUMNS,
    "Handoff_Lock": HANDOFF_LOCK_COLUMNS,
    "Peer_Benchmarks": PEER_BENCHMARK_COLUMNS,
    "Run_Metadata": RUN_METADATA_COLUMNS,
    "REF_Method": ("Key", "Value"),
    # ── the client's own facts (research stage; Client Profile §1, §6, §7) ──
    "Firmographics": FIRMOGRAPHICS_COLUMNS,
    "Focus_Areas": FOCUS_AREAS_COLUMNS,
    "Issue_Register": ISSUE_REGISTER_COLUMNS,
    "Enrichment_Needed": ENRICHMENT_NEEDED_COLUMNS,
    # Added 2026-09-03 (contract v7). The five-year trajectory the gold
    # standard carries had no sheet and no writer; `engine.profile financial`
    # fills it in PRELIM and both reports render it.
    "Financial_Trends": FINANCIAL_TRENDS_COLUMNS,
    # ── the scoring stage (assessment; the report's §1-§5, §8) ────────────
    "Executive_Summary": EXECUTIVE_SUMMARY_COLUMNS,
    "Subcap_Scores": SUBCAP_SCORES_COLUMNS,
    "Pillar_Rollup": PILLAR_ROLLUP_COLUMNS,
    "Category_Rollup": CATEGORY_ROLLUP_COLUMNS,
    "Pillar_Weights": PILLAR_WEIGHTS_COLUMNS,
    "Maturity_Rubric": MATURITY_RUBRIC_COLUMNS,
    "Catalogue_Meta": CATALOGUE_META_COLUMNS,
    "Cap_Triggers": CAP_TRIGGERS_COLUMNS,
    "Caps_Applied_Log": CAPS_APPLIED_LOG_COLUMNS,
    "Coverage_Map": COVERAGE_MAP_COLUMNS,
    "Capability_Definitions": CAPABILITY_DEFINITIONS_COLUMNS,
    "Solution_Catalogue": SOLUTION_CATALOGUE_COLUMNS,
    "Platform_Peer_Adoption": PLATFORM_PEER_ADOPTION_COLUMNS,
}

PILLAR_SHEETS = tuple(f"{p}_Subcap_Scoring" for p in ("P1", "P2", "P3", "P4"))

#: Sheets whose absence is rule 1. Every sheet in SHEETS is required — the
#: archive's validator required a set that the pinned template had already
#: retired (AUD-0012, AUD-0061), so the required set and the generated set
#: are now the same object and cannot disagree.
REQUIRED_SHEETS = tuple(SHEETS)

#: The literal the contract forbids in Source_URLs, in any casing (rule 6).
BANNED_URL_PLACEHOLDERS = ("multiple searches", "see report", "various",
                           "n/a", "tbd", "see above")

NO_EVIDENCE = "NO_EVIDENCE"

TIERS = ("T1", "T2", "T3", "T4", "T5", NO_EVIDENCE)
CLAIM_LABELS = ("FACT", "INFERENCE", "HYPOTHESIS", "CEILING_ESTIMATE")
FACETS = ("works", "fails", "value", "contradicts", "corroborates")

#: The AI overlay, measured from the pinned workbook's own DQ_Bank: its
#: question set is "4255 catalogue + 2553 AI overlay" — 851 x 5 facet probes
#: and 851 x 3 overlay questions exactly, so EVERY subcap carries all three.
AI_FACETS = ("ai_deployment", "ai_data", "ai_constraint")
ALL_FACETS = FACETS + AI_FACETS

#: The DQ that seeds the rest: the Pillar Scoring Toolkit's own Diagnostic
#: Question for the subcap (column H), regraded to an open form where the
#: toolkit wrote it closed.
PRIMARY_FACET = "primary"
DQ_FACETS = (PRIMARY_FACET,) + ALL_FACETS

#: How a run gathers evidence. PUBLIC = web only; INTERNAL = client-supplied
#: documents only; HYBRID = both. The mode decides WHICH diagnostic
#: questions are answerable and which are deferred to discovery — it never
#: silently drops one.
ASSESSMENT_MODES = ("PUBLIC", "INTERNAL", "HYBRID")

#: What a DQ's own toolkit row declares about where its answer lives.
#: Measured from Pillar1_Scoring_Toolkit.xlsx (Source Type column): the
#: values shipped are "Both" and "Public Only"; "Internal Only" is the third
#: member the vocabulary admits. Normalised here to a closed enum.
DQ_MODE_FIT = ("PUBLIC", "INTERNAL", "BOTH")

#: mode -> the mode_fit values whose DQs are ANSWERABLE in that mode.
#: Everything else is DEFERRED: emitted as an INT-Q (or PUB-Q) discovery
#: question with the reason, never silently dropped.
MODE_ANSWERABLE = {
    "PUBLIC": ("PUBLIC", "BOTH"),
    "INTERNAL": ("INTERNAL", "BOTH"),
    "HYBRID": ("PUBLIC", "INTERNAL", "BOTH"),
}

#: The evidence-recency ladder. Undated evidence is UNVERIFIED, never
#: current (invariant 9) — so UNVERIFIED is a member here, not an absence.
RECENCY_LADDER = (
    ("CURRENT", 12), ("RECENT", 24), ("DATED", 36), ("LEGACY", 48),
)
RECENCY_ARCHIVAL = "ARCHIVAL"
RECENCY_UNVERIFIED = "UNVERIFIED"


def catalogue_hash() -> str:
    """A stable digest of the catalogue this run is pinned to.

    AUD-0060: the Client Profile asserts "the hash recorded here is written
    into Handoff_Lock; the assessment stage compares against it and refuses
    to score if the catalogue has moved" — and nothing computed a hash. This
    is that hash. It digests the cell ids and their tiers, so a renamed cell
    or a moved sub-vertical changes it and a reformatted file does not."""
    import hashlib
    t = taxonomy()
    payload = "\n".join(f"{c}\t{t.tier[c]}" for c in t.cells)
    return hashlib.sha256(payload.encode()).hexdigest()


def counts() -> dict:
    """Every count the prose used to hardcode, measured.

    Renderers and skills interpolate from here. AUD-0062's rule — "never
    write one as a literal in prose" — is enforceable only if there is a
    call to make instead."""
    t = taxonomy()
    return {
        "catalogue_version": t.version,
        "catalogue_hash": catalogue_hash(),
        "pillars": t.n_pillars,
        "categories": t.n_categories,
        "capabilities": t.n_capabilities,
        "cells": t.n_cells,
        "universal": t.n_universal,
        "sub_vertical_variants": t.n_variants,
        "bands": len(BANDS),
        "sub_verticals": len(t.sub_vertical_codes()),
    }


def compare_to_workbook(path) -> list[str]:
    """Divergences between a real workbook and this contract, as sentences.

    Exists because the pinned template is a Drive document no code path can
    resolve (AUD-0069). When someone does export it, this reports what
    differs instead of a validator silently accepting whichever shape it
    happens to meet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    out = []
    have = set(wb.sheetnames)
    for name in REQUIRED_SHEETS:
        if name not in have:
            out.append(f"sheet missing: {name}")
    for extra in sorted(have - set(SHEETS)):
        out.append(f"sheet not in contract: {extra}")
    for name, cols in SHEETS.items():
        if name not in have:
            continue
        ws = wb[name]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        got = tuple((c or "").strip() for c in row if c is not None)
        if got != tuple(cols):
            out.append(
                f"{name}: header differs — contract {list(cols)} vs workbook {list(got)}")
    wb.close()
    return out


if __name__ == "__main__":  # a one-line answer to "what is the shape?"
    import argparse
    import sys
    ap = argparse.ArgumentParser(
        description=__doc__.split("\n")[0],
        epilog="With no --workbook this prints the counts, measured from the "
               "catalogue. With one, it prints how that workbook diverges "
               "from the contract — which is the honest answer when the "
               "pinned template is exported and compared (AUD-0069).")
    ap.add_argument("--workbook", help="compare this workbook to the contract")
    a = ap.parse_args()
    if a.workbook:
        for line in compare_to_workbook(a.workbook) or ["contract: no divergence"]:
            print(line)
        sys.exit(0)
    print(json.dumps(counts(), indent=2))
