#!/usr/bin/env python3
"""The migration the template's own changelog prescribes and that was never written.

    patch_validator.py --workbook WB.xlsx [--apply]

WHY THIS EXISTS. AUD-0011 / AUD-0061: the pinned template's changelog row for
version 2.0.0 records that merging `Subcap_Synthesis` and `Negative_Findings`
onto the pillar tabs "Breaks: YES: run patch_validator.py". `find / -name
'patch_validator*'` returned zero files. So the workbook shape moved, the
validator did not, and there was no documented path from the template to a
passing validation — the authority artefact could not pass the gate meant to
admit it.

This is that path. It reports what a pre-v3 workbook needs, and with --apply
it performs the migration: the two retired sheets are folded into the pillar
sheets' working area and the missing contract sheets are created. It is
idempotent — a workbook already at v3 reports "no change" and touches nothing.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

from . import contract as C

RETIRED_SHEETS = ("Subcap_Synthesis", "Negative_Findings")

#: Where a retired sheet's columns land in the working area. Anything not
#: named here is reported as UNMAPPED rather than dropped — a migration that
#: silently loses a column is the defect it exists to prevent.
FOLD = {
    "Subcap_Synthesis": {
        "dominant_claim": "Dominant_Claim", "claim_label": "Claim_Label",
        "what_we_found": "What_We_Found", "triangulation": "Triangulation",
        "ceiling_reasoning": "Ceiling_Reasoning",
        "why_it_matters": "Why_It_Matters", "dma_impact": "DMA_Impact",
        "facet_coverage": "Facet_Coverage",
        "contradiction_disposition": "Contradiction_Disposition",
        "discovery_questions": "Discovery_Questions",
        "ceiling_band": "Ceiling_Band", "uncertainty": "Uncertainty",
    },
    "Negative_Findings": {
        "proxy_log": "Proxy_Log", "ladder": "Negative_Ladder",
        "negative_ladder": "Negative_Ladder",
        "absence_claimed": "Absence_Claimed",
    },
}


def _norm(s) -> str:
    return "".join(ch for ch in str(s or "").lower() if ch.isalnum() or ch == "_")


def plan(workbook) -> dict:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    try:
        have = set(wb.sheetnames)
        actions, unmapped = [], []
        for s in RETIRED_SHEETS:
            if s in have:
                ws = wb[s]
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                cols = [_norm(c) for c in hdr if c is not None]
                fold = FOLD.get(s, {})
                actions.append({"fold_sheet": s, "rows": max(0, ws.max_row - 1),
                                "into": "pillar working area"})
                unmapped += [f"{s}.{c}" for c in cols
                             if c and c not in fold and c not in
                             ("subcap_id", "subcapid", "id")]
        missing = [s for s in C.REQUIRED_SHEETS if s not in have]
        widen = []
        for s in C.PILLAR_SHEETS:
            if s in have:
                n = wb[s].max_column
                if n < len(C.PILLAR_COLUMNS):
                    widen.append({"sheet": s, "from": n,
                                  "to": len(C.PILLAR_COLUMNS)})
        return {
            "workbook": str(workbook),
            "already_v3": not actions and not missing and not widen,
            "fold": actions, "create_sheets": missing, "widen": widen,
            "unmapped_columns": sorted(set(unmapped)),
        }
    finally:
        wb.close()


def apply(workbook) -> dict:
    p = plan(workbook)
    if p["already_v3"]:
        return {**p, "applied": False, "note": "no change"}
    if p["unmapped_columns"]:
        raise SystemExit(
            "REFUSED: these retired-sheet columns have no home in the "
            f"contract-v3 working area: {p['unmapped_columns']}. Map them in "
            "FOLD, or the migration would drop them silently — which is the "
            "class of defect this script exists to close.")
    wb = openpyxl.load_workbook(workbook)
    # widen and create first, so the fold has somewhere to land
    for name in p["create_sheets"]:
        ws = wb.create_sheet(name)
        ws.append(list(C.SHEETS[name]))
    for w in p["widen"]:
        ws = wb[w["sheet"]]
        for i in range(w["from"] + 1, w["to"] + 1):
            ws.cell(row=1, column=i, value=C.PILLAR_COLUMNS[i - 1])
    idx = {c: i + 1 for i, c in enumerate(C.PILLAR_COLUMNS)}
    folded = 0
    for s in RETIRED_SHEETS:
        if s not in wb.sheetnames:
            continue
        src = wb[s]
        hdr = [_norm(c) for c in
               next(src.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        fold = FOLD.get(s, {})
        for row in src.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(hdr, row))
            sub = str(rec.get("subcap_id") or rec.get("subcapid")
                      or rec.get("id") or "").strip()
            if not sub:
                continue
            sheet = f"{sub[:2]}_Subcap_Scoring"
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=1).value or "").strip() == sub:
                    for k, dest in fold.items():
                        if rec.get(k) is not None and dest in idx:
                            ws.cell(row=r, column=idx[dest], value=rec[k])
                    folded += 1
                    break
        del wb[s]
    wb.save(workbook)
    wb.close()
    return {**p, "applied": True, "rows_folded": folded}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args(argv)
    out = apply(a.workbook) if a.apply else plan(a.workbook)
    print(json.dumps(out, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
