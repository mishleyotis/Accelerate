#!/usr/bin/env python3
"""The substrate. Every research step records here, as it goes.

WHY THIS EXISTS. AUD-0001: `grep -rln 'openpyxl|load_workbook' scripts/engine/`
returned zero files — no research, synthesis, floors, followup or ledger step
could touch a sheet. The workbook had exactly one producer, which built a
fresh `Workbook()` and performed its single `.save()` at the end, from a
parallel JSON plane. Four other findings were siblings of that one root: the
empty CHAIN INTEGRITY block, the checksum nothing wrote, the governance
auditor reading sheets that differ from the export, and the resume that
recovered nothing.

This module is the answer, and it is deliberately the ONLY writer. Open the
run's workbook, append the row, save. A crash loses the current row and
nothing before it. A fresh container opens the same file and reads the same
state — which is what makes the documented resume (AUD-0010) possible at all.

Two disciplines are enforced here rather than left to callers:

  * **Append, never rebuild.** `open_or_create` opens an existing workbook.
    There is no code path that constructs a second `Workbook()` over a run.
  * **Atomic save.** Written to a sibling temp file and `os.replace`d, so a
    killed process leaves the previous good workbook rather than a truncated
    one. `os.replace` is atomic within a filesystem; the temp file is placed
    beside the target for exactly that reason.
"""
from __future__ import annotations

# Runnable both ways. `python3 -m engine.<mod>` is the documented invocation,
# but every audit and every operator reaches for `python3 <path> --help`
# first, and a relative import dies there. Binding __package__ makes the two
# equivalent instead of making one of them a trap.
if __package__ in (None, ""):  # noqa: E402  (must precede the relative imports)
    import os as _os
    import sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.dirname(
        _os.path.abspath(__file__))))
    __package__ = "engine"

import datetime as _dt
import os
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from . import contract as C


def _utcnow() -> str:
    return _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


class WorkbookError(RuntimeError):
    pass


class RunWorkbook:
    """One run's scoring workbook, held open for append.

    Not a context manager by accident: a `with` block invites "do all the
    writes, then save", which is the batching that made the workbook a
    terminal export in the first place. Each `append_*` saves."""

    def __init__(self, path: str | Path, *, autosave: bool = True):
        self.path = Path(path)
        self.autosave = autosave
        if not self.path.exists():
            raise WorkbookError(
                f"{self.path} does not exist — use RunWorkbook.create()")
        self._wb = openpyxl.load_workbook(self.path)
        notes = self._upgrade_shape()
        missing = [s for s in C.REQUIRED_SHEETS if s not in self._wb.sheetnames]
        if missing:
            raise WorkbookError(
                f"{self.path} is not a contract-{C.WORKBOOK_CONTRACT} workbook; "
                f"missing sheets: {', '.join(missing)}")
        if notes:
            self._record_upgrade(notes)

    # ── shape upgrade: expand, migrate, contract ─────────────────────────

    def _upgrade_shape(self) -> list[str]:
        """Bring an earlier-contract workbook up to the current shape.

        A run's workbook lives in the client's Drive folder for as long as
        the engagement does, so a contract bump that REFUSED every workbook
        written before it would strand every open run — the engine would
        have a shape nothing in the field could satisfy. Additive changes
        are therefore migrated in place, on open, once.

        Sheets are REWRITTEN in contract order rather than having columns
        appended, because `contract.verify` compares the header row as an
        ordered tuple: a column inserted in the middle of the contract (as
        `Providers` was) cannot be reconciled by appending. Values move by
        NAME, so nothing is read positionally and a new column arrives
        empty, which is a readable state.
        """
        notes = []
        for name, cols in C.SHEETS.items():
            if name not in self._wb.sheetnames:
                self._wb.create_sheet(name).append(list(cols))
                notes.append(f"added sheet {name}")
                continue
            ws = self._wb[name]
            head = [str(c.value).strip() if c.value is not None else ""
                    for c in next(ws.iter_rows(min_row=1, max_row=1), ())]
            if tuple(h for h in head if h) == tuple(cols):
                continue
            body = [{head[i]: r[i] for i in range(min(len(head), len(r)))}
                    for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(c is not None for c in r)]
            added = [c for c in cols if c not in head]
            dropped = [h for h in head if h and h not in cols]
            idx = self._wb.sheetnames.index(name)
            del self._wb[name]
            new = self._wb.create_sheet(name, idx)
            new.append(list(cols))
            for row in body:
                new.append([row.get(c) for c in cols])
            notes.append(
                f"{name}: " + ", ".join(
                    ([f"added {', '.join(added)}"] if added else [])
                    + ([f"no longer in contract: {', '.join(dropped)}"]
                       if dropped else [])
                    + ([f"reordered to contract order"]
                       if not added and not dropped else []))
                + f" ({len(body)} row(s) moved by name)")
        return notes

    def _record_upgrade(self, notes: list[str]) -> None:
        """Say so, in the workbook, where the next reader will see it."""
        stamp = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        for note in notes:
            self.append("00_README",
                        {"Key": f"shape_upgraded_{stamp}", "Value": note},
                        save=False)
        # The lock recorded the contract this workbook was BUILT under. It is
        # now this one, and an upgrade that left the lock behind would read
        # as undetected drift to `verify_handoff_lock` — which is exactly the
        # signal that must stay meaningful for real drift.
        for row in self._wb["Handoff_Lock"].iter_rows(min_row=2):
            if str(row[0].value) == "workbook_contract":
                row[1].value = C.WORKBOOK_CONTRACT
        self.save()

    # ── construction ─────────────────────────────────────────────────────

    @classmethod
    def create(cls, path, *, run_id: str, entity_name: str, entity_id: str,
               sub_vertical: str | None, scope_mode: str,
               reference_date: str, selected: list[str] | None = None,
               evidence_mode: str = "PUBLIC",
               sv_basis: str | None = None, mode_basis: str | None = None,
               lob_census: str | None = None,
               overwrite: bool = False) -> "RunWorkbook":
        """Build the workbook once, at the START of the run, with its
        metadata resolved.

        AUD-0010 found the two anchors a resumed run would need shipping as
        unfilled template tokens — run_id `{{RUN_ID}}` and kg_checksum
        `{{CHECKSUM}}`. Neither can be a token here: both are arguments or
        computed, and `_write_metadata` refuses a value that still looks
        like a template placeholder."""
        path = Path(path)
        if path.exists() and not overwrite:
            raise WorkbookError(f"{path} already exists; opening is the "
                                f"resume path, not creating")
        tax = C.taxonomy()
        if selected is None:
            selected = list(tax.selected(sub_vertical, scope_mode))
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name, cols in C.SHEETS.items():
            ws = wb.create_sheet(name)
            ws.append(list(cols))
            ws.freeze_panes = "A2"
            for i, col in enumerate(cols, start=1):
                ws.column_dimensions[get_column_letter(i)].width = \
                    max(12, min(40, len(col) + 4))
        path.parent.mkdir(parents=True, exist_ok=True)
        _atomic_save(wb, path)
        self = cls(path)
        self._write_readme()
        self._write_metadata(run_id=run_id, entity_name=entity_name,
                             entity_id=entity_id, sub_vertical=sub_vertical,
                             scope_mode=scope_mode,
                             reference_date=reference_date,
                             selected=selected, evidence_mode=evidence_mode,
                             sv_basis=sv_basis, mode_basis=mode_basis,
                             lob_census=lob_census)
        self._seed_scoring_rows(selected)
        self._write_handoff_lock()
        self.recompute_coverage()
        return self

    # ── low-level ────────────────────────────────────────────────────────

    def _sheet(self, name: str):
        if name not in self._wb.sheetnames:
            raise WorkbookError(f"no sheet {name!r}")
        return self._wb[name]

    def save(self) -> None:
        _atomic_save(self._wb, self.path)

    def _touch(self) -> None:
        self.set_metadata("last_written_at", _utcnow(), save=False)

    def append(self, sheet: str, row: dict, *, save: bool | None = None) -> int:
        """Append one row to `sheet`, keyed by the contract's column names.

        A key the contract does not declare is refused rather than dropped:
        AUD-0067's whole cost was a column that existed, was censused as
        present, and was read by nobody."""
        cols = C.SHEETS[sheet]
        unknown = [k for k in row if k not in cols]
        if unknown:
            raise WorkbookError(
                f"{sheet}: no such column(s) {unknown}; contract is {list(cols)}")
        ws = self._sheet(sheet)
        ws.append([_cell(row.get(c)) for c in cols])
        self._touch()
        if save if save is not None else self.autosave:
            self.save()
        return ws.max_row

    def rows(self, sheet: str) -> list[dict]:
        """Every data row of `sheet` as a dict keyed by the contract."""
        cols = C.SHEETS[sheet]
        ws = self._sheet(sheet)
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in r):
                continue
            out.append({c: r[i] if i < len(r) else None
                        for i, c in enumerate(cols)})
        return out

    def update_row(self, sheet: str, key_col: str, key: str, values: dict,
                   *, save: bool | None = None) -> int:
        """Set named cells on the row whose `key_col` equals `key`."""
        cols = list(C.SHEETS[sheet])
        unknown = [k for k in values if k not in cols]
        if unknown:
            raise WorkbookError(f"{sheet}: no such column(s) {unknown}")
        ws = self._sheet(sheet)
        kidx = cols.index(key_col) + 1
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=kidx).value or "").strip() == key:
                for k, v in values.items():
                    ws.cell(row=r, column=cols.index(k) + 1, value=_cell(v))
                self._touch()
                if save if save is not None else self.autosave:
                    self.save()
                return r
        raise WorkbookError(f"{sheet}: no row where {key_col} == {key!r}")

    def update_row_where(self, sheet: str, match: dict, values: dict,
                         *, save: bool | None = None) -> int:
        """Set named cells on the row matching EVERY key in `match`.

        `update_row` keys on one column, which is a silent corruption for
        any sheet whose row identity is composite. Report_Narrative is
        exactly that: both reports number their sections "1"–"8", so
        rewriting the research profile's §1 walked the sheet, found the
        assessment's §1 first, and overwrote it — relabelling the victim as
        belonging to the other report, because the values dict carries
        `Report` too.
        """
        cols = list(C.SHEETS[sheet])
        unknown = [k for k in list(values) + list(match) if k not in cols]
        if unknown:
            raise WorkbookError(f"{sheet}: no such column(s) {unknown}")
        ws = self._sheet(sheet)
        idx = {k: cols.index(k) + 1 for k in match}
        for r in range(2, ws.max_row + 1):
            if all(str(ws.cell(row=r, column=i).value or "").strip()
                   == str(match[k] or "").strip() for k, i in idx.items()):
                for k, v in values.items():
                    ws.cell(row=r, column=cols.index(k) + 1, value=_cell(v))
                self._touch()
                if save if save is not None else self.autosave:
                    self.save()
                return r
        raise WorkbookError(f"{sheet}: no row where {match!r}")

    # ── metadata, and the two anti-drift anchors ─────────────────────────

    _TOKENISH = ("{{", "}}", "TODO", "TBD")

    #: The honest value for a binding rationale nobody recorded. It says so
    #: in words rather than posing as a rationale, so a reader (and the
    #: resume report) can tell "bound with a reason" from "bound bare
    #: through the API".
    _UNSTATED_BASIS = ("UNSTATED — bound through the Python API without a "
                       "recorded rationale (engine.cli start records one)")

    def _write_metadata(self, *, run_id, entity_name, entity_id, sub_vertical,
                        scope_mode, reference_date, selected,
                        evidence_mode="PUBLIC", sv_basis=None,
                        mode_basis=None, lob_census=None) -> None:
        if evidence_mode not in C.ASSESSMENT_MODES:
            raise WorkbookError(
                f"evidence_mode {evidence_mode!r} is not one of "
                f"{C.ASSESSMENT_MODES}. The mode decides which diagnostic "
                f"questions are answerable, so a run cannot start without "
                f"declaring one.")
        tax = C.taxonomy()
        vals = {
            "run_id": run_id,
            "entity_name": entity_name,
            "entity_id": entity_id,
            "sub_vertical": sub_vertical or "",
            "scope_mode": scope_mode,
            "catalogue_version": tax.version,
            "catalogue_hash": C.catalogue_hash(),
            "taxonomy_pillars": tax.n_pillars,
            "taxonomy_categories": tax.n_categories,
            "taxonomy_capabilities": tax.n_capabilities,
            "taxonomy_cells": tax.n_cells,
            "subcaps_selected": len(selected),
            "reference_date": reference_date,
            "engine_version": C.ENGINE_VERSION,
            "workbook_contract": C.WORKBOOK_CONTRACT,
            # RESEARCH until something scores it. Column D is empty at this
            # stage by contract rule 4, and the three grain tabs the
            # assessment fills are NOT_APPLICABLE until it does.
            "stage": "research",
            "evidence_mode": evidence_mode,
            # The binding provenance. Which sub-vertical the entity was bound
            # to and which evidence mode the engagement runs under are the
            # two choices that decide WHAT gets researched (165 variant cells
            # ride on the first; every DQ's askability on the second) — so
            # each records WHY, verbatim from the caller. "UNSTATED" is the
            # honest value for an API-created run; the CLI path requires a
            # real rationale and runstate refuses a placeholder one.
            "sv_basis": sv_basis or self._UNSTATED_BASIS,
            "mode_basis": mode_basis or self._UNSTATED_BASIS,
            "lob_census": lob_census or "",
            # Written by kg.build once the DQ bank is seeded; empty is the
            # honest value for a run whose KG has not been built, and the
            # resume path REPORTS it rather than treating it as fine.
            "kg_checksum": "",
            "created_at": _utcnow(),
            "last_written_at": _utcnow(),
            "checkpoint": "",
            # Filled by the phases that own them, and EMPTY is a readable
            # state rather than a missing key: an unopened client folder, an
            # unrun preflight and an open PRELIM each have a gate that names
            # them, so a blank here is a finding, not a crash.
            "preflight_sha": "",
            "client_folder": "",
            "client_folder_opened_at": "",
            "prelim_status": "OPEN",
            "prelim_completed_at": "",
            "empty_sheet_reasons": "",
        }
        for k in ("run_id", "entity_name", "entity_id", "reference_date"):
            v = str(vals[k])
            if not v.strip() or any(t in v for t in self._TOKENISH):
                raise WorkbookError(
                    f"Run_Metadata.{k} = {v!r} is an unresolved placeholder. "
                    f"A resumed run reads these two anchors and nothing else; "
                    f"an unfilled token is the AUD-0010 failure.")
        ws = self._sheet("Run_Metadata")
        for k in C.RUN_METADATA_KEYS:
            ws.append([k, _cell(vals[k])])
        self.save()

    def metadata(self) -> dict:
        return {str(r["Key"]): r["Value"] for r in self.rows("Run_Metadata")}

    def set_metadata(self, key: str, value, *, save: bool = True) -> None:
        if key not in C.RUN_METADATA_KEYS:
            raise WorkbookError(f"Run_Metadata has no key {key!r}")
        ws = self._sheet("Run_Metadata")
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "") == key:
                ws.cell(row=r, column=2, value=_cell(value))
                break
        else:
            ws.append([key, _cell(value)])
        if save:
            self.save()

    def _write_handoff_lock(self) -> None:
        """The lock the Client Profile asserts and nothing built (AUD-0060).

        The assessment stage compares these values and refuses to score if
        the catalogue has moved. It is a real comparison because the hash is
        a real digest of the cells and their tiers."""
        md = self.metadata()
        for k, v in (
            ("catalogue_version", md.get("catalogue_version")),
            ("catalogue_hash", md.get("catalogue_hash")),
            ("workbook_contract", C.WORKBOOK_CONTRACT),
            ("engine_version", C.ENGINE_VERSION),
            ("run_id", md.get("run_id")),
            ("locked_at", _utcnow()),
            ("handoff_status", "IN_PROGRESS"),
        ):
            self.append("Handoff_Lock", {"Key": k, "Value": v}, save=False)
        self.save()

    def lock_peer_set(self, peers: list[str], *, basis: str) -> dict:
        """Freeze the peer set into Handoff_Lock, once.

        AUD-0043: `Handoff_Lock` is the peer-set immutability mechanism both
        pinned templates depend on, and it existed nowhere — 0 hits across
        the archive and 0 across the repository. Without it the assessment
        stage can silently choose a different cohort from the one the
        research stage compared against, and every peer figure in the report
        is then about a set nobody can name.

        Locking is idempotent for the SAME set and refused for a different
        one. A cohort that changes mid-assessment is a decision, and a
        decision needs a person, not a second write."""
        if basis not in C.PEER_BASIS:
            raise WorkbookError(f"peer basis {basis!r} not in {C.PEER_BASIS}")
        wanted = sorted({str(p).strip() for p in peers if str(p).strip()})
        lock = self.handoff_lock()
        have = [p for p in str(lock.get("locked_peer_set") or "").split("|") if p]
        if have and have != wanted:
            raise WorkbookError(
                f"the peer set is already locked to {have} and this call "
                f"names {wanted}. A cohort that changes mid-assessment "
                f"invalidates every peer figure already written against it; "
                f"withdraw the run and re-research, or keep the locked set.")
        if not have:
            self.append("Handoff_Lock",
                        {"Key": "locked_peer_set", "Value": "|".join(wanted)},
                        save=False)
            self.append("Handoff_Lock",
                        {"Key": "peer_basis", "Value": basis}, save=False)
            self.append("Handoff_Lock",
                        {"Key": "peer_n", "Value": len(wanted)})
        return {"locked_peer_set": wanted, "peer_basis": basis,
                "peer_n": len(wanted), "already_locked": bool(have)}

    def handoff_lock(self) -> dict:
        return {str(r["Key"]): r["Value"] for r in self.rows("Handoff_Lock")}

    def verify_handoff_lock(self) -> list[str]:
        """Divergences between the lock and the catalogue as it is NOW."""
        lock = self.handoff_lock()
        out = []
        if lock.get("catalogue_hash") != C.catalogue_hash():
            out.append(
                f"catalogue has moved since this run was locked: workbook "
                f"{lock.get('catalogue_hash')!r} vs current "
                f"{C.catalogue_hash()!r}")
        if lock.get("workbook_contract") != C.WORKBOOK_CONTRACT:
            out.append(f"workbook contract {lock.get('workbook_contract')!r} "
                       f"!= engine's {C.WORKBOOK_CONTRACT!r}")
        return out

    def _write_readme(self) -> None:
        for k, v in (
            ("artefact", "DMA Scoring Workbook"),
            ("contract", C.WORKBOOK_CONTRACT),
            ("substrate", "This workbook IS the record. Every research step "
                          "appends here as it goes; it is not an export."),
            ("scores", "Column D is EMPTY at the research stage. A value in D "
                       "before assessment is contract rule 4."),
            ("working_area", "Columns L..AG carry the synthesis. Strip them "
                             "with engine/strip_working_area.py, which first "
                             "proves the three analysis fields survive."),
            ("engine", C.ENGINE_VERSION),
        ):
            self.append("00_README", {"Key": k, "Value": v}, save=False)
        self._write_ref_method()
        self.save()

    def _write_ref_method(self) -> None:
        """The method the workbook is read under, IN the workbook.

        `REF_Method` was a declared sheet with no writer — the completeness
        gate found it empty in every run ever produced, which is what that
        gate is for. A reader who opens this file without the six design
        docs beside it needs the vocabularies the columns are scored against;
        rendering them from the contract means they cannot drift from it."""
        bands = " · ".join(
            f"<{i + 2} {b}" if i < 3 else f">={i + 1} {b}"
            for i, b in enumerate(C.BANDS))
        ladder = ", ".join(f"{n} <= {m}mo" for n, m in C.RECENCY_LADDER)
        for k, v in (
            ("catalogue", f"{C.taxonomy().version} — "
                          f"{C.counts()['cells']} cells, "
                          f"{C.counts()['categories']} categories, "
                          f"{C.counts()['pillars']} pillars"),
            ("catalogue_hash", C.catalogue_hash()),
            ("bands", f"{bands} (strict less-than, on the RAW score before "
                      f"display rounding; null score = no band)"),
            ("evidence_tiers", " > ".join(t for t in C.TIERS
                                          if t != C.NO_EVIDENCE)
                               + f"; {C.NO_EVIDENCE} where none was found"),
            ("recency", f"{ladder}, then {C.RECENCY_ARCHIVAL}; undated "
                        f"evidence is {C.RECENCY_UNVERIFIED}, never current"),
            ("claim_labels", ", ".join(C.CLAIM_LABELS)),
            ("evidence_modes", ", ".join(C.ASSESSMENT_MODES)
                               + " — decides which diagnostic questions are "
                                 "answerable and which defer to discovery"),
            ("challenge_dimensions", ", ".join(C.CHALLENGE_DIMENSIONS)),
            ("scores", "written by the ASSESSMENT stage, never by research; "
                       "a score present at research time is contract rule 4"),
        ):
            self.append("REF_Method", {"Key": k, "Value": v}, save=False)

    # ── the scoring rows ─────────────────────────────────────────────────

    def _seed_scoring_rows(self, selected: list[str]) -> None:
        """One row per SELECTED cell, on its pillar's sheet, scores empty.

        Seeding at creation is what makes scope knowable downstream. AUD-0014
        measured the app labelling 44 of 49 in-scope-but-unscored rows
        `toggled_out` — 'variant cells excluded by the toggle cascade' —
        because a row that is present and unscored was indistinguishable
        from a row that does not apply. A seeded row IS the scope
        declaration; `Evidence_IDs` starts at the literal NO_EVIDENCE so
        rule 5 is never vacuous on a blank (AUD-0064)."""
        tax = C.taxonomy()
        sv = str(self.metadata().get("sub_vertical") or "") or None
        by_sheet: dict[str, list[str]] = {s: [] for s in C.PILLAR_SHEETS}
        foreign = []
        for cell in selected:
            if cell not in tax.tier:
                raise WorkbookError(f"{cell} is not in catalogue {tax.version}")
            # A variant cell belongs to ONE sub-vertical. Seeding another
            # sub-vertical's variant produces a run that asks an institution
            # about capabilities defined for a different kind of institution,
            # and the app then — correctly — reports those rows as toggled
            # out, so the run looks 3 cells smaller than it is and nothing
            # says why. AUD-0077's family: the binder validated neither its
            # sub-vertical nor its scope.
            tier = tax.tier.get(cell, "")
            if "-" in tier and tier.split("-", 1)[1] != (sv or ""):
                foreign.append((cell, tier))
                continue
            by_sheet[f"{cell[:2]}_Subcap_Scoring"].append(cell)
        if foreign:
            raise WorkbookError(
                f"this run is sub-vertical {sv!r} and the engagement set names "
                f"{len(foreign)} variant cell(s) belonging to another: "
                f"{foreign[:6]}. Select with Taxonomy.selected(sv, scope), "
                f"which resolves the overlay for THIS sub-vertical and "
                f"withdraws the base cells it supersedes.")
        for sheet, cells in by_sheet.items():
            ws = self._sheet(sheet)
            for cell in sorted(cells):
                row = {c: None for c in C.PILLAR_COLUMNS}
                row["SubCap_ID"] = cell
                row["Category"] = cell.split(".")[0]
                row["Evidence_IDs"] = C.NO_EVIDENCE
                row["Proxy_Searched"] = "NOT_RUN"
                ws.append([_cell(row[c]) for c in C.PILLAR_COLUMNS])
        self.save()

    def scoring_rows(self) -> list[dict]:
        out = []
        for s in C.PILLAR_SHEETS:
            out.extend(self.rows(s))
        return out

    def scoring_row(self, subcap: str) -> dict | None:
        sheet = f"{subcap[:2]}_Subcap_Scoring"
        for r in self.rows(sheet):
            if str(r["SubCap_ID"]).strip() == subcap:
                return r
        return None

    def selected_subcaps(self) -> list[str]:
        """The engagement set, read back from the workbook itself.

        This is the answer to "in scope and unscored" vs "out of scope": a
        seeded row is in scope, full stop."""
        return [str(r["SubCap_ID"]).strip() for r in self.scoring_rows()
                if r.get("SubCap_ID")]

    def set_scoring(self, subcap: str, values: dict, *, save: bool = True) -> int:
        sheet = f"{subcap[:2]}_Subcap_Scoring"
        return self.update_row(sheet, "SubCap_ID", subcap, values, save=save)

    # ── coverage, computed ───────────────────────────────────────────────

    def recompute_coverage(self) -> list[dict]:
        """Rewrite Coverage from the pillar sheets and the evidence register.

        AUD-0059: the pinned template's seven CHAIN INTEGRITY verdict cells
        held the literal token `{{OK | INVESTIGATE}}`, two of its counts read
        sheets that did not exist and cached to zero forever, and one was the
        tautology `=4255/851`. Here the Verdict is a real value computed from
        real counts, recomputed on every call, and the counts come from rows
        that exist. Invariant 8: counts are computed, never stored where a
        source of truth exists."""
        tax = C.taxonomy()
        ev_by_subcap: dict[str, int] = {}
        for e in self.rows("Evidence_Detail"):
            for s in _split_ids(e.get("SubCap_IDs")):
                ev_by_subcap[s] = ev_by_subcap.get(s, 0) + 1
        stats: dict[str, dict] = {}
        for r in self.scoring_rows():
            cell = str(r.get("SubCap_ID") or "").strip()
            if not cell:
                continue
            cat = cell.split(".")[0]
            d = stats.setdefault(cat, {"selected": 0, "researched": 0,
                                       "items": 0, "floor_pass": 0,
                                       "synthesised": 0})
            d["selected"] += 1
            n = ev_by_subcap.get(cell, 0)
            d["items"] += n
            if n:
                d["researched"] += 1
            if n >= FLOOR_ITEMS:
                d["floor_pass"] += 1
            if _nonempty(r.get("Dominant_Claim")):
                d["synthesised"] += 1
        ws = self._sheet("Coverage")
        ws.delete_rows(2, max(0, ws.max_row - 1))
        out = []
        for cat in sorted(stats):
            d = stats[cat]
            pct = (d["floor_pass"] / d["selected"]) if d["selected"] else None
            verdict = _coverage_verdict(d, pct)
            row = {
                "Category_ID": cat, "Selected": d["selected"],
                "Researched": d["researched"], "Items": d["items"],
                "Floor_Pass": d["floor_pass"],
                "Floor_Pass_Pct": None if pct is None else round(pct, 4),
                "Synthesised": d["synthesised"], "Verdict": verdict,
            }
            ws.append([_cell(row[c]) for c in C.COVERAGE_COLUMNS])
            out.append(row)
        self.save()
        return out

    def coverage(self) -> list[dict]:
        return self.rows("Coverage")

    # ── evidence ─────────────────────────────────────────────────────────

    def next_evidence_id(self) -> str:
        """The next E-id, allocated from the workbook's own register.

        AUD-0133: the archive minted ids under an fcntl lock on a local
        counter file — safe only among processes on one host sharing one
        filesystem, and $RUN could not be shared at all. The workbook is the
        shared object, so the id comes from it. Two writers to one workbook
        is not a supported topology and never was; one writer reading its own
        maximum is correct and needs no lock."""
        n = 0
        for r in self.rows("Evidence_Detail"):
            eid = str(r.get("E_ID") or "")
            if eid.startswith("E-") and eid[2:].isdigit():
                n = max(n, int(eid[2:]))
        return f"E-{n + 1:03d}"

    def evidence_index(self) -> dict[str, dict]:
        return {str(r["E_ID"]): r for r in self.rows("Evidence_Detail")
                if r.get("E_ID")}


FLOOR_ITEMS = 3          # the per-subcap evidence floor
FLOOR_CATEGORY_ITEMS = 20  # the per-category minimum (AUD-0022)


def _coverage_verdict(d: dict, pct) -> str:
    """A verdict word, computed — never a template token."""
    if d["selected"] == 0:
        return "EMPTY"
    if d["items"] < FLOOR_CATEGORY_ITEMS:
        return "INVESTIGATE:category_items_below_floor"
    if pct is not None and pct < 0.8:
        return "INVESTIGATE:floor_pass_below_80pct"
    if d["synthesised"] < d["floor_pass"]:
        return "INVESTIGATE:synthesis_missing"
    return "OK"


def _split_ids(v) -> list[str]:
    if v is None:
        return []
    return [s.strip() for s in str(v).replace(";", ",").split(",") if s.strip()]


def _nonempty(v) -> bool:
    return v is not None and str(v).strip() != ""


def _cell(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        import json
        return json.dumps(v, separators=(",", ":"), sort_keys=True)
    return str(v)


def _atomic_save(wb, path: Path) -> None:
    """Save beside the target, then replace. A killed process leaves the
    previous good workbook, never a truncated one."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".xlsx.part")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


if __name__ == "__main__":  # a library, but it must answer --help
    import argparse as _ap
    _ap.ArgumentParser(
        prog=__file__.rsplit("/", 1)[-1],
        description=__doc__.split("\n")[0],
        epilog="A library module: import it, or run the modules that do have "
               "a command line (cli, orient, floors_gate, validator, handoff, "
               "reports, strip_working_area, patch_validator, watchdog).",
    ).parse_args()
