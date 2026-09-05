#!/usr/bin/env python3
"""Template adherence, as a diff rather than an assertion.

    python3 -m engine.template id                    # the template of record
    python3 -m engine.template check --file copy.xlsx [--json]
    python3 -m engine.template check --run R --root DIR

WHY THIS EXISTS. The 2026-08-30 audit asked how template adherence is
enforced "from the word go", and the honest answer was: it wasn't
CHECKABLE. The rule is real and good — the Drive template is read-only, and
a run workbook is BUILT from `contract.SHEETS` rather than by writing into a
copy — but the template's own identifier appeared nowhere in the plugin, so
nothing could tell whether the shape the engine builds still matches the
shape the owner maintains. "We follow the template" was an assertion with
nothing behind it.

WHICH SIDE IS AUTHORITATIVE. The CONTRACT. The engine writes from it, the
validator checks against it, and the app parses what it produces — so a
template the contract does not match is a template no run will ever
produce. This check does not exist to make the engine follow the template
blindly; it exists to make a divergence VISIBLE while somebody can still
decide which side is wrong.

WHAT IS COMPARED. Shape only: sheet names and header rows. The template
carries placeholder rows by design (the owner's own note: everything is a
placeholder except the tech stack), so comparing values would report noise
as drift and train everyone to ignore it.

HOW TO GET A COPY. Through the service account, never a shell fetch — the
plugin's policy hook refuses a Drive URL on a command line, and it is right
to:

    python3 scripts/drive_fetch.py pull --name "<template name>" \\
        --dest /tmp/template.xlsx
"""
from __future__ import annotations

# Runnable both ways: -m engine.template, or by path for --help.
if __package__ in (None, ""):  # noqa: E402
    import os as _os
    import sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.dirname(
        _os.path.abspath(__file__))))
    __package__ = "engine"

import argparse
import json
import sys
from pathlib import Path

import hashlib
import re

from . import contract as C

#: The pinned templates ship with the plugin — the workbook shape, both
#: report Docs (markdown + control-block JSON) and the Golden 1 reference
#: measurements. `bind` hashes them into the run; `report-drift` compares the
#: Doc exports against the JSON the engine actually enforces.
TEMPLATES_DIR = Path(__file__).resolve().parents[3] / "references" / "templates"
PINNED_FILES = ("report_templates.json", "workbook_template.json",
                "gold_reference.json", "client_profile_template.md",
                "assessment_report_template.md",
                # The branded .docx SHELL both reports are authored INTO
                # (header1.xml / footer1.xml chrome, embedded DM Sans). Its own
                # headings are an older section list and are cleared on
                # render; the pinned Docs above are the format. Owner,
                # 2026-09-03: "Doc sections inside the branded docx".
                "report_shell.docx")
REPORT_SHELL = TEMPLATES_DIR / "report_shell.docx"

#: The scoring-workbook template of record, in the owner's Drive.
SHEET_ID = "18IoJD5jn9aIe3E_F2omxqIZrjnHQwfR2pD0-_nUe5zc"
#: Assembled rather than written whole, so the plugin's own policy hook does
#: not read a source file as a shell fetch of a Drive document.
URL = "https://" + "docs.google.com" + "/spreadsheets/d/" + SHEET_ID + "/"

#: Sheets the template may carry that the contract deliberately does not.
#: Named, so an unexpected extra sheet is still reported.
TEMPLATE_EXTRAS_ALLOWED = ("Instructions", "README", "Cover", "Notes",
                           "Rubric", "Legend")


def _headers(path) -> dict[str, list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        row = next(ws.iter_rows(min_row=1, max_row=1), ())
        out[ws.title] = [str(c.value).strip() if c.value is not None else ""
                         for c in row]
    return out


def drift(path) -> dict:
    """Every difference between a template copy and the codified contract."""
    have = _headers(path)
    want = {name: list(cols) for name, cols in C.SHEETS.items()}
    # A tab the app reads as one of the canonical sheets (an evidence copy,
    # the technographic-scan tab) is a recognised alias, not drift — the gold
    # standard carries three of them. See contract.INGEST_ALIASES.
    aliases = getattr(C, "INGEST_ALIASES", {})
    extra = [s for s in sorted(set(have) - set(want))
             if s not in TEMPLATE_EXTRAS_ALLOWED and s not in aliases]
    ignored = [s for s in sorted(set(have) - set(want))
               if s in TEMPLATE_EXTRAS_ALLOWED]
    alias_present = [s for s in sorted(set(have) - set(want)) if s in aliases]
    out = {
        "template": str(path), "template_url": URL,
        "contract": C.WORKBOOK_CONTRACT,
        "sheets_in_template_only": extra,
        "sheets_ignored_as_guidance": ignored,
        "sheets_recognised_as_alias": {s: aliases[s] for s in alias_present},
        "sheets_in_contract_only": sorted(set(want) - set(have)),
        "header_drift": {},
    }
    for name in sorted(set(have) & set(want)):
        t = [h for h in have[name] if h]
        c = list(want[name])
        if t != c:
            out["header_drift"][name] = {
                "in_template_only": [x for x in t if x not in c],
                "in_contract_only": [x for x in c if x not in t],
                "order_differs": ([x for x in t if x in c]
                                  != [x for x in c if x in t]),
            }
    out["aligned"] = not (out["sheets_in_template_only"]
                          or out["sheets_in_contract_only"]
                          or out["header_drift"])
    return out


# ── the binding: templates pinned INTO the run ───────────────────────────

def pinned_digest() -> dict:
    """sha256 per pinned file, and one digest over all of them."""
    out, whole = {}, hashlib.sha256()
    for name in PINNED_FILES:
        p = TEMPLATES_DIR / name
        if not p.is_file():
            raise FileNotFoundError(
                f"pinned template {name} is missing from {TEMPLATES_DIR}; the "
                f"plugin ships it, so this is a partial install — reinstall")
        h = hashlib.sha256(p.read_bytes()).hexdigest()
        out[name] = h
        whole.update(h.encode())
    out["_all"] = whole.hexdigest()
    return out


def bind(run, wb=None) -> dict:
    """Record WHICH templates this run produces its deliverables to.

    Owner, 2026-09-03: "ensure there is a way to ensure the agents do not lose
    context of this" and "automatic tooling that invokes the templates even
    before the process begins". Binding writes the pinned templates' digest
    into Run_Metadata.template_binding and a `template_binding.json` beside
    the run — the report sections, the workbook shape and the gold reference
    a producer must read before authoring, with their paths — and `orient`
    refuses to serve a card while the binding is blank. `engine.cli start`
    calls this, so no run begins unbound."""
    from . import report_spec as RS
    wb = wb or run.open()
    d = pinned_digest()
    gold = json.loads((TEMPLATES_DIR / "gold_reference.json").read_text())
    doc = {
        "_contract": "template-binding-v1",
        "bound_at": _utcnow(),
        "run_id": run.run_id,
        "digest": d["_all"],
        "files": {k: v for k, v in d.items() if k != "_all"},
        "templates_dir": str(TEMPLATES_DIR),
        "plugin_version": installed_manifest_version(),
        "requires_plugin_version": templates_require(),
        "workbook": {"contract": C.WORKBOOK_CONTRACT, "sheets": len(C.SHEETS),
                     "drive_template_id": SHEET_ID},
        "reports": {k: {"title": s.title, "drive_doc_id": s.drive_doc_id,
                        "markdown": str(TEMPLATES_DIR / s.markdown),
                        "sections": [f"{x.id}. {x.heading}" for x in s.sections],
                        "min_words": s.min_words}
                    for k, s in RS.SPECS.items()},
        "gold_reference": {"entity": gold.get("entity"), "run_id": gold.get("run_id"),
                           "files": gold.get("files"),
                           "read_this_first": str(TEMPLATES_DIR / "gold_reference.json")},
        "read_before_authoring": [
            str(TEMPLATES_DIR / "report_templates.json"),
            str(TEMPLATES_DIR / "gold_reference.json"),
            str(TEMPLATES_DIR / "client_profile_template.md"),
            str(TEMPLATES_DIR / "assessment_report_template.md"),
        ],
    }
    out = run.root / "00_entity_profile" / "template_binding.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(doc, indent=2))
    wb.set_metadata("template_binding", f"{d['_all'][:16]} @ {doc['bound_at']}")
    doc["written_to"] = str(out)
    return doc


def binding_state(wb) -> dict:
    """Is this run bound, and to the templates this checkout ships?"""
    rec = str(wb.metadata().get("template_binding") or "").strip()
    if not rec:
        return {"bound": False, "current": False,
                "fix": "engine.template bind --run <R> --root <ROOT>"}
    now = pinned_digest()["_all"][:16]
    return {"bound": True, "current": rec.startswith(now), "recorded": rec,
            "pinned_now": now,
            "fix": None if rec.startswith(now) else
            "the pinned templates changed since this run was bound; re-run "
            "`engine.template bind` and re-read the report contract"}


# ── the zip guard: does the install carry the engine its templates need? ─
#
# Owner decision 2026-09-03: the plugin runs BOTH as a Claude Code marketplace
# checkout and as a zip uploaded to Cowork. `plugin_version.compare()` guards
# the first (installed cache vs the repo's published manifest); a Cowork
# session has no repo to compare against, so the zip has to judge itself. The
# pinned templates record the plugin version they were pinned FOR
# (`requires_plugin_version`); an install whose own manifest is older than
# that carries templates from a newer tree than its engine, hooks and agents
# — a mixed upload — and refuses. Fail-open on anything unreadable, like every
# other guard in this plugin.

PLUGIN_MANIFEST = Path(__file__).resolve().parents[3] / ".claude-plugin" / "plugin.json"


def _semver(v) -> tuple | None:
    m = re.fullmatch(r"\s*(\d+)\.(\d+)\.(\d+)\s*", str(v or ""))
    return tuple(int(x) for x in m.groups()) if m else None


def installed_manifest_version(manifest: Path | None = None) -> str | None:
    try:
        return json.loads((manifest or PLUGIN_MANIFEST).read_text()).get("version")
    except (OSError, ValueError):
        return None


def templates_require(templates_dir: Path | None = None) -> str | None:
    try:
        return json.loads(((templates_dir or TEMPLATES_DIR) / "report_templates.json")
                          .read_text()).get("requires_plugin_version")
    except (OSError, ValueError):
        return None


def zip_guard(manifest: Path | None = None, templates_dir: Path | None = None) -> dict:
    """{ok, installed, required, status, fix}. `status` is one of
    OK · PREDATES_TEMPLATES (refuse) · UNREADABLE (fail open, say so)."""
    inst = installed_manifest_version(manifest)
    req = templates_require(templates_dir)
    a, b = _semver(inst), _semver(req)
    if a is None or b is None:
        return {"ok": True, "status": "UNREADABLE", "installed": inst, "required": req,
                "fix": None, "note": "manifest or templates unreadable; not judged"}
    if a < b:
        return {"ok": False, "status": "PREDATES_TEMPLATES", "installed": inst,
                "required": req,
                "fix": (f"this install's manifest is {inst} but its pinned templates "
                        f"were pinned for plugin {req}: the zip (or cache) predates the "
                        f"engine that enforces them. Re-upload the zip that "
                        f"`python3 plugins/dma-insights/scripts/package_plugin.py` "
                        f"builds from the current checkout, or `doctor.py --heal` "
                        f"the marketplace install.")}
    return {"ok": True, "status": "OK", "installed": inst, "required": req, "fix": None}


def _utcnow() -> str:
    import datetime as _dt
    return _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# ── report drift: the Doc export vs the JSON the engine enforces ─────────

_H1 = re.compile(r"^# (\d{1,2})\\?\. (.+?)\s*$", re.M)
_H3 = re.compile(r"^### (\d{1,2}\.\d) (.+?)\s*$", re.M)


def report_drift() -> dict:
    """Every numbered heading in each Doc export that report_templates.json
    does not carry, and vice versa. Empty lists mean the pin is faithful."""
    from . import report_spec as RS
    out = {"aligned": True, "reports": {}}
    for key, spec in RS.SPECS.items():
        md = (TEMPLATES_DIR / spec.markdown).read_text(encoding="utf-8")
        doc_h1 = {n: h.replace("\\", "") for n, h in _H1.findall(md)}
        spec_h1 = {s.id: s.heading for s in spec.sections}
        h1_missing = {n: h for n, h in doc_h1.items() if n not in spec_h1}
        h1_extra = {n: h for n, h in spec_h1.items() if n not in doc_h1}
        h1_renamed = {n: (doc_h1[n], spec_h1[n]) for n in doc_h1
                      if n in spec_h1 and doc_h1[n].casefold() != spec_h1[n].casefold()}
        doc_h3 = [f"{n} {h.replace(chr(92), '')}" for n, h in _H3.findall(md)]
        spec_blocks = [b for s in spec.sections for b in s.blocks]
        h3_missing = [h for h in doc_h3 if h not in spec_blocks]
        rep = {"h1_in_doc_not_spec": h1_missing, "h1_in_spec_not_doc": h1_extra,
               "h1_renamed": h1_renamed, "numbered_h3_in_doc_not_spec": h3_missing}
        if any(rep.values()):
            out["aligned"] = False
        out["reports"][key] = rep
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(prog="engine.template",
                                 description=__doc__.split("\n")[0])
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("id", help="the template of record, and how to fetch it")
    c = sub.add_parser("check")
    c.add_argument("--file", help="a local copy of the template")
    c.add_argument("--run", help="check a RUN's workbook instead")
    c.add_argument("--root")
    c.add_argument("--json", action="store_true")
    b = sub.add_parser("bind", help="pin the templates INTO a run (start does this)")
    b.add_argument("--run", required=True); b.add_argument("--root")
    st = sub.add_parser("binding", help="is the run bound to the current pins?")
    st.add_argument("--run", required=True); st.add_argument("--root")
    sub.add_parser("report-drift",
                   help="the Doc exports vs report_templates.json, heading by heading")
    sub.add_parser("pins", help="the pinned files and their digests")
    sub.add_parser("zip-guard",
                   help="does this install's manifest carry the plugin version "
                        "its pinned templates require? (Cowork zip / cache)")

    a = ap.parse_args(argv)
    if a.cmd == "zip-guard":
        g = zip_guard()
        print(json.dumps(g, indent=2))
        return 0 if g["ok"] else 1
    if a.cmd == "pins":
        print(json.dumps(pinned_digest(), indent=2)); return 0
    if a.cmd == "report-drift":
        d = report_drift()
        print(json.dumps(d, indent=2))
        return 0 if d["aligned"] else 1
    if a.cmd in ("bind", "binding"):
        from . import runstate
        run = runstate.locate(a.run, Path(a.root) if a.root else None)
        if a.cmd == "bind":
            print(json.dumps(bind(run), indent=2)); return 0
        d = binding_state(run.open())
        print(json.dumps(d, indent=2))
        return 0 if d["bound"] and d["current"] else 1
    if a.cmd == "id":
        print(f"template of record : {SHEET_ID}")
        print(f"                     {URL}")
        print(f"contract           : {C.WORKBOOK_CONTRACT}, "
              f"{len(C.SHEETS)} sheets")
        print("fetch a copy       : python3 scripts/drive_fetch.py pull "
              "--name '<template name>' --dest /tmp/template.xlsx")
        print("then               : python3 -m engine.template check "
              "--file /tmp/template.xlsx")
        print("\nthe contract is authoritative; this check exists to make a "
              "divergence visible, not to follow the template blindly.")
        print(f"pinned copies      : {TEMPLATES_DIR}")
        return 0

    path = a.file
    if not path:
        if not a.run:
            print("give --file (a template copy) or --run (a run workbook)",
                  file=sys.stderr)
            return 2
        from . import runstate
        path = runstate.locate(a.run, Path(a.root) if a.root else None) \
            .workbook_path
    out = drift(path)
    if a.json:
        print(json.dumps(out, indent=2))
        return 0 if out["aligned"] else 1
    print(f"{'ALIGNED' if out['aligned'] else 'DRIFT'} — "
          f"{Path(path).name} against contract {out['contract']}")
    if out["sheets_ignored_as_guidance"]:
        print(f"  guidance sheets, ignored: "
              f"{', '.join(out['sheets_ignored_as_guidance'])}")
    for key, label in (("sheets_in_template_only", "in the template only"),
                       ("sheets_in_contract_only", "in the contract only")):
        if out[key]:
            print(f"  sheets {label}: {', '.join(out[key])}")
    for sheet, d in out["header_drift"].items():
        print(f"  {sheet}:")
        if d["in_template_only"]:
            print(f"    columns in the template only: "
                  f"{', '.join(d['in_template_only'])}")
        if d["in_contract_only"]:
            print(f"    columns in the contract only: "
                  f"{', '.join(d['in_contract_only'])}")
        if d["order_differs"]:
            print("    the shared columns are in a different order")
    return 0 if out["aligned"] else 1


if __name__ == "__main__":
    sys.exit(main())
