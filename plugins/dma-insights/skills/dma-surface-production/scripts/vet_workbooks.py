#!/usr/bin/env python3
"""Workbook hygiene, before the parser sees anything.

Every check here corresponds to a defect that reached a rendered page. The script
REPORTS; you decide whether to refuse. Read `02-inputs/4-vetting.md` for what each
finding does downstream — the consequence is the reason the check exists.

    python scripts/vet_workbooks.py <package-dir> [--subvertical CU]
    python scripts/vet_workbooks.py <scoring.xlsx> [research.xlsx]

Give it the entity's sub-vertical code and it names the variant cells the workbook
scored for somebody else — they render nowhere, and 59 of them reached a credit
union's promoted heatmap.

Exit 0 clean · 1 findings that need a decision · 2 could not read the input.
"""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:                                            # pragma: no cover
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

# Headers that are STATISTICS, not peer institutions. A parser that treats an
# unrecognised score column as a peer invented 54 rows of institutions literally
# named "Median", and the median cross-check then never ran.
STAT_HEADERS = {"median", "p25", "p75", "mean", "average", "avg", "stdev",
                "std", "min", "max", "count", "n", "quartile"}

# Columns whose header contains "score" and which are NOT 1–5 maturity
# scores. Measured across 111 corpus clients, which carry 130 distinct
# *score* headers between them — matching on the substring alone refused
# Houlihan Lokey for 26 "scores outside 1.0–5.0" that were a COUNT of
# sub-capabilities scored (14–25) and a recommendation PRIORITY on its own
# scale (6.0–7.05). Neither is a maturity score and neither was dirty.
NON_MATURITY_SCORE = (
    "scored",        # subcaps_scored, scored_count, categories_scored — counts
    "priority",      # priority_score — its own ranking scale
    "ers",           # ers_score — evidence strength, a different scale
    "delta",         # score_delta — a difference, legitimately negative
    "max_", "max ",  # max_score — the scale's ceiling, not a measurement
    "target",        # target_score — an aspiration
    "weighted",      # weighted_score — score x weight, exceeds 5 by design
    "rationale",     # score_rationale — prose
    "impact",        # impact_score — recommendation prioritisation, own scale
    "effort",        # effort_score — the same, measured 5.2-6.8 on a real
                     # Recommendations sheet and refused as dirty maturity
    "confidence",    # confidence_score — a 0-1 or 0-100 scale
    "coverage", "readiness", "risk_score", "fit",
    "count", "total", "/",
)


def is_maturity_score_column(header: str) -> bool:
    h = (header or "").strip().lower()
    if "score" not in h or "peer" in h:
        return False
    return not any(m in h for m in NON_MATURITY_SCORE)

#: How a package says "I searched and found nothing". Measured spellings
#: from the corpus; a blank cell means the same thing.
ABSENCE_RE = re.compile(
    r"^\s*(\[?(no[_ ]?evidence|not[_ ]?found|none|n/?a|absent|"
    r"no public evidence|not available|not evidenced)\]?)\b", re.I)

#: A column that carries the CONTENT of an evidence row — what a second
#: definition would have to contradict. Linkage columns (subcap, tier,
#: confidence, weight) are deliberately absent: they vary across the rows of
#: one id by design.
DEFINING_COL_RE = re.compile(
    r"excerpt|anchor_quote|quote|passage|verbatim|url|link|source_doc|"
    r"source_name|source_title|title|publisher|published", re.I)

CELL_RE = re.compile(r"^P[1-4]C\d+(\.\d+)*(\.[A-Z]{2,3}\d+)?$", re.I)
EID_RE = re.compile(r"^E[-_][A-Z0-9]+[-_]?\d*(:F\d+)?$", re.I)

# The suffix codes that name exactly ONE sub-vertical. A family or product code
# (BK depository, WM wealth, PEN retirement) serves every entity and is not
# evidence that a cell belongs to somebody else.
SUBVERTICAL_CODES = {"RB", "CU", "CL", "CIB", "FC", "AM", "RIA", "IC", "IB"}
VARIANT_RE = re.compile(r"^([A-Z]{2,3})(\d+)$")

findings: list[tuple[str, str]] = []
entity_sv: str | None = None


#: THE CLOSED LIST. A package enters the system unless one of these is true.
#:
#: Owner, 2026-08-23: a firing refused its client and both reserves and
#: produced nobody. One of those refusals was a missing sheet whose contents
#: sat in a column; another ("103 cell names mismatched against the
#: catalogue") is raised by no check in this repository at all — an agent
#: reasoned its way to it. An agent that may refuse for any reason it can
#: articulate will, and no amount of fixing individual checks bounds that.
#:
#: WHY A PERMISSIVE VETTER IS SAFE, and this is the argument a future
#: tightening must answer: the vetter is a PRE-FILTER, not the last line.
#: Fabricated content cannot reach a client through a lenient vetter,
#: because promotion is gated independently — evidence is fail-closed
#: (invariant 4: every cited id resolves, belongs to this run, and carries a
#: verbatim 50-500 character excerpt), Gate M fails a run whose citations
#: cannot be opened, and the AG/SG/ET/CG families run at submit. A package
#: that gets past this list still cannot promote a score it cannot evidence.
#: What a false refusal costs, by contrast, is the whole firing.
SCRIPT_REFUSALS = {
    "V1": "the workbook has too few tabs to be a generation the parser knows",
    "V2": "maturity scores fall outside 1.0-5.0",
    "V3": "one evidence id is defined twice with DIFFERENT content",
    "V4": "scored rows carry no source_cell, which cannot be backfilled",
    "V5": "excerpts are under 50 characters and will be refused at registration",
    "V7": "no research workbook exists anywhere in the tree",
}

#: Conditions no script can check, already stated in package-vetter.md. An
#: agent may refuse for one of these and must quote the code when it does.
AGENT_REFUSALS = {
    "V8": "the run names no catalogue version, so its cell names come from nowhere",
    "V9": "a score was taken from the research workbook rather than the scoring one",
    "V10": "undated evidence was dated to today rather than left UNVERIFIED",
    "V11": "entity identity is PENDING_REVIEW and unadjudicated",
}

#: V6 was "no excerpt column found". It is NOT here: the script already
#: raises that as a WARN, deliberately, and a registry entry with no call
#: site is the same defect as a guard that checks nothing — it advertises a
#: protection that does not exist. The number is retired rather than reused,
#: because a code that changes meaning is worse than a gap.
REFUSALS = {**SCRIPT_REFUSALS, **AGENT_REFUSALS}


def note(level: str, msg: str, code: str | None = None) -> None:
    """A REFUSE must name its code. Anything a code cannot be found for is a
    finding that travels with the run, never a refusal."""
    if level == "REFUSE":
        if code not in REFUSALS:
            raise ValueError(
                f"refusal without a listed code: {code!r}. Add it to REFUSALS "
                f"with its criterion, or emit this as WARN — an unlisted "
                f"refusal is how a firing loses its client slot to a rule "
                f"nobody agreed to")
        msg = f"{code}: {msg}"
    findings.append((level, msg))


#: Where a cap can be recorded. A cap is a scoring ceiling the assessment
#: applied, and it is written wherever that assessment kept its issue log —
#: a workbook sheet, a CSV, a JSON ledger, a column on the scoring detail.
CAPS_RE = re.compile(r"caps?[_ ]?applied|caps?[_ ]?log|issues?|"
                     r"contradiction", re.I)

#: A `Caps_Applied` cell saying, in the ways packages say it, "none".
NO_CAP = {"", "-", "--", "n/a", "na", "none", "no", "no cap", "no caps",
          "not applied", "nil", "0", "0.0", "false"}


#: Defining columns grouped into the family they state, so `url` and
#: `source_url` compare against each other rather than past each other.
def _field_family(header: str) -> str:
    h = header.lower()
    if "url" in h or "link" in h:
        return "url"
    if "date" in h or "published" in h:
        return "date"
    if "source" in h or "title" in h or "publisher" in h:
        return "source"
    if "excerpt" in h or "quote" in h or "verbatim" in h or "passage" in h:
        return "excerpt"
    return h


def _disagreeing(vals: list) -> bool:
    """Do these values genuinely contradict, or restate one another?

    Measured false positives this replaces: a tab that TRUNCATES its sibling
    ("…Vibe CU, charte" beside "…charter 61522" — same call report, one cell
    cut short) read as different content. One value being a prefix of the
    other is the same statement at two lengths; the longer row survives an
    ON CONFLICT and nothing is lost. Trailing-slash and case differences on
    a URL are likewise the same address.
    """
    norm = []
    for v in set(vals):
        n = re.sub(r"\s+", " ", v.strip().lower()).rstrip("/")
        if n:
            norm.append(n)
    norm = sorted(set(norm), key=len)
    for i, a in enumerate(norm):
        for b in norm[i + 1:]:
            if not b.startswith(a):
                return True
    return False


def _is_cap_value(v) -> bool:
    return str(v if v is not None else "").strip().lower() not in NO_CAP


def _sheet_has_scores(rows) -> bool:
    """Does this sheet resolve to scored rows — a subcap-id column beside a
    1-5 score — whatever the sheet is called?

    Shape, not name. The corpus carries at least three sheet-naming
    generations plus flattened single-sheet workbooks, and every check that
    matched on a name refused one of them.
    """
    if not rows:
        return False
    hi, hdr = header_row(rows)
    if hi is None:
        return False
    low = [str(c or "").strip().lower() for c in hdr]
    has_id = any("subcap" in h and "id" in h for h in low) or any(
        CELL_RE.match(str(c or "").strip())
        for r in rows[hi + 1:hi + 40] for c in (r or []) if isinstance(c, str))
    if not has_id:
        return False
    for j, h in enumerate(low):
        if "score" not in h or not is_maturity_score_column(hdr[j]):
            continue
        for r in rows[hi + 1:]:
            if j < len(r) and isinstance(r[j], (int, float)) \
                    and 1.0 <= float(r[j]) <= 5.0:
                return True
    return False


def scan_caps(root: Path, pm: dict, books: list) -> dict:
    """Every place this package could have recorded a cap, and what is in them.

    THE RULE THIS EXISTS TO ENFORCE (owner, 2026-08-23): "Caps applied may
    even exist in the scoring and research workbook and usually relate to
    the issue log or issues raised in the client research report, or an
    issue log in csv or any other format. If no caps were applied, then
    there were no issues."

    Both halves matter. Caps are not confined to a `Caps_Applied_Log` sheet,
    so looking only there and finding nothing proves nothing — and an EMPTY
    result is a real answer, not a missing one. On 2026-08-23 a vetter
    refused three consecutive packages for a missing sheet and the routine
    burned its entire reserve list in one firing on a state that means "this
    assessment raised no issues".

    So this returns what it found and where it looked, and the caller reports
    both. Nothing here refuses.
    """
    checked, sources, records = [], [], 0
    for book in [b for b in books if b]:
        try:
            tabs = sheets_of(book)
        except Exception as exc:                               # noqa: BLE001
            # NOT a silent drop. This `continue` left the workbook out of
            # `checked` entirely, so report_caps could print "no caps sheet,
            # log or column exists anywhere in this package" about a book it
            # never opened — inside the one function rewritten to make an
            # absence honest.
            checked.append(f"{Path(str(book)).name}: UNREADABLE "
                           f"({type(exc).__name__}) — NOT SEARCHED")
            continue
        rel = str(book)
        counted_sheets = set()
        for name, rows in tabs.items():
            if CAPS_RE.search(name):
                checked.append(f"{Path(rel).name}[{name}]")
                counted_sheets.add(name)
                body = [r for r in rows[1:] if any(
                    str(c or "").strip() for c in r)]
                if body:
                    sources.append(f"{Path(rel).name}[{name}]: {len(body)} row(s)")
                    records += len(body)
        # The COLUMN, which is where a cap is recorded per scored row and
        # which no "is the sheet present" check can see. A sheet already
        # counted whole is skipped: an Issue_Log sheet with an `Issue`
        # column would otherwise be counted once as rows and again as cells.
        for name, rows in tabs.items():
            if not rows or name in counted_sheets:
                continue
            hdr = [str(c or "").strip().lower() for c in rows[0]]
            cols = [i for i, h in enumerate(hdr) if CAPS_RE.search(h)]
            if not cols:
                continue
            checked.append(f"{Path(rel).name}[{name}].{hdr[cols[0]]}")
            hits = sum(1 for r in rows[1:] for i in cols
                       if i < len(r) and _is_cap_value(r[i]))
            if hits:
                sources.append(f"{Path(rel).name}[{name}] column "
                               f"{hdr[cols[0]]!r}: {hits} capped row(s)")
                records += hits

    # Files, in any format the package chose.
    for rel in (pm.get("governance") or []) + (pm.get("other") or []) + \
            (pm.get("evidence_tables") or []):
        if not CAPS_RE.search(rel):
            continue
        p = root / rel
        try:
            if p.suffix.lower() in (".csv", ".tsv"):
                lines = [ln for ln in p.read_text(
                    errors="replace").splitlines() if ln.strip()]
                n = max(0, len(lines) - 1)
            elif p.suffix.lower() in (".json", ".jsonl"):
                import json                                    # noqa: PLC0415
                if p.suffix.lower() == ".jsonl":
                    n = sum(1 for ln in p.read_text(
                        errors="replace").splitlines() if ln.strip())
                else:
                    d = json.loads(p.read_text(errors="replace"))
                    n = len(d) if isinstance(d, list) else len(
                        next((v for v in d.values() if isinstance(v, list)), []))
            else:
                checked.append(f"{rel}: format not parsed here — NOT SEARCHED")
                continue
        except Exception as exc:                               # noqa: BLE001
            checked.append(f"{rel}: UNPARSEABLE ({type(exc).__name__}) "
                           f"— NOT SEARCHED")
            continue
        checked.append(rel)
        if n:
            sources.append(f"{rel}: {n} row(s)")
            records += n

    # A report is where a human reads the issues; it is named, never parsed.
    prose = [r for r in (pm.get("reports") or []) if CAPS_RE.search(r)]
    return {"records": records, "sources": sources, "checked": checked,
            "prose": prose}


def report_caps(caps: dict) -> None:
    """Say what was found, and say plainly that nothing found is an answer."""
    if caps["records"]:
        note("PIN", f"caps applied: {caps['records']} record(s) across "
                    f"{len(caps['sources'])} source(s) — "
                    f"{'; '.join(caps['sources'][:4])}. Every cap is a "
                    f"scoring ceiling the assessment applied; it belongs in "
                    f"the payload's caps[] and is NOT a safeguard gate")
        return
    where = (f"looked in {len(caps['checked'])} place(s): "
             f"{', '.join(caps['checked'][:6])}"
             if caps["checked"] else
             "no caps sheet, log or column exists anywhere in this package")
    note("PIN", f"NO CAPS APPLIED — {where}. This is a valid state and NEVER "
                f"a refusal (owner, 2026-08-23): if no caps were applied, "
                f"then there were no issues. Serve caps[] empty and say so; "
                f"do not hunt for a Caps_Applied_Log that a clean assessment "
                f"had no reason to write")
    if caps["prose"]:
        note("PIN", f"issue narrative to read if a cap is later claimed: "
                    f"{', '.join(caps['prose'][:3])}")


def sheets_of(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {ws.title: [list(r) for r in ws.iter_rows(max_row=400, values_only=True)]
                for ws in wb.worksheets}
    finally:
        wb.close()


def header_row(rows):
    """The first row with three or more non-empty string cells."""
    for i, r in enumerate(rows[:20]):
        strs = [c for c in r if isinstance(c, str) and c.strip()]
        if len(strs) >= 3:
            return i, [str(c).strip() if c is not None else "" for c in r]
    return None, []


def vet_scoring(path: Path) -> None:
    print(f"\n=== scoring workbook · {path.name}")
    tabs = sheets_of(path)
    print(f"tabs ({len(tabs)}): {', '.join(list(tabs)[:12])}"
          + (" …" if len(tabs) > 12 else ""))
    if len(tabs) <= 3:
        # COUNT WHAT THE TABS CONTAIN, NOT HOW MANY THERE ARE. A flattened
        # single-sheet generation is real and carries everything: measured on
        # a corpus package refused for "only 3 tab(s)" whose one
        # `Scoring_Workbook` sheet held 160 rows under the full canonical
        # header — SubCap_ID, Score, Confidence, Evidence_IDs, Source_URLs,
        # Caps_Applied. Refusing on cardinality refuses a shape, not a defect.
        scored = sum(1 for rows in tabs.values() if _sheet_has_scores(rows))
        if scored:
            note("PIN", f"only {len(tabs)} tab(s) — a flattened generation. "
                        f"{scored} of them resolve to scored rows, so the "
                        f"shape is unfamiliar, not empty")
        else:
            note("REFUSE", f"only {len(tabs)} tab(s) and NONE resolves to "
                           f"scored rows (a subcap-id column beside a 1-5 "
                           f"score). Name the tabs in your refusal.", code="V1")

    cells: list[str] = []
    scores: list[float] = []
    e_ids: list[str] = []
    ev_defs: dict[str, dict] = {}      # eid -> field family -> [values]
    missing_source_cell = 0
    saw_source_cell_col = False

    for name, rows in tabs.items():
        hi, hdr = header_row(rows)
        if hi is None:
            continue
        low = [h.lower() for h in hdr]
        # A DEFINITION TAB DEFINES; A LINKAGE TAB REFERS. Matching on the
        # tab NAME alone swept in Evidence_Linkage, Evidence_Index and
        # Absent_Evidence_Log — tables where one evidence id legitimately
        # appears once per subcap it supports. The vetter's own note says a
        # reference is never a defect; this is how references were reaching
        # the check anyway. A definition tab is one that carries the CONTENT
        # being defined.
        defining_cols = [i for i, h in enumerate(low)
                         if DEFINING_COL_RE.search(h)]
        is_register = "evidence" in name.lower() and bool(defining_cols)

        # peer columns that are really statistics
        for h in hdr:
            hl = h.strip().lower()
            if hl in STAT_HEADERS:
                note("WARN", f"{name}: column {h!r} is a STATISTIC. Confirm the "
                             f"parser does not read it as a peer institution.")
            if "peer" in hl and any(sh in hl for sh in ("median", "p25", "p75")):
                pass  # correctly qualified

        if "source_cell" in low:
            saw_source_cell_col = True
        idx_src = low.index("source_cell") if "source_cell" in low else None

        for r in rows[hi + 1:]:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            row_eids: list[str] = []
            for c in r:
                if isinstance(c, str):
                    t = c.strip()
                    if CELL_RE.match(t):
                        cells.append(t.upper())
                    elif EID_RE.match(t):
                        e_ids.append(t.upper())
                        row_eids.append(t.upper())
                elif isinstance(c, (int, float)):
                    if 0 < float(c) <= 5.0 or float(c) == 0:
                        pass
            if is_register and len(row_eids) == 1:
                # PER FIELD, not per row-fingerprint. A fingerprint built by
                # joining whichever defining columns a tab happens to carry
                # refused three June packages for "58 duplicates with
                # different content" whose content was IDENTICAL: one tab
                # stated (name, url, date) and its sibling (name, url), so
                # every id "differed" by the projection. Only a FIELD that
                # carries two genuinely disagreeing values loses a row under
                # ON CONFLICT DO NOTHING.
                for i in defining_cols:
                    if i >= len(r) or r[i] is None:
                        continue
                    v = re.sub(r"\s+", " ", str(r[i]).strip())
                    if not v or v.upper() == row_eids[0]:
                        continue
                    ev_defs.setdefault(row_eids[0], {}).setdefault(
                        _field_family(low[i]), []).append(v)
            if idx_src is not None and (idx_src >= len(r) or r[idx_src] in (None, "")):
                missing_source_cell += 1

        # score range, per column named like a score
        for j, h in enumerate(low):
            if "score" not in h or "peer" in h:
                continue
            col = []
            for r in rows[hi + 1:]:
                if r is None or j >= len(r):
                    continue
                v = r[j]
                if isinstance(v, (int, float)):
                    col.append(float(v))
            if not col:
                continue
            if not is_maturity_score_column(h):
                note("WARN", f"{name}: column {hdr[j]!r} is named like a "
                             f"score and is not one (a count, a different "
                             f"scale, or prose). Its range is NOT checked "
                             f"against 1.0-5.0.")
                continue
            live = [v for v in col if v != 0]
            in_range = [v for v in live if 1.0 <= v <= 5.0]
            if live and not in_range:
                # EVERY value out of range is a misidentified column, not
                # 26 dirty measurements. Refusing here is how a package is
                # halted for a header this script did not recognise.
                note("WARN", f"{name}: column {hdr[j]!r} holds no value in "
                             f"1.0-5.0 at all ({len(live)} values, e.g. "
                             f"{sorted(set(live))[:4]}) — it is very likely "
                             f"not a maturity score. Name it here if it is.")
                continue
            scores.extend(col)

    bad = [v for v in scores if v != 0 and not (1.0 <= v <= 5.0)]
    if bad:
        note("REFUSE", f"{len(bad)} score(s) outside 1.0–5.0 "
                       f"(e.g. {sorted(set(bad))[:5]}). A 0 bands as Activating "
                       f"and looks assessed.", code="V2")
    zeros = [v for v in scores if v == 0]
    if zeros:
        note("WARN", f"{len(zeros)} score(s) are exactly 0 — confirm these are "
                     f"blanks, not measurements.")

    # Evidence ids are unique PER CLIENT, and one id cited from many tabs is
    # a reference, not a defect (owner adjudication 2026-08-20: 43 false
    # REFUSEs on the first live vetting were exactly this). The defect that
    # loses rows silently under ON CONFLICT DO NOTHING is one id DEFINED
    # more than once with DIFFERENT content in a register tab — duplicate
    # by content decides, never duplicate by id alone.
    conflicting = {}
    for k, fields in ev_defs.items():
        bad = {fam: vals for fam, vals in fields.items()
               if _disagreeing(vals)}
        if bad:
            conflicting[k] = [f"{fam}: {' <> '.join(sorted(set(vals))[:2])[:90]}"
                              for fam, vals in bad.items()]
    repeated = {k: max(len(v) for v in fields.values())
                for k, fields in ev_defs.items()
                if any(len(v) > 1 for v in fields.values())
                and k not in conflicting}
    if conflicting:
        worst = sorted(conflicting.items(), key=lambda kv: -len(kv[1]))[:6]
        note("REFUSE", f"{len(conflicting)} evidence id(s) defined more than "
                       f"once with DIFFERENT content "
                       f"({', '.join(f'{k}×{len(v)}' for k, v in worst)}). One "
                       f"of each pair would be lost silently — adjudicate "
                       f"which row is real before parsing.", code="V3")
    if repeated:
        note("WARN", f"{len(repeated)} evidence id(s) re-defined with "
                     f"identical content — benign repetition; dedup is by "
                     f"content hash. Ids are unique per client only: any "
                     f"cross-client ledger entry carries the client slug as "
                     f"a prefix (e.g. t-rowe-price-group-inc:E-017).")

    # catalogue version, from the category count
    cats = {c.split(".")[0] for c in cells if CELL_RE.match(c)}
    cats = {c for c in cats if re.fullmatch(r"P[1-4]C\d+", c, re.I)}
    if cats:
        print(f"categories seen: {len(cats)}")
        if len(cats) == 17:
            note("PIN", "17 categories → this is a v5.0-shaped assessment. Pin "
                        "runs.ccg_catalog_version to v5.0, or every cell name "
                        "joins against v7.0 and comes back NULL.")
        elif len(cats) == 16:
            note("PIN", "16 categories → v7.0. Confirm the run is pinned.")
        else:
            note("WARN", f"{len(cats)} categories — matches neither v7.0 (16) nor "
                         f"v5.0 (17). State what you inferred and from what.")
    print(f"cells seen: {len(set(cells))} · evidence ids seen: {len(set(e_ids))}")

    # variant cells the workbook scored for OTHER sub-verticals. They are the
    # catalogue's, not this entity's, and the serve layer drops them — so a payload
    # that cites one cites a cell that renders nowhere.
    variants = Counter()
    for c in set(cells):
        m = VARIANT_RE.match(c.rsplit(".", 1)[-1])
        if m and m.group(1) in SUBVERTICAL_CODES:
            variants[m.group(1)] += 1
    if variants:
        print("variant cells by sub-vertical: "
              + " · ".join(f"{k}×{n}" for k, n in sorted(variants.items())))
        if entity_sv:
            foreign = {k: n for k, n in variants.items() if k != entity_sv}
            if foreign:
                note("WARN", f"{sum(foreign.values())} variant cell(s) belong to another "
                             f"sub-vertical on a {entity_sv} run "
                             f"({', '.join(f'{k}×{n}' for k, n in sorted(foreign.items()))}). "
                             f"They stay in the workbook and out of the payload — cite one "
                             f"and it resolves here and renders nowhere.")
        elif len(variants) > 1:
            note("WARN", f"variant cells span {len(variants)} sub-verticals. Pass "
                         f"--subvertical to name the entity's, or the payload will cite "
                         f"cells the run cannot serve.")
    if saw_source_cell_col and missing_source_cell:
        note("REFUSE", f"{missing_source_cell} row(s) have no source_cell. It "
                       f"cannot be backfilled after the scan.", code="V4")


def vet_research(path: Path, evidence_stores: list | None = None) -> None:
    print(f"\n=== research workbook · {path.name}")
    tabs = sheets_of(path)
    print(f"tabs ({len(tabs)}): {', '.join(list(tabs)[:12])}"
          + (" …" if len(tabs) > 12 else ""))
    excerpts: list[str] = []
    dated = 0
    ers = 0
    rows_seen = 0
    for name, rows in tabs.items():
        hi, hdr = header_row(rows)
        if hi is None:
            continue
        low = [h.lower() for h in hdr]
        col = lambda *keys: next(  # noqa: E731
            (low.index(k) for k in keys if k in low), None)
        # anchor_quote is the research workbook's OWN primary verbatim
        # column — package_map names it, evidence_normalize ranks it
        # first. Omitting it here meant a workbook that used only that
        # spelling reported "no excerpt column found" and got ZERO
        # excerpt vetting, on the one artefact that carries real
        # quotations. Latent on today's corpus (those tabs carry both
        # spellings); a one-word gap all the same.
        i_ex = col("evidence_excerpt", "excerpt", "anchor_quote",
                   "quote", "verbatim", "passage", "snippet")
        i_dt = col("date_published", "published_date", "publish_date", "date")
        i_er = col("ers_total", "ers", "ers_score")
        for r in rows[hi + 1:]:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            rows_seen += 1
            if i_ex is not None and i_ex < len(r) and isinstance(r[i_ex], str):
                excerpts.append(r[i_ex].strip())
            if i_dt is not None and i_dt < len(r) and r[i_dt] not in (None, ""):
                dated += 1
            if i_er is not None and i_er < len(r) and r[i_er] not in (None, ""):
                ers += 1
    if excerpts:
        lens = sorted(len(e) for e in excerpts)
        med = lens[len(lens) // 2]
        # AN EXPLICIT ABSENCE IS NOT A SHORT EXCERPT. Measured on a real
        # package: all 28 "short excerpts" that refused it were the literal
        # string "[NO_EVIDENCE] No public evidence found." — the assessment
        # declaring it searched and found nothing, which 02-inputs/4-vetting.md
        # requires ("Do not exclude a cell for having no evidence"). A blank
        # cell is the same statement with less ceremony. Neither is a
        # 0-character quotation, and neither will be offered at registration.
        absent = sum(1 for e in excerpts if not e.strip() or ABSENCE_RE.match(e))
        short = sum(1 for e in excerpts
                    if e.strip() and not ABSENCE_RE.match(e)
                    and len(e.strip()) < 50)
        empty = sum(1 for e in excerpts if not e)
        print(f"excerpts: {len(excerpts)} · median {med} chars · "
              f"{short} under the 50-char floor · {empty} empty")
        if med < 120:
            note("WARN", f"excerpt median is {med} chars. An excerpt that clears "
                         f"the floor and says nothing passes every gate and helps "
                         f"no reader. Take the whole claim.")
        if absent:
            note("PIN", f"{absent} row(s) record NO evidence rather than a "
                        f"quotation — an explicit absence marker or a blank. "
                        f"They go out as GAPs, never as fabrications, and "
                        f"they are not short excerpts")
        if short:
            note("REFUSE", f"{short} POPULATED excerpt(s) under 50 characters "
                           f"will be refused at registration (absences are "
                           f"counted separately and are not this).", code="V5")
    else:
        note("WARN", "no excerpt column found — the evidence tier's authority is "
                     "this workbook; confirm the tab names.")
    if rows_seen:
        print(f"rows: {rows_seen} · with a date: {dated} · with ERS: {ers}")
        if dated < rows_seen * 0.5:
            note("WARN", f"only {dated} of {rows_seen} rows carry a publication "
                         f"date. Undated items band UNVERIFIED and the recency "
                         f"ladder cannot rank them.")


def main(argv: list[str]) -> int:
    global entity_sv
    argv = list(argv)
    if any(a in ("-h", "--help") for a in argv[1:]):
        print(__doc__ or "mechanical vetting of an assessment package")
        print("usage: vet_workbooks.py <package-dir> [--subvertical <CODE>]")
        return 0
    if "--subvertical" in argv:
        i = argv.index("--subvertical")
        if i + 1 >= len(argv):
            print("--subvertical needs a code "
                  f"({' '.join(sorted(SUBVERTICAL_CODES))})", file=sys.stderr)
            return 2
        entity_sv = argv[i + 1].strip().upper()
        del argv[i:i + 2]
        if entity_sv not in SUBVERTICAL_CODES:
            print(f"unknown sub-vertical code {entity_sv!r} — expected one of "
                  f"{' '.join(sorted(SUBVERTICAL_CODES))}", file=sys.stderr)
            return 2
    if len(argv) < 2:
        print(__doc__)
        return 2
    target = Path(argv[1])
    scoring: Path | None = None
    research: Path | None = None
    evidence_stores: list = []
    if target.is_dir():
        # Discovery goes through package_map: packages come in at least four
        # structure generations (wrappers, 03_Assessment, workbooks in
        # 08_appendices, version stacks with INTERIM copies beside the live
        # workbook — measured across the 178-client corpus, 2026-08-20).
        # Naive rglob picked whichever xlsx sorted first.
        sys.path.insert(0, str(Path(__file__).resolve().parents[3]
                               / "scripts"))
        import package_map  # noqa: PLC0415
        pm = package_map.map_package(target)
        scoring = Path(pm["scoring"]["primary"]) \
            if pm["scoring"]["primary"] else None
        research = Path(pm["research"]["primary"]) \
            if pm["research"]["primary"] else None
        for amb in pm["ambiguities"]:
            note("WARN", f"package_map: {amb}")
        for aux in pm["auxiliary_xlsx"]:
            note("PIN", f"auxiliary workbook (not vetted as scoring): {aux}")
        if pm["evidence_tables"]:
            note("PIN", f"{len(pm['evidence_tables'])} evidence stores "
                        f"beyond the workbooks — evidence_normalize.py "
                        f"merges them; vet gaps there, not here")
        evidence_stores = list(pm["evidence_tables"])
        report_caps(scan_caps(target, pm, [scoring, research]))
    else:
        scoring = target
        research = Path(argv[2]) if len(argv) > 2 else None

    if scoring is None:
        print("no scoring workbook found — package_map classified the tree; "
              "a briefing- or research-only folder is not a synthesis input",
              file=sys.stderr)
        return 2
    try:
        vet_scoring(scoring)
        if research:
            vet_research(research, evidence_stores)
        elif evidence_stores:
            # THE RESEARCH WORKBOOK IS THE USUAL AUTHORITY, NOT THE ONLY ONE.
            # Measured across the corpus: two packages refused here carried 49
            # and 4 evidence stores respectively — JSON registers, issue
            # registers, search logs, inventory CSVs — which
            # evidence_normalize.py merges. This script says so itself sixteen
            # lines above, and then refused anyway.
            note("WARN", f"no research workbook, but {len(evidence_stores)} "
                         f"evidence store(s) exist ({', '.join(evidence_stores[:3])}"
                         f"{' …' if len(evidence_stores) > 3 else ''}). "
                         f"evidence_normalize.py merges them; excerpts and "
                         f"dates are vetted THERE, not here. Expect gaps to "
                         f"go out as GAPs, never as inventions")
        else:
            note("REFUSE", "no research workbook AND no evidence store of any "
                           "format anywhere in the tree — nothing carries an "
                           "excerpt, so every item would band UNVERIFIED.",
                 code="V7")
    except Exception as e:                                     # noqa: BLE001
        print(f"could not read: {type(e).__name__}: {e}", file=sys.stderr)
        return 2

    print("\n=== findings")
    if not findings:
        print("clean — nothing to decide.")
        return 0
    for level, msg in findings:
        print(f"[{level}] {msg}")
    print("\nA REFUSE line means: do not hand this to the parser. Say what is "
          "dirty, in which tab and column, and how many rows.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
