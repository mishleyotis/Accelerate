#!/usr/bin/env python3
"""Does this workbook carry the tabs the app reads — and which surface starves?

## Why this is derived from code, not from a sample file

Owner, 2026-09-03: "there is an issue with the agents getting the wrong
templates, even missing simple things such as formatting the scoring workbook
and the reports."

The obvious check is to diff a client's workbook against a golden sample. It
is the wrong check: a sample ages, and the moment it does the diff reports
differences that do not matter and misses the one that does. The app already
declares what it needs — `_TAB_TARGET` in the worker's parser maps every tab
it reads to the SURFACE that tab feeds, 29 of them — and that map cannot
drift from the code because it IS the code.

So this asks the only question that has consequences: for each tab the app
reads, is it present, and if not, what goes empty on the client's screen.

## What it reports

    MISSING   a tab the app reads is absent. The line names the surface that
              will render an empty state because of it.
    EMPTY     the tab exists and holds no data rows — which reaches the app
              identically to absent, and is the more common failure: a
              template copied but never filled.
    UNREAD    a tab the workbook ships that nothing reads. Not an error;
              worth knowing, because a producer filling it is doing work the
              app will never show.

Exit 1 when any tab the app reads is missing or empty.

    check_template.py <workbook.xlsx> [--json OUT] [--quiet]

Measured on three real workbooks, which is why the EMPTY rung exists:

    Golden 1 CU  scoring workbook   43 tabs   the reference shape
    BOTR         assessment         20 tabs   688 scored, composite 1.71
    BOTR         scoring (v5)       23 tabs   688 rows, ZERO scored — column
                                              D empty by contract at the
                                              research stage
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl

WORKER = Path(__file__).resolve().parents[5] / "apps" / "worker"
sys.path.insert(0, str(WORKER))

try:
    from dma_worker.workbook_parser import _TAB_TARGET, _tab_key
except Exception as exc:                       # pragma: no cover
    raise SystemExit(
        f"cannot import the worker's tab map ({exc!r}) — this script reports "
        "what the APP reads, and a second copy of that list here would be "
        "wrong the first time the app changed")


def _surface(v) -> str:
    tgt = v[0] if isinstance(v, (tuple, list)) else v
    if isinstance(tgt, (tuple, list)):
        return ", ".join(str(x) for x in tgt)
    return str(tgt)


def inspect(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        present = {_tab_key(n): n for n in wb.sheetnames}
        missing, empty, ok = [], [], []
        for tab, target in sorted(_TAB_TARGET.items()):
            key = _tab_key(tab)
            if key not in present:
                missing.append((tab, _surface(target)))
                continue
            ws = wb[present[key]]
            rows = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if any(c is not None and str(c).strip() for c in row):
                    rows += 1
                if rows > 1:
                    break
            (ok if rows > 1 else empty).append((tab, _surface(target)))
        read_keys = {_tab_key(t) for t in _TAB_TARGET}
        unread = sorted(n for k, n in present.items() if k not in read_keys)
        return {"tabs": len(wb.sheetnames), "ok": ok, "missing": missing,
                "empty": empty, "unread": unread}
    finally:
        wb.close()


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("workbook")
    ap.add_argument("--json", dest="out", default=None)
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args(argv)

    r = inspect(a.workbook)
    name = os.path.basename(a.workbook)
    print(f"{name}: {r['tabs']} tabs — {len(r['ok'])} of "
          f"{len(_TAB_TARGET)} read-tabs carry data")

    for label, group in (("MISSING", r["missing"]), ("EMPTY", r["empty"])):
        if not group:
            continue
        print(f"\n  {label} ({len(group)}) — the surface each one starves:")
        for tab, surface in group:
            print(f"    {tab:28s} -> {surface[:60]}")
    if r["unread"] and not a.quiet:
        print(f"\n  UNREAD ({len(r['unread'])}) — shipped, nothing reads them:")
        print("    " + ", ".join(r["unread"][:14])
              + (" …" if len(r["unread"]) > 14 else ""))

    if a.out:
        Path(a.out).write_text(json.dumps(r, indent=1), encoding="utf-8")
        print(f"\nwritten to {a.out}")
    bad = len(r["missing"]) + len(r["empty"])
    print(f"\n{bad} read-tab(s) missing or empty")
    return 1 if bad else 0


if __name__ == "__main__":
    raise SystemExit(main())
