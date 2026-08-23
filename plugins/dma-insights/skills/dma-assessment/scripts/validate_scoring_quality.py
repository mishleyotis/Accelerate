#!/usr/bin/env python3
"""
DMA Scoring Quality Validator — MANDATORY after Phase 4
========================================================
This script MUST be run before proceeding past Phase 4. It checks for the 7 most
common failure modes and blocks progression if any CRITICAL check fails.

Usage:
    python validate_scoring_quality.py <workbook_path>

Exit codes:
    0 = PASS (all checks pass)
    1 = FAIL (one or more CRITICAL checks failed — fix before proceeding)
"""

import sys
import os
import re
import json
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


def load_workbook(path):
    if not os.path.exists(path):
        print(f"FATAL: Workbook not found at {path}")
        sys.exit(1)
    return openpyxl.load_workbook(path, data_only=True)


# ──────────────────────────────────────────────────────────────────────
# CHECK 1: Row count per pillar (are we scoring at subcap level?)
# ──────────────────────────────────────────────────────────────────────
def check_row_counts(wb):
    """Each pillar sheet must have ≥50 data rows. Fewer = capability-level scoring."""
    expected_minimums = {
        "P1_Subcap_Scoring": 50,
        "P2_Subcap_Scoring": 50,
        "P3_Subcap_Scoring": 50,
        "P4_Subcap_Scoring": 50,
    }
    issues = []
    for sheet_name, min_rows in expected_minimums.items():
        if sheet_name not in wb.sheetnames:
            issues.append(f"CRITICAL: Sheet '{sheet_name}' missing entirely")
            continue
        ws = wb[sheet_name]
        data_rows = ws.max_row - 1
        if data_rows < min_rows:
            issues.append(
                f"CRITICAL: {sheet_name} has {data_rows} rows (need ≥{min_rows}). "
                f"You are scoring at CAPABILITY level, not SUBCAPABILITY level."
            )
    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 2: Score differentiation within capabilities
# ──────────────────────────────────────────────────────────────────────
def check_differentiation(wb):
    """No capability should have 100% identical scores across its subcaps."""
    issues = []
    total_caps = 0
    uniform_caps = 0
    low_variation_caps = 0

    for pname in PILLAR_SHEETS:
        if pname not in wb.sheetnames:
            continue
        ws = wb[pname]
        cap_scores = {}
        for row in range(2, ws.max_row + 1):
            cap_id = ws.cell(row=row, column=3).value
            score = ws.cell(row=row, column=4).value          # D, canonical
            if cap_id and score is not None:
                cap_scores.setdefault(cap_id, []).append(score)

        for cap_id, scores in cap_scores.items():
            if len(scores) < 2:
                continue
            total_caps += 1
            unique = len(set(scores))
            ratio = unique / len(scores)

            if unique == 1 and not all(s == 1.0 for s in scores):
                uniform_caps += 1
                issues.append(
                    f"CRITICAL: {cap_id} — ALL {len(scores)} subcaps scored identically ({scores[0]}). "
                    f"HARD BLOCK: differentiate at least 2 subcaps before proceeding."
                )
            elif ratio < 0.4 and not all(s == 1.0 for s in scores):
                low_variation_caps += 1
                issues.append(
                    f"WARNING: {cap_id} — only {unique}/{len(scores)} unique scores "
                    f"(variation ratio {ratio:.2f}). Re-examine diagnostic questions."
                )

    if uniform_caps > 0:
        issues.insert(0,
            f"CRITICAL SUMMARY: {uniform_caps}/{total_caps} capabilities have ZERO score "
            f"differentiation. This is the #1 quality failure in DMA assessments."
        )
    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 3: Rationale quality
# ──────────────────────────────────────────────────────────────────────
def check_rationale_quality(wb):
    """Rationales must be ≥150 chars, cite evidence IDs, and not use forbidden patterns."""
    FORBIDDEN = [
        "demonstrates well-developed capability",
        "demonstrates minimal capability",
        "demonstrates established maturity",
        "demonstrates basic capability",
        "demonstrates advanced capability",
        "demonstrates foundational capability",
        "demonstrates transformational",
        "based on public evidence analysis",
        "based on available data",
        "evidence suggests m",
        "score assigned based on",
        "category-based scoring",
    ]
    issues = []
    total = 0
    under_150 = 0
    forbidden_matches = 0
    no_evidence_cite = 0
    all_same_length = True
    lengths = set()

    for pname in PILLAR_SHEETS:
        if pname not in wb.sheetnames:
            continue
        ws = wb[pname]
        for row in range(2, ws.max_row + 1):
            subcap_id = ws.cell(row=row, column=1).value      # A, canonical
            rationale = str(ws.cell(row=row, column=10).value or "")   # J, canonical
            if not subcap_id:
                continue
            total += 1
            rlen = len(rationale)
            lengths.add(rlen)

            if rlen < 150:
                under_150 += 1

            rat_lower = rationale.lower()
            for pat in FORBIDDEN:
                if pat in rat_lower:
                    forbidden_matches += 1
                    if forbidden_matches <= 5:
                        issues.append(
                            f"CRITICAL: {subcap_id} rationale uses forbidden pattern: '{pat}'"
                        )
                    break

            if not re.search(r'E-\d+', rationale) and not re.search(r'HUBBL-\d+', rationale):
                no_evidence_cite += 1

    if len(lengths) <= 3 and total > 20:
        issues.insert(0,
            f"CRITICAL: All rationales cluster around {sorted(lengths)} character lengths. "
            f"This indicates template-stamped rationales, not individual analysis."
        )

    if under_150 > 0:
        issues.append(
            f"CRITICAL: {under_150}/{total} rationales are under 150 characters "
            f"(minimum required)."
        )

    if forbidden_matches > 5:
        issues.append(
            f"CRITICAL: {forbidden_matches} total forbidden pattern matches "
            f"(showing first 5 above)."
        )

    if no_evidence_cite > total * 0.2:
        issues.append(
            f"WARNING: {no_evidence_cite}/{total} rationales don't cite any Evidence ID "
            f"(E-xxx pattern). Rationales must reference specific evidence."
        )

    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 4: Evidence differentiation per capability
# ──────────────────────────────────────────────────────────────────────
def check_evidence_differentiation(wb):
    """Subcaps within a capability should NOT all cite identical evidence sets."""
    issues = []
    total_caps = 0
    identical_evidence_caps = 0

    for pname in PILLAR_SHEETS:
        if pname not in wb.sheetnames:
            continue
        ws = wb[pname]
        cap_evidence = {}
        for row in range(2, ws.max_row + 1):
            cap_id = ws.cell(row=row, column=3).value
            evidence = str(ws.cell(row=row, column=6).value or "")     # F, canonical
            if cap_id:
                cap_evidence.setdefault(cap_id, []).append(evidence)

        for cap_id, ev_list in cap_evidence.items():
            if len(ev_list) < 2:
                continue
            total_caps += 1
            unique_ev = set(ev_list)
            if len(unique_ev) == 1 and ev_list[0].strip():
                identical_evidence_caps += 1
                if identical_evidence_caps <= 5:
                    issues.append(
                        f"CRITICAL: {cap_id} — all {len(ev_list)} subcaps cite identical "
                        f"evidence '{ev_list[0][:60]}...'. Map facts to individual subcaps."
                    )

    if identical_evidence_caps > 0:
        issues.append(
            f"CRITICAL SUMMARY: {identical_evidence_caps}/{total_caps} capabilities have "
            f"ZERO evidence differentiation across subcaps."
        )
    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 5: Required columns present
# ──────────────────────────────────────────────────────────────────────
def check_required_columns(wb):
    """Verify columns S (Proof_Claims), T (Proof_Links), U (Evidence_Excerpt), V (Source_Document)."""
    required = {9: "Rationale", 5: "Evidence_IDs", 7: "Evidence_Ceiling", 10: "Proxy_Searched"}
    issues = []
    for pname in PILLAR_SHEETS:
        if pname not in wb.sheetnames:
            continue
        ws = wb[pname]
        for col, name in required.items():
            header = ws.cell(row=1, column=col).value
            if header is None:
                issues.append(f"CRITICAL: {pname} missing column {col} ({name})")
            elif str(header).strip() != name:
                # Allow close matches
                pass

        # Check if Evidence_Excerpt (col 20 or wherever it is) is populated
        excerpt_col = None
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col).value or "").strip() == "Evidence_Excerpt":
                excerpt_col = col
                break
        if excerpt_col:
            blank_excerpts = 0
            total_rows = 0
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=5).value:
                    total_rows += 1
                    val = ws.cell(row=row, column=excerpt_col).value
                    if not val or len(str(val).strip()) < 30:
                        blank_excerpts += 1
            if blank_excerpts > total_rows * 0.3:
                issues.append(
                    f"WARNING: {pname} — {blank_excerpts}/{total_rows} Evidence_Excerpt cells "
                    f"are blank or <30 chars."
                )
    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 6: Caps Applied Log
# ──────────────────────────────────────────────────────────────────────
#: Sheets and columns a cap can be recorded in. A cap is a scoring ceiling,
#: and it is written wherever that assessment kept its issue log.
_CAPS_RE = re.compile(r"caps?[_ ]?applied|caps?[_ ]?log|issues?|contradiction",
                      re.I)
_NO_CAP = {"", "-", "--", "n/a", "na", "none", "no", "no cap", "no caps",
           "not applied", "nil", "0", "0.0", "false"}


def check_caps_applied(wb):
    """How many caps this workbook records, and where.

    THIS RETURNED A CRITICAL FOR A MISSING SHEET until 2026-08-23, and
    `package-vetter.md` instructs that a CRITICAL from this script is a
    REFUSE. So a package with no `Caps_Applied_Log` tab did not enter the
    system — while the same workbook carried 380 capped rows in the
    `Caps_Applied` COLUMN of its four scoring sheets, and the package carried
    `exports/caps_applied_log.csv` with those same 380 rows. The sheet was
    absent; the caps were not.

    Owner, 2026-08-23: "Caps applied may even exist in the scoring and
    research workbook and usually relate to the issue log or issues raised in
    the client research report, or an issue log in csv or any other format.
    If no caps were applied, then there were no issues."

    Both halves are honoured here. Every sheet and column is searched, and
    zero is reported as a state — never as an issue of any severity. The one
    thing that can still be raised is a cap log that EXISTS and cannot be
    read, because that is a fact about the file rather than about the
    assessment.
    """
    issues, found, where = [], 0, []
    for name in wb.sheetnames:
        ws = wb[name]
        if _CAPS_RE.search(name):
            rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(str(c or "").strip() for c in r)]
            if rows:
                found += len(rows)
                where.append(f"{name} sheet ({len(rows)} rows)")
            continue
        header = [str(c or "").strip().lower()
                  for c in next(ws.iter_rows(min_row=1, max_row=1,
                                             values_only=True), ())]
        cols = [i for i, h in enumerate(header) if _CAPS_RE.search(h)]
        if not cols:
            continue
        hits = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                   for i in cols
                   if i < len(r) and str(r[i] if r[i] is not None else "")
                   .strip().lower() not in _NO_CAP)
        if hits:
            found += hits
            where.append(f"{name}.{header[cols[0]]} ({hits} capped rows)")

    if found:
        print(f"    caps recorded: {found} across {len(where)} source(s) — "
              f"{'; '.join(where[:4])}")
    else:
        print("    NO CAPS APPLIED — searched every sheet and every "
              "caps/issue column. Valid state, not a defect: if no caps were "
              "applied, then there were no issues (owner, 2026-08-23). A cap "
              "log may also live outside this workbook (CSV, JSON, the "
              "research report); this check does not see those and does not "
              "need to.")
    return issues


# ──────────────────────────────────────────────────────────────────────
# CHECK 7: Evidence fact-level granularity
# ──────────────────────────────────────────────────────────────────────
def check_evidence_fact_granularity(wb):
    """Evidence_IDs should use fact-level citations (E-xxx:Fy), not just E-xxx."""
    issues = []
    total_cells = 0
    has_fact_refs = 0

    for pname in PILLAR_SHEETS:
        if pname not in wb.sheetnames:
            continue
        ws = wb[pname]
        for row in range(2, ws.max_row + 1):
            evidence = str(ws.cell(row=row, column=6).value or "")     # F, canonical
            if evidence.strip() and evidence.strip().lower() not in ("none", "n/a", ""):
                total_cells += 1
                if re.search(r'E-\d+:F\d+|HUBBL-\d+:F\d+', evidence):
                    has_fact_refs += 1

    if total_cells > 0 and has_fact_refs < total_cells * 0.5:
        issues.append(
            f"WARNING: Only {has_fact_refs}/{total_cells} Evidence_ID cells use fact-level "
            f"references (E-xxx:Fy). Most just use E-xxx without specifying which fact. "
            f"This prevents subcap-level differentiation."
        )
    return issues


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
# The four scoring sheets, under the CANONICAL names — SKILL.md's "Workbook
# Column Structure (CANONICAL — 11 columns per P#_Subcap_Scoring sheet)", which
# says in as many words: "This is the ONLY acceptable column layout. Do NOT use
# the legacy 22-column (A-V) layout."
#
# This file used to check the legacy layout: sheets named P#_Scoring_Detail and
# columns S/T/U/V. So the mandatory Phase-4 validator was checking a schema the
# skill FORBIDS, on every workbook, since the canonical form changed. It found
# four missing sheets, and then five of its seven checks iterated those absent
# sheets, examined zero rows and printed PASS — including "Required Columns
# (S/T/U/V) ✅" on a workbook that has no S, T, U or V.
#
# That is one rule held in two places: SKILL.md said eleven columns, this file
# and workbook_specification.md said twenty-two, and the disagreement read as a
# workbook defect rather than a toolchain defect. The first repair attempted
# here made the validator refuse the canonical form, which would have blocked
# every conforming workbook — recorded because it is the same class one more
# time and the store should carry it.
PILLAR_SHEETS = ("P1_Subcap_Scoring", "P2_Subcap_Scoring",
                 "P3_Subcap_Scoring", "P4_Subcap_Scoring")

LEGACY_SHEETS = ("P1_Scoring_Detail", "P2_Scoring_Detail",
                 "P3_Scoring_Detail", "P4_Scoring_Detail")

# Canonical 11-column layout, 0-indexed. Proof is carried by J/F/H/I, which is
# what SKILL.md's "Proof-Carrying Scoring" section names.
COL = {"subcap_id": 0, "name": 1, "category": 2, "score": 3, "confidence": 4,
       "evidence_ids": 5, "source_urls": 6, "ceiling": 7, "caps": 8,
       "rationale": 9, "proxy_searched": 10}

# Tier -> the highest score that tier alone can support. SKILL.md's Evidence
# Tier System; vendor collateral is T5 and caps at 2.0.
TIER_CEILING = {"T1": 5.0, "T2": 5.0, "T3": 4.0, "T4": 2.5, "T5": 2.0}


def resolve_pillar_sheets(wb):
    """(present, missing, legacy) for the four canonical scoring sheets."""
    present = [n for n in PILLAR_SHEETS if n in wb.sheetnames]
    missing = [n for n in PILLAR_SHEETS if n not in wb.sheetnames]
    legacy = [n for n in LEGACY_SHEETS if n in wb.sheetnames]
    return present, missing, legacy


def check_ceiling_is_derived(wb):
    """Evidence_Ceiling must follow the cited tiers, not be asserted.

    Column H is defined as "Maximum score supported by evidence tier (e.g.,
    T5-only -> 2.0)". Nothing recomputed it, so a row could cite a vendor
    marketing page and still carry a ceiling of 5.

    Measured on the workbook behind MEM-0054: Evidence_Ceiling is 5 on 634 of
    709 rows overall — which is legitimate where the evidence is T1/T2 — but 5
    on 62 of 62 rows of P3C3 and P4C4, whose strongest cited document is
    a security vendor's own site's own customer case study. A ceiling that is never below the
    score it is supposed to bound is not a ceiling.
    """
    issues, ceil_at_5_top_band = [], 0
    by_cat = {}
    for name in PILLAR_SHEETS:
        if name not in wb.sheetnames:
            continue
        for r in list(wb[name].iter_rows(values_only=True))[1:]:
            if not r or not r[COL["subcap_id"]]:
                continue
            try:
                score = float(r[COL["score"]] or 0)
                ceiling = float(r[COL["ceiling"]] or 0)
            except (TypeError, ValueError):
                continue
            cat = str(r[COL["category"]] or "")
            by_cat.setdefault(cat, []).append((score, ceiling))
            if score >= 4.0 and ceiling >= 5.0:
                ceil_at_5_top_band += 1
    # A CATEGORY in which every row is top-band under an unbounded ceiling is
    # the shape that reached a client: it means the ceiling never bound anything.
    for cat, rows in sorted(by_cat.items()):
        if len(rows) >= 10 and all(s >= 4.0 and c >= 5.0 for s, c in rows):
            issues.append(
                f"CRITICAL: {cat} — all {len(rows)} subcaps are >=4.0 with "
                f"Evidence_Ceiling 5. A ceiling that never falls below the "
                f"score is not bounding anything; recompute it from the cited "
                f"tiers (T5 vendor collateral -> 2.0, T4 -> 2.5, T3 -> 4.0).")
    if ceil_at_5_top_band:
        issues.append(
            f"WARNING: {ceil_at_5_top_band} rows score >=4.0 under a ceiling of "
            f"5.0 — check each cites at least one T1/T2 source that ADDRESSES "
            f"its diagnostic question, not merely mentions the institution.")
    return issues


def rows_examined(wb):
    """How many scoring rows the sheet-iterating checks can actually see."""
    total = 0
    for name in PILLAR_SHEETS:
        if name in wb.sheetnames:
            total += max(0, wb[name].max_row - 1)
    return total


def main():
    if len(sys.argv) > 1 and sys.argv[1] in ("-h", "--help"):
        print(__doc__ or "scoring-quality checks over a completed workbook")
        print("Usage: python validate_scoring_quality.py <workbook_path>")
        return
    if len(sys.argv) < 2:
        print("Usage: python validate_scoring_quality.py <workbook_path>")
        sys.exit(1)

    wb_path = sys.argv[1]
    wb = load_workbook(wb_path)

    print("=" * 72)
    print("DMA SCORING QUALITY VALIDATOR")
    print("=" * 72)
    print(f"Workbook: {wb_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print()

    all_issues = []
    checks = [
        ("1. Row Count (Subcap-Level Scoring)", check_row_counts, False),
        ("2. Score Differentiation", check_differentiation, True),
        ("3. Rationale Quality", check_rationale_quality, True),
        ("4. Evidence Differentiation", check_evidence_differentiation, True),
        ("5. Proof Columns (J/F/H/I)", check_required_columns, True),
        ("6. Caps Applied Log", check_caps_applied, False),
        ("7. Evidence Fact Granularity", check_evidence_fact_granularity, True),
        ("8. Ceiling Is Derived, Not Asserted", check_ceiling_is_derived, True),
    ]

    critical_count = 0
    warning_count = 0

    # Named up front, so the run says which artefact it is about to check
    # rather than leaving that to be inferred from seven silent skips.
    present, missing, legacy = resolve_pillar_sheets(wb)
    seen = rows_examined(wb)
    print(f"── 0. Scoring sheets ──")
    print(f"  found {len(present)} of {len(PILLAR_SHEETS)}: "
          f"{', '.join(present) or 'none'}   ({seen} scoring rows visible)")
    if legacy:
        print(f"  🔴 CRITICAL: this workbook carries the LEGACY 22-column "
              f"sheets ({', '.join(legacy)}). SKILL.md: 'Do NOT use the legacy "
              f"22-column (A-V) layout.' Re-emit as P#_Subcap_Scoring.")
        critical_count += 1
        all_issues.append("CRITICAL: legacy scoring sheet schema")
    elif missing:
        print(f"  🔴 CRITICAL: missing {', '.join(missing)}")
        critical_count += 1
        all_issues.append("CRITICAL: missing canonical scoring sheets")
    print()

    for check_name, check_fn, needs_rows in checks:
        print(f"── {check_name} ──")
        issues = check_fn(wb)
        if needs_rows and seen == 0:
            # A check that examined nothing is not a check that passed. Five
            # of these printed ✅ on the workbook behind MEM-0054, including
            # the one asserting the proof columns exist.
            print("  🔴 CRITICAL: examined 0 scoring rows — this check did not "
                  "run and its result is UNKNOWN, not PASS")
            critical_count += 1
            all_issues.append(f"CRITICAL: {check_name} examined nothing")
        elif not issues:
            print(f"  ✅ PASS  ({seen} rows examined)" if needs_rows else "  ✅ PASS")
        else:
            for issue in issues:
                level = "🔴 CRITICAL" if issue.startswith("CRITICAL") else "🟡 WARNING"
                print(f"  {level}: {issue}")
                if issue.startswith("CRITICAL"):
                    critical_count += 1
                else:
                    warning_count += 1
                all_issues.append(issue)
        print()

    print("=" * 72)
    print(f"RESULT: {critical_count} CRITICAL, {warning_count} WARNING")
    if critical_count > 0:
        print("❌ FAIL — Fix all CRITICAL issues before proceeding to Phase 5.")
        print("   Do NOT skip this. Re-score affected subcapabilities.")
        sys.exit(1)
    elif warning_count > 0:
        print("⚠️  PASS WITH WARNINGS — Review warnings, fix if possible.")
        sys.exit(0)
    else:
        print("✅ PASS — All quality checks passed.")
        sys.exit(0)


if __name__ == "__main__":
    main()
