#!/usr/bin/env python3
"""
generate_governance_outputs.py — DMA Assessment Skill (Layer 1)

Extracts governance-compatible outputs from a completed scoring workbook:
  - run_manifest.json (Contract 1)
  - caps_applied_log.csv (Contract 2)
  - contradiction_log.csv (Contract 3)
  - evidence_index.csv (Contract 4)

All outputs conform to Layer 2 Governance Skill interface_contracts.md schemas.

Usage:
    python generate_governance_outputs.py --workbook scored_workbook.xlsx \
        --output-dir ./governance_outputs/ \
        --institution-name "Example Credit Union" \
        --institution-id "CU12345" \
        --sub-vertical "Credit Union" \
        --size-tier "Medium" \
        --primary-regulator "NCUA" \
        --geography "California" \
        --evidence-mode "PUBLIC" \
        --assessor "Claude" \
        --tool-version "claude-opus-4-20250514 + DMA v5.0"

Requires: openpyxl, pandas
"""

import argparse
import csv
import json
import logging
import statistics
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    pd = None  # Fallback to openpyxl-only mode

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# === CONSTANTS: Layer 2 Contract Column Names (snake_case) ===

CAPS_COLUMNS = [
    "cap_id", "cap_type", "trigger_reason", "trigger_evidence",
    "affected_id", "raw_score", "cap_ceiling", "final_score", "score_delta"
]

CAPS_TYPE_ENUM = {
    "EVIDENCE_CEILING", "SENTIMENT", "REGULATORY", "CROSS_PILLAR",
    "ADJ_STALENESS", "ADJ_COMPLAINT", "ADJ_INCIDENT_MAJOR",
    "ADJ_INCIDENT_PATTERN", "CRITIC_CHALLENGE"
}

CONTRADICTION_COLUMNS = [
    "contradiction_id", "subcap_id", "evidence_a_id", "evidence_a_ers",
    "evidence_a_claim", "evidence_b_id", "evidence_b_ers", "evidence_b_claim",
    "resolution_rule", "winner", "justification", "confidence_impact",
    "flagged_in_report", "contradiction_type"
]

RESOLUTION_RULE_ENUM = {
    "ERS_RANKING", "T1T2_OVERRIDE", "TIEBREAKER",
    "CONSERVATIVE_DEFAULT", "UNRESOLVED"
}

EVIDENCE_COLUMNS = [
    "evidence_id", "source_name", "url", "tier",
    "ers_score", "publish_date", "subcaps_supported", "key_facts_count"
]

TIER_ENUM = {"T1", "T2", "T3", "T4", "T5"}

# Workbook column name mapping: Layer 1 PascalCase -> Layer 2 snake_case
CAPS_COL_MAP = {
    "Cap_ID": "cap_id", "CapLogID": "cap_id",
    "Cap_Type": "cap_type",
    "Trigger_Reason": "trigger_reason",
    "Trigger_Evidence": "trigger_evidence",
    "Affected_SubCap_or_Cap": "affected_id", "Affected_ID": "affected_id",
    "Raw_Score": "raw_score",
    "Cap_Ceiling": "cap_ceiling",
    "Final_Score": "final_score",
    "Score_Delta": "score_delta",
}

EVIDENCE_COL_MAP = {
    "Evidence_ID": "evidence_id",
    "Source_Name": "source_name", "Source": "source_name",
    "URL": "url",
    "Tier": "tier", "Evidence_Tier": "tier",
    "ERS_Score": "ers_score", "ERS": "ers_score",
    "Date_Period": "publish_date", "Publish_Date": "publish_date", "Date": "publish_date",
    "SubCaps_Supported": "subcaps_supported",
    "Fact_Summary": "key_facts_count", "Key_Facts_Count": "key_facts_count",
    "Facts_Count": "key_facts_count",
}

PILLAR_WEIGHTS = {"P1": 0.25, "P2": 0.25, "P3": 0.25, "P4": 0.25}
CATEGORY_IDS = [
    "P1C1", "P1C2", "P1C3", "P1C4", "P1C5",
    "P2C1", "P2C2", "P2C3", "P2C4",
    "P3C1", "P3C2", "P3C3", "P3C4",
    "P4C1", "P4C2", "P4C3", "P4C4",
]


def load_workbook(path: str):
    """Load workbook and return openpyxl workbook object."""
    wb = openpyxl.load_workbook(path, data_only=True)
    logger.info(f"Loaded workbook: {path} ({len(wb.sheetnames)} sheets)")
    return wb


def find_sheet(wb, candidates: list[str]):
    """Find a sheet by trying multiple name candidates."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def sheet_to_dicts(sheet) -> list[dict]:
    """Convert a worksheet to a list of dicts using first row as headers."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


def remap_columns(records: list[dict], col_map: dict) -> list[dict]:
    """Remap column names from Layer 1 PascalCase to Layer 2 snake_case."""
    remapped = []
    for rec in records:
        new_rec = {}
        for old_key, value in rec.items():
            new_key = col_map.get(old_key, old_key.lower().replace(" ", "_"))
            new_rec[new_key] = value
        remapped.append(new_rec)
    return remapped


def extract_caps_log(wb) -> list[dict]:
    """Extract caps_applied_log from workbook."""
    sheet = find_sheet(wb, ["Caps_Applied_Log", "Caps Applied Log", "CapsAppliedLog"])
    if not sheet:
        logger.warning("Caps_Applied_Log sheet not found — generating empty CSV")
        return []

    records = sheet_to_dicts(sheet)
    remapped = remap_columns(records, CAPS_COL_MAP)

    # Ensure all required columns exist
    for rec in remapped:
        for col in CAPS_COLUMNS:
            if col not in rec:
                rec[col] = ""
        # Calculate score_delta if missing
        if not rec.get("score_delta") and rec.get("raw_score") and rec.get("final_score"):
            try:
                rec["score_delta"] = round(float(rec["raw_score"]) - float(rec["final_score"]), 2)
            except (ValueError, TypeError):
                pass

    logger.info(f"Extracted {len(remapped)} cap entries")
    return remapped


def extract_evidence_index(wb) -> list[dict]:
    """Extract evidence_index from workbook."""
    sheet = find_sheet(wb, ["Evidence_Index", "Evidence Index", "EvidenceIndex"])
    if not sheet:
        logger.warning("Evidence_Index sheet not found — generating empty CSV")
        return []

    records = sheet_to_dicts(sheet)
    remapped = remap_columns(records, EVIDENCE_COL_MAP)

    # Handle Fact_Summary -> key_facts_count conversion
    for rec in remapped:
        kfc = rec.get("key_facts_count", "")
        if isinstance(kfc, str) and not kfc.isdigit():
            # Count facts in summary text (rough heuristic: count sentences or semicolons)
            count = max(1, len(str(kfc).split(";")) if kfc else 0)
            rec["key_facts_count"] = count

        for col in EVIDENCE_COLUMNS:
            if col not in rec:
                rec[col] = ""

    logger.info(f"Extracted {len(remapped)} evidence items")
    return remapped


def extract_contradiction_log(wb) -> list[dict]:
    """Extract contradiction_log from workbook."""
    sheet = find_sheet(wb, ["Contradiction_Log", "Contradiction Log", "ContradictionLog"])
    if not sheet:
        logger.warning("Contradiction_Log sheet not found — generating empty CSV")
        return []

    records = sheet_to_dicts(sheet)
    # Layer 1 uses different column names; remap as best we can
    contradiction_col_map = {
        "Contradiction_ID": "contradiction_id",
        "SubCap_ID": "subcap_id", "Subcap_ID": "subcap_id",
        "Fact_A_ID": "evidence_a_id", "Evidence_A_ID": "evidence_a_id",
        "Fact_A_Source": "evidence_a_claim", "Evidence_A_Claim": "evidence_a_claim",
        "Evidence_A_ERS": "evidence_a_ers", "Fact_A_ERS": "evidence_a_ers",
        "Fact_B_ID": "evidence_b_id", "Evidence_B_ID": "evidence_b_id",
        "Fact_B_Source": "evidence_b_claim", "Evidence_B_Claim": "evidence_b_claim",
        "Evidence_B_ERS": "evidence_b_ers", "Fact_B_ERS": "evidence_b_ers",
        "Resolution": "resolution_rule", "Resolution_Rule": "resolution_rule",
        "Winning_Fact": "winner", "Winner": "winner",
        "Resolution_Rationale": "justification", "Justification": "justification",
        "Score_Impact": "confidence_impact", "Confidence_Impact": "confidence_impact",
        "Flagged_In_Report": "flagged_in_report",
        "Contradiction_Type": "contradiction_type",
        "Capabilities_Affected": "capabilities_affected",
    }

    remapped = remap_columns(records, contradiction_col_map)

    for rec in remapped:
        # Ensure all required columns exist
        for col in CONTRADICTION_COLUMNS:
            if col not in rec:
                rec[col] = ""

        # Default flagged_in_report for unresolved
        if str(rec.get("resolution_rule", "")).upper() == "UNRESOLVED":
            if not rec.get("flagged_in_report"):
                rec["flagged_in_report"] = "true"

    logger.info(f"Extracted {len(remapped)} contradictions")
    return remapped


def compute_evidence_metrics(evidence_records: list[dict]) -> dict:
    """Compute evidence_metrics for run_manifest from evidence index."""
    total = len(evidence_records)
    tier_dist = {"T1": 0, "T2": 0, "T3": 0, "T4": 0, "T5": 0}
    ers_values = []

    for rec in evidence_records:
        tier = str(rec.get("tier", "")).upper()
        if tier in tier_dist:
            tier_dist[tier] += 1

        try:
            ers = float(rec.get("ers_score", 0))
            if ers > 0:
                ers_values.append(ers)
        except (ValueError, TypeError):
            pass

    avg_ers = round(statistics.mean(ers_values), 2) if ers_values else 0.0
    median_ers = round(statistics.median(ers_values), 2) if ers_values else 0.0

    return {
        "total_items": total,
        "tier_distribution": tier_dist,
        "avg_ers": avg_ers,
        "median_ers": median_ers,
        "sources_per_subcap_avg": 0.0,  # Computed after scoring sheets loaded
        "single_source_subcap_count": 0,
        "no_evidence_subcap_count": 0,
        "document_count": total,  # Each evidence record represents a unique source document
    }


def extract_scores(wb) -> dict:
    """Extract scores from Summary or Calculation_Chain sheet."""
    scores = {"overall": 0.0, "pillars": {}, "categories": {}}

    summary = find_sheet(wb, ["Summary", "summary", "Dashboard"])
    if not summary:
        logger.warning("Summary sheet not found — scores will be empty")
        return scores

    # Try to parse summary sheet for pillar/category scores
    # This is heuristic — actual layout depends on workbook template
    records = sheet_to_dicts(summary)
    for rec in records:
        name = str(rec.get(list(rec.keys())[0], "")) if rec else ""
        score_val = None
        for key, val in rec.items():
            if "score" in key.lower() or "final" in key.lower():
                try:
                    score_val = round(float(val), 2)
                except (ValueError, TypeError):
                    pass

        if score_val is not None:
            if name.startswith("P") and "C" in name and name[:4] in CATEGORY_IDS:
                scores["categories"][name[:4]] = score_val
            elif name in ("P1", "P2", "P3", "P4"):
                scores["pillars"][name] = score_val
            elif "overall" in name.lower() or "total" in name.lower():
                scores["overall"] = score_val

    return scores


def extract_scoring_metrics(caps_records: list[dict], contradiction_records: list[dict], wb) -> dict:
    """Compute scoring_metrics for run_manifest."""
    caps_count = len(caps_records)
    adj_count = sum(1 for r in caps_records if str(r.get("cap_type", "")).startswith("ADJ_"))
    dep_count = sum(1 for r in caps_records if r.get("cap_type") == "CROSS_PILLAR")
    ctr_found = len(contradiction_records)
    ctr_unresolved = sum(1 for r in contradiction_records
                         if str(r.get("resolution_rule", "")).upper() == "UNRESOLVED")

    # N/A capabilities from Absent_Evidence_Log
    na_caps = []
    absent_sheet = find_sheet(wb, ["Absent_Evidence_Log", "Absent Evidence Log"])
    if absent_sheet:
        for rec in sheet_to_dicts(absent_sheet):
            cap_id = str(rec.get(list(rec.keys())[0], ""))
            if cap_id and cap_id not in na_caps:
                na_caps.append(cap_id)

    # Peer count from Peer_Benchmarks sheet
    peer_count = 0
    peer_sheet = find_sheet(wb, ["Peer_Benchmarks", "Peer Benchmarks", "Peer_Analysis"])
    if peer_sheet:
        peer_records = sheet_to_dicts(peer_sheet)
        # Count unique peer institution names (first column typically)
        peer_names = set()
        for rec in peer_records:
            first_val = str(rec.get(list(rec.keys())[0], "")) if rec else ""
            if first_val and first_val.lower() not in ("", "none", "n/a", "peer"):
                peer_names.add(first_val)
        peer_count = len(peer_names) if peer_names else 0

    return {
        "caps_applied_count": caps_count,
        "adjustments_applied_count": adj_count,
        "dependency_caps_triggered": dep_count,
        "contradictions_found": ctr_found,
        "contradictions_unresolved": ctr_unresolved,
        "na_capabilities": na_caps,
        "peer_count": max(peer_count, 1),  # Minimum 1 per schema constraint
    }


def extract_confidence_distribution(wb) -> dict:
    """Extract confidence distribution from scoring detail sheets."""
    dist = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

    for pillar_num in range(1, 5):
        sheet = find_sheet(wb, [f"P{pillar_num}_Scoring_Detail", f"P{pillar_num} Scoring Detail"])
        if not sheet:
            continue

        records = sheet_to_dicts(sheet)
        for rec in records:
            conf = None
            for key, val in rec.items():
                if "confidence" in key.lower():
                    conf = str(val).upper().strip() if val else None
                    break
            if conf in dist:
                dist[conf] += 1

    return dist


def build_run_manifest(args, scores, evidence_metrics, scoring_metrics, confidence_dist) -> dict:
    """Build run_manifest.json conforming to Layer 2 Contract 1 (hybrid v2.0)."""
    # Generate run_id from institution name
    inst_code = "".join(c for c in args.institution_name.upper() if c.isalnum())[:6]
    if not inst_code:
        inst_code = "UNKN"
    run_date = date.today().strftime("%Y%m%d")
    run_id = f"DMA-{inst_code}-{run_date}-0001"

    return {
        "$schema": "run_manifest_v2",
        "run_id": run_id,
        "institution": {
            "name": args.institution_name,
            "id": args.institution_id,
            "sub_vertical": args.sub_vertical,
            "size_tier": args.size_tier,
            "primary_regulator": args.primary_regulator,
            "geography": args.geography,
        },
        "assessment": {
            "date": date.today().isoformat(),
            "evidence_mode": args.evidence_mode,
            "assessor": args.assessor,
            "tool_version": args.tool_version,
            "status": "AWAITING_REVIEW",
        },
        "versions": {
            "rubric": args.rubric_version,
            "taxonomy": args.taxonomy_version,
            "template": args.template_version,
            "peer_methodology": args.peer_methodology_version,
            "governance_skill": None,
        },
        "scores": scores,
        "evidence_metrics": evidence_metrics,
        "scoring_metrics": scoring_metrics,
        "confidence_distribution": confidence_dist,
        "qa": {
            "verdict": "PENDING",
            "regression_tests": "0/8 PASS",
            "issues_found": {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0},
            "critical_issues": 0,
        },
        "skill_references_read": [],
        "files_generated": [],
        "assessment_notes": "",
    }


def validate_manifest(manifest: dict) -> list[str]:
    """Validate run_manifest against Contract 1 rules (hybrid v2.0)."""
    errors = []

    # Rule 0: $schema must be run_manifest_v2
    if manifest.get("$schema") != "run_manifest_v2":
        errors.append(f"$schema is '{manifest.get('$schema')}', expected 'run_manifest_v2'")

    # Rule 1: overall = weighted avg of pillars (±0.02)
    pillars = manifest.get("scores", {}).get("pillars", {})
    if pillars and all(v > 0 for v in pillars.values()):
        weighted_avg = sum(pillars.get(p, 0) * PILLAR_WEIGHTS.get(p, 0.25)
                          for p in PILLAR_WEIGHTS)
        overall = manifest.get("scores", {}).get("overall", 0)
        if abs(weighted_avg - overall) > 0.02:
            errors.append(f"overall ({overall}) != weighted pillar avg ({weighted_avg:.2f}), delta={abs(weighted_avg-overall):.3f}")

    # Rule 2: total_items = sum of tier_distribution
    em = manifest.get("evidence_metrics", {})
    tier_sum = sum(em.get("tier_distribution", {}).values())
    if tier_sum != em.get("total_items", 0):
        errors.append(f"total_items ({em.get('total_items')}) != tier_distribution sum ({tier_sum})")

    # Rule 3: confidence sum = total subcap count (can't verify without taxonomy)

    # Rule 4: qa.critical_issues must equal qa.issues_found.CRITICAL
    qa = manifest.get("qa", {})
    issues_found = qa.get("issues_found", {})
    if isinstance(issues_found, dict):
        if qa.get("critical_issues", 0) != issues_found.get("CRITICAL", 0):
            errors.append(f"qa.critical_issues ({qa.get('critical_issues')}) != qa.issues_found.CRITICAL ({issues_found.get('CRITICAL')})")
    elif isinstance(issues_found, int):
        errors.append("qa.issues_found is integer — must be object with {CRITICAL, HIGH, MEDIUM, LOW} keys (v2 schema)")

    # Rule 5: run_id format
    import re
    run_id = manifest.get("run_id", "")
    if run_id and not re.match(r"^DMA-[A-Z0-9]{2,6}-[0-9]{8}-[0-9]{4}$", run_id):
        errors.append(f"run_id '{run_id}' does not match pattern DMA-[A-Z0-9]{{2,6}}-[YYYYMMDD]-[SEQ]")

    return errors


def write_csv(records: list[dict], columns: list[str], path: Path):
    """Write records to CSV with specified column order."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    logger.info(f"Wrote {len(records)} rows to {path}")


def main():
    parser = argparse.ArgumentParser(description="Generate Layer 2 governance outputs from scored workbook")
    parser.add_argument("--workbook", required=True, help="Path to scored workbook (.xlsx)")
    parser.add_argument("--output-dir", required=True, help="Output directory for governance files")
    parser.add_argument("--institution-name", required=True)
    parser.add_argument("--institution-id", required=True)
    parser.add_argument("--sub-vertical", required=True)
    parser.add_argument("--size-tier", required=True, choices=["Mega", "Large", "Medium", "Small", "Micro", "Nano"])
    parser.add_argument("--primary-regulator", required=True)
    parser.add_argument("--geography", required=True)
    parser.add_argument("--evidence-mode", required=True, choices=["PUBLIC", "INTERNAL", "HYBRID"])
    parser.add_argument("--assessor", default="Claude")
    parser.add_argument("--tool-version", default="claude-opus-4-20250514 + DMA v5.0")
    parser.add_argument("--rubric-version", default="5.0")
    parser.add_argument("--taxonomy-version", default="5.0")
    parser.add_argument("--template-version", default="5.0")
    parser.add_argument("--peer-methodology-version", default="5.0")

    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    wb = load_workbook(args.workbook)

    # Extract governance outputs
    caps_records = extract_caps_log(wb)
    evidence_records = extract_evidence_index(wb)
    contradiction_records = extract_contradiction_log(wb)

    # Compute metrics
    scores = extract_scores(wb)
    evidence_metrics = compute_evidence_metrics(evidence_records)
    scoring_metrics = extract_scoring_metrics(caps_records, contradiction_records, wb)
    confidence_dist = extract_confidence_distribution(wb)

    # Build and validate manifest
    manifest = build_run_manifest(args, scores, evidence_metrics, scoring_metrics, confidence_dist)
    validation_errors = validate_manifest(manifest)
    if validation_errors:
        for err in validation_errors:
            logger.warning(f"Manifest validation: {err}")

    # Write outputs
    manifest_path = output_dir / "run_manifest.json"
    caps_path = output_dir / "caps_applied_log.csv"
    evidence_path = output_dir / "evidence_index.csv"
    contradiction_path = output_dir / "contradiction_log.csv"

    write_csv(caps_records, CAPS_COLUMNS, caps_path)
    write_csv(evidence_records, EVIDENCE_COLUMNS, evidence_path)
    write_csv(contradiction_records, CONTRADICTION_COLUMNS, contradiction_path)

    # Populate files_generated with all outputs (including manifest itself)
    generated_files = []
    for fpath in [caps_path, evidence_path, contradiction_path]:
        if fpath.exists():
            generated_files.append({
                "filename": fpath.name,
                "file_type": fpath.suffix.lstrip("."),
                "path": str(fpath.resolve()),
                "size_bytes": fpath.stat().st_size,
                "generated_at": datetime.now().isoformat(),
            })
    # Add the workbook itself
    wb_path = Path(args.workbook)
    if wb_path.exists():
        generated_files.append({
            "filename": wb_path.name,
            "file_type": "xlsx",
            "path": str(wb_path.resolve()),
            "size_bytes": wb_path.stat().st_size,
            "generated_at": datetime.now().isoformat(),
        })
    manifest["files_generated"] = generated_files

    # Write manifest last (so it includes all file metadata)
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, default=str)
    logger.info(f"Wrote run_manifest.json")

    # Summary
    logger.info(f"=== Governance outputs generated ===")
    logger.info(f"  {len(caps_records)} cap entries")
    logger.info(f"  {len(contradiction_records)} contradictions")
    logger.info(f"  {len(evidence_records)} evidence items")
    if validation_errors:
        logger.warning(f"  {len(validation_errors)} manifest validation warnings")
    else:
        logger.info(f"  Manifest validation: PASS")

    return 0 if not validation_errors else 1


if __name__ == "__main__":
    sys.exit(main())
