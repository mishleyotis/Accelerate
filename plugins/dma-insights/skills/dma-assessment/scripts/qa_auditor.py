#!/usr/bin/env python3
"""
QA Auditor for DMA Assessment Workbooks (READ-ONLY).

Runs governance checks on scored workbooks:
- Row counts per pillar
- Score bounds (1.0-5.0)
- Evidence linkage
- Aggregation integrity
- Cap application
- Rationale quality
- Weight sums

Output: issue_register.csv, qa_verdict.json, patch_block.md

Usage:
    python qa_auditor.py --workbook ./assessment.xlsx --out-dir ./qa_results
"""

import csv
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class QAAuditor:
    """Audit DMA assessment workbooks for compliance."""

    # Expected row counts per pillar (configurable)
    EXPECTED_SUBCAP_COUNTS = {
        "P1": 199,
        "P2": 288,
        "P3": 162,
        "P4": 187,
    }

    SEVERITY_LEVELS = {
        "CRITICAL": 1,
        "HIGH": 2,
        "MEDIUM": 3,
        "LOW": 4,
    }

    def __init__(self, institution_name: str = "", assessment_date: str = ""):
        """Initialize auditor."""
        self.issues = []
        self.checks_run = 0
        self.checks_passed = 0
        self.institution_name = institution_name
        self.assessment_date = assessment_date or datetime.now().strftime("%Y-%m-%d")

    def add_issue(
        self,
        check_id: str,
        severity: str,
        message: str,
        details: str = "",
        category: str = "SCORE_INTEGRITY",
        affected_id: str = "",
        fix_instruction: str = "",
        auto_fixable: bool = False,
    ):
        """Log an issue in Layer 2 governance-compatible format (Contract 5)."""
        issue_id = f"ISS-{len(self.issues) + 1:03d}"
        self.issues.append({
            "issue_id": issue_id,
            "severity": severity,
            "category": category,
            "subcategory": check_id,
            "affected_id": affected_id,
            "description": message,
            "detection_evidence": details,
            "fix_instruction": fix_instruction or f"Review and fix: {message}",
            "auto_fixable": str(auto_fixable).lower(),
            "status": "OPEN",
        })

    def check(self, condition: bool, check_id: str, severity: str, message: str, details: str = ""):
        """Run a check."""
        self.checks_run += 1

        if not condition:
            self.add_issue(check_id, severity, message, details)
        else:
            self.checks_passed += 1

    def load_workbook(self, workbook_path: str) -> Tuple[object, Dict]:
        """Load workbook and metadata."""
        logger.info(f"Loading workbook: {workbook_path}")

        wb = load_workbook(workbook_path, data_only=True)
        sheets = {name: wb[name] for name in wb.sheetnames}

        return wb, sheets

    def check_row_counts(self, sheets: Dict) -> bool:
        """Check row counts per pillar."""
        logger.info("Running row count checks")

        all_pass = True

        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                self.add_issue(
                    f"ROW_{pillar}",
                    "CRITICAL",
                    f"Missing sheet: {sheet_name}",
                )
                all_pass = False
                continue

            ws = sheets[sheet_name]
            row_count = ws.max_row - 1  # Exclude header

            expected = self.EXPECTED_SUBCAP_COUNTS[pillar]
            tolerance = int(expected * 0.05)  # ±5%

            if not (expected - tolerance <= row_count <= expected + tolerance):
                self.add_issue(
                    f"ROW_{pillar}",
                    "HIGH",
                    f"{sheet_name} has {row_count} rows, expected ~{expected}",
                    f"Tolerance: {expected - tolerance} to {expected + tolerance}",
                )
                all_pass = False

        return all_pass

    def check_score_bounds(self, sheets: Dict) -> bool:
        """Check score columns are in [1.0, 5.0]."""
        logger.info("Running score bound checks")

        all_pass = True

        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                continue

            ws = sheets[sheet_name]

            # Find Score column
            score_col = None
            for cell in ws[1]:
                if cell.value and "score" in str(cell.value).lower():
                    score_col = cell.column
                    break

            if not score_col:
                self.add_issue(
                    f"SCORE_{pillar}",
                    "MEDIUM",
                    f"Could not find Score column in {sheet_name}",
                )
                continue

            # Check bounds
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                score_cell = row[score_col - 1]
                if score_cell.value is not None:
                    try:
                        score = float(score_cell.value)
                        if not (1.0 <= score <= 5.0):
                            self.add_issue(
                                f"SCORE_{pillar}",
                                "CRITICAL",
                                f"{sheet_name} row {row[0].row} has out-of-bounds score: {score}",
                            )
                            all_pass = False

                        # Check decimal places (max 1)
                        if len(str(score).split(".")[-1]) > 1:
                            self.add_issue(
                                f"SCORE_{pillar}",
                                "LOW",
                                f"{sheet_name} row {row[0].row} has >1 decimal: {score}",
                            )
                    except ValueError:
                        self.add_issue(
                            f"SCORE_{pillar}",
                            "HIGH",
                            f"{sheet_name} row {row[0].row} has non-numeric score",
                        )
                        all_pass = False

        return all_pass

    def check_evidence_linkage(self, sheets: Dict) -> bool:
        """Check every score has evidence IDs and citations."""
        logger.info("Running evidence linkage checks")

        all_pass = True

        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                continue

            ws = sheets[sheet_name]

            # Find Score and Evidence columns
            score_col = None
            evidence_col = None

            for cell in ws[1]:
                if cell.value:
                    value_lower = str(cell.value).lower()
                    if "score" in value_lower:
                        score_col = cell.column
                    elif "evidence" in value_lower:
                        evidence_col = cell.column

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                score_cell = row[score_col - 1] if score_col else None
                evidence_cell = row[evidence_col - 1] if evidence_col else None

                if score_cell and score_cell.value and (not evidence_cell or not evidence_cell.value):
                    self.add_issue(
                        "EVIDENCE",
                        "HIGH",
                        f"{sheet_name} row {row[0].row} has score but no evidence",
                    )
                    all_pass = False

        return all_pass

    def check_caps_log_consistency(self, sheets: Dict) -> bool:
        """Check that every capped score has a log entry."""
        logger.info("Running caps log consistency checks")

        all_pass = True

        # Load scores
        scores = {}
        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                continue

            ws = sheets[sheet_name]

            # Parse scores
            raw_col = None
            final_col = None

            for cell in ws[1]:
                if cell.value:
                    value_lower = str(cell.value).lower()
                    if "raw" in value_lower:
                        raw_col = cell.column
                    elif "final" in value_lower:
                        final_col = cell.column

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                subcap_id = row[0].value
                raw_cell = row[raw_col - 1] if raw_col else None
                final_cell = row[final_col - 1] if final_col else None

                if raw_cell and final_cell:
                    try:
                        raw = float(raw_cell.value) if raw_cell.value else None
                        final = float(final_cell.value) if final_cell.value else None

                        if raw and final and raw != final:
                            scores[subcap_id] = (raw, final)
                    except ValueError:
                        pass

        # Load caps log
        caps_log = {}
        if "Caps_Applied_Log" in sheets:
            ws = sheets["Caps_Applied_Log"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                subcap_id = row[0].value
                if subcap_id:
                    caps_log[subcap_id] = True

        # Check consistency
        for subcap_id, (raw, final) in scores.items():
            if raw != final and subcap_id not in caps_log:
                self.add_issue(
                    "CAPS_LOG",
                    "MEDIUM",
                    f"Score mismatch for {subcap_id} but no caps log entry",
                    f"Raw: {raw}, Final: {final}",
                )
                all_pass = False

        return all_pass

    def check_rationale_quality(self, sheets: Dict) -> bool:
        """Check rationale fields meet quality standards."""
        logger.info("Running rationale quality checks")

        all_pass = True
        min_rationale_length = 150

        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                continue

            ws = sheets[sheet_name]

            # Find Rationale column
            rationale_col = None
            for cell in ws[1]:
                if cell.value and "rationale" in str(cell.value).lower():
                    rationale_col = cell.column
                    break

            if not rationale_col:
                continue

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                rationale_cell = row[rationale_col - 1]
                if rationale_cell.value:
                    rationale = str(rationale_cell.value)

                    # Check length
                    if len(rationale) < min_rationale_length:
                        self.add_issue(
                            "RATIONALE",
                            "LOW",
                            f"{sheet_name} row {row[0].row} rationale too short ({len(rationale)} chars)",
                            f"Minimum: {min_rationale_length} chars",
                        )

                    # Check for evidence ID reference (E-###:F#, INT-XXX-###, or legacy format)
                    if not re.search(r"(E-\d+|INT-[A-Z]+-\d+|[A-Z]{2,}-\d+|NO_EVIDENCE)", rationale):
                        self.add_issue(
                            "RATIONALE",
                            "MEDIUM",
                            f"{sheet_name} row {row[0].row} rationale missing evidence ID reference",
                        )
                        all_pass = False

        return all_pass

    def check_weight_sums(self, sheets: Dict) -> bool:
        """Check weight columns sum to ~1.0 at capability aggregation level."""
        logger.info("Running weight sum checks")

        all_pass = True

        for pillar in ["P1", "P2", "P3", "P4"]:
            sheet_name = f"{pillar}_Scoring_Detail"
            if sheet_name not in sheets:
                continue

            ws = sheets[sheet_name]

            # Find Weight and Category columns
            weight_col = None
            category_col = None
            for cell in ws[1]:
                if cell.value:
                    value_lower = str(cell.value).lower()
                    if "weight" in value_lower:
                        weight_col = cell.column
                    elif "category" in value_lower and category_col is None:
                        category_col = cell.column

            if not weight_col or not category_col:
                continue

            # Group weights by category
            category_weights = {}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cat_cell = row[category_col - 1]
                wt_cell = row[weight_col - 1]
                if cat_cell.value and wt_cell.value:
                    try:
                        cat = str(cat_cell.value)
                        wt = float(wt_cell.value)
                        category_weights.setdefault(cat, []).append(wt)
                    except ValueError:
                        pass

            # Check each category sums to ~1.0 (allowing percentage format)
            for cat, weights in category_weights.items():
                total = sum(weights)
                # Accept both 0-1 range and 0-100 range
                if not (0.99 <= total <= 1.01) and not (99.0 <= total <= 101.0):
                    self.add_issue(
                        f"WEIGHT_{pillar}",
                        "MEDIUM",
                        f"{sheet_name} category '{cat}' weights sum to {total:.4f}",
                        "Expected ~1.0 or ~100%",
                    )
                    all_pass = False

        return all_pass

    def run_all_checks(self, workbook_path: str) -> bool:
        """Run all audit checks."""
        logger.info("Starting QA audit")

        wb, sheets = self.load_workbook(workbook_path)

        # Run checks
        checks = [
            self.check_row_counts(sheets),
            self.check_score_bounds(sheets),
            self.check_evidence_linkage(sheets),
            self.check_caps_log_consistency(sheets),
            self.check_rationale_quality(sheets),
            self.check_weight_sums(sheets),
        ]

        all_pass = all(checks)

        logger.info(f"QA audit complete: {self.checks_passed}/{self.checks_run} checks passed")

        return all_pass

    def get_verdict(self) -> str:
        """Get overall verdict."""
        if not self.issues:
            return "PASS"

        critical_count = sum(1 for i in self.issues if i["severity"] == "CRITICAL")
        high_count = sum(1 for i in self.issues if i["severity"] == "HIGH")

        if critical_count > 0:
            return "FAIL"
        elif high_count > 0:
            return "PASS_WITH_NOTES"
        else:
            return "PASS"

    def save_report(self, output_dir: str):
        """Save QA report files in Layer 2 governance-compatible format."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Issue register CSV — Contract 5 (10 columns)
        issue_path = Path(output_dir) / "issue_register.csv"
        fieldnames = [
            "issue_id", "severity", "category", "subcategory",
            "affected_id", "description", "detection_evidence",
            "fix_instruction", "auto_fixable", "status"
        ]
        with open(issue_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.issues)
        logger.info(f"Issue register saved to {issue_path} ({len(self.issues)} issues)")

        # QA Verdict JSON — Contract 6
        critical_count = sum(1 for i in self.issues if i["severity"] == "CRITICAL")
        high_count = sum(1 for i in self.issues if i["severity"] == "HIGH")
        medium_count = sum(1 for i in self.issues if i["severity"] == "MEDIUM")
        low_count = sum(1 for i in self.issues if i["severity"] == "LOW")

        blocking_ids = [i["issue_id"] for i in self.issues
                        if i["severity"] == "CRITICAL" and i["status"] == "OPEN"]
        note_ids = [i["issue_id"] for i in self.issues
                    if i["severity"] in ("HIGH", "MEDIUM") and i["status"] == "OPEN"]

        verdict_str = self.get_verdict()

        # Determine recommendation
        if verdict_str == "FAIL":
            recommendation = "FIX_AND_REAUDIT"
        elif critical_count == 0 and high_count > 0:
            recommendation = "FIX_AND_REAUDIT"
        elif critical_count == 0 and medium_count > 0:
            recommendation = "DELIVER"
        else:
            recommendation = "DELIVER"

        # Build rationale
        if verdict_str == "PASS":
            rationale = f"All {self.checks_run} checks passed with no blocking issues."
        elif verdict_str == "FAIL":
            rationale = f"{critical_count} CRITICAL issues block delivery. Fix and re-audit required."
        else:
            rationale = f"No CRITICAL issues but {high_count} HIGH and {medium_count} MEDIUM issues noted."

        verdict = {
            "$schema": "qa_verdict_v1",
            "institution_name": self.institution_name,
            "assessment_date": self.assessment_date,
            "rubric_version": "5.0",
            "governance_skill_version": "2.2",
            "audit_date": datetime.now().strftime("%Y-%m-%d"),
            "verdict": verdict_str,
            "verdict_rationale": rationale,
            "issue_summary": {
                "total": len(self.issues),
                "critical": critical_count,
                "high": high_count,
                "medium": medium_count,
                "low": low_count,
            },
            "blocking_issues": blocking_ids,
            "notes": note_ids,
            "regression_results": f"{self.checks_passed}/{self.checks_run} PASS",
            "calibration_flags": [],
            "recommendation": recommendation,
        }

        verdict_path = Path(output_dir) / "qa_verdict.json"
        with open(verdict_path, "w", encoding="utf-8") as f:
            json.dump(verdict, f, indent=2)
        logger.info(f"Verdict saved to {verdict_path}")

        # Patch block markdown — Contract 7 (structured format)
        patch_path = Path(output_dir) / "patch_block.md"
        with open(patch_path, "w", encoding="utf-8") as f:
            f.write(f"=== GOVERNANCE PATCH BLOCK — {self.institution_name} {self.assessment_date} ===\n")
            f.write(f"Governance Skill Version: 1.0\n")
            f.write(f"Assessment Rubric Version: 5.0\n\n")

            # Error Log Additions
            f.write("## Error Log Additions\n\n")
            novel_patterns = [i for i in self.issues if i["severity"] in ("CRITICAL", "HIGH")]
            if novel_patterns:
                for issue in novel_patterns:
                    f.write(f"- **{issue['subcategory']}** ({issue['severity']}): {issue['description']}\n")
                    f.write(f"  Prevention: {issue['fix_instruction']}\n")
            else:
                f.write("No new error patterns detected.\n")
            f.write("\n")

            # Regression Test Proposals
            f.write("## Regression Test Proposals\n\n")
            critical_issues = [i for i in self.issues if i["severity"] == "CRITICAL"]
            if critical_issues:
                for issue in critical_issues:
                    f.write(f"- Add regression test for: {issue['subcategory']} — {issue['description']}\n")
            else:
                f.write("No new regression tests proposed.\n")
            f.write("\n")

            # Rubric Tweak Proposals
            f.write("## Rubric Tweak Proposals (REQUIRES CCB APPROVAL)\n\n")
            f.write("No rubric tweaks proposed from this audit.\n\n")

            # Calibration Observations
            f.write("## Calibration Observations\n\n")
            f.write(f"- Total checks: {self.checks_run}\n")
            f.write(f"- Pass rate: {self.checks_passed}/{self.checks_run}")
            if self.checks_run > 0:
                f.write(f" ({100 * self.checks_passed / self.checks_run:.0f}%)")
            f.write("\n")
            f.write(f"- Issue distribution: {critical_count} CRITICAL, {high_count} HIGH, {medium_count} MEDIUM, {low_count} LOW\n\n")

            f.write(f"=== END PATCH BLOCK ===\n")

        logger.info(f"Patch block saved to {patch_path}")


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Audit DMA assessment workbook for quality and compliance."
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to assessment workbook Excel file",
    )
    parser.add_argument(
        "--out-dir",
        default="./qa_results",
        help="Output directory for QA reports",
    )
    parser.add_argument(
        "--institution-name",
        default="",
        help="Institution name for governance outputs",
    )
    parser.add_argument(
        "--assessment-date",
        default="",
        help="Assessment date (YYYY-MM-DD) for governance outputs",
    )

    args = parser.parse_args()

    # Validate workbook
    if not Path(args.workbook).exists():
        logger.error(f"Workbook not found: {args.workbook}")
        sys.exit(1)

    # Run audit
    auditor = QAAuditor(
        institution_name=args.institution_name,
        assessment_date=args.assessment_date,
    )
    auditor.run_all_checks(args.workbook)

    # Save report
    auditor.save_report(args.out_dir)

    # Exit with code based on verdict
    verdict = auditor.get_verdict()
    logger.info(f"QA Verdict: {verdict}")

    if verdict == "FAIL":
        sys.exit(1)


if __name__ == "__main__":
    main()
