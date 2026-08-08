#!/usr/bin/env python3
"""
Taxonomy-Aware Query Generation for DMA Assessment.

Loads Pillar XLSX files, generates multi-angle retrieval queries for each subcapability.

Usage:
    python subcap_query_builder.py --pillar-dir /path/to/pillars/ --out subcap_queries.json
"""

import json
import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class SubcapabilityQueryBuilder:
    """Build taxonomy-aware queries from Pillar XLSX files."""

    def __init__(self):
        """Initialize builder."""
        self.queries = {}
        self.taxonomy = {}

    def load_pillar_file(self, pillar_path: str) -> Dict[str, List[dict]]:
        """Load taxonomy from a Pillar XLSX file.

        Expected structure:
        - Worksheet with columns: Category, Capability, Subcapability, ID, ...
        """
        logger.info(f"Loading {pillar_path}")

        wb = load_workbook(pillar_path, data_only=True)
        ws = wb.active

        # Find header row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and any(row):
                headers = [h.lower() if h else "" for h in row]
                if any(h in headers for h in ["category", "capability", "subcapability"]):
                    break

        if not headers:
            logger.warning(f"Could not find header row in {pillar_path}")
            return {}

        logger.info(f"Headers: {headers}")

        # Extract data
        data = []
        for row in ws.iter_rows(min_row=len(headers) + 1, values_only=True):
            if row and any(row):
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                data.append(row_dict)

        return {
            "file": Path(pillar_path).name,
            "data": data,
            "headers": headers,
        }

    def generate_queries_for_subcap(
        self,
        subcap_id: str,
        subcap_name: str,
        capability_name: Optional[str] = None,
        category_name: Optional[str] = None,
    ) -> Dict[str, List[str]]:
        """Generate multi-angle queries for a subcapability.

        Returns dict with query keys (canonical, controls, metrics, operating_model)
        """
        queries = {}

        # Canonical query: subcap name + capability + category
        canonical_parts = [subcap_name]
        if capability_name:
            canonical_parts.append(capability_name)
        if category_name:
            canonical_parts.append(category_name)
        queries["canonical"] = " ".join(canonical_parts)

        # Controls query: governance, policy, framework
        queries["controls"] = f"{subcap_name} controls governance policy framework"

        # Metrics query: KPI, dashboard, measurement, analytics
        queries["metrics"] = f"{subcap_name} KPI metrics measurement dashboard analytics"

        # Operating model query: process, workflow, automation, orchestration
        queries["operating_model"] = f"{subcap_name} process workflow automation orchestration"

        return queries

    def process_pillars(self, pillar_dir: str) -> Dict[str, Dict[str, List[str]]]:
        """Process all Pillar XLSX files in directory."""
        logger.info(f"Processing pillars from {pillar_dir}")

        pillar_files = sorted(Path(pillar_dir).glob("Pillar*.xlsx"))

        if not pillar_files:
            logger.warning(f"No Pillar*.xlsx files found in {pillar_dir}")
            return {}

        all_queries = {}

        for pillar_file in pillar_files:
            logger.info(f"Processing {pillar_file.name}")

            try:
                pillar_data = self.load_pillar_file(str(pillar_file))

                if not pillar_data.get("data"):
                    logger.warning(f"No data extracted from {pillar_file.name}")
                    continue

                for row in pillar_data["data"]:
                    # Extract IDs and names
                    subcap_id = row.get("id") or row.get("subcap_id") or row.get("Sub_ID")
                    subcap_name = row.get("subcapability") or row.get("Sub_Capability")
                    capability_name = row.get("capability") or row.get("Capability")
                    category_name = row.get("category") or row.get("Category")

                    if not subcap_id or not subcap_name:
                        continue

                    # Ensure string
                    subcap_id = str(subcap_id).strip()
                    subcap_name = str(subcap_name).strip()

                    if subcap_id and subcap_name:
                        queries = self.generate_queries_for_subcap(
                            subcap_id,
                            subcap_name,
                            capability_name,
                            category_name,
                        )
                        all_queries[subcap_id] = queries

            except Exception as e:
                logger.error(f"Failed to process {pillar_file.name}: {e}")
                continue

        logger.info(f"Generated queries for {len(all_queries)} subcapabilities")
        return all_queries

    def save_queries(self, queries: Dict, output_path: str):
        """Save queries to JSON."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w") as f:
            json.dump(queries, f, indent=2)

        logger.info(f"Saved {len(queries)} query sets to {output_path}")


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate taxonomy-aware queries from Pillar XLSX files."
    )
    parser.add_argument(
        "--pillar-dir",
        required=True,
        help="Directory containing Pillar*.xlsx files",
    )
    parser.add_argument(
        "--out",
        default="subcap_queries.json",
        help="Output JSON file path",
    )

    args = parser.parse_args()

    # Validate directory
    if not Path(args.pillar_dir).is_dir():
        logger.error(f"Pillar directory not found: {args.pillar_dir}")
        sys.exit(1)

    # Build queries
    builder = SubcapabilityQueryBuilder()
    queries = builder.process_pillars(args.pillar_dir)

    if not queries:
        logger.error("No queries generated")
        sys.exit(1)

    # Save
    builder.save_queries(queries, args.out)


if __name__ == "__main__":
    main()
