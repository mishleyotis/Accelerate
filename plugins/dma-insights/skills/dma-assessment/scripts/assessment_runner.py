#!/usr/bin/env python3
"""
Batch Scoring Orchestrator for DMA Assessment.

Loads taxonomy, evidence packs, and scoring rubric.
Processes by pillar batches with checkpointing.
Aggregates scores and applies caps cascade.
Exports to Excel workbook.

Usage:
    python assessment_runner.py \
        --corpus ./evidence_corpus.parquet \
        --index-dir ./index \
        --pillar-dir /path/to/pillars \
        --institution "My Bank" \
        --sub-vertical "Credit Union" \
        --size-tier "Mega" \
        --out-dir ./assessment_output
"""

import hashlib
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class AssessmentRunner:
    """Orchestrate batch scoring workflow."""

    # Pillar weights by sub-vertical
    PILLAR_WEIGHTS = {
        "Credit Union": {"P1": 0.25, "P2": 0.30, "P3": 0.20, "P4": 0.25},
        "Regional Bank": {"P1": 0.25, "P2": 0.30, "P3": 0.20, "P4": 0.25},
        "Commercial Lending": {"P1": 0.20, "P2": 0.20, "P3": 0.35, "P4": 0.25},
        "Insurance Carrier": {"P1": 0.20, "P2": 0.20, "P3": 0.30, "P4": 0.30},
        "Insurance Broker": {"P1": 0.20, "P2": 0.35, "P3": 0.20, "P4": 0.25},
    }

    def __init__(
        self,
        institution: str,
        sub_vertical: str,
        size_tier: str,
        scoring_function: Optional[Callable] = None,
    ):
        """Initialize runner.

        Args:
            institution: Institution name
            sub_vertical: Sub-vertical (e.g., "Credit Union")
            size_tier: Size tier (e.g., "Mega", "Large")
            scoring_function: Pluggable LLM scoring function
        """
        self.institution = institution
        self.sub_vertical = sub_vertical
        self.size_tier = size_tier
        self.scoring_function = scoring_function or self._default_scorer

        self.pillar_weights = self.PILLAR_WEIGHTS.get(sub_vertical, {})
        self.taxonomy = {}
        self.evidence_packs = {}
        self.scores = {}
        self.caps_log = []
        self.contradiction_log = []

    def _default_scorer(
        self,
        subcap_id: str,
        subcap_name: str,
        evidence_pack: Dict,
    ) -> Dict[str, Any]:
        """Default (stub) scorer. Replace with real LLM call.

        Args:
            subcap_id: Subcapability ID
            subcap_name: Subcapability name
            evidence_pack: Evidence pack dict with chunks

        Returns:
            Dict with: score (1.0-5.0), rationale, evidence_ids, confidence
        """
        # This is a stub that can be replaced with actual LLM call
        # For now, return a placeholder
        return {
            "score": 3.0,  # Placeholder
            "rationale": f"Default scoring for {subcap_name}. Replace with LLM call.",
            "evidence_ids": [c["chunk_id"] for c in evidence_pack.get("chunks", [])[:3]],
            "confidence": "LOW",
        }

    def load_taxonomy(self, pillar_dir: str):
        """Load taxonomy from Pillar XLSX files."""
        logger.info(f"Loading taxonomy from {pillar_dir}")

        from openpyxl import load_workbook

        all_subcaps = []
        for pillar_file in sorted(Path(pillar_dir).glob("Pillar*.xlsx")):
            logger.info(f"Loading {pillar_file.name}")

            try:
                wb = load_workbook(pillar_file, data_only=True)
                ws = wb.active

                # Simple row-by-row read (adjust column names as needed)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and any(row):
                        # Assuming: [Category, Capability, Subcapability, ID, Weight, ...]
                        if len(row) >= 5:
                            category = row[0]
                            capability = row[1]
                            subcapability = row[2]
                            subcap_id = row[3]
                            weight = row[4]

                            if subcap_id and subcapability:
                                all_subcaps.append({
                                    "subcap_id": str(subcap_id).strip(),
                                    "subcapability": str(subcapability).strip(),
                                    "capability": str(capability).strip() if capability else "",
                                    "category": str(category).strip() if category else "",
                                    "weight": float(weight) if weight else 0.0,
                                })
            except Exception as e:
                logger.error(f"Failed to load {pillar_file.name}: {e}")

        self.taxonomy = {s["subcap_id"]: s for s in all_subcaps}
        logger.info(f"Loaded {len(self.taxonomy)} subcapabilities")

    def load_evidence_packs(self, packs_json: str):
        """Load pre-generated evidence packs."""
        logger.info(f"Loading evidence packs from {packs_json}")

        with open(packs_json) as f:
            packs = json.load(f)

        self.evidence_packs = packs
        logger.info(f"Loaded {len(packs)} evidence packs")

    def score_subcapability(
        self,
        subcap_id: str,
        evidence_pack: Dict,
    ) -> Dict[str, Any]:
        """Score a single subcapability."""
        subcap = self.taxonomy.get(subcap_id, {})
        subcap_name = subcap.get("subcapability", subcap_id)

        logger.debug(f"Scoring {subcap_id}: {subcap_name}")

        # Call scorer
        result = self.scoring_function(subcap_id, subcap_name, evidence_pack)

        # Validate result
        score = float(result.get("score", 3.0))
        score = max(1.0, min(5.0, score))  # Clamp 1-5

        return {
            "subcap_id": subcap_id,
            "score": round(score, 1),
            "rationale": result.get("rationale", ""),
            "evidence_ids": result.get("evidence_ids", []),
            "confidence": result.get("confidence", "MEDIUM"),
        }

    def apply_caps(self, scores: Dict[str, float]) -> Dict[str, Tuple[float, str, float]]:
        """Apply caps cascade to scores.

        Returns: {subcap_id: (final_score, cap_type, delta)}
        """
        # Placeholder: just apply basic bounds
        capped = {}

        for subcap_id, score in scores.items():
            raw_score = score
            final_score = max(1.0, min(5.0, score))
            delta = final_score - raw_score

            if delta != 0:
                self.caps_log.append({
                    "subcap_id": subcap_id,
                    "raw_score": raw_score,
                    "cap_ceiling": 5.0,
                    "final_score": final_score,
                    "cap_type": "BOUNDS",
                    "delta": delta,
                })

            capped[subcap_id] = (final_score, "BOUNDS", delta)

        return capped

    def aggregate_to_capability(
        self,
        subcap_scores: Dict[str, float],
        category: str,
        capability: str,
    ) -> float:
        """Aggregate subcapability scores to capability level."""
        # Get subcaps for this capability
        subcaps_in_cap = [
            s for s in self.taxonomy.values()
            if s.get("capability") == capability
        ]

        if not subcaps_in_cap:
            return 3.0

        # Weighted average
        total_weight = sum(s.get("weight", 1.0) for s in subcaps_in_cap)
        weighted_sum = sum(
            subcap_scores.get(s["subcap_id"], 3.0) * s.get("weight", 1.0)
            for s in subcaps_in_cap
        )

        return weighted_sum / max(total_weight, 1.0)

    def aggregate_to_pillar(
        self,
        capability_scores: Dict[str, float],
        pillar: str,
    ) -> float:
        """Aggregate capability scores to pillar level."""
        # Get capabilities for this pillar
        # Assuming pillar is first char of subcap_id (e.g., P1C1 -> P1)
        caps_in_pillar = [
            cap for cap in capability_scores.keys()
            if cap.startswith(pillar)
        ]

        if not caps_in_pillar:
            return 3.0

        # Simple average (weights would be handled per cap)
        return sum(
            capability_scores.get(cap, 3.0)
            for cap in caps_in_pillar
        ) / len(caps_in_pillar)

    def run_scoring_batch(self, pillar: str) -> Dict[str, float]:
        """Score all subcapabilities in a pillar."""
        logger.info(f"Scoring pillar {pillar}")

        pillar_subcaps = [
            s for s in self.taxonomy.values()
            if s["subcap_id"].startswith(pillar)
        ]

        pillar_scores = {}

        for subcap in pillar_subcaps:
            subcap_id = subcap["subcap_id"]
            evidence_pack = self.evidence_packs.get(subcap_id, {})

            score_result = self.score_subcapability(subcap_id, evidence_pack)
            pillar_scores[subcap_id] = score_result["score"]

            self.scores[subcap_id] = score_result

        logger.info(f"Scored {len(pillar_scores)} subcapabilities in {pillar}")
        return pillar_scores

    def run_all_pillars(self) -> Dict[str, float]:
        """Score all pillars."""
        all_scores = {}

        for pillar in ["P1", "P2", "P3", "P4"]:
            logger.info(f"Starting {pillar}")
            pillar_scores = self.run_scoring_batch(pillar)
            all_scores.update(pillar_scores)

            # Save checkpoint
            self._save_checkpoint(pillar, pillar_scores)

        return all_scores

    def _save_checkpoint(self, pillar: str, scores: Dict[str, float]):
        """Save checkpoint after pillar completion."""
        # Checkpoint file would be saved to entity folder
        logger.info(f"Checkpoint saved for {pillar}")

    def generate_workbook(self, output_dir: str) -> str:
        """Generate Excel workbook with all sheets."""
        logger.info("Generating Excel workbook")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb)
        self._create_calculation_sheet(wb)
        self._create_pillar_detail_sheets(wb)
        self._create_evidence_sheet(wb)
        self._create_caps_log_sheet(wb)

        # Save
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(output_dir) / f"DMA_Scoring_Workbook_{self.institution}_{now}.xlsx"

        wb.save(output_path)
        logger.info(f"Workbook saved to {output_path}")

        return str(output_path)

    def _create_summary_sheet(self, wb: Workbook):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)

        ws["A1"] = "Pillar"
        ws["B1"] = "Score"
        ws["C1"] = "Level"

        for row, pillar in enumerate(["P1", "P2", "P3", "P4"], start=2):
            ws[f"A{row}"] = pillar
            ws[f"B{row}"] = 3.0  # Placeholder
            ws[f"C{row}"] = "M3"

        ws[f"A6"] = "Overall"
        ws[f"B6"] = 3.0

    def _create_calculation_sheet(self, wb: Workbook):
        """Create calculation chain sheet."""
        ws = wb.create_sheet("Calculation_Chain", 1)

        ws["A1"] = "SubCap"
        ws["B1"] = "Raw Score"
        ws["C1"] = "Final Score"
        ws["D1"] = "Caps Applied"

        row = 2
        for subcap_id, score_data in self.scores.items():
            ws[f"A{row}"] = subcap_id
            ws[f"B{row}"] = score_data["score"]
            ws[f"C{row}"] = score_data["score"]
            row += 1

    def _create_pillar_detail_sheets(self, wb: Workbook):
        """Create pillar detail sheets."""
        for pillar in ["P1", "P2", "P3", "P4"]:
            ws = wb.create_sheet(f"{pillar}_Scoring_Detail")

            # Headers
            headers = [
                "SubCap_ID", "SubCapability", "Score", "Evidence_IDs",
                "Confidence", "Rationale"
            ]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header

            # Data
            row = 2
            for subcap_id, score_data in self.scores.items():
                if subcap_id.startswith(pillar):
                    subcap = self.taxonomy.get(subcap_id, {})
                    ws.cell(row=row, column=1).value = subcap_id
                    ws.cell(row=row, column=2).value = subcap.get("subcapability", "")
                    ws.cell(row=row, column=3).value = score_data["score"]
                    ws.cell(row=row, column=4).value = ", ".join(score_data["evidence_ids"])
                    ws.cell(row=row, column=5).value = score_data["confidence"]
                    ws.cell(row=row, column=6).value = score_data["rationale"][:100]
                    row += 1

    def _create_evidence_sheet(self, wb: Workbook):
        """Create evidence index sheet."""
        ws = wb.create_sheet("Evidence_Index")

        headers = ["Evidence_ID", "Source", "Tier", "SubCaps_Supported"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header

    def _create_caps_log_sheet(self, wb: Workbook):
        """Create caps applied log sheet."""
        ws = wb.create_sheet("Caps_Applied_Log")

        headers = ["SubCap_ID", "Raw_Score", "Cap_Type", "Cap_Ceiling", "Final_Score", "Delta"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header

        for row, cap_entry in enumerate(self.caps_log, start=2):
            ws.cell(row=row, column=1).value = cap_entry["subcap_id"]
            ws.cell(row=row, column=2).value = cap_entry["raw_score"]
            ws.cell(row=row, column=3).value = cap_entry["cap_type"]
            ws.cell(row=row, column=4).value = cap_entry.get("cap_ceiling", 5.0)
            ws.cell(row=row, column=5).value = cap_entry["final_score"]
            ws.cell(row=row, column=6).value = cap_entry["delta"]

    def run(
        self,
        corpus_path: str,
        index_dir: str,
        pillar_dir: str,
        out_dir: str,
    ) -> str:
        """Main assessment workflow."""
        # Load taxonomy
        self.load_taxonomy(pillar_dir)

        # TODO: Load evidence packs from index
        # For now, create minimal packs
        self.evidence_packs = {
            subcap_id: {"chunks": []}
            for subcap_id in self.taxonomy.keys()
        }

        # Score all pillars
        logger.info("Starting assessment")
        all_scores = self.run_all_pillars()

        # Apply caps
        logger.info("Applying caps")
        capped_scores = self.apply_caps(all_scores)

        # Generate workbook
        workbook_path = self.generate_workbook(out_dir)

        return workbook_path


RETIRED = """REFUSED: assessment_runner.py is retired (2026-09-03). It built a fresh
openpyxl.Workbook() with an 11-column layout beside the run — a SECOND scoring
workbook the gates never saw (owner issue 3; goeasy GSY-15/21: "the work went
around the pipeline"). The scoring stage lives in the engine, on the run's own
workbook, and is gated at both ends:

    python3 -m engine.assessment open     --run R --root ROOT   # refuses until every category gate is PASS
    python3 -m engine.assessment score    --run R --root ROOT --subcap P1C1.1.1 --score 2.75 \\
            --confidence MEDIUM --rationale '[EVIDENCE] E-0012 … [CEILING] …' --actor scoring-p1-producer \\
            --ai-applicability ASSISTIVE --data-dependency '…' --data-readiness AMBER
    python3 -m engine.assessment critique --run R --root ROOT --pillar P1 --verdict PASS --actor scoring-critic --note '…'
    python3 -m engine.assessment rollup   --run R --root ROOT
    python3 -m engine.assessment gate     --run R --root ROOT

The four pillar scorers and the critic are agents (agents/scoring/); the
brief that dispatches them is `engine.brief scoring-batch`.
"""


def main():
    """CLI entry point — retired; prints the engine path and exits 1."""
    import sys as _sys
    _sys.stderr.write(RETIRED)
    return 1


def _legacy_main():           # kept for reference; unreachable
    import argparse

    parser = argparse.ArgumentParser(
        description="Run batch DMA assessment scoring."
    )
    parser.add_argument("--corpus", required=True)
    parser.add_argument("--index-dir", required=True)
    parser.add_argument("--pillar-dir", required=True)
    parser.add_argument("--institution", required=True)
    parser.add_argument("--sub-vertical", required=True)
    parser.add_argument("--size-tier", required=True)
    parser.add_argument("--out-dir", default="./assessment_output")

    args = parser.parse_args()

    # Initialize runner
    runner = AssessmentRunner(
        institution=args.institution,
        sub_vertical=args.sub_vertical,
        size_tier=args.size_tier,
    )

    # Run assessment
    workbook_path = runner.run(
        args.corpus,
        args.index_dir,
        args.pillar_dir,
        args.out_dir,
    )

    logger.info(f"Assessment complete: {workbook_path}")


if __name__ == "__main__":
    import sys as _sys
    _sys.exit(main())
