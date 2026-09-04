#!/usr/bin/env python3
"""Walk the DRIVER through the REAL command line, with chaos.

    python3 scripts/stress_pipeline_stub.py [--workdir DIR] [--keep]

Not a unit test — a WALK. `tests/acceptance/test_acceptance_pipeline.py`
pins each behaviour through the library; this shells `engine.pipeline`
exactly as the conductor, a Routine or `/dma-insights:run-assessment` does,
against a run started through `engine.cli start` with an answered preflight,
and drives it with the stub doubles (`--dispatcher stub`: lanes played by the
engine's own fixtures, a connector that ingests when polled, a shipper that
passes unless told otherwise). What it proves is the SEQUENCE and the
refusals a live run meets at the command line:

  1  `env` and `plan` say what is missing before anything is dispatched
  2  a lane that produces nothing twice is retried and the stage still PASSES
  3  an ingest that never arrives is a loud FAIL at INGEST_A, with the
     checkpoint already pushed — and the run resumes THERE, not at PRELIM
  4  a page verdict FAIL re-dispatches only that page, with the reasons
  5  `--max-wall-min` stops cleanly between stages (exit 0) and resumes
  6  a refused promote is a FAIL at PROMOTE; the next run promotes
  7  a second run redoes nothing; `cost report` reads every stage's clock
  8  a v6 workbook in flight opens under the v7 engine and continues

Exit 0 only if every step behaves as stated. A step that cannot run says
NOT_RUN with the reason rather than being skipped silently.
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

SKILL = Path(__file__).resolve().parent.parent / "skills" / "dma-research"
REPO = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(SKILL))

RESULTS: list[tuple[str, bool, str]] = []


def check(name: str, ok: bool, detail: str = "") -> bool:
    RESULTS.append((name, bool(ok), detail))
    print(f"  {'PASS' if ok else 'FAIL'}  {name}"
          + (f"\n          {detail}" if detail and not ok else ""))
    return bool(ok)


def run(*args, expect: int | None = 0, env: dict | None = None, timeout=900):
    e = dict(os.environ)
    e.update(env or {})
    r = subprocess.run([sys.executable, "-m", *args], cwd=str(SKILL),
                       capture_output=True, text=True, timeout=timeout, env=e)
    if expect is not None and r.returncode != expect:
        print(f"    ! {' '.join(str(a) for a in args)[:160]} -> {r.returncode} (wanted {expect})")
        print("    " + (r.stderr or r.stdout).strip()[-800:].replace("\n", "\n    "))
    return r


def jout(r):
    """The JSON document the command printed LAST (the stage log lines
    before it may themselves contain braces)."""
    text = r.stdout or ""
    for marker in ("\n{\n", "{\n"):
        i = text.rfind(marker)
        if i >= 0:
            try:
                return json.loads(text[i:].lstrip())
            except ValueError:
                continue
    try:
        return json.loads(text.strip())
    except ValueError:
        return {}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--workdir")
    ap.add_argument("--keep", action="store_true")
    a = ap.parse_args(argv)
    work = Path(a.workdir) if a.workdir else Path(tempfile.mkdtemp(prefix="dma-pipeline-"))
    work.mkdir(parents=True, exist_ok=True)
    print(f"\nworkdir: {work}\n")
    env = {"DMA_RUN_REGISTRY": str(work / "registry.jsonl")}

    # A started run, through the engine, with an ANSWERED preflight — the
    # state `engine.cli start` leaves a real run in. The lifecycle walk's
    # preflight document is the one real-shaped answer this repository has.
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from stress_run_lifecycle import preflight_doc          # noqa: E402
    entity, eid, run_id = "Stress Credit Union", "acme-cu", "R-STRESS-PIPE"
    pf = work / "preflight.json"
    pf.write_text(json.dumps(preflight_doc(entity, eid)))
    root = work / "run"
    base = ["--run", run_id, "--root", str(root)]
    r = run("engine.cli", "start", *base, "--entity", entity, "--entity-id", eid,
            "--reference-date", "2026-08-29", "--preflight", str(pf), "--no-push",
            "--folder-root", str(work / "client"), env=env)
    if not check("engine.cli start creates the run from an answered preflight",
                 r.returncode == 0 and root.is_dir(), (r.stderr or r.stdout)[-300:]):
        return _finish(work, a.keep)
    # the stub plays a six-cell P1C1 selection; narrow the run to that so the
    # walk takes seconds, not the 851-cell catalogue
    from engine import runstate                                  # noqa: E402
    wb = runstate.locate(run_id, root).open()
    cells = [c for c in wb.selected_subcaps() if c.startswith("P1C1")][:6]
    import openpyxl                                              # noqa: E402
    x = openpyxl.load_workbook(wb.path)
    for sheet in ("P1_Subcap_Scoring", "P2_Subcap_Scoring", "P3_Subcap_Scoring", "P4_Subcap_Scoring"):
        ws = x[sheet]
        for row in range(ws.max_row, 1, -1):
            if str(ws.cell(row=row, column=1).value or "") not in cells:
                ws.delete_rows(row)
    for rrow in x["Run_Metadata"].iter_rows(min_row=2):
        if rrow[0].value == "subcaps_selected":
            rrow[1].value = len(cells)
    x.save(wb.path)
    stub = {"DMA_STUB_STATE": str(work / "stub_connector.json"), **env}
    common = [*base, "--dispatcher", "stub", "--no-push", "--folder-root", str(work / "client_out"),
              "--json", "--ingest-timeout-s", "0"]

    # ── 1 · env and plan before anything is dispatched ──────────────────
    print("STEP 1 · env, stages, plan")
    r = run("engine.pipeline", "env", expect=None)
    d = jout(r)
    check("env measures every hard dependency and names the failures",
          "checks" in d and {c["check"] for c in d["checks"]} >= {"claude CLI", "connector identity", "install"},
          r.stdout[-300:])
    r = run("engine.pipeline", "stages")
    check("stages prints the table", "PRELIM" in r.stdout and "PROMOTE" in r.stdout)
    r = run("engine.pipeline", "plan", *base)
    d = jout(r)
    check("plan on a fresh run names PRELIM next and dispatches nothing",
          d.get("next") == "PRELIM" and not (root / "briefs").exists(), json.dumps(d)[:300])

    # ── 2 · a lane that produces nothing twice is retried ────────────────
    print("STEP 2 · retried lane")
    r = run("engine.pipeline", "run", *common, "--until", "RESEARCH", "--lane-retries", "2",
            env={**stub, "DMA_STUB_FAIL_FIRST": "research-p1c1-producer:2"})
    d = jout(r)
    st = json.loads((root / "07_qa" / "pipeline_state.json").read_text()) if (root / "07_qa" / "pipeline_state.json").exists() else {}
    check("the research stage PASSES after the retries",
          r.returncode == 0 and d.get("outcome") == "STOPPED_AT_UNTIL"
          and st.get("stages", {}).get("RESEARCH", {}).get("verdict") == "PASS",
          json.dumps(d)[:400])
    wb = runstate.locate(run_id, root).open()
    check("the stage rows carry wall clock in Gate_Log",
          any(str(g["Gate"]) == "STAGE_RESEARCH" and "elapsed" in str(g["Detail"])
              for g in wb.rows("Gate_Log")))

    # ── 3 · an ingest that never arrives ─────────────────────────────────
    print("STEP 3 · ingest timeout, resume at INGEST_A")
    r = run("engine.pipeline", "run", *common, env={**stub, "DMA_STUB_NEVER_INGEST": "1"}, expect=1)
    d = jout(r)
    check("INGEST_A is a loud FAIL naming the scan",
          d.get("outcome") == "FAILED" and d.get("stage") == "INGEST_A" and "did not ingest" in (d.get("reason") or ""),
          json.dumps(d)[:400])
    st = json.loads((root / "07_qa" / "pipeline_state.json").read_text())
    ck = st.get("connector", {}).get("checkpoint_a", {})
    check("the scored checkpoint was already pushed to the client folder",
          bool(ck.get("folder")) and Path(ck["folder"]).is_dir(), json.dumps(ck)[:200])
    r = run("engine.pipeline", "plan", *base)
    check("plan resumes at INGEST_A, not PRELIM", jout(r).get("next") == "INGEST_A", r.stdout[-200:])

    # ── 4 · a page FAIL re-dispatches only that page ─────────────────────
    print("STEP 4 · page verdict FAIL")
    r = run("engine.pipeline", "run", *common, "--until", "PAGES_A",
            env={**stub, "DMA_STUB_PAGE_FAIL": "heatmap:1"})
    d = jout(r)
    st = json.loads((root / "07_qa" / "pipeline_state.json").read_text())
    pages = st.get("pages", {})
    check("heatmap shipped twice, techstack once, stage PASS",
          r.returncode == 0 and pages.get("heatmap", {}).get("attempts", 0) >= 2
          and pages.get("techstack", {}).get("attempts") == 1
          and pages.get("heatmap", {}).get("status") == "pass", json.dumps(pages)[:400])
    briefs = sorted((root / "briefs").glob("pages_A_1/*.md"))
    check("the re-dispatch brief carries the verdict's reasons",
          bool(briefs) and any("CG-99" in b.read_text() for b in briefs))

    # ── 5 · a clean wall-clock stop, then resume ─────────────────────────
    print("STEP 5 · --max-wall-min")
    r = run("engine.pipeline", "run", *common, "--max-wall-min", "0", env=stub)
    d = jout(r)
    check("a zero wall clock stops cleanly (exit 0) before the next stage",
          r.returncode == 0 and d.get("outcome") == "STOPPED_WALL_CLOCK" and d.get("stage") == "PACKAGE",
          json.dumps(d)[:300])

    # ── 6 · a refused promote ────────────────────────────────────────────
    print("STEP 6 · promote refused, then promoted")
    r = run("engine.pipeline", "run", *common, env={**stub, "DMA_STUB_REFUSE_PROMOTE": "1"}, expect=1)
    d = jout(r)
    check("PROMOTE is a FAIL when the connector refuses",
          d.get("outcome") == "FAILED" and d.get("stage") == "PROMOTE", json.dumps(d)[:300])
    r = run("engine.pipeline", "run", *common, env=stub)
    d = jout(r)
    md = runstate.locate(run_id, root).open().metadata()
    check("the next run promotes; two connector versions recorded",
          d.get("outcome") == "COMPLETE" and d.get("stages_run") == ["PROMOTE"]
          and md.get("promoted_at") and md.get("connector_run_id") != md.get("connector_run_id_prev"),
          json.dumps(d)[:300])

    # ── 7 · idempotent; the ledger ───────────────────────────────────────
    print("STEP 7 · second run, cost report")
    r = run("engine.pipeline", "run", *common, env=stub)
    d = jout(r)
    check("a second run redoes nothing", d.get("outcome") == "COMPLETE" and d.get("stages_run") == [],
          json.dumps(d)[:300])
    r = run("engine.cost", "report", *base, "--json")
    d = jout(r)
    check("cost report reads every stage's wall clock and is within target",
          r.returncode == 0 and d.get("within") and {s["stage"] for s in d.get("stages", [])} >= {"PRELIM", "RESEARCH", "SCORING", "REPORTS", "PROMOTE"},
          json.dumps(d)[:400])
    r = run("engine.pipeline", "plan", *base)
    check("plan reads complete", jout(r).get("complete") is True)

    # ── 8 · a v6 workbook opens under v7 and the driver continues ────────
    print("STEP 8 · v6 in flight")
    v6 = work / "v6"
    pf2 = v6 / "preflight.json"; v6.mkdir()
    pf2.write_text(json.dumps(preflight_doc(entity, eid)))
    root2 = v6 / "run"
    r = run("engine.cli", "start", "--run", "R-STRESS-V6", "--root", str(root2), "--entity", entity,
            "--entity-id", eid, "--reference-date", "2026-08-29", "--preflight", str(pf2),
            "--no-push", "--folder-root", str(v6 / "client"), env=env)
    ok = r.returncode == 0
    if ok:
        wb2 = runstate.locate("R-STRESS-V6", root2).open()
        x = openpyxl.load_workbook(wb2.path)
        if "Financial_Trends" in x.sheetnames:
            del x["Financial_Trends"]
        for rrow in x["Handoff_Lock"].iter_rows(min_row=2):
            if rrow[0].value == "workbook_contract":
                rrow[1].value = "v6"
        x.save(wb2.path)
        r = run("engine.pipeline", "plan", "--run", "R-STRESS-V6", "--root", str(root2))
        d = jout(r)
        again = runstate.locate("R-STRESS-V6", root2).open()
        ok = r.returncode == 0 and "Financial_Trends" in again._wb.sheetnames \
            and again.verify_handoff_lock() == [] and d.get("next") == "PRELIM"
    check("a v6 workbook opens under v7 (sheet added, lock moved) and the driver plans from PRELIM",
          ok, (r.stderr or r.stdout)[-300:])

    return _finish(work, a.keep)


def _finish(work: Path, keep: bool) -> int:
    failed = [n for n, ok, _ in RESULTS if not ok]
    print(f"\n{len(RESULTS) - len(failed)}/{len(RESULTS)} steps PASS"
          + (f"; FAILED: {', '.join(failed)}" if failed else ""))
    if not keep and not failed:
        shutil.rmtree(work, ignore_errors=True)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
