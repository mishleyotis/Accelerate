#!/usr/bin/env python3
"""Deep search across ONE client's pulled package corpus.

Owner instruction, 2026-08-20: "if it is not in a workbook, where else can
it be found? How can it do deep client corpus searches?" — this is the
mechanical answer. It extracts text from every synthesis-input file in the
pulled package (workbooks tab by tab, CSVs, reports via their DOCX XML,
manifests, markdown) into a local index, then answers ranked queries with
file + line + snippet. Producers search HERE FIRST — the package usually
already holds the URL, date or excerpt a gap needs — and reach for web
enrichment only when the corpus comes back empty (that refusal is what a
`search_requests` entry is for).

  index  --package DIR            build/refresh DIR/.corpus_index/
  search --package DIR --query "core banking nCino" [--limit 8] [--json]
  search --package DIR --eid E-017            # exact evidence-id mode

Slides are excluded by policy (the corpus rule); PDFs are indexed only when
a text extractor is importable — canonical packages carry a DOCX twin of
every report, which indexes fully. Skipped files are named in the manifest,
never silently dropped.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from package_map import SLIDES_RE  # noqa: E402

INDEX_DIR = ".corpus_index"
TAG_RE = re.compile(r"<[^>]+>")
WS_RE = re.compile(r"[ \t]+")
TEXT_EXT = {".csv", ".md", ".txt", ".json", ".jsonl", ".log"}


def _docx_text(path: Path) -> str:
    out = []
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("word/") and name.endswith(".xml") \
                    and ("document" in name or "header" in name
                         or "footer" in name):
                xml = z.read(name).decode("utf-8", errors="replace")
                xml = xml.replace("</w:p>", "\n")
                out.append(WS_RE.sub(" ", TAG_RE.sub(" ", xml)))
    return "\n".join(out)


def _xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    try:
        for ws in wb.worksheets:
            out.append(f"## tab: {ws.title}")
            for row in ws.iter_rows(max_row=3000, values_only=True):
                cells = [str(c) for c in row if c not in (None, "")]
                if cells:
                    out.append(" | ".join(cells))
    finally:
        wb.close()
    return "\n".join(out)


def _pdf_text(path: Path):
    """(text, None) on success, or (None, why) — and `why` is the truth.

    This returned a bare None for BOTH "pypdf is not installed" and "this
    PDF would not parse", and the caller then recorded the reason as "no
    extractor for this type" for either. That sentence is false in both
    cases and consequential in the first: the reports never enter the index,
    `search` returns nothing, the CLI says "the corpus cannot answer this;
    a search_requests entry (web) is the next rung", and a producer goes to
    the web for a document sitting in the package. pypdf was also undeclared
    in dma-deps until 2026-08-23, so nothing checked for it either.
    """
    try:
        import pypdf                                   # noqa: PLC0415
    except ImportError:
        return None, ("extractor unavailable: pypdf is not installed — run "
                      "scripts/dma-deps install. THIS FILE WAS NOT READ; it "
                      "is not a file without an extractor")
    try:
        return "\n".join((pg.extract_text() or "")
                         for pg in pypdf.PdfReader(str(path)).pages), None
    except Exception as exc:                           # noqa: BLE001
        return None, (f"unreadable PDF: {type(exc).__name__} — encrypted, "
                      f"scanned or corrupt. THIS FILE WAS NOT READ")


def _extract(path: Path) -> tuple:
    low = path.name.lower()
    ext = path.suffix.lower()
    if SLIDES_RE.search(str(path).lower()):
        return None, "slides excluded by the corpus rule"   # policy
    if ext in TEXT_EXT:
        raw = path.read_bytes()
        if ext == ".csv":
            try:
                rows = list(csv.reader(io.StringIO(
                    raw.decode("utf-8-sig", errors="replace"))))
                return "\n".join(" | ".join(r) for r in rows), None
            except Exception:                          # noqa: BLE001
                pass
        return raw.decode("utf-8", errors="replace"), None
    if ext in (".xlsx", ".xlsm"):
        try:
            return _xlsx_text(path), None
        except Exception as e:                         # noqa: BLE001
            # NOT indexed as an error string. A workbook that would not open
            # used to be written into the index as "[unreadable workbook: …]",
            # counted as indexed, and answer nothing — "I read it and it says
            # nothing" reported for "I could not read it".
            return None, f"unreadable workbook: {type(e).__name__}"
    if ext == ".docx":
        try:
            return _docx_text(path), None
        except Exception as e:                         # noqa: BLE001
            return None, f"unreadable docx: {type(e).__name__}"
    if ext == ".pdf":
        return _pdf_text(path)
    if low == ".ds_store" or ext in (".png", ".jpg", ".jpeg", ".gif",
                                     ".zip", ".pptx", ".ppt"):
        return None, "binary type, nothing to index"
    if ext in (".doc", ".xls", ".rtf", ".eml", ".msg", ".numbers", ".key"):
        # These used to fall through to the byte decode below, land in the
        # index as replacement-character mush, count as INDEXED with a
        # character count, and answer no query. A legacy .xls evidence
        # register reported as read.
        return None, (f"legacy binary format ({ext}) — no extractor. Convert "
                      f"it or read it by hand; it is NOT empty")
    try:
        return path.read_bytes().decode("utf-8", errors="replace"), None
    except OSError as exc:
        return None, f"unreadable: {type(exc).__name__}"


def _unread_files(package: Path) -> list:
    """Files the index skipped, from the manifest it already writes."""
    try:
        m = json.loads((package / INDEX_DIR / "manifest.json").read_text())
    except (OSError, ValueError):
        return []
    return m.get("skipped") or []


def build_index(package: Path) -> dict:
    idx = package / INDEX_DIR
    idx.mkdir(exist_ok=True)
    manifest = {"indexed": [], "skipped": []}
    for p in sorted(package.rglob("*")):
        if not p.is_file() or INDEX_DIR in p.parts:
            continue
        rel = str(p.relative_to(package))
        text, why = _extract(p)
        if text is None:
            manifest["skipped"].append({"path": rel, "reason": why or
                                        "no extractor for this type"})
            continue
        out = idx / (rel.replace("/", "__") + ".txt")
        out.write_text(text, encoding="utf-8")
        manifest["indexed"].append({"path": rel, "chars": len(text)})
    (idx / "manifest.json").write_text(json.dumps(manifest, indent=1))
    return manifest


def search(package: Path, query: str, limit: int = 8,
           exact: bool = False) -> list:
    idx = package / INDEX_DIR
    if not (idx / "manifest.json").is_file():
        build_index(package)
    if exact:
        pattern = re.compile(re.escape(query), re.I)
        terms = [query.lower()]
    else:
        terms = [t for t in re.split(r"\W+", query.lower()) if len(t) > 1]
        pattern = re.compile("|".join(re.escape(t) for t in terms), re.I)
    hits = []
    for f in idx.glob("*.txt"):
        text = f.read_text(encoding="utf-8", errors="replace")
        low = text.lower()
        score = sum(low.count(t) for t in terms)
        if not score:
            continue
        lines = []
        for i, line in enumerate(text.splitlines(), 1):
            if pattern.search(line):
                snippet = line.strip()
                if len(snippet) > 240:
                    m = pattern.search(snippet)
                    s = max(0, (m.start() if m else 0) - 100)
                    snippet = "…" + snippet[s:s + 220] + "…"
                lines.append({"line": i, "snippet": snippet})
                if len(lines) >= 3:
                    break
        hits.append({"file": f.name[:-4].replace("__", "/"),
                     "score": score, "matches": lines})
    hits.sort(key=lambda h: -h["score"])
    return hits[:limit]


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = ap.add_subparsers(dest="cmd", required=True)
    p_i = sub.add_parser("index")
    p_i.add_argument("--package", required=True)
    p_s = sub.add_parser("search")
    p_s.add_argument("--package", required=True)
    p_s.add_argument("--query")
    p_s.add_argument("--eid")
    p_s.add_argument("--limit", type=int, default=8)
    p_s.add_argument("--json", action="store_true")
    a = ap.parse_args(argv)
    package = Path(a.package)
    if not package.is_dir():
        print(f"not a directory: {package}", file=sys.stderr)
        return 2
    if a.cmd == "index":
        m = build_index(package)
        print(f"indexed {len(m['indexed'])} files, "
              f"skipped {len(m['skipped'])} "
              f"({sum(1 for s in m['skipped'] if 'slides' in s['reason'])} "
              f"by the slides rule)")
        return 0
    q = a.eid or a.query
    if not q:
        ap.error("search needs --query or --eid")
    hits = search(package, q, a.limit, exact=bool(a.eid))
    if a.json:
        print(json.dumps(hits, indent=1))
    else:
        if not hits:
            # "THE CORPUS CANNOT ANSWER" IS A CLAIM ABOUT THE CORPUS, and it
            # is only true if the corpus was read. A file skipped for a
            # missing extractor is a file nobody looked in, and concluding
            # otherwise sends a producer to the web for a document sitting
            # in the package — the mechanism behind an invented excerpt.
            unread = [x for x in _unread_files(package)
                      if "NOT READ" in x.get("reason", "")]
            if unread:
                print(f"no corpus hits for {q!r} — BUT {len(unread)} file(s) "
                      f"were never read, so this is not evidence the corpus "
                      f"lacks the answer:")
                for x in unread[:5]:
                    print(f"    {x['path']}: {x['reason']}")
                print("  fix the extractor and re-index before treating this "
                      "as a gap; the web is NOT the next rung yet")
            else:
                print(f"no corpus hits for {q!r} — the corpus cannot answer "
                      f"this; a search_requests entry (web) is the next rung")
        for h in hits:
            print(f"{h['score']:5d}  {h['file']}")
            for m in h["matches"]:
                print(f"        L{m['line']}: {m['snippet'][:200]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
