#!/usr/bin/env python3
"""Survey the intake corpus for what packages ACTUALLY look like.

Owner instruction, 2026-08-20: the workflow assumed uniform packages and
evidence ids that live only in workbooks — an ideal. This script measures
the mess instead: it walks the intake tree's client folders and reports the
structure generations, where the workbooks really are, where evidence ids
and URLs really live, and which folders a uniform parser would break on.
Safeguards are built from this survey's output, not from the spec.

  survey            top-level structure of every client folder (metadata only)
  deep --client X   full recursive inventory + workbook/CSV content profile
                    (downloads only .xlsx/.csv, never decks or PDFs)

Output is JSON on stdout; findings prose on stderr. Nothing is written to
Drive. No token value is ever printed.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import drive_fetch  # noqa: E402

EID_RE = re.compile(r"^E[-_][A-Z0-9]+[-_]?\d*(:F\d+)?$", re.I)
URL_RE = re.compile(r"^https?://", re.I)
CANONICAL_FOLDERS = {"01_evidence", "02_research_workbook",
                     "03_scoring_workbook", "04_reports", "05_narrative_deck",
                     "06_peers", "07_governance", "08_appendices"}


def survey(limit: int | None = None) -> dict:
    tok = drive_fetch._token()
    clients = [f for f in drive_fetch._list_children(
        tok, drive_fetch.INTAKE_FOLDER_ID)
        if f["mimeType"] == drive_fetch.FOLDER_MIME]
    out = []
    for f in clients[:limit]:
        kids = drive_fetch._list_children(tok, f["id"])
        folders = sorted(k["name"] for k in kids
                         if k["mimeType"] == drive_fetch.FOLDER_MIME)
        files = sorted(k["name"] for k in kids
                       if k["mimeType"] != drive_fetch.FOLDER_MIME)
        canonical = sum(1 for n in folders if n in CANONICAL_FOLDERS)
        shape = ("canonical" if canonical >= 6 else
                 "partial" if canonical >= 2 else
                 "flat" if files and not folders else "other")
        out.append({"client": f["name"], "shape": shape,
                    "folders": folders, "top_files": files,
                    "canonical_folders": canonical})
    shapes = Counter(c["shape"] for c in out)
    return {"clients": len(out), "shapes": dict(shapes), "rows": out}


def _grab(tok, file_id: str) -> bytes:
    url = (f"{drive_fetch.API}/files/{file_id}"
           f"?alt=media&supportsAllDrives=true")
    with drive_fetch._req(tok, url) as resp:
        return resp.read()


def _walk(tok, folder_id: str, prefix: str = "", depth: int = 0):
    if depth > 6:
        return
    for f in drive_fetch._list_children(tok, folder_id):
        path = f"{prefix}/{f['name']}" if prefix else f["name"]
        if f["mimeType"] == drive_fetch.FOLDER_MIME:
            yield from _walk(tok, f["id"], path, depth + 1)
        else:
            yield path, f


def _profile_csv(raw: bytes) -> dict:
    try:
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    except Exception as e:                                  # noqa: BLE001
        return {"error": f"unreadable csv: {e}"}
    if not rows:
        return {"rows": 0}
    hdr = [h.strip().lower() for h in rows[0]]
    body = rows[1:]
    ids = urls = dated = excerpts = 0
    for r in body:
        for cell in r:
            c = (cell or "").strip()
            if EID_RE.match(c):
                ids += 1
            elif URL_RE.match(c):
                urls += 1
    date_cols = [i for i, h in enumerate(hdr) if "date" in h]
    ex_cols = [i for i, h in enumerate(hdr)
               if "excerpt" in h or "quote" in h]
    for r in body:
        if any(i < len(r) and (r[i] or "").strip() for i in date_cols):
            dated += 1
        if any(i < len(r) and len((r[i] or "").strip()) >= 50
               for i in ex_cols):
            excerpts += 1
    return {"rows": len(body), "header": hdr[:12], "eid_cells": ids,
            "url_cells": urls, "rows_with_date": dated,
            "rows_with_50char_excerpt": excerpts}


def deep(client: str) -> dict:
    tok = drive_fetch._token()
    folder = drive_fetch._find_client_folder(tok, client)
    inv, csvs, xlsx = [], {}, []
    for path, f in _walk(tok, folder["id"]):
        size = int(f.get("size") or 0)
        inv.append({"path": path, "mime": f["mimeType"], "size": size})
        low = path.lower()
        if low.endswith(".csv") and size < 5_000_000:
            csvs[path] = _profile_csv(_grab(tok, f["id"]))
        elif low.endswith((".xlsx", ".xlsm")):
            xlsx.append(path)
    return {"client": folder["name"], "files": len(inv),
            "inventory": inv, "workbooks": xlsx, "csv_profiles": csvs}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = ap.add_subparsers(dest="cmd", required=True)
    p_s = sub.add_parser("survey")
    p_s.add_argument("--limit", type=int, default=None)
    p_d = sub.add_parser("deep")
    p_d.add_argument("--client", required=True)
    p_c = sub.add_parser("corpus")
    p_c.add_argument("--out", required=True,
                     help="JSONL sink, one client per line, flushed as it goes")
    p_c.add_argument("--limit", type=int, default=None)
    p_t = sub.add_parser("trends")
    p_t.add_argument("--from", dest="src", required=True)
    p_w = sub.add_parser("workbooks")
    p_w.add_argument("--out", required=True,
                     help="JSONL sink, APPENDED to so a run can be resumed")
    p_w.add_argument("--limit", type=int, default=None)
    p_w.add_argument("--skip", type=int, default=0)
    a = ap.parse_args(argv)
    if a.cmd == "workbooks":
        print(json.dumps(workbooks(a.out, a.limit, a.skip)))
        return 0
    if a.cmd == "survey":
        print(json.dumps(survey(a.limit), indent=1))
        return 0
    if a.cmd == "deep":
        print(json.dumps(deep(a.client), indent=1))
        return 0
    if a.cmd == "corpus":
        print(json.dumps(corpus(a.out, a.limit)))
        return 0
    if a.cmd == "trends":
        print(json.dumps(trends(a.src), indent=1))
        return 0
    return 2


# ── full-corpus survey: every client, recursive, with source-location map ──

KIND_RULES = (
    ("scoring_workbook", re.compile(r"scor|assessment.*workbook", re.I),
     (".xlsx", ".xlsm")),
    ("research_workbook", re.compile(r"research", re.I), (".xlsx", ".xlsm")),
    ("evidence_table", re.compile(r"evidence|inventory", re.I), (".csv",)),
    ("governance_log", re.compile(r"caps_applied|contradiction|issue", re.I),
     (".csv",)),
    ("peer_table", re.compile(r"peer", re.I), (".csv",)),
    ("deck", re.compile(r"presentation|deck|slides", re.I),
     (".pptx", ".ppt", "")),
    ("report", re.compile(r"report|assessment|profile", re.I),
     (".docx", ".pdf", "")),
    ("manifest", re.compile(r"manifest", re.I), (".json",)),
    ("memory", re.compile(r"synthesis memory", re.I), (".md",)),
)


def _kind(path: str) -> str:
    name = path.rsplit("/", 1)[-1]
    low = name.lower()
    ext = ("." + low.rsplit(".", 1)[-1]) if "." in low else ""
    for kind, rx, exts in KIND_RULES:
        if rx.search(low) and (ext in exts or ("" in exts and ext == "")):
            return kind
    if ext in (".xlsx", ".xlsm"):
        return "other_xlsx"
    if ext == ".csv":
        return "other_csv"
    if ext == ".pptx":
        return "deck"
    if ext in (".docx", ".pdf", ".md", ".txt"):
        return "document"
    return "other"


def corpus(out_path: str, limit: int | None = None) -> dict:
    tok = drive_fetch._token()
    clients = [f for f in drive_fetch._list_children(
        tok, drive_fetch.INTAKE_FOLDER_ID)
        if f["mimeType"] == drive_fetch.FOLDER_MIME]
    rows = []
    with open(out_path, "w") as sink:
        for i, cf in enumerate(clients[:limit]):
            try:
                files = list(_walk(tok, cf["id"]))
            except Exception as e:                          # noqa: BLE001
                rows.append({"client": cf["name"], "error": str(e)})
                sink.write(json.dumps(rows[-1]) + "\n")
                sink.flush()
                continue
            top = drive_fetch._list_children(tok, cf["id"])
            top_folders = [k["name"] for k in top
                           if k["mimeType"] == drive_fetch.FOLDER_MIME]
            top_files = [k for k in top
                         if k["mimeType"] != drive_fetch.FOLDER_MIME]
            wrapper = (len(top_folders) == 1 and not
                       [t for t in top_files if t["name"] != ".DS_Store"]
                       and top_folders[0] not in CANONICAL_FOLDERS)
            inv = []
            ev_profiles = {}
            for path, f in files:
                kind = _kind(path)
                inv.append({"path": path, "kind": kind,
                            "size": int(f.get("size") or 0)})
                if (kind in ("evidence_table", "governance_log") and
                        int(f.get("size") or 0) < 2_000_000):
                    try:
                        ev_profiles[path] = _profile_csv(_grab(tok, f["id"]))
                    except Exception as e:                  # noqa: BLE001
                        ev_profiles[path] = {"error": str(e)}
            row = {"client": cf["name"], "files": len(inv),
                   "wrapper": wrapper, "top_folders": top_folders,
                   "inventory": inv, "evidence_profiles": ev_profiles}
            rows.append(row)
            sink.write(json.dumps(row) + "\n")
            sink.flush()
            print(f"[{i+1}/{len(clients)}] {cf['name']}: {len(inv)} files",
                  file=sys.stderr)
    return {"clients": len(rows)}



def trends(jsonl_path: str) -> dict:
    """Aggregate a corpus run into the trends that drive safeguards."""
    rows = [json.loads(ln) for ln in open(jsonl_path)
            if ln.strip()]
    ok = [r for r in rows if "error" not in r]
    kinds = Counter()
    ev_locations = Counter()
    header_vocab = Counter()
    scoring_dirs = Counter()
    url_rows = date_rows = excerpt_rows = table_rows = 0
    per_client = []
    for r in ok:
        inv = r["inventory"]
        k = Counter(i["kind"] for i in inv)
        kinds.update(k)
        scoring = [i["path"] for i in inv if i["kind"] == "scoring_workbook"]
        for s in scoring:
            scoring_dirs[s.rsplit("/", 1)[0] if "/" in s else "(top)"] += 1
        for path, prof in (r.get("evidence_profiles") or {}).items():
            if "error" in prof or not prof.get("rows"):
                continue
            ev_locations[path.rsplit("/", 1)[0] if "/" in path
                         else "(top)"] += 1
            header_vocab.update(prof.get("header") or [])
            table_rows += prof["rows"]
            url_rows += min(prof.get("url_cells", 0), prof["rows"])
            date_rows += prof.get("rows_with_date", 0)
            excerpt_rows += prof.get("rows_with_50char_excerpt", 0)
        per_client.append({
            "client": r["client"], "files": r["files"],
            "wrapper": r.get("wrapper", False),
            "scoring_workbooks": len(scoring),
            "score_exports": any(
                i["path"].lower().endswith("export_scoring_detail.csv")
                for i in inv),
            "has_evidence_table": any(
                i["kind"] == "evidence_table" for i in inv),
            "interim": any("interim" in i["path"].lower()
                           or "draft" in i["path"].lower() for i in inv),
        })
    n = len(ok)
    return {
        "clients_surveyed": n,
        "errors": [r["client"] for r in rows if "error" in r],
        "wrapper_clients": sum(1 for c in per_client if c["wrapper"]),
        "export_only_scoring": sorted(
            c["client"] for c in per_client
            if not c["scoring_workbooks"] and c["score_exports"]),
        "no_scoring_artifacts": sorted(
            c["client"] for c in per_client
            if not c["scoring_workbooks"] and not c["score_exports"]),
        "multi_scoring_workbook": sorted(
            c["client"] for c in per_client if c["scoring_workbooks"] > 1),
        "interim_or_draft_present": sum(1 for c in per_client if c["interim"]),
        "no_evidence_table": sorted(c["client"] for c in per_client
                                    if not c["has_evidence_table"])[:20],
        "file_kinds": dict(kinds.most_common()),
        "scoring_workbook_locations": dict(scoring_dirs.most_common(12)),
        "evidence_table_locations": dict(ev_locations.most_common(15)),
        "evidence_header_vocab": dict(header_vocab.most_common(30)),
        "evidence_rows_profiled": table_rows,
        "evidence_rows_with_date": date_rows,
        "evidence_rows_with_50char_excerpt": excerpt_rows,
    }


# ── workbook-shape survey: what a register ACTUALLY looks like ──
#
# Added 2026-08-22 after Houlihan Lokey. The `corpus` mode above profiles
# CSVs and never opens a workbook, so every rule about workbook ROLE, tab
# naming and excerpt columns was written from two or three examples. That
# is how `02_research_workbook/DMA_Scoring_Workbook_HL.xlsx` came to be
# classified by its filename, and how `Fact_Summary` came to be in no
# synonym tuple. This mode opens every workbook in the corpus and records
# the three things those rules need: where the file sits, what its tabs are
# called, and which column actually carries verbatim text.

XLSX_MAX_BYTES = 60_000_000
EXCERPT_FLOOR = 50


def _header_row(rows):
    """The same rule the live parsers use: first row with 3+ non-empty
    strings. Measuring with a different rule would measure a fiction."""
    for i, r in enumerate(rows[:20]):
        strs = [c for c in r if isinstance(c, str) and c.strip()]
        if len(strs) >= 3:
            return i, [str(c).strip() if c is not None else "" for c in r]
    return None, []


def _profile_workbook(raw: bytes) -> dict:
    from openpyxl import load_workbook
    bio = io.BytesIO(raw)
    try:
        wb = load_workbook(bio, read_only=True, data_only=True)
    except Exception as e:                                  # noqa: BLE001
        return {"error": f"unreadable workbook: {e}"}
    tabs = []
    try:
        for ws in wb.worksheets:
            try:
                rows = [list(r) for r in
                        ws.iter_rows(max_row=1200, values_only=True)]
            except Exception as e:                          # noqa: BLE001
                tabs.append({"tab": ws.title, "error": str(e)})
                continue
            hi, hdr = _header_row(rows)
            if hi is None:
                tabs.append({"tab": ws.title, "rows": 0, "header": []})
                continue
            body = rows[hi + 1:]
            body = [r for r in body
                    if r and any(c is not None and str(c).strip() for c in r)]
            # which column holds evidence ids, and which holds long text
            eid_col = None
            long_cols = {}
            for j, h in enumerate(hdr):
                if not h:
                    continue
                vals = [r[j] for r in body if j < len(r)]
                strs = [str(v).strip() for v in vals
                        if isinstance(v, str) and str(v).strip()]
                if eid_col is None and strs and \
                        sum(1 for s in strs if EID_RE.match(s)) >= max(
                            3, len(strs) // 2):
                    eid_col = h.lower()
                if strs:
                    over = sum(1 for s in strs if len(s) >= EXCERPT_FLOOR)
                    if over:
                        long_cols[h.lower()] = {
                            "n_over_floor": over,
                            "median_len": sorted(len(s) for s in strs)[
                                len(strs) // 2]}
            tabs.append({"tab": ws.title, "rows": len(body),
                         "header": [h.lower() for h in hdr if h][:24],
                         "eid_column": eid_col,
                         "long_text_columns": long_cols})
    finally:
        wb.close()
    return {"tabs": tabs}


def workbooks(out_path: str, limit: int | None = None,
              skip: int = 0) -> dict:
    """One JSONL line per client: every workbook, where it sits, its tabs."""
    tok = drive_fetch._token()
    clients = [f for f in drive_fetch._list_children(
        tok, drive_fetch.INTAKE_FOLDER_ID)
        if f["mimeType"] == drive_fetch.FOLDER_MIME]
    sel = clients[skip:None if limit is None else skip + limit]
    n = 0
    with open(out_path, "a") as sink:
        for i, cf in enumerate(sel):
            row = {"client": cf["name"], "workbooks": []}
            try:
                files = list(_walk(tok, cf["id"]))
            except Exception as e:                          # noqa: BLE001
                row["error"] = str(e)
                sink.write(json.dumps(row) + "\n")
                sink.flush()
                continue
            row["folders"] = sorted({p.rsplit("/", 1)[0]
                                     for p, _ in files if "/" in p})
            for path, f in files:
                if not path.lower().endswith((".xlsx", ".xlsm")):
                    continue
                size = int(f.get("size") or 0)
                entry = {"path": path, "size": size,
                         "dir": path.rsplit("/", 1)[0] if "/" in path
                         else "(top)"}
                if size > XLSX_MAX_BYTES:
                    entry["error"] = f"skipped, {size} bytes"
                else:
                    try:
                        entry.update(_profile_workbook(_grab(tok, f["id"])))
                    except Exception as e:                  # noqa: BLE001
                        entry["error"] = str(e)
                row["workbooks"].append(entry)
            sink.write(json.dumps(row) + "\n")
            sink.flush()
            n += 1
            print(f"[{skip + i + 1}] {cf['name']}: "
                  f"{len(row['workbooks'])} workbook(s)", file=sys.stderr)
    return {"clients": n}


if __name__ == "__main__":
    sys.exit(main())
