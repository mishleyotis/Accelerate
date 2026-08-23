#!/usr/bin/env python3
"""Merge every evidence store in a package into schema-fit records.

Owner instruction, 2026-08-20: the pipeline must "retrieve the evidence
URLs, or even enrich where required to fit the required schema formats".
The corpus survey measured why: evidence lives in up to ten stores per
package (workbook register tabs, CSVs under three different folders,
JSON/JSONL ledgers), no CSV in the corpus carries a 50-char excerpt, one
generation has no URL column at all, another has URLs but no dates.

This module makes the package answer for itself before anything reaches
the web:

  1. MERGE — every store package_map names, keyed by evidence id; fields
     unified through a header-synonym map; same id + conflicting values is
     a CONFLICT row (adjudicated, never averaged).
  2. CORPUS FILL — a record missing url/date/excerpt searches the client's
     OWN corpus first (corpus_search): a fill found there is package data
     with provenance, not enrichment.
  3. GAPS OUT — whatever the corpus cannot answer becomes a ready
     `search_requests` entry for the top session's connectors. A row is
     never registered bare, never dropped silently; undated stays
     UNVERIFIED until a date is actually found.

Usage:
  evidence_normalize.py --package DIR [--client SLUG] [--out normalized.jsonl]

Output: one JSON record per evidence id; gaps + conflicts to stdout.
Cross-client ledger form of every id: <client>:<eid>.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import corpus_search  # noqa: E402
import package_map  # noqa: E402

EID_RE = re.compile(r"^(E[-_][A-Z0-9]+[-_]?\d*)(:F\d+)?$", re.I)
URL_IN_TEXT = re.compile(r"https?://[^\s\"'|<>\)\]]+", re.I)
DATE_IN_TEXT = re.compile(r"\b(20[12]\d[-/]\d{1,2}[-/]\d{1,2}|"
                          r"\d{1,2}/\d{1,2}/20[12]\d)\b")

# A QUOTATION and a PARAPHRASE are not interchangeable, and the corpus keeps
# them in differently-named columns (measured 2026-08-22 across the intake
# tree): the research workbook's Evidence_Detail carries `Excerpt` and
# `Anchor_Quote` — real spans of the source — while the scoring workbook's
# Evidence_Master carries `Fact_Summary`, and Lawley's carries `key_finding`.
# Those are assessors' prose ABOUT the source. Invariant 4 wants a verbatim
# excerpt, so a summary column must be recognised in order to be REFUSED for
# that purpose, not left out of the vocabulary where it gets silently missed
# and then replaced by something worse.
#
# Which side each column falls on was MEASURED, not inferred from its name.
# The corpus keeps 899 facts carrying both a `text` and an `anchor_quote`,
# and ZERO of them are identical: `text` is the assessor writing "HL
# explicitly positions its advisory-only independence as its core strategic
# differentiator: …", `anchor_quote` is the span the filing actually
# contains. So the whole `fact*` family is paraphrase however long and
# quotable it looks — and it looks very quotable, which is the trap.
# Only a column that NAMES a quotation is treated as one.
VERBATIM_COLS = ("evidence_excerpt", "excerpt", "anchor_quote", "quote",
                 "verbatim", "passage", "snippet")
SUMMARY_COLS = ("fact_summary", "key_finding", "key_finding_summary",
                "key_extract", "key_facts", "key_fact", "fact_text",
                "lead_fact", "top_fact", "sample_fact", "fact_preview",
                "facts", "fact", "text", "claim", "summary", "finding",
                "description", "note")

SYN = {
    # PLURALS ARE THE SCORING-DETAIL SPELLING. Those tabs write
    # `Evidence_IDs`, `Evidence_URLs`, `Source_URLs` — one row citing
    # several — and they were unreachable for as long as the tab itself was
    # (22,501 rows across 19 corpus packages). Reading the tab and then not
    # recognising its columns recovers the row and loses its content.
    "eid": ("evidence_id", "evidence_ids", "e_id", "fact_id", "id"),
    "source": ("source_name", "source_names", "source_title", "source",
               "sources", "publisher", "kb_source_id", "title",
               "source_document", "source_documents"),
    # `url_or_citation` is the scoring workbook's column on 6 register tabs
    # in the corpus; without it those clients' URLs are invisible.
    "url": ("url", "urls", "link", "links", "source_url", "source_urls",
            "evidence_url", "evidence_urls", "url_or_citation",
            "citation_url", "source_link"),
    # ordered best-first: a publication-flavoured column outranks a bare
    # "date", which in fact-level rows carries EVENT dates (E-083's 1979
    # is a timeline fact, not when its source was published — measured)
    "date": ("date_published", "publish_date", "publication_date",
             "published_date", "as_of_date", "date"),
    "excerpt": VERBATIM_COLS,
    "summary": SUMMARY_COLS,
    "tier": ("tier",),
    "ers": ("ers_total", "ers_score", "ers", "ers_core"),
    "subcaps": ("subcaps_supported", "subcaps", "subcap_id", "subcap_ids",
                "subcaps_cited", "subcaps_mapped", "mapped_subcaps",
                "subcaps_covered", "subcap_mappings",
                "categories_referenced", "pillars_mapped"),
}


def _norm_key(k) -> str:
    """`Source Name` and `source_name` are one column; the corpus writes it
    both ways, and lowercasing alone left `source name` matching nothing.
    A trailing parenthetical is decoration — `Key Facts (F1..)` is the
    `key_facts` column with a note about its contents stuck to the header."""
    s = re.sub(r"\(.*?\)", " ", str(k or "").lower())
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


FIELDS =("source", "url", "date", "excerpt", "summary", "tier", "ers",
          "subcaps")

# Stores nest their real rows one level down. Houlihan Lokey's ledger.jsonl
# keeps one record per evidence id whose verbatim spans live in
# `facts: [{fact_id, text}, …]` — flattening only the top level found no
# `text` key, reported the excerpt as absent, and sent the row to be filled
# from somewhere else. The fact ids (`E-001:F1`) fold back onto their base id.
NESTED_ROW_KEYS = ("facts", "items", "rows", "records", "entries",
                   "evidence", "excerpts")

# A serialized record is not a quotation, whatever column it arrives in.
# This is the shape of the Houlihan Lokey fabrication — 306 of 462 excerpts
# opened `{"evidence_id": "E-001", "source_name": …` — and it is cheap to
# refuse at the door. Keyed on JSON/dict-literal punctuation only: an
# earlier version also treated pipes as structure and refused E-482, whose
# excerpt is a real job posting reading "Platform Engineer | Solutions
# Engineer | GCP, AWS, Snowflake". Prose does use pipes; prose does not use
# `"key": "value"`.
SERIALIZED_RE = re.compile(r"\"\s*:\s*[\"\[{]|'\s*:\s*['\[{]|^\s*[\[{]")


def _pick(row: dict, field: str):
    """(value, rank, column) — the column is kept so provenance can name it.

    A reader downstream has to be able to tell a value the package stated
    from a value this pipeline reasoned its way to, and "which column" is
    the whole of that answer for a register row.
    """
    for rank, k in enumerate(SYN[field]):
        v = row.get(k)
        if v not in (None, ""):
            return str(v).strip(), rank, k
    return None, None, None


def _rows_from_csv(path: Path):
    try:
        text = path.read_bytes().decode("utf-8-sig", errors="replace")
        rd = csv.DictReader(io.StringIO(text))
        for r in rd:
            yield {_norm_key(k): (v or "") for k, v in r.items()}
    except Exception as e:                                  # noqa: BLE001
        print(f"unreadable csv {path.name}: {e}", file=sys.stderr)


def _flat(d: dict) -> dict:
    """Scalar cells only — a nested list is not a value, it is more rows."""
    return {_norm_key(k): ("" if v is None else v) for k, v in d.items()
            if not isinstance(v, (list, dict))}


def _expand(item: dict):
    """The record itself, then each nested sub-row carrying its parent's id.

    A sub-row wins on its own fields and inherits the rest, so a fact's
    verbatim `text` reaches the excerpt column while the parent's url,
    source and publish_date still travel with it.
    """
    parent = _flat(item)
    yield parent
    for key in NESTED_ROW_KEYS:
        kids = item.get(key)
        if not isinstance(kids, list):
            continue
        for kid in kids:
            if isinstance(kid, dict):
                yield {**parent, **_flat(kid)}


def _rows_from_json(path: Path):
    try:
        raw = path.read_bytes().decode("utf-8", errors="replace")
        if path.suffix == ".jsonl":
            items = [json.loads(ln) for ln in raw.splitlines() if ln.strip()]
        else:
            data = json.loads(raw)
            items = (data if isinstance(data, list) else
                     next((v for v in data.values()
                           if isinstance(v, list) and v
                           and isinstance(v[0], dict)), []))
        for it in items:
            if isinstance(it, dict):
                yield from _expand(it)
    except Exception as e:                                  # noqa: BLE001
        print(f"unreadable json {path.name}: {e}", file=sys.stderr)


def _rows_from_xlsx(path: Path):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:                                  # noqa: BLE001
        print(f"unreadable workbook {path.name}: {e}", file=sys.stderr)
        return
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(max_row=3000, values_only=True)
            hdr = None
            for r in rows:
                cells = ["" if c is None else str(c).strip() for c in r]
                if hdr is None:
                    if sum(1 for c in cells if c) >= 3:
                        hdr = [_norm_key(c) for c in cells]
                        # A TAB IS EVIDENCE-SHAPED, NOT EVIDENCE-NAMED. This
                        # read `if "evidence" not in ws.title.lower(): continue`
                        # and skipped 134 tabs across 19 corpus packages —
                        # 22,501 rows, every one of them carrying a URL column
                        # and 9,552 of them an excerpt column. They are the
                        # P1..P4 _Scoring_Detail and _Subcap_Scoring tabs,
                        # whose headers are Evidence_IDs, Evidence_URLs,
                        # Evidence_Excerpt. One client merged 120 records with
                        # ZERO excerpts while Evidence_Excerpt sat in four
                        # unread tabs. package_map already accepts three words
                        # for a CSV; this accepted one for a worksheet.
                        if not _evidence_shaped(hdr):
                            hdr = None
                            break
                    continue
                if any(cells):
                    yield dict(zip(hdr, cells))
    finally:
        wb.close()


#: What an evidence id looks like when a package does not spell it `E-`.
#: Measured across the corpus: `EV-P1C1-001`, `EV-CONN-001` (1,013 rows in
#: one client, across five stores that all agree), `INT-BRIEF-*`, `US-*`,
#: `PX-P1C1.1.1-1`. `EID_RE` rejected every one, `merge` returned 0 records,
#: and the tool printed `"records": 0` — which reads as an empty package
#: rather than as an id vocabulary nothing recognised.
EID_WIDE_RE = re.compile(
    r"^([A-Z]{1,6}[-_][A-Z0-9][A-Z0-9._-]*\d)(:F\d+)?$", re.I)

#: Never an evidence id, whatever their shape: catalogue cell ids and the
#: variant suffixes that decorate them.
NOT_AN_EID_RE = re.compile(r"^P[1-4]C\d", re.I)


def _base_eid(value: str, wide: bool = False) -> str | None:
    """The canonical id, or None.

    `wide` is used ONLY when a column is already known to be the evidence-id
    column: the value in it IS the id, so the pattern is a guard against
    junk rather than a vocabulary. Value-scanning a whole row still uses the
    strict pattern, or every product code in the row becomes evidence.
    """
    v = value.strip()
    if NOT_AN_EID_RE.match(v):
        return None
    m = EID_RE.match(v) or (EID_WIDE_RE.match(v) if wide else None)
    return m.group(1).upper() if m else None


def _evidence_shaped(hdr: list) -> bool:
    """Does this header row carry evidence: an id column beside a url,
    excerpt or source column? Shape, not name."""
    has_id = any(h in SYN["eid"] or "evidence_id" in h or h == "e_id"
                 for h in hdr)
    has_content = any(h in SYN["url"] or h in VERBATIM_COLS
                      or h in SYN.get("source", ()) or "url" in h
                      or "excerpt" in h for h in hdr)
    return has_id and has_content


def merge(package: Path) -> tuple[dict, list, dict]:
    pm = package_map.map_package(package)
    stores = list(pm["evidence_tables"])
    if pm["research"]["primary"]:
        stores.append(str(Path(pm["research"]["primary"]).relative_to(package)))
    if pm["scoring"]["primary"]:
        stores.append(str(Path(pm["scoring"]["primary"]).relative_to(package)))
    records: dict[str, dict] = {}
    conflicts: list[dict] = []
    unrecognised: dict[str, list] = {}
    for rel in stores:
        path = package / rel
        if path.suffix in (".csv",):
            rows = _rows_from_csv(path)
        elif path.suffix in (".json", ".jsonl"):
            rows = _rows_from_json(path)
        elif path.suffix in (".xlsx", ".xlsm"):
            rows = _rows_from_xlsx(path)
        else:
            continue
        for row in rows:
            raw_id, _, _ = _pick(row, "eid")
            named_col = bool(raw_id)
            if not raw_id:
                raw_id = next((v for v in row.values()
                               if isinstance(v, str) and _base_eid(v)), None)
            if not raw_id:
                continue
            eid = _base_eid(str(raw_id), wide=named_col)
            if not eid:
                # NOT SILENT. This `continue` dropped 1,029 URL-carrying rows
                # across four packages with no counter anywhere. A catalogue
                # CELL id is excluded from the count: it is a known non-id
                # shape, not an evidence row nobody recognised, and reporting
                # 712 of them buries the handful that matter.
                if not NOT_AN_EID_RE.match(str(raw_id).strip()):
                    unrecognised.setdefault(rel, []).append(str(raw_id)[:40])
                continue
            rec = records.setdefault(eid, {"eid": eid, "provenance": []})
            if rel not in rec["provenance"]:
                rec["provenance"].append(rel)
            for field in FIELDS:
                v, rank, col = _pick(row, field)
                if v is None:
                    continue
                if field == "excerpt" and (len(v) < 50
                                          or SERIALIZED_RE.search(v)):
                    continue        # below the floor, or not a quotation
                old = rec.get(field)
                if old is None:
                    rec[field] = v
                    rec.setdefault("_rank", {})[field] = rank
                    rec.setdefault("_store", {})[field] = rel
                    rec.setdefault("field_provenance", {})[field] = {
                        "how": "register", "store": rel, "column": col}
                    continue
                if field == "url" and old.rstrip("/") != v.rstrip("/"):
                    conflicts.append({"eid": eid, "field": "url",
                                      "values": [old, v], "store": rel})
                elif field == "date" and old != v:
                    old_rank = rec.get("_rank", {}).get("date", 99)
                    if rank < old_rank:
                        rec["date"] = v          # better column wins quietly
                        rec["_rank"]["date"] = rank
                        rec["_store"]["date"] = rel
                        rec.setdefault("field_provenance", {})["date"] = {
                            "how": "register", "store": rel, "column": col}
                    elif (rank == old_rank
                          and rec.get("_store", {}).get("date") != rel):
                        conflicts.append({"eid": eid, "field": "date",
                                          "values": [old, v], "store": rel})
    return records, conflicts, unrecognised


#: Fields a corpus SCAN may fill, and the only one on the list.
#:
#: This used to include `excerpt` and `date`, and the consequence was
#: measured on Houlihan Lokey (2026-08-22): 462 of 462 excerpts were
#: fabricated here. The rule was "any corpus line ≥50 chars that mentions
#: this evidence id is its excerpt", and the line that most reliably
#: mentions an evidence id is the LEDGER RECORD THAT DEFINES IT — so the
#: pipeline harvested its own bookkeeping and served it as a quotation from
#: a 10-K. E-002's "excerpt" was a spreadsheet row rendered with pipes.
#:
#: A quotation cannot be inferred from proximity: it is verbatim or it is
#: not an excerpt, and no amount of nearness makes a ledger line into a span
#: of the source. A publication date cannot be inferred from proximity
#: either — E-022 acquired 2026-07-27, the package's own build stamp,
#: scraped off a workbook row that merely mentioned it.
#:
#: A URL is different in kind, and that difference is the whole reason it
#: survives here: it is an opaque identifier that appears LITERALLY, so
#: finding one on the id's own line is recognition, not inference. It is
#: still marked `corpus_scan` so no reader mistakes it for a register value.
SCANNABLE_FIELDS = ("url",)


def corpus_fill(package: Path, records: dict) -> int:
    filled = 0
    for eid, rec in records.items():
        missing = [f for f in SCANNABLE_FIELDS if not rec.get(f)]
        if not missing:
            continue
        hits = corpus_search.search(package, eid, limit=4, exact=True)
        for h in hits:
            for m in h["matches"]:
                line = m["snippet"]
                if "url" in missing and not rec.get("url"):
                    u = URL_IN_TEXT.search(line)
                    if u:
                        rec["url"] = u.group(0)
                        rec.setdefault("fills", []).append(
                            {"field": "url", "from": h["file"],
                             "line": m["line"]})
                        rec.setdefault("field_provenance", {})["url"] = {
                            "how": "corpus_scan", "store": h["file"],
                            "line": m["line"]}
                        missing.remove("url")
                        filled += 1
            if not missing:
                break
    return filled


DATE_IN_NAME = re.compile(
    r"(20\d{2})[-_.]?(0[1-9]|1[0-2])(?:[-_.]?(0[1-9]|[12]\d|3[01]))?")
ISO_DATE = re.compile(r"20\d{2}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01])")


def collection_date(package: Path) -> tuple:
    """(iso_date, basis) — the package's own latest date stamp, or (None,
    reason). WHY (owner, 2026-08-20: 'for most evidence, recency is tagged
    unverified — improve this'): most ingested rows carry no publication
    date, but every row in an assessment package was OBSERVED no later than
    the package's own stamp, and 'observed as of the assessment date' is a
    real, defensible date with explicit provenance — strictly more honest
    than UNVERIFIED for a fact the assessors verified when they wrote it.
    File mtimes are NEVER used (a Drive pull stamps download time); only
    dates the package itself states: file/folder names first, then the
    heads of export CSVs and JSON stores. A publication date still outranks
    it — gaps_out keeps emitting dating search_requests for these rows."""
    hits = []
    for f in package.rglob("*"):
        if f.is_dir():
            continue
        m = DATE_IN_NAME.search(f.name)
        if m:
            y, mo, d = m.group(1), m.group(2), m.group(3) or "01"
            hits.append((f"{y}-{mo}-{d}", f"file name {f.name!r}"))
        if f.suffix.lower() in (".csv", ".json", ".jsonl"):
            try:
                head = f.open(errors="ignore").read(2048)
            except OSError:
                continue
            for iso in ISO_DATE.findall(head):
                hits.append((iso, f"head of {f.name!r}"))
    from datetime import date as _date
    today = _date.today().isoformat()
    hits = [h for h in hits if h[0] <= today]     # never a future stamp
    if not hits:
        return None, ("no date stamp in package file names or store heads — "
                      "rows without dates stay UNVERIFIED")
    best = max(hits)
    return best[0], f"latest package stamp: {best[1]}"


def apply_collection_date(records: dict, cdate: str, basis: str) -> int:
    """Date the still-dateless rows at collection, provenance explicit."""
    n = 0
    for rec in records.values():
        if not rec.get("date"):
            rec["date"] = cdate
            rec["date_provenance"] = "collection"
            rec["date_basis"] = basis
            rec.setdefault("field_provenance", {})["date"] = {
                "how": "collection", "basis": basis}
            n += 1
    return n


def gaps_out(records: dict, client: str | None) -> list:
    out = []
    for eid, rec in sorted(records.items()):
        missing = [f for f in ("url", "date", "excerpt")
                   if not rec.get(f)]
        if not rec.get("date"):
            rec["recency"] = "UNVERIFIED"
        elif rec.get("date_provenance") == "collection":
            missing = sorted(set(missing) | {"publication_date"})
        if not missing:
            continue
        src = rec.get("source") or ""
        # A summary is the right thing to SEARCH with and the wrong thing to
        # cite, so it steers the query when no verbatim excerpt exists.
        lead = (rec.get("excerpt") or rec.get("summary") or "")[:60]
        q = " ".join(x for x in (client or "", src, lead) if x)
        if rec.get("date_provenance") == "collection":
            note = ("dated at collection ({}) — a publication date outranks "
                    "it; connector search upgrades the row".format(
                        rec.get("date_basis", "package stamp")))
        elif "excerpt" in missing:
            note = ("no verbatim excerpt in any register this package ships. "
                    "Retrieve the source and take the span it actually says — "
                    "a summary, a ledger line or a nearby sentence is NOT an "
                    "excerpt, and the row stays uncitable until a real one "
                    "lands")
        else:
            note = ("corpus exhausted — web retrieval through the session's "
                    "connectors; undated stays UNVERIFIED until a real date "
                    "lands")
        out.append({"eid": (f"{client}:{eid}" if client else eid),
                    "missing": missing,
                    "query": q.strip() or eid,
                    "note": note})
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--package", required=True)
    ap.add_argument("--client", default=None,
                    help="client slug — prefixes ids in the gaps output")
    ap.add_argument("--out", default=None,
                    help="write normalized records as JSONL here")
    ap.add_argument("--assessment-date", default=None,
                    help="ISO date the package was collected; overrides "
                         "auto-derivation from the package's own stamps")
    a = ap.parse_args(argv)
    package = Path(a.package)
    if not package.is_dir():
        print(f"not a directory: {package}", file=sys.stderr)
        return 2
    records, conflicts, unrecognised = merge(package)
    filled = corpus_fill(package, records)
    if a.assessment_date:
        cdate, basis = a.assessment_date, "owner-supplied --assessment-date"
    else:
        cdate, basis = collection_date(package)
    dated_at_collection = (
        apply_collection_date(records, cdate, basis) if cdate else 0)
    gaps = gaps_out(records, a.client)
    if a.out:
        with open(a.out, "w") as fh:
            for eid in sorted(records):
                rec = {k: v for k, v in records[eid].items()
                       if not k.startswith("_")}
                fh.write(json.dumps(rec) + "\n")
    full = sum(1 for r in records.values()
               if r.get("url") and r.get("date") and r.get("excerpt"))
    # AN UNRECOGNISED ID VOCABULARY IS NOT AN EMPTY PACKAGE. Reported at the
    # top of the output, because "records: 0" was previously the only trace
    # of 1,013 rows that five agreeing stores all carried.
    dropped = sum(len(v) for v in unrecognised.values())
    print(json.dumps({
        "records": len(records), "schema_complete": full,
        "unrecognised_ids": dropped,
        "unrecognised_by_store": {k: {"rows": len(v), "examples": v[:3]}
                                  for k, v in unrecognised.items()} or None,
        "corpus_fills": filled,
        "collection_date": cdate, "collection_basis": basis,
        "dated_at_collection": dated_at_collection,
        "conflicts": conflicts[:20],
        "conflict_count": len(conflicts),
        "gaps": gaps[:25], "gap_count": len(gaps)}, indent=1))
    return 0


if __name__ == "__main__":
    sys.exit(main())
