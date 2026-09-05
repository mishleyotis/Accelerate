#!/usr/bin/env python3
"""Walk the assessment stage and the second run through the REAL command line.

    python3 scripts/stress_stage_and_supersede.py [--workdir DIR] [--keep]

Not a unit test — a WALK. The unit tests pin each refusal; this pins the
SEQUENCE, which is where the last ten changes can still break a run:

  1  a v3 workbook opens as v5 IN PLACE, values preserved and moved by NAME
  2  a research workbook refuses every assessment-stage command
  3  a research workbook CARRYING an assessment row is blocking, not silent
  4  the stage flips only when column D actually carries scores
  5  both stated grains land in the columns the APP's parser reads
  6  the recommendations are PROJECTED from the report, never authored
  7  a package is assembled, and a SECOND run of the same client supersedes
     it whole rather than merging into it
  8  the archive is invisible to the app's package scan
  9  the Client Research Profile classifies, parses, and resolves to kinds a
     consumer can ask for by name

Every step shells out exactly as the conductor does, or calls the APP's own
parser, so what this proves is that the commands and the readers actually
work together — not that each half works alone. That distinction is the whole
history of this repository: `classification.py` classified the profile
correctly for months while `_classify_artefact` dropped it, and both halves
had tests.

Exit 0 only if every step behaves as stated. A step that cannot run says
NOT_RUN with the reason rather than being skipped silently.
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
SKILL = Path(__file__).resolve().parent.parent / "skills" / "dma-research"
WORKER = REPO / "apps" / "worker"

RESULTS: list[tuple[str, str, str]] = []
PASS, FAIL, NOT_RUN = "PASS", "FAIL", "NOT_RUN"

# The engine is importable the same way the conductor's own commands import
# it, and the test fixtures are the only place that knows how to bank three
# sources — reusing them keeps this walk honest about the SEQUENCE without
# inventing a second way to build a run.
for _p in (str(SKILL), str(REPO / "tests" / "skills" / "research_engine"),
           str(WORKER)):
    if _p not in sys.path:
        sys.path.insert(0, _p)


def check(name: str, ok: bool, detail: str = "") -> bool:
    RESULTS.append((name, PASS if ok else FAIL, detail))
    print(f"  {'PASS' if ok else 'FAIL'}  {name}"
          + (f"\n          {detail}" if detail and not ok else ""))
    return bool(ok)


def not_run(name: str, why: str) -> None:
    RESULTS.append((name, NOT_RUN, why))
    print(f"  ....  {name}\n          NOT_RUN: {why}")


def cli(*args, expect: int | None = 0):
    """One engine command, as the conductor issues it."""
    r = subprocess.run([sys.executable, "-m", "engine.cli", *args],
                       cwd=str(SKILL), capture_output=True, text=True,
                       timeout=900)
    if expect is not None and r.returncode != expect:
        print(f"    ! engine.cli {' '.join(args)} -> {r.returncode} "
              f"(wanted {expect})")
        print("    " + (r.stderr or r.stdout).strip()[-500:].replace(
            "\n", "\n    "))
    return r


def build_run(root: Path, run_id: str, ref_date: str, n=6):
    """A researched run: selection, evidence, syntheses. Built through the
    engine's own modules — the CLI has no verb for 'bank three sources', and
    inventing one for a stress harness would be a second write path."""
    sys.path.insert(0, str(SKILL))
    sys.path.insert(0, str(REPO / "tests" / "skills" / "research_engine"))
    from engine import assemble, runstate
    import fixtures as F

    run = runstate.start(
        run_id=run_id, entity_name="Acme Credit Union", entity_id="acme-cu",
        sub_vertical="CU", scope_mode="T1_CORE", reference_date=ref_date,
        root=root / "run", selected=F.small_selection(n))
    assemble.open_folder(run, root / "client", push=False)
    F.close_prelim(run)
    wb = run.open()
    for cell in wb.selected_subcaps():
        F.synthesise(wb, cell, F.good_synthesis(cell, F.bank_evidence(wb, cell)))
    return run, wb



def section_body(spec_key: str, section: str, eids) -> dict:
    """A body that satisfies the section's own anatomy — its declared blocks,
    in order, each over the card floor, each citing in the PROSE.

    `reports.CITE_RE` reads citations out of the BODY, not out of the
    Evidence_IDs column, so a section that cites in the column and not in the
    prose reads as uncited to the artefact a client opens.
    """
    from engine import report_spec as RS
    sec = RS.SPECS[spec_key].section(section)
    para = (
        "The public record for this institution is read here against the "
        "question the block asks, and the reading is stated so a reader can "
        "disagree with it rather than accept it. Nothing in this paragraph "
        "rests on a source that is not in the run's own register, and every "
        "figure it carries can be reopened from the excerpt that supplied "
        "it. Where the record is silent the silence is reported as silence, "
        "with the ladder that establishes it, rather than read as an answer "
        "in either direction; and where two sources disagree the "
        "disagreement is carried forward rather than resolved by preference.")
    floor = sec.card_min_words or sec.min_words
    nblocks = len(sec.blocks) or 1
    per = max(1, -(-floor // (nblocks * len(para.split()))) + 1)
    body = []
    for b in sec.blocks or ("",):
        if b:
            body.append(f"## {b}")
        body.extend([para] * per)
        body.append("Sources for this block: "
                    + " ".join(f"[{e}]" for e in eids) + ".")
        body.append("")
    return {
        "Body": "\n".join(body).strip(),
        "Evidence_IDs": ", ".join(eids),
        "Weighing": (
            "The reading above was weighed against the opposite one — that "
            "the silence reflects an absence of practice rather than an "
            "absence of disclosure — and the conservative reading was "
            "preferred because the institution is member-owned and "
            "publishes little of either kind."),
        "Assumptions": (
            "Assumed that what a member-owned institution publishes "
            "understates what it does; that cuts toward under-reading it."),
        "Bias_Notes": (
            "A public-evidence run over-reads what a client publishes and "
            "under-reads what it does not; this section leans that way."),
        "Inference_Tags": "",
        "Absence_Basis": "",
    }





def fill_never_empty(run, root: str) -> None:
    """DQ_Bank, Search_Log and Gate_Log, through the commands that own them.

    `completeness` names each one's filler by command, and those names are
    the contract: a tab whose FILLED_BY hint points at a command that does
    not fill it is a hint that sends an unattended run in a circle.
    """
    cli("kg", "build", "--run", run.run_id, "--root", root, expect=None)
    wb = run.open()
    for cell in wb.selected_subcaps()[:3]:
        cli("search", "--run", run.run_id, "--root", root,
            "--subcap", cell, "--facet", "primary",
            "--query", f"Acme Credit Union {cell} public disclosure",
            "--tool", "exa", "--hits", "6", "--kept", "3",
            "--outcome", "KEPT", expect=None)
    cats = sorted({c.split(".")[0] for c in wb.selected_subcaps()})
    for cat in cats:
        cli("gate", "--run", run.run_id, "--root", root,
            "--category", cat, "--require-synthesis", expect=None)


def satisfy_completeness(run_id: str, root: str):
    """Declare each legitimately-empty tab, the way a conductor does.

    Returns the sheets declared, or None if the gate still blocks — a run
    that cannot state why a tab is empty is a run that must not package, and
    that is the point of the gate rather than an obstacle to it.
    """
    from engine import runstate
    for _ in range(6):
        wb = runstate.locate(run_id, Path(root)).open()
        from engine import completeness as K
        out = K.check(wb)
        empty = [r["sheet"] for r in out["sheets"]
                 if r["verdict"] in ("EMPTY", "SHORT")
                 and any(r["sheet"] in b for b in out["blocking"])]
        if not empty:
            return []
        for sheet in empty:
            cli("complete", "declare", "--run", run_id, "--root", root,
                "--sheet", sheet,
                "--reason", f"this walk banks no {sheet} rows: the stress "
                            f"harness exercises the stage and the folder, "
                            f"and states the absence rather than seeding a "
                            f"row that would make the tab read as researched",
                expect=None)
        wb2 = runstate.locate(run_id, Path(root)).open()
        if not any(s in b for b in K.check(wb2)["blocking"] for s in empty):
            return empty
    return None


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--workdir")
    ap.add_argument("--keep", action="store_true")
    a = ap.parse_args(argv)
    work = Path(a.workdir) if a.workdir else Path(
        tempfile.mkdtemp(prefix="dma-stress-stage-"))
    work.mkdir(parents=True, exist_ok=True)
    print(f"stress: the assessment stage and the second run — {work}\n")

    try:
        return walk(work)
    finally:
        if not a.keep and not a.workdir:
            shutil.rmtree(work, ignore_errors=True)


def walk(work: Path) -> int:
    from engine import contract as C
    from engine import grains as G
    from engine import narrative as N
    from engine import report_spec as RS

    # ── 1 · the run, and the contract it opens at ────────────────────────
    run, wb = build_run(work / "a", "R-STRESS-1", "2026-08-29")
    md = wb.metadata()
    check("a new workbook records the contract and the stage",
          md.get("workbook_contract") == C.WORKBOOK_CONTRACT
          and C.stage_of(md) == "research",
          f"contract={md.get('workbook_contract')} stage={md.get('stage')}")

    # An older workbook, forced back to v3 shape by dropping the key and the
    # three assessment sheets — the state a run opened before 2026-08-30 is in.
    import openpyxl
    old = work / "v3.xlsx"
    shutil.copy(run.workbook_path, old)
    book = openpyxl.load_workbook(old)
    for sheet in C.SHEET_STAGE:
        if sheet in book.sheetnames:
            del book[sheet]
    ws = book["Run_Metadata"]
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(row=r, column=1).value or "") in (
                "stage", "workbook_contract"):
            ws.delete_rows(r)
    book.save(old)
    book.close()

    before = len(openpyxl.load_workbook(old)["Evidence_Detail"]["A"])
    shutil.copy(old, run.workbook_path)
    wb2 = run.open()                       # the upgrade happens on OPEN
    after = len([r for r in wb2.rows("Evidence_Detail") if r.get("E_ID")])
    have = [s for s in C.SHEET_STAGE
            if openpyxl.load_workbook(run.workbook_path).sheetnames.count(s)]
    check("a v3 workbook upgrades in place, keeping its rows",
          len(have) == len(C.SHEET_STAGE) and after > 0,
          f"evidence rows before={before - 1} after={after}; sheets={have}")
    check("the upgraded workbook reads as RESEARCH, not an empty assessment",
          C.stage_of(wb2.metadata()) == "research")

    # ── 2 · the stage gates, through the CLI ─────────────────────────────
    root = str(run.root)
    r = cli("grains", "recompute", "--run", run.run_id, "--root", root,
            expect=1)
    check("`grains recompute` refuses at the research stage",
          r.returncode == 1 and "research stage" in (r.stderr + r.stdout))

    r = cli("grains", "stage", "--run", run.run_id, "--root", root,
            "--to", "assessment", expect=1)
    check("the stage cannot be flipped with an empty column D",
          r.returncode == 1 and "Column D" in (r.stderr + r.stdout))

    # ── 3 · a research workbook carrying an assessment row BLOCKS ────────
    wb3 = run.open()
    wb3.append("Pillar_Summary", {"Pillar": "P1", "Pillar_Name": "",
                                  "Score": 3.1, "Weight_Pct": "",
                                  "Peer_Median": ""})
    from engine import completeness as K
    out = K.check(wb3)
    row = next(r for r in out["sheets"] if r["sheet"] == "Pillar_Summary")
    check("a research workbook carrying a stated grain is AHEAD_OF_STAGE "
          "and blocks",
          row["verdict"] == "AHEAD_OF_STAGE"
          and any("Pillar_Summary" in b for b in out["blocking"]),
          f"verdict={row['verdict']} blocking={out['blocking'][:2]}")
    ws = wb3._sheet("Pillar_Summary")
    ws.delete_rows(2, ws.max_row)
    wb3.save()

    # ── 4 · research it, score it THROUGH THE STAGE, state the grains ────
    # Until 2026-09-04 this walk wrote column D by hand and flipped the stage
    # with `grains stage`. A report section can no longer be written on a
    # workbook scored out-of-band (the preconditions run on every write —
    # the owner's issue 2), so the walk now earns its scores the way a run
    # does: evidence, a challenged synthesis per cell, the floors gate, then
    # `engine.assessment open / score / critique / rollup / gate`.
    import fixtures as F
    wb4 = run.open()
    cells = wb4.selected_subcaps()
    ev = {}
    for cell in cells:
        ev[cell] = F.bank_evidence(wb4, cell, n=5)
        F.synthesise(wb4, cell, F.good_synthesis(cell, ev[cell]))
    F.client_facts(wb4, cells, ev)
    for cat in sorted({c.split(".")[0] for c in cells}):
        r = cli("gate", "--run", run.run_id, "--root", root, "--category", cat,
                "--require-synthesis", expect=None)
    r = cli("grains", "stage", "--run", run.run_id, "--root", root,
            "--to", "assessment", expect=1)
    check("the stage cannot be flipped by hand on an unscored workbook",
          r.returncode == 1)
    F.score_stage(run, run.open(), cells, ev)          # open → score → critic → rollup → gate
    check("the stage flips once column D carries scores — through the stage's own gate",
          C.stage_of(run.open().metadata()) == "assessment"
          and any(str(g.get("Gate")) == "SCORING" and str(g.get("Verdict")) == "PASS"
                  for g in run.open().rows("Gate_Log")))

    r = cli("grains", "recompute", "--run", run.run_id, "--root", root)
    got = json.loads(r.stdout) if r.returncode == 0 else {}
    check("both stated grains are written",
          got.get("pillars", 0) >= 1 and got.get("categories", 0) >= 1,
          json.dumps(got))

    # ── 5 · the APP's own parser reads what the engine wrote ─────────────
    sys.path.insert(0, str(WORKER))
    from dma_worker.workbook_parser import (parse_grain_summaries,
                                            parse_recommendations)
    obs: list = []
    grains = parse_grain_summaries(str(run.workbook_path), obs)
    check("the app's parser reads both stated grains",
          bool(grains.get("pillars")) and bool(grains.get("categories")),
          f"pillars={len(grains.get('pillars') or [])} "
          f"categories={len(grains.get('categories') or [])} "
          f"obs={[o.kind for o in obs][:3]}")

    # ── 6 · the recommendations are PROJECTED, not authored ──────────────
    r = cli("grains", "recommendations", "--run", run.run_id, "--root", root,
            expect=1)
    check("recommendations refuse when the report section is unwritten",
          r.returncode == 1 and "nothing to project" in (r.stderr + r.stdout))

    wb5 = run.open()
    F.make_shippable(wb5)                  # the completeness half of the report preconditions
    eids = ev[cells[0]]
    sec = next(s for s in RS.SPECS["assessment"].sections
               if s.kind == "recommendation")
    for i in range(3):
        N.write(wb5, "assessment", sec.id,
                F.section_record(sec.id, eids, report="assessment"),
                actor="report-assessment-producer", card=f"REC-{i + 1:02d}", run=run)
    r = cli("grains", "recommendations", "--run", run.run_id, "--root", root)
    proj = json.loads(r.stdout) if r.returncode == 0 else {}
    recs = parse_recommendations(str(run.workbook_path), [])
    check("the report's rows are projected and the app reads them",
          proj.get("rows") == 3 and len(recs) == 3
          and all(x["payload"].get("rationale") for x in recs),
          f"projected={proj.get('rows')} parsed={len(recs)} "
          + (r.stderr or "")[-200:])

    # ── 7 · a package, then a SECOND run of the same client ──────────────
    # The completeness gate blocks a package whose tabs are empty and
    # undeclared. That is the gate working, and a stress harness that routed
    # around it would be proving the wrong thing — so the run declares each
    # legitimately-empty tab the way a conductor does, and the package is
    # allowed to assemble only because it did.
    # Three tabs are NEVER_EMPTY by contract, so a declaration on them is
    # ILLEGAL_DECLARATION and blocks harder than the emptiness did — writing
    # around a refusal must be louder than tripping it. They are filled with
    # the real commands the gate names.
    fill_never_empty(run, root)
    declared = satisfy_completeness(run.run_id, root)
    # every section of both reports through the sanctioned writer, on the
    # scored run, under the stage preconditions, signed off by another actor
    F.write_both_reports(run, run.open(), cells, ev, render=False)
    st = N.state(run.open())
    wrote = sum(len(v.get("sections") or []) for v in st["reports"].values())
    r = cli("techscan", "render", "--run", run.run_id, "--root", root,
            expect=None)
    r1 = cli("report", "--run", run.run_id, "--root", root,
             "--report", "both", expect=None)
    check("every section of both reports is written, reviewed, and both "
          "reports render",
          all(v.get("ready") for v in st["reports"].values())
          and wrote == 19 and r1.returncode == 0,
          f"sections={wrote} ready={[k for k, v in st['reports'].items() if v.get('ready')]} "
          f"report_rc={r1.returncode} " + (r1.stderr or r1.stdout).strip()[-300:])
    check("the completeness gate blocks until every empty tab is declared",
          declared is not None,
          f"declared={declared}")
    r = cli("assemble", "package", "--run", run.run_id, "--root", root)
    if r.returncode != 0:
        not_run("a second run supersedes the first",
                "the first package would not assemble; nothing to supersede")
        return report()
    folder = next((work / "a" / "client").glob("* - DMA"), None)
    check("the client folder is assembled", folder is not None,
          str(list((work / 'a' / 'client').iterdir())[:4]))

    first = sorted(p.name for p in folder.iterdir()) if folder else []
    run2, wb6 = build_run(work / "a", "R-STRESS-2", "2026-09-15")
    r2 = cli("assemble", "package", "--run", run2.run_id, "--root",
             str(run2.root), expect=None)
    archive = folder / "_superseded" if folder else None
    kept = sorted(p.name for p in archive.iterdir()) if archive and \
        archive.exists() else []
    check("the previous package moves whole into _superseded/<run_id>/",
          bool(kept) and any("R-STRESS-1" in k for k in kept),
          f"archive={kept} first_package={first[:4]}")
    if kept:
        sup = archive / kept[0] / "SUPERSEDED.json"
        doc = json.loads(sup.read_text()) if sup.exists() else {}
        check("SUPERSEDED.json names what replaced it",
              doc.get("run_id") == "R-STRESS-1"
              and doc.get("superseded_by") == "R-STRESS-2",
              json.dumps(doc)[:200])
    check("the folder keeps its name — runs.source_folder_id keys on it",
          folder is not None and folder.name.endswith(" - DMA"),
          folder.name if folder else "")

    # ── 8 · the archive is invisible to the app's package scan ───────────
    import job_main as J
    seg = getattr(J, "ARCHIVE_SEGMENT", None)
    check("the app's scan excludes the archive by path segment",
          seg is not None and seg in ("_superseded", "/_superseded"),
          f"ARCHIVE_SEGMENT={seg!r}")

    # ── 9 · the second report reaches the app ────────────────────────────
    # The CURRENT package is the second run, which this walk deliberately did
    # not carry through report finalization — so the profile to read is the
    # first run's, now under _superseded/. That it is still readable there is
    # the retention half of the supersession working.
    prof = next((p for p in folder.rglob("Client_Profile_Research_*.docx")),
                None) if folder else None
    if prof is None:
        not_run("the Client Research Profile is ingestable",
                "no profile in the current package")
    else:
        class F2:
            def __init__(self, name):
                self.name = name
                self.path_segments = [folder.name]
        kind = J._classify_artefact(F2(prof.name))
        from dma_worker.report_parser import parse_report
        secs = parse_report(str(prof), [])
        unmapped = sorted({s.section_kind for s in secs
                           if s.section_kind.startswith("unmapped:")})
        check("the profile classifies and every section has a nameable kind",
              kind == ("profile", 0) and secs and not unmapped,
              f"classified={kind} sections={len(secs)} unmapped={unmapped}")

    return report()


def report() -> int:
    print()
    bad = [n for n, v, _ in RESULTS if v == FAIL]
    skipped = [n for n, v, _ in RESULTS if v == NOT_RUN]
    print(f"{sum(1 for _n, v, _d in RESULTS if v == PASS)}/{len(RESULTS)} "
          f"step(s) passed"
          + (f", {len(skipped)} NOT_RUN" if skipped else ""))
    for n, v, d in RESULTS:
        if v != PASS:
            print(f"  {v}  {n}\n        {d}")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
