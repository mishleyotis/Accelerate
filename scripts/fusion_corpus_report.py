#!/usr/bin/env python3
"""Does rank fusion change platform selection on REAL client gap surfaces?

Owner, 2026-08-23: "reciprocal rank fusion for platform selection tested
against client corpus."

`apps/api/tests/test_platform_fit_fusion.py` proves the properties — the
band holds, the placings survive rescaling, the fusion is neither decorative
nor in charge — over GENERATED runs. Generated runs answer "is the mechanism
sound"; they cannot answer "does it fire on the shapes clients actually
have", and the difference is not academic: the first probe of this fusion
used two candidates, found sixty-four near-ties and reordered none, and read
exactly like a fusion that does nothing. It was the probe that was wrong.

So this is the instrument. It reads each package's own scoring workbook —
the real per-cell scores, the real category grain, the real confidence
distribution — builds candidate platforms from them, and reports what
fusion does.

WHAT IS REAL HERE AND WHAT IS NOT, because a corpus report that blurs the
two is worth less than no report:

  REAL, read from the client's workbook
    · every cell's score, and therefore every gap
    · the category each cell belongs to, and therefore the interconnect
      surface and the candidate grouping
    · the per-cell confidence, mapped to evidence strength
    · how many cells each client actually carries, which is the variable
      that decides whether near-ties happen at all

  SWEPT, because it is a producer's judgement and no workbook holds it
    · readiness (green/amber/red) and strategic alignment (0..1)
    · greenfield, which comes from the technology register

  Sweeping rather than inventing one value per client is the point: a single
  invented pair would make the whole report a function of a number this
  script chose. The sweep reports the DISTRIBUTION over a fixed grid, so the
  figure quoted is "across N clients x M judgement settings", and the grid
  is in the output.

    python3 scripts/fusion_corpus_report.py
    python3 scripts/fusion_corpus_report.py --root /root/.dma/packages --json out.json
    python3 scripts/fusion_corpus_report.py --min-fire 0.05   # exit 1 below

Exit codes: 0 measured, 1 a stated floor was missed, 2 nothing to measure.
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from itertools import product
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "packages" / "shared"))

import platform_fit as pf  # noqa: E402

DEFAULT_ROOT = Path("/root/.dma/packages")

#: Confidence -> evidence strength. The same ladder the engine's own tier
#: weights use in spirit: a HIGH-confidence cell is bankable, a LOW one is
#: discounted rather than discarded. A blank confidence takes the engine's
#: neutral prior rather than a guess.
CONFIDENCE = {"HIGH": 0.92, "MEDIUM": 0.72, "LOW": 0.48}

#: The judgement grid. Five readiness patterns and four alignment patterns,
#: applied across a client's candidates by position — so the sweep covers
#: "all green", "the leader is red", "alignment agrees with gap size",
#: "alignment disagrees with gap size", and so on.
READINESS_PATTERNS = {
    "all-green": ["green"] * 8,
    "leader-red": ["red"] + ["green"] * 7,
    "leader-amber": ["amber"] + ["green"] * 7,
    "mixed": ["green", "amber", "red", "green", "amber", "green", "red", "amber"],
    "all-amber": ["amber"] * 8,
}
ALIGNMENT_PATTERNS = {
    "agrees-with-gap": [0.95, 0.80, 0.65, 0.50, 0.35, 0.30, 0.25, 0.20],
    "disagrees-with-gap": [0.20, 0.30, 0.50, 0.65, 0.80, 0.95, 0.90, 0.85],
    "flat": [0.6] * 8,
    "unknown": [None] * 8,
}
GREENFIELD_PATTERNS = {
    "none": [False] * 8,
    "tail-absent": [False, False, False, True, True, True, True, True],
    "leader-absent": [True] + [False] * 7,
}

CANDIDATES_PER_CLIENT = 5


# ── reading a real package ────────────────────────────────────────────

def scoring_workbook(pkg: Path):
    """The scoring workbook, or None. A package without one was never a
    synthesis input and is reported separately rather than as a failure."""
    for d in (pkg / "03_scoring_workbook", pkg):
        if d.is_dir():
            for f in sorted(d.glob("*.xlsx")):
                if not f.name.startswith("~$"):
                    return f
    return None


def _num(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def cells_of(path: Path) -> list:
    """Every scored subcap in the workbook, as engine Cells.

    Reads the four `P?_Subcap_Scoring` sheets by HEADER NAME rather than by
    column index: the header order differs between workbook generations, and
    a positional read is how a column of scores gets read as a column of
    confidences without anything looking wrong.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    try:
        for name in wb.sheetnames:
            if not name.endswith("_Subcap_Scoring"):
                continue
            ws = wb[name]
            head = None
            for row in ws.iter_rows(values_only=True):
                if head is None:
                    if row and any(str(c or "").strip() == "SubCap_ID" for c in row):
                        head = {str(c or "").strip(): i for i, c in enumerate(row)}
                    continue
                def get(col):
                    i = head.get(col)
                    return row[i] if i is not None and i < len(row) else None
                sid = str(get("SubCap_ID") or "").strip()
                if not sid:
                    continue
                score = _num(get("Score"))
                if score is None:
                    continue
                conf = str(get("Confidence") or "").strip().upper()
                out.append(pf.Cell(
                    subcap_id=sid,
                    current_score=score,
                    category_id=str(get("Category") or "").strip() or None,
                    severities=(),          # the issue register is a separate
                                            # sheet and often empty; a neutral
                                            # weight is the engine's own
                                            # answer, not this script's
                    evidence_strength=CONFIDENCE.get(conf)))
    finally:
        wb.close()
    return out


def candidates_of(cells, n=CANDIDATES_PER_CLIENT) -> list:
    """Group the client's real cells into candidate platforms by category.

    Category is the grain the catalogue's L3 areas cluster at, and it is the
    grain interconnect is computed over — so a candidate built this way has
    a real cell set, a real gap distribution and a real adjacency, which are
    the three inputs fusion actually operates on. The n largest categories
    are taken because that is the shape of a real page: a handful of areas
    with substantial surface, not every category the workbook touches.
    """
    by_cat = {}
    for c in cells:
        by_cat.setdefault(c.category_id or "?", []).append(c)
    big = sorted(by_cat.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    return [(cat, group) for cat, group in big[:n]
            if len(group) >= pf.MIN_CELLS]


# ── the measurement ───────────────────────────────────────────────────

def measure(pkg_cells, name):
    """Sweep the judgement grid over one client's real cells."""
    groups = candidates_of(pkg_cells)
    if len(groups) < 2:
        return None
    rows = []
    grid = product(READINESS_PATTERNS.items(), ALIGNMENT_PATTERNS.items(),
                   GREENFIELD_PATTERNS.items())
    for (rk, rv), (ak, av), (gk, gv) in grid:
        cands = [
            pf.Candidate(platform=f"{name}:{cat}", l3_area=cat, cells=group,
                         family_absent=gv[i], readiness=rv[i], alignment=av[i])
            for i, (cat, group) in enumerate(groups)]
        ranked = pf.rank(cands, pkg_cells)
        fits = [r["fit_score"] for r in ranked]
        moved = [r for r in ranked if r["rank_basis"].startswith("rank fusion")]
        # THE BOUND, checked here too and not only in the unit tests: this
        # runs over real shapes the generators do not produce.
        violations = 0
        for i, a in enumerate(ranked):
            for b in ranked[i + 1:]:
                if a["rank_basis"].startswith("sequenced"):
                    continue
                if a["fit_score"] < b["fit_score"] and \
                   (b["fit_score"] - a["fit_score"]) > pf.FUSION_BAND:
                    violations += 1
        rows.append({
            "client": name,
            "setting": f"{rk}/{ak}/{gk}",
            "candidates": len(ranked),
            "near_tie": any(fits[i] - fits[i + 1] <= pf.FUSION_BAND
                            for i in range(len(fits) - 1)),
            "fusion_fired": bool(moved),
            "cards_moved": len(moved),
            "top_card_changed": bool(moved and ranked[0] in moved),
            "violations": violations,
            "fit_spread": round(max(fits) - min(fits), 1) if fits else 0.0,
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=str(DEFAULT_ROOT), type=Path)
    ap.add_argument("--json", type=Path)
    ap.add_argument("--min-fire", type=float, default=None,
                    help="fail below this share of settings where fusion fired")
    args = ap.parse_args()

    if not args.root.is_dir():
        print(f"no package root at {args.root} — nothing to measure. Pull "
              f"packages with drive_fetch.py first.", file=sys.stderr)
        return 2

    packages = sorted(p for p in args.root.iterdir() if p.is_dir())
    all_rows, read, no_workbook, too_few = [], [], [], []
    for pkg in packages:
        wbk = scoring_workbook(pkg)
        if wbk is None:
            no_workbook.append(pkg.name)
            continue
        try:
            cells = cells_of(wbk)
        except Exception as exc:                            # noqa: BLE001
            no_workbook.append(f"{pkg.name} (unreadable: {exc})")
            continue
        rows = measure(cells, pkg.name)
        if rows is None:
            too_few.append(f"{pkg.name} ({len(cells)} cells)")
            continue
        read.append((pkg.name, len(cells)))
        all_rows += rows

    if not all_rows:
        print("no client produced a measurable candidate set.", file=sys.stderr)
        for n in no_workbook:
            print(f"  no scoring workbook: {n}", file=sys.stderr)
        return 2

    n = len(all_rows)
    fired = sum(1 for r in all_rows if r["fusion_fired"])
    near = sum(1 for r in all_rows if r["near_tie"])
    tops = sum(1 for r in all_rows if r["top_card_changed"])
    cards = sum(r["candidates"] for r in all_rows)
    moved = sum(r["cards_moved"] for r in all_rows)
    viol = sum(r["violations"] for r in all_rows)

    print(f"FUSION AGAINST THE CLIENT CORPUS")
    print(f"  clients read              {len(read)}"
          f"  ({sum(c for _, c in read)} scored cells)")
    if no_workbook:
        print(f"  no scoring workbook       {len(no_workbook)}"
              f"   (not a refusal — a briefing folder is not a synthesis input)")
    if too_few:
        print(f"  too few candidates        {len(too_few)}"
              f"   (fewer than 2 categories with >= {pf.MIN_CELLS} cells)")
    print(f"  judgement settings/client {n // max(len(read), 1)}"
          f"   ({len(READINESS_PATTERNS)} readiness x "
          f"{len(ALIGNMENT_PATTERNS)} alignment x "
          f"{len(GREENFIELD_PATTERNS)} greenfield)")
    print(f"  runs measured             {n}")
    print()
    print(f"  contained a near-tie      {near:5}  {near / n:6.1%}")
    print(f"  fusion reordered a card   {fired:5}  {fired / n:6.1%}")
    print(f"  cards moved               {moved:5}  {moved / cards:6.1%} of all cards")
    print(f"  top card changed          {tops:5}  {tops / n:6.1%}")
    print(f"  band violations           {viol:5}  <- must be 0")
    print()
    worst = Counter(r["client"] for r in all_rows if r["fusion_fired"])
    if worst:
        print("  clients where fusion fires most (real gap surfaces):")
        for cname, c in worst.most_common(8):
            print(f"    {c:4} / {n // max(len(read), 1)}  {cname}")

    if args.json:
        args.json.write_text(json.dumps({
            "clients": [{"name": a, "cells": b} for a, b in read],
            "no_workbook": no_workbook, "too_few": too_few,
            "runs": n, "near_tie": near, "fusion_fired": fired,
            "cards": cards, "cards_moved": moved,
            "top_card_changed": tops, "violations": viol,
            "k": pf.RRF_K, "band": pf.FUSION_BAND,
            "rows": all_rows,
        }, indent=2))
        print(f"\n  wrote {args.json}")

    if viol:
        print(f"\nFAIL: {viol} orderings broke the {pf.FUSION_BAND}-point band.",
              file=sys.stderr)
        return 1
    if args.min_fire is not None and (fired / n) < args.min_fire:
        print(f"\nFAIL: fusion fired on {fired / n:.1%} of runs, under the "
              f"{args.min_fire:.0%} floor — it is decorative on this corpus.",
              file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
